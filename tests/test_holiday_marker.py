"""
非工作日交易标记模块单元测试

覆盖场景：
1. 日期分类 - 工作日、周末、法定节假日、调休工作日正确识别
2. 交易标记 - 非工作日交易正确标记并统计
3. 打标回写 - apply_holiday_tags 正确回写标签到记录
4. 报告导出 - Excel 报告正确生成（含各 Sheet）
5. 边界场景 - 空数据、无交易日期、日期格式兼容
6. 多主体多银行混合场景
7. 调休工作日正确识别（周末上班）
8. generate 函数的 from_records / from_total 路径
"""
import os
import sys
import tempfile
import shutil
from datetime import datetime

import openpyxl
import pandas as pd
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck
from bankcheck import (
    classify_date,
    mark_non_workday_transactions,
    apply_holiday_tags,
    export_holiday_check_result,
    generate_holiday_check_from_records,
    generate_holiday_check_from_total,
    HolidayCheckResult,
    HolidayMarkedRecord,
    CHINESE_HOLIDAYS,
    CHINESE_WORKDAY_ADJUSTMENTS,
    HOLIDAY_TAG_WORKDAY,
    HOLIDAY_TAG_WEEKEND,
    HOLIDAY_TAG_HOLIDAY,
    HOLIDAY_TAG_ADJUSTED_WORKDAY,
    _parse_trade_date,
)


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='holiday_marker_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _record(unique_id, account, date, payment=None, receipt=None, balance=0.0,
            bank='北京银行', subject='北京XX科技有限公司',
            summary='', counterpart='', txn_id=''):
    return {
        '唯一id': unique_id,
        '银行': bank,
        '银行账号': account,
        '主体': subject,
        '交易日期': date,
        '付款': payment,
        '收款': receipt,
        '摘要': summary,
        '对方户名': counterpart,
        '余额': balance,
        '交易流水号': txn_id or unique_id,
    }


class TestClassifyDate:

    def test_workday_monday(self):
        tag, name = classify_date('2024-01-08')
        assert tag == HOLIDAY_TAG_WORKDAY
        assert name == ''

    def test_workday_friday(self):
        tag, name = classify_date('2024-01-12')
        assert tag == HOLIDAY_TAG_WORKDAY
        assert name == ''

    def test_weekend_saturday(self):
        tag, name = classify_date('2024-01-06')
        assert tag == HOLIDAY_TAG_WEEKEND
        assert name == ''

    def test_weekend_sunday(self):
        tag, name = classify_date('2024-01-07')
        assert tag == HOLIDAY_TAG_WEEKEND
        assert name == ''

    def test_holiday_new_year_2024(self):
        tag, name = classify_date('2024-01-01')
        assert tag == HOLIDAY_TAG_HOLIDAY
        assert name == '元旦'

    def test_holiday_spring_festival_2024(self):
        tag, name = classify_date('2024-02-10')
        assert tag == HOLIDAY_TAG_HOLIDAY
        assert '春节' in name

    def test_holiday_national_day_2024(self):
        tag, name = classify_date('2024-10-01')
        assert tag == HOLIDAY_TAG_HOLIDAY
        assert '国庆节' in name

    def test_adjusted_workday_2024(self):
        tag, name = classify_date('2024-02-04')
        assert tag == HOLIDAY_TAG_ADJUSTED_WORKDAY
        assert name == ''
        dt = datetime(2024, 2, 4)
        assert dt.weekday() == 6

    def test_adjusted_workday_saturday_2024(self):
        tag, name = classify_date('2024-02-18')
        assert tag == HOLIDAY_TAG_ADJUSTED_WORKDAY
        assert name == ''
        dt = datetime(2024, 2, 18)
        assert dt.weekday() == 6

    def test_none_date(self):
        tag, name = classify_date(None)
        assert tag == HOLIDAY_TAG_WORKDAY
        assert name == ''

    def test_empty_date(self):
        tag, name = classify_date('')
        assert tag == HOLIDAY_TAG_WORKDAY
        assert name == ''

    def test_datetime_object(self):
        tag, name = classify_date(datetime(2024, 5, 1))
        assert tag == HOLIDAY_TAG_HOLIDAY
        assert '劳动节' in name

    def test_pandas_timestamp(self):
        ts = pd.Timestamp('2024-05-01')
        tag, name = classify_date(ts)
        assert tag == HOLIDAY_TAG_HOLIDAY
        assert '劳动节' in name

    def test_various_date_formats(self):
        tag1, _ = classify_date('2024-01-01')
        tag2, _ = classify_date(datetime(2024, 1, 1))
        assert tag1 == tag2 == HOLIDAY_TAG_HOLIDAY

    def test_holiday_coverage_2020_2026(self):
        assert len(CHINESE_HOLIDAYS) > 0
        for year in range(2020, 2027):
            year_holidays = [k for k in CHINESE_HOLIDAYS if k.startswith(str(year))]
            assert len(year_holidays) > 0, f'缺少 {year} 年节假日数据'

    def test_adjusted_workday_coverage(self):
        assert len(CHINESE_WORKDAY_ADJUSTMENTS) > 0


class TestParseTradeDate:

    def test_standard_format(self):
        dt = _parse_trade_date('2024-01-15')
        assert dt is not None
        assert dt.year == 2024 and dt.month == 1 and dt.day == 15

    def test_slash_format(self):
        dt = _parse_trade_date('2024/01/15')
        assert dt is not None
        assert dt.year == 2024

    def test_compact_format(self):
        dt = _parse_trade_date('20240115')
        assert dt is not None
        assert dt.year == 2024

    def test_none_value(self):
        assert _parse_trade_date(None) is None

    def test_empty_string(self):
        assert _parse_trade_date('') is None


class TestMarkNonWorkdayTransactions:

    def test_all_workday(self):
        records = [
            _record('1', '622001', '2024-01-08', payment=100),
            _record('2', '622001', '2024-01-09', receipt=200),
            _record('3', '622001', '2024-01-10', payment=300),
        ]
        result = mark_non_workday_transactions(records)
        assert result.total_records == 3
        assert result.workday_count == 3
        assert result.weekend_count == 0
        assert result.holiday_count == 0
        assert len(result.non_workday_records) == 0

    def test_mixed_dates(self):
        records = [
            _record('1', '622001', '2024-01-08', payment=100),
            _record('2', '622001', '2024-01-06', payment=200),
            _record('3', '622001', '2024-01-01', payment=300),
            _record('4', '622001', '2024-02-04', payment=400),
        ]
        result = mark_non_workday_transactions(records)
        assert result.total_records == 4
        assert result.workday_count == 1
        assert result.weekend_count == 1
        assert result.holiday_count == 1
        assert result.adjusted_workday_count == 1
        assert len(result.non_workday_records) == 2

    def test_holiday_name_stats(self):
        records = [
            _record('1', '622001', '2024-01-01', payment=100),
            _record('2', '622001', '2024-10-01', payment=200),
        ]
        result = mark_non_workday_transactions(records)
        assert '元旦' in result.holiday_name_stats
        assert '国庆节' in result.holiday_name_stats

    def test_subject_stats(self):
        records = [
            _record('1', '622001', '2024-01-06', payment=100, subject='公司A'),
            _record('2', '622001', '2024-01-07', payment=200, subject='公司A'),
            _record('3', '622001', '2024-01-01', payment=300, subject='公司B'),
        ]
        result = mark_non_workday_transactions(records)
        assert '公司A' in result.subject_stats
        assert result.subject_stats['公司A'][HOLIDAY_TAG_WEEKEND] == 2
        assert '公司B' in result.subject_stats
        assert result.subject_stats['公司B'][HOLIDAY_TAG_HOLIDAY] == 1

    def test_check_summary(self):
        records = [
            _record('1', '622001', '2024-01-08', payment=100),
            _record('2', '622001', '2024-01-06', payment=200),
        ]
        result = mark_non_workday_transactions(records)
        assert result.check_summary['total_records'] == 2
        assert result.check_summary['non_workday_count'] == 1
        assert result.check_summary['non_workday_rate'] == 50.0

    def test_empty_records(self):
        result = mark_non_workday_transactions([])
        assert result.total_records == 0
        assert result.check_summary.get('status') == '无数据'

    def test_non_workday_rate_calculation(self):
        records = [
            _record('1', '622001', '2024-01-08', payment=100),
            _record('2', '622001', '2024-01-09', payment=200),
            _record('3', '622001', '2024-01-06', payment=300),
            _record('4', '622001', '2024-01-07', payment=400),
        ]
        result = mark_non_workday_transactions(records)
        assert result.check_summary['non_workday_rate'] == 50.0


class TestApplyHolidayTags:

    def test_basic_tagging(self):
        records = [
            _record('1', '622001', '2024-01-08', payment=100),
            _record('2', '622001', '2024-01-06', payment=200),
            _record('3', '622001', '2024-01-01', payment=300),
        ]
        tagged, summary = apply_holiday_tags(records)
        assert tagged[0]['非工作日标签'] == HOLIDAY_TAG_WORKDAY
        assert tagged[0]['节假日名称'] == ''
        assert tagged[1]['非工作日标签'] == HOLIDAY_TAG_WEEKEND
        assert tagged[1]['节假日名称'] == ''
        assert tagged[2]['非工作日标签'] == HOLIDAY_TAG_HOLIDAY
        assert tagged[2]['节假日名称'] == '元旦'

    def test_summary_counts(self):
        records = [
            _record('1', '622001', '2024-01-08', payment=100),
            _record('2', '622001', '2024-01-06', payment=200),
            _record('3', '622001', '2024-01-01', payment=300),
        ]
        _, summary = apply_holiday_tags(records)
        assert summary['total_records'] == 3
        assert summary['tagged_count'] == 2
        assert summary['workday_count'] == 1
        assert summary['weekend_count'] == 1
        assert summary['holiday_count'] == 1

    def test_empty_records(self):
        tagged, summary = apply_holiday_tags([])
        assert tagged == []
        assert summary['tagged_count'] == 0

    def test_records_mutated_in_place(self):
        records = [
            _record('1', '622001', '2024-01-06', payment=100),
        ]
        tagged, _ = apply_holiday_tags(records)
        assert tagged is records
        assert records[0]['非工作日标签'] == HOLIDAY_TAG_WEEKEND


class TestExportHolidayCheckResult:

    def test_export_with_non_workday(self, tmp_dir):
        records = [
            _record('1', '622001', '2024-01-08', payment=100),
            _record('2', '622001', '2024-01-06', payment=200),
            _record('3', '622001', '2024-01-01', payment=300),
        ]
        result = mark_non_workday_transactions(records)
        output_path = os.path.join(tmp_dir, 'holiday_report.xlsx')
        export_holiday_check_result(result, output_path)

        assert os.path.exists(output_path)
        wb = openpyxl.load_workbook(output_path)
        assert '标记总览' in wb.sheetnames
        assert '非工作日交易明细' in wb.sheetnames
        assert '节假日类型分布' in wb.sheetnames
        assert '主体分布' in wb.sheetnames

    def test_export_no_non_workday(self, tmp_dir):
        records = [
            _record('1', '622001', '2024-01-08', payment=100),
            _record('2', '622001', '2024-01-09', receipt=200),
        ]
        result = mark_non_workday_transactions(records)
        output_path = os.path.join(tmp_dir, 'holiday_report_empty.xlsx')
        export_holiday_check_result(result, output_path)

        assert os.path.exists(output_path)
        wb = openpyxl.load_workbook(output_path)
        assert '标记总览' in wb.sheetnames

    def test_export_with_source_info(self, tmp_dir):
        records = [
            _record('1', '622001', '2024-01-06', payment=100),
        ]
        result = mark_non_workday_transactions(records)
        output_path = os.path.join(tmp_dir, 'holiday_report_src.xlsx')
        source_info = {'数据来源': '单元测试', '记录数': 1}
        export_holiday_check_result(result, output_path, source_info)

        df = pd.read_excel(output_path, sheet_name='标记总览')
        assert '数据来源' in df['检测项'].values

    def test_overview_sheet_content(self, tmp_dir):
        records = [
            _record('1', '622001', '2024-01-08', payment=100),
            _record('2', '622001', '2024-01-06', payment=200),
            _record('3', '622001', '2024-01-01', payment=300),
        ]
        result = mark_non_workday_transactions(records)
        output_path = os.path.join(tmp_dir, 'holiday_overview.xlsx')
        export_holiday_check_result(result, output_path)

        df = pd.read_excel(output_path, sheet_name='标记总览')
        keys = df['检测项'].tolist()
        assert '总记录数' in keys
        assert '工作日交易数' in keys
        assert '周末交易数' in keys
        assert '法定节假日交易数' in keys
        assert '非工作日交易占比(%)' in keys

    def test_detail_sheet_content(self, tmp_dir):
        records = [
            _record('1', '622001', '2024-01-06', payment=200, subject='公司A'),
            _record('2', '622001', '2024-01-01', payment=300, subject='公司A'),
        ]
        result = mark_non_workday_transactions(records)
        output_path = os.path.join(tmp_dir, 'holiday_detail.xlsx')
        export_holiday_check_result(result, output_path)

        df = pd.read_excel(output_path, sheet_name='非工作日交易明细')
        assert len(df) == 2
        assert '日期类型' in df.columns
        assert '节假日名称' in df.columns
        assert '主体' in df.columns


class TestGenerateHolidayCheck:

    def test_from_records(self, tmp_dir):
        records = [
            _record('1', '622001', '2024-01-08', payment=100),
            _record('2', '622001', '2024-01-06', payment=200),
        ]
        result_path = generate_holiday_check_from_records(records, tmp_dir)
        assert result_path is not None
        assert os.path.exists(result_path)
        assert '非工作日交易标记报告' in os.path.basename(result_path)

    def test_from_records_empty(self):
        result_path = generate_holiday_check_from_records([], '/tmp')
        assert result_path is None

    def test_from_total(self, tmp_dir):
        records = [
            _record('1', '622001', '2024-01-08', payment=100),
            _record('2', '622001', '2024-01-06', payment=200),
        ]
        df = pd.DataFrame(records)
        total_path = os.path.join(tmp_dir, 'total.xlsx')
        df.to_excel(total_path, index=False, engine='openpyxl')

        result_path = generate_holiday_check_from_total(total_path, tmp_dir)
        assert result_path is not None
        assert os.path.exists(result_path)

    def test_from_total_empty(self, tmp_dir):
        df = pd.DataFrame()
        total_path = os.path.join(tmp_dir, 'empty_total.xlsx')
        df.to_excel(total_path, index=False, engine='openpyxl')

        result_path = generate_holiday_check_from_total(total_path, tmp_dir)
        assert result_path is None

    def test_from_total_nonexistent(self):
        result_path = generate_holiday_check_from_total('/nonexistent/path.xlsx')
        assert result_path is None


class TestHolidayDataConsistency:

    def test_adjusted_workdays_are_weekends(self):
        for date_str in CHINESE_WORKDAY_ADJUSTMENTS:
            dt = datetime.strptime(date_str, '%Y-%m-%d')
            assert dt.weekday() >= 5, f'调休工作日 {date_str} 应为周末, 实为周{dt.weekday() + 1}'

    def test_holidays_not_in_adjusted_workdays(self):
        overlap = set(CHINESE_HOLIDAYS.keys()) & CHINESE_WORKDAY_ADJUSTMENTS
        assert len(overlap) == 0, f'节假日与调休工作日存在重叠: {overlap}'

    def test_classify_date_consistency_with_data(self):
        for date_str, name in CHINESE_HOLIDAYS.items():
            tag, holiday_name = classify_date(date_str)
            assert tag == HOLIDAY_TAG_HOLIDAY, f'{date_str} 应为法定节假日, 实为 {tag}'
            assert holiday_name == name

        for date_str in CHINESE_WORKDAY_ADJUSTMENTS:
            tag, _ = classify_date(date_str)
            assert tag == HOLIDAY_TAG_ADJUSTED_WORKDAY, f'{date_str} 应为调休工作日, 实为 {tag}'


class TestMultiSubjectScenario:

    def test_multiple_subjects_banks(self):
        records = [
            _record('1', '622001', '2024-01-06', payment=100, subject='公司A', bank='北京银行'),
            _record('2', '622002', '2024-01-07', payment=200, subject='公司A', bank='工商银行'),
            _record('3', '622003', '2024-01-01', payment=300, subject='公司B', bank='北京银行'),
            _record('4', '622004', '2024-01-08', receipt=400, subject='公司B', bank='建设银行'),
        ]
        result = mark_non_workday_transactions(records)
        assert result.total_records == 4
        assert len(result.non_workday_records) == 3
        assert '公司A' in result.subject_stats
        assert '公司B' in result.subject_stats
        assert result.subject_stats['公司A'][HOLIDAY_TAG_WEEKEND] == 2
        assert result.subject_stats['公司B'][HOLIDAY_TAG_HOLIDAY] == 1
