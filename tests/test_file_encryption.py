import os
import sys
import shutil
import tempfile
import json

import openpyxl
import pandas as pd
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import file_encryption
from conftest import _create_beijing_bank_excel, _create_lookup_table
import bankcheck


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='encryption_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


@pytest.fixture
def sample_excel(tmp_dir):
    path = os.path.join(tmp_dir, 'test_output.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = '银行'
    ws['B1'] = '金额'
    ws['A2'] = '北京银行'
    ws['B2'] = 50000
    ws['A3'] = '建设银行'
    ws['B3'] = 80000
    wb.save(path)
    wb.close()
    return path


@pytest.fixture
def sample_text_file(tmp_dir):
    path = os.path.join(tmp_dir, 'test_report.txt')
    with open(path, 'w', encoding='utf-8') as f:
        f.write('检验报告内容\n余额连续性校验通过\n')
    return path


class TestPasswordValidation:
    def test_empty_password_fails(self):
        valid, msg = file_encryption.validate_password_strength('')
        assert valid is False
        assert '空' in msg

    def test_short_password_fails(self):
        valid, msg = file_encryption.validate_password_strength('ab1')
        assert valid is False
        assert '6' in msg

    def test_long_password_fails(self):
        valid, msg = file_encryption.validate_password_strength('a' * 129 + '1')
        assert valid is False
        assert '128' in msg

    def test_letters_only_fails(self):
        valid, msg = file_encryption.validate_password_strength('abcdef')
        assert valid is False
        assert '数字' in msg

    def test_digits_only_fails(self):
        valid, msg = file_encryption.validate_password_strength('123456')
        assert valid is False
        assert '字母' in msg

    def test_valid_password_passes(self):
        valid, msg = file_encryption.validate_password_strength('abc123')
        assert valid is True
        assert msg == ''

    def test_strong_password_passes(self):
        valid, msg = file_encryption.validate_password_strength('MyP@ss2024!')
        assert valid is True

    def test_chinese_password_passes(self):
        valid, msg = file_encryption.validate_password_strength('密码test123')
        assert valid is True


class TestAESEncryption:
    def test_encrypt_file_basic(self, sample_excel):
        result = file_encryption.encrypt_file_aes(sample_excel, 'test123')

        assert result.success is True
        assert result.mode == 'aes256gcm'
        assert result.encrypted_path == sample_excel + '.enc'
        assert os.path.exists(result.encrypted_path)
        assert result.encrypted_size > 0
        assert result.original_size > 0
        assert result.file_hash is not None
        assert result.timestamp is not None

    def test_encrypt_file_custom_output(self, sample_excel, tmp_dir):
        output = os.path.join(tmp_dir, 'custom_output.enc')
        result = file_encryption.encrypt_file_aes(sample_excel, 'test123', output_path=output)

        assert result.success is True
        assert result.encrypted_path == output
        assert os.path.exists(output)

    def test_encrypt_nonexistent_file(self):
        result = file_encryption.encrypt_file_aes('/nonexistent/file.xlsx', 'test123')
        assert result.success is False
        assert '不存在' in result.error

    def test_encrypt_weak_password(self, sample_excel):
        result = file_encryption.encrypt_file_aes(sample_excel, 'abc')
        assert result.success is False
        assert result.error is not None

    def test_encrypt_text_file(self, sample_text_file):
        result = file_encryption.encrypt_file_aes(sample_text_file, 'pass123')

        assert result.success is True
        assert os.path.exists(sample_text_file + '.enc')

    def test_encrypted_file_has_marker(self, sample_excel):
        file_encryption.encrypt_file_aes(sample_excel, 'test123')

        with open(sample_excel + '.enc', 'rb') as f:
            header = f.read(len(file_encryption.ENCRYPTION_MARKER))

        assert header == file_encryption.ENCRYPTION_MARKER

    def test_encrypted_file_larger_than_original(self, sample_excel):
        original_size = os.path.getsize(sample_excel)
        result = file_encryption.encrypt_file_aes(sample_excel, 'test123')

        assert result.encrypted_size > original_size


class TestAESDecryption:
    def test_decrypt_roundtrip(self, sample_excel):
        file_encryption.encrypt_file_aes(sample_excel, 'test123')

        enc_path = sample_excel + '.enc'
        result = file_encryption.decrypt_file_aes(enc_path, 'test123')

        assert result.success is True
        assert result.encrypted_path == sample_excel

        original_data = open(sample_excel, 'rb').read()
        decrypted_data = open(result.encrypted_path, 'rb').read()
        assert original_data == decrypted_data

    def test_decrypt_with_wrong_password(self, sample_excel):
        file_encryption.encrypt_file_aes(sample_excel, 'test123')

        enc_path = sample_excel + '.enc'
        result = file_encryption.decrypt_file_aes(enc_path, 'wrong456')

        assert result.success is False
        assert '密码错误' in result.error or '错误' in result.error

    def test_decrypt_non_encrypted_file(self, sample_excel):
        result = file_encryption.decrypt_file_aes(sample_excel, 'test123')
        assert result.success is False
        assert '格式' in result.error or '非本工具' in result.error

    def test_decrypt_nonexistent_file(self):
        result = file_encryption.decrypt_file_aes('/nonexistent/file.enc', 'test123')
        assert result.success is False

    def test_decrypt_custom_output(self, sample_excel, tmp_dir):
        file_encryption.encrypt_file_aes(sample_excel, 'test123')

        output = os.path.join(tmp_dir, 'decrypted.xlsx')
        result = file_encryption.decrypt_file_aes(
            sample_excel + '.enc', 'test123', output_path=output
        )

        assert result.success is True
        assert result.encrypted_path == output
        assert os.path.exists(output)

    def test_decrypt_auto_output_path(self, tmp_dir):
        src = os.path.join(tmp_dir, 'report.txt')
        with open(src, 'w') as f:
            f.write('sensitive data')

        file_encryption.encrypt_file_aes(src, 'pass123')

        enc_path = src + '.enc'
        result = file_encryption.decrypt_file_aes(enc_path, 'pass123')

        assert result.success is True
        assert result.encrypted_path == src

        with open(src, 'r') as f:
            assert f.read() == 'sensitive data'

    def test_encrypt_decrypt_excel_integrity(self, tmp_dir):
        path = os.path.join(tmp_dir, 'financial.xlsx')
        df = pd.DataFrame({
            '银行': ['北京银行', '建设银行'],
            '金额': [50000, 80000],
            '摘要': ['采购付款', '销售收款'],
        })
        df.to_excel(path, index=False, engine='openpyxl')

        file_encryption.encrypt_file_aes(path, 'Secure123')

        enc_path = path + '.enc'
        dec_result = file_encryption.decrypt_file_aes(enc_path, 'Secure123')

        assert dec_result.success is True

        df_dec = pd.read_excel(dec_result.encrypted_path, engine='openpyxl')
        assert len(df_dec) == 2
        assert df_dec['银行'].tolist() == ['北京银行', '建设银行']
        assert df_dec['金额'].tolist() == [50000, 80000]


class TestIsEncryptedFile:
    def test_encrypted_file_detected(self, sample_excel):
        file_encryption.encrypt_file_aes(sample_excel, 'test123')

        assert file_encryption.is_encrypted_file(sample_excel + '.enc') is True

    def test_plain_file_not_detected(self, sample_excel):
        assert file_encryption.is_encrypted_file(sample_excel) is False

    def test_nonexistent_file(self):
        assert file_encryption.is_encrypted_file('/nonexistent') is False


class TestGetEncryptionInfo:
    def test_plain_excel_info(self, sample_excel):
        info = file_encryption.get_encryption_info(sample_excel)

        assert info['exists'] is True
        assert info['encrypted'] is False
        assert info['file_hash'] is not None
        assert info['file_size'] > 0

    def test_aes_encrypted_info(self, sample_excel):
        file_encryption.encrypt_file_aes(sample_excel, 'test123')

        info = file_encryption.get_encryption_info(sample_excel + '.enc')
        assert info['exists'] is True
        assert info['encrypted'] is True
        assert info['encryption_mode'] == 'aes256gcm'

    def test_nonexistent_file_info(self):
        info = file_encryption.get_encryption_info('/nonexistent')
        assert info['exists'] is False


class TestBatchEncryption:
    def test_batch_encrypt_excel_password_mode(self, tmp_dir):
        files = []
        for i in range(3):
            path = os.path.join(tmp_dir, f'report_{i}.xlsx')
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'] = f'Report {i}'
            wb.save(path)
            wb.close()
            files.append(path)

        result = file_encryption.encrypt_output_files(
            files, password='Test123', mode='aes256gcm'
        )

        assert result.total_files == 3
        assert result.success_count == 3
        assert result.failure_count == 0
        assert len(result.results) == 3

        for r in result.results:
            assert r.success is True
            assert r.encrypted_path is not None
            assert os.path.exists(r.encrypted_path)

    def test_batch_encrypt_with_output_dir(self, tmp_dir):
        files = []
        for i in range(2):
            path = os.path.join(tmp_dir, f'data_{i}.xlsx')
            pd.DataFrame({'A': [i]}).to_excel(path, index=False)
            files.append(path)

        output_dir = os.path.join(tmp_dir, 'encrypted')
        result = file_encryption.encrypt_output_files(
            files, password='Test123', mode='aes256gcm',
            output_dir=output_dir,
        )

        assert result.success_count == 2
        assert os.path.isdir(output_dir)

    def test_batch_encrypt_with_nonexistent_file(self, tmp_dir):
        path = os.path.join(tmp_dir, 'real.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(path, index=False)

        result = file_encryption.encrypt_output_files(
            [path, '/nonexistent/file.xlsx'],
            password='Test123',
            mode='aes256gcm',
        )

        assert result.total_files == 2
        assert result.success_count == 1
        assert result.failure_count == 1

    def test_batch_encrypt_weak_password(self, tmp_dir):
        path = os.path.join(tmp_dir, 'data.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(path, index=False)

        result = file_encryption.encrypt_output_files(
            [path], password='abc', mode='aes256gcm',
        )

        assert result.failure_count >= 1

    def test_batch_encrypt_unsupported_mode(self, tmp_dir):
        path = os.path.join(tmp_dir, 'data.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(path, index=False)

        result = file_encryption.encrypt_output_files(
            [path], password='Test123', mode='unknown_mode',
        )

        assert result.failure_count == 1

    def test_batch_result_to_dict(self, tmp_dir):
        path = os.path.join(tmp_dir, 'data.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(path, index=False)

        result = file_encryption.encrypt_output_files(
            [path], password='Test123', mode='aes256gcm',
        )

        d = result.to_dict()
        assert 'results' in d
        assert 'total_files' in d
        assert 'success_count' in d
        assert 'timestamp' in d


class TestEncryptionResultDataclass:
    def test_encryption_result_defaults(self):
        r = file_encryption.EncryptionResult(file_path='/test.xlsx')
        assert r.encrypted_path is None
        assert r.mode == ''
        assert r.success is False
        assert r.error is None
        assert r.original_size == 0
        assert r.encrypted_size == 0

    def test_encryption_result_to_dict(self):
        r = file_encryption.EncryptionResult(
            file_path='/test.xlsx',
            encrypted_path='/test.xlsx.enc',
            mode='aes256gcm',
            success=True,
            original_size=100,
            encrypted_size=150,
        )
        d = r.to_dict()
        assert d['file_path'] == '/test.xlsx'
        assert d['success'] is True
        assert d['mode'] == 'aes256gcm'

    def test_batch_result_defaults(self):
        r = file_encryption.BatchEncryptionResult()
        assert r.results == []
        assert r.total_files == 0
        assert r.success_count == 0
        assert r.failure_count == 0


class TestSaveEncryptionRecord:
    def test_save_record_creates_file(self, tmp_dir):
        path = os.path.join(tmp_dir, 'data.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(path, index=False)

        enc_result = file_encryption.encrypt_file_aes(path, 'Test123')
        batch = file_encryption.BatchEncryptionResult(
            results=[enc_result],
            total_files=1,
            success_count=1,
            failure_count=0,
            mode='aes256gcm',
            timestamp='2024-01-01 12:00:00',
        )

        log_path = file_encryption.save_encryption_record(batch, script_dir=tmp_dir)
        assert log_path is not None
        assert os.path.exists(log_path)
        assert log_path.endswith('.json')

        with open(log_path, 'r', encoding='utf-8') as f:
            log_data = json.load(f)

        assert log_data['mode'] == 'aes256gcm'
        assert log_data['total_files'] == 1
        assert log_data['success_count'] == 1
        assert len(log_data['results']) == 1

    def test_save_record_empty_results(self, tmp_dir):
        batch = file_encryption.BatchEncryptionResult()
        log_path = file_encryption.save_encryption_record(batch, script_dir=tmp_dir)
        assert log_path is None


class TestExcelPasswordEncryption:
    @pytest.mark.skipif(not file_encryption.HAS_MSOFFCRYPTO,
                        reason='msoffcrypto-python not installed')
    def test_excel_password_encrypt(self, sample_excel):
        result = file_encryption.encrypt_excel_with_password(
            sample_excel, 'Test123'
        )

        assert result.mode == 'excel_password'
        if result.success:
            assert result.encrypted_path is not None
            assert os.path.exists(result.encrypted_path)
            assert '加密版' in result.encrypted_path

    @pytest.mark.skipif(not file_encryption.HAS_MSOFFCRYPTO,
                        reason='msoffcrypto-python not installed')
    def test_excel_password_custom_output(self, sample_excel, tmp_dir):
        output = os.path.join(tmp_dir, 'protected.xlsx')
        result = file_encryption.encrypt_excel_with_password(
            sample_excel, 'Test123', output_path=output
        )

        if result.success:
            assert result.encrypted_path == output

    def test_excel_password_nonexistent_file(self):
        result = file_encryption.encrypt_excel_with_password(
            '/nonexistent.xlsx', 'Test123'
        )
        assert result.success is False

    def test_excel_password_non_excel_file(self, sample_text_file):
        result = file_encryption.encrypt_excel_with_password(
            sample_text_file, 'Test123'
        )
        assert result.success is False
        assert 'Excel' in result.error

    def test_excel_password_weak_password(self, sample_excel):
        result = file_encryption.encrypt_excel_with_password(
            sample_excel, 'abc'
        )
        assert result.success is False


class TestPipelineWithEncryption:
    def _setup_folder(self, tmp_dir, script_dir):
        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))
        return source_folder

    def test_pipeline_with_aes_encryption(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script_enc')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        result = bankcheck.run_pipeline(
            source, script_dir,
            enable_encryption=True,
            encryption_password='Secure123',
            encryption_mode='aes256gcm',
        )

        assert len(result.all_rows) > 0
        assert result.output_path is not None
        assert result.encryption_result is not None
        assert len(result.encrypted_files) > 0

        for enc_file in result.encrypted_files:
            assert os.path.exists(enc_file)
            assert file_encryption.is_encrypted_file(enc_file)

    def test_pipeline_without_encryption(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script_noenc')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        result = bankcheck.run_pipeline(source, script_dir)

        assert result.encryption_result is None
        assert result.encrypted_files == []

    def test_pipeline_encryption_no_password(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script_nopass')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        result = bankcheck.run_pipeline(
            source, script_dir,
            enable_encryption=True,
            encryption_password=None,
        )

        assert result.encryption_result is None
        assert result.encrypted_files == []

    def test_pipeline_encryption_weak_password(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script_weak')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        result = bankcheck.run_pipeline(
            source, script_dir,
            enable_encryption=True,
            encryption_password='abc',
            encryption_mode='aes256gcm',
        )

        assert result.encryption_result is None or result.encryption_result.failure_count > 0

    def test_pipeline_encrypted_files_are_decryptable(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script_dec')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        password = 'Secure123'
        result = bankcheck.run_pipeline(
            source, script_dir,
            enable_encryption=True,
            encryption_password=password,
            encryption_mode='aes256gcm',
        )

        assert len(result.encrypted_files) > 0

        for enc_file in result.encrypted_files:
            dec_result = file_encryption.decrypt_file_aes(enc_file, password)
            assert dec_result.success is True
            assert os.path.exists(dec_result.encrypted_path)

    def test_pipeline_with_encryption_result_fields(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script_fields')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        result = bankcheck.run_pipeline(
            source, script_dir,
            enable_encryption=True,
            encryption_password='Test1234',
            encryption_mode='aes256gcm',
        )

        assert hasattr(result, 'encryption_result')
        assert hasattr(result, 'encrypted_files')
        assert isinstance(result.encrypted_files, list)

    def test_format_result_message_includes_encryption(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script_msg')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        result = bankcheck.run_pipeline(
            source, script_dir,
            enable_encryption=True,
            encryption_password='Test1234',
            encryption_mode='aes256gcm',
        )

        msg = bankcheck.format_result_message(result)

        if result.encrypted_files:
            assert '加密' in msg


class TestPipelineWithOptionsEncryption:
    def _setup_folder(self, tmp_dir, script_dir):
        source_folder = os.path.join(tmp_dir, '流水文件夹_opt')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))
        return source_folder

    def test_pipeline_with_options_encryption(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script_opt_enc')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        result = bankcheck.run_pipeline_with_options(
            source, script_dir,
            enable_encryption=True,
            encryption_password='Secure123',
            encryption_mode='aes256gcm',
        )

        assert len(result.all_rows) > 0
        assert hasattr(result, 'encrypted_files')


class TestEncryptionLogIntegrity:
    def test_encryption_log_contains_all_info(self, tmp_dir):
        path = os.path.join(tmp_dir, 'financial.xlsx')
        pd.DataFrame({
            '银行': ['北京银行'],
            '金额': [50000],
        }).to_excel(path, index=False, engine='openpyxl')

        enc_result = file_encryption.encrypt_file_aes(path, 'Test123')
        batch = file_encryption.BatchEncryptionResult(
            results=[enc_result],
            total_files=1,
            success_count=1,
            failure_count=0,
            mode='aes256gcm',
            timestamp='2024-01-01 12:00:00',
        )

        log_path = file_encryption.save_encryption_record(batch, script_dir=tmp_dir)
        assert log_path is not None

        with open(log_path, 'r', encoding='utf-8') as f:
            log_data = json.load(f)

        result_entry = log_data['results'][0]
        assert result_entry['file_path'] == path
        assert result_entry['success'] is True
        assert result_entry['mode'] == 'aes256gcm'
        assert result_entry['original_size'] > 0
        assert result_entry['encrypted_size'] > 0
        assert result_entry['file_hash'] is not None
        assert result_entry['timestamp'] is not None


class TestEncryptionModuleConstants:
    def test_encryption_marker_is_bytes(self):
        assert isinstance(file_encryption.ENCRYPTION_MARKER, bytes)

    def test_encrypted_extension(self):
        assert file_encryption.ENCRYPTED_EXTENSION == '.enc'

    def test_aes_key_size(self):
        assert file_encryption.AES_KEY_SIZE == 32

    def test_salt_size(self):
        assert file_encryption.SALT_SIZE == 16

    def test_nonce_size(self):
        assert file_encryption.NONCE_SIZE == 12

    def test_pbkdf2_iterations(self):
        assert file_encryption.PBKDF2_ITERATIONS >= 100000


class TestMultipleEncryptionRounds:
    def test_encrypt_decrypt_multiple_times(self, sample_excel):
        password = 'Test123'

        for _ in range(3):
            enc_result = file_encryption.encrypt_file_aes(
                sample_excel, password,
                output_path=sample_excel + '.enc',
            )
            assert enc_result.success is True

            dec_result = file_encryption.decrypt_file_aes(
                sample_excel + '.enc', password,
                output_path=sample_excel + '.dec.xlsx',
            )
            assert dec_result.success is True

            original_hash = file_encryption._compute_file_hash(sample_excel)
            decrypted_hash = file_encryption._compute_file_hash(sample_excel + '.dec.xlsx')
            assert original_hash == decrypted_hash

            os.remove(sample_excel + '.enc')
            os.remove(sample_excel + '.dec.xlsx')


class TestEdgeCases:
    def test_encrypt_empty_file(self, tmp_dir):
        path = os.path.join(tmp_dir, 'empty.xlsx')
        with open(path, 'wb') as f:
            pass

        result = file_encryption.encrypt_file_aes(path, 'Test123')
        assert result.success is True

        dec_result = file_encryption.decrypt_file_aes(
            result.encrypted_path, 'Test123',
            output_path=path + '.dec',
        )
        assert dec_result.success is True

    def test_encrypt_file_with_chinese_name(self, tmp_dir):
        path = os.path.join(tmp_dir, '银行流水检验报告.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '中文内容测试'
        wb.save(path)
        wb.close()

        result = file_encryption.encrypt_file_aes(path, 'Test123')
        assert result.success is True

        dec_result = file_encryption.decrypt_file_aes(
            result.encrypted_path, 'Test123',
        )
        assert dec_result.success is True

    def test_password_with_special_characters(self, sample_excel):
        password = 'P@$$w0rd!#%^&*()'
        result = file_encryption.encrypt_file_aes(sample_excel, password)
        assert result.success is True

        dec_result = file_encryption.decrypt_file_aes(
            result.encrypted_path, password,
        )
        assert dec_result.success is True

    def test_password_with_unicode(self, sample_excel):
        password = '密码Test123'
        result = file_encryption.encrypt_file_aes(sample_excel, password)
        assert result.success is True

        dec_result = file_encryption.decrypt_file_aes(
            result.encrypted_path, password,
        )
        assert dec_result.success is True
