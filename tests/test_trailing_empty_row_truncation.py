import os
import sys
import tempfile

import openpyxl
import pytest

import bankcheck
from bankcheck import (
    is_row_empty,
    get_effective_max_row,
    iter_sheet_rows,
    iter_sheet_records,
    DEFAULT_EMPTY_ROW_THRESHOLD,
    open_workbook_compat,
)


class TestIsRowEmpty:
    def test_empty_row_all_columns(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'data'
        ws['A3'] = 'data'
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path)
        ws2 = wb2.active
        assert is_row_empty(ws2, 1) is False
        assert is_row_empty(ws2, 2) is True
        assert is_row_empty(ws2, 3) is False
        wb2.close()

    def test_empty_row_with_check_columns(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'data'
        ws['B2'] = 'data'
        ws['A3'] = 'data'
        ws['B3'] = 'data'
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path)
        ws2 = wb2.active
        assert is_row_empty(ws2, 1, check_columns=[1]) is False
        assert is_row_empty(ws2, 1, check_columns=[2]) is True
        assert is_row_empty(ws2, 2, check_columns=[1]) is True
        assert is_row_empty(ws2, 2, check_columns=[2]) is False
        assert is_row_empty(ws2, 2, check_columns=[1, 2]) is False
        wb2.close()

    def test_empty_row_whitespace_only(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '   '
        ws['A2'] = '\t\n'
        ws['A3'] = 'data'
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path)
        ws2 = wb2.active
        assert is_row_empty(ws2, 1) is True
        assert is_row_empty(ws2, 2) is True
        assert is_row_empty(ws2, 3) is False
        wb2.close()

    def test_empty_row_zero_and_empty_string(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 0
        ws['A2'] = ''
        ws['A3'] = None
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path)
        ws2 = wb2.active
        assert is_row_empty(ws2, 1) is False
        assert is_row_empty(ws2, 2) is True
        assert is_row_empty(ws2, 3) is True
        wb2.close()

    def test_empty_row_out_of_range(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'data'
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path)
        ws2 = wb2.active
        assert is_row_empty(ws2, 100) is True
        wb2.close()


class TestGetEffectiveMaxRow:
    def _create_sheet_with_trailing_empty(self, tmp_dir, data_rows, trailing_empty_rows):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, data_rows + 1):
            ws.cell(row=i, column=1, value=f'data_{i}')
        for i in range(data_rows + 1, data_rows + trailing_empty_rows + 1):
            ws.cell(row=i, column=1, value=None)
        wb.save(path)
        wb.close()
        return path

    def test_no_trailing_empty_rows(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 100, 0)
        wb, tmp_path = open_workbook_compat(path)
        ws = wb.active

        result = get_effective_max_row(ws, consecutive_empty_threshold=10)
        assert result == 100
        wb.close()

    def test_few_trailing_empty_rows_below_threshold(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 100, 5)
        wb, tmp_path = open_workbook_compat(path)
        ws = wb.active

        result = get_effective_max_row(ws, consecutive_empty_threshold=10)
        assert result == 100
        wb.close()

    def test_many_trailing_empty_rows_above_threshold(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 100, 1000)
        wb, tmp_path = open_workbook_compat(path)
        ws = wb.active

        result = get_effective_max_row(ws, consecutive_empty_threshold=50)
        assert result == 100
        wb.close()

    def test_with_start_row(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 100, 200)
        wb, tmp_path = open_workbook_compat(path)
        ws = wb.active

        result = get_effective_max_row(ws, start_row=50, consecutive_empty_threshold=50)
        assert result == 100
        wb.close()

    def test_with_end_row(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 100, 200)
        wb, tmp_path = open_workbook_compat(path)
        ws = wb.active

        result = get_effective_max_row(ws, end_row=150, consecutive_empty_threshold=10)
        assert result == 100
        wb.close()

    def test_with_check_columns(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 101):
            ws.cell(row=i, column=1, value=f'data_{i}')
            ws.cell(row=i, column=2, value=f'other_{i}')
        for i in range(101, 201):
            ws.cell(row=i, column=1, value=None)
            ws.cell(row=i, column=2, value=f'other_{i}')
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path)
        ws2 = wb2.active

        result_col1 = get_effective_max_row(ws2, check_columns=[1], consecutive_empty_threshold=10)
        assert result_col1 == 100

        result_col2 = get_effective_max_row(ws2, check_columns=[2], consecutive_empty_threshold=10)
        assert result_col2 == 200
        wb2.close()

    def test_empty_sheet(self, tmp_dir):
        path = os.path.join(tmp_dir, 'empty.xlsx')
        wb = openpyxl.Workbook()
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path)
        ws2 = wb2.active
        result = get_effective_max_row(ws2)
        assert result == 0
        wb2.close()

    def test_only_empty_rows(self, tmp_dir):
        path = os.path.join(tmp_dir, 'empty_rows.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 101):
            ws.cell(row=i, column=1, value=None)
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path)
        ws2 = wb2.active
        result = get_effective_max_row(ws2, consecutive_empty_threshold=10)
        assert result == 0
        wb2.close()

    def test_start_row_greater_than_end_row(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 100, 50)
        wb, tmp_path = open_workbook_compat(path)
        ws = wb.active

        result = get_effective_max_row(ws, start_row=200, end_row=100)
        assert result == 199
        wb.close()

    def test_single_data_row(self, tmp_dir):
        path = os.path.join(tmp_dir, 'single.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'only_data'
        for i in range(2, 101):
            ws.cell(row=i, column=1, value=None)
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path)
        ws2 = wb2.active
        result = get_effective_max_row(ws2, consecutive_empty_threshold=10)
        assert result == 1
        wb2.close()

    def test_data_interspersed_with_empty_rows(self, tmp_dir):
        path = os.path.join(tmp_dir, 'interspersed.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 51):
            ws.cell(row=i * 2 - 1, column=1, value=f'data_{i}')
            ws.cell(row=i * 2, column=1, value=None)
        for i in range(101, 201):
            ws.cell(row=i, column=1, value=None)
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path)
        ws2 = wb2.active
        result = get_effective_max_row(ws2, consecutive_empty_threshold=10)
        assert result == 99
        wb2.close()

    def test_threshold_zero_uses_default(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 50, 100)
        wb, tmp_path = open_workbook_compat(path)
        ws = wb.active

        result = get_effective_max_row(ws, consecutive_empty_threshold=0)
        assert result == 50
        wb.close()

    def test_very_large_trailing_empty(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 100, 5000)
        wb, tmp_path = open_workbook_compat(path)
        ws = wb.active

        result = get_effective_max_row(ws, consecutive_empty_threshold=50)
        assert result == 100
        wb.close()

    def test_default_threshold_value(self):
        assert DEFAULT_EMPTY_ROW_THRESHOLD == 50
        assert isinstance(DEFAULT_EMPTY_ROW_THRESHOLD, int)
        assert DEFAULT_EMPTY_ROW_THRESHOLD > 0


class TestIterSheetRowsTrimTrailingEmpty:
    def _create_sheet_with_trailing_empty(self, tmp_dir, data_rows, trailing_empty_rows):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, data_rows + 1):
            ws.cell(row=i, column=1, value=f'data_{i}')
            ws.cell(row=i, column=2, value=i * 10)
        for i in range(data_rows + 1, data_rows + trailing_empty_rows + 1):
            ws.cell(row=i, column=1, value=None)
            ws.cell(row=i, column=2, value=None)
        wb.save(path)
        wb.close()
        return path

    def test_trim_trailing_empty_enabled(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 50, 500)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws, trim_trailing_empty=True,
                                      consecutive_empty_threshold=20))
        total_rows = sum(len(chunk) for chunk in chunks)
        assert total_rows == 50
        wb.close()

    def test_trim_trailing_empty_disabled(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 50, 100)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws, trim_trailing_empty=False))
        total_rows = sum(len(chunk) for chunk in chunks)
        assert total_rows == 150
        wb.close()

    def test_trim_trailing_empty_with_check_columns(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 51):
            ws.cell(row=i, column=1, value=f'data_{i}')
            ws.cell(row=i, column=2, value=i * 10)
        for i in range(51, 151):
            ws.cell(row=i, column=1, value=None)
            ws.cell(row=i, column=2, value=f'other_{i}')
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path, read_only=True)
        ws2 = wb2.active

        chunks_col1 = list(iter_sheet_rows(
            ws2, trim_trailing_empty=True, check_columns=[1],
            consecutive_empty_threshold=20))
        total_rows_col1 = sum(len(chunk) for chunk in chunks_col1)
        assert total_rows_col1 == 50

        chunks_col2 = list(iter_sheet_rows(
            ws2, trim_trailing_empty=True, check_columns=[2],
            consecutive_empty_threshold=20))
        total_rows_col2 = sum(len(chunk) for chunk in chunks_col2)
        assert total_rows_col2 == 150

        wb2.close()

    def test_trim_trailing_empty_with_start_row(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 100, 200)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(
            ws, start_row=50, trim_trailing_empty=True,
            consecutive_empty_threshold=20))
        total_rows = sum(len(chunk) for chunk in chunks)
        assert total_rows == 51
        assert chunks[0][0][0] == 'data_50'
        wb.close()

    def test_trim_trailing_empty_with_end_row(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 100, 200)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(
            ws, end_row=80, trim_trailing_empty=True,
            consecutive_empty_threshold=20))
        total_rows = sum(len(chunk) for chunk in chunks)
        assert total_rows == 80
        wb.close()

    def test_trim_trailing_empty_empty_sheet(self, tmp_dir):
        path = os.path.join(tmp_dir, 'empty.xlsx')
        wb = openpyxl.Workbook()
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path, read_only=True)
        ws2 = wb2.active
        chunks = list(iter_sheet_rows(ws2, trim_trailing_empty=True))
        assert len(chunks) == 0
        wb2.close()


class TestIterSheetRecordsTrimTrailingEmpty:
    def _create_sheet_with_trailing_empty(self, tmp_dir, data_rows, trailing_empty_rows):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'name'
        ws['B1'] = 'value'
        for i in range(2, data_rows + 2):
            ws.cell(row=i, column=1, value=f'item_{i-1}')
            ws.cell(row=i, column=2, value=i * 10)
        for i in range(data_rows + 2, data_rows + trailing_empty_rows + 2):
            ws.cell(row=i, column=1, value=None)
            ws.cell(row=i, column=2, value=None)
        wb.save(path)
        wb.close()
        return path

    def test_trim_trailing_empty_enabled(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 50, 500)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        columns_map = {'name': 1, 'value': 2}
        chunks = list(iter_sheet_records(
            ws, columns_map, start_row=2,
            trim_trailing_empty=True, consecutive_empty_threshold=20))
        total_records = sum(len(chunk) for chunk in chunks)
        assert total_records == 50
        wb.close()

    def test_trim_trailing_empty_disabled(self, tmp_dir):
        path = self._create_sheet_with_trailing_empty(tmp_dir, 50, 100)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        columns_map = {'name': 1, 'value': 2}
        chunks = list(iter_sheet_records(
            ws, columns_map, start_row=2, trim_trailing_empty=False))
        total_records = sum(len(chunk) for chunk in chunks)
        assert total_records == 150
        wb.close()

    def test_trim_trailing_empty_with_check_columns(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'name'
        ws['B1'] = 'value'
        for i in range(2, 52):
            ws.cell(row=i, column=1, value=f'item_{i-1}')
            ws.cell(row=i, column=2, value=i * 10)
        for i in range(52, 152):
            ws.cell(row=i, column=1, value=None)
            ws.cell(row=i, column=2, value=i * 10)
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path, read_only=True)
        ws2 = wb2.active

        columns_map = {'name': 1, 'value': 2}
        chunks_col1 = list(iter_sheet_records(
            ws2, columns_map, start_row=2,
            trim_trailing_empty=True, check_columns=[1],
            consecutive_empty_threshold=20))
        total_col1 = sum(len(chunk) for chunk in chunks_col1)
        assert total_col1 == 50

        chunks_col2 = list(iter_sheet_records(
            ws2, columns_map, start_row=2,
            trim_trailing_empty=True, check_columns=[2],
            consecutive_empty_threshold=20))
        total_col2 = sum(len(chunk) for chunk in chunks_col2)
        assert total_col2 == 150

        wb2.close()


class TestBankProcessorsWithTrailingEmpty:
    def test_parse_sheet_with_trailing_empty_rows(self, tmp_dir):
        from bankcheck import BANK_PROCESSORS
        from conftest import _create_bank_excel, _create_lookup_table_all_banks

        data_rows = [
            [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A公司', '622001234', '工商银行', '转账', '001', '采购付款', None, None, None, 'BJ20240105001'],
            [2, '2024-01-10', 'CNY', None, 80000, 1580000, '客户B公司', '622005678', '建设银行', '转账', '002', '销售收款', None, None, None, 'BJ20240110002'],
        ]
        for _ in range(1000):
            data_rows.append([None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None])

        path = os.path.join(tmp_dir, 'trailing_empty_test.xlsx')
        _create_bank_excel(path, '北京银行', rows=data_rows)
        lookup = _create_lookup_table_all_banks(os.path.join(tmp_dir, 'lookup.xlsx'))

        processor = BANK_PROCESSORS['北京银行']
        records = processor(path, lookup)

        assert len(records) == 2
        assert records[0]['交易日期'] == '2024-01-05'
        assert records[1]['交易日期'] == '2024-01-10'

    def test_parse_sheet_with_lots_of_trailing_empty(self, tmp_dir):
        from bankcheck import BANK_PROCESSORS
        from conftest import _create_bank_excel, _create_lookup_table_all_banks

        data_rows = [
            [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A公司', '622001234', '工商银行', '转账', '001', '采购付款', None, None, None, 'BJ20240105001'],
        ]
        for _ in range(5000):
            data_rows.append([None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None])

        path = os.path.join(tmp_dir, 'lots_of_empty.xlsx')
        _create_bank_excel(path, '北京银行', rows=data_rows)
        lookup = _create_lookup_table_all_banks(os.path.join(tmp_dir, 'lookup.xlsx'))

        processor = BANK_PROCESSORS['北京银行']
        records = processor(path, lookup)

        assert len(records) == 1
        assert records[0]['交易流水号'] == 'BJ20240105001'

    def test_parse_sheet_no_data_only_empty_rows(self, tmp_dir):
        from bankcheck import BANK_PROCESSORS
        from conftest import _create_bank_excel, _create_lookup_table_all_banks

        data_rows = []
        for _ in range(100):
            data_rows.append([None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None])

        path = os.path.join(tmp_dir, 'no_data.xlsx')
        _create_bank_excel(path, '北京银行', rows=data_rows)
        lookup = _create_lookup_table_all_banks(os.path.join(tmp_dir, 'lookup.xlsx'))

        processor = BANK_PROCESSORS['北京银行']
        records = processor(path, lookup)

        assert len(records) == 0
