import os
import sys
import shutil
import tempfile

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck
from bankcheck import BankRule, GenericBankParser, HeaderValidationError


@pytest.fixture(autouse=True)
def init_logging():
    bankcheck.setup_logging()


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='bankcheck_header_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _make_rule(columns=None, expected_headers=None, header_validation='warn',
               start_row=4, bank_name='测试银行', account_cell='B2'):
    if columns is None:
        columns = {
            'trade_date': 2,
            'payment': 4,
            'receipt': 5,
            'summary': 12,
            'counterpart': 7,
            'balance': 6,
            'transaction_id': 16,
        }
    return BankRule(
        bank_name=bank_name,
        account_cell=account_cell,
        start_row=start_row,
        columns=columns,
        expected_headers=expected_headers or {},
        header_validation=header_validation,
    )


def _create_excel_with_headers(path, headers, account='01090312345678901',
                               start_row=4, rows=None):
    if rows is None:
        rows = [
            [1, '2024-01-05', 'CNY', 50000, None, 1500000,
             '供应商A', '622001', '工商银行', '转账', '001',
             '采购付款', None, None, None, 'BJ20240105001'],
        ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '交易明细'
    ws['A1'] = '测试银行交易明细'
    ws['B2'] = account
    header_row = start_row - 1
    for c, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=c, value=h)
    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=start_row + i, column=j + 1, value=val)
    wb.save(path)
    wb.close()
    return path


BEIJING_HEADERS = [
    '序号', '交易日期', '币种', '支出金额', '收入金额', '余额',
    '对方户名', '对方账号', '对方行名', '凭证种类', '凭证号码',
    '摘要', '备注1', '备注2', '备注3', '交易流水号',
]

BEIJING_EXPECTED = {
    'trade_date': ['交易日期'],
    'payment': ['支出金额'],
    'receipt': ['收入金额'],
    'summary': ['摘要'],
    'counterpart': ['对方户名'],
    'balance': ['余额'],
    'transaction_id': ['交易流水号'],
}


class TestValidateHeadersMatch:
    def test_all_headers_match(self, tmp_dir):
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), BEIJING_HEADERS)
        rule = _make_rule(expected_headers=BEIJING_EXPECTED)
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert mismatches == []

    def test_partial_expected_headers_match(self, tmp_dir):
        expected = {
            'trade_date': ['交易日期'],
            'payment': ['支出金额'],
        }
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), BEIJING_HEADERS)
        rule = _make_rule(expected_headers=expected)
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert mismatches == []


class TestValidateHeadersMismatch:
    def test_single_column_mismatch(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '付款金额'
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(expected_headers=BEIJING_EXPECTED)
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert 'payment' in mismatches
        assert len(mismatches) == 1

    def test_multiple_columns_mismatch(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '付款金额'
        modified_headers[4] = '收款金额'
        modified_headers[5] = '账户余额'
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(expected_headers=BEIJING_EXPECTED)
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert 'payment' in mismatches
        assert 'receipt' in mismatches
        assert 'balance' in mismatches

    def test_empty_header_cell(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = None
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(expected_headers=BEIJING_EXPECTED)
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert 'payment' in mismatches

    def test_whitespace_in_header(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[1] = ' 交易日期 '
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(expected_headers=BEIJING_EXPECTED)
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert mismatches == []


class TestValidateHeadersAcceptableNames:
    def test_multiple_acceptable_names_first_match(self, tmp_dir):
        expected = {
            'trade_date': ['交易日期', '日期'],
            'payment': ['支出金额', '付款金额', '借方金额'],
        }
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), BEIJING_HEADERS)
        rule = _make_rule(expected_headers=expected)
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert mismatches == []

    def test_multiple_acceptable_names_second_match(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '付款金额'
        expected = {
            'payment': ['支出金额', '付款金额', '借方金额'],
        }
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(expected_headers=expected)
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert mismatches == []

    def test_acceptable_names_no_match(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '借方发生额'
        expected = {
            'payment': ['支出金额', '付款金额', '借方金额'],
        }
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(expected_headers=expected)
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert 'payment' in mismatches


class TestValidateHeadersOffMode:
    def test_off_mode_skips_validation(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '完全不同的表头'
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(
            expected_headers=BEIJING_EXPECTED, header_validation='off')
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert mismatches == []

    def test_no_expected_headers_skips_validation(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '完全不同的表头'
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(expected_headers={}, header_validation='warn')
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert mismatches == []


class TestHeaderValidationStrictMode:
    def test_strict_mode_raises_on_mismatch(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '付款金额'
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(
            expected_headers=BEIJING_EXPECTED, header_validation='strict')
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        with pytest.raises(HeaderValidationError, match='payment'):
            parser._parse_sheet(ws, path, ws.title, None)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)

    def test_strict_mode_passes_on_match(self, tmp_dir):
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), BEIJING_HEADERS)
        rule = _make_rule(
            expected_headers=BEIJING_EXPECTED, header_validation='strict')
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        rows = parser._parse_sheet(ws, path, ws.title, None)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert len(rows) >= 1

    def test_strict_mode_error_message_contains_details(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '错误列名'
        modified_headers[5] = '错误余额'
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(
            expected_headers=BEIJING_EXPECTED, header_validation='strict')
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        with pytest.raises(HeaderValidationError) as exc_info:
            parser._parse_sheet(ws, path, ws.title, None)
        msg = str(exc_info.value)
        assert 'payment' in msg
        assert 'balance' in msg
        wb.close()
        bankcheck.cleanup_temp_file(tmp)


class TestHeaderValidationWarnMode:
    def test_warn_mode_continues_on_mismatch(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '付款金额'
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(
            expected_headers=BEIJING_EXPECTED, header_validation='warn')
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        rows = parser._parse_sheet(ws, path, ws.title, None)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert len(rows) >= 1

    def test_warn_mode_returns_mismatch_list(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '付款金额'
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(
            expected_headers=BEIJING_EXPECTED, header_validation='warn')
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert 'payment' in mismatches


class TestHeaderValidationEastAsiaBank:
    EAST_ASIA_COLUMNS = {
        'trade_date': 1,
        'payment': 4,
        'receipt': 5,
        'summary': 12,
        'counterpart': 12,
        'balance': 9,
        'transaction_id': 11,
    }

    EAST_ASIA_EXPECTED = {
        'trade_date': ['交易日期'],
        'payment': ['支出金额'],
        'receipt': ['收入金额'],
        'summary': ['摘要', '交易描述/对方'],
        'counterpart': ['对方户名', '交易描述/对方'],
        'balance': ['余额'],
        'transaction_id': ['交易流水号'],
    }

    EAST_ASIA_HEADERS = [
        '交易日期', '交易时间', '币种', '支出金额', '收入金额',
        '手续费', '利息', '税费', '余额', '交易类型',
        '交易流水号', '交易描述/对方',
    ]

    def test_east_asia_headers_match(self, tmp_dir):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '账号'
        ws['B1'] = '38812345678'
        ws['A2'] = '东亚银行交易明细'
        ws['A3'] = '查询期间'
        for c, h in enumerate(self.EAST_ASIA_HEADERS, 1):
            ws.cell(row=4, column=c, value=h)
        ws.cell(row=5, column=1, value='2024-01-03')
        ws.cell(row=5, column=4, value=20000)
        ws.cell(row=5, column=9, value=480000)
        ws.cell(row=5, column=11, value='EA001')
        ws.cell(row=5, column=12, value='付款')
        path = os.path.join(tmp_dir, 'east_asia.xlsx')
        wb.save(path)
        wb.close()

        rule = _make_rule(
            columns=self.EAST_ASIA_COLUMNS,
            expected_headers=self.EAST_ASIA_EXPECTED,
            start_row=5,
            account_cell='B1',
            bank_name='东亚银行',
        )
        parser = GenericBankParser(rule)
        wb2, tmp = bankcheck.open_workbook_compat(path)
        ws2 = wb2.active
        mismatches = parser.validate_headers(ws2, path, ws2.title)
        wb2.close()
        bankcheck.cleanup_temp_file(tmp)
        assert mismatches == []

    def test_east_asia_shared_column_acceptable_names(self, tmp_dir):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '账号'
        ws['B1'] = '38812345678'
        modified_headers = list(self.EAST_ASIA_HEADERS)
        modified_headers[11] = '摘要'
        for c, h in enumerate(modified_headers, 1):
            ws.cell(row=4, column=c, value=h)
        ws.cell(row=5, column=1, value='2024-01-03')
        ws.cell(row=5, column=4, value=20000)
        ws.cell(row=5, column=9, value=480000)
        ws.cell(row=5, column=11, value='EA001')
        ws.cell(row=5, column=12, value='付款')
        path = os.path.join(tmp_dir, 'east_asia_summary.xlsx')
        wb.save(path)
        wb.close()

        rule = _make_rule(
            columns=self.EAST_ASIA_COLUMNS,
            expected_headers=self.EAST_ASIA_EXPECTED,
            start_row=5,
            account_cell='B1',
            bank_name='东亚银行',
        )
        parser = GenericBankParser(rule)
        wb2, tmp = bankcheck.open_workbook_compat(path)
        ws2 = wb2.active
        mismatches = parser.validate_headers(ws2, path, ws2.title)
        wb2.close()
        bankcheck.cleanup_temp_file(tmp)
        assert 'counterpart' in mismatches
        assert 'summary' not in mismatches


class TestHeaderValidationIntegration:
    def test_parse_with_correct_headers(self, tmp_dir):
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), BEIJING_HEADERS)
        rule = _make_rule(
            expected_headers=BEIJING_EXPECTED, header_validation='warn')
        parser = GenericBankParser(rule)
        rows = parser.parse(path, None)
        assert len(rows) >= 1

    def test_parse_with_wrong_headers_warn_mode(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '错误列名'
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(
            expected_headers=BEIJING_EXPECTED, header_validation='warn')
        parser = GenericBankParser(rule)
        rows = parser.parse(path, None)
        assert len(rows) >= 1

    def test_parse_with_wrong_headers_strict_mode(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '错误列名'
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(
            expected_headers=BEIJING_EXPECTED, header_validation='strict')
        parser = GenericBankParser(rule)
        with pytest.raises(HeaderValidationError):
            parser.parse(path, None)

    def test_parse_without_expected_headers(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = '错误列名'
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(expected_headers={}, header_validation='warn')
        parser = GenericBankParser(rule)
        rows = parser.parse(path, None)
        assert len(rows) >= 1


class TestHeaderValidationEdgeCases:
    def test_start_row_at_1(self, tmp_dir):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '交易日期'
        ws['B1'] = '支出金额'
        ws['A2'] = '2024-01-05'
        ws['B2'] = 50000
        path = os.path.join(tmp_dir, 'edge.xlsx')
        wb.save(path)
        wb.close()

        columns = {'trade_date': 1, 'payment': 2}
        expected = {'trade_date': ['交易日期'], 'payment': ['支出金额']}
        rule = _make_rule(
            columns=columns, expected_headers=expected,
            start_row=2, account_cell='A1')
        parser = GenericBankParser(rule)
        wb2, tmp = bankcheck.open_workbook_compat(path)
        ws2 = wb2.active
        mismatches = parser.validate_headers(ws2, path, ws2.title)
        wb2.close()
        bankcheck.cleanup_temp_file(tmp)
        assert mismatches == []

    def test_numeric_header_value(self, tmp_dir):
        modified_headers = list(BEIJING_HEADERS)
        modified_headers[3] = 12345
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), modified_headers)
        rule = _make_rule(expected_headers=BEIJING_EXPECTED)
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert 'payment' in mismatches

    def test_expected_header_key_not_in_columns(self, tmp_dir):
        expected = {
            'trade_date': ['交易日期'],
            'nonexistent_column': ['不存在'],
        }
        path = _create_excel_with_headers(
            os.path.join(tmp_dir, 'test.xlsx'), BEIJING_HEADERS)
        rule = _make_rule(expected_headers=expected)
        parser = GenericBankParser(rule)
        wb, tmp = bankcheck.open_workbook_compat(path)
        ws = wb.active
        mismatches = parser.validate_headers(ws, path, ws.title)
        wb.close()
        bankcheck.cleanup_temp_file(tmp)
        assert 'nonexistent_column' not in mismatches
        assert 'trade_date' not in mismatches

    def test_strict_mode_with_empty_sheet(self, tmp_dir):
        wb = openpyxl.Workbook()
        ws = wb.active
        path = os.path.join(tmp_dir, 'empty.xlsx')
        wb.save(path)
        wb.close()

        rule = _make_rule(
            expected_headers=BEIJING_EXPECTED, header_validation='strict')
        parser = GenericBankParser(rule)
        with pytest.raises(HeaderValidationError):
            parser.parse(path, None)


class TestBankRuleDataclass:
    def test_default_expected_headers_empty(self):
        rule = BankRule(
            bank_name='测试', account_cell='A1', start_row=2,
            columns={'trade_date': 1})
        assert rule.expected_headers == {}
        assert rule.header_validation == 'warn'

    def test_custom_expected_headers(self):
        rule = BankRule(
            bank_name='测试', account_cell='A1', start_row=2,
            columns={'trade_date': 1},
            expected_headers={'trade_date': ['日期', '交易日期']},
            header_validation='strict')
        assert rule.expected_headers == {'trade_date': ['日期', '交易日期']}
        assert rule.header_validation == 'strict'


class TestHeaderValidationConfigLoading:
    def test_load_config_with_expected_headers(self):
        import tempfile
        import yaml
        config_data = {
            'banks': [{
                'bank_name': '配置测试银行',
                'account_cell': 'A1',
                'start_row': 2,
                'columns': {'trade_date': 1, 'payment': 2},
                'expected_headers': {
                    'trade_date': ['交易日期'],
                    'payment': '支出金额',
                },
                'header_validation': 'strict',
                'enabled': True,
            }]
        }
        fd, config_path = tempfile.mkstemp(suffix='.yaml', prefix='bank_rules_test_')
        os.close(fd)
        try:
            with open(config_path, 'w', encoding='utf-8') as f:
                yaml.dump(config_data, f, allow_unicode=True)

            bankcheck.BankRuleConfig._instance = None
            config = bankcheck.BankRuleConfig(config_path=config_path)
            rule = config.get_rule('配置测试银行')
            assert rule is not None
            assert rule.expected_headers == {
                'trade_date': ['交易日期'],
                'payment': ['支出金额'],
            }
            assert rule.header_validation == 'strict'
        finally:
            bankcheck.BankRuleConfig._instance = None
            os.unlink(config_path)
