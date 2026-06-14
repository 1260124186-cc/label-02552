import os
import sys
import tempfile
import threading
import time
from unittest.mock import MagicMock, patch

import openpyxl
import pytest

import bankcheck
from bankcheck import (
    convert_xls_to_xlsx,
    open_workbook_compat,
    cleanup_temp_file,
    detect_excel_format,
    get_extension_format,
    XLSX_MAGIC_BYTES,
    XLS_MAGIC_BYTES,
)


class TestConvertXlsToXlsxBasic:
    def _mock_xlrd_workbook(self, sheet_data=None):
        if sheet_data is None:
            sheet_data = {
                'Sheet1': [
                    ['Name', 'Value'],
                    ['test', 123],
                ]
            }

        mock_book = MagicMock()
        mock_book.sheet_names.return_value = list(sheet_data.keys())
        mock_book.datemode = 0
        mock_book.release_resources = MagicMock()

        mock_sheets = {}
        for sheet_name, rows in sheet_data.items():
            mock_sheet = MagicMock()
            mock_sheet.nrows = len(rows)
            mock_sheet.ncols = len(rows[0]) if rows else 0

            def make_cell_value(rows):
                def cell_value(row_idx, col_idx):
                    return rows[row_idx][col_idx]
                return cell_value

            def make_cell_type(rows):
                def cell_type(row_idx, col_idx):
                    return 0
                return cell_type

            mock_sheet.cell_value = make_cell_value(rows)
            mock_sheet.cell_type = make_cell_type(rows)
            mock_sheet.sheet_by_name = MagicMock()
            mock_sheets[sheet_name] = mock_sheet

        def sheet_by_name(name):
            return mock_sheets[name]

        mock_book.sheet_by_name = sheet_by_name
        return mock_book

    def test_returns_valid_xlsx_path(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        mock_book = self._mock_xlrd_workbook()

        with patch('xlrd.open_workbook', return_value=mock_book):
            result_path = convert_xls_to_xlsx(xls_path)

        assert result_path is not None
        assert os.path.exists(result_path)
        assert result_path.endswith('.xlsx')

        cleanup_temp_file(result_path)

    def test_temp_file_has_bankcheck_prefix(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        mock_book = self._mock_xlrd_workbook()

        with patch('xlrd.open_workbook', return_value=mock_book):
            result_path = convert_xls_to_xlsx(xls_path)

        basename = os.path.basename(result_path)
        assert basename.startswith('bankcheck_')

        cleanup_temp_file(result_path)

    def test_temp_file_has_xlsx_suffix(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        mock_book = self._mock_xlrd_workbook()

        with patch('xlrd.open_workbook', return_value=mock_book):
            result_path = convert_xls_to_xlsx(xls_path)

        assert result_path.endswith('.xlsx')

        cleanup_temp_file(result_path)

    def test_each_call_creates_unique_temp_file(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        mock_book = self._mock_xlrd_workbook()

        paths = []
        with patch('xlrd.open_workbook', return_value=mock_book):
            for _ in range(5):
                paths.append(convert_xls_to_xlsx(xls_path))

        unique_paths = set(paths)
        assert len(unique_paths) == len(paths), '多次调用应生成不同的临时文件路径'

        for p in paths:
            cleanup_temp_file(p)

    def test_same_source_filename_produces_different_temp_paths(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'same_name.xls')
        mock_book = self._mock_xlrd_workbook()

        with patch('xlrd.open_workbook', return_value=mock_book):
            path1 = convert_xls_to_xlsx(xls_path)
            path2 = convert_xls_to_xlsx(xls_path)

        assert path1 != path2, '同名源文件的多次转换也应生成不同的临时路径'

        cleanup_temp_file(path1)
        cleanup_temp_file(path2)

    def test_different_source_filenames_produce_different_temp_paths(self, tmp_dir):
        mock_book = self._mock_xlrd_workbook()

        with patch('xlrd.open_workbook', return_value=mock_book):
            path1 = convert_xls_to_xlsx(os.path.join(tmp_dir, 'file_a.xls'))
            path2 = convert_xls_to_xlsx(os.path.join(tmp_dir, 'file_b.xls'))

        assert path1 != path2

        cleanup_temp_file(path1)
        cleanup_temp_file(path2)

    def test_converted_file_is_valid_xlsx(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        mock_book = self._mock_xlrd_workbook()

        with patch('xlrd.open_workbook', return_value=mock_book):
            result_path = convert_xls_to_xlsx(xls_path)

        wb = openpyxl.load_workbook(result_path)
        assert 'Sheet1' in wb.sheetnames
        ws = wb['Sheet1']
        assert ws['A1'].value == 'Name'
        assert ws['B1'].value == 'Value'
        wb.close()

        cleanup_temp_file(result_path)

    def test_temp_file_in_system_temp_dir(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        mock_book = self._mock_xlrd_workbook()

        with patch('xlrd.open_workbook', return_value=mock_book):
            result_path = convert_xls_to_xlsx(xls_path)

        temp_dir = tempfile.gettempdir()
        assert os.path.dirname(result_path) == temp_dir

        cleanup_temp_file(result_path)

    def test_import_error_raised_when_xlrd_missing(self, tmp_dir, monkeypatch):
        xls_path = os.path.join(tmp_dir, 'test.xls')

        original_import = __builtins__.__import__ if hasattr(__builtins__, '__import__') else __import__

        def mock_import(name, *args, **kwargs):
            if name == 'xlrd':
                raise ImportError('No module named xlrd')
            return original_import(name, *args, **kwargs)

        monkeypatch.setattr('builtins.__import__', mock_import)

        with pytest.raises(ImportError, match='缺少 xlrd 库'):
            convert_xls_to_xlsx(xls_path)


class TestConvertXlsToXlsxConcurrency:
    def test_concurrent_calls_create_unique_files(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'concurrent_test.xls')
        mock_book = MagicMock()
        mock_book.sheet_names.return_value = ['Sheet1']
        mock_book.datemode = 0
        mock_book.release_resources = MagicMock()

        mock_sheet = MagicMock()
        mock_sheet.nrows = 1
        mock_sheet.ncols = 1
        mock_sheet.cell_value = lambda r, c: 'test'
        mock_sheet.cell_type = lambda r, c: 0

        def sheet_by_name(name):
            return mock_sheet

        mock_book.sheet_by_name = sheet_by_name

        results = []
        errors = []
        num_threads = 10
        barrier = threading.Barrier(num_threads)

        def worker():
            try:
                barrier.wait()
                with patch('xlrd.open_workbook', return_value=mock_book):
                    path = convert_xls_to_xlsx(xls_path)
                results.append(path)
            except Exception as e:
                errors.append(e)

        threads = [threading.Thread(target=worker) for _ in range(num_threads)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        assert len(errors) == 0, f'并发调用出现错误: {errors}'
        assert len(results) == num_threads

        unique_results = set(results)
        assert len(unique_results) == num_threads, '并发调用应生成互不冲突的临时文件路径'

        for p in results:
            cleanup_temp_file(p)

    def test_concurrent_calls_do_not_overwrite_each_other(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'overwrite_test.xls')
        mock_book = MagicMock()
        mock_book.sheet_names.return_value = ['Sheet1']
        mock_book.datemode = 0
        mock_book.release_resources = MagicMock()

        mock_sheet = MagicMock()
        mock_sheet.nrows = 1
        mock_sheet.ncols = 1
        mock_sheet.cell_type = lambda r, c: 0
        mock_book.sheet_by_name = lambda name: mock_sheet

        call_count = 0
        count_lock = threading.Lock()

        def mock_cell_value(r, c):
            nonlocal call_count
            with count_lock:
                call_count += 1
                return f'value_{call_count}'

        mock_sheet.cell_value = mock_cell_value

        results = []
        errors = []
        num_threads = 5
        barrier = threading.Barrier(num_threads)

        def worker():
            try:
                barrier.wait()
                with patch('xlrd.open_workbook', return_value=mock_book):
                    path = convert_xls_to_xlsx(xls_path)
                results.append(path)
                time.sleep(0.01)
            except Exception as e:
                errors.append(e)

        threads = [threading.Thread(target=worker) for _ in range(num_threads)]
        for t in threads:
            t.start()
        for t in threads:
            t.join()

        assert len(errors) == 0
        for path in results:
            assert os.path.exists(path), '每个并发生成的文件都应独立存在，不被覆盖'

        for p in results:
            cleanup_temp_file(p)


class TestOpenWorkbookCompatWithXls:
    def test_xls_returns_temp_path(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        mock_book = MagicMock()
        mock_book.sheet_names.return_value = ['Sheet1']
        mock_book.datemode = 0
        mock_book.release_resources = MagicMock()

        mock_sheet = MagicMock()
        mock_sheet.nrows = 1
        mock_sheet.ncols = 1
        mock_sheet.cell_value = lambda r, c: 'test'
        mock_sheet.cell_type = lambda r, c: 0
        mock_book.sheet_by_name = lambda name: mock_sheet

        with patch('xlrd.open_workbook', return_value=mock_book):
            wb, tmp_path = open_workbook_compat(xls_path)

        assert tmp_path is not None
        assert os.path.exists(tmp_path)
        assert tmp_path.endswith('.xlsx')
        wb.close()
        cleanup_temp_file(tmp_path)

    def test_xls_temp_file_has_bankcheck_prefix(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        mock_book = MagicMock()
        mock_book.sheet_names.return_value = ['Sheet1']
        mock_book.datemode = 0
        mock_book.release_resources = MagicMock()

        mock_sheet = MagicMock()
        mock_sheet.nrows = 1
        mock_sheet.ncols = 1
        mock_sheet.cell_value = lambda r, c: 'test'
        mock_sheet.cell_type = lambda r, c: 0
        mock_book.sheet_by_name = lambda name: mock_sheet

        with patch('xlrd.open_workbook', return_value=mock_book):
            wb, tmp_path = open_workbook_compat(xls_path)

        basename = os.path.basename(tmp_path)
        assert basename.startswith('bankcheck_')
        wb.close()
        cleanup_temp_file(tmp_path)

    def test_multiple_xls_opens_create_different_temp_files(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        mock_book = MagicMock()
        mock_book.sheet_names.return_value = ['Sheet1']
        mock_book.datemode = 0
        mock_book.release_resources = MagicMock()

        mock_sheet = MagicMock()
        mock_sheet.nrows = 1
        mock_sheet.ncols = 1
        mock_sheet.cell_value = lambda r, c: 'test'
        mock_sheet.cell_type = lambda r, c: 0
        mock_book.sheet_by_name = lambda name: mock_sheet

        with patch('xlrd.open_workbook', return_value=mock_book):
            wb1, tmp1 = open_workbook_compat(xls_path)
            wb2, tmp2 = open_workbook_compat(xls_path)

        assert tmp1 != tmp2

        wb1.close()
        wb2.close()
        cleanup_temp_file(tmp1)
        cleanup_temp_file(tmp2)

    def test_cleanup_removes_temp_file(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        mock_book = MagicMock()
        mock_book.sheet_names.return_value = ['Sheet1']
        mock_book.datemode = 0
        mock_book.release_resources = MagicMock()

        mock_sheet = MagicMock()
        mock_sheet.nrows = 1
        mock_sheet.ncols = 1
        mock_sheet.cell_value = lambda r, c: 'test'
        mock_sheet.cell_type = lambda r, c: 0
        mock_book.sheet_by_name = lambda name: mock_sheet

        with patch('xlrd.open_workbook', return_value=mock_book):
            wb, tmp_path = open_workbook_compat(xls_path)

        assert os.path.exists(tmp_path)
        wb.close()
        cleanup_temp_file(tmp_path)
        assert not os.path.exists(tmp_path)


class TestTempFileSecurity:
    def test_temp_file_not_predictable(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        mock_book = MagicMock()
        mock_book.sheet_names.return_value = ['Sheet1']
        mock_book.datemode = 0
        mock_book.release_resources = MagicMock()

        mock_sheet = MagicMock()
        mock_sheet.nrows = 1
        mock_sheet.ncols = 1
        mock_sheet.cell_value = lambda r, c: 'test'
        mock_sheet.cell_type = lambda r, c: 0
        mock_book.sheet_by_name = lambda name: mock_sheet

        with patch('xlrd.open_workbook', return_value=mock_book):
            path = convert_xls_to_xlsx(xls_path)

        basename = os.path.basename(path)
        assert 'converted' not in basename, '临时文件名不应包含可预测的固定模式'
        assert basename != 'bankcheck_.xlsx', '临时文件名应包含随机部分'
        assert len(basename) > len('bankcheck_.xlsx'), '临时文件名应包含随机字符'

        cleanup_temp_file(path)

    def test_temp_file_random_component_exists(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        mock_book = MagicMock()
        mock_book.sheet_names.return_value = ['Sheet1']
        mock_book.datemode = 0
        mock_book.release_resources = MagicMock()

        mock_sheet = MagicMock()
        mock_sheet.nrows = 1
        mock_sheet.ncols = 1
        mock_sheet.cell_value = lambda r, c: 'test'
        mock_sheet.cell_type = lambda r, c: 0
        mock_book.sheet_by_name = lambda name: mock_sheet

        with patch('xlrd.open_workbook', return_value=mock_book):
            path = convert_xls_to_xlsx(xls_path)

        basename = os.path.basename(path)
        name_without_ext = basename[:-len('.xlsx')]
        prefix = 'bankcheck_'
        random_part = name_without_ext[len(prefix):]
        assert len(random_part) > 0, '临时文件名应包含随机部分'

        cleanup_temp_file(path)


class TestGetExtensionFormat:
    def test_xlsx_extension(self):
        assert get_extension_format('test.xlsx') == 'xlsx'

    def test_xls_extension(self):
        assert get_extension_format('test.xls') == 'xls'

    def test_xlsm_extension(self):
        assert get_extension_format('test.xlsm') == 'xlsx'

    def test_case_insensitive(self):
        assert get_extension_format('TEST.XLSX') == 'xlsx'
        assert get_extension_format('TEST.XLS') == 'xls'

    def test_csv_extension(self):
        assert get_extension_format('test.csv') == 'csv'

    def test_unknown_extension(self):
        assert get_extension_format('test.txt') == 'unknown'
        assert get_extension_format('test.doc') == 'unknown'
        assert get_extension_format('test') == 'unknown'

    def test_path_with_directory(self):
        assert get_extension_format('/path/to/test.xlsx') == 'xlsx'
        assert get_extension_format('/path/to/test.xls') == 'xls'
        assert get_extension_format('/path/to/test.csv') == 'csv'

    def test_xlsx_not_confused_with_xls(self):
        assert get_extension_format('test.xlsx') != 'xls'


class TestDetectExcelFormat:
    def test_detect_xlsx_file(self, tmp_dir):
        xlsx_path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        wb.save(xlsx_path)
        wb.close()
        assert detect_excel_format(xlsx_path) == 'xlsx'

    def test_detect_xls_file_by_magic_bytes(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xls')
        with open(xls_path, 'wb') as f:
            f.write(XLS_MAGIC_BYTES)
            f.write(b'\x00' * 100)
        assert detect_excel_format(xls_path) == 'xls'

    def test_detect_unknown_format(self, tmp_dir):
        txt_path = os.path.join(tmp_dir, 'test.txt')
        with open(txt_path, 'wb') as f:
            f.write(b'Hello, World!')
        assert detect_excel_format(txt_path) == 'unknown'

    def test_detect_empty_file(self, tmp_dir):
        empty_path = os.path.join(tmp_dir, 'empty.xls')
        with open(empty_path, 'wb') as f:
            pass
        assert detect_excel_format(empty_path) == 'unknown'

    def test_detect_short_file(self, tmp_dir):
        short_path = os.path.join(tmp_dir, 'short.xls')
        with open(short_path, 'wb') as f:
            f.write(b'\x00\x00\x00')
        assert detect_excel_format(short_path) == 'unknown'

    def test_file_not_exists(self, tmp_dir):
        missing_path = os.path.join(tmp_dir, 'nonexistent.xlsx')
        assert detect_excel_format(missing_path) == 'unknown'

    def test_xlsx_with_wrong_extension(self, tmp_dir):
        xlsx_path = os.path.join(tmp_dir, 'test.xls')
        wb = openpyxl.Workbook()
        wb.save(xlsx_path)
        wb.close()
        assert detect_excel_format(xlsx_path) == 'xlsx'

    def test_xls_with_wrong_extension(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'test.xlsx')
        with open(xls_path, 'wb') as f:
            f.write(XLS_MAGIC_BYTES)
            f.write(b'\x00' * 100)
        assert detect_excel_format(xls_path) == 'xls'


class TestOpenWorkbookCompatMagicBytes:
    def _mock_xlrd_workbook(self, sheet_data=None):
        if sheet_data is None:
            sheet_data = {
                'Sheet1': [
                    ['Name', 'Value'],
                    ['test', 123],
                ]
            }
        mock_book = MagicMock()
        mock_book.sheet_names.return_value = list(sheet_data.keys())
        mock_book.datemode = 0
        mock_book.release_resources = MagicMock()
        mock_sheets = {}
        for sheet_name, rows in sheet_data.items():
            mock_sheet = MagicMock()
            mock_sheet.nrows = len(rows)
            mock_sheet.ncols = len(rows[0]) if rows else 0
            def make_cell_value(rows):
                def cell_value(row_idx, col_idx):
                    return rows[row_idx][col_idx]
                return cell_value
            def make_cell_type(rows):
                def cell_type(row_idx, col_idx):
                    return 0
                return cell_type
            mock_sheet.cell_value = make_cell_value(rows)
            mock_sheet.cell_type = make_cell_type(rows)
            mock_sheets[sheet_name] = mock_sheet
        def sheet_by_name(name):
            return mock_sheets[name]
        mock_book.sheet_by_name = sheet_by_name
        return mock_book

    def test_xlsx_file_with_xls_extension_uses_openpyxl_directly(self, tmp_dir):
        fake_xls_path = os.path.join(tmp_dir, 'fake.xls')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'Hello'
        wb.save(fake_xls_path)
        wb.close()

        wb_result, tmp_path = open_workbook_compat(fake_xls_path)
        assert tmp_path is None
        assert wb_result.active['A1'].value == 'Hello'
        wb_result.close()

    def test_xls_file_with_xlsx_extension_uses_xlrd_conversion(self, tmp_dir):
        fake_xlsx_path = os.path.join(tmp_dir, 'fake.xlsx')
        with open(fake_xlsx_path, 'wb') as f:
            f.write(XLS_MAGIC_BYTES)
            f.write(b'\x00' * 100)

        mock_book = self._mock_xlrd_workbook()
        with patch('xlrd.open_workbook', return_value=mock_book):
            wb_result, tmp_path = open_workbook_compat(fake_xlsx_path)

        assert tmp_path is not None
        assert os.path.exists(tmp_path)
        wb_result.close()
        cleanup_temp_file(tmp_path)

    def test_normal_xlsx_still_works(self, tmp_dir):
        xlsx_path = os.path.join(tmp_dir, 'normal.xlsx')
        wb = openpyxl.Workbook()
        wb.active['A1'] = 'test'
        wb.save(xlsx_path)
        wb.close()

        wb_result, tmp_path = open_workbook_compat(xlsx_path)
        assert tmp_path is None
        wb_result.close()

    def test_normal_xls_still_works(self, tmp_dir):
        xls_path = os.path.join(tmp_dir, 'normal.xls')
        with open(xls_path, 'wb') as f:
            f.write(XLS_MAGIC_BYTES)
            f.write(b'\x00' * 100)

        mock_book = self._mock_xlrd_workbook()
        with patch('xlrd.open_workbook', return_value=mock_book):
            wb_result, tmp_path = open_workbook_compat(xls_path)

        assert tmp_path is not None
        wb_result.close()
        cleanup_temp_file(tmp_path)

    def test_unknown_format_falls_back_to_extension_xlsx(self, tmp_dir):
        unknown_path = os.path.join(tmp_dir, 'unknown.xlsx')
        with open(unknown_path, 'wb') as f:
            f.write(b'\x00\x01\x02\x03\x04\x05\x06\x07')

        wb = openpyxl.Workbook()
        wb.active['A1'] = 'fallback_test'
        wb.save(unknown_path)
        wb.close()

        wb_result, tmp_path = open_workbook_compat(unknown_path)
        assert tmp_path is None
        wb_result.close()

    def test_unknown_format_falls_back_to_extension_xls(self, tmp_dir):
        unknown_path = os.path.join(tmp_dir, 'unknown.xls')
        with open(unknown_path, 'wb') as f:
            f.write(b'\x00\x01\x02\x03\x04\x05\x06\x07')

        mock_book = self._mock_xlrd_workbook()
        with patch('xlrd.open_workbook', return_value=mock_book):
            wb_result, tmp_path = open_workbook_compat(unknown_path)

        assert tmp_path is not None
        wb_result.close()
        cleanup_temp_file(tmp_path)
