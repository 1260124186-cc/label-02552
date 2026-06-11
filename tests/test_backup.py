import os
import re
import time
import shutil

import openpyxl
import pandas as pd
import pytest

import bankcheck
from bankcheck import (
    backup_existing_file,
    export_summary_to_csv,
    merge_and_export_summary,
    OUTPUT_FORMAT_XLSX,
    OUTPUT_FORMAT_CSV,
)


class TestBackupExistingFile:
    def test_file_not_exists_returns_none(self, tmp_dir):
        target = os.path.join(tmp_dir, '不存在的文件.xlsx')
        result = backup_existing_file(target)
        assert result is None

    def test_none_path_returns_none(self):
        assert backup_existing_file(None) is None

    def test_empty_string_path_returns_none(self):
        assert backup_existing_file('') is None

    def test_backup_xlsx_success(self, tmp_dir):
        original = os.path.join(tmp_dir, '银行流水总表.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '原始数据'
        ws['A2'] = 100
        wb.save(original)
        wb.close()
        assert os.path.exists(original)

        backup_path = backup_existing_file(original)

        assert backup_path is not None
        assert os.path.exists(backup_path)
        assert os.path.exists(original)

        base = os.path.basename(backup_path)
        assert base.startswith('银行流水总表_')
        assert base.endswith('.xlsx')
        assert re.search(r'_\d{8}_\d{6}', base)

    def test_backup_csv_success(self, tmp_dir):
        original = os.path.join(tmp_dir, '银行流水总表.csv')
        with open(original, 'w', encoding='utf-8-sig') as f:
            f.write('col1,col2\n')
            f.write('a,1\n')

        backup_path = backup_existing_file(original)

        assert backup_path is not None
        assert os.path.exists(backup_path)
        base = os.path.basename(backup_path)
        assert base.startswith('银行流水总表_')
        assert base.endswith('.csv')
        assert re.search(r'_\d{8}_\d{6}', base)

    def test_backup_preserves_content(self, tmp_dir):
        original = os.path.join(tmp_dir, '银行流水总表.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '唯一id'
        ws['B1'] = '银行'
        ws['A2'] = 'id_001'
        ws['B2'] = '北京银行'
        wb.save(original)
        wb.close()

        backup_path = backup_existing_file(original)

        df_orig = pd.read_excel(original, engine='openpyxl')
        df_backup = pd.read_excel(backup_path, engine='openpyxl')
        pd.testing.assert_frame_equal(df_orig, df_backup)

    def test_same_timestamp_conflict_appends_counter(self, tmp_dir, monkeypatch):
        original = os.path.join(tmp_dir, '银行流水总表.xlsx')
        wb = openpyxl.Workbook()
        wb.save(original)
        wb.close()

        fixed_ts = '20260611_120000'

        class MockDatetimeInstance:
            def strftime(self, fmt):
                return fixed_ts

        class MockDatetime:
            @staticmethod
            def now():
                return MockDatetimeInstance()

        monkeypatch.setattr('bankcheck.datetime', MockDatetime)

        backup1 = backup_existing_file(original)
        assert backup1 is not None
        backup1_base = os.path.basename(backup1)
        assert fixed_ts in backup1_base
        assert not re.search(rf'{fixed_ts}_\d+\.xlsx$', backup1_base)

        wb2 = openpyxl.Workbook()
        wb2.save(original)
        wb2.close()

        backup2 = backup_existing_file(original)
        assert backup2 is not None
        backup2_base = os.path.basename(backup2)
        assert re.search(rf'{fixed_ts}_1\.xlsx$', backup2_base)

        wb3 = openpyxl.Workbook()
        wb3.save(original)
        wb3.close()

        backup3 = backup_existing_file(original)
        assert backup3 is not None
        backup3_base = os.path.basename(backup3)
        assert re.search(rf'{fixed_ts}_2\.xlsx$', backup3_base)

    def test_backup_different_extensions(self, tmp_dir):
        for ext in ['.xlsx', '.csv', '.xls', '.txt']:
            fname = f'test_file{ext}'
            fpath = os.path.join(tmp_dir, fname)
            with open(fpath, 'w') as f:
                f.write(f'dummy{ext}')

            result = backup_existing_file(fpath)
            assert result is not None
            assert os.path.basename(result).startswith(f'test_file_')
            assert os.path.basename(result).endswith(ext)
            assert os.path.exists(result)

    def test_exception_during_backup_returns_none(self, tmp_dir, monkeypatch):
        original = os.path.join(tmp_dir, '银行流水总表.xlsx')
        wb = openpyxl.Workbook()
        wb.save(original)
        wb.close()

        def mock_copy2(*args, **kwargs):
            raise OSError('磁盘已满')

        monkeypatch.setattr('bankcheck.shutil.copy2', mock_copy2)

        result = backup_existing_file(original)
        assert result is None
        assert os.path.exists(original)

    def test_backup_in_nested_directory(self, tmp_dir):
        nested = os.path.join(tmp_dir, 'level1', 'level2')
        os.makedirs(nested)
        original = os.path.join(nested, '银行流水总表.xlsx')
        wb = openpyxl.Workbook()
        wb.save(original)
        wb.close()

        backup_path = backup_existing_file(original)

        assert backup_path is not None
        assert os.path.dirname(backup_path) == nested
        assert os.path.exists(backup_path)


class TestExportSummaryToCsvBackup:
    def test_first_run_no_backup(self, tmp_dir, monkeypatch):
        monkeypatch.setattr('bankcheck.get_output_dir', lambda: tmp_dir)
        records = [{'唯一id': '1', '银行': '测试银行', '银行账号': '001',
                    '主体': '测试主体', '交易日期': '2024-01-01',
                    '付款': None, '收款': 100, '摘要': '测试',
                    '对方户名': '对方', '余额': 100, '交易流水号': 'txn001'}]

        result = export_summary_to_csv(records, output_dir=tmp_dir)

        assert result is not None
        backup_files = [f for f in os.listdir(tmp_dir) if re.match(r'银行流水总表_\d{8}_\d{6}.csv', f)]
        assert len(backup_files) == 0

    def test_second_run_creates_backup(self, tmp_dir, monkeypatch):
        monkeypatch.setattr('bankcheck.get_output_dir', lambda: tmp_dir)
        target = os.path.join(tmp_dir, '银行流水总表.csv')

        with open(target, 'w', encoding='utf-8-sig') as f:
            f.write('唯一id,银行,银行账号\n')
            f.write('old1,旧银行,999\n')

        records = [{'唯一id': 'new1', '银行': '新银行', '银行账号': '001',
                    '主体': '测试主体', '交易日期': '2024-01-01',
                    '付款': None, '收款': 100, '摘要': '测试',
                    '对方户名': '对方', '余额': 100, '交易流水号': 'txn001'}]

        export_summary_to_csv(records, output_dir=tmp_dir)

        backup_files = [f for f in os.listdir(tmp_dir) if re.match(r'银行流水总表_\d{8}_\d{6}.csv', f)]
        assert len(backup_files) == 1

        df_backup = pd.read_csv(os.path.join(tmp_dir, backup_files[0]))
        assert df_backup.iloc[0]['银行'] == '旧银行'

        df_new = pd.read_csv(target)
        assert df_new.iloc[0]['银行'] == '新银行'


class TestMergeAndExportSummaryBackup:
    def test_first_run_no_backup_xlsx(self, tmp_dir, monkeypatch):
        monkeypatch.setattr('bankcheck.get_output_dir', lambda: tmp_dir)
        records = [{'唯一id': '1', '银行': '北京银行', '银行账号': '010',
                    '主体': '测试', '交易日期': '2024-01-01',
                    '付款': None, '收款': 1000, '摘要': '货款',
                    '对方户名': '供应商', '余额': 5000, '交易流水号': 't1'}]

        paths = merge_and_export_summary(
            existing_records=[],
            incremental_rows=records,
            output_dir=tmp_dir,
            output_formats=[OUTPUT_FORMAT_XLSX],
        )

        assert OUTPUT_FORMAT_XLSX in paths
        backup_files = [f for f in os.listdir(tmp_dir) if re.match(r'银行流水总表_\d{8}_\d{6}.xlsx', f)]
        assert len(backup_files) == 0

    def test_second_run_backup_xlsx(self, tmp_dir, monkeypatch):
        monkeypatch.setattr('bankcheck.get_output_dir', lambda: tmp_dir)
        target = os.path.join(tmp_dir, '银行流水总表.xlsx')

        old_wb = openpyxl.Workbook()
        old_ws = old_wb.active
        old_ws['A1'] = '唯一id'
        old_ws['B1'] = '银行'
        old_ws['A2'] = 'old_id'
        old_ws['B2'] = '旧银行'
        old_wb.save(target)
        old_wb.close()

        new_records = [{'唯一id': 'new_id', '银行': '新银行', '银行账号': '001',
                        '主体': '测试', '交易日期': '2024-02-01',
                        '付款': None, '收款': 2000, '摘要': '收款',
                        '对方户名': '客户', '余额': 7000, '交易流水号': 't_new'}]

        paths = merge_and_export_summary(
            existing_records=[],
            incremental_rows=new_records,
            output_dir=tmp_dir,
            output_formats=[OUTPUT_FORMAT_XLSX],
        )

        backup_files = [f for f in os.listdir(tmp_dir) if re.match(r'银行流水总表_\d{8}_\d{6}.xlsx', f)]
        assert len(backup_files) == 1

        df_backup = pd.read_excel(os.path.join(tmp_dir, backup_files[0]), engine='openpyxl')
        assert df_backup.iloc[0]['银行'] == '旧银行'

        df_new = pd.read_excel(target, engine='openpyxl')
        assert df_new.iloc[0]['银行'] == '新银行'

    def test_both_xlsx_and_csv_backup(self, tmp_dir, monkeypatch):
        monkeypatch.setattr('bankcheck.get_output_dir', lambda: tmp_dir)
        xlsx_target = os.path.join(tmp_dir, '银行流水总表.xlsx')
        csv_target = os.path.join(tmp_dir, '银行流水总表.csv')

        wb = openpyxl.Workbook()
        wb.save(xlsx_target)
        wb.close()
        with open(csv_target, 'w', encoding='utf-8-sig') as f:
            f.write('col1\n')

        records = [{'唯一id': '1', '银行': '测试', '银行账号': '001',
                    '主体': '主体', '交易日期': '2024-01-01',
                    '付款': None, '收款': 500, '摘要': '摘要',
                    '对方户名': '对方', '余额': 500, '交易流水号': 'txn1'}]

        merge_and_export_summary(
            existing_records=[],
            incremental_rows=records,
            output_dir=tmp_dir,
            output_formats=[OUTPUT_FORMAT_XLSX, OUTPUT_FORMAT_CSV],
        )

        xlsx_backups = [f for f in os.listdir(tmp_dir) if re.match(r'银行流水总表_\d{8}_\d{6}.xlsx', f)]
        csv_backups = [f for f in os.listdir(tmp_dir) if re.match(r'银行流水总表_\d{8}_\d{6}.csv', f)]
        assert len(xlsx_backups) == 1
        assert len(csv_backups) == 1

    def test_incremental_with_existing_history_backup(self, tmp_dir, monkeypatch):
        monkeypatch.setattr('bankcheck.get_output_dir', lambda: tmp_dir)
        target = os.path.join(tmp_dir, '银行流水总表.xlsx')

        history = [
            {'唯一id': 'h1', '银行': '历史银行', '银行账号': 'h01',
             '主体': '历史主体', '交易日期': '2024-01-01',
             '付款': None, '收款': 100, '摘要': '历史',
             '对方户名': '对方', '余额': 100, '交易流水号': 'ht1'},
        ]
        df_hist = pd.DataFrame(history, columns=[
            '唯一id', '银行', '银行账号', '主体', '交易日期',
            '付款', '收款', '摘要', '对方户名', '余额', '交易流水号',
        ])
        df_hist.to_excel(target, index=False, engine='openpyxl')

        incremental = [
            {'唯一id': 'n1', '银行': '新增银行', '银行账号': 'n01',
             '主体': '新增主体', '交易日期': '2024-02-01',
             '付款': None, '收款': 200, '摘要': '新增',
             '对方户名': '新对方', '余额': 300, '交易流水号': 'nt1'},
        ]

        merge_and_export_summary(
            existing_records=history,
            incremental_rows=incremental,
            output_dir=tmp_dir,
            output_formats=[OUTPUT_FORMAT_XLSX],
        )

        backup_files = [f for f in os.listdir(tmp_dir) if re.match(r'银行流水总表_\d{8}_\d{6}.xlsx', f)]
        assert len(backup_files) == 1

        df_final = pd.read_excel(target, engine='openpyxl')
        assert len(df_final) == 2
        assert list(df_final['银行']) == ['历史银行', '新增银行']
