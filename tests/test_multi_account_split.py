import os
import shutil
import tempfile

import openpyxl
import pytest

from conftest import (
    _create_beijing_bank_excel,
    _create_beijing_bank_multi_account_excel,
    _create_lookup_table,
)
import bankcheck
from bankcheck import GenericBankParser, BankRule


_BEIJING_RULE_KWARGS = dict(
    bank_name='北京银行',
    account_cell='B2',
    start_row=4,
    columns={
        'trade_date': 2, 'payment': 4, 'receipt': 5,
        'summary': 12, 'counterpart': 7, 'balance': 6,
        'transaction_id': 16,
    },
    payment_sign='negative',
    multi_account=True,
)


class TestDetectAccountBlocks:

    def test_single_account_returns_one_block(self, tmp_dir):
        filepath = _create_beijing_bank_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        blocks = parser._detect_account_blocks(ws)
        wb.close()
        assert len(blocks) == 1
        assert blocks[0]['account'] == '01090312345678901'

    def test_two_accounts_returns_two_blocks(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        blocks = parser._detect_account_blocks(ws)
        wb.close()
        assert len(blocks) == 2
        assert blocks[0]['account'] == '01090312345678901'
        assert blocks[1]['account'] == '01090399999999999'

    def test_block_row_ranges(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        blocks = parser._detect_account_blocks(ws)
        wb.close()
        assert blocks[0]['header_row'] == 3
        assert blocks[0]['data_start_row'] == 4
        assert blocks[1]['header_row'] == 8
        assert blocks[1]['data_start_row'] == 9
        assert blocks[0]['data_end_row'] == 6
        assert blocks[1]['data_end_row'] == ws.max_row

    def test_three_accounts(self, tmp_dir):
        accounts = [
            {
                'account': '01090311111111111',
                'rows': [
                    [1, '2024-01-05', 'CNY', 10000, None, 1490000, 'A', '1', 'ICBC', 'T', '001', 'P1', None, None, None, 'T1'],
                ],
            },
            {
                'account': '01090322222222222',
                'rows': [
                    [1, '2024-02-01', 'CNY', 20000, None, 1470000, 'B', '2', 'CCB', 'T', '002', 'P2', None, None, None, 'T2'],
                ],
            },
            {
                'account': '01090333333333333',
                'rows': [
                    [1, '2024-03-01', 'CNY', None, 30000, 1500000, 'C', '3', 'BOC', 'T', '003', 'P3', None, None, None, 'T3'],
                ],
            },
        ]
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'), accounts=accounts)
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        blocks = parser._detect_account_blocks(ws)
        wb.close()
        assert len(blocks) == 3
        assert blocks[0]['account'] == '01090311111111111'
        assert blocks[1]['account'] == '01090322222222222'
        assert blocks[2]['account'] == '01090333333333333'

    def test_empty_sheet_returns_empty(self):
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.Workbook()
        ws = wb.active
        blocks = parser._detect_account_blocks(ws)
        wb.close()
        assert blocks == []

    def test_no_account_like_values_returns_empty(self):
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='标题')
        ws.cell(row=2, column=2, value='无账号')
        blocks = parser._detect_account_blocks(ws)
        wb.close()
        assert blocks == []


class TestParseSheetMultiAccount:

    def test_two_accounts_correct_subjects(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        mappings = [
            ('北京XX科技有限公司', '01090312345678901'),
            ('北京YY贸易有限公司', '01090399999999999'),
        ]
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'),
                                       mappings=mappings)
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = parser._parse_sheet(ws, filepath, ws.title, lookup)
        wb.close()
        assert len(rows) == 4
        acct1_rows = [r for r in rows if r['银行账号'] == '01090312345678901']
        acct2_rows = [r for r in rows if r['银行账号'] == '01090399999999999']
        assert len(acct1_rows) == 2
        assert len(acct2_rows) == 2
        assert acct1_rows[0]['主体'] == '北京XX科技有限公司'
        assert acct2_rows[0]['主体'] == '北京YY贸易有限公司'

    def test_two_accounts_payment_receipt_values(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        mappings = [
            ('北京XX科技有限公司', '01090312345678901'),
            ('北京YY贸易有限公司', '01090399999999999'),
        ]
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'),
                                       mappings=mappings)
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = parser._parse_sheet(ws, filepath, ws.title, lookup)
        wb.close()
        acct1_rows = [r for r in rows if r['银行账号'] == '01090312345678901']
        acct2_rows = [r for r in rows if r['银行账号'] == '01090399999999999']
        assert acct1_rows[0]['付款'] == -50000.0
        assert acct1_rows[1]['收款'] == 80000.0
        assert acct2_rows[0]['付款'] == -20000.0
        assert acct2_rows[1]['收款'] == 60000.0

    def test_each_row_has_unique_id(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        mappings = [
            ('北京XX科技有限公司', '01090312345678901'),
            ('北京YY贸易有限公司', '01090399999999999'),
        ]
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'),
                                       mappings=mappings)
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = parser._parse_sheet(ws, filepath, ws.title, lookup)
        wb.close()
        ids = [r['唯一id'] for r in rows]
        assert len(set(ids)) == len(ids)

    def test_all_rows_have_correct_bank_name(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'))
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = parser._parse_sheet(ws, filepath, ws.title, lookup)
        wb.close()
        for row in rows:
            assert row['银行'] == '北京银行'

    def test_standard_columns_present(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'))
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = parser._parse_sheet(ws, filepath, ws.title, lookup)
        wb.close()
        expected_keys = {'唯一id', '银行', '银行账号', '主体', '交易日期',
                         '付款', '收款', '摘要', '对方户名', '余额', '交易流水号'}
        for row in rows:
            assert expected_keys <= set(row.keys())

    def test_single_account_in_multi_account_mode(self, tmp_dir):
        filepath = _create_beijing_bank_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'))
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = parser._parse_sheet(ws, filepath, ws.title, lookup)
        wb.close()
        assert len(rows) == 2
        assert rows[0]['银行账号'] == '01090312345678901'
        assert rows[0]['主体'] == '北京XX科技有限公司'
        assert rows[0]['付款'] == -50000.0

    def test_empty_account_block_skipped(self, tmp_dir):
        accounts = [
            {
                'account': '01090312345678901',
                'rows': [
                    [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A', '622001', '工商银行', '转账', '001', '付款', None, None, None, 'BJ001'],
                ],
            },
            {
                'account': '01090399999999999',
                'rows': [],
            },
        ]
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'), accounts=accounts)
        mappings = [
            ('北京XX科技有限公司', '01090312345678901'),
            ('北京YY贸易有限公司', '01090399999999999'),
        ]
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'),
                                       mappings=mappings)
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = parser._parse_sheet(ws, filepath, ws.title, lookup)
        wb.close()
        assert len(rows) == 1
        assert rows[0]['银行账号'] == '01090312345678901'


class TestParseFullFileMultiAccount:

    def test_parse_two_account_file(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        mappings = [
            ('北京XX科技有限公司', '01090312345678901'),
            ('北京YY贸易有限公司', '01090399999999999'),
        ]
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'),
                                       mappings=mappings)
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        rows = parser.parse(filepath, lookup)
        assert len(rows) == 4
        acct1_rows = [r for r in rows if r['银行账号'] == '01090312345678901']
        acct2_rows = [r for r in rows if r['银行账号'] == '01090399999999999']
        assert len(acct1_rows) == 2
        assert len(acct2_rows) == 2
        assert acct1_rows[0]['主体'] == '北京XX科技有限公司'
        assert acct2_rows[0]['主体'] == '北京YY贸易有限公司'

    def test_parse_three_account_file(self, tmp_dir):
        accounts = [
            {
                'account': '01090311111111111',
                'rows': [
                    [1, '2024-01-05', 'CNY', 10000, None, 1490000, 'A', '1', 'ICBC', 'T', '001', 'P1', None, None, None, 'T1'],
                ],
            },
            {
                'account': '01090322222222222',
                'rows': [
                    [1, '2024-02-01', 'CNY', 20000, None, 1470000, 'B', '2', 'CCB', 'T', '002', 'P2', None, None, None, 'T2'],
                ],
            },
            {
                'account': '01090333333333333',
                'rows': [
                    [1, '2024-03-01', 'CNY', None, 30000, 1500000, 'C', '3', 'BOC', 'T', '003', 'P3', None, None, None, 'T3'],
                ],
            },
        ]
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'), accounts=accounts)
        mappings = [
            ('公司A', '01090311111111111'),
            ('公司B', '01090322222222222'),
            ('公司C', '01090333333333333'),
        ]
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'),
                                       mappings=mappings)
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        rows = parser.parse(filepath, lookup)
        assert len(rows) == 3
        subjects = {r['主体'] for r in rows}
        assert subjects == {'公司A', '公司B', '公司C'}


class TestMultiAccountBackwardCompat:

    def test_multi_account_false_keeps_existing_behavior(self, tmp_dir):
        filepath = _create_beijing_bank_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'))
        rule = BankRule(
            bank_name='北京银行',
            account_cell='B2',
            start_row=4,
            columns={
                'trade_date': 2, 'payment': 4, 'receipt': 5,
                'summary': 12, 'counterpart': 7, 'balance': 6,
                'transaction_id': 16,
            },
            payment_sign='negative',
            multi_account=False,
        )
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = parser._parse_sheet(ws, filepath, ws.title, lookup)
        wb.close()
        assert len(rows) == 2
        assert rows[0]['银行账号'] == '01090312345678901'
        assert rows[0]['主体'] == '北京XX科技有限公司'

    def test_multi_account_false_with_multi_account_file_uses_first_account(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        mappings = [
            ('北京XX科技有限公司', '01090312345678901'),
            ('北京YY贸易有限公司', '01090399999999999'),
        ]
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'),
                                       mappings=mappings)
        rule = BankRule(
            bank_name='北京银行',
            account_cell='B2',
            start_row=4,
            columns={
                'trade_date': 2, 'payment': 4, 'receipt': 5,
                'summary': 12, 'counterpart': 7, 'balance': 6,
                'transaction_id': 16,
            },
            payment_sign='negative',
            multi_account=False,
        )
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = parser._parse_sheet(ws, filepath, ws.title, lookup)
        wb.close()
        for row in rows:
            assert row['银行账号'] == '01090312345678901'

    def test_default_multi_account_is_false(self):
        rule = BankRule(
            bank_name='测试银行',
            account_cell='B2',
            start_row=4,
            columns={'trade_date': 1, 'payment': 2, 'receipt': 3,
                     'summary': 4, 'counterpart': 5, 'balance': 6,
                     'transaction_id': 7},
        )
        assert rule.multi_account is False


class TestMultiAccountPipelineIntegration:

    def _setup_multi_account_folder(self, tmp_dir, script_dir):
        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)

        _create_beijing_bank_multi_account_excel(
            os.path.join(source_folder, '北京银行_流水.xlsx'))

        mappings = [
            ('北京XX科技有限公司', '01090312345678901'),
            ('北京YY贸易有限公司', '01090399999999999'),
        ]
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'),
                              mappings=mappings)
        return source_folder

    def test_pipeline_with_multi_account(self, tmp_dir):
        original_multi_account = {}
        for bank_name in bankcheck.BANK_PROCESSORS:
            config = bankcheck.get_bank_config()
            rule = config.get_rule(bank_name)
            if rule:
                original_multi_account[bank_name] = rule.multi_account

        try:
            config = bankcheck.get_bank_config()
            rule = config.get_rule('北京银行')
            if rule:
                rule.multi_account = True

            script_dir = os.path.join(tmp_dir, 'script')
            os.makedirs(script_dir, exist_ok=True)
            source = self._setup_multi_account_folder(tmp_dir, script_dir)

            result = bankcheck.run_pipeline(source, script_dir)

            assert len(result.all_rows) == 4
            acct1_rows = [r for r in result.all_rows if r['银行账号'] == '01090312345678901']
            acct2_rows = [r for r in result.all_rows if r['银行账号'] == '01090399999999999']
            assert len(acct1_rows) == 2
            assert len(acct2_rows) == 2
            assert acct1_rows[0]['主体'] == '北京XX科技有限公司'
            assert acct2_rows[0]['主体'] == '北京YY贸易有限公司'
        finally:
            config = bankcheck.get_bank_config()
            for bank_name, orig_val in original_multi_account.items():
                rule = config.get_rule(bank_name)
                if rule:
                    rule.multi_account = orig_val


class TestMultiAccountHeaderValidation:

    def test_headers_validated_per_block(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'))
        rule = BankRule(
            **{**_BEIJING_RULE_KWARGS,
               'expected_headers': {
                   'trade_date': ['交易日期'],
                   'payment': ['支出金额'],
                   'receipt': ['收入金额'],
               },
               'header_validation': 'warn'}
        )
        parser = GenericBankParser(rule)
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active
        rows = parser._parse_sheet(ws, filepath, ws.title, lookup)
        wb.close()
        assert len(rows) == 4

    def test_strict_validation_fails_on_bad_header(self, tmp_dir):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '交易明细'
        ws.cell(row=1, column=1, value='北京银行交易明细')
        ws.cell(row=2, column=2, value='01090312345678901')
        ws.cell(row=3, column=2, value='错误表头')
        ws.cell(row=4, column=2, value='2024-01-05')
        ws.cell(row=5, column=1, value='北京银行交易明细')
        ws.cell(row=6, column=2, value='01090399999999999')
        ws.cell(row=7, column=2, value='交易日期')
        ws.cell(row=8, column=2, value='2024-02-01')
        filepath = os.path.join(tmp_dir, '北京银行_坏表头.xlsx')
        wb.save(filepath)
        wb.close()

        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'))
        rule = BankRule(
            **{**_BEIJING_RULE_KWARGS,
               'expected_headers': {
                   'trade_date': ['交易日期'],
               },
               'header_validation': 'strict'}
        )
        parser = GenericBankParser(rule)
        with pytest.raises(bankcheck.HeaderValidationError):
            parser._parse_sheet_multi_account(ws, filepath, ws.title, lookup)


class TestMultiAccountUnmatchedSubject:

    def test_unmatched_account_gets_empty_subject(self, tmp_dir):
        accounts = [
            {
                'account': '01090312345678901',
                'rows': [
                    [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A', '622001', '工商银行', '转账', '001', '付款', None, None, None, 'BJ001'],
                ],
            },
            {
                'account': '999999999999',
                'rows': [
                    [1, '2024-02-01', 'CNY', 20000, None, 800000, '供应商B', '622002', '建设银行', '转账', '002', '采购', None, None, None, 'BJ002'],
                ],
            },
        ]
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'), accounts=accounts)
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'))
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        rows = parser.parse(filepath, lookup)
        assert len(rows) == 2
        matched = [r for r in rows if r['主体'] == '北京XX科技有限公司']
        unmatched = [r for r in rows if r['银行账号'] == '999999999999']
        assert len(matched) == 1
        assert len(unmatched) == 1
        assert unmatched[0]['主体'] == ''


class TestMultiAccountTransactionDetails:

    def test_transaction_ids_preserved(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        mappings = [
            ('北京XX科技有限公司', '01090312345678901'),
            ('北京YY贸易有限公司', '01090399999999999'),
        ]
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'),
                                       mappings=mappings)
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        rows = parser.parse(filepath, lookup)
        tx_ids = [r['交易流水号'] for r in rows]
        assert 'BJ20240105001' in tx_ids
        assert 'BJ20240110002' in tx_ids
        assert 'BJ20240201003' in tx_ids
        assert 'BJ20240215004' in tx_ids

    def test_counterpart_preserved(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        mappings = [
            ('北京XX科技有限公司', '01090312345678901'),
            ('北京YY贸易有限公司', '01090399999999999'),
        ]
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'),
                                       mappings=mappings)
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        rows = parser.parse(filepath, lookup)
        counterparts = {r['对方户名'] for r in rows}
        assert '供应商A公司' in counterparts
        assert '供应商C公司' in counterparts

    def test_balance_preserved(self, tmp_dir):
        filepath = _create_beijing_bank_multi_account_excel(
            os.path.join(tmp_dir, '北京银行_流水.xlsx'))
        mappings = [
            ('北京XX科技有限公司', '01090312345678901'),
            ('北京YY贸易有限公司', '01090399999999999'),
        ]
        lookup = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'),
                                       mappings=mappings)
        rule = BankRule(**_BEIJING_RULE_KWARGS)
        parser = GenericBankParser(rule)
        rows = parser.parse(filepath, lookup)
        acct1_rows = [r for r in rows if r['银行账号'] == '01090312345678901']
        acct2_rows = [r for r in rows if r['银行账号'] == '01090399999999999']
        assert acct1_rows[0]['余额'] == 1500000
        assert acct2_rows[0]['余额'] == 800000
