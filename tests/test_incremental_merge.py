import os
import tempfile
import shutil

import openpyxl
import pandas as pd
import pytest

from conftest import _create_lookup_table
import bankcheck


SUMMARY_COLUMNS = [
    '唯一id', '银行', '银行账号', '主体', '交易日期',
    '付款', '收款', '摘要', '对方户名', '余额', '交易流水号',
]


def _build_summary_df(rows):
    return pd.DataFrame(rows, columns=SUMMARY_COLUMNS)


def _save_summary_excel(df, path):
    df.to_excel(path, index=False, engine='openpyxl')
    return path


def _base_rows():
    return [
        {
            '唯一id': bankcheck.generate_unique_id(),
            '银行': '北京银行',
            '银行账号': '01090312345678901',
            '主体': '北京XX科技有限公司',
            '交易日期': '2024-01-05',
            '付款': -50000.0,
            '收款': None,
            '摘要': '采购付款',
            '对方户名': '供应商A公司',
            '余额': 1500000,
            '交易流水号': 'BJ20240105001',
        },
        {
            '唯一id': bankcheck.generate_unique_id(),
            '银行': '北京银行',
            '银行账号': '01090312345678901',
            '主体': '北京XX科技有限公司',
            '交易日期': '2024-01-10',
            '付款': None,
            '收款': 80000.0,
            '摘要': '销售收款',
            '对方户名': '客户B公司',
            '余额': 1580000,
            '交易流水号': 'BJ20240110002',
        },
    ]


def _new_rows():
    return [
        {
            '唯一id': bankcheck.generate_unique_id(),
            '银行': '东亚银行',
            '银行账号': '38812345678',
            '主体': '上海YY贸易有限公司',
            '交易日期': '2024-01-15',
            '付款': None,
            '收款': 30000.0,
            '摘要': '服务费',
            '对方户名': '客户C公司',
            '余额': 545000,
            '交易流水号': 'EA20240115003',
        },
    ]


class TestLoadExistingKeys:
    def test_file_not_exists(self):
        keys, records = bankcheck.load_existing_keys('/nonexistent/file.xlsx')
        assert keys == set()
        assert records == []

    def test_empty_file(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            empty_path = os.path.join(tmp_dir, 'empty.xlsx')
            df = pd.DataFrame(columns=SUMMARY_COLUMNS)
            df.to_excel(empty_path, index=False, engine='openpyxl')

            keys, records = bankcheck.load_existing_keys(empty_path)
            assert keys == set()
            assert records == []
        finally:
            shutil.rmtree(tmp_dir)

    def test_missing_columns(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            bad_path = os.path.join(tmp_dir, 'bad.xlsx')
            df = pd.DataFrame([{'col1': 'val1'}])
            df.to_excel(bad_path, index=False, engine='openpyxl')

            keys, records = bankcheck.load_existing_keys(bad_path)
            assert keys == set()
            assert records == []
        finally:
            shutil.rmtree(tmp_dir)

    def test_valid_summary_file(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            rows = _base_rows()
            df = _build_summary_df(rows)
            summary_path = os.path.join(tmp_dir, 'summary.xlsx')
            _save_summary_excel(df, summary_path)

            keys, records = bankcheck.load_existing_keys(summary_path)

            assert len(keys) == 2
            assert len(records) == 2

            for row in rows:
                expected_key = bankcheck._make_match_key(row)
                assert expected_key in keys
        finally:
            shutil.rmtree(tmp_dir)

    def test_with_transaction_id_and_fallback(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            rows = [
                {
                    '唯一id': bankcheck.generate_unique_id(),
                    '银行': '北京银行',
                    '银行账号': '01090312345678901',
                    '主体': '北京XX科技有限公司',
                    '交易日期': '2024-01-05',
                    '付款': -50000.0,
                    '收款': None,
                    '摘要': '采购付款',
                    '对方户名': '供应商A公司',
                    '余额': 1500000,
                    '交易流水号': 'BJ20240105001',
                },
                {
                    '唯一id': bankcheck.generate_unique_id(),
                    '银行': '北京银行',
                    '银行账号': '01090312345678901',
                    '主体': '北京XX科技有限公司',
                    '交易日期': '2024-01-10',
                    '付款': None,
                    '收款': 80000.0,
                    '摘要': '销售收款',
                    '对方户名': '客户B公司',
                    '余额': 1580000,
                    '交易流水号': None,
                },
            ]
            df = _build_summary_df(rows)
            summary_path = os.path.join(tmp_dir, 'summary.xlsx')
            _save_summary_excel(df, summary_path)

            keys, records = bankcheck.load_existing_keys(summary_path)

            assert len(keys) == 2
            assert len(records) == 2
        finally:
            shutil.rmtree(tmp_dir)


class TestFilterIncrementalRecords:
    def test_no_existing_keys(self):
        new_rows = _base_rows()
        existing_keys = set()

        incremental, duplicate = bankcheck.filter_incremental_records(new_rows, existing_keys)

        assert len(incremental) == 2
        assert duplicate == 0

    def test_all_duplicate(self):
        base_rows = _base_rows()
        existing_keys = {bankcheck._make_match_key(r) for r in base_rows}
        original_key_count = len(existing_keys)

        incremental, duplicate = bankcheck.filter_incremental_records(base_rows, existing_keys)

        assert len(incremental) == 0
        assert duplicate == 2
        assert len(existing_keys) == original_key_count

    def test_mixed_duplicate_and_new(self):
        base_rows = _base_rows()
        new_rows = base_rows + _new_rows()
        existing_keys = {bankcheck._make_match_key(r) for r in base_rows}

        incremental, duplicate = bankcheck.filter_incremental_records(new_rows, existing_keys)

        assert len(incremental) == 1
        assert duplicate == 2
        assert incremental[0]['交易流水号'] == 'EA20240115003'

    def test_empty_new_rows(self):
        existing_keys = {'key1', 'key2'}
        incremental, duplicate = bankcheck.filter_incremental_records([], existing_keys)
        assert incremental == []
        assert duplicate == 0
        assert existing_keys == {'key1', 'key2'}

    def test_fallback_key_matching(self):
        row_with_id = {
            '唯一id': bankcheck.generate_unique_id(),
            '银行': '北京银行',
            '银行账号': '01090312345678901',
            '主体': '北京XX科技有限公司',
            '交易日期': '2024-01-05',
            '付款': -50000.0,
            '收款': None,
            '摘要': '采购付款',
            '对方户名': '供应商A公司',
            '余额': 1500000,
            '交易流水号': 'BJ20240105001',
        }
        row_without_id = {
            '唯一id': bankcheck.generate_unique_id(),
            '银行': '北京银行',
            '银行账号': '01090312345678901',
            '主体': '北京XX科技有限公司',
            '交易日期': '2024-01-05',
            '付款': -50000.0,
            '收款': None,
            '摘要': '采购付款',
            '对方户名': '供应商A公司',
            '余额': 1500000,
            '交易流水号': None,
        }

        existing_keys = {bankcheck._make_match_key(row_without_id)}

        incremental, duplicate = bankcheck.filter_incremental_records([row_without_id], existing_keys)
        assert duplicate == 1
        assert len(incremental) == 0


class TestMergeAndExportSummary:
    def test_merge_existing_and_incremental(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            existing = _base_rows()
            incremental = _new_rows()

            output_path = bankcheck.merge_and_export_summary(existing, incremental, tmp_dir)

            assert output_path is not None
            assert os.path.exists(output_path)

            df = pd.read_excel(output_path, engine='openpyxl')
            assert len(df) == 3

            transaction_ids = df['交易流水号'].tolist()
            assert 'BJ20240105001' in transaction_ids
            assert 'BJ20240110002' in transaction_ids
            assert 'EA20240115003' in transaction_ids
        finally:
            shutil.rmtree(tmp_dir)

    def test_only_existing_records(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            existing = _base_rows()

            output_path = bankcheck.merge_and_export_summary(existing, [], tmp_dir)

            assert output_path is not None
            df = pd.read_excel(output_path, engine='openpyxl')
            assert len(df) == 2
        finally:
            shutil.rmtree(tmp_dir)

    def test_only_incremental_records(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            incremental = _new_rows()

            output_path = bankcheck.merge_and_export_summary([], incremental, tmp_dir)

            assert output_path is not None
            df = pd.read_excel(output_path, engine='openpyxl')
            assert len(df) == 1
        finally:
            shutil.rmtree(tmp_dir)

    def test_no_records(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            output_path = bankcheck.merge_and_export_summary([], [], tmp_dir)
            assert output_path is None
        finally:
            shutil.rmtree(tmp_dir)

    def test_preserves_existing_unique_ids(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            existing = _base_rows()
            existing_ids = [r['唯一id'] for r in existing]
            incremental = _new_rows()

            output_path = bankcheck.merge_and_export_summary(existing, incremental, tmp_dir)
            df = pd.read_excel(output_path, engine='openpyxl')

            saved_existing_ids = df[df['交易流水号'].isin(['BJ20240105001', 'BJ20240110002'])]['唯一id'].tolist()
            for orig_id in existing_ids:
                assert orig_id in saved_existing_ids
        finally:
            shutil.rmtree(tmp_dir)


class TestGetSummaryTablePath:
    def test_correct_path(self):
        script_dir = '/path/to/script'
        path = bankcheck.get_summary_table_path(script_dir)
        assert path == os.path.join(script_dir, '银行流水总表.xlsx')


class TestFormatResultMessage:
    def test_incremental_mode_with_new_records(self):
        result = bankcheck.ProcessingResult(
            all_rows=_base_rows() + _new_rows(),
            incremental_mode=True,
            existing_record_count=2,
            new_record_count=1,
            duplicate_record_count=1,
            output_path='/path/to/summary.xlsx',
        )
        msg = bankcheck.format_result_message(result)
        assert '增量合并处理完成' in msg
        assert '运行模式：增量合并' in msg
        assert '历史总记录数：2' in msg
        assert '重复记录（已跳过）：1' in msg
        assert '新增记录（已追加）：1' in msg
        assert '总表当前总记录数：3' in msg

    def test_full_mode(self):
        result = bankcheck.ProcessingResult(
            all_rows=_base_rows(),
            incremental_mode=False,
            output_path='/path/to/summary.xlsx',
        )
        msg = bankcheck.format_result_message(result)
        assert '运行模式：全量覆盖' in msg
        assert '提取记录数：2' in msg

    def test_incremental_mode_no_new_records(self):
        result = bankcheck.ProcessingResult(
            all_rows=_base_rows(),
            incremental_mode=True,
            existing_record_count=2,
            new_record_count=0,
            duplicate_record_count=3,
            output_path='/path/to/summary.xlsx',
        )
        msg = bankcheck.format_result_message(result)
        assert '重复记录（已跳过）：3' in msg
        assert '新增记录（已追加）：0' in msg

    def test_incremental_mode_empty_extraction_with_history(self):
        result = bankcheck.ProcessingResult(
            all_rows=_base_rows(),
            incremental_mode=True,
            existing_record_count=2,
            new_record_count=0,
            duplicate_record_count=0,
            output_path='/path/to/summary.xlsx',
        )
        msg = bankcheck.format_result_message(result)
        assert '增量合并处理完成' in msg
        assert '历史总记录数：2' in msg
        assert '新增记录（已追加）：0' in msg


class TestRunPipelineIncremental:
    def test_full_mode_first_run(self, tmpdir):
        script_dir = str(tmpdir)
        folder = tmpdir.mkdir('input')
        lookup_path = _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['B2'] = '01090312345678901'
        ws['B4'] = '2024-01-05'
        ws['E4'] = 10000
        ws['L4'] = '测试收款'
        ws['G4'] = '测试公司'
        ws['F4'] = 10000
        ws['P4'] = 'TEST001'
        test_file = folder.join('北京银行_测试.xlsx')
        wb.save(str(test_file))

        result = bankcheck.run_pipeline(str(folder), script_dir, incremental=False)

        assert result.incremental_mode is False
        assert result.existing_record_count == 0
        assert result.new_record_count >= 1
        assert result.duplicate_record_count == 0
        assert os.path.exists(result.output_path)

    def test_incremental_mode_with_history(self, tmpdir):
        script_dir = str(tmpdir)
        folder = tmpdir.mkdir('input')
        lookup_path = _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        historical_rows = _base_rows()
        df = _build_summary_df(historical_rows)
        summary_path = bankcheck.get_summary_table_path(script_dir)
        _save_summary_excel(df, summary_path)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['B1'] = '38812345678'
        ws['A5'] = '2024-01-15'
        ws['E5'] = 30000
        ws['L5'] = '服务费'
        ws['I5'] = 545000
        ws['K5'] = 'EA20240115003'
        test_file = folder.join('东亚银行_测试.xlsx')
        wb.save(str(test_file))

        result = bankcheck.run_pipeline(str(folder), script_dir, incremental=True)

        assert result.incremental_mode is True
        assert result.existing_record_count == 2
        assert result.new_record_count == 1
        assert result.duplicate_record_count == 0
        assert len(result.all_rows) == 3

        df_output = pd.read_excel(result.output_path, engine='openpyxl')
        assert len(df_output) == 3
        transaction_ids = df_output['交易流水号'].tolist()
        assert 'EA20240115003' in transaction_ids

    def test_incremental_mode_with_duplicates(self, tmpdir):
        script_dir = str(tmpdir)
        folder = tmpdir.mkdir('input')
        lookup_path = _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        historical_rows = _base_rows()
        df = _build_summary_df(historical_rows)
        summary_path = bankcheck.get_summary_table_path(script_dir)
        _save_summary_excel(df, summary_path)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['B2'] = '01090312345678901'
        ws['B4'] = '2024-01-05'
        ws['D4'] = 50000
        ws['L4'] = '采购付款'
        ws['G4'] = '供应商A公司'
        ws['F4'] = 1500000
        ws['P4'] = 'BJ20240105001'
        test_file = folder.join('北京银行_测试.xlsx')
        wb.save(str(test_file))

        result = bankcheck.run_pipeline(str(folder), script_dir, incremental=True)

        assert result.incremental_mode is True
        assert result.existing_record_count == 2
        assert result.new_record_count == 0
        assert result.duplicate_record_count == 1
        assert len(result.all_rows) == 2

    def test_incremental_mode_no_history(self, tmpdir):
        script_dir = str(tmpdir)
        folder = tmpdir.mkdir('input')
        lookup_path = _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['B2'] = '01090312345678901'
        ws['B4'] = '2024-01-05'
        ws['E4'] = 10000
        ws['L4'] = '测试收款'
        ws['G4'] = '测试公司'
        ws['F4'] = 10000
        ws['P4'] = 'TEST001'
        test_file = folder.join('北京银行_测试.xlsx')
        wb.save(str(test_file))

        result = bankcheck.run_pipeline(str(folder), script_dir, incremental=True)

        assert result.incremental_mode is False
        assert result.existing_record_count == 0
        assert result.new_record_count >= 1
        assert os.path.exists(result.output_path)
