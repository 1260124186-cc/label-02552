"""
跨账号内部划转识别模块单元测试
"""
import os
import sys
import tempfile
import shutil

import openpyxl
import pandas as pd
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='internal_transfer_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _rec(uid, subject, bank, account, trade_date, payment, receipt,
         counterpart, summary='测试', balance=0.0, tid=''):
    """快速创建一条交易记录"""
    return {
        '唯一id': uid,
        '银行': bank,
        '银行账号': account,
        '主体': subject,
        '交易日期': trade_date,
        '付款': payment,
        '收款': receipt,
        '摘要': summary,
        '对方户名': counterpart,
        '余额': balance,
        '交易流水号': tid or uid,
    }


class TestDataClasses:
    """测试数据类默认值"""

    def test_internal_transfer_match_defaults(self):
        m = bankcheck.InternalTransferMatch(
            match_id='', amount=0.0, out_record_id='', in_record_id='',
            out_subject='', in_subject='', out_bank='', in_bank='',
            out_date='', in_date='', days_diff=0,
            out_counterparty='', in_counterparty='',
        )
        assert m.match_id == ''
        assert m.amount == 0.0
        assert m.days_diff == 0

    def test_internal_transfer_result_defaults(self):
        r = bankcheck.InternalTransferResult(marked_records=[], matches=[])
        assert r.total_records == 0
        assert r.match_pairs == 0
        assert r.marked_out_count == 0
        assert r.marked_in_count == 0
        assert r.involved_subjects == []
        assert r.involved_banks == []


class TestHelperFunctions:
    """测试辅助函数"""

    def test_abs_amount_equal_exact(self):
        assert bankcheck._abs_amount_equal(100.0, 100.0) is True
        assert bankcheck._abs_amount_equal(-100.0, 100.0) is True
        assert bankcheck._abs_amount_equal(100.0, -100.0) is True

    def test_abs_amount_equal_within_tolerance(self):
        assert bankcheck._abs_amount_equal(100.0, 100.005, tolerance=0.01) is True
        assert bankcheck._abs_amount_equal(100.0, 99.995, tolerance=0.01) is True

    def test_abs_amount_equal_outside_tolerance(self):
        assert bankcheck._abs_amount_equal(100.0, 100.02, tolerance=0.01) is False
        assert bankcheck._abs_amount_equal(100.0, 99.98, tolerance=0.01) is False

    def test_build_subject_index(self):
        records = [
            _rec('A1', '主体甲', '银行A', '111', '2024-01-01', None, 100, '外部'),
            _rec('A2', '主体甲', '银行B', '222', '2024-01-02', -50, None, '外部'),
            _rec('B1', '主体乙', '银行A', '333', '2024-01-03', None, 200, '外部'),
        ]
        idx = bankcheck._build_subject_index(records)
        assert '主体甲' in idx
        assert '主体乙' in idx
        assert len(idx['主体甲']) == 2
        assert len(idx['主体乙']) == 1
        assert idx['主体甲'] == [0, 1]

    def test_build_counterparty_subject_index(self):
        records = [
            _rec('A1', '主体甲', '银行A', '111', '2024-01-01', None, 100, '主体乙'),
            _rec('A2', '主体甲', '银行B', '222', '2024-01-02', -50, None, '无关公司'),
            _rec('B1', '主体乙', '银行A', '333', '2024-01-03', None, 200, '主体甲'),
        ]
        idx = bankcheck._build_counterparty_subject_index(records)
        assert '主体乙' in idx
        assert '主体甲' in idx
        assert '无关公司' not in idx
        assert idx['主体乙'] == [0]
        assert idx['主体甲'] == [2]


class TestIdentifyInternalTransfers:
    """测试核心识别函数"""

    def test_empty_records(self):
        result = bankcheck.identify_internal_transfers([])
        assert result.total_records == 0
        assert result.match_pairs == 0
        assert result.matches == []
        assert result.marked_records == []

    def test_no_internal_transfer_pure_external(self):
        """纯外部交易，不应被识别"""
        records = [
            _rec('A1', '主体甲', '银行A', '111', '2024-01-05', -50000, None, '供应商公司'),
            _rec('B1', '主体乙', '银行B', '222', '2024-01-06', None, 50000, '客户公司'),
        ]
        result = bankcheck.identify_internal_transfers(records)
        assert result.match_pairs == 0
        assert result.marked_out_count == 0
        assert result.marked_in_count == 0
        for r in result.marked_records:
            assert r.get('内部划转标记') == ''

    def test_strict_match_success(self):
        """严格模式下，双向对方户名匹配、金额对称、时间接近"""
        records = [
            _rec('OUT01', '主体甲', '北京银行', '111', '2024-01-10',
                 -100000.0, None, '主体乙', summary='内部资金调拨'),
            _rec('IN01', '主体乙', '招商银行', '222', '2024-01-11',
                 None, 100000.0, '主体甲', summary='收到内部拨款'),
        ]
        result = bankcheck.identify_internal_transfers(
            records, strict_counterparty_match=True,
        )
        assert result.match_pairs == 1
        assert result.marked_out_count == 1
        assert result.marked_in_count == 1
        assert len(result.matches) == 1
        m = result.matches[0]
        assert m.amount == 100000.0
        assert m.out_subject == '主体甲'
        assert m.in_subject == '主体乙'
        assert m.days_diff == 1

        out = [r for r in result.marked_records if r['唯一id'] == 'OUT01'][0]
        assert out['内部划转标记'] == '是'
        assert out['内部划转方向'] == '划出'
        assert out['内部划转配对ID'] != ''

        inn = [r for r in result.marked_records if r['唯一id'] == 'IN01'][0]
        assert inn['内部划转标记'] == '是'
        assert inn['内部划转方向'] == '划入'
        assert inn['内部划转配对ID'] == out['内部划转配对ID']

    def test_amount_mismatch_not_matched(self):
        """金额不匹配，不能识别"""
        records = [
            _rec('OUT', '主体甲', '银行A', '111', '2024-01-10', -100000, None, '主体乙'),
            _rec('IN', '主体乙', '银行B', '222', '2024-01-11', None, 99999, '主体甲'),
        ]
        result = bankcheck.identify_internal_transfers(records)
        assert result.match_pairs == 0

    def test_time_window_exceeded_not_matched(self):
        """超出时间窗口，不能识别"""
        records = [
            _rec('OUT', '主体甲', '银行A', '111', '2024-01-01', -50000, None, '主体乙'),
            _rec('IN', '主体乙', '银行B', '222', '2024-01-15', None, 50000, '主体甲'),
        ]
        result = bankcheck.identify_internal_transfers(records, time_window_days=7)
        assert result.match_pairs == 0

    def test_time_window_larger_success(self):
        """扩大时间窗口后，可以识别"""
        records = [
            _rec('OUT', '主体甲', '银行A', '111', '2024-01-01', -50000, None, '主体乙'),
            _rec('IN', '主体乙', '银行B', '222', '2024-01-15', None, 50000, '主体甲'),
        ]
        result = bankcheck.identify_internal_transfers(records, time_window_days=20)
        assert result.match_pairs == 1

    def test_counterparty_one_sided_strict_fails(self):
        """严格模式下，单向对方户名匹配失败"""
        records = [
            _rec('OUT', '主体甲', '银行A', '111', '2024-01-10', -80000, None, '主体乙'),
            _rec('IN', '主体乙', '银行B', '222', '2024-01-11', None, 80000, '某第三方'),
        ]
        result = bankcheck.identify_internal_transfers(
            records, strict_counterparty_match=True,
        )
        assert result.match_pairs == 0

    def test_counterparty_one_sided_relaxed_success(self):
        """宽松模式下，单向对方户名匹配成功"""
        records = [
            _rec('OUT', '主体甲', '银行A', '111', '2024-01-10', -80000, None, '主体乙'),
            _rec('IN', '主体乙', '银行B', '222', '2024-01-11', None, 80000, '某第三方'),
        ]
        result = bankcheck.identify_internal_transfers(
            records, strict_counterparty_match=False,
        )
        assert result.match_pairs == 1

    def test_multiple_pairs_multiple_subjects(self):
        """多主体、多对划转"""
        records = [
            _rec('O1', '甲', '银行A', '1', '2024-02-01', -30000, None, '乙'),
            _rec('I1', '乙', '银行B', '2', '2024-02-02', None, 30000, '甲'),
            _rec('O2', '乙', '银行B', '2', '2024-02-05', -15000, None, '丙'),
            _rec('I2', '丙', '银行C', '3', '2024-02-06', None, 15000, '乙'),
            _rec('O3', '丙', '银行C', '3', '2024-02-10', -7500, None, '甲'),
            _rec('I3', '甲', '银行A', '1', '2024-02-11', None, 7500, '丙'),
            _rec('X1', '甲', '银行A', '1', '2024-02-20', -500, None, '外部供应商'),
        ]
        result = bankcheck.identify_internal_transfers(records)
        assert result.match_pairs == 3
        assert result.marked_out_count == 3
        assert result.marked_in_count == 3
        assert len(result.matches) == 3
        amounts = sorted(m.amount for m in result.matches)
        assert amounts == [7500.0, 15000.0, 30000.0]
        assert '甲' in result.involved_subjects
        assert '乙' in result.involved_subjects
        assert '丙' in result.involved_subjects

        x1 = [r for r in result.marked_records if r['唯一id'] == 'X1'][0]
        assert x1.get('内部划转标记') == ''

    def test_amount_tolerance_success(self):
        """容忍度内的微小差异可匹配（如手续费）"""
        records = [
            _rec('OUT', '主体甲', '银行A', '111', '2024-01-10', -50000.0, None, '主体乙'),
            _rec('IN', '主体乙', '银行B', '222', '2024-01-11', None, 49999.99, '主体甲'),
        ]
        result = bankcheck.identify_internal_transfers(
            records, amount_tolerance=0.05,
        )
        assert result.match_pairs == 1

    def test_involved_subjects_and_banks(self):
        """涉及主体和银行统计"""
        records = [
            _rec('O1', '甲公司', '招商银行', '1001', '2024-03-01', -25000, None, '乙公司'),
            _rec('I1', '乙公司', '工商银行', '2001', '2024-03-02', None, 25000, '甲公司'),
        ]
        result = bankcheck.identify_internal_transfers(records)
        assert '甲公司' in result.involved_subjects
        assert '乙公司' in result.involved_subjects
        assert '招商银行' in result.involved_banks
        assert '工商银行' in result.involved_banks

    def test_same_day_transfer(self):
        """当日划转，时间差为0"""
        records = [
            _rec('O', '甲', '银行A', '1', '2024-05-01', -9999, None, '乙'),
            _rec('I', '乙', '银行B', '2', '2024-05-01', None, 9999, '甲'),
        ]
        result = bankcheck.identify_internal_transfers(records)
        assert result.match_pairs == 1
        assert result.matches[0].days_diff == 0

    def test_reverse_direction_transfer(self):
        """划入日期早于划出日期（银行记账时间差异），仍能识别"""
        records = [
            _rec('O', '甲', '银行A', '1', '2024-05-02', -6666, None, '乙'),
            _rec('I', '乙', '银行B', '2', '2024-05-01', None, 6666, '甲'),
        ]
        result = bankcheck.identify_internal_transfers(records, time_window_days=3)
        assert result.match_pairs == 1
        assert result.matches[0].days_diff == 1

    def test_identify_and_tag_convenience_wrapper(self):
        """便捷函数 identify_and_tag_internal_transfers"""
        records = [
            _rec('O', '甲', '银行A', '1', '2024-01-10', -1000, None, '乙'),
            _rec('I', '乙', '银行B', '2', '2024-01-11', None, 1000, '甲'),
        ]
        marked, summary, result = bankcheck.identify_and_tag_internal_transfers(records)
        assert summary['match_pairs'] == 1
        assert summary['marked_out_count'] == 1
        assert summary['marked_in_count'] == 1
        assert summary['total_transfer_amount'] == 1000.0
        assert '甲' in summary['involved_subjects']
        assert len(marked) == 2
        for r in marked:
            assert '内部划转标记' in r


class TestFilterInternalTransfers:
    """测试汇总过滤函数"""

    def _sample_with_internal(self):
        return [
            _rec('OUT', '甲', '银行A', '1', '2024-01-10', -50000, None, '乙'),
            _rec('IN', '乙', '银行B', '2', '2024-01-11', None, 50000, '甲'),
            _rec('X1', '甲', '银行A', '1', '2024-01-15', -10000, None, '外部供应商'),
            _rec('X2', '乙', '银行B', '2', '2024-01-16', None, 20000, '外部客户'),
        ]

    def test_exclude_internal(self):
        records = self._sample_with_internal()
        result = bankcheck.identify_internal_transfers(records)
        filtered = bankcheck.filter_internal_transfers_for_summary(
            result.marked_records, exclude=True,
        )
        ids = [r['唯一id'] for r in filtered]
        assert 'OUT' not in ids
        assert 'IN' not in ids
        assert 'X1' in ids
        assert 'X2' in ids
        assert len(filtered) == 2

    def test_include_only_internal(self):
        records = self._sample_with_internal()
        result = bankcheck.identify_internal_transfers(records)
        internal_only = bankcheck.filter_internal_transfers_for_summary(
            result.marked_records, exclude=False,
        )
        ids = [r['唯一id'] for r in internal_only]
        assert 'OUT' in ids
        assert 'IN' in ids
        assert 'X1' not in ids
        assert 'X2' not in ids
        assert len(internal_only) == 2

    def test_no_internal_tagged_all_kept(self):
        records = [
            _rec('X1', '甲', '银行A', '1', '2024-01-01', -500, None, '外部'),
            _rec('X2', '乙', '银行B', '2', '2024-01-02', None, 500, '外部'),
        ]
        result = bankcheck.identify_internal_transfers(records)
        filtered = bankcheck.filter_internal_transfers_for_summary(
            result.marked_records, exclude=True,
        )
        assert len(filtered) == 2


class TestSummarizeWithInternalTransfers:
    """测试汇总时排除内部划转"""

    def test_summary_excludes_internal_by_default(self):
        records = [
            _rec('OUT', '甲', '银行A', '1', '2024-01-10', -50000, None, '乙'),
            _rec('IN', '乙', '银行B', '2', '2024-01-11', None, 50000, '甲'),
            _rec('X1', '甲', '银行A', '1', '2024-01-15', -10000, None, '外部供应商'),
            _rec('X2', '乙', '银行B', '2', '2024-01-16', None, 20000, '外部客户'),
        ]
        it_result = bankcheck.identify_internal_transfers(records)
        summary = bankcheck.summarize_transactions(
            it_result.marked_records, exclude_internal_transfers=True,
        )
        overall = summary.overall_summary
        assert overall['transaction_count'] == 2
        assert overall['income_count'] == 1
        assert overall['expense_count'] == 1
        assert overall['total_income'] == pytest.approx(20000.0)
        assert overall['total_expense'] == pytest.approx(10000.0)
        assert overall['net_amount'] == pytest.approx(10000.0)

        by_subject = {r['subject']: r for r in summary.by_subject}
        assert by_subject['甲']['total_expense'] == pytest.approx(10000.0)
        assert by_subject['甲']['total_income'] == pytest.approx(0.0)
        assert by_subject['乙']['total_income'] == pytest.approx(20000.0)
        assert by_subject['乙']['total_expense'] == pytest.approx(0.0)

    def test_summary_include_internal_when_disabled(self):
        records = [
            _rec('OUT', '甲', '银行A', '1', '2024-01-10', -50000, None, '乙'),
            _rec('IN', '乙', '银行B', '2', '2024-01-11', None, 50000, '甲'),
            _rec('X1', '甲', '银行A', '1', '2024-01-15', -10000, None, '外部供应商'),
            _rec('X2', '乙', '银行B', '2', '2024-01-16', None, 20000, '外部客户'),
        ]
        it_result = bankcheck.identify_internal_transfers(records)
        summary = bankcheck.summarize_transactions(
            it_result.marked_records, exclude_internal_transfers=False,
        )
        overall = summary.overall_summary
        assert overall['transaction_count'] == 4
        assert overall['total_income'] == pytest.approx(70000.0)
        assert overall['total_expense'] == pytest.approx(60000.0)


class TestExportInternalTransferReport:
    """测试报告导出功能"""

    def test_export_creates_file(self, tmp_dir):
        records = [
            _rec('OUT', '甲', '招商银行', '1', '2024-01-10', -50000.0, None, '乙'),
            _rec('IN', '乙', '工商银行', '2', '2024-01-11', None, 50000.0, '甲'),
        ]
        it_result = bankcheck.identify_internal_transfers(records)
        output_path = os.path.join(tmp_dir, '内部划转报告.xlsx')
        result = bankcheck.export_internal_transfer_report(
            it_result, output_path, source_info={'测试来源': '单元测试'}
        )
        assert result == output_path
        assert os.path.exists(output_path)
        assert os.path.getsize(output_path) > 0

    def test_export_sheet_names(self, tmp_dir):
        records = [
            _rec('OUT', '甲', '银行A', '1', '2024-01-10', -50000, None, '乙'),
            _rec('IN', '乙', '银行B', '2', '2024-01-11', None, 50000, '甲'),
        ]
        it_result = bankcheck.identify_internal_transfers(records)
        output_path = os.path.join(tmp_dir, 'report.xlsx')
        bankcheck.export_internal_transfer_report(it_result, output_path)

        wb = openpyxl.load_workbook(output_path)
        sheets = wb.sheetnames
        assert '识别总览' in sheets
        assert '配对明细' in sheets
        assert '标记明细' in sheets

        ws_pair = wb['配对明细']
        assert ws_pair.max_row == 2
        headers = [c.value for c in ws_pair[1]]
        assert '配对ID' in headers
        assert '划转金额(元)' in headers
        assert '划出主体' in headers
        assert '划入主体' in headers
        assert '时间差(天)' in headers

        wb.close()

    def test_export_no_matches_skips_pair_sheet(self, tmp_dir):
        records = [
            _rec('X1', '甲', '银行A', '1', '2024-01-10', -100, None, '外部'),
            _rec('X2', '乙', '银行B', '2', '2024-01-11', None, 100, '外部'),
        ]
        it_result = bankcheck.identify_internal_transfers(records)
        output_path = os.path.join(tmp_dir, 'report_no_match.xlsx')
        bankcheck.export_internal_transfer_report(it_result, output_path)

        wb = openpyxl.load_workbook(output_path)
        sheets = wb.sheetnames
        assert '识别总览' in sheets
        assert '配对明细' not in sheets
        assert '标记明细' not in sheets

        ws_overview = wb['识别总览']
        data = {}
        for row in ws_overview.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                data[row[0]] = row[1]
        assert data.get('识别配对数') == 0
        wb.close()

    def test_generate_from_records_creates_report(self, tmp_dir):
        records = [
            _rec('OUT', '甲', '银行A', '1', '2024-01-10', -88888, None, '乙'),
            _rec('IN', '乙', '银行B', '2', '2024-01-11', None, 88888, '甲'),
        ]
        path = bankcheck.generate_internal_transfer_from_records(
            records, output_dir=tmp_dir,
            source_info={'数据来源': '测试'},
        )
        assert path is not None
        assert os.path.exists(path)
        assert '内部划转识别报告' in os.path.basename(path)

    def test_generate_from_records_no_matches_returns_none(self, tmp_dir):
        records = [
            _rec('X1', '甲', '银行A', '1', '2024-01-10', -100, None, '外部'),
        ]
        path = bankcheck.generate_internal_transfer_from_records(
            records, output_dir=tmp_dir,
        )
        assert path is None


class TestIntegrationWithPipeline:
    """集成测试：内部划转识别与主体汇总协同工作"""

    def test_full_pipeline_subject_summary_excludes_internal(self, tmp_dir):
        """完整流程：识别内部划转 -> 汇总时排除 -> 验证汇总净额正确"""
        records = [
            _rec('T1', '北京XX科技有限公司', '北京银行', '01090312345678901',
                 '2024-01-05', -50000.0, None, '上海YY贸易有限公司',
                 summary='内部资金调拨-货款', balance=950000.0),
            _rec('T2', '上海YY贸易有限公司', '东亚银行', '38812345678',
                 '2024-01-06', None, 50000.0, '北京XX科技有限公司',
                 summary='收到北京调拨', balance=560000.0),
            _rec('T3', '北京XX科技有限公司', '北京银行', '01090312345678901',
                 '2024-01-10', None, 80000.0, '外部客户公司',
                 summary='销售收款', balance=1030000.0),
            _rec('T4', '上海YY贸易有限公司', '东亚银行', '38812345678',
                 '2024-01-20', -1500.0, None, '银行手续费',
                 summary='手续费', balance=558500.0),
            _rec('T5', '北京XX科技有限公司', '北京银行', '01090312345678901',
                 '2024-02-05', -20000.0, None, '外部供应商',
                 summary='采购付款', balance=1010000.0),
        ]
        marked, summary, _ = bankcheck.identify_and_tag_internal_transfers(records)
        assert summary['match_pairs'] == 1

        output_path = os.path.join(tmp_dir, '主体汇总.xlsx')
        sum_result = bankcheck.summarize_transactions(marked)
        bankcheck.export_subject_summary(sum_result, output_path)

        overall = sum_result.overall_summary
        assert overall['transaction_count'] == 3
        assert overall['total_income'] == pytest.approx(80000.0)
        assert overall['total_expense'] == pytest.approx(21500.0)

        by_subject = {r['subject']: r for r in sum_result.by_subject}
        assert by_subject['北京XX科技有限公司']['total_income'] == pytest.approx(80000.0)
        assert by_subject['北京XX科技有限公司']['total_expense'] == pytest.approx(20000.0)
        assert by_subject['上海YY贸易有限公司']['total_expense'] == pytest.approx(1500.0)
        assert by_subject['上海YY贸易有限公司']['total_income'] == pytest.approx(0.0)

        wb = openpyxl.load_workbook(output_path)
        assert '汇总总览' in wb.sheetnames
        assert '按主体汇总' in wb.sheetnames
        wb.close()
