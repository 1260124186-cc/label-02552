# -*- coding: utf-8 -*-
"""
大额与高频交易筛查模块测试
"""

import os
import sys
import tempfile
import shutil
from datetime import datetime, timedelta

import pytest
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import transaction_screening as ts
from transaction_screening import (
    ScreeningConfig,
    CounterpartyFrequencyRule,
    RiskAlert,
    RiskType,
    ScreeningResult,
    screen_large_amount,
    screen_high_frequency,
    screen_counterparty_frequency,
    run_screening,
    export_screening_result,
    screening_from_records,
)


@pytest.fixture(autouse=True)
def init_logging():
    import logging
    logging.basicConfig(level=logging.WARNING)


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='screening_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _create_test_records():
    """创建测试交易记录"""
    base_date = datetime(2024, 1, 15)
    records = []

    for i in range(5):
        records.append({
            '唯一id': f'REC{i:04d}',
            '银行': '测试银行',
            '银行账号': '1234567890',
            '主体': '测试主体',
            '交易日期': (base_date + timedelta(days=i)).strftime('%Y-%m-%d'),
            '付款': -10000.0,
            '收款': None,
            '摘要': f'日常支出{i}',
            '对方户名': f'供应商{i+1}',
            '余额': 1000000.0 - i * 10000,
            '交易流水号': f'TXN{i:04d}',
        })

    records.append({
        '唯一id': 'REC_LARGE01',
        '银行': '测试银行',
        '银行账号': '1234567890',
        '主体': '测试主体',
        '交易日期': '2024-01-20',
        '付款': -600000.0,
        '收款': None,
        '摘要': '大额采购',
        '对方户名': '大供应商A',
        '余额': 400000.0,
        '交易流水号': 'TXN_LARGE01',
    })

    records.append({
        '唯一id': 'REC_LARGE02',
        '银行': '测试银行',
        '银行账号': '1234567890',
        '主体': '测试主体',
        '交易日期': '2024-01-21',
        '付款': None,
        '收款': 1200000.0,
        '摘要': '大额收款',
        '对方户名': '大客户B',
        '余额': 1600000.0,
        '交易流水号': 'TXN_LARGE02',
    })

    for i in range(60):
        records.append({
            '唯一id': f'REC_HIGH{i:04d}',
            '银行': '测试银行',
            '银行账号': '1234567890',
            '主体': '测试主体',
            '交易日期': '2024-01-25',
            '付款': -100.0,
            '收款': None,
            '摘要': f'小额支出{i}',
            '对方户名': f'零散商户{i}',
            '余额': 1000000.0 - i * 100,
            '交易流水号': f'TXN_HIGH{i:04d}',
        })

    for i in range(25):
        records.append({
            '唯一id': f'REC_FREQ{i:04d}',
            '银行': '测试银行',
            '银行账号': '1234567890',
            '主体': '测试主体',
            '交易日期': (base_date + timedelta(days=i)).strftime('%Y-%m-%d'),
            '付款': -5000.0,
            '收款': None,
            '摘要': f'采购付款{i}',
            '对方户名': '高频供应商X',
            '余额': 1000000.0 - i * 5000,
            '交易流水号': f'TXN_FREQ{i:04d}',
        })

    return records


class TestCounterpartyFrequencyRule:
    """测试对方户名频次规则"""

    def test_default_rule(self):
        rule = CounterpartyFrequencyRule()
        assert rule.name == ''
        assert rule.window_days == 30
        assert rule.min_frequency == 10
        assert rule.min_total_amount is None
        assert rule.enabled is True

    def test_custom_rule(self):
        rule = CounterpartyFrequencyRule(
            name='测试规则',
            window_days=7,
            min_frequency=5,
            min_total_amount=10000.0,
            enabled=False,
        )
        assert rule.name == '测试规则'
        assert rule.window_days == 7
        assert rule.min_frequency == 5
        assert rule.min_total_amount == 10000.0
        assert rule.enabled is False

    def test_to_dict_and_from_dict(self):
        rule = CounterpartyFrequencyRule(
            name='往返转换测试',
            window_days=15,
            min_frequency=8,
            min_total_amount=50000.0,
            enabled=True,
        )
        rule_dict = rule.to_dict()
        assert isinstance(rule_dict, dict)
        assert rule_dict['name'] == '往返转换测试'
        assert rule_dict['window_days'] == 15

        rule2 = CounterpartyFrequencyRule.from_dict(rule_dict)
        assert rule2.name == rule.name
        assert rule2.window_days == rule.window_days
        assert rule2.min_frequency == rule.min_frequency
        assert rule2.min_total_amount == rule.min_total_amount
        assert rule2.enabled == rule.enabled


class TestScreeningConfig:
    """测试筛查配置"""

    def test_default_config(self):
        config = ScreeningConfig.default()
        assert config.config_name == '默认风险筛查配置'
        assert config.enabled is True
        assert config.single_amount_threshold == 500000.0
        assert config.daily_count_threshold == 50
        assert len(config.counterparty_rules) == 2

    def test_empty_config(self):
        config = ScreeningConfig()
        assert config.single_amount_threshold is None
        assert config.daily_count_threshold is None
        assert len(config.counterparty_rules) == 0

    def test_custom_config(self):
        config = ScreeningConfig(
            config_name='自定义配置',
            single_amount_threshold=100000.0,
            daily_count_threshold=20,
            counterparty_rules=[
                CounterpartyFrequencyRule(name='规则1', window_days=7, min_frequency=10),
            ],
        )
        assert config.config_name == '自定义配置'
        assert config.single_amount_threshold == 100000.0
        assert config.daily_count_threshold == 20
        assert len(config.counterparty_rules) == 1

    def test_to_dict_and_from_dict(self):
        config = ScreeningConfig(
            config_name='转换测试',
            single_amount_threshold=200000.0,
            daily_count_threshold=30,
            counterparty_rules=[
                CounterpartyFrequencyRule(name='规则A', window_days=14, min_frequency=15, min_total_amount=10000.0),
            ],
        )
        config_dict = config.to_dict()
        assert isinstance(config_dict, dict)
        assert config_dict['config_name'] == '转换测试'
        assert config_dict['single_amount_threshold'] == 200000.0
        assert len(config_dict['counterparty_rules']) == 1

        config2 = ScreeningConfig.from_dict(config_dict)
        assert config2.config_name == config.config_name
        assert config2.single_amount_threshold == config.single_amount_threshold
        assert config2.daily_count_threshold == config.daily_count_threshold
        assert len(config2.counterparty_rules) == len(config.counterparty_rules)
        assert config2.counterparty_rules[0].name == config.counterparty_rules[0].name


class TestRiskAlert:
    """测试风险提示项"""

    def test_default_alert(self):
        alert = RiskAlert()
        assert alert.risk_type == ''
        assert alert.risk_level == 'medium'
        assert alert.review_status == 'pending'

    def test_custom_alert(self):
        alert = RiskAlert(
            risk_id='TEST001',
            risk_type=RiskType.LARGE_AMOUNT,
            risk_level='high',
            description='测试风险',
            transaction_id='TXN001',
            trade_date='2024-01-15',
            amount=1000000.0,
            counterpart='测试对方',
            bank='测试银行',
            subject='测试主体',
            matched_rule='测试规则',
            rule_details={'threshold': 500000, 'actual': 1000000},
        )
        assert alert.risk_id == 'TEST001'
        assert alert.risk_type == RiskType.LARGE_AMOUNT
        assert alert.risk_level == 'high'
        assert alert.amount == 1000000.0

    def test_to_dict_and_from_dict(self):
        alert = RiskAlert(
            risk_id='TEST002',
            risk_type=RiskType.HIGH_FREQUENCY,
            risk_level='medium',
            description='转换测试',
            rule_details={'key': 'value', 'number': 123},
        )
        alert_dict = alert.to_dict()
        assert isinstance(alert_dict, dict)
        assert alert_dict['risk_id'] == 'TEST002'
        assert 'rule_details' in alert_dict

        alert2 = RiskAlert.from_dict(alert_dict)
        assert alert2.risk_id == alert.risk_id
        assert alert2.risk_type == alert.risk_type
        assert alert2.rule_details == {'key': 'value', 'number': 123}


class TestScreeningResult:
    """测试筛查结果"""

    def test_empty_result(self):
        config = ScreeningConfig.default()
        result = ScreeningResult(config=config)
        assert result.total_records == 0
        assert result.alert_count == 0
        assert result.high_risk_count == 0
        assert result.medium_risk_count == 0
        assert result.low_risk_count == 0

    def test_result_with_alerts(self):
        config = ScreeningConfig.default()
        result = ScreeningResult(
            config=config,
            total_records=100,
            large_amount_alerts=[
                RiskAlert(risk_type=RiskType.LARGE_AMOUNT, risk_level='high'),
                RiskAlert(risk_type=RiskType.LARGE_AMOUNT, risk_level='medium'),
            ],
            high_frequency_alerts=[
                RiskAlert(risk_type=RiskType.HIGH_FREQUENCY, risk_level='low'),
            ],
            counterparty_alerts=[
                RiskAlert(risk_type=RiskType.COUNTERPARTY_FREQUENCY, risk_level='high'),
            ],
        )
        assert result.alert_count == 4
        assert result.high_risk_count == 2
        assert result.medium_risk_count == 1
        assert result.low_risk_count == 1
        assert len(result.all_alerts) == 4

    def test_to_dict(self):
        config = ScreeningConfig.default()
        result = ScreeningResult(
            config=config,
            total_records=100,
            large_amount_alerts=[RiskAlert(risk_type=RiskType.LARGE_AMOUNT)],
        )
        result_dict = result.to_dict()
        assert isinstance(result_dict, dict)
        assert result_dict['total_records'] == 100
        assert 'summary' in result_dict
        assert result_dict['summary']['total_alerts'] == 1


class TestScreenLargeAmount:
    """测试大额交易筛查"""

    def test_no_threshold(self):
        records = _create_test_records()
        alerts = screen_large_amount(records, None)
        assert len(alerts) == 0

    def test_zero_threshold(self):
        records = _create_test_records()
        alerts = screen_large_amount(records, 0)
        assert len(alerts) == 0

    def test_negative_threshold(self):
        records = _create_test_records()
        alerts = screen_large_amount(records, -1000)
        assert len(alerts) == 0

    def test_large_amount_detection(self):
        records = _create_test_records()
        alerts = screen_large_amount(records, 500000.0)

        large_alerts = [a for a in alerts if a.risk_type == RiskType.LARGE_AMOUNT]
        assert len(large_alerts) >= 2

        amount_alert = [a for a in large_alerts if a.transaction_id == 'TXN_LARGE01']
        assert len(amount_alert) == 1
        assert amount_alert[0].amount == 600000.0
        assert amount_alert[0].counterpart == '大供应商A'
        assert amount_alert[0].risk_level == 'low'

        receipt_alert = [a for a in large_alerts if a.transaction_id == 'TXN_LARGE02']
        assert len(receipt_alert) == 1
        assert receipt_alert[0].amount == 1200000.0
        assert receipt_alert[0].counterpart == '大客户B'
        assert receipt_alert[0].risk_level == 'medium'

    def test_risk_level_calculation(self):
        records = [
            {'唯一id': '1', '付款': -600000.0, '收款': None, '交易日期': '2024-01-01'},
            {'唯一id': '2', '付款': -1200000.0, '收款': None, '交易日期': '2024-01-02'},
            {'唯一id': '3', '付款': -3000000.0, '收款': None, '交易日期': '2024-01-03'},
        ]
        alerts = screen_large_amount(records, 500000.0)

        assert len(alerts) == 3
        assert alerts[0].risk_level == 'low'
        assert alerts[1].risk_level == 'medium'
        assert alerts[2].risk_level == 'high'

    def test_empty_records(self):
        alerts = screen_large_amount([], 500000.0)
        assert len(alerts) == 0


class TestScreenHighFrequency:
    """测试高频交易筛查"""

    def test_no_threshold(self):
        records = _create_test_records()
        alerts = screen_high_frequency(records, None)
        assert len(alerts) == 0

    def test_zero_threshold(self):
        records = _create_test_records()
        alerts = screen_high_frequency(records, 0)
        assert len(alerts) == 0

    def test_high_frequency_detection(self):
        records = _create_test_records()
        alerts = screen_high_frequency(records, 50)

        assert len(alerts) == 1
        alert = alerts[0]
        assert alert.risk_type == RiskType.HIGH_FREQUENCY
        assert alert.trade_date == '2024-01-25'
        assert alert.rule_details['actual_count'] == 61
        assert alert.risk_level == 'low'
        assert len(alert.related_transaction_ids) == 61

    def test_risk_level_calculation(self):
        base_records = []
        for i in range(10):
            base_records.append({'交易日期': '2024-01-01', '付款': -100, '银行': '测试'})
        for i in range(20):
            base_records.append({'交易日期': '2024-01-02', '付款': -100, '银行': '测试'})
        for i in range(50):
            base_records.append({'交易日期': '2024-01-03', '付款': -100, '银行': '测试'})

        alerts = screen_high_frequency(base_records, 10)
        assert len(alerts) == 3

        alerts_sorted = sorted(alerts, key=lambda a: a.trade_date)
        assert alerts_sorted[0].risk_level == 'low'
        assert alerts_sorted[1].risk_level == 'medium'
        assert alerts_sorted[2].risk_level == 'high'

    def test_empty_records(self):
        alerts = screen_high_frequency([], 50)
        assert len(alerts) == 0


class TestScreenCounterpartyFrequency:
    """测试对方户名频次筛查"""

    def test_no_rules(self):
        records = _create_test_records()
        alerts = screen_counterparty_frequency(records, [])
        assert len(alerts) == 0

    def test_disabled_rules(self):
        records = _create_test_records()
        rules = [
            CounterpartyFrequencyRule(
                name='禁用规则', window_days=30, min_frequency=10, enabled=False
            )
        ]
        alerts = screen_counterparty_frequency(records, rules)
        assert len(alerts) == 0

    def test_counterparty_frequency_detection(self):
        records = _create_test_records()
        rules = [
            CounterpartyFrequencyRule(
                name='30天内超过20笔',
                window_days=30,
                min_frequency=20,
                min_total_amount=50000.0,
                enabled=True,
            )
        ]
        alerts = screen_counterparty_frequency(records, rules)

        assert len(alerts) >= 1

        x_alerts = [a for a in alerts if a.counterpart == '高频供应商X']
        assert len(x_alerts) >= 1

        max_alert = max(x_alerts, key=lambda a: a.rule_details['actual_count'])
        assert max_alert.risk_type == RiskType.COUNTERPARTY_FREQUENCY
        assert max_alert.matched_rule == '30天内超过20笔'
        assert max_alert.rule_details['actual_count'] == 25
        assert max_alert.rule_details['actual_total_amount'] == 125000.0
        assert len(max_alert.related_transaction_ids) == 25

    def test_min_total_amount_filter(self):
        records = _create_test_records()
        rules = [
            CounterpartyFrequencyRule(
                name='高金额规则',
                window_days=30,
                min_frequency=5,
                min_total_amount=1000000.0,
                enabled=True,
            )
        ]
        alerts = screen_counterparty_frequency(records, rules)

        for alert in alerts:
            assert alert.amount >= 1000000.0

    def test_empty_records(self):
        rules = [CounterpartyFrequencyRule(name='测试', window_days=30, min_frequency=10)]
        alerts = screen_counterparty_frequency([], rules)
        assert len(alerts) == 0


class TestRunScreening:
    """测试完整筛查流程"""

    def test_disabled_config(self):
        records = _create_test_records()
        config = ScreeningConfig(enabled=False)
        result = run_screening(records, config)
        assert result.alert_count == 0

    def test_empty_records(self):
        config = ScreeningConfig.default()
        result = run_screening([], config)
        assert result.alert_count == 0
        assert result.total_records == 0

    def test_full_screening(self):
        records = _create_test_records()
        config = ScreeningConfig.default()
        source_info = {'数据来源': '测试数据'}
        result = run_screening(records, config, source_info)

        assert result.total_records == len(records)
        assert result.source_info == source_info
        assert len(result.large_amount_alerts) >= 2
        assert len(result.high_frequency_alerts) == 1
        assert len(result.counterparty_alerts) >= 1
        assert result.alert_count >= 4

    def test_partial_config(self):
        records = _create_test_records()
        config = ScreeningConfig(
            config_name='仅大额筛查',
            single_amount_threshold=500000.0,
        )
        result = run_screening(records, config)

        assert len(result.large_amount_alerts) >= 2
        assert len(result.high_frequency_alerts) == 0
        assert len(result.counterparty_alerts) == 0


class TestExportScreeningResult:
    """测试报告导出"""

    def test_no_alerts(self, tmp_dir):
        config = ScreeningConfig.default()
        result = ScreeningResult(config=config, total_records=100)

        output_path = os.path.join(tmp_dir, 'test_report.xlsx')
        result_path = export_screening_result(result, output_path)

        assert result_path is None
        assert not os.path.exists(output_path)

    def test_export_with_alerts(self, tmp_dir):
        config = ScreeningConfig.default()
        result = ScreeningResult(
            config=config,
            total_records=100,
            large_amount_alerts=[
                RiskAlert(
                    risk_id='LARGE001',
                    risk_type=RiskType.LARGE_AMOUNT,
                    risk_level='high',
                    description='大额交易测试',
                    transaction_id='TXN001',
                    trade_date='2024-01-15',
                    amount=600000.0,
                    counterpart='测试供应商',
                    bank='测试银行',
                    subject='测试主体',
                    matched_rule='单笔金额阈值 500,000.00 元',
                ),
            ],
            high_frequency_alerts=[
                RiskAlert(
                    risk_id='HIGH001',
                    risk_type=RiskType.HIGH_FREQUENCY,
                    risk_level='medium',
                    description='高频交易测试',
                    trade_date='2024-01-20',
                    amount=50000.0,
                    matched_rule='单日笔数阈值 50 笔',
                    related_transaction_ids=[f'TXN{i:03d}' for i in range(60)],
                ),
            ],
            counterparty_alerts=[
                RiskAlert(
                    risk_id='COUNTER001',
                    risk_type=RiskType.COUNTERPARTY_FREQUENCY,
                    risk_level='low',
                    description='对方户名频次异常测试',
                    counterpart='高频商户',
                    trade_date='2024-01-25',
                    amount=100000.0,
                    matched_rule='30天内交易超过20笔',
                    related_transaction_ids=[f'TXN{i:03d}' for i in range(25)],
                ),
            ],
            source_info={'数据来源': '测试数据'},
        )

        output_path = os.path.join(tmp_dir, 'test_report.xlsx')
        result_path = export_screening_result(result, output_path)

        assert result_path is not None
        assert os.path.exists(output_path)

        wb = openpyxl.load_workbook(output_path)

        assert '筛查总览' in wb.sheetnames
        assert '全部风险提示' in wb.sheetnames
        assert '大额交易风险' in wb.sheetnames
        assert '高频交易风险' in wb.sheetnames
        assert '对方户名频次异常' in wb.sheetnames
        assert '数据来源' in wb.sheetnames
        assert '筛查规则配置' in wb.sheetnames

        ws_all = wb['全部风险提示']
        assert ws_all.max_row == 4

        ws_summary = wb['筛查总览']
        assert ws_summary.max_row >= 10

        wb.close()

    def test_export_without_source_info(self, tmp_dir):
        config = ScreeningConfig.default()
        result = ScreeningResult(
            config=config,
            total_records=50,
            large_amount_alerts=[
                RiskAlert(
                    risk_id='LARGE001',
                    risk_type=RiskType.LARGE_AMOUNT,
                    risk_level='medium',
                    description='测试',
                    amount=600000.0,
                ),
            ],
        )

        output_path = os.path.join(tmp_dir, 'test_report2.xlsx')
        result_path = export_screening_result(result, output_path)

        assert result_path is not None
        wb = openpyxl.load_workbook(output_path)
        assert '数据来源' not in wb.sheetnames
        wb.close()


class TestScreeningFromRecords:
    """测试从交易记录列表筛查"""

    def test_with_default_config(self, tmp_dir):
        records = _create_test_records()
        result, report_path = screening_from_records(records, output_dir=tmp_dir)

        assert result is not None
        assert result.alert_count >= 4
        assert report_path is not None
        assert os.path.exists(report_path)
        assert '风险提示清单' in os.path.basename(report_path)

    def test_with_custom_config(self, tmp_dir):
        records = _create_test_records()
        config = ScreeningConfig(
            config_name='严格配置',
            single_amount_threshold=100000.0,
            daily_count_threshold=10,
        )
        result, report_path = screening_from_records(
            records, config=config, output_dir=tmp_dir
        )

        assert result is not None
        assert result.config.config_name == '严格配置'
        assert len(result.large_amount_alerts) >= 2
        assert len(result.high_frequency_alerts) == 1

    def test_no_alerts(self, tmp_dir):
        records = [
            {'唯一id': '1', '付款': -1000, '交易日期': '2024-01-01'},
            {'唯一id': '2', '收款': 2000, '交易日期': '2024-01-02'},
        ]
        config = ScreeningConfig(
            single_amount_threshold=100000.0,
            daily_count_threshold=100,
        )
        result, report_path = screening_from_records(
            records, config=config, output_dir=tmp_dir
        )

        assert result is not None
        assert result.alert_count == 0
        assert report_path is None


class TestRiskTypeLabels:
    """测试风险类型标签"""

    def test_risk_type_labels(self):
        assert RiskType.LABELS[RiskType.LARGE_AMOUNT] == '大额交易'
        assert RiskType.LABELS[RiskType.HIGH_FREQUENCY] == '高频交易'
        assert RiskType.LABELS[RiskType.COUNTERPARTY_FREQUENCY] == '对方户名频次异常'


class TestHelperFunctions:
    """测试辅助函数"""

    def test_to_float(self):
        assert ts.to_float(100) == 100.0
        assert ts.to_float('123.45') == 123.45
        assert ts.to_float(None) is None
        assert ts.to_float('abc') is None
        assert ts.to_float('') is None

    def test_get_amount(self):
        assert ts._get_amount({'付款': -50000}) == 50000.0
        assert ts._get_amount({'收款': 80000}) == 80000.0
        assert ts._get_amount({'付款': 0, '收款': 0}) is None
        assert ts._get_amount({}) is None

    def test_get_trade_date(self):
        assert ts._get_trade_date({'交易日期': '2024-01-15'}) == '2024-01-15'
        assert ts._get_trade_date({'交易日期': '2024-01-15 10:30:00'}) == '2024-01-15'
        assert ts._get_trade_date({'交易日期': datetime(2024, 1, 15)}) == '2024-01-15'
        assert ts._get_trade_date({}) is None

    def test_get_counterpart(self):
        assert ts._get_counterpart({'对方户名': '测试公司'}) == '测试公司'
        assert ts._get_counterpart({'对方户名': '  测试公司  '}) == '测试公司'
        assert ts._get_counterpart({}) is None

    def test_record_to_dict(self):
        class TestRecord:
            def to_dict(self):
                return {'key': 'value'}

        assert ts._record_to_dict({'a': 1}) == {'a': 1}
        assert ts._record_to_dict(TestRecord()) == {'key': 'value'}


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
