import os
import tempfile
import shutil

import openpyxl
import pandas as pd
import pytest

from conftest import _create_lookup_table
import bankcheck


SUMMARY_COLUMNS = [
    '唯一id', '银行', '银行账号', '主体', '交易日期',
    '付款', '收款', '摘要', '对方户名', '余额', '交易流水号',
]


def _build_summary_df(rows):
    return pd.DataFrame(rows, columns=SUMMARY_COLUMNS)


def _save_summary_excel(df, path):
    df.to_excel(path, index=False, engine='openpyxl')
    return path


def _base_rows():
    return [
        {
            '唯一id': bankcheck.generate_unique_id(),
            '银行': '北京银行',
            '银行账号': '01090312345678901',
            '主体': '北京XX科技有限公司',
            '交易日期': '2024-01-05',
            '付款': -50000.0,
            '收款': None,
            '摘要': '采购付款',
            '对方户名': '供应商A公司',
            '余额': 1500000,
            '交易流水号': 'BJ20240105001',
        },
        {
            '唯一id': bankcheck.generate_unique_id(),
            '银行': '北京银行',
            '银行账号': '01090312345678901',
            '主体': '北京XX科技有限公司',
            '交易日期': '2024-01-10',
            '付款': None,
            '收款': 80000.0,
            '摘要': '销售收款',
            '对方户名': '客户B公司',
            '余额': 1580000,
            '交易流水号': 'BJ20240110002',
        },
    ]


class TestMakeMatchKey:
    def test_with_transaction_id(self):
        row = {'银行账号': '01090312345678901', '交易流水号': 'BJ20240105001'}
        key = bankcheck._make_match_key(row)
        assert key == '1090312345678901::BJ20240105001'

    def test_without_transaction_id_fallback(self):
        row = {
            '银行账号': '01090312345678901',
            '交易流水号': None,
            '交易日期': '2024-01-05',
            '付款': -50000.0,
            '收款': None,
        }
        key = bankcheck._make_match_key(row)
        assert '1090312345678901' in key
        assert '2024-01-05' in key

    def test_empty_transaction_id_fallback(self):
        row = {
            '银行账号': '01090312345678901',
            '交易流水号': '  ',
            '交易日期': '2024-01-05',
            '付款': -50000.0,
            '收款': None,
        }
        key = bankcheck._make_match_key(row)
        assert '2024-01-05' in key


class TestValuesEqual:
    def test_amount_equal(self):
        assert bankcheck._values_equal(50000.0, 50000.0, '付款') is True

    def test_amount_close(self):
        assert bankcheck._values_equal(50000.001, 50000.002, '付款') is True

    def test_amount_different(self):
        assert bankcheck._values_equal(50000.0, 60000.0, '付款') is False

    def test_amount_none_both(self):
        assert bankcheck._values_equal(None, None, '收款') is True

    def test_amount_none_one(self):
        assert bankcheck._values_equal(None, 100.0, '收款') is False

    def test_string_equal(self):
        assert bankcheck._values_equal('abc', 'abc', '摘要') is True

    def test_string_different(self):
        assert bankcheck._values_equal('abc', 'def', '摘要') is False

    def test_string_strip(self):
        assert bankcheck._values_equal(' abc ', 'abc', '摘要') is True


class TestDiffTransactions:
    def test_identical_tables(self):
        rows = _base_rows()
        old_df = _build_summary_df(rows)
        new_df = _build_summary_df(rows)

        result = bankcheck.diff_transactions(old_df, new_df)

        assert result.unchanged_count == 2
        assert result.added_count == 0
        assert result.deleted_count == 0
        assert result.changed_count == 0

    def test_added_transaction(self):
        old_rows = _base_rows()
        new_rows = old_rows + [{
            '唯一id': bankcheck.generate_unique_id(),
            '银行': '东亚银行',
            '银行账号': '38812345678',
            '主体': '上海YY贸易有限公司',
            '交易日期': '2024-01-15',
            '付款': None,
            '收款': 30000.0,
            '摘要': '服务费',
            '对方户名': '客户C公司',
            '余额': 545000,
            '交易流水号': 'EA20240115003',
        }]
        old_df = _build_summary_df(old_rows)
        new_df = _build_summary_df(new_rows)

        result = bankcheck.diff_transactions(old_df, new_df)

        assert result.added_count == 1
        assert result.unchanged_count == 2
        added_records = [r for r in result.records if r.change_type == '新增']
        assert added_records[0].new_row['交易流水号'] == 'EA20240115003'

    def test_deleted_transaction(self):
        old_rows = _base_rows()
        new_rows = [old_rows[0]]
        old_df = _build_summary_df(old_rows)
        new_df = _build_summary_df(new_rows)

        result = bankcheck.diff_transactions(old_df, new_df)

        assert result.deleted_count == 1
        assert result.unchanged_count == 1
        deleted_records = [r for r in result.records if r.change_type == '删除']
        assert deleted_records[0].old_row['交易流水号'] == 'BJ20240110002'

    def test_amount_changed(self):
        old_rows = _base_rows()
        new_rows = _base_rows()
        new_rows[1] = dict(old_rows[1])
        new_rows[1]['收款'] = 90000.0
        new_rows[1]['余额'] = 1590000

        old_df = _build_summary_df(old_rows)
        new_df = _build_summary_df(new_rows)

        result = bankcheck.diff_transactions(old_df, new_df)

        assert result.changed_count == 1
        assert result.unchanged_count == 1
        changed = [r for r in result.records if r.change_type == '变更']
        assert '收款' in changed[0].changed_fields
        assert '余额' in changed[0].changed_fields

    def test_non_amount_field_changed(self):
        old_rows = _base_rows()
        new_rows = _base_rows()
        new_rows[0] = dict(old_rows[0])
        new_rows[0]['摘要'] = '采购退款'

        old_df = _build_summary_df(old_rows)
        new_df = _build_summary_df(new_rows)

        result = bankcheck.diff_transactions(old_df, new_df)

        assert result.changed_count == 1
        changed = [r for r in result.records if r.change_type == '变更']
        assert '摘要' in changed[0].changed_fields

    def test_empty_tables(self):
        old_df = pd.DataFrame(columns=SUMMARY_COLUMNS)
        new_df = pd.DataFrame(columns=SUMMARY_COLUMNS)

        result = bankcheck.diff_transactions(old_df, new_df)

        assert len(result.records) == 0
        assert result.added_count == 0
        assert result.deleted_count == 0

    def test_old_empty_new_has_data(self):
        old_df = pd.DataFrame(columns=SUMMARY_COLUMNS)
        new_rows = _base_rows()
        new_df = _build_summary_df(new_rows)

        result = bankcheck.diff_transactions(old_df, new_df)

        assert result.added_count == 2

    def test_old_has_data_new_empty(self):
        old_rows = _base_rows()
        old_df = _build_summary_df(old_rows)
        new_df = pd.DataFrame(columns=SUMMARY_COLUMNS)

        result = bankcheck.diff_transactions(old_df, new_df)

        assert result.deleted_count == 2


class TestDiffRecord:
    def test_added_to_flat_dict(self):
        rec = bankcheck.DiffRecord(
            change_type='新增',
            match_key='key1',
            new_row={'银行': '北京银行', '交易流水号': 'BJ001', '付款': -50000.0,
                     '收款': None, '余额': 1500000, '银行账号': '123',
                     '主体': 'A公司', '交易日期': '2024-01-05',
                     '摘要': '付款', '对方户名': 'B公司'},
        )
        flat = rec.to_flat_dict()
        assert flat['变更类型'] == '新增'
        assert flat['银行'] == '北京银行'
        assert flat['变更明细'] == ''

    def test_changed_to_flat_dict(self):
        rec = bankcheck.DiffRecord(
            change_type='变更',
            match_key='key1',
            old_row={'银行': '北京银行', '交易流水号': 'BJ001', '付款': -50000.0,
                     '收款': None, '余额': 1500000, '银行账号': '123',
                     '主体': 'A公司', '交易日期': '2024-01-05',
                     '摘要': '付款', '对方户名': 'B公司'},
            new_row={'银行': '北京银行', '交易流水号': 'BJ001', '付款': -60000.0,
                     '收款': None, '余额': 1490000, '银行账号': '123',
                     '主体': 'A公司', '交易日期': '2024-01-05',
                     '摘要': '付款', '对方户名': 'B公司'},
            changed_fields=['付款', '余额'],
        )
        flat = rec.to_flat_dict()
        assert flat['变更类型'] == '变更'
        assert '付款' in flat['变更明细']
        assert '余额' in flat['变更明细']
        assert flat['付款(旧)'] == -50000.0
        assert flat['付款(新)'] == -60000.0

    def test_deleted_to_flat_dict(self):
        rec = bankcheck.DiffRecord(
            change_type='删除',
            match_key='key1',
            old_row={'银行': '北京银行', '交易流水号': 'BJ001', '付款': -50000.0,
                     '收款': None, '余额': 1500000, '银行账号': '123',
                     '主体': 'A公司', '交易日期': '2024-01-05',
                     '摘要': '付款', '对方户名': 'B公司'},
        )
        flat = rec.to_flat_dict()
        assert flat['变更类型'] == '删除'
        assert flat['银行'] == '北京银行'


class TestExportDiffResult:
    def test_export_creates_file(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            rows = _base_rows()
            old_df = _build_summary_df(rows)
            new_df = _build_summary_df(rows[:1])

            diff_result = bankcheck.diff_transactions(old_df, new_df)
            output_path = os.path.join(tmp_dir, 'diff_output.xlsx')

            result_path = bankcheck.export_diff_result(diff_result, output_path)

            assert result_path is not None
            assert os.path.exists(result_path)
        finally:
            shutil.rmtree(tmp_dir)

    def test_export_has_correct_columns(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            rows = _base_rows()
            new_rows = _base_rows()
            new_rows[1] = dict(rows[1])
            new_rows[1]['收款'] = 90000.0

            old_df = _build_summary_df(rows)
            new_df = _build_summary_df(new_rows)

            diff_result = bankcheck.diff_transactions(old_df, new_df)
            output_path = os.path.join(tmp_dir, 'diff_output.xlsx')

            bankcheck.export_diff_result(diff_result, output_path)

            df_read = pd.read_excel(output_path, engine='openpyxl')
            assert '变更类型' in df_read.columns
            assert '变更明细' in df_read.columns
        finally:
            shutil.rmtree(tmp_dir)

    def test_export_highlight_colors(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            old_rows = _base_rows()
            new_rows = _base_rows()
            new_rows.append({
                '唯一id': bankcheck.generate_unique_id(),
                '银行': '东亚银行',
                '银行账号': '38812345678',
                '主体': '上海YY贸易有限公司',
                '交易日期': '2024-01-15',
                '付款': None,
                '收款': 30000.0,
                '摘要': '服务费',
                '对方户名': '客户C公司',
                '余额': 545000,
                '交易流水号': 'EA20240115003',
            })

            old_df = _build_summary_df(old_rows[:1])
            new_df = _build_summary_df(new_rows)

            diff_result = bankcheck.diff_transactions(old_df, new_df)
            output_path = os.path.join(tmp_dir, 'diff_output.xlsx')

            bankcheck.export_diff_result(diff_result, output_path)

            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            green_found = False
            for row_idx in range(2, ws.max_row + 1):
                cell_type = ws.cell(row=row_idx, column=1).value
                cell_fill = ws.cell(row=row_idx, column=1).fill
                if cell_type == '新增':
                    assert cell_fill.start_color.rgb == '00C6EFCE'
                    green_found = True
            wb.close()
            assert green_found
        finally:
            shutil.rmtree(tmp_dir)

    def test_export_empty_result(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            diff_result = bankcheck.DiffResult()
            output_path = os.path.join(tmp_dir, 'diff_output.xlsx')

            result = bankcheck.export_diff_result(diff_result, output_path)

            assert result is None
            assert not os.path.exists(output_path)
        finally:
            shutil.rmtree(tmp_dir)


class TestRunDiff:
    def test_run_diff_creates_output(self):
        tmp_dir = tempfile.mkdtemp()
        try:
            rows = _base_rows()
            old_df = _build_summary_df(rows)
            new_df = _build_summary_df(rows[:1])

            old_path = os.path.join(tmp_dir, 'old.xlsx')
            new_path = os.path.join(tmp_dir, 'new.xlsx')
            _save_summary_excel(old_df, old_path)
            _save_summary_excel(new_df, new_path)

            output_dir = os.path.join(tmp_dir, 'output')
            result = bankcheck.run_diff(old_path, new_path, output_dir)

            assert result.output_path is not None
            assert os.path.exists(result.output_path)
            assert result.deleted_count == 1
            assert result.unchanged_count == 1
        finally:
            shutil.rmtree(tmp_dir)

    def test_run_diff_file_not_found(self):
        with pytest.raises(FileNotFoundError):
            bankcheck.run_diff('/nonexistent/old.xlsx', '/nonexistent/new.xlsx')


class TestFormatDiffMessage:
    def test_with_records(self):
        result = bankcheck.DiffResult(
            records=[bankcheck.DiffRecord('新增', 'k1')],
            added_count=1, deleted_count=0, changed_count=0, unchanged_count=0,
            output_path='/path/to/diff.xlsx',
        )
        msg = bankcheck.format_diff_message(result)
        assert '新增交易：1' in msg
        assert '总记录数：1' in msg
        assert '/path/to/diff.xlsx' in msg

    def test_empty_records(self):
        result = bankcheck.DiffResult()
        msg = bankcheck.format_diff_message(result)
        assert '无数据可对比' in msg
