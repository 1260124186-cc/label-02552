"""
派生列功能单元测试
测试 apply_derived_columns、get_enabled_derived_columns、内置计算器，
以及 merge_and_export_summary 与派生列的集成。
"""
import os
import sys
import tempfile
import shutil

import openpyxl
import pandas as pd
import pytest
import yaml

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='derived_columns_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


@pytest.fixture
def sample_records():
    return [
        {
            '唯一id': 'D001',
            '银行': '北京银行',
            '银行账号': '01090312345678901',
            '主体': '测试公司A',
            '交易日期': '2024-01-05',
            '付款': -50000.0,
            '收款': None,
            '摘要': '采购付款',
            '对方户名': '供应商X',
            '余额': 1500000.0,
            '交易流水号': 'BJ20240105001',
            '异常标记': '正常',
            '异常详情': '',
        },
        {
            '唯一id': 'D002',
            '银行': '东亚银行',
            '银行账号': '38812345678',
            '主体': '测试公司B',
            '交易日期': '2024-01-10',
            '付款': None,
            '收款': 80000.0,
            '摘要': '销售收款',
            '对方户名': '客户Y',
            '余额': 580000.0,
            '交易流水号': 'EA20240110002',
            '异常标记': '正常',
            '异常详情': '',
        },
        {
            '唯一id': 'D003',
            '银行': '北京银行',
            '银行账号': '01090312345678901',
            '主体': '测试公司A',
            '交易日期': '2024-02-15',
            '付款': -30000.0,
            '收款': 10000.0,
            '摘要': '部分退款',
            '对方户名': '供应商X',
            '余额': 1480000.0,
            '交易流水号': 'BJ20240215003',
            '异常标记': '正常',
            '异常详情': '',
        },
        {
            '唯一id': 'D004',
            '银行': '工商银行',
            '银行账号': '6222021234567890',
            '主体': '测试公司C',
            '交易日期': '2024-03-01',
            '付款': None,
            '收款': None,
            '摘要': '备注记录',
            '对方户名': '',
            '余额': 200000.0,
            '交易流水号': 'IC20240301004',
            '异常标记': '正常',
            '异常详情': '',
        },
    ]


class TestDerivedColumnCalculators:
    """测试内置计算器"""

    def test_net_amount_payment_only(self):
        rec = {'付款': -50000.0, '收款': None}
        assert bankcheck._calc_net_amount(rec) == -50000.0

    def test_net_amount_receipt_only(self):
        rec = {'付款': None, '收款': 80000.0}
        assert bankcheck._calc_net_amount(rec) == 80000.0

    def test_net_amount_both(self):
        rec = {'付款': -30000.0, '收款': 10000.0}
        assert bankcheck._calc_net_amount(rec) == -20000.0

    def test_net_amount_none(self):
        rec = {'付款': None, '收款': None}
        assert bankcheck._calc_net_amount(rec) is None

    def test_net_amount_zero(self):
        rec = {'付款': 0.0, '收款': 0.0}
        assert bankcheck._calc_net_amount(rec) == 0.0

    def test_transaction_direction_income(self):
        rec = {'付款': None, '收款': 80000.0}
        assert bankcheck._calc_transaction_direction(rec) == '收入'

    def test_transaction_direction_expense(self):
        rec = {'付款': -50000.0, '收款': None}
        assert bankcheck._calc_transaction_direction(rec) == '支出'

    def test_transaction_direction_both(self):
        rec = {'付款': -30000.0, '收款': 10000.0}
        assert bankcheck._calc_transaction_direction(rec) == '收支'

    def test_transaction_direction_none(self):
        rec = {'付款': None, '收款': None}
        assert bankcheck._calc_transaction_direction(rec) == '无流向'

    def test_transaction_direction_zero_values(self):
        rec = {'付款': 0.0, '收款': 0.0}
        assert bankcheck._calc_transaction_direction(rec) == '无流向'

    def test_accounting_period_valid_date(self):
        rec = {'交易日期': '2024-01-15'}
        result = bankcheck._calc_accounting_period(rec)
        assert result is not None
        assert '2024' in result

    def test_accounting_period_no_date(self):
        rec = {'交易日期': None}
        assert bankcheck._calc_accounting_period(rec) == '未知期间'

    def test_accounting_period_invalid_date(self):
        rec = {'交易日期': 'invalid'}
        assert bankcheck._calc_accounting_period(rec) == '未知期间'


class TestGetEnabledDerivedColumns:
    """测试 get_enabled_derived_columns"""

    def test_default_config_all_disabled(self):
        result = bankcheck.get_enabled_derived_columns(bankcheck.DEFAULT_SUMMARY_CONFIG)
        assert result == []

    def test_enabled_net_amount(self):
        config = {
            'derived_columns': [
                {'name': '净额', 'calculator': 'net_amount', 'enabled': True},
                {'name': '交易方向', 'calculator': 'transaction_direction', 'enabled': False},
            ]
        }
        result = bankcheck.get_enabled_derived_columns(config)
        assert len(result) == 1
        assert result[0]['name'] == '净额'

    def test_multiple_enabled(self):
        config = {
            'derived_columns': [
                {'name': '净额', 'calculator': 'net_amount', 'enabled': True},
                {'name': '交易方向', 'calculator': 'transaction_direction', 'enabled': True},
            ]
        }
        result = bankcheck.get_enabled_derived_columns(config)
        assert len(result) == 2

    def test_empty_config(self):
        result = bankcheck.get_enabled_derived_columns({})
        assert result == []

    def test_none_uses_default(self):
        result = bankcheck.get_enabled_derived_columns(None)
        assert result == []


class TestApplyDerivedColumns:
    """测试 apply_derived_columns"""

    def test_no_enabled_columns(self, sample_records):
        config = bankcheck._deep_copy_default_summary_config()
        result = bankcheck.apply_derived_columns(sample_records, config)
        for rec in result:
            assert '净额' not in rec
            assert '交易方向' not in rec

    def test_apply_net_amount(self, sample_records):
        config = {
            'derived_columns': [
                {'name': '净额', 'calculator': 'net_amount', 'enabled': True},
            ]
        }
        result = bankcheck.apply_derived_columns(sample_records, config)
        assert result[0]['净额'] == -50000.0
        assert result[1]['净额'] == 80000.0
        assert result[2]['净额'] == -20000.0
        assert result[3]['净额'] is None

    def test_apply_transaction_direction(self, sample_records):
        config = {
            'derived_columns': [
                {'name': '交易方向', 'calculator': 'transaction_direction', 'enabled': True},
            ]
        }
        result = bankcheck.apply_derived_columns(sample_records, config)
        assert result[0]['交易方向'] == '支出'
        assert result[1]['交易方向'] == '收入'
        assert result[2]['交易方向'] == '收支'
        assert result[3]['交易方向'] == '无流向'

    def test_apply_multiple_derived(self, sample_records):
        config = {
            'derived_columns': [
                {'name': '净额', 'calculator': 'net_amount', 'enabled': True},
                {'name': '交易方向', 'calculator': 'transaction_direction', 'enabled': True},
            ]
        }
        result = bankcheck.apply_derived_columns(sample_records, config)
        assert '净额' in result[0]
        assert '交易方向' in result[0]

    def test_empty_records(self):
        config = {
            'derived_columns': [
                {'name': '净额', 'calculator': 'net_amount', 'enabled': True},
            ]
        }
        result = bankcheck.apply_derived_columns([], config)
        assert result == []

    def test_unknown_calculator_skipped(self, sample_records):
        config = {
            'derived_columns': [
                {'name': '自定义列', 'calculator': 'nonexistent_calc', 'enabled': True},
            ]
        }
        result = bankcheck.apply_derived_columns(sample_records, config)
        for rec in result:
            assert rec.get('自定义列') is None

    def test_apply_accounting_period(self, sample_records):
        config = {
            'derived_columns': [
                {'name': '会计期间', 'calculator': 'accounting_period', 'enabled': True},
            ]
        }
        result = bankcheck.apply_derived_columns(sample_records, config)
        assert result[0]['会计期间'] is not None
        assert '2024' in result[0]['会计期间']


class TestGetSummaryColumnsWithDerived:
    """测试 get_summary_columns 包含派生列"""

    def test_derived_columns_included_when_enabled(self, sample_records):
        config = {
            'derived_columns': [
                {'name': '净额', 'calculator': 'net_amount', 'enabled': True},
                {'name': '交易方向', 'calculator': 'transaction_direction', 'enabled': True},
            ],
            'columns': {
                'order': list(bankcheck.DEFAULT_SUMMARY_CONFIG['columns']['order']),
                'enabled': dict(bankcheck.DEFAULT_SUMMARY_CONFIG['columns']['enabled']),
            },
        }
        columns = bankcheck.get_summary_columns(sample_records, config=config)
        assert '净额' in columns
        assert '交易方向' in columns

    def test_derived_columns_excluded_when_disabled(self, sample_records):
        config = bankcheck._deep_copy_default_summary_config()
        columns = bankcheck.get_summary_columns(sample_records, config=config)
        assert '净额' not in columns
        assert '交易方向' not in columns

    def test_derived_respects_column_enabled(self, sample_records):
        config = {
            'derived_columns': [
                {'name': '净额', 'calculator': 'net_amount', 'enabled': True},
                {'name': '交易方向', 'calculator': 'transaction_direction', 'enabled': True},
            ],
            'columns': {
                'order': list(bankcheck.DEFAULT_SUMMARY_CONFIG['columns']['order']) + ['净额', '交易方向'],
                'enabled': {
                    **bankcheck.DEFAULT_SUMMARY_CONFIG['columns']['enabled'],
                    '净额': True,
                    '交易方向': False,
                },
            },
        }
        columns = bankcheck.get_summary_columns(sample_records, config=config)
        assert '净额' in columns
        assert '交易方向' not in columns

    def test_derived_column_order_configurable(self, sample_records):
        config = {
            'derived_columns': [
                {'name': '净额', 'calculator': 'net_amount', 'enabled': True},
            ],
            'columns': {
                'order': ['交易日期', '净额', '付款', '收款'],
                'enabled': {
                    '交易日期': True, '净额': True, '付款': True, '收款': True,
                    '唯一id': False, '银行': False, '银行账号': False,
                    '主体': False, '摘要': False, '对方户名': False,
                    '对方账号': False, '余额': False, '交易流水号': False,
                    '票据号': False, '结算号': False, '凭证号': False,
                    '异常标记': False, '异常详情': False,
                },
            },
        }
        columns = bankcheck.get_summary_columns(sample_records, config=config)
        assert columns.index('净额') == 1
        assert columns[0] == '交易日期'


class TestMergeExportWithDerivedColumns:
    """测试 merge_and_export_summary 集成派生列"""

    def test_export_with_net_amount(self, tmp_dir, sample_records):
        custom_config = {
            'derived_columns': [
                {'name': '净额', 'calculator': 'net_amount', 'enabled': True},
            ],
            'columns': {
                'order': ['交易日期', '付款', '收款', '净额'],
                'enabled': {
                    '交易日期': True, '付款': True, '收款': True, '净额': True,
                    '唯一id': False, '银行': False, '银行账号': False,
                    '主体': False, '摘要': False, '对方户名': False,
                    '对方账号': False, '余额': False, '交易流水号': False,
                    '票据号': False, '结算号': False, '凭证号': False,
                    '异常标记': False, '异常详情': False,
                },
            },
            'excel_style': {
                'freeze_header': True,
                'auto_column_width': {'enabled': False},
            },
        }

        output_path = bankcheck.merge_and_export_summary(
            existing_records=[],
            incremental_rows=sample_records,
            script_dir=tmp_dir,
            output_dir=tmp_dir,
            config=custom_config,
        )

        assert output_path is not None
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert '净额' in headers
        net_col_idx = headers.index('净额') + 1

        assert ws.cell(row=2, column=net_col_idx).value == -50000.0
        assert ws.cell(row=3, column=net_col_idx).value == 80000.0
        assert ws.cell(row=4, column=net_col_idx).value == -20000.0
        assert ws.cell(row=5, column=net_col_idx).value is None
        wb.close()

    def test_export_with_direction(self, tmp_dir, sample_records):
        custom_config = {
            'derived_columns': [
                {'name': '交易方向', 'calculator': 'transaction_direction', 'enabled': True},
            ],
            'columns': {
                'order': ['交易日期', '付款', '收款', '交易方向'],
                'enabled': {
                    '交易日期': True, '付款': True, '收款': True, '交易方向': True,
                    '唯一id': False, '银行': False, '银行账号': False,
                    '主体': False, '摘要': False, '对方户名': False,
                    '对方账号': False, '余额': False, '交易流水号': False,
                    '票据号': False, '结算号': False, '凭证号': False,
                    '异常标记': False, '异常详情': False,
                },
            },
            'excel_style': {
                'freeze_header': True,
                'auto_column_width': {'enabled': False},
            },
        }

        output_path = bankcheck.merge_and_export_summary(
            existing_records=[],
            incremental_rows=sample_records,
            script_dir=tmp_dir,
            output_dir=tmp_dir,
            config=custom_config,
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        dir_col_idx = headers.index('交易方向') + 1
        assert ws.cell(row=2, column=dir_col_idx).value == '支出'
        assert ws.cell(row=3, column=dir_col_idx).value == '收入'
        assert ws.cell(row=4, column=dir_col_idx).value == '收支'
        assert ws.cell(row=5, column=dir_col_idx).value == '无流向'
        wb.close()

    def test_export_without_derived(self, tmp_dir, sample_records):
        config = bankcheck._deep_copy_default_summary_config()
        output_path = bankcheck.merge_and_export_summary(
            existing_records=[],
            incremental_rows=sample_records,
            script_dir=tmp_dir,
            output_dir=tmp_dir,
            config=config,
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert '净额' not in headers
        assert '交易方向' not in headers
        wb.close()


class TestDerivedColumnsConfigPersistence:
    """测试派生列配置的持久化"""

    def test_save_and_load_config(self, tmp_dir):
        config_path = os.path.join(tmp_dir, bankcheck.SUMMARY_CONFIG_FILENAME)
        user_cfg = {
            'derived_columns': [
                {'name': '净额', 'calculator': 'net_amount', 'enabled': True},
                {'name': '交易方向', 'calculator': 'transaction_direction', 'enabled': True},
            ],
        }
        with open(config_path, 'w', encoding='utf-8') as f:
            yaml.dump(user_cfg, f, allow_unicode=True)

        loaded = bankcheck.load_summary_config(tmp_dir)
        derived = loaded.get('derived_columns', [])
        assert len(derived) == 2
        assert derived[0]['name'] == '净额'
        assert derived[0]['enabled'] is True
        assert derived[1]['name'] == '交易方向'
        assert derived[1]['enabled'] is True

    def test_merge_preserves_derived(self, tmp_dir):
        default = bankcheck._deep_copy_default_summary_config()
        user_cfg = {
            'derived_columns': [
                {'name': '净额', 'calculator': 'net_amount', 'enabled': True},
            ],
        }
        merged = bankcheck._merge_summary_config(default, user_cfg)
        derived = merged['derived_columns']
        assert len(derived) == 1
        assert derived[0]['enabled'] is True

    def test_merge_forward_compatible_new_key(self):
        default = bankcheck._deep_copy_default_summary_config()
        user_cfg = {
            'new_feature_key': {'some': 'value'},
        }
        merged = bankcheck._merge_summary_config(default, user_cfg)
        assert 'new_feature_key' in merged
        assert merged['new_feature_key'] == {'some': 'value'}


class TestDerivedColumnRegistry:
    """测试派生列计算器注册表"""

    def test_builtin_calculators_registered(self):
        assert 'net_amount' in bankcheck.DERIVED_COLUMN_CALCULATORS
        assert 'transaction_direction' in bankcheck.DERIVED_COLUMN_CALCULATORS
        assert 'accounting_period' in bankcheck.DERIVED_COLUMN_CALCULATORS

    def test_custom_calculator_registration(self):
        def _calc_test(rec):
            return 'test_value'

        bankcheck._register_derived_calculator('test_calc', _calc_test)
        assert 'test_calc' in bankcheck.DERIVED_COLUMN_CALCULATORS
        assert bankcheck.DERIVED_COLUMN_CALCULATORS['test_calc']({'a': 1}) == 'test_value'
        del bankcheck.DERIVED_COLUMN_CALCULATORS['test_calc']
