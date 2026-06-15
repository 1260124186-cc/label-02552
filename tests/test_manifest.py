import os
import sys
import json
import csv
import shutil
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck


@pytest.fixture(autouse=True)
def init_logging():
    bankcheck.setup_logging()


def _create_test_bank_excel(path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '交易明细'
    ws['A1'] = '北京银行交易明细'
    ws['B2'] = '01090312345678901'
    headers = ['序号', '交易日期', '币种', '支出金额', '收入金额', '余额',
               '对方户名', '对方账号', '对方行名', '凭证种类', '凭证号码',
               '摘要', '备注1', '备注2', '备注3', '交易流水号']
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    rows = [
        [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A公司', '622001234', '工商银行',
         '转账', '001', '采购付款', None, None, None, 'BJ20240105001'],
        [2, '2024-01-10', 'CNY', None, 80000, 1580000, '客户B公司', '622005678', '建设银行',
         '转账', '002', '销售收款', None, None, None, 'BJ20240110002'],
    ]
    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=4 + i, column=j + 1, value=val)
    wb.save(path)
    wb.close()
    return path


def _create_lookup_table(path):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '主体映射'
    ws['A1'] = '主体名称'
    ws['B1'] = '银行账号'
    ws['A2'] = '北京XX科技有限公司'
    ws['B2'] = '01090312345678901'
    wb.save(path)
    wb.close()
    return path


class TestFileProcessingRecord:
    def test_to_dict(self):
        record = bankcheck.FileProcessingRecord(
            file_path='/tmp/test.xlsx',
            file_name='test.xlsx',
            status='success',
            bank_type='北京银行',
            record_count=10,
            error_message=None,
            processed_at='2024-01-01 00:00:00',
        )
        d = record.to_dict()
        assert d['file_path'] == '/tmp/test.xlsx'
        assert d['file_name'] == 'test.xlsx'
        assert d['status'] == 'success'
        assert d['bank_type'] == '北京银行'
        assert d['record_count'] == 10
        assert d['error_message'] is None
        assert d['processed_at'] == '2024-01-01 00:00:00'


class TestGenerateProcessingManifest:
    def test_generate_json_and_csv(self):
        tmp_dir = tempfile.mkdtemp(prefix='manifest_test_')
        try:
            script_dir = os.path.join(os.path.dirname(__file__), '..', 'backend')
            records = [
                bankcheck.FileProcessingRecord(
                    file_path='/tmp/a.xlsx',
                    file_name='a.xlsx',
                    status='success',
                    bank_type='北京银行',
                    record_count=5,
                    error_message=None,
                    processed_at='2024-01-01 00:00:00',
                ),
                bankcheck.FileProcessingRecord(
                    file_path='/tmp/b.xlsx',
                    file_name='b.xlsx',
                    status='error',
                    bank_type='东亚银行',
                    record_count=0,
                    error_message='读取失败',
                    processed_at='2024-01-01 00:00:01',
                ),
                bankcheck.FileProcessingRecord(
                    file_path='/tmp/c.xlsx',
                    file_name='c.xlsx',
                    status='unrecognized',
                    bank_type=None,
                    record_count=0,
                    error_message='无法识别银行类型',
                    processed_at='2024-01-01 00:00:02',
                ),
            ]

            json_path, csv_path = bankcheck.generate_processing_manifest(
                records, script_dir, output_dir=tmp_dir,
                extra_metadata={'test': True},
            )

            assert json_path is not None
            assert csv_path is not None
            assert os.path.isfile(json_path)
            assert os.path.isfile(csv_path)

            with open(json_path, 'r', encoding='utf-8') as f:
                data = json.load(f)

            assert '_metadata' in data
            assert 'files' in data
            assert len(data['files']) == 3
            assert data['_metadata']['total_files'] == 3
            assert data['_metadata']['total_records_extracted'] == 5
            assert data['_metadata']['status_summary']['success'] == 1
            assert data['_metadata']['status_summary']['error'] == 1
            assert data['_metadata']['status_summary']['unrecognized'] == 1
            assert data['_metadata']['test'] is True

            with open(csv_path, 'r', encoding='utf-8-sig') as f:
                reader = csv.DictReader(f)
                rows = list(reader)

            assert len(rows) == 3
            assert rows[0]['file_name'] == 'a.xlsx'
            assert rows[0]['status'] == 'success'
            assert rows[0]['record_count'] == '5'
            assert rows[1]['file_name'] == 'b.xlsx'
            assert rows[1]['error_message'] == '读取失败'

        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

    def test_empty_records(self):
        tmp_dir = tempfile.mkdtemp(prefix='manifest_test_')
        try:
            script_dir = os.path.join(os.path.dirname(__file__), '..', 'backend')
            json_path, csv_path = bankcheck.generate_processing_manifest(
                [], script_dir, output_dir=tmp_dir,
            )
            assert json_path is None
            assert csv_path is None
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)

    def test_dry_run_skips_write(self):
        tmp_dir = tempfile.mkdtemp(prefix='manifest_test_')
        try:
            script_dir = os.path.join(os.path.dirname(__file__), '..', 'backend')
            records = [
                bankcheck.FileProcessingRecord(
                    file_path='/tmp/a.xlsx',
                    file_name='a.xlsx',
                    status='success',
                    bank_type='北京银行',
                    record_count=5,
                    error_message=None,
                    processed_at='2024-01-01 00:00:00',
                ),
            ]
            json_path, csv_path = bankcheck.generate_processing_manifest(
                records, script_dir, output_dir=tmp_dir, dry_run=True,
            )
            assert json_path is None
            assert csv_path is None
            assert len(os.listdir(tmp_dir)) == 0
        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)


class TestPipelineManifest:
    def test_run_pipeline_generates_manifest(self):
        tmp_dir = tempfile.mkdtemp(prefix='pipeline_manifest_test_')
        try:
            script_dir = os.path.join(os.path.dirname(__file__), '..', 'backend')

            input_folder = os.path.join(tmp_dir, 'input')
            os.makedirs(input_folder)

            _create_test_bank_excel(os.path.join(input_folder, '北京银行_流水.xlsx'))

            unrecognized_path = os.path.join(input_folder, '未知银行_文件.xlsx')
            import openpyxl
            wb = openpyxl.Workbook()
            wb.active['A1'] = 'test'
            wb.save(unrecognized_path)
            wb.close()

            _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

            try:
                result = bankcheck.run_pipeline(
                    input_folder, script_dir, incremental=False,
                    folder_strategy='in_place',
                )

                assert result.file_records is not None
                assert len(result.file_records) >= 2

                success_records = [r for r in result.file_records if r.status == 'success']
                unrecognized_records = [r for r in result.file_records if r.status == 'unrecognized']
                assert len(success_records) >= 1
                assert len(unrecognized_records) >= 1

                assert result.manifest_json_path is not None
                assert result.manifest_csv_path is not None
                assert os.path.isfile(result.manifest_json_path)
                assert os.path.isfile(result.manifest_csv_path)

                with open(result.manifest_json_path, 'r', encoding='utf-8') as f:
                    manifest = json.load(f)
                assert manifest['_metadata']['total_files'] >= 2
                assert len(manifest['files']) >= 2

                msg = bankcheck.format_result_message(result)
                assert '处理清单（机器可读）' in msg
                assert 'JSON' in msg or 'CSV' in msg

            finally:
                lookup_path = os.path.join(script_dir, '主体查找表.xlsx')
                if os.path.exists(lookup_path):
                    os.remove(lookup_path)

        finally:
            shutil.rmtree(tmp_dir, ignore_errors=True)
