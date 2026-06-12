import os
import sys
import shutil
import tempfile
import sqlite3
import json

import openpyxl
import pandas as pd
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck
from conftest import _create_beijing_bank_excel, _create_lookup_table


@pytest.fixture
def signature_script_dir(tmp_dir):
    """创建带签名密钥的脚本目录"""
    script_dir = os.path.join(tmp_dir, 'script_signature')
    os.makedirs(script_dir, exist_ok=True)
    _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))
    db_path = bankcheck.get_audit_db_path(script_dir)
    bankcheck.init_audit_db(db_path)
    return script_dir


class TestCryptographyAvailability:
    def test_cryptography_available(self):
        assert bankcheck.HAS_CRYPTOGRAPHY is True

    def test_signature_constants_defined(self):
        assert bankcheck.SIGNATURE_KEY_FILENAME == 'bankcheck_signing_key.pem'
        assert bankcheck.SIGNATURE_PUB_FILENAME == 'bankcheck_signing_key.pub.pem'
        assert bankcheck.SIGNATURE_ALGORITHM == 'RSA-SHA256'


class TestGenerateKeyPair:
    def test_generate_key_pair_creates_files(self, signature_script_dir):
        private_path, public_path = bankcheck.generate_signing_key_pair(signature_script_dir)

        assert os.path.exists(private_path)
        assert os.path.exists(public_path)
        assert private_path.endswith(bankcheck.SIGNATURE_KEY_FILENAME)
        assert public_path.endswith(bankcheck.SIGNATURE_PUB_FILENAME)

    def test_generate_key_pair_with_password(self, signature_script_dir):
        password = 'test_password_123'
        private_path, public_path = bankcheck.generate_signing_key_pair(
            signature_script_dir, password=password
        )

        assert os.path.exists(private_path)
        assert os.path.exists(public_path)

    def test_generate_key_pair_returns_paths(self, signature_script_dir):
        result = bankcheck.generate_signing_key_pair(signature_script_dir)
        assert isinstance(result, tuple)
        assert len(result) == 2
        assert all(isinstance(p, str) for p in result)

    def test_private_key_permissions(self, signature_script_dir):
        private_path, _ = bankcheck.generate_signing_key_pair(signature_script_dir)
        import stat
        file_stat = os.stat(private_path)
        permissions = stat.S_IMODE(file_stat.st_mode)
        assert permissions <= 0o600


class TestLoadKeys:
    def test_load_private_key_without_password(self, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        private_key = bankcheck.load_private_key(signature_script_dir)
        assert private_key is not None

    def test_load_private_key_with_password(self, signature_script_dir):
        password = 'secure_pass'
        bankcheck.generate_signing_key_pair(signature_script_dir, password=password)
        private_key = bankcheck.load_private_key(signature_script_dir, password=password)
        assert private_key is not None

    def test_load_private_key_wrong_password(self, signature_script_dir):
        password = 'correct_pass'
        bankcheck.generate_signing_key_pair(signature_script_dir, password=password)
        private_key = bankcheck.load_private_key(signature_script_dir, password='wrong_pass')
        assert private_key is None

    def test_load_private_key_no_key_file(self, tmp_dir):
        private_key = bankcheck.load_private_key(tmp_dir)
        assert private_key is None

    def test_load_public_key(self, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        public_key = bankcheck.load_public_key(signature_script_dir)
        assert public_key is not None

    def test_load_public_key_from_custom_path(self, signature_script_dir):
        _, public_path = bankcheck.generate_signing_key_pair(signature_script_dir)
        public_key = bankcheck.load_public_key(public_key_path=public_path)
        assert public_key is not None

    def test_load_public_key_no_file(self, tmp_dir):
        public_key = bankcheck.load_public_key(tmp_dir)
        assert public_key is None


class TestBuildSignaturePayload:
    def test_build_payload_contains_all_fields(self):
        payload_bytes = bankcheck.build_signature_payload(
            file_hash='abc123',
            username='test_user',
            operation_time='2024-01-01 12:00:00',
            input_directory='/test/path'
        )

        assert isinstance(payload_bytes, bytes)
        payload = json.loads(payload_bytes)
        assert payload['file_hash'] == 'abc123'
        assert payload['username'] == 'test_user'
        assert payload['operation_time'] == '2024-01-01 12:00:00'
        assert payload['input_directory'] == '/test/path'
        assert payload['algorithm'] == 'RSA-SHA256'

    def test_build_payload_with_extra_data(self):
        extra = {'record_count': 100, 'mode': 'incremental'}
        payload_bytes = bankcheck.build_signature_payload(
            file_hash='abc123',
            username='test_user',
            operation_time='2024-01-01 12:00:00',
            input_directory='/test/path',
            extra_data=extra
        )

        payload = json.loads(payload_bytes)
        assert payload['extra_data'] == extra

    def test_build_payload_consistent(self):
        kwargs = dict(
            file_hash='abc123',
            username='test_user',
            operation_time='2024-01-01 12:00:00',
            input_directory='/test/path'
        )
        p1 = bankcheck.build_signature_payload(**kwargs)
        p2 = bankcheck.build_signature_payload(**kwargs)
        assert p1 == p2


class TestSignAndVerify:
    def test_sign_and_verify_data(self, signature_script_dir):
        private_path, public_path = bankcheck.generate_signing_key_pair(signature_script_dir)
        private_key = bankcheck.load_private_key(signature_script_dir)
        public_key = bankcheck.load_public_key(signature_script_dir)

        data = b'test data for signing'
        signature = bankcheck.sign_data(private_key, data)

        assert signature is not None
        assert isinstance(signature, str)
        assert len(signature) > 0

        is_valid = bankcheck.verify_signature(public_key, signature, data)
        assert is_valid is True

    def test_verify_wrong_data_fails(self, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        private_key = bankcheck.load_private_key(signature_script_dir)
        public_key = bankcheck.load_public_key(signature_script_dir)

        data = b'original data'
        signature = bankcheck.sign_data(private_key, data)

        is_valid = bankcheck.verify_signature(public_key, signature, b'tampered data')
        assert is_valid is False

    def test_verify_wrong_signature_fails(self, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        public_key = bankcheck.load_public_key(signature_script_dir)

        data = b'test data'
        fake_signature = 'a' * 100 + '='

        is_valid = bankcheck.verify_signature(public_key, fake_signature, data)
        assert is_valid is False

    def test_sign_with_none_key_returns_none(self):
        signature = bankcheck.sign_data(None, b'data')
        assert signature is None

    def test_verify_with_none_key_returns_false(self):
        is_valid = bankcheck.verify_signature(None, 'sig', b'data')
        assert is_valid is False


class TestSignOutputFile:
    def _setup_test_file(self, tmp_dir):
        test_file = os.path.join(tmp_dir, 'test_output.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'test data'
        wb.save(test_file)
        wb.close()
        return test_file

    def test_sign_output_file_success(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        test_file = self._setup_test_file(tmp_dir)

        result = bankcheck.sign_output_file(
            output_path=test_file,
            script_dir=signature_script_dir,
            username='test_signer',
            input_directory='/input/path'
        )

        assert result is not None
        assert result['file_hash'] == bankcheck.compute_file_hash(test_file)
        assert result['username'] == 'test_signer'
        assert result['input_directory'] == '/input/path'
        assert result['algorithm'] == 'RSA-SHA256'
        assert 'signature' in result
        assert 'payload' in result
        assert 'signed_at' in result

    def test_sign_output_file_nonexistent_returns_none(self, signature_script_dir):
        result = bankcheck.sign_output_file(
            output_path='/nonexistent/file.xlsx',
            script_dir=signature_script_dir
        )
        assert result is None

    def test_sign_output_file_no_key_returns_none(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'no_key_dir')
        os.makedirs(script_dir)
        test_file = self._setup_test_file(tmp_dir)

        result = bankcheck.sign_output_file(
            output_path=test_file,
            script_dir=script_dir
        )
        assert result is None

    def test_verify_output_file_signature(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        test_file = self._setup_test_file(tmp_dir)

        signature_info = bankcheck.sign_output_file(
            output_path=test_file,
            script_dir=signature_script_dir,
            username='verifier',
            input_directory='/test'
        )

        is_valid = bankcheck.verify_output_file_signature(
            output_path=test_file,
            signature_info=signature_info,
            script_dir=signature_script_dir
        )
        assert is_valid is True

    def test_verify_tampered_file_fails(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        test_file = self._setup_test_file(tmp_dir)

        signature_info = bankcheck.sign_output_file(
            output_path=test_file,
            script_dir=signature_script_dir,
            username='test',
            input_directory='/test'
        )

        wb = openpyxl.load_workbook(test_file)
        wb.active['A1'] = 'tampered data'
        wb.save(test_file)
        wb.close()

        is_valid = bankcheck.verify_output_file_signature(
            output_path=test_file,
            signature_info=signature_info,
            script_dir=signature_script_dir
        )
        assert is_valid is False


class TestSignatureHelpers:
    def test_has_signing_key_false_when_no_key(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'no_key')
        os.makedirs(script_dir)
        assert bankcheck.has_signing_key(script_dir) is False

    def test_has_signing_key_true_when_key_exists(self, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        assert bankcheck.has_signing_key(signature_script_dir) is True

    def test_ensure_signing_key_auto_generates(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'auto_gen')
        os.makedirs(script_dir)
        result = bankcheck.ensure_signing_key(script_dir, auto_generate=True)
        assert result is True
        assert bankcheck.has_signing_key(script_dir) is True

    def test_ensure_signing_key_no_auto_generate(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'no_auto')
        os.makedirs(script_dir)
        result = bankcheck.ensure_signing_key(script_dir, auto_generate=False)
        assert result is False

    def test_get_signature_key_path(self, signature_script_dir):
        private_path = bankcheck.get_signature_key_path(signature_script_dir, public=False)
        public_path = bankcheck.get_signature_key_path(signature_script_dir, public=True)

        assert private_path.endswith(bankcheck.SIGNATURE_KEY_FILENAME)
        assert public_path.endswith(bankcheck.SIGNATURE_PUB_FILENAME)
        assert signature_script_dir in private_path
        assert signature_script_dir in public_path


class TestSignatureDatabase:
    def test_digital_signatures_table_created(self, signature_script_dir):
        db_path = bankcheck.get_audit_db_path(signature_script_dir)
        bankcheck._ensure_signature_tables(db_path)

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='digital_signatures'")
        table = cursor.fetchone()
        conn.close()

        assert table is not None

    def test_signature_table_columns(self, signature_script_dir):
        db_path = bankcheck.get_audit_db_path(signature_script_dir)
        bankcheck._ensure_signature_tables(db_path)

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(digital_signatures)")
        columns = {row[1] for row in cursor.fetchall()}
        conn.close()

        expected_cols = {
            'id', 'signature_id', 'audit_id', 'file_path', 'file_hash',
            'username', 'operation_time', 'input_directory', 'algorithm',
            'signature', 'payload', 'signed_at', 'public_key_path',
            'is_verified', 'verified_at'
        }
        assert expected_cols.issubset(columns)

    def test_signature_indexes_created(self, signature_script_dir):
        db_path = bankcheck.get_audit_db_path(signature_script_dir)
        bankcheck._ensure_signature_tables(db_path)

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='index'")
        indexes = {row[0] for row in cursor.fetchall()}
        conn.close()

        assert 'idx_signatures_audit' in indexes
        assert 'idx_signatures_file' in indexes
        assert 'idx_signatures_hash' in indexes
        assert 'idx_signatures_signed' in indexes


class TestSaveAndQuerySignatures:
    def _create_signature_info(self, test_file, username='test_user'):
        return {
            'file_hash': bankcheck.compute_file_hash(test_file),
            'username': username,
            'operation_time': '2024-01-01 12:00:00',
            'input_directory': '/test/path',
            'algorithm': 'RSA-SHA256',
            'signature': 'test_signature_abc123',
            'payload': '{"test": "data"}',
            'signed_at': '2024-01-01 12:00:01',
            'file_path': test_file,
        }

    def test_save_signature_record(self, tmp_dir, signature_script_dir):
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1, 2, 3]}).to_excel(test_file, index=False)

        signature_info = self._create_signature_info(test_file)
        signature_id = bankcheck.save_signature_record(
            signature_info, script_dir=signature_script_dir
        )

        assert signature_id is not None
        assert signature_id.startswith('SIG')

    def test_save_signature_with_audit_id(self, tmp_dir, signature_script_dir):
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(test_file, index=False)

        signature_info = self._create_signature_info(test_file)
        signature_id = bankcheck.save_signature_record(
            signature_info, audit_id='AUD123', script_dir=signature_script_dir
        )

        signatures = bankcheck.query_signatures(
            script_dir=signature_script_dir, audit_id='AUD123'
        )
        assert len(signatures) == 1
        assert signatures[0]['signature_id'] == signature_id
        assert signatures[0]['audit_id'] == 'AUD123'

    def test_query_signatures_by_file_path(self, tmp_dir, signature_script_dir):
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(test_file, index=False)

        signature_info = self._create_signature_info(test_file)
        bankcheck.save_signature_record(signature_info, script_dir=signature_script_dir)

        signatures = bankcheck.query_signatures(
            script_dir=signature_script_dir, file_path=test_file
        )
        assert len(signatures) >= 1
        assert signatures[0]['file_path'] == test_file

    def test_query_signatures_by_username(self, tmp_dir, signature_script_dir):
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(test_file, index=False)

        signature_info = self._create_signature_info(test_file, username='specific_user')
        bankcheck.save_signature_record(signature_info, script_dir=signature_script_dir)

        signatures = bankcheck.query_signatures(
            script_dir=signature_script_dir, username='specific_user'
        )
        assert len(signatures) >= 1
        assert all(s['username'] == 'specific_user' for s in signatures)

    def test_query_signatures_by_file_hash(self, tmp_dir, signature_script_dir):
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(test_file, index=False)
        file_hash = bankcheck.compute_file_hash(test_file)

        signature_info = self._create_signature_info(test_file)
        bankcheck.save_signature_record(signature_info, script_dir=signature_script_dir)

        signatures = bankcheck.query_signatures(
            script_dir=signature_script_dir, file_hash=file_hash
        )
        assert len(signatures) >= 1
        assert signatures[0]['file_hash'] == file_hash

    def test_query_signatures_empty_db(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'empty')
        os.makedirs(script_dir)
        signatures = bankcheck.query_signatures(script_dir=script_dir)
        assert signatures == []

    def test_query_signatures_is_verified_bool(self, tmp_dir, signature_script_dir):
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(test_file, index=False)

        signature_info = self._create_signature_info(test_file)
        bankcheck.save_signature_record(signature_info, script_dir=signature_script_dir)

        signatures = bankcheck.query_signatures(script_dir=signature_script_dir)
        assert len(signatures) >= 1
        assert isinstance(signatures[0]['is_verified'], bool)


class TestVerifyAndUpdateSignature:
    def test_verify_and_update_valid_signature(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1, 2, 3]}).to_excel(test_file, index=False)

        signature_info = bankcheck.sign_output_file(
            output_path=test_file,
            script_dir=signature_script_dir,
            username='verify_test',
            input_directory='/test'
        )
        signature_id = bankcheck.save_signature_record(
            signature_info, script_dir=signature_script_dir
        )

        is_valid = bankcheck.verify_and_update_signature(
            signature_id, script_dir=signature_script_dir
        )
        assert is_valid is True

        signatures = bankcheck.query_signatures(
            script_dir=signature_script_dir, signature_id=signature_id
        )
        assert signatures[0]['is_verified'] is True
        assert signatures[0]['verified_at'] is not None

    def test_verify_and_update_invalid_signature(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1, 2, 3]}).to_excel(test_file, index=False)

        signature_info = bankcheck.sign_output_file(
            output_path=test_file,
            script_dir=signature_script_dir,
            username='verify_test',
            input_directory='/test'
        )
        signature_id = bankcheck.save_signature_record(
            signature_info, script_dir=signature_script_dir
        )

        pd.DataFrame({'A': [999]}).to_excel(test_file, index=False)

        is_valid = bankcheck.verify_and_update_signature(
            signature_id, script_dir=signature_script_dir
        )
        assert is_valid is False

        signatures = bankcheck.query_signatures(
            script_dir=signature_script_dir, signature_id=signature_id
        )
        assert signatures[0]['is_verified'] is False

    def test_verify_nonexistent_signature(self, signature_script_dir):
        is_valid = bankcheck.verify_and_update_signature(
            'NONEXISTENT_SIG', script_dir=signature_script_dir
        )
        assert is_valid is False


class TestVerifyFileByPath:
    def test_verify_file_by_path_valid(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1, 2, 3]}).to_excel(test_file, index=False)

        signature_info = bankcheck.sign_output_file(
            output_path=test_file,
            script_dir=signature_script_dir,
            username='file_verify_test',
            input_directory='/test'
        )
        bankcheck.save_signature_record(signature_info, script_dir=signature_script_dir)

        result = bankcheck.verify_file_by_path(test_file, script_dir=signature_script_dir)
        assert result['exists'] is True
        assert result['integrity_valid'] is True
        assert result['signature_valid'] is True
        assert result['error'] is None

    def test_verify_file_by_path_tampered(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1, 2, 3]}).to_excel(test_file, index=False)

        signature_info = bankcheck.sign_output_file(
            output_path=test_file,
            script_dir=signature_script_dir,
            username='file_verify_test',
            input_directory='/test'
        )
        bankcheck.save_signature_record(signature_info, script_dir=signature_script_dir)

        pd.DataFrame({'A': [999]}).to_excel(test_file, index=False)

        result = bankcheck.verify_file_by_path(test_file, script_dir=signature_script_dir)
        assert result['exists'] is True
        assert result['integrity_valid'] is False
        assert result['signature_valid'] is False
        assert result['error'] is not None

    def test_verify_file_by_path_nonexistent(self, signature_script_dir):
        result = bankcheck.verify_file_by_path(
            '/nonexistent/file.xlsx', script_dir=signature_script_dir
        )
        assert result['exists'] is False
        assert result['error'] is not None

    def test_verify_file_no_signature_record(self, tmp_dir, signature_script_dir):
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(test_file, index=False)

        result = bankcheck.verify_file_by_path(test_file, script_dir=signature_script_dir)
        assert result['exists'] is True
        assert result['error'] is not None
        assert '未找到该文件的签名记录' in result['error']


class TestExportSignatures:
    def test_export_signatures_creates_excel(self, tmp_dir, signature_script_dir):
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(test_file, index=False)

        signature_info = {
            'file_hash': bankcheck.compute_file_hash(test_file),
            'username': 'export_test',
            'operation_time': '2024-01-01 12:00:00',
            'input_directory': '/test/path',
            'algorithm': 'RSA-SHA256',
            'signature': 'test_sig',
            'payload': '{}',
            'signed_at': '2024-01-01 12:00:01',
            'file_path': test_file,
        }
        bankcheck.save_signature_record(signature_info, script_dir=signature_script_dir)

        output_path = os.path.join(tmp_dir, 'signature_export.xlsx')
        result = bankcheck.export_signatures(output_path, script_dir=signature_script_dir)

        assert result == output_path
        assert os.path.exists(output_path)

        df = pd.read_excel(output_path, engine='openpyxl')
        assert len(df) >= 1
        assert 'signature_id' in df.columns
        assert 'file_path' in df.columns
        assert 'file_hash' in df.columns
        assert 'username' in df.columns
        assert 'signature' in df.columns
        assert 'is_verified' in df.columns

    def test_export_signatures_no_data_returns_none(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'empty')
        os.makedirs(script_dir)
        output_path = os.path.join(tmp_dir, 'export.xlsx')
        result = bankcheck.export_signatures(output_path, script_dir=script_dir)
        assert result is None


class TestExportSignatureManifest:
    def test_export_signature_manifest(self, tmp_dir, signature_script_dir):
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(test_file, index=False)

        signature_info = {
            'file_hash': bankcheck.compute_file_hash(test_file),
            'username': 'manifest_test',
            'operation_time': '2024-01-01 12:00:00',
            'input_directory': '/test/path',
            'algorithm': 'RSA-SHA256',
            'signature': 'test_sig',
            'payload': '{}',
            'signed_at': '2024-01-01 12:00:01',
            'file_path': test_file,
        }
        bankcheck.save_signature_record(signature_info, script_dir=signature_script_dir)

        output_path = os.path.join(tmp_dir, 'manifest.json')
        result = bankcheck.export_signature_manifest(output_path, script_dir=signature_script_dir)

        assert result == output_path
        assert os.path.exists(output_path)

        with open(output_path, 'r', encoding='utf-8') as f:
            manifest = json.load(f)

        assert 'exported_at' in manifest
        assert manifest['signature_count'] >= 1
        assert manifest['algorithm'] == 'RSA-SHA256'
        assert 'public_key_path' in manifest
        assert 'signatures' in manifest
        assert len(manifest['signatures']) >= 1


class TestAuditLogsWithSignature:
    def test_audit_logs_has_signature_columns(self, signature_script_dir):
        logs = bankcheck.query_audit_logs(script_dir=signature_script_dir, limit=1)
        assert logs == []

        db_path = bankcheck.get_audit_db_path(signature_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(audit_logs)")
        columns = {row[1] for row in cursor.fetchall()}
        conn.close()

        signature_cols = {'signature_id', 'digital_signature', 'signature_algorithm', 'signed_at'}
        assert signature_cols.issubset(columns)

    def test_query_audit_logs_with_signature_filter(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)

        source_folder = os.path.join(tmp_dir, '流水')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))

        with bankcheck.AuditLogger('pipeline', signature_script_dir) as audit:
            audit.record_input(source_folder)
            result = bankcheck.run_pipeline(
                source_folder, signature_script_dir,
                enable_signature=True
            )
            audit.record_result(result)

        logs_with_sig = bankcheck.query_audit_logs(
            script_dir=signature_script_dir, has_signature=True
        )
        assert len(logs_with_sig) >= 1
        assert logs_with_sig[0]['digital_signature'] is not None

        logs_without_sig = bankcheck.query_audit_logs(
            script_dir=signature_script_dir, has_signature=False
        )
        assert len(logs_without_sig) == 0

    def test_audit_logger_record_signature(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        test_file = os.path.join(tmp_dir, 'test.xlsx')
        pd.DataFrame({'A': [1]}).to_excel(test_file, index=False)

        with bankcheck.AuditLogger('test_op', signature_script_dir, username='sig_test') as audit:
            signature_info = bankcheck.sign_output_file(
                output_path=test_file,
                script_dir=signature_script_dir,
                username='sig_test',
                input_directory='/test'
            )
            audit.record_signature(signature_info)

        logs = bankcheck.query_audit_logs(
            script_dir=signature_script_dir, username='sig_test'
        )
        assert len(logs) == 1
        assert logs[0]['signature_id'] is not None
        assert logs[0]['digital_signature'] is not None
        assert logs[0]['signature_algorithm'] == 'RSA-SHA256'
        assert logs[0]['signed_at'] is not None


class TestPipelineWithSignature:
    def test_run_pipeline_with_signature(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)

        source_folder = os.path.join(tmp_dir, '流水')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))

        result = bankcheck.run_pipeline(
            source_folder, signature_script_dir,
            enable_signature=True
        )

        assert result.signature_id is not None
        assert result.signature_info is not None
        assert result.output_hash is not None
        assert result.signature_id.startswith('SIG')
        assert result.signature_info['algorithm'] == 'RSA-SHA256'
        assert result.signature_info['username'] == bankcheck.get_current_user()
        assert result.signature_info['input_directory'] == source_folder

        signatures = bankcheck.query_signatures(
            script_dir=signature_script_dir, file_path=result.output_path
        )
        assert len(signatures) >= 1

    def test_run_pipeline_with_signature_auto_gen_key(self, tmp_dir, signature_script_dir):
        source_folder = os.path.join(tmp_dir, '流水')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))

        result = bankcheck.run_pipeline(
            source_folder, signature_script_dir,
            enable_signature=True,
            auto_generate_key=True
        )

        assert result.signature_id is not None
        assert bankcheck.has_signing_key(signature_script_dir) is True

    def test_run_pipeline_without_signature(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)

        source_folder = os.path.join(tmp_dir, '流水')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))

        result = bankcheck.run_pipeline(
            source_folder, signature_script_dir,
            enable_signature=False
        )

        assert result.signature_id is None
        assert result.signature_info is None
        assert result.output_hash is None

    def test_run_pipeline_signature_verification(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)

        source_folder = os.path.join(tmp_dir, '流水')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))

        result = bankcheck.run_pipeline(
            source_folder, signature_script_dir,
            enable_signature=True
        )

        verify_result = bankcheck.verify_file_by_path(
            result.output_path, script_dir=signature_script_dir
        )

        assert verify_result['exists'] is True
        assert verify_result['integrity_valid'] is True
        assert verify_result['signature_valid'] is True

    def test_run_pipeline_tampered_file_fails_verification(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)

        source_folder = os.path.join(tmp_dir, '流水')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))

        result = bankcheck.run_pipeline(
            source_folder, signature_script_dir,
            enable_signature=True
        )

        wb = openpyxl.load_workbook(result.output_path)
        ws = wb.active
        ws['A2'] = 'tampered_data'
        wb.save(result.output_path)
        wb.close()

        verify_result = bankcheck.verify_file_by_path(
            result.output_path, script_dir=signature_script_dir
        )

        assert verify_result['integrity_valid'] is False
        assert verify_result['signature_valid'] is False
        assert verify_result['error'] is not None


class TestDatabaseMigration:
    def test_migrate_audit_db_adds_columns(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'migrate_test')
        os.makedirs(script_dir)

        db_path = bankcheck.get_audit_db_path(script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        cursor.execute('''
            CREATE TABLE IF NOT EXISTS audit_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                audit_id TEXT NOT NULL UNIQUE,
                username TEXT NOT NULL,
                operation_type TEXT NOT NULL,
                status TEXT NOT NULL,
                started_at TEXT NOT NULL
            )
        ''')
        conn.commit()
        conn.close()

        bankcheck._migrate_audit_db_columns(db_path)

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(audit_logs)")
        columns = {row[1] for row in cursor.fetchall()}
        conn.close()

        new_cols = {'signature_id', 'digital_signature', 'signature_algorithm', 'signed_at'}
        assert new_cols.issubset(columns)

    def test_migrate_idempotent(self, signature_script_dir):
        db_path = bankcheck.get_audit_db_path(signature_script_dir)

        bankcheck._migrate_audit_db_columns(db_path)
        bankcheck._migrate_audit_db_columns(db_path)
        bankcheck._migrate_audit_db_columns(db_path)

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(audit_logs)")
        columns = {row[1] for row in cursor.fetchall()}
        conn.close()

        assert 'signature_id' in columns


class TestExportAuditLogsWithSignatures:
    def test_export_audit_logs_includes_signature_columns(self, tmp_dir, signature_script_dir):
        bankcheck.generate_signing_key_pair(signature_script_dir)
        source_folder = os.path.join(tmp_dir, '流水')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))

        with bankcheck.AuditLogger('pipeline', signature_script_dir) as audit:
            audit.record_input(source_folder)
            result = bankcheck.run_pipeline(
                source_folder, signature_script_dir,
                enable_signature=True
            )
            audit.record_result(result)

        output_path = os.path.join(tmp_dir, 'audit_export.xlsx')
        result = bankcheck.export_audit_logs(output_path, script_dir=signature_script_dir)

        assert result == output_path
        df = pd.read_excel(output_path, engine='openpyxl')

        signature_cols = ['signature_id', 'digital_signature', 'signature_algorithm', 'signed_at']
        for col in signature_cols:
            assert col in df.columns
