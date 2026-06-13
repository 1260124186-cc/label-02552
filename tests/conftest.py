import os
import sys
import shutil
import tempfile

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck


BANK_TEST_CONFIGS = {
    '北京银行': {
        'bank_name': '北京银行',
        'account': '01090312345678901',
        'header_row': 3,
        'start_row': 4,
        'sheet_title': '交易明细',
        'headers': [
            '序号', '交易日期', '币种', '支出金额', '收入金额', '余额',
            '对方户名', '对方账号', '对方行名', '凭证种类', '凭证号码',
            '摘要', '备注1', '备注2', '备注3', '交易流水号',
        ],
        'cells': {'A1': '北京银行交易明细', 'B2': '{account}'},
        'default_rows': [
            [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A公司', '622001234', '工商银行', '转账', '001', '采购付款', None, None, None, 'BJ20240105001'],
            [2, '2024-01-10', 'CNY', None, 80000, 1580000, '客户B公司', '622005678', '建设银行', '转账', '002', '销售收款', None, None, None, 'BJ20240110002'],
        ],
        'skip_empty_date_rows': [
            [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A公司', '622001', '工商银行', '转账', '001', '付款', None, None, None, 'BJ001'],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [2, '2024-01-10', 'CNY', None, 80000, 1580000, '客户B公司', '622002', '建设银行', '转账', '002', '收款', None, None, None, 'BJ002'],
        ],
        'expected': {
            'row_count': 2,
            'bank_name': '北京银行',
            'account': '01090312345678901',
            'subject': '北京XX科技有限公司',
            'first_payment': -50000.0,
            'second_receipt': 80000.0,
            'first_receipt': None,
            'second_payment': None,
            'trade_dates': ['2024-01-05', '2024-01-10'],
            'first_summary': '采购付款',
            'second_summary': '销售收款',
            'first_counterpart': '供应商A公司',
            'second_counterpart': '客户B公司',
            'first_balance': 1500000,
            'second_balance': 1580000,
            'transaction_ids': ['BJ20240105001', 'BJ20240110002'],
            'expected_keys': {'唯一id', '银行', '银行账号', '主体', '交易日期', '付款', '收款', '摘要', '对方户名', '余额', '交易流水号'},
        },
    },
    '东亚银行': {
        'bank_name': '东亚银行',
        'account': '38812345678',
        'header_row': 4,
        'start_row': 5,
        'sheet_title': '交易明细',
        'headers': [
            '交易日期', '交易时间', '币种', '支出金额', '收入金额',
            '手续费', '利息', '税费', '余额', '交易类型', '交易流水号', '交易描述/对方',
        ],
        'cells': {'A1': '账号', 'B1': '{account}', 'A2': '东亚银行交易明细', 'A3': '查询期间：2024-01-01 至 2024-01-31'},
        'default_rows': [
            ['2024-01-03', '09:30:00', 'CNY', 20000, None, 100, 0, 0, 480000, '转账', 'EA20240103001', '向 张三 付款-货款'],
            ['2024-01-08', '14:15:00', 'CNY', None, 35000, 0, 0, 0, 515000, '转账', 'EA20240108002', '收到 李四 汇款-服务费'],
        ],
        'skip_empty_date_rows': [
            ['2024-01-03', '09:30:00', 'CNY', 20000, None, 100, 0, 0, 480000, '转账', 'EA001', '付款'],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            ['2024-01-08', '14:15:00', 'CNY', None, 35000, 0, 0, 0, 515000, '转账', 'EA002', '收款'],
        ],
        'expected': {
            'row_count': 2,
            'bank_name': '东亚银行',
            'account': '38812345678',
            'subject': '上海YY贸易有限公司',
            'first_payment': -20000.0,
            'second_receipt': 35000.0,
            'first_receipt': None,
            'second_payment': None,
            'trade_dates': ['2024-01-03', '2024-01-08'],
            'first_summary': '向 张三 付款-货款',
            'counterpart_same_as_summary': True,
            'first_balance': 480000,
            'second_balance': 515000,
            'transaction_ids': ['EA20240103001', 'EA20240108002'],
            'expected_keys': {'唯一id', '银行', '银行账号', '主体', '交易日期', '付款', '收款', '摘要', '对方户名', '余额', '交易流水号'},
        },
    },
    '工商银行': {
        'bank_name': '工商银行',
        'account': '6222021234567890123',
        'header_row': 5,
        'start_row': 6,
        'sheet_title': '交易明细',
        'headers': [
            '交易日期', '交易时间', '币种', '收入金额', '支出金额',
            '余额', '对方户名', '摘要', '备注', '交易流水号',
        ],
        'cells': {'A1': '中国工商银行账户明细', 'B4': '{account}', 'A3': '查询期间：2024-01-01 至 2024-01-31'},
        'default_rows': [
            ['2024-01-02', '10:00:00', 'CNY', None, 45000, 1200000, '供应商甲公司', '材料采购付款', None, 'ICBC20240102001'],
            ['2024-01-09', '11:30:00', 'CNY', 75000, None, 1275000, '客户乙公司', '销售款入账', None, 'ICBC20240109002'],
        ],
        'skip_empty_date_rows': [
            ['2024-01-02', '10:00:00', 'CNY', None, 45000, 1200000, '供应商甲', '付款', None, 'ICBC001'],
            [None, None, None, None, None, None, None, None, None, None],
            ['2024-01-09', '11:30:00', 'CNY', 75000, None, 1275000, '客户乙', '收款', None, 'ICBC002'],
        ],
        'expected': {
            'row_count': 2,
            'bank_name': '工商银行',
            'account': '6222021234567890123',
            'subject': '深圳ZZ科技有限公司',
            'first_payment': -45000.0,
            'second_receipt': 75000.0,
            'first_receipt': None,
            'second_payment': None,
            'trade_dates': ['2024-01-02', '2024-01-09'],
            'first_summary': '材料采购付款',
            'first_counterpart': '供应商甲公司',
            'second_counterpart': '客户乙公司',
            'first_balance': 1200000,
            'second_balance': 1275000,
            'transaction_ids': ['ICBC20240102001', 'ICBC20240109002'],
            'expected_keys': {'唯一id', '银行', '银行账号', '主体', '交易日期', '付款', '收款', '摘要', '对方户名', '余额', '交易流水号'},
        },
    },
    '建设银行': {
        'bank_name': '建设银行',
        'account': '6227001234567890123',
        'header_row': 3,
        'start_row': 4,
        'sheet_title': '交易明细',
        'headers': [
            '序号', '交易日期', '交易时间', '收入金额', '支出金额',
            '余额', '对方户名', '对方账号', '摘要', '备注', '交易流水号',
        ],
        'cells': {'A1': '账户信息', 'A2': '{account}', 'A3': '中国建设银行个人活期账户交易明细'},
        'default_rows': [
            [1, '2024-01-04', '08:45:00', None, 32000, 968000, '供应商丙公司', '62170000001', '项目工程款', None, 'CCB20240104001'],
            [2, '2024-01-11', '16:20:00', 58000, None, 1026000, '客户丁公司', '62170000002', '合同回款', None, 'CCB20240111002'],
        ],
        'skip_empty_date_rows': [
            [1, '2024-01-04', '08:45:00', None, 32000, 968000, '供应商丙', '6217001', '工程款', None, 'CCB001'],
            [None, None, None, None, None, None, None, None, None, None, None],
            [2, '2024-01-11', '16:20:00', 58000, None, 1026000, '客户丁', '6217002', '回款', None, 'CCB002'],
        ],
        'expected': {
            'row_count': 2,
            'bank_name': '建设银行',
            'account': '6227001234567890123',
            'subject': '广州WW建设集团有限公司',
            'first_payment': -32000.0,
            'second_receipt': 58000.0,
            'first_receipt': None,
            'second_payment': None,
            'trade_dates': ['2024-01-04', '2024-01-11'],
            'first_summary': '项目工程款',
            'first_counterpart': '供应商丙公司',
            'second_counterpart': '客户丁公司',
            'first_balance': 968000,
            'second_balance': 1026000,
            'transaction_ids': ['CCB20240104001', 'CCB20240111002'],
            'expected_keys': {'唯一id', '银行', '银行账号', '主体', '交易日期', '付款', '收款', '摘要', '对方户名', '余额', '交易流水号'},
        },
    },
    '招商银行': {
        'bank_name': '招商银行',
        'account': '6225881234567890',
        'header_row': 5,
        'start_row': 6,
        'sheet_title': '活期交易明细',
        'headers': [
            '交易日期', '交易时间', '收入金额', '支出金额', '余额',
            '币种', '对方账号', '对方行名', '对方户名', '摘要', '备注', '交易流水号',
        ],
        'cells': {'A1': '招商银行', 'A2': '账户活期交易明细查询', 'B3': '{account}', 'A4': '查询周期：2024-01-01 至 2024-01-31'},
        'default_rows': [
            ['2024-01-06', '09:15:00', None, 28000, 750000, 'CNY', '622600000001', '民生银行', '供应商戊公司', '采购设备款', None, 'CMB20240106001'],
            ['2024-01-13', '15:40:00', 92000, None, 842000, 'CNY', '622600000002', '浦发银行', '客户己公司', '服务收入', None, 'CMB20240113002'],
        ],
        'skip_empty_date_rows': [
            ['2024-01-06', '09:15:00', None, 28000, 750000, 'CNY', '6226001', '民生银行', '供应商戊', '采购款', None, 'CMB001'],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            ['2024-01-13', '15:40:00', 92000, None, 842000, 'CNY', '6226002', '浦发银行', '客户己', '收入款', None, 'CMB002'],
        ],
        'expected': {
            'row_count': 2,
            'bank_name': '招商银行',
            'account': '6225881234567890',
            'subject': '杭州VV网络技术有限公司',
            'first_payment': -28000.0,
            'second_receipt': 92000.0,
            'first_receipt': None,
            'second_payment': None,
            'trade_dates': ['2024-01-06', '2024-01-13'],
            'first_summary': '采购设备款',
            'first_counterpart': '供应商戊公司',
            'second_counterpart': '客户己公司',
            'first_balance': 750000,
            'second_balance': 842000,
            'transaction_ids': ['CMB20240106001', 'CMB20240113002'],
            'expected_keys': {'唯一id', '银行', '银行账号', '主体', '交易日期', '付款', '收款', '摘要', '对方户名', '余额', '交易流水号'},
        },
    },
}


_UNSET = object()


def _create_bank_excel(path, bank_name, account=_UNSET, rows=None):
    cfg = BANK_TEST_CONFIGS[bank_name]
    if account is _UNSET:
        account = cfg['account']
    if rows is None:
        rows = cfg['default_rows']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = cfg['sheet_title']
    for cell_ref, template in cfg['cells'].items():
        if account is not None and '{account}' in template:
            ws[cell_ref] = template.replace('{account}', str(account))
        else:
            ws[cell_ref] = template
    for c, h in enumerate(cfg['headers'], 1):
        ws.cell(row=cfg['header_row'], column=c, value=h)
    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=cfg['start_row'] + i, column=j + 1, value=val)
    wb.save(path)
    wb.close()
    return path


@pytest.fixture(autouse=True)
def init_logging():
    bankcheck.setup_logging()


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='bankcheck_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _create_beijing_bank_excel(path, account=_UNSET, rows=None):
    if account is _UNSET:
        account = BANK_TEST_CONFIGS['北京银行']['account']
    return _create_bank_excel(path, '北京银行', account=account, rows=rows)


def _create_beijing_bank_multi_sheet_excel(path, sheets=None):
    if sheets is None:
        sheets = [
            {
                'title': '1月',
                'account': '01090312345678901',
                'rows': [
                    [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A公司', '622001234', '工商银行', '转账', '001', '采购付款', None, None, None, 'BJ20240105001'],
                    [2, '2024-01-10', 'CNY', None, 80000, 1580000, '客户B公司', '622005678', '建设银行', '转账', '002', '销售收款', None, None, None, 'BJ20240110002'],
                ],
            },
            {
                'title': '2月',
                'account': '01090312345678901',
                'rows': [
                    [1, '2024-02-03', 'CNY', 30000, None, 1550000, '供应商C公司', '622009999', '农业银行', '转账', '003', '材料采购', None, None, None, 'BJ20240203003'],
                    [2, '2024-02-15', 'CNY', None, 60000, 1610000, '客户D公司', '622008888', '中国银行', '转账', '004', '服务收款', None, None, None, 'BJ20240215004'],
                ],
            },
        ]
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    headers = ['序号', '交易日期', '币种', '支出金额', '收入金额', '余额', '对方户名', '对方账号', '对方行名', '凭证种类', '凭证号码', '摘要', '备注1', '备注2', '备注3', '交易流水号']
    for sheet_data in sheets:
        ws = wb.create_sheet(title=sheet_data['title'])
        ws['A1'] = '北京银行交易明细'
        ws['B2'] = sheet_data.get('account')
        for c, h in enumerate(headers, 1):
            ws.cell(row=3, column=c, value=h)
        for i, row_data in enumerate(sheet_data.get('rows', [])):
            for j, val in enumerate(row_data):
                ws.cell(row=4 + i, column=j + 1, value=val)
    wb.save(path)
    wb.close()
    return path


def _create_east_asia_bank_excel(path, account=_UNSET, rows=None):
    if account is _UNSET:
        account = BANK_TEST_CONFIGS['东亚银行']['account']
    return _create_bank_excel(path, '东亚银行', account=account, rows=rows)


def _create_beijing_bank_excel_with_extra_sheets(path, data_rows=None, extra_sheet_titles=None):
    if data_rows is None:
        data_rows = [
            [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A公司', '622001234', '工商银行', '转账', '001', '采购付款', None, None, None, 'BJ20240105001'],
        ]
    if extra_sheet_titles is None:
        extra_sheet_titles = ['说明', '汇总']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '交易明细'
    ws['A1'] = '北京银行交易明细'
    ws['B2'] = '01090312345678901'
    headers = ['序号', '交易日期', '币种', '支出金额', '收入金额', '余额', '对方户名', '对方账号', '对方行名', '凭证种类', '凭证号码', '摘要', '备注1', '备注2', '备注3', '交易流水号']
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    for i, row_data in enumerate(data_rows):
        for j, val in enumerate(row_data):
            ws.cell(row=4 + i, column=j + 1, value=val)
    for title in extra_sheet_titles:
        extra_ws = wb.create_sheet(title=title)
        extra_ws['A1'] = '此表为说明页，无流水数据'
    wb.save(path)
    wb.close()
    return path


def _create_lookup_table(path, mappings=None):
    if mappings is None:
        mappings = [
            ('北京XX科技有限公司', '01090312345678901'),
            ('上海YY贸易有限公司', '38812345678'),
        ]
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '主体映射'
    ws['A1'] = '主体名称'
    ws['B1'] = '银行账号'
    for i, (subject, account) in enumerate(mappings, 2):
        ws.cell(row=i, column=1, value=subject)
        ws.cell(row=i, column=2, value=account)
    wb.save(path)
    wb.close()
    return path


def _create_lookup_table_all_banks(path):
    """创建包含全部5个银行映射的查找表"""
    mappings = [
        ('北京XX科技有限公司', '01090312345678901'),
        ('上海YY贸易有限公司', '38812345678'),
        ('深圳ZZ科技有限公司', '6222021234567890123'),
        ('广州WW建设集团有限公司', '6227001234567890123'),
        ('杭州VV网络技术有限公司', '6225881234567890'),
    ]
    return _create_lookup_table(path, mappings)


def _create_beijing_bank_multi_account_excel(path, accounts=None):
    """
    创建包含多个账号区块的北京银行 Excel 文件（同一工作表内纵向堆叠）。

    accounts 参数格式：
    [
        {
            'account': '01090312345678901',
            'rows': [
                [1, '2024-01-05', 'CNY', 50000, None, 1500000, ...],
            ],
        },
        {
            'account': '01090399999999999',
            'rows': [
                [1, '2024-02-01', 'CNY', 20000, None, 800000, ...],
            ],
        },
    ]
    """
    if accounts is None:
        accounts = [
            {
                'account': '01090312345678901',
                'rows': [
                    [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A公司', '622001234', '工商银行', '转账', '001', '采购付款', None, None, None, 'BJ20240105001'],
                    [2, '2024-01-10', 'CNY', None, 80000, 1580000, '客户B公司', '622005678', '建设银行', '转账', '002', '销售收款', None, None, None, 'BJ20240110002'],
                ],
            },
            {
                'account': '01090399999999999',
                'rows': [
                    [1, '2024-02-01', 'CNY', 20000, None, 800000, '供应商C公司', '622009999', '农业银行', '转账', '003', '材料采购', None, None, None, 'BJ20240201003'],
                    [2, '2024-02-15', 'CNY', None, 60000, 860000, '客户D公司', '622008888', '中国银行', '转账', '004', '服务收款', None, None, None, 'BJ20240215004'],
                ],
            },
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '交易明细'

    headers = ['序号', '交易日期', '币种', '支出金额', '收入金额', '余额',
               '对方户名', '对方账号', '对方行名', '凭证种类', '凭证号码',
               '摘要', '备注1', '备注2', '备注3', '交易流水号']

    current_row = 1
    for acct_info in accounts:
        ws.cell(row=current_row, column=1, value='北京银行交易明细')
        ws.cell(row=current_row + 1, column=2, value=acct_info['account'])
        for c, h in enumerate(headers, 1):
            ws.cell(row=current_row + 2, column=c, value=h)
        for i, row_data in enumerate(acct_info['rows']):
            for j, val in enumerate(row_data):
                ws.cell(row=current_row + 3 + i, column=j + 1, value=val)
        current_row += 3 + len(acct_info['rows'])

    wb.save(path)
    wb.close()
    return path
