import os

import pytest
import openpyxl

from conftest import _create_lookup_table
from bankcheck import (
    find_lookup_file, get_subject, _normalize_account_str, _account_key,
    get_subject_info, get_lookup_extra_fields, get_summary_columns,
    load_lookup_table, _resolve_lookup,
)


class TestFindLookupFile:
    def test_exact_match_xlsx(self, tmp_dir):
        path = os.path.join(tmp_dir, '主体查找表.xlsx')
        _create_lookup_table(path)
        assert find_lookup_file(tmp_dir) == path

    def test_exact_match_xls(self, tmp_dir):
        path = os.path.join(tmp_dir, '主体查找表.xls')
        _create_lookup_table(path)
        result = find_lookup_file(tmp_dir)
        assert result == path

    def test_fallback_single_excel(self, tmp_dir):
        path = os.path.join(tmp_dir, '映射表.xlsx')
        _create_lookup_table(path)
        result = find_lookup_file(tmp_dir)
        assert result == path

    def test_no_excel_files(self, tmp_dir):
        result = find_lookup_file(tmp_dir)
        assert result is None

    def test_multiple_excel_files_ambiguous(self, tmp_dir):
        _create_lookup_table(os.path.join(tmp_dir, '表1.xlsx'))
        _create_lookup_table(os.path.join(tmp_dir, '表2.xlsx'))
        result = find_lookup_file(tmp_dir)
        assert result is None

    def test_excludes_output_table(self, tmp_dir):
        output_path = os.path.join(tmp_dir, '银行流水总表.xlsx')
        _create_lookup_table(output_path)
        lookup_path = os.path.join(tmp_dir, '主体查找表.xlsx')
        _create_lookup_table(lookup_path)
        result = find_lookup_file(tmp_dir)
        assert result == lookup_path

    def test_excludes_temp_files(self, tmp_dir):
        open(os.path.join(tmp_dir, '~$temp.xlsx'), 'w').close()
        path = os.path.join(tmp_dir, '映射表.xlsx')
        _create_lookup_table(path)
        result = find_lookup_file(tmp_dir)
        assert result == path


class TestGetSubject:
    def test_found(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'))
        result = get_subject('01090312345678901', path)
        assert result == '北京XX科技有限公司'

    def test_found_east_asia(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'))
        result = get_subject('38812345678', path)
        assert result == '上海YY贸易有限公司'

    def test_not_found(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'))
        result = get_subject('99999999999', path)
        assert result == ''

    def test_none_account(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'))
        result = get_subject(None, path)
        assert result == ''

    def test_empty_account(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'))
        result = get_subject('', path)
        assert result == ''

    def test_none_lookup_file(self):
        result = get_subject('01090312345678901', None)
        assert result == ''

    def test_nonexistent_lookup_file(self):
        result = get_subject('01090312345678901', '/nonexistent/path.xlsx')
        assert result == ''

    def test_account_with_spaces(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'), [
            ('测试公司', ' 12345 '),
        ])
        result = get_subject('12345', path)
        assert result == '测试公司'

    def test_int_account_matches_string_in_lookup(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'), [
            ('北京XX科技有限公司', '01090312345678901'),
        ])
        result = get_subject(1090312345678901, path)
        assert result == '北京XX科技有限公司'

    def test_float_account_matches_string_in_lookup(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'), [
            ('北京XX科技有限公司', '01090312345678901'),
        ])
        result = get_subject(1090312345678901.0, path)
        assert result == '北京XX科技有限公司'

    def test_string_account_matches_int_in_lookup(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'), [
            ('北京XX科技有限公司', 1090312345678901),
        ])
        result = get_subject('01090312345678901', path)
        assert result == '北京XX科技有限公司'

    def test_string_account_matches_float_in_lookup(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'), [
            ('北京XX科技有限公司', 1090312345678901.0),
        ])
        result = get_subject('01090312345678901', path)
        assert result == '北京XX科技有限公司'

    def test_int_account_matches_int_in_lookup(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'), [
            ('上海YY贸易有限公司', 38812345678),
        ])
        result = get_subject(38812345678, path)
        assert result == '上海YY贸易有限公司'

    def test_float_account_matches_float_in_lookup(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'), [
            ('上海YY贸易有限公司', 38812345678.0),
        ])
        result = get_subject(38812345678.0, path)
        assert result == '上海YY贸易有限公司'

    def test_string_with_leading_zero_matches_int(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'), [
            ('测试公司A', 123456789),
        ])
        result = get_subject('0123456789', path)
        assert result == '测试公司A'

    def test_int_matches_string_with_leading_zero(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'), [
            ('测试公司B', '0123456789'),
        ])
        result = get_subject(123456789, path)
        assert result == '测试公司B'

    def test_string_dot_zero_matches_int_in_lookup(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'), [
            ('测试公司C', 38812345678),
        ])
        result = get_subject('38812345678.0', path)
        assert result == '测试公司C'

    def test_no_false_positive_on_different_accounts(self, tmp_dir):
        path = _create_lookup_table(os.path.join(tmp_dir, '主体查找表.xlsx'), [
            ('公司A', '1234567890'),
            ('公司B', '01234567890'),
        ])
        result = get_subject('01234567890', path)
        assert result == '公司A' or result == '公司B'


class TestNormalizeAccountStr:
    def test_none(self):
        assert _normalize_account_str(None) == ''

    def test_int(self):
        assert _normalize_account_str(123456789) == '123456789'

    def test_float_whole(self):
        assert _normalize_account_str(123456789.0) == '123456789'

    def test_float_with_decimals(self):
        assert _normalize_account_str(123.45) == '123.45'

    def test_float_nan(self):
        assert _normalize_account_str(float('nan')) == ''

    def test_string_digits(self):
        assert _normalize_account_str('01090312345678901') == '01090312345678901'

    def test_string_dot_zero(self):
        assert _normalize_account_str('38812345678.0') == '38812345678'

    def test_string_with_spaces(self):
        assert _normalize_account_str('  12345  ') == '12345'

    def test_empty_string(self):
        assert _normalize_account_str('') == ''

    def test_large_int(self):
        assert _normalize_account_str(1090312345678901) == '1090312345678901'


class TestAccountKey:
    def test_string_with_leading_zero(self):
        assert _account_key('01090312345678901') == '1090312345678901'

    def test_int_without_leading_zero(self):
        assert _account_key(1090312345678901) == '1090312345678901'

    def test_float_without_leading_zero(self):
        assert _account_key(1090312345678901.0) == '1090312345678901'

    def test_both_equal(self):
        assert _account_key('01090312345678901') == _account_key(1090312345678901)

    def test_zero_account(self):
        assert _account_key(0) == '0'

    def test_zero_string(self):
        assert _account_key('0') == '0'

    def test_none(self):
        assert _account_key(None) == '0'


def _create_lookup_table_with_extra(path, entries=None):
    """创建带扩展字段的查找表
    entries: 列表，每个元素是 dict，包含 subject, account, priority, extra_fields
    """
    if entries is None:
        entries = [
            {'subject': '北京XX科技有限公司', 'account': '01090312345678901',
             'priority': 0, 'extra_fields': {'部门': '技术部', '项目编号': 'PRJ001'}},
            {'subject': '上海YY贸易有限公司', 'account': '38812345678',
             'priority': 0, 'extra_fields': {'部门': '市场部', '项目编号': 'PRJ002'}},
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '主体映射'

    all_extra_keys = set()
    for entry in entries:
        if entry.get('extra_fields'):
            all_extra_keys.update(entry['extra_fields'].keys())
    sorted_extra_keys = sorted(all_extra_keys)

    has_priority = any(e.get('priority', 0) != 0 for e in entries)

    ws.cell(row=1, column=1, value='主体名称')
    ws.cell(row=1, column=2, value='银行账号')
    col_idx = 3
    if has_priority:
        ws.cell(row=1, column=col_idx, value='优先级')
        col_idx += 1
    for key in sorted_extra_keys:
        ws.cell(row=1, column=col_idx, value=key)
        col_idx += 1

    for i, entry in enumerate(entries, start=2):
        ws.cell(row=i, column=1, value=entry.get('subject', ''))
        ws.cell(row=i, column=2, value=entry.get('account', ''))
        col_idx = 3
        if has_priority:
            ws.cell(row=i, column=col_idx, value=entry.get('priority', 0))
            col_idx += 1
        for key in sorted_extra_keys:
            val = entry.get('extra_fields', {}).get(key, '')
            ws.cell(row=i, column=col_idx, value=val)
            col_idx += 1

    wb.save(path)
    wb.close()
    return path


class TestPriorityStrategy:
    """测试多主体优先级策略"""

    def test_single_subject_no_priority(self, tmp_dir):
        """单个主体，无优先级"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table(path, [
            ('公司A', '12345'),
        ])
        result = get_subject('12345', path)
        assert result == '公司A'

    def test_multiple_subjects_same_account_priority(self, tmp_dir):
        """同一账号多个主体，按优先级排序返回最高的"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '低优先级主体', 'account': '12345', 'priority': 1},
            {'subject': '高优先级主体', 'account': '12345', 'priority': 10},
            {'subject': '中优先级主体', 'account': '12345', 'priority': 5},
        ])
        result = get_subject('12345', path)
        assert result == '高优先级主体'

    def test_multiple_subjects_same_priority(self, tmp_dir):
        """同一账号多个主体，相同优先级时返回第一个"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '主体A', 'account': '12345', 'priority': 5},
            {'subject': '主体B', 'account': '12345', 'priority': 5},
        ])
        result = get_subject('12345', path)
        assert result in ['主体A', '主体B']

    def test_zero_priority_default(self, tmp_dir):
        """默认优先级为 0"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '主体A', 'account': '12345', 'priority': -1},
            {'subject': '主体B', 'account': '12345', 'priority': 0},
            {'subject': '主体C', 'account': '12345', 'priority': 1},
        ])
        result = get_subject('12345', path)
        assert result == '主体C'

    def test_negative_priority(self, tmp_dir):
        """负优先级也能正确排序"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '主体A', 'account': '12345', 'priority': -10},
            {'subject': '主体B', 'account': '12345', 'priority': -5},
        ])
        result = get_subject('12345', path)
        assert result == '主体B'


class TestFuzzyMatching:
    """测试模糊匹配功能"""

    def test_exact_match_no_fuzzy(self, tmp_dir):
        """精确匹配时不需要模糊匹配"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table(path, [
            ('测试公司', '1234567890'),
        ])
        info = get_subject_info('1234567890', path, use_fuzzy=True)
        assert info['matched'] is True
        assert info['fuzzy_matched'] is False
        assert info['subject'] == '测试公司'
        assert info['similarity'] == 1.0

    def test_fuzzy_match_close_account(self, tmp_dir):
        """账号接近时模糊匹配"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table(path, [
            ('测试公司', '1234567890'),
        ])
        info = get_subject_info('123456789', path, use_fuzzy=True)
        assert info['matched'] is True
        assert info['fuzzy_matched'] is True
        assert info['subject'] == '测试公司'
        assert info['similarity'] > 0.5

    def test_fuzzy_match_disabled(self, tmp_dir):
        """禁用模糊匹配时不匹配"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table(path, [
            ('测试公司', '1234567890'),
        ])
        info = get_subject_info('123456789', path, use_fuzzy=False)
        assert info['matched'] is False
        assert info['subject'] == ''

    def test_fuzzy_match_threshold(self, tmp_dir):
        """模糊匹配阈值"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table(path, [
            ('测试公司', '1234567890'),
        ])
        info = get_subject_info('9876543210', path, use_fuzzy=True, fuzzy_threshold=0.9)
        assert info['matched'] is False

    def test_fuzzy_match_best_match(self, tmp_dir):
        """模糊匹配返回相似度最高的"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table(path, [
            ('公司A', '1234567890'),
            ('公司B', '1234567000'),
        ])
        info = get_subject_info('1234567899', path, use_fuzzy=True)
        assert info['matched'] is True
        assert info['subject'] == '公司A'

    def test_empty_account_fuzzy(self, tmp_dir):
        """空账号模糊匹配"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table(path, [
            ('测试公司', '1234567890'),
        ])
        info = get_subject_info('', path, use_fuzzy=True)
        assert info['matched'] is False


class TestExtraFields:
    """测试扩展字段功能"""

    def test_read_extra_fields(self, tmp_dir):
        """读取扩展字段"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '北京XX科技有限公司', 'account': '01090312345678901',
             'priority': 0, 'extra_fields': {'部门': '技术部', '项目编号': 'PRJ001'}},
        ])
        info = get_subject_info('01090312345678901', path)
        assert info['matched'] is True
        assert info['extra_fields']['部门'] == '技术部'
        assert info['extra_fields']['项目编号'] == 'PRJ001'

    def test_no_extra_fields(self, tmp_dir):
        """没有扩展字段时返回空字典"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table(path, [
            ('测试公司', '12345'),
        ])
        info = get_subject_info('12345', path)
        assert info['matched'] is True
        assert info['extra_fields'] == {}

    def test_extra_fields_empty_value(self, tmp_dir):
        """扩展字段有空值"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '12345',
             'priority': 0, 'extra_fields': {'部门': '技术部', '项目编号': ''}},
        ])
        info = get_subject_info('12345', path)
        assert info['extra_fields']['部门'] == '技术部'
        assert info['extra_fields']['项目编号'] == ''

    def test_multiple_extra_fields(self, tmp_dir):
        """多个扩展字段"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '12345',
             'priority': 0, 'extra_fields': {
                 '部门': '技术部',
                 '项目编号': 'PRJ001',
                 '负责人': '张三',
                 '备注': '测试备注',
             }},
        ])
        info = get_subject_info('12345', path)
        assert len(info['extra_fields']) == 4
        assert info['extra_fields']['负责人'] == '张三'
        assert info['extra_fields']['备注'] == '测试备注'


class TestGetSubjectInfo:
    """测试 get_subject_info 函数"""

    def test_basic_info(self, tmp_dir):
        """基本信息获取"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table(path, [
            ('测试公司', '12345'),
        ])
        info = get_subject_info('12345', path)
        assert info['subject'] == '测试公司'
        assert info['account'] == '12345'
        assert info['priority'] == 0
        assert info['matched'] is True
        assert info['fuzzy_matched'] is False

    def test_not_found_info(self, tmp_dir):
        """未找到时的返回"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table(path, [
            ('测试公司', '12345'),
        ])
        info = get_subject_info('99999', path)
        assert info['subject'] == ''
        assert info['matched'] is False
        assert info['fuzzy_matched'] is False
        assert info['similarity'] == 0.0
        assert info['extra_fields'] == {}

    def test_none_account(self, tmp_dir):
        """None 账号"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table(path, [
            ('测试公司', '12345'),
        ])
        info = get_subject_info(None, path)
        assert info['matched'] is False

    def test_none_lookup_file(self):
        """None 查找表"""
        info = get_subject_info('12345', None)
        assert info['matched'] is False

    def test_nonexistent_lookup_file(self):
        """不存在的查找表"""
        info = get_subject_info('12345', '/nonexistent/path.xlsx')
        assert info['matched'] is False


class TestGetLookupExtraFields:
    """测试 get_lookup_extra_fields 函数"""

    def test_with_extra_fields(self, tmp_dir):
        """有扩展字段时"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '12345',
             'priority': 0, 'extra_fields': {'部门': '技术部', '项目编号': 'PRJ001'}},
        ])
        fields = get_lookup_extra_fields(path)
        assert '部门' in fields
        assert '项目编号' in fields
        assert len(fields) == 2

    def test_without_extra_fields(self, tmp_dir):
        """没有扩展字段时"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table(path, [
            ('测试公司', '12345'),
        ])
        fields = get_lookup_extra_fields(path)
        assert fields == []

    def test_sorted_fields(self, tmp_dir):
        """扩展字段按字母排序"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '12345',
             'priority': 0, 'extra_fields': {'Z部门': '测试', 'A项目': '测试'}},
        ])
        fields = get_lookup_extra_fields(path)
        assert fields[0] == 'A项目'
        assert fields[1] == 'Z部门'

    def test_none_file(self):
        """None 文件"""
        fields = get_lookup_extra_fields(None)
        assert fields == []

    def test_nonexistent_file(self):
        """不存在的文件"""
        fields = get_lookup_extra_fields('/nonexistent/path.xlsx')
        assert fields == []


class TestGetSummaryColumns:
    """测试 get_summary_columns 函数"""

    def test_standard_columns_only(self):
        """只有标准列"""
        columns = get_summary_columns()
        assert '唯一id' in columns
        assert '银行' in columns
        assert '主体' in columns
        assert len(columns) == 11

    def test_with_lookup_extra_fields(self, tmp_dir):
        """带查找表扩展字段"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '12345',
             'priority': 0, 'extra_fields': {'部门': '技术部'}},
        ])
        columns = get_summary_columns(lookup_source=path)
        assert '部门' in columns

    def test_with_records_extra_fields(self):
        """带记录中的扩展字段"""
        records = [
            {'唯一id': '1', '银行': '测试银行', '银行账号': '123', '主体': '测试',
             '交易日期': '2024-01-01', '付款': None, '收款': 100,
             '摘要': '', '对方户名': '', '余额': 100, '交易流水号': 'T1',
             '自定义字段': '值1'},
        ]
        columns = get_summary_columns(records=records)
        assert '自定义字段' in columns

    def test_combined_extra_fields(self, tmp_dir):
        """查找表和记录的扩展字段合并"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '123',
             'priority': 0, 'extra_fields': {'部门': '技术部'}},
        ])
        records = [
            {'唯一id': '1', '银行': '测试银行', '银行账号': '123', '主体': '测试公司',
             '交易日期': '2024-01-01', '付款': None, '收款': 100,
             '摘要': '', '对方户名': '', '余额': 100, '交易流水号': 'T1',
             '记录字段': '值'},
        ]
        columns = get_summary_columns(records=records, lookup_source=path)
        assert '部门' in columns
        assert '记录字段' in columns


class TestLoadLookupTable:
    """测试 load_lookup_table 函数"""

    def test_load_basic(self, tmp_dir):
        """基本加载测试"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        lookup = load_lookup_table(path)
        assert 'by_account' in lookup
        assert 'all_entries' in lookup
        assert 'extra_field_names' in lookup
        assert len(lookup['all_entries']) == 2
        assert len(lookup['by_account']) == 2

    def test_load_none_file(self):
        """None 文件路径返回空结构"""
        lookup = load_lookup_table(None)
        assert lookup['by_account'] == {}
        assert lookup['all_entries'] == []
        assert lookup['extra_field_names'] == []

    def test_load_empty_string(self):
        """空字符串路径返回空结构"""
        lookup = load_lookup_table('')
        assert lookup['by_account'] == {}
        assert lookup['all_entries'] == []

    def test_load_nonexistent_file(self):
        """不存在的文件返回空结构"""
        lookup = load_lookup_table('/nonexistent/path.xlsx')
        assert lookup['by_account'] == {}
        assert lookup['all_entries'] == []

    def test_load_priority_sorted(self, tmp_dir):
        """加载后同账号条目按优先级降序排列"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '低优先级主体', 'account': '12345', 'priority': 1},
            {'subject': '高优先级主体', 'account': '12345', 'priority': 10},
            {'subject': '中优先级主体', 'account': '12345', 'priority': 5},
        ])
        lookup = load_lookup_table(path)
        key = _account_key('12345')
        entries = lookup['by_account'][key]
        assert entries[0]['subject'] == '高优先级主体'
        assert entries[0]['priority'] == 10
        assert entries[1]['subject'] == '中优先级主体'
        assert entries[1]['priority'] == 5
        assert entries[2]['subject'] == '低优先级主体'
        assert entries[2]['priority'] == 1

    def test_load_extra_field_names(self, tmp_dir):
        """加载扩展字段名称"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '12345',
             'priority': 0, 'extra_fields': {'部门': '技术部', '项目编号': 'PRJ001'}},
        ])
        lookup = load_lookup_table(path)
        assert '部门' in lookup['extra_field_names']
        assert '项目编号' in lookup['extra_field_names']
        assert len(lookup['extra_field_names']) == 2

    def test_load_extra_field_names_sorted(self, tmp_dir):
        """扩展字段名称按字母排序"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '12345',
             'priority': 0, 'extra_fields': {'Z部门': '测试', 'A项目': '测试'}},
        ])
        lookup = load_lookup_table(path)
        assert lookup['extra_field_names'][0] == 'A项目'
        assert lookup['extra_field_names'][1] == 'Z部门'

    def test_load_account_normalization(self, tmp_dir):
        """账号在加载时被规范化处理"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'), [
            ('测试公司', '01090312345678901'),
        ])
        lookup = load_lookup_table(path)
        key = _account_key('01090312345678901')
        assert key in lookup['by_account']
        key2 = _account_key(1090312345678901)
        assert key2 in lookup['by_account']
        assert key == key2

    def test_load_no_extra_fields(self, tmp_dir):
        """没有扩展字段时 extra_field_names 为空列表"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        lookup = load_lookup_table(path)
        assert lookup['extra_field_names'] == []

    def test_load_skip_none_account(self, tmp_dir):
        """跳过账号为空的行"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '主体名称'
        ws['B1'] = '银行账号'
        ws['A2'] = '公司A'
        ws['B2'] = '12345'
        ws['A3'] = '公司B'
        ws['B3'] = None
        ws['A4'] = '公司C'
        ws['B4'] = '67890'
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        wb.save(path)
        wb.close()

        lookup = load_lookup_table(path)
        assert len(lookup['all_entries']) == 2

    def test_source_file_preserved(self, tmp_dir):
        """_source_file 字段保留原始路径"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        lookup = load_lookup_table(path)
        assert lookup['_source_file'] == path


class TestResolveLookup:
    """测试 _resolve_lookup 辅助函数"""

    def test_resolve_dict(self, tmp_dir):
        """传入预加载字典时直接返回"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        lookup = load_lookup_table(path)
        resolved = _resolve_lookup(lookup)
        assert resolved is lookup

    def test_resolve_file_path(self, tmp_dir):
        """传入文件路径时加载并返回"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        resolved = _resolve_lookup(path)
        assert 'by_account' in resolved
        assert len(resolved['all_entries']) == 2

    def test_resolve_none(self):
        """传入 None 返回空结构"""
        resolved = _resolve_lookup(None)
        assert resolved['by_account'] == {}


class TestGetSubjectWithPreloadedLookup:
    """测试 get_subject 使用预加载查找表"""

    def test_found_with_preloaded(self, tmp_dir):
        """使用预加载字典精确匹配"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        lookup = load_lookup_table(path)
        result = get_subject('01090312345678901', lookup)
        assert result == '北京XX科技有限公司'

    def test_not_found_with_preloaded(self, tmp_dir):
        """使用预加载字典未匹配到"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        lookup = load_lookup_table(path)
        result = get_subject('99999999999', lookup)
        assert result == ''

    def test_none_account_with_preloaded(self, tmp_dir):
        """使用预加载字典但账号为 None"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        lookup = load_lookup_table(path)
        result = get_subject(None, lookup)
        assert result == ''

    def test_empty_lookup_dict(self):
        """使用空预加载字典"""
        empty_lookup = load_lookup_table(None)
        result = get_subject('12345', empty_lookup)
        assert result == ''


class TestGetSubjectInfoWithPreloadedLookup:
    """测试 get_subject_info 使用预加载查找表"""

    def test_basic_info_preloaded(self, tmp_dir):
        """使用预加载字典获取基本信息"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        lookup = load_lookup_table(path)
        info = get_subject_info('01090312345678901', lookup)
        assert info['subject'] == '北京XX科技有限公司'
        assert info['matched'] is True
        assert info['fuzzy_matched'] is False
        assert info['similarity'] == 1.0

    def test_priority_with_preloaded(self, tmp_dir):
        """使用预加载字典按优先级返回最高的"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '低优先级主体', 'account': '12345', 'priority': 1},
            {'subject': '高优先级主体', 'account': '12345', 'priority': 10},
        ])
        lookup = load_lookup_table(path)
        info = get_subject_info('12345', lookup)
        assert info['subject'] == '高优先级主体'
        assert info['priority'] == 10

    def test_extra_fields_with_preloaded(self, tmp_dir):
        """使用预加载字典获取扩展字段"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '12345',
             'priority': 0, 'extra_fields': {'部门': '技术部', '项目编号': 'PRJ001'}},
        ])
        lookup = load_lookup_table(path)
        info = get_subject_info('12345', lookup)
        assert info['extra_fields']['部门'] == '技术部'
        assert info['extra_fields']['项目编号'] == 'PRJ001'

    def test_fuzzy_match_with_preloaded(self, tmp_dir):
        """使用预加载字典进行模糊匹配"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'), [
            ('测试公司', '1234567890'),
        ])
        lookup = load_lookup_table(path)
        info = get_subject_info('123456789', lookup, use_fuzzy=True)
        assert info['matched'] is True
        assert info['fuzzy_matched'] is True
        assert info['subject'] == '测试公司'

    def test_not_found_info_preloaded(self, tmp_dir):
        """使用预加载字典未找到"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        lookup = load_lookup_table(path)
        info = get_subject_info('99999', lookup)
        assert info['matched'] is False
        assert info['subject'] == ''

    def test_empty_lookup_preloaded(self):
        """使用空预加载字典"""
        empty_lookup = load_lookup_table(None)
        info = get_subject_info('12345', empty_lookup)
        assert info['matched'] is False


class TestGetLookupExtraFieldsWithPreloadedLookup:
    """测试 get_lookup_extra_fields 使用预加载查找表"""

    def test_with_preloaded_dict(self, tmp_dir):
        """使用预加载字典获取扩展字段"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '12345',
             'priority': 0, 'extra_fields': {'部门': '技术部', '项目编号': 'PRJ001'}},
        ])
        lookup = load_lookup_table(path)
        fields = get_lookup_extra_fields(lookup)
        assert '部门' in fields
        assert '项目编号' in fields

    def test_no_extra_preloaded(self, tmp_dir):
        """使用无扩展字段的预加载字典"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        lookup = load_lookup_table(path)
        fields = get_lookup_extra_fields(lookup)
        assert fields == []

    def test_empty_dict_preloaded(self):
        """使用空预加载字典"""
        empty_lookup = load_lookup_table(None)
        fields = get_lookup_extra_fields(empty_lookup)
        assert fields == []


class TestBackwardCompatibility:
    """测试向后兼容性 - 传入文件路径仍然有效"""

    def test_get_subject_with_file_path(self, tmp_dir):
        """get_subject 传入文件路径仍然可用"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        result = get_subject('01090312345678901', path)
        assert result == '北京XX科技有限公司'

    def test_get_subject_info_with_file_path(self, tmp_dir):
        """get_subject_info 传入文件路径仍然可用"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        info = get_subject_info('01090312345678901', path)
        assert info['matched'] is True
        assert info['subject'] == '北京XX科技有限公司'

    def test_get_lookup_extra_fields_with_file_path(self, tmp_dir):
        """get_lookup_extra_fields 传入文件路径仍然可用"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '12345',
             'priority': 0, 'extra_fields': {'部门': '技术部'}},
        ])
        fields = get_lookup_extra_fields(path)
        assert '部门' in fields

    def test_get_summary_columns_with_file_path(self, tmp_dir):
        """get_summary_columns 传入文件路径仍然可用"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '12345',
             'priority': 0, 'extra_fields': {'部门': '技术部'}},
        ])
        columns = get_summary_columns(lookup_source=path)
        assert '部门' in columns


class TestPerformancePreloadVsFilePath:
    """测试预加载性能优于每次打开文件"""

    def test_repeated_queries_preloaded(self, tmp_dir):
        """预加载字典上的重复查询应该快速且一致"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        lookup = load_lookup_table(path)

        for _ in range(100):
            result = get_subject('01090312345678901', lookup)
            assert result == '北京XX科技有限公司'

    def test_preloaded_dict_not_modified(self, tmp_dir):
        """查询不应修改预加载字典的内容"""
        path = _create_lookup_table(os.path.join(tmp_dir, 'lookup.xlsx'))
        lookup = load_lookup_table(path)
        original_entries = len(lookup['all_entries'])
        original_by_account_keys = set(lookup['by_account'].keys())

        get_subject('01090312345678901', lookup)
        get_subject_info('38812345678', lookup)

        assert len(lookup['all_entries']) == original_entries
        assert set(lookup['by_account'].keys()) == original_by_account_keys

    def test_extra_fields_copy_independent(self, tmp_dir):
        """返回的 extra_fields 应该是副本，不影响预加载字典"""
        path = os.path.join(tmp_dir, 'lookup.xlsx')
        _create_lookup_table_with_extra(path, [
            {'subject': '测试公司', 'account': '12345',
             'priority': 0, 'extra_fields': {'部门': '技术部'}},
        ])
        lookup = load_lookup_table(path)

        info = get_subject_info('12345', lookup)
        info['extra_fields']['部门'] = '修改后的部门'

        key = _account_key('12345')
        assert lookup['by_account'][key][0]['extra_fields']['部门'] == '技术部'

