# -*- coding: utf-8 -*-
"""
变更影响评估模块单元测试
"""

import os
import sys
import shutil
import tempfile
import json

import pytest
import openpyxl
import yaml

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import lookup_manager as lm
import change_impact_evaluator as cie
from change_impact_evaluator import (
    _normalize_value,
    _values_equal,
    FieldDiff,
    RecordDiff,
    ImpactReport,
    HistorySampleManager,
    _compare_records,
    ChangeImpactEvaluator,
)


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='cie_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


@pytest.fixture
def sample_manager(tmp_dir):
    return HistorySampleManager(samples_dir=tmp_dir)


@pytest.fixture
def evaluator(tmp_dir):
    return ChangeImpactEvaluator(samples_dir=tmp_dir)


def _make_record(trade_date='2024-01-05', payment=None, receipt=None,
                 summary='测试', counterpart='对方A', balance=1000,
                 transaction_id='TX001', bank='北京银行', account='01090312345678901',
                 subject='测试公司', **kwargs):
    rec = {
        '唯一id': 'test-id-1',
        '银行': bank,
        '银行账号': account,
        '主体': subject,
        '交易日期': trade_date,
        '付款': payment,
        '收款': receipt,
        '摘要': summary,
        '对方户名': counterpart,
        '余额': balance,
        '交易流水号': transaction_id,
        '来源文件名': 'test.xlsx',
        '来源相对路径': 'test.xlsx',
        '处理时间': '2024-01-01 00:00:00',
    }
    rec.update(kwargs)
    return rec


def _create_test_bank_excel(filepath: str):
    """创建一个测试用的北京银行格式Excel文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '流水'

    ws['A1'] = '户名'
    ws['B1'] = '账号'
    ws['A2'] = '测试公司'
    ws['B2'] = '01090312345678901'

    headers = ['交易日期', '凭证号', '对方账号', '对方户名',
               '借方发生额', '贷方发生额', '余额', '摘要', '交易流水号']
    for col, header in enumerate(headers, 1):
        ws.cell(row=4, column=col, value=header)

    ws.cell(row=5, column=1, value='2024-01-15')
    ws.cell(row=5, column=2, value='P001')
    ws.cell(row=5, column=3, value='1234567890')
    ws.cell(row=5, column=4, value='供应商A')
    ws.cell(row=5, column=5, value=1000.00)
    ws.cell(row=5, column=6, value=None)
    ws.cell(row=5, column=7, value=9000.00)
    ws.cell(row=5, column=8, value='采购办公用品')
    ws.cell(row=5, column=9, value='TX202401150001')

    ws.cell(row=6, column=1, value='2024-01-16')
    ws.cell(row=6, column=2, value='P002')
    ws.cell(row=6, column=3, value='9876543210')
    ws.cell(row=6, column=4, value='客户B')
    ws.cell(row=6, column=5, value=None)
    ws.cell(row=6, column=6, value=5000.00)
    ws.cell(row=6, column=7, value=14000.00)
    ws.cell(row=6, column=8, value='销售收入')
    ws.cell(row=6, column=9, value='TX202401160002')

    wb.save(filepath)
    wb.close()


class TestNormalizeValue:
    def test_none(self):
        assert _normalize_value(None) is None

    def test_int(self):
        assert _normalize_value(123) == 123

    def test_float_whole(self):
        assert _normalize_value(123.0) == 123

    def test_float_with_decimals(self):
        assert _normalize_value(123.45) == 123.45

    def test_string_strip(self):
        assert _normalize_value('  hello  ') == 'hello'

    def test_empty_string(self):
        assert _normalize_value('   ') is None

    def test_empty_string_no_space(self):
        assert _normalize_value('') is None


class TestValuesEqual:
    def test_both_none(self):
        assert _values_equal(None, None) is True

    def test_one_none(self):
        assert _values_equal(None, 'a') is False
        assert _values_equal('a', None) is False

    def test_same_string(self):
        assert _values_equal('hello', 'hello') is True

    def test_different_string(self):
        assert _values_equal('hello', 'world') is False

    def test_int_and_string(self):
        assert _values_equal(123, '123') is True

    def test_float_and_int_equal(self):
        assert _values_equal(123.0, 123) is True

    def test_stripped_vs_not(self):
        assert _values_equal('  hello  ', 'hello') is True

    def test_empty_vs_none(self):
        assert _values_equal('', None) is True
        assert _values_equal(None, '   ') is True


class TestFieldDiff:
    def test_creation(self):
        fd = FieldDiff(field_name='主体', old_value='公司A', new_value='公司B', record_id='TX001')
        assert fd.field_name == '主体'
        assert fd.old_value == '公司A'
        assert fd.new_value == '公司B'
        assert fd.record_id == 'TX001'

    def test_default_record_id(self):
        fd = FieldDiff(field_name='摘要', old_value='old', new_value='new')
        assert fd.record_id == ''


class TestRecordDiff:
    def test_creation(self):
        fd = FieldDiff(field_name='主体', old_value='A', new_value='B')
        rd = RecordDiff(record_id='TX001', field_diffs=[fd])
        assert rd.record_id == 'TX001'
        assert len(rd.field_diffs) == 1

    def test_default_field_diffs(self):
        rd = RecordDiff(record_id='TX001')
        assert rd.field_diffs == []


class TestImpactReport:
    def test_creation(self):
        report = ImpactReport(
            change_type='lookup_table',
            change_target='主体查找表',
            total_records=100,
            affected_records=10,
            unaffected_records=85,
            added_records=3,
            removed_records=2,
        )
        assert report.change_type == 'lookup_table'
        assert report.change_target == '主体查找表'
        assert report.total_records == 100
        assert report.affected_records == 10

    def test_to_dict(self):
        report = ImpactReport(
            change_type='bank_rule',
            change_target='北京银行',
            total_records=10,
            affected_records=2,
            field_diff_summary={'主体': 2},
            field_diff_examples=[{'field_name': '主体', 'old_value': 'A', 'new_value': 'B'}],
        )
        d = report.to_dict()
        assert d['change_type'] == 'bank_rule'
        assert d['total_records'] == 10
        assert '主体' in d['field_diff_summary']
        assert len(d['field_diff_examples']) == 1

    def test_to_markdown_with_impact(self):
        report = ImpactReport(
            change_type='bank_rule',
            change_target='北京银行',
            timestamp='2024-01-01 12:00:00',
            total_records=100,
            affected_records=10,
            unaffected_records=90,
            added_records=0,
            removed_records=0,
            field_diff_summary={'摘要': 5, '对方户名': 3},
            field_diff_examples=[
                {'record_id': 'TX001', 'field_name': '摘要', 'old_value': '旧摘要', 'new_value': '新摘要'},
            ],
        )
        md = report.to_markdown()
        assert '# 变更影响评估报告' in md
        assert '北京银行' in md
        assert '| 受影响记录数 | 10 |' in md
        assert '## 字段差异统计' in md
        assert '## 字段差异示例' in md
        assert '## ⚠️ 风险提示' in md

    def test_to_markdown_no_impact(self):
        report = ImpactReport(
            change_type='lookup_table',
            change_target='主体查找表',
            timestamp='2024-01-01 12:00:00',
            total_records=100,
            affected_records=0,
            unaffected_records=100,
        )
        md = report.to_markdown()
        assert '## ✅ 评估结论' in md
        assert '未影响任何记录' in md

    def test_save_report(self, tmp_dir):
        report = ImpactReport(
            change_type='bank_rule',
            change_target='北京银行',
            total_records=10,
            affected_records=2,
        )
        md_path = report.save_report(output_dir=tmp_dir)
        assert os.path.exists(md_path)
        assert md_path.endswith('.md')

        json_path = md_path.replace('.md', '.json')
        assert os.path.exists(json_path)

        with open(md_path, 'r', encoding='utf-8') as f:
            content = f.read()
        assert '北京银行' in content


class TestHistorySampleManager:
    def test_save_and_load_sample(self, sample_manager, tmp_dir):
        records = [
            _make_record(transaction_id='TX001'),
            _make_record(transaction_id='TX002'),
        ]
        src_file = os.path.join(tmp_dir, 'test_source.xlsx')
        openpyxl.Workbook().save(src_file)

        saved_path = sample_manager.save_sample(src_file, '北京银行', records)
        assert saved_path is not None
        assert os.path.exists(saved_path)

        loaded = sample_manager.load_sample(saved_path)
        assert loaded is not None
        assert loaded['bank_name'] == '北京银行'
        assert loaded['record_count'] == 2
        assert len(loaded['records']) == 2
        assert loaded['source_full_path'] == os.path.abspath(src_file)
        assert 'source_hash' in loaded
        assert loaded['source_hash'] != ''

        meta = sample_manager.list_samples('北京银行')
        assert len(meta) == 1
        assert meta[0]['source_full_path'] == os.path.abspath(src_file)
        assert 'source_hash' in meta[0]

    def test_save_with_nonexistent_source_file(self, sample_manager, tmp_dir):
        """测试源文件不存在时仍能保存样本（哈希为空）"""
        records = [_make_record(transaction_id='TX001')]
        nonexistent_file = os.path.join(tmp_dir, 'not_exists.xlsx')

        saved_path = sample_manager.save_sample(nonexistent_file, '测试银行', records)
        assert saved_path is not None

        loaded = sample_manager.load_sample(saved_path)
        assert loaded is not None
        assert loaded['source_full_path'] == os.path.abspath(nonexistent_file)
        assert loaded['source_hash'] == ''

    def test_save_empty_records(self, sample_manager, tmp_dir):
        src_file = os.path.join(tmp_dir, 'empty.xlsx')
        openpyxl.Workbook().save(src_file)
        result = sample_manager.save_sample(src_file, '测试银行', [])
        assert result is None

    def test_list_samples(self, sample_manager, tmp_dir):
        src_file = os.path.join(tmp_dir, 'src.xlsx')
        openpyxl.Workbook().save(src_file)

        sample_manager.save_sample(src_file, '北京银行', [_make_record()])
        sample_manager.save_sample(src_file, '东亚银行', [_make_record()])

        all_samples = sample_manager.list_samples()
        assert len(all_samples) == 2

        bj_samples = sample_manager.list_samples('北京银行')
        assert len(bj_samples) == 1
        assert bj_samples[0]['bank_name'] == '北京银行'

    def test_load_samples_for_bank(self, sample_manager, tmp_dir):
        src_file = os.path.join(tmp_dir, 'src.xlsx')
        openpyxl.Workbook().save(src_file)

        for i in range(3):
            sample_manager.save_sample(src_file, '北京银行', [_make_record(transaction_id=f'TX{i:03d}')])

        samples = sample_manager.load_samples_for_bank('北京银行', max_samples=2)
        assert len(samples) == 2

    def test_load_nonexistent_sample(self, sample_manager):
        result = sample_manager.load_sample('/nonexistent/path.json')
        assert result is None

    def test_load_all_samples(self, sample_manager, tmp_dir):
        src_file = os.path.join(tmp_dir, 'src.xlsx')
        openpyxl.Workbook().save(src_file)

        sample_manager.save_sample(src_file, '北京银行', [_make_record()])
        sample_manager.save_sample(src_file, '东亚银行', [_make_record()])
        sample_manager.save_sample(src_file, '工商银行', [_make_record()])

        samples = sample_manager.load_all_samples(max_per_bank=1)
        assert len(samples) == 3


class TestCompareRecords:
    def test_identical_records(self):
        old = [_make_record(transaction_id='TX001')]
        new = [_make_record(transaction_id='TX001')]
        diffs, total, affected, unaffected, added, removed = _compare_records(old, new)
        assert total == 1
        assert affected == 0
        assert unaffected == 1
        assert added == 0
        assert removed == 0
        assert len(diffs) == 0

    def test_different_subject(self):
        old = [_make_record(transaction_id='TX001', subject='公司A')]
        new = [_make_record(transaction_id='TX001', subject='公司B')]
        diffs, total, affected, unaffected, added, removed = _compare_records(old, new)
        assert total == 1
        assert affected == 1
        assert len(diffs) == 1
        assert any(fd.field_name == '主体' for fd in diffs[0].field_diffs)

    def test_added_record(self):
        old = [_make_record(transaction_id='TX001')]
        new = [_make_record(transaction_id='TX001'), _make_record(transaction_id='TX002')]
        diffs, total, affected, unaffected, added, removed = _compare_records(old, new)
        assert added == 1

    def test_removed_record(self):
        old = [_make_record(transaction_id='TX001'), _make_record(transaction_id='TX002')]
        new = [_make_record(transaction_id='TX001')]
        diffs, total, affected, unaffected, added, removed = _compare_records(old, new)
        assert removed == 1

    def test_multiple_field_diffs(self):
        old = [_make_record(transaction_id='TX001', subject='公司A', balance=1000)]
        new = [_make_record(transaction_id='TX001', subject='公司B', balance=2000)]
        diffs, total, affected, unaffected, added, removed = _compare_records(old, new)
        assert affected == 1
        assert len(diffs[0].field_diffs) == 2

    def test_empty_records(self):
        diffs, total, affected, unaffected, added, removed = _compare_records([], [])
        assert total == 0
        assert affected == 0

    def test_ignores_traceability_fields(self):
        old = [_make_record(transaction_id='TX001')]
        old[0]['处理时间'] = '2024-01-01 00:00:00'
        new = [_make_record(transaction_id='TX001')]
        new[0]['处理时间'] = '2024-12-31 23:59:59'
        new[0]['唯一id'] = 'different-id'

        diffs, total, affected, unaffected, added, removed = _compare_records(old, new)
        assert unaffected == 1


class TestChangeImpactEvaluator:
    def test_evaluate_lookup_change_with_samples(self, evaluator, tmp_dir):
        records = [
            _make_record(account='01090312345678901', subject='旧公司名', transaction_id='TX001'),
            _make_record(account='38812345678', subject='上海YY贸易有限公司', transaction_id='TX002'),
            _make_record(account='99999999999', subject='', transaction_id='TX003'),
        ]
        sample_data = {
            'source_file': 'test.xlsx',
            'bank_name': '测试银行',
            'records': records,
        }
        evaluator.sample_manager.save_sample(
            os.path.join(tmp_dir, 'src.xlsx'), '测试银行', records
        )

        old_lookup = os.path.join(tmp_dir, 'old_lookup.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '主体名称'
        ws['B1'] = '银行账号'
        ws['A2'] = '旧公司名'
        ws['B2'] = '01090312345678901'
        ws['A3'] = '上海YY贸易有限公司'
        ws['B3'] = '38812345678'
        wb.save(old_lookup)
        wb.close()

        new_entries = [
            lm.LookupEntry(subject='新公司名', account='01090312345678901'),
            lm.LookupEntry(subject='上海YY贸易有限公司', account='38812345678'),
            lm.LookupEntry(subject='新增公司', account='99999999999'),
        ]

        report = evaluator.evaluate_lookup_change(
            old_lookup_file=old_lookup,
            new_lookup_entries=new_entries,
        )

        assert report.change_type == 'lookup_table'
        assert report.total_records >= 1
        assert report.affected_records >= 1
        assert '主体' in report.field_diff_summary
        assert len(report.field_diff_examples) >= 1

    def test_evaluate_lookup_change_no_samples(self, evaluator, tmp_dir):
        new_entries = [
            lm.LookupEntry(subject='公司A', account='12345'),
        ]
        report = evaluator.evaluate_lookup_change(None, new_entries)
        assert '未找到历史样本' in report.details

    def test_evaluate_bank_rule_change_with_samples(self, evaluator, tmp_dir):
        records = [
            _make_record(bank='北京银行', account='01090312345678901',
                        transaction_id='TX001', summary='采购付款', counterpart='供应商A'),
        ]
        evaluator.sample_manager.save_sample(
            os.path.join(tmp_dir, 'src.xlsx'), '北京银行', records
        )

        old_rule = {
            'bank_name': '北京银行',
            'account_cell': 'B2',
            'start_row': 4,
            'columns': {
                'trade_date': 2,
                'payment': 4,
                'receipt': 5,
                'summary': 12,
                'counterpart': 7,
                'balance': 6,
                'transaction_id': 16,
            },
        }
        new_rule = dict(old_rule)
        new_rule['columns'] = dict(old_rule['columns'])
        new_rule['columns']['summary'] = 13
        new_rule['columns']['counterpart'] = 12

        report = evaluator.evaluate_bank_rule_change('北京银行', old_rule, new_rule)
        assert report.change_type == 'bank_rule'
        assert report.total_records >= 1

    def test_evaluate_bank_rule_change_no_samples(self, evaluator):
        report = evaluator.evaluate_bank_rule_change('不存在银行', {}, {})
        assert '未找到该银行的历史样本' in report.details

    def test_save_sample_from_records(self, evaluator, tmp_dir):
        src = os.path.join(tmp_dir, 'test.xlsx')
        openpyxl.Workbook().save(src)
        records = [_make_record()]
        path = evaluator.save_sample_from_records(src, '测试银行', records)
        assert path is not None
        assert os.path.exists(path)

    def test_evaluate_bank_rule_change_real_reparse(self, evaluator, tmp_dir):
        """测试银行规则变更时使用真实样本文件重跑"""
        import bankcheck

        test_xlsx = os.path.join(tmp_dir, '北京银行_test.xlsx')
        _create_test_bank_excel(test_xlsx)

        rule_data = {
            'bank_name': '北京银行',
            'account_cell': 'B2',
            'start_row': 5,
            'payment_sign': 'negative',
            'enabled': True,
            'columns': {
                'trade_date': 1,
                'counterpart': 4,
                'payment': 5,
                'receipt': 6,
                'balance': 7,
                'summary': 8,
                'transaction_id': 9,
            },
        }

        rule = bankcheck.BankRule(**rule_data)
        parser = bankcheck.GenericBankParser(rule)
        lookup_data = bankcheck.load_lookup_table(None)
        records = parser.parse(test_xlsx, lookup_data, base_dir=tmp_dir)

        assert len(records) == 2
        assert records[0]['摘要'] == '采购办公用品'
        assert records[0]['对方户名'] == '供应商A'

        evaluator.save_sample_from_records(test_xlsx, '北京银行', records)

        old_rule = dict(rule_data)
        new_rule = dict(rule_data)
        new_rule['columns'] = dict(rule_data['columns'])
        new_rule['columns']['summary'] = 2
        new_rule['columns']['counterpart'] = 3

        report = evaluator.evaluate_bank_rule_change('北京银行', old_rule, new_rule)

        assert report.change_type == 'bank_rule'
        assert report.total_records >= 2
        assert '真实重跑' in report.details
        assert '模拟解析' in report.details

        if report.affected_records > 0:
            assert '摘要' in report.field_diff_summary or '对方户名' in report.field_diff_summary

    def test_evaluate_bank_rule_change_fallback_to_simulated(self, evaluator, tmp_dir):
        """测试原始文件不存在时自动降级到模拟解析"""
        nonexistent_xlsx = os.path.join(tmp_dir, '已删除_北京银行.xlsx')

        records = [
            _make_record(bank='北京银行', account='01090312345678901',
                        transaction_id='TX001', summary='测试摘要', counterpart='测试对方'),
        ]

        sample_data = {
            'source_file': os.path.basename(nonexistent_xlsx),
            'source_full_path': os.path.abspath(nonexistent_xlsx),
            'source_hash': '',
            'bank_name': '北京银行',
            'saved_at': '20240101_000000',
            'record_count': 1,
            'records': records,
        }

        sample_filename = '北京银行_20240101_000000.json'
        sample_path = os.path.join(evaluator.sample_manager.samples_dir, sample_filename)
        with open(sample_path, 'w', encoding='utf-8') as f:
            json.dump(sample_data, f, ensure_ascii=False, default=str)

        evaluator.sample_manager._load_metadata()
        metadata = [{
            'filename': sample_filename,
            'bank_name': '北京银行',
            'source_file': os.path.basename(nonexistent_xlsx),
            'source_full_path': os.path.abspath(nonexistent_xlsx),
            'source_hash': '',
            'saved_at': '20240101_000000',
            'record_count': 1,
            'sample_path': sample_path,
        }]
        evaluator.sample_manager._save_metadata(metadata)

        old_rule = {
            'bank_name': '北京银行',
            'account_cell': 'B2',
            'start_row': 5,
            'columns': {
                'trade_date': 1,
                'summary': 8,
                'counterpart': 4,
            },
        }
        new_rule = dict(old_rule)
        new_rule['columns'] = dict(old_rule['columns'])
        new_rule['columns']['summary'] = 9

        report = evaluator.evaluate_bank_rule_change('北京银行', old_rule, new_rule)

        assert report.change_type == 'bank_rule'
        assert report.total_records >= 1
        assert '模拟解析' in report.details


class TestIntegrationWithExistingCode:
    def test_bankcheck_save_rule_returns_tuple(self, tmp_dir):
        import bankcheck
        rule_yaml = os.path.join(tmp_dir, 'bank_rules_test.yaml')
        config_data = {
            'banks': [{
                'bank_name': '测试银行',
                'account_cell': 'A1',
                'start_row': 2,
                'payment_sign': 'negative',
                'enabled': True,
                'columns': {'trade_date': 1},
            }]
        }
        with open(rule_yaml, 'w', encoding='utf-8') as f:
            yaml.dump(config_data, f, allow_unicode=True)

        config = bankcheck.BankRuleConfig(config_path=rule_yaml)
        new_rule = {
            'bank_name': '测试银行',
            'account_cell': 'B2',
            'start_row': 3,
            'columns': {'trade_date': 1, 'payment': 2},
        }
        result = config.save_rule(new_rule, run_impact_eval=False)
        assert isinstance(result, tuple)
        assert len(result) == 2
        assert result[0] is True

    def test_lookup_manager_save_entries_returns_tuple(self, tmp_dir):
        import lookup_manager as lm
        path = os.path.join(tmp_dir, 'test_lookup.xlsx')
        entries = [
            lm.LookupEntry(subject='公司A', account='111'),
        ]
        result = lm.save_lookup_entries(entries, path, run_impact_eval=False)
        assert isinstance(result, tuple)
        assert len(result) == 2
        assert result[0] is True

        loaded = lm.read_lookup_entries(path)
        assert len(loaded) == 1
        assert loaded[0].subject == '公司A'
