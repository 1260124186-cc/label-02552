"""
主体维度汇总分析模块单元测试
"""
import os
import sys
import tempfile
import shutil
from datetime import datetime

import openpyxl
import pandas as pd
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='subject_summary_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _create_test_records():
    """创建测试用交易记录，覆盖多主体、多银行、多月份"""
    return [
        {
            '唯一id': 'TEST001', '银行': '北京银行', '银行账号': '01090312345678901',
            '主体': '北京XX科技有限公司', '交易日期': '2024-01-05',
            '付款': -50000.0, '收款': None, '摘要': '采购付款-办公设备',
            '对方户名': '供应商A公司', '余额': 1500000.0, '交易流水号': 'BJ20240105001',
        },
        {
            '唯一id': 'TEST002', '银行': '北京银行', '银行账号': '01090312345678901',
            '主体': '北京XX科技有限公司', '交易日期': '2024-01-10',
            '付款': None, '收款': 80000.0, '摘要': '销售收款-产品销售',
            '对方户名': '客户B公司', '余额': 1580000.0, '交易流水号': 'BJ20240110002',
        },
        {
            '唯一id': 'TEST003', '银行': '北京银行', '银行账号': '01090312345678901',
            '主体': '北京XX科技有限公司', '交易日期': '2024-02-05',
            '付款': -20000.0, '收款': None, '摘要': '采购付款-原材料',
            '对方户名': '供应商C公司', '余额': 1560000.0, '交易流水号': 'BJ20240205003',
        },
        {
            '唯一id': 'TEST004', '银行': '北京银行', '银行账号': '01090312345678901',
            '主体': '北京XX科技有限公司', '交易日期': '2024-02-15',
            '付款': None, '收款': 60000.0, '摘要': '销售收款-服务收入',
            '对方户名': '客户D公司', '余额': 1620000.0, '交易流水号': 'BJ20240215004',
        },
        {
            '唯一id': 'TEST005', '银行': '东亚银行', '银行账号': '38812345678',
            '主体': '上海YY贸易有限公司', '交易日期': '2024-01-15',
            '付款': -1500.0, '收款': None, '摘要': '手续费',
            '对方户名': '银行手续费', '余额': 513500.0, '交易流水号': 'EA20240115005',
        },
        {
            '唯一id': 'TEST006', '银行': '东亚银行', '银行账号': '38812345678',
            '主体': '上海YY贸易有限公司', '交易日期': '2024-01-20',
            '付款': -3000.0, '收款': None, '摘要': '差旅费报销',
            '对方户名': '员工张三', '余额': 510500.0, '交易流水号': 'EA20240120006',
        },
        {
            '唯一id': 'TEST007', '银行': '东亚银行', '银行账号': '38812345678',
            '主体': '上海YY贸易有限公司', '交易日期': '2024-02-10',
            '付款': None, '收款': 120000.0, '摘要': '贸易收入',
            '对方户名': '海外客户E', '余额': 630500.0, '交易流水号': 'EA20240210007',
        },
        {
            '唯一id': 'TEST008', '银行': '招商银行', '银行账号': '6225880112345678',
            '主体': '北京XX科技有限公司', '交易日期': '2024-03-01',
            '付款': -10000.0, '收款': None, '摘要': '工资发放',
            '对方户名': '员工工资', '余额': 1610000.0, '交易流水号': 'CMB20240301008',
        },
        {
            '唯一id': 'TEST009', '银行': '招商银行', '银行账号': '6225880112345678',
            '主体': '深圳ZZ互联网有限公司', '交易日期': '2024-03-15',
            '付款': None, '收款': 200000.0, '摘要': '投资款',
            '对方户名': '投资人F', '余额': 200000.0, '交易流水号': 'CMB20240315009',
        },
    ]


def _create_test_total_table(path, records=None):
    """创建测试用总表文件"""
    if records is None:
        records = _create_test_records()
    columns = ['唯一id', '银行', '银行账号', '主体', '交易日期',
               '付款', '收款', '摘要', '对方户名', '余额', '交易流水号']
    df = pd.DataFrame(records, columns=columns)
    df.to_excel(path, index=False, engine='openpyxl')
    return path


class TestDataClasses:
    """测试数据类"""

    def test_subject_dimension_summary_defaults(self):
        """测试 SubjectDimensionSummary 默认值"""
        s = bankcheck.SubjectDimensionSummary()
        assert s.subject == ''
        assert s.bank == ''
        assert s.year_month == ''
        assert s.total_income == 0.0
        assert s.total_expense == 0.0
        assert s.net_amount == 0.0
        assert s.transaction_count == 0
        assert s.income_count == 0
        assert s.expense_count == 0

    def test_subject_summary_result_defaults(self):
        """测试 SubjectSummaryResult 默认值"""
        r = bankcheck.SubjectSummaryResult()
        assert r.by_subject == []
        assert r.by_subject_bank == []
        assert r.by_subject_month == []
        assert r.by_subject_bank_month == []
        assert r.by_bank == []
        assert r.by_month == []
        assert r.overall_summary == {}


class TestExtractYearMonth:
    """测试 _extract_year_month 函数"""

    def test_yyyy_mm_dd_format(self):
        """测试 YYYY-MM-DD 格式"""
        assert bankcheck._extract_year_month('2024-01-15') == '2024-01'
        assert bankcheck._extract_year_month('2024-12-31') == '2024-12'

    def test_yyyy_mm_dd_slash_format(self):
        """测试 YYYY/MM/DD 格式"""
        assert bankcheck._extract_year_month('2024/03/20') == '2024-03'

    def test_datetime_object(self):
        """测试 datetime 对象"""
        dt = datetime(2024, 5, 10)
        assert bankcheck._extract_year_month(dt) == '2024-05'

    def test_pandas_timestamp(self):
        """测试 pandas Timestamp"""
        ts = pd.Timestamp('2024-06-15')
        assert bankcheck._extract_year_month(ts) == '2024-06'

    def test_none_value(self):
        """测试 None 值"""
        assert bankcheck._extract_year_month(None) == '未知'

    def test_empty_string(self):
        """测试空字符串"""
        assert bankcheck._extract_year_month('') == '未知'

    def test_yyyymmdd_numeric(self):
        """测试 YYYYMMDD 纯数字格式"""
        assert bankcheck._extract_year_month('20240720') == '2024-07'


class TestSummarizeTransactions:
    """测试 summarize_transactions 核心汇总函数"""

    def test_empty_records(self):
        """测试空记录列表"""
        result = bankcheck.summarize_transactions([])
        assert result.overall_summary.get('transaction_count') == 0
        assert result.by_subject == []
        assert result.by_bank == []

    def test_none_records(self):
        """测试 None 输入"""
        result = bankcheck.summarize_transactions(None)
        assert result.overall_summary.get('transaction_count') == 0

    def test_overall_summary_correct(self):
        """测试总体汇总统计"""
        records = _create_test_records()
        result = bankcheck.summarize_transactions(records)

        overall = result.overall_summary
        assert overall['transaction_count'] == 9
        assert overall['income_count'] == 4
        assert overall['expense_count'] == 5
        assert overall['total_income'] == pytest.approx(460000.0)
        assert overall['total_expense'] == pytest.approx(84500.0)
        assert overall['net_amount'] == pytest.approx(375500.0)
        assert overall['subject_count'] == 3
        assert overall['bank_count'] == 3
        assert overall['month_count'] == 3

    def test_by_subject_aggregation(self):
        """测试按主体汇总"""
        records = _create_test_records()
        result = bankcheck.summarize_transactions(records)

        by_subject = {row['subject']: row for row in result.by_subject}

        assert '北京XX科技有限公司' in by_subject
        bj = by_subject['北京XX科技有限公司']
        assert bj['total_income'] == pytest.approx(140000.0)
        assert bj['total_expense'] == pytest.approx(80000.0)
        assert bj['net_amount'] == pytest.approx(60000.0)
        assert bj['transaction_count'] == 5
        assert bj['income_count'] == 2
        assert bj['expense_count'] == 3

        assert '上海YY贸易有限公司' in by_subject
        sh = by_subject['上海YY贸易有限公司']
        assert sh['total_income'] == pytest.approx(120000.0)
        assert sh['total_expense'] == pytest.approx(4500.0)
        assert sh['net_amount'] == pytest.approx(115500.0)
        assert sh['transaction_count'] == 3

        assert '深圳ZZ互联网有限公司' in by_subject
        sz = by_subject['深圳ZZ互联网有限公司']
        assert sz['total_income'] == pytest.approx(200000.0)
        assert sz['total_expense'] == pytest.approx(0.0)
        assert sz['net_amount'] == pytest.approx(200000.0)
        assert sz['transaction_count'] == 1

    def test_by_bank_aggregation(self):
        """测试按银行汇总"""
        records = _create_test_records()
        result = bankcheck.summarize_transactions(records)

        by_bank = {row['bank']: row for row in result.by_bank}

        assert '北京银行' in by_bank
        bj = by_bank['北京银行']
        assert bj['total_income'] == pytest.approx(140000.0)
        assert bj['total_expense'] == pytest.approx(70000.0)
        assert bj['transaction_count'] == 4

        assert '东亚银行' in by_bank
        ea = by_bank['东亚银行']
        assert ea['total_income'] == pytest.approx(120000.0)
        assert ea['total_expense'] == pytest.approx(4500.0)
        assert ea['transaction_count'] == 3

        assert '招商银行' in by_bank
        cmb = by_bank['招商银行']
        assert cmb['total_income'] == pytest.approx(200000.0)
        assert cmb['total_expense'] == pytest.approx(10000.0)
        assert cmb['transaction_count'] == 2

    def test_by_month_aggregation(self):
        """测试按月份汇总"""
        records = _create_test_records()
        result = bankcheck.summarize_transactions(records)

        by_month = {row['year_month']: row for row in result.by_month}

        assert '2024-01' in by_month
        jan = by_month['2024-01']
        assert jan['total_income'] == pytest.approx(80000.0)
        assert jan['total_expense'] == pytest.approx(54500.0)
        assert jan['transaction_count'] == 4

        assert '2024-02' in by_month
        feb = by_month['2024-02']
        assert feb['total_income'] == pytest.approx(180000.0)
        assert feb['total_expense'] == pytest.approx(20000.0)
        assert feb['transaction_count'] == 3

        assert '2024-03' in by_month
        mar = by_month['2024-03']
        assert mar['total_income'] == pytest.approx(200000.0)
        assert mar['total_expense'] == pytest.approx(10000.0)
        assert mar['transaction_count'] == 2

    def test_by_subject_bank_month_3d_aggregation(self):
        """测试主体+银行+月份三维交叉汇总"""
        records = _create_test_records()
        result = bankcheck.summarize_transactions(records)

        bj_tech_bj_bank_jan = [
            r for r in result.by_subject_bank_month
            if r['subject'] == '北京XX科技有限公司'
            and r['bank'] == '北京银行'
            and r['year_month'] == '2024-01'
        ]
        assert len(bj_tech_bj_bank_jan) == 1
        row = bj_tech_bj_bank_jan[0]
        assert row['total_income'] == pytest.approx(80000.0)
        assert row['total_expense'] == pytest.approx(50000.0)
        assert row['net_amount'] == pytest.approx(30000.0)
        assert row['transaction_count'] == 2
        assert row['income_count'] == 1
        assert row['expense_count'] == 1

    def test_missing_subject_handled(self):
        """测试缺失主体字段的处理"""
        records = [
            {
                '唯一id': 'MISS001', '银行': '测试银行', '银行账号': '123',
                '主体': None, '交易日期': '2024-01-05',
                '付款': None, '收款': 1000.0, '摘要': '测试',
                '对方户名': '', '余额': 1000.0, '交易流水号': 'T001',
            },
        ]
        result = bankcheck.summarize_transactions(records)
        subjects = [r['subject'] for r in result.by_subject]
        assert '未指定主体' in subjects

    def test_zero_amounts_ignored(self):
        """测试零金额记录被忽略"""
        records = [
            {
                '唯一id': 'ZERO001', '银行': '测试银行', '银行账号': '123',
                '主体': '测试主体', '交易日期': '2024-01-05',
                '付款': 0.0, '收款': 0.0, '摘要': '零测试',
                '对方户名': '', '余额': 0.0, '交易流水号': 'Z001',
            },
        ]
        result = bankcheck.summarize_transactions(records)
        assert result.overall_summary['transaction_count'] == 0


class TestExportSubjectSummary:
    """测试 Excel 导出功能"""

    def test_export_creates_file(self, tmp_dir):
        """测试导出能正确创建文件"""
        records = _create_test_records()
        summary = bankcheck.summarize_transactions(records)
        output_path = os.path.join(tmp_dir, '汇总分析.xlsx')

        result = bankcheck.export_subject_summary(summary, output_path)

        assert result == output_path
        assert os.path.exists(output_path)
        assert os.path.getsize(output_path) > 0

    def test_export_sheet_names(self, tmp_dir):
        """测试导出的 Excel 包含正确的 Sheet"""
        records = _create_test_records()
        summary = bankcheck.summarize_transactions(records)
        output_path = os.path.join(tmp_dir, '汇总分析.xlsx')
        bankcheck.export_subject_summary(summary, output_path)

        wb = openpyxl.load_workbook(output_path)
        sheet_names = wb.sheetnames

        assert '汇总总览' in sheet_names
        assert '按主体汇总' in sheet_names
        assert '按主体+银行' in sheet_names
        assert '按主体+月份' in sheet_names
        assert '主体+银行+月份' in sheet_names
        assert '按银行汇总' in sheet_names
        assert '按月份汇总' in sheet_names

    def test_export_overview_sheet_content(self, tmp_dir):
        """测试汇总总览 Sheet 的内容"""
        records = _create_test_records()
        summary = bankcheck.summarize_transactions(records)
        output_path = os.path.join(tmp_dir, '汇总分析.xlsx')
        bankcheck.export_subject_summary(summary, output_path)

        df = pd.read_excel(output_path, sheet_name='汇总总览')
        data = dict(zip(df['统计项'], df['数值']))

        assert data['交易总笔数'] == 9
        assert data['收入笔数'] == 4
        assert data['支出笔数'] == 5
        assert data['收入总额(元)'] == pytest.approx(460000.0)
        assert data['支出总额(元)'] == pytest.approx(84500.0)
        assert data['净额(元)'] == pytest.approx(375500.0)
        assert data['涉及主体数'] == 3
        assert data['涉及银行数'] == 3
        assert data['覆盖月份数'] == 3

    def test_export_by_subject_sheet_content(self, tmp_dir):
        """测试按主体汇总 Sheet 内容"""
        records = _create_test_records()
        summary = bankcheck.summarize_transactions(records)
        output_path = os.path.join(tmp_dir, '汇总分析.xlsx')
        bankcheck.export_subject_summary(summary, output_path)

        df = pd.read_excel(output_path, sheet_name='按主体汇总')
        assert '主体' in df.columns
        assert '收入总额(元)' in df.columns
        assert '支出总额(元)' in df.columns
        assert '净额(元)' in df.columns
        assert '交易笔数' in df.columns
        assert len(df) == 3

    def test_export_with_source_info(self, tmp_dir):
        """测试携带数据源信息的导出"""
        records = _create_test_records()
        summary = bankcheck.summarize_transactions(records)
        output_path = os.path.join(tmp_dir, '汇总分析.xlsx')
        source_info = {
            '数据来源文件': '银行流水总表.xlsx',
            '总表记录数': 9,
            '生成时间': '2024-03-20 10:00:00',
        }

        bankcheck.export_subject_summary(summary, output_path, source_info)

        df = pd.read_excel(output_path, sheet_name='汇总总览')
        data = dict(zip(df['统计项'], df['数值']))

        assert '数据来源文件' in data
        assert data['数据来源文件'] == '银行流水总表.xlsx'
        assert data['总表记录数'] == 9

    def test_export_empty_result(self, tmp_dir):
        """测试空数据的导出"""
        summary = bankcheck.SubjectSummaryResult()
        output_path = os.path.join(tmp_dir, '空汇总.xlsx')

        bankcheck.export_subject_summary(summary, output_path)

        assert os.path.exists(output_path)
        wb = openpyxl.load_workbook(output_path)
        assert '汇总总览' in wb.sheetnames


class TestGenerateFromRecords:
    """测试从记录直接生成汇总"""

    def test_generate_from_records(self, tmp_dir):
        """测试 generate_subject_summary_from_records"""
        records = _create_test_records()
        result_path = bankcheck.generate_subject_summary_from_records(
            records, tmp_dir, {'来源': '测试'}
        )

        assert result_path is not None
        assert os.path.exists(result_path)
        assert '主体维度汇总分析' in os.path.basename(result_path)

    def test_generate_from_empty_records(self, tmp_dir):
        """测试空记录返回 None"""
        result_path = bankcheck.generate_subject_summary_from_records([], tmp_dir)
        assert result_path is None


class TestGenerateFromTotal:
    """测试从总表文件生成汇总"""

    def test_generate_from_total_file(self, tmp_dir):
        """测试 generate_subject_summary_from_total"""
        total_path = os.path.join(tmp_dir, '总表.xlsx')
        _create_test_total_table(total_path)

        result_path = bankcheck.generate_subject_summary_from_total(total_path, tmp_dir)

        assert result_path is not None
        assert os.path.exists(result_path)

        df = pd.read_excel(result_path, sheet_name='汇总总览')
        data = dict(zip(df['统计项'], df['数值']))
        assert data['交易总笔数'] == 9

    def test_generate_from_nonexistent_file(self, tmp_dir):
        """测试总表不存在的情况"""
        result = bankcheck.generate_subject_summary_from_total(
            os.path.join(tmp_dir, '不存在.xlsx'), tmp_dir
        )
        assert result is None


class TestNumberFormat:
    """测试金额计算和格式化"""

    def test_net_amount_calculation(self):
        """测试净额 = 收入 - 支出"""
        records = [
            {
                '唯一id': 'A1', '银行': '测试银行', '银行账号': '1',
                '主体': '测试主体', '交易日期': '2024-01-01',
                '付款': None, '收款': 10000.0, '摘要': '收入',
                '对方户名': '', '余额': 10000.0, '交易流水号': 'A1',
            },
            {
                '唯一id': 'A2', '银行': '测试银行', '银行账号': '1',
                '主体': '测试主体', '交易日期': '2024-01-02',
                '付款': -3500.5, '收款': None, '摘要': '支出',
                '对方户名': '', '余额': 6499.5, '交易流水号': 'A2',
            },
        ]
        result = bankcheck.summarize_transactions(records)
        row = result.by_subject[0]
        assert row['total_income'] == pytest.approx(10000.0)
        assert row['total_expense'] == pytest.approx(3500.5)
        assert row['net_amount'] == pytest.approx(6499.5)

    def test_rounding_to_two_decimals(self):
        """测试金额四舍五入到两位小数"""
        records = [
            {
                '唯一id': 'R1', '银行': '测试银行', '银行账号': '1',
                '主体': '测试主体', '交易日期': '2024-01-01',
                '付款': None, '收款': 100.123, '摘要': '',
                '对方户名': '', '余额': 0, '交易流水号': 'R1',
            },
            {
                '唯一id': 'R2', '银行': '测试银行', '银行账号': '1',
                '主体': '测试主体', '交易日期': '2024-01-02',
                '付款': -50.456, '收款': None, '摘要': '',
                '对方户名': '', '余额': 0, '交易流水号': 'R2',
            },
        ]
        result = bankcheck.summarize_transactions(records)
        row = result.by_subject[0]
        assert str(row['total_income']).split('.')[-1] in ['12', '13'] or row['total_income'] == pytest.approx(100.12)
        assert len(str(row['net_amount']).split('.')[-1]) <= 2


class TestUnmatchedAccounts:
    """测试未匹配账号汇总功能"""

    def test_no_unmatched_accounts_when_all_matched(self):
        """测试所有账号都匹配时，未匹配账号列表为空"""
        records = [
            {
                '唯一id': 'T1', '银行': '测试银行', '银行账号': '123456',
                '主体': '匹配主体A', '交易日期': '2024-01-01',
                '付款': None, '收款': 1000.0, '摘要': '收入',
                '对方户名': '', '余额': 1000.0, '交易流水号': 'T1',
            },
        ]
        result = bankcheck.summarize_transactions(records)
        assert result.unmatched_accounts == []
        assert result.overall_summary.get('unmatched_account_count', 0) == 0

    def test_single_unmatched_account(self):
        """测试单个未匹配账号的收集"""
        records = [
            {
                '唯一id': 'U1', '银行': '测试银行', '银行账号': '999999',
                '主体': '', '交易日期': '2024-01-05',
                '付款': None, '收款': 5000.0, '摘要': '收入',
                '对方户名': '', '余额': 5000.0, '交易流水号': 'U1',
            },
            {
                '唯一id': 'U2', '银行': '测试银行', '银行账号': '999999',
                '主体': '', '交易日期': '2024-01-10',
                '付款': -2000.0, '收款': None, '摘要': '支出',
                '对方户名': '', '余额': 3000.0, '交易流水号': 'U2',
            },
        ]
        result = bankcheck.summarize_transactions(records)

        assert len(result.unmatched_accounts) == 1
        ua = result.unmatched_accounts[0]
        assert ua['account'] == '999999'
        assert ua['banks'] == '测试银行'
        assert ua['first_date'] == '2024-01-05'
        assert ua['last_date'] == '2024-01-10'
        assert ua['total_income'] == pytest.approx(5000.0)
        assert ua['total_expense'] == pytest.approx(2000.0)
        assert ua['net_amount'] == pytest.approx(3000.0)
        assert ua['transaction_count'] == 2
        assert result.overall_summary.get('unmatched_account_count') == 1

    def test_multiple_unmatched_accounts(self):
        """测试多个未匹配账号的收集，按交易笔数降序排列"""
        records = [
            {
                '唯一id': 'A1', '银行': '银行A', '银行账号': '111111',
                '主体': '', '交易日期': '2024-01-01',
                '付款': None, '收款': 100.0, '摘要': '',
                '对方户名': '', '余额': 100.0, '交易流水号': 'A1',
            },
            {
                '唯一id': 'B1', '银行': '银行B', '银行账号': '222222',
                '主体': '', '交易日期': '2024-01-01',
                '付款': None, '收款': 200.0, '摘要': '',
                '对方户名': '', '余额': 200.0, '交易流水号': 'B1',
            },
            {
                '唯一id': 'B2', '银行': '银行B', '银行账号': '222222',
                '主体': '', '交易日期': '2024-01-02',
                '付款': None, '收款': 300.0, '摘要': '',
                '对方户名': '', '余额': 500.0, '交易流水号': 'B2',
            },
            {
                '唯一id': 'B3', '银行': '银行B', '银行账号': '222222',
                '主体': '', '交易日期': '2024-01-03',
                '付款': None, '收款': 400.0, '摘要': '',
                '对方户名': '', '余额': 900.0, '交易流水号': 'B3',
            },
        ]
        result = bankcheck.summarize_transactions(records)

        assert len(result.unmatched_accounts) == 2
        assert result.unmatched_accounts[0]['account'] == '222222'
        assert result.unmatched_accounts[0]['transaction_count'] == 3
        assert result.unmatched_accounts[1]['account'] == '111111'
        assert result.unmatched_accounts[1]['transaction_count'] == 1
        assert result.overall_summary.get('unmatched_account_count') == 2

    def test_unmatched_account_multiple_banks(self):
        """测试同一账号出现在多家银行的情况"""
        records = [
            {
                '唯一id': 'M1', '银行': '银行A', '银行账号': '888888',
                '主体': '', '交易日期': '2024-01-01',
                '付款': None, '收款': 1000.0, '摘要': '',
                '对方户名': '', '余额': 1000.0, '交易流水号': 'M1',
            },
            {
                '唯一id': 'M2', '银行': '银行B', '银行账号': '888888',
                '主体': '', '交易日期': '2024-01-02',
                '付款': None, '收款': 2000.0, '摘要': '',
                '对方户名': '', '余额': 2000.0, '交易流水号': 'M2',
            },
        ]
        result = bankcheck.summarize_transactions(records)

        assert len(result.unmatched_accounts) == 1
        ua = result.unmatched_accounts[0]
        assert ua['account'] == '888888'
        assert '银行A' in ua['banks']
        assert '银行B' in ua['banks']
        assert ua['first_date'] == '2024-01-01'
        assert ua['last_date'] == '2024-01-02'
        assert ua['total_income'] == pytest.approx(3000.0)
        assert ua['transaction_count'] == 2

    def test_empty_account_not_collected(self):
        """测试空银行账号不会被收集为未匹配账号"""
        records = [
            {
                '唯一id': 'E1', '银行': '测试银行', '银行账号': '',
                '主体': '', '交易日期': '2024-01-01',
                '付款': None, '收款': 500.0, '摘要': '',
                '对方户名': '', '余额': 500.0, '交易流水号': 'E1',
            },
        ]
        result = bankcheck.summarize_transactions(records)
        assert result.unmatched_accounts == []

    def test_matched_and_unmatched_mixed(self):
        """测试匹配和未匹配账号混合的情况"""
        records = [
            {
                '唯一id': 'M1', '银行': '测试银行', '银行账号': '111',
                '主体': '匹配主体', '交易日期': '2024-01-01',
                '付款': None, '收款': 1000.0, '摘要': '',
                '对方户名': '', '余额': 1000.0, '交易流水号': 'M1',
            },
            {
                '唯一id': 'U1', '银行': '测试银行', '银行账号': '999',
                '主体': '', '交易日期': '2024-01-02',
                '付款': None, '收款': 500.0, '摘要': '',
                '对方户名': '', '余额': 500.0, '交易流水号': 'U1',
            },
        ]
        result = bankcheck.summarize_transactions(records)

        subjects = [r['subject'] for r in result.by_subject]
        assert '匹配主体' in subjects
        assert '未指定主体' in subjects

        assert len(result.unmatched_accounts) == 1
        assert result.unmatched_accounts[0]['account'] == '999'
        assert result.overall_summary.get('unmatched_account_count') == 1

    def test_export_creates_unmatched_sheet(self, tmp_dir):
        """测试导出时创建未匹配账号汇总 Sheet"""
        records = [
            {
                '唯一id': 'U1', '银行': '测试银行', '银行账号': '999999',
                '主体': '', '交易日期': '2024-01-05',
                '付款': None, '收款': 5000.0, '摘要': '收入',
                '对方户名': '', '余额': 5000.0, '交易流水号': 'U1',
            },
        ]
        summary = bankcheck.summarize_transactions(records)
        output_path = os.path.join(tmp_dir, '未匹配测试.xlsx')

        bankcheck.export_subject_summary(summary, output_path)

        assert os.path.exists(output_path)
        wb = openpyxl.load_workbook(output_path)
        assert '未匹配账号汇总' in wb.sheetnames

        df = pd.read_excel(output_path, sheet_name='未匹配账号汇总')
        assert '银行账号' in df.columns
        assert '涉及银行' in df.columns
        assert '首次交易日期' in df.columns
        assert '最后交易日期' in df.columns
        assert len(df) == 1
        assert str(df.iloc[0]['银行账号']) == '999999'

    def test_export_no_unmatched_sheet_when_none(self, tmp_dir):
        """测试没有未匹配账号时不创建该 Sheet"""
        records = [
            {
                '唯一id': 'M1', '银行': '测试银行', '银行账号': '111',
                '主体': '匹配主体', '交易日期': '2024-01-01',
                '付款': None, '收款': 1000.0, '摘要': '',
                '对方户名': '', '余额': 1000.0, '交易流水号': 'M1',
            },
        ]
        summary = bankcheck.summarize_transactions(records)
        output_path = os.path.join(tmp_dir, '无未匹配测试.xlsx')

        bankcheck.export_subject_summary(summary, output_path)

        wb = openpyxl.load_workbook(output_path)
        assert '未匹配账号汇总' not in wb.sheetnames

    def test_overview_shows_unmatched_count(self, tmp_dir):
        """测试汇总总览中显示未匹配账号数量"""
        records = [
            {
                '唯一id': 'U1', '银行': '测试银行', '银行账号': '111',
                '主体': '', '交易日期': '2024-01-01',
                '付款': None, '收款': 1000.0, '摘要': '',
                '对方户名': '', '余额': 1000.0, '交易流水号': 'U1',
            },
            {
                '唯一id': 'U2', '银行': '测试银行', '银行账号': '222',
                '主体': '', '交易日期': '2024-01-02',
                '付款': None, '收款': 2000.0, '摘要': '',
                '对方户名': '', '余额': 2000.0, '交易流水号': 'U2',
            },
        ]
        summary = bankcheck.summarize_transactions(records)
        output_path = os.path.join(tmp_dir, '总览测试.xlsx')
        bankcheck.export_subject_summary(summary, output_path)

        df = pd.read_excel(output_path, sheet_name='汇总总览')
        data = dict(zip(df['统计项'], df['数值']))
        assert '未匹配账号数' in data
        assert data['未匹配账号数'] == 2

    def test_generate_from_records_includes_unmatched(self, tmp_dir):
        """测试 generate_subject_summary_from_records 包含未匹配账号"""
        records = [
            {
                '唯一id': 'G1', '银行': '测试银行', '银行账号': '777',
                '主体': '', '交易日期': '2024-01-01',
                '付款': None, '收款': 1000.0, '摘要': '',
                '对方户名': '', '余额': 1000.0, '交易流水号': 'G1',
            },
        ]
        result_path = bankcheck.generate_subject_summary_from_records(
            records, tmp_dir, {'来源': '测试'}
        )

        assert result_path is not None
        wb = openpyxl.load_workbook(result_path)
        assert '未匹配账号汇总' in wb.sheetnames

        df = pd.read_excel(result_path, sheet_name='未匹配账号汇总')
        assert len(df) == 1
        assert str(df.iloc[0]['银行账号']) == '777'
