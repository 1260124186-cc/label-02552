"""
总表导出配置模块单元测试
测试列顺序自定义、可选列开关、Excel 表头样式（冻结首行、列宽自适应）
"""
import os
import sys
import tempfile
import shutil

import openpyxl
import pandas as pd
import pytest
import yaml

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='summary_config_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


@pytest.fixture
def sample_records():
    return [
        {
            '唯一id': 'TEST001',
            '银行': '北京银行',
            '银行账号': '01090312345678901',
            '主体': '北京XX科技有限公司',
            '交易日期': '2024-01-05',
            '付款': -50000.0,
            '收款': None,
            '摘要': '采购付款-办公设备',
            '对方户名': '供应商A公司',
            '对方账号': '6222021234567890',
            '余额': 1500000.0,
            '交易流水号': 'BJ20240105001',
            '票据号': 'PJ001',
            '异常标记': '正常',
            '异常详情': '',
        },
        {
            '唯一id': 'TEST002',
            '银行': '东亚银行',
            '银行账号': '38812345678',
            '主体': '上海YY贸易有限公司',
            '交易日期': '2024-01-10',
            '付款': None,
            '收款': 80000.0,
            '摘要': '销售收款-产品销售',
            '对方户名': '客户B公司',
            '对方账号': '6222029876543210',
            '余额': 580000.0,
            '交易流水号': 'EA20240110002',
            '结算号': 'JS002',
            '异常标记': '正常',
            '异常详情': '',
        },
    ]


class TestDefaultConfig:
    """测试默认配置"""

    def test_default_summary_config_exists(self):
        """测试默认配置常量已定义"""
        assert hasattr(bankcheck, 'DEFAULT_SUMMARY_CONFIG')
        assert 'columns' in bankcheck.DEFAULT_SUMMARY_CONFIG
        assert 'excel_style' in bankcheck.DEFAULT_SUMMARY_CONFIG

    def test_default_columns_structure(self):
        """测试默认列配置结构"""
        cols_cfg = bankcheck.DEFAULT_SUMMARY_CONFIG['columns']
        assert 'order' in cols_cfg
        assert 'enabled' in cols_cfg
        assert isinstance(cols_cfg['order'], list)
        assert isinstance(cols_cfg['enabled'], dict)
        assert '唯一id' in cols_cfg['order']
        assert '银行' in cols_cfg['order']

    def test_default_excel_style_structure(self):
        """测试默认 Excel 样式配置结构"""
        style_cfg = bankcheck.DEFAULT_SUMMARY_CONFIG['excel_style']
        assert 'freeze_header' in style_cfg
        assert 'auto_column_width' in style_cfg
        assert style_cfg['freeze_header'] is True
        assert style_cfg['auto_column_width']['enabled'] is True
        assert style_cfg['auto_column_width']['min_width'] == 8
        assert style_cfg['auto_column_width']['max_width'] == 50
        assert style_cfg['auto_column_width']['padding'] == 2

    def test_deep_copy_default_is_independent(self):
        """测试深拷贝默认配置相互独立"""
        cfg1 = bankcheck._deep_copy_default_summary_config()
        cfg2 = bankcheck._deep_copy_default_summary_config()
        cfg1['columns']['order'].append('测试列')
        cfg1['excel_style']['freeze_header'] = False
        assert '测试列' not in cfg2['columns']['order']
        assert cfg2['excel_style']['freeze_header'] is True


class TestConfigMerge:
    """测试配置合并"""

    def test_merge_with_empty_user_config(self):
        """用户空配置时返回默认配置"""
        default = bankcheck._deep_copy_default_summary_config()
        merged = bankcheck._merge_summary_config(default, {})
        assert merged['excel_style']['freeze_header'] is True
        assert merged['columns']['order'][0] == '唯一id'

    def test_merge_partial_user_config(self):
        """用户部分配置正确覆盖默认值"""
        default = bankcheck._deep_copy_default_summary_config()
        user_cfg = {
            'excel_style': {
                'freeze_header': False,
            }
        }
        merged = bankcheck._merge_summary_config(default, user_cfg)
        assert merged['excel_style']['freeze_header'] is False
        assert merged['excel_style']['auto_column_width']['enabled'] is True
        assert merged['excel_style']['auto_column_width']['min_width'] == 8

    def test_merge_nested_config(self):
        """嵌套配置正确合并"""
        default = bankcheck._deep_copy_default_summary_config()
        user_cfg = {
            'excel_style': {
                'auto_column_width': {
                    'min_width': 10,
                    'max_width': 60,
                }
            }
        }
        merged = bankcheck._merge_summary_config(default, user_cfg)
        assert merged['excel_style']['auto_column_width']['min_width'] == 10
        assert merged['excel_style']['auto_column_width']['max_width'] == 60
        assert merged['excel_style']['auto_column_width']['padding'] == 2
        assert merged['excel_style']['auto_column_width']['enabled'] is True

    def test_merge_column_enabled_config(self):
        """列启用配置正确合并"""
        default = bankcheck._deep_copy_default_summary_config()
        user_cfg = {
            'columns': {
                'enabled': {
                    '对方账号': False,
                    '票据号': False,
                }
            }
        }
        merged = bankcheck._merge_summary_config(default, user_cfg)
        assert merged['columns']['enabled']['对方账号'] is False
        assert merged['columns']['enabled']['票据号'] is False
        assert merged['columns']['enabled']['唯一id'] is True


class TestLoadConfig:
    """测试配置文件加载"""

    def test_load_config_without_file(self, tmp_dir):
        """无配置文件时返回默认配置"""
        cfg = bankcheck.load_summary_config(tmp_dir)
        assert cfg['excel_style']['freeze_header'] is True
        assert cfg['columns']['order'][0] == '唯一id'

    def test_load_config_from_file(self, tmp_dir):
        """正确加载 YAML 配置文件"""
        config_path = os.path.join(tmp_dir, bankcheck.SUMMARY_CONFIG_FILENAME)
        user_cfg = {
            'columns': {
                'order': ['交易日期', '摘要', '付款', '收款'],
                'enabled': {'交易日期': True, '摘要': True, '付款': True, '收款': True},
            },
            'excel_style': {
                'freeze_header': False,
                'auto_column_width': {'enabled': False},
            },
        }
        with open(config_path, 'w', encoding='utf-8') as f:
            yaml.dump(user_cfg, f, allow_unicode=True)

        cfg = bankcheck.load_summary_config(tmp_dir)
        assert cfg['excel_style']['freeze_header'] is False
        assert cfg['excel_style']['auto_column_width']['enabled'] is False
        assert cfg['columns']['order'] == ['交易日期', '摘要', '付款', '收款']

    def test_load_config_invalid_yaml(self, tmp_dir):
        """无效 YAML 格式回退到默认配置"""
        config_path = os.path.join(tmp_dir, bankcheck.SUMMARY_CONFIG_FILENAME)
        with open(config_path, 'w', encoding='utf-8') as f:
            f.write('{{invalid yaml content')
        cfg = bankcheck.load_summary_config(tmp_dir)
        assert cfg['excel_style']['freeze_header'] is True

    def test_load_config_non_dict(self, tmp_dir):
        """非字典类型配置回退到默认配置"""
        config_path = os.path.join(tmp_dir, bankcheck.SUMMARY_CONFIG_FILENAME)
        with open(config_path, 'w', encoding='utf-8') as f:
            yaml.dump([1, 2, 3], f)
        cfg = bankcheck.load_summary_config(tmp_dir)
        assert cfg['excel_style']['freeze_header'] is True

    def test_get_summary_config_path(self, tmp_dir):
        """测试配置文件路径"""
        path = bankcheck.get_summary_config_path(tmp_dir)
        expected = os.path.join(tmp_dir, bankcheck.SUMMARY_CONFIG_FILENAME)
        assert path == expected


class TestApplyColumnConfig:
    """测试列配置应用（排序和启用开关）"""

    def test_empty_columns(self):
        """空列列表返回空"""
        result = bankcheck.apply_column_config([], {'order': ['A', 'B'], 'enabled': {'A': True}})
        assert result == []

    def test_column_order_customized(self):
        """列顺序按照配置重排"""
        columns = ['唯一id', '银行', '交易日期', '摘要']
        config = {
            'order': ['交易日期', '摘要', '银行', '唯一id'],
            'enabled': {'唯一id': True, '银行': True, '交易日期': True, '摘要': True},
        }
        result = bankcheck.apply_column_config(columns, config)
        assert result == ['交易日期', '摘要', '银行', '唯一id']

    def test_column_disabled_hidden(self):
        """禁用的列被过滤掉"""
        columns = ['唯一id', '银行', '银行账号', '对方账号', '交易日期']
        config = {
            'order': ['唯一id', '银行', '银行账号', '对方账号', '交易日期'],
            'enabled': {
                '唯一id': True, '银行': True, '银行账号': True,
                '对方账号': False, '交易日期': True,
            },
        }
        result = bankcheck.apply_column_config(columns, config)
        assert '对方账号' not in result
        assert result == ['唯一id', '银行', '银行账号', '交易日期']

    def test_partial_config_order(self):
        """配置中部分列排序，剩余列按原顺序追加"""
        columns = ['唯一id', '银行', '交易日期', '摘要', '付款', '收款']
        config = {
            'order': ['交易日期', '摘要'],
            'enabled': {
                '唯一id': True, '银行': True, '交易日期': True,
                '摘要': True, '付款': True, '收款': True,
            },
        }
        result = bankcheck.apply_column_config(columns, config)
        assert result[0] == '交易日期'
        assert result[1] == '摘要'
        assert '唯一id' in result[2:]
        assert '银行' in result[2:]

    def test_extra_columns_not_in_config(self):
        """数据中存在但配置中不存在的列按原顺序追加，受 enabled 控制"""
        columns = ['唯一id', '银行', '自定义字段1', '自定义字段2']
        config = {
            'order': ['唯一id', '银行'],
            'enabled': {
                '唯一id': True, '银行': True,
                '自定义字段1': True, '自定义字段2': False,
            },
        }
        result = bankcheck.apply_column_config(columns, config)
        assert '自定义字段1' in result
        assert '自定义字段2' not in result

    def test_get_summary_columns_with_config(self, sample_records):
        """get_summary_columns 应用列配置"""
        custom_config = {
            'columns': {
                'order': ['交易日期', '摘要', '付款', '收款', '银行', '主体'],
                'enabled': {
                    '交易日期': True, '摘要': True, '付款': True,
                    '收款': True, '银行': True, '主体': True,
                    '唯一id': False, '银行账号': False,
                },
            },
            'excel_style': {'freeze_header': True, 'auto_column_width': {'enabled': True}},
        }
        columns = bankcheck.get_summary_columns(sample_records, config=custom_config)
        assert columns[0] == '交易日期'
        assert columns[1] == '摘要'
        assert '唯一id' not in columns
        assert '银行账号' not in columns

    def test_get_summary_columns_without_config(self, sample_records):
        """get_summary_columns 无配置时保持默认行为"""
        columns = bankcheck.get_summary_columns(sample_records)
        assert '唯一id' in columns
        assert '银行' in columns


class TestDisplayWidth:
    """测试显示宽度计算"""

    def test_none_value(self):
        assert bankcheck._get_display_width(None) == 0

    def test_ascii_text(self):
        assert bankcheck._get_display_width('hello') == 5

    def test_chinese_text(self):
        assert bankcheck._get_display_width('你好') == 4

    def test_mixed_text(self):
        assert bankcheck._get_display_width('你好world') == 4 + 5

    def test_numeric_value(self):
        assert bankcheck._get_display_width(12345) == 5

    def test_float_value(self):
        assert bankcheck._get_display_width(123.45) == 6


class TestApplyExcelStyle:
    """测试 Excel 样式应用"""

    def _create_test_workbook(self, tmp_dir):
        """创建测试用 Excel 文件"""
        path = os.path.join(tmp_dir, 'test_style.xlsx')
        df = pd.DataFrame({
            '唯一id': ['TEST001', 'TEST002'],
            '银行': ['北京银行', '东亚银行'],
            '摘要': ['采购付款-办公设备采购', '销售收款'],
            '付款': [-50000.0, None],
            '收款': [None, 80000.0],
        })
        df.to_excel(path, index=False, engine='openpyxl')
        return path

    def test_freeze_header_enabled(self, tmp_dir):
        """冻结首行配置生效"""
        path = self._create_test_workbook(tmp_dir)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        config = {'freeze_header': True, 'auto_column_width': {'enabled': False}}
        bankcheck.apply_excel_style(ws, config)
        wb.save(path)
        wb.close()

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        assert ws2.freeze_panes is not None
        assert str(ws2.freeze_panes) == 'A2'
        wb2.close()

    def test_freeze_header_disabled(self, tmp_dir):
        """禁用冻结首行时不设置"""
        path = self._create_test_workbook(tmp_dir)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        config = {'freeze_header': False, 'auto_column_width': {'enabled': False}}
        bankcheck.apply_excel_style(ws, config)
        wb.save(path)
        wb.close()

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        assert ws2.freeze_panes is None
        wb2.close()

    def test_auto_column_width_enabled(self, tmp_dir):
        """列宽自适应配置生效"""
        path = self._create_test_workbook(tmp_dir)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        before_widths = {}
        for col in ['A', 'B', 'C', 'D', 'E']:
            before_widths[col] = ws.column_dimensions[col].width

        config = {
            'freeze_header': False,
            'auto_column_width': {
                'enabled': True, 'min_width': 8, 'max_width': 50, 'padding': 2,
            },
        }
        bankcheck.apply_excel_style(ws, config)
        wb.save(path)
        wb.close()

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        c_width = ws2.column_dimensions['C'].width
        assert c_width > before_widths.get('C', 0) or before_widths.get('C') is None
        assert c_width <= 50
        assert c_width >= 8
        wb2.close()

    def test_auto_column_width_disabled(self, tmp_dir):
        """禁用列宽自适应时不修改列宽"""
        path = self._create_test_workbook(tmp_dir)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        before_widths = {}
        for col in ['A', 'B', 'C', 'D', 'E']:
            before_widths[col] = ws.column_dimensions[col].width

        config = {
            'freeze_header': False,
            'auto_column_width': {'enabled': False},
        }
        bankcheck.apply_excel_style(ws, config)
        wb.save(path)
        wb.close()

        wb2 = openpyxl.load_workbook(path)
        ws2 = wb2.active
        for col in ['A', 'B', 'C', 'D', 'E']:
            assert ws2.column_dimensions[col].width == before_widths[col]
        wb2.close()

    def test_apply_style_on_empty_worksheet(self):
        """空工作表不报错"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.delete_rows(1, ws.max_row)
        config = {'freeze_header': True, 'auto_column_width': {'enabled': True}}
        bankcheck.apply_excel_style(ws, config)
        wb.close()

    def test_apply_style_on_none_worksheet(self):
        """None 工作表不报错"""
        bankcheck.apply_excel_style(None, {'freeze_header': True})


class TestExportWithConfig:
    """测试总表导出时应用配置"""

    def test_merge_and_export_summary_with_custom_config(self, tmp_dir, sample_records):
        """merge_and_export_summary 应用自定义配置"""
        custom_config = {
            'columns': {
                'order': ['交易日期', '摘要', '付款', '收款', '银行'],
                'enabled': {
                    '交易日期': True, '摘要': True, '付款': True,
                    '收款': True, '银行': True,
                    '唯一id': False, '银行账号': False,
                    '主体': False, '对方户名': False, '对方账号': False,
                    '余额': False, '交易流水号': False,
                    '票据号': False, '结算号': False, '凭证号': False,
                    '异常标记': False, '异常详情': False,
                },
            },
            'excel_style': {
                'freeze_header': True,
                'auto_column_width': {
                    'enabled': True, 'min_width': 8, 'max_width': 30, 'padding': 2,
                },
            },
        }

        output_path = bankcheck.merge_and_export_summary(
            existing_records=[],
            incremental_rows=sample_records,
            script_dir=tmp_dir,
            output_dir=tmp_dir,
            config=custom_config,
        )

        assert output_path is not None
        assert os.path.exists(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        headers = [cell.value for cell in ws[1]]
        assert headers[0] == '交易日期'
        assert headers[1] == '摘要'
        assert '唯一id' not in headers
        assert '银行账号' not in headers
        assert len(headers) == 5

        assert ws.freeze_panes is not None
        assert str(ws.freeze_panes) == 'A2'

        for col_letter in ['A', 'B', 'C', 'D', 'E']:
            assert ws.column_dimensions[col_letter].width is not None
            assert ws.column_dimensions[col_letter].width >= 8
            assert ws.column_dimensions[col_letter].width <= 30

        wb.close()

    def test_merge_and_export_summary_auto_load_config(self, tmp_dir, sample_records):
        """merge_and_export_summary 自动从配置文件加载"""
        config_path = os.path.join(tmp_dir, bankcheck.SUMMARY_CONFIG_FILENAME)
        user_cfg = {
            'excel_style': {
                'freeze_header': False,
                'auto_column_width': {'enabled': False},
            },
        }
        with open(config_path, 'w', encoding='utf-8') as f:
            yaml.dump(user_cfg, f, allow_unicode=True)

        output_path = bankcheck.merge_and_export_summary(
            existing_records=[],
            incremental_rows=sample_records,
            script_dir=tmp_dir,
            output_dir=tmp_dir,
        )

        assert output_path is not None
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        assert ws.freeze_panes is None
        wb.close()

    def test_export_masked_summary_with_config(self, tmp_dir, sample_records):
        """export_masked_summary 应用配置"""
        custom_config = {
            'columns': {
                'order': ['交易日期', '摘要', '银行'],
                'enabled': {
                    '交易日期': True, '摘要': True, '银行': True,
                    '唯一id': False, '银行账号': False, '主体': False,
                    '付款': False, '收款': False, '对方户名': False,
                    '对方账号': False, '余额': False, '交易流水号': False,
                    '票据号': False, '结算号': False, '凭证号': False,
                    '异常标记': False, '异常详情': False,
                },
            },
            'excel_style': {
                'freeze_header': True,
                'auto_column_width': {'enabled': True, 'min_width': 10, 'max_width': 40, 'padding': 2},
            },
        }

        output_path = bankcheck.export_masked_summary(
            records=sample_records,
            script_dir=tmp_dir,
            output_dir=tmp_dir,
            config=custom_config,
        )

        assert output_path is not None
        assert os.path.exists(output_path)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]
        assert headers == ['交易日期', '摘要', '银行']
        assert ws.freeze_panes is not None
        wb.close()

    def test_df_to_excel_text_safe_with_style(self, tmp_dir):
        """df_to_excel_text_safe 应用样式"""
        df = pd.DataFrame({
            '列A': ['数据1很长的内容用于测试宽度', '短'],
            '列B': [100, 200],
        })
        output_path = os.path.join(tmp_dir, 'styled.xlsx')
        style_config = {
            'freeze_header': True,
            'auto_column_width': {
                'enabled': True, 'min_width': 8, 'max_width': 50, 'padding': 2,
            },
        }
        bankcheck.df_to_excel_text_safe(
            df, output_path, index=False, excel_style_config=style_config,
        )

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        assert ws.freeze_panes is not None
        a_width = ws.column_dimensions['A'].width
        assert a_width > 10
        assert a_width <= 50
        wb.close()

    def test_df_to_excel_text_safe_without_style(self, tmp_dir):
        """df_to_excel_text_safe 无样式配置时不应用"""
        df = pd.DataFrame({'列A': ['测试数据'], '列B': [100]})
        output_path = os.path.join(tmp_dir, 'no_style.xlsx')
        bankcheck.df_to_excel_text_safe(df, output_path, index=False)

        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        assert ws.freeze_panes is None
        wb.close()
