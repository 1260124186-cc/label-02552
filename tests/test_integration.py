import os
import sys
import shutil
import json

import openpyxl
import pandas as pd
import pytest

from conftest import _create_beijing_bank_excel, _create_east_asia_bank_excel, _create_lookup_table
import bankcheck


class TestRunPipeline:
    def _setup_folder(self, tmp_dir, script_dir, files=None):
        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)

        if files is None:
            files = ['北京银行', '东亚银行']

        for bank in files:
            if bank == '北京银行':
                _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))
            elif bank == '东亚银行':
                _create_east_asia_bank_excel(os.path.join(source_folder, '东亚银行_流水.xlsx'))
            elif bank == '未知':
                wb = openpyxl.Workbook()
                ws = wb.active
                ws['A1'] = '未知银行数据'
                wb.save(os.path.join(source_folder, '未知银行_流水.xlsx'))
                wb.close()

        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))
        return source_folder

    def test_pipeline_beijing(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir, ['北京银行'])

        result = bankcheck.run_pipeline(source, script_dir)

        assert len(result.all_rows) == 2
        assert result.all_rows[0]['银行'] == '北京银行'
        assert result.all_rows[0]['主体'] == '北京XX科技有限公司'
        assert result.all_rows[0]['付款'] == -50000.0
        assert len(result.processed_files) == 1
        assert not result.folder_empty
        assert result.output_path is not None

    def test_pipeline_east_asia(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir, ['东亚银行'])

        result = bankcheck.run_pipeline(source, script_dir)

        assert len(result.all_rows) == 2
        assert result.all_rows[0]['银行'] == '东亚银行'
        assert result.all_rows[0]['主体'] == '上海YY贸易有限公司'

    def test_pipeline_both_banks(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir, ['北京银行', '东亚银行'])

        result = bankcheck.run_pipeline(source, script_dir)

        assert len(result.all_rows) == 4
        beijing_rows = [r for r in result.all_rows if r['银行'] == '北京银行']
        east_asia_rows = [r for r in result.all_rows if r['银行'] == '东亚银行']
        assert len(beijing_rows) == 2
        assert len(east_asia_rows) == 2

    def test_output_table(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir, ['北京银行', '东亚银行'])

        result = bankcheck.run_pipeline(source, script_dir)

        assert result.output_path is not None
        assert os.path.exists(result.output_path)
        df_read = pd.read_excel(result.output_path, engine='openpyxl')
        assert len(df_read) == 4
        columns = ['唯一id', '银行', '银行账号', '主体', '交易日期', '付款', '收款', '摘要', '对方户名', '余额', '交易流水号']
        assert list(df_read.columns) == columns

    def test_unprocessed_files_kept(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir, ['北京银行', '未知'])

        result = bankcheck.run_pipeline(source, script_dir)

        assert len(result.unprocessed_files) == 1
        assert '未知银行' in os.path.basename(result.unprocessed_files[0])

    def test_processed_files_deleted(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir, ['北京银行', '未知'])

        result = bankcheck.run_pipeline(source, script_dir)

        new_folder = source + '＋检验版'
        remaining = bankcheck.scan_excel_files(new_folder)
        assert len(remaining) == 1
        assert '未知银行' in os.path.basename(remaining[0])

    def test_original_folder_untouched(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir, ['北京银行'])

        original_count = len(os.listdir(source))
        bankcheck.run_pipeline(source, script_dir)

        assert len(os.listdir(source)) == original_count
        assert os.path.exists(os.path.join(source, '北京银行_流水.xlsx'))

    def test_empty_folder(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = os.path.join(tmp_dir, '空文件夹')
        os.makedirs(source, exist_ok=True)
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        result = bankcheck.run_pipeline(source, script_dir)

        assert result.folder_empty is True
        assert len(result.all_rows) == 0
        assert result.output_path is None

    def test_lookup_missing(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir, ['北京银行'])
        os.remove(os.path.join(script_dir, '主体查找表.xlsx'))

        result = bankcheck.run_pipeline(source, script_dir)

        assert result.lookup_missing is True
        assert result.all_rows[0]['主体'] == ''

    def test_error_files_kept_after_pipeline(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)

        corrupt_path = os.path.join(source_folder, '北京银行_坏.xlsx')
        with open(corrupt_path, 'wb') as f:
            f.write(b'not a valid excel file')

        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        result = bankcheck.run_pipeline(source_folder, script_dir)

        assert len(result.error_files) == 1
        error_filepath = result.error_files[0][0]
        assert '北京银行_坏' in os.path.basename(error_filepath)

        new_folder = source_folder + '＋检验版'
        remaining = bankcheck.scan_excel_files(new_folder)
        assert len(remaining) == 1
        assert '北京银行_坏' in os.path.basename(remaining[0])

    def test_mixed_success_error_unprocessed(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)

        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_正常.xlsx'))
        corrupt_path = os.path.join(source_folder, '北京银行_坏.xlsx')
        with open(corrupt_path, 'wb') as f:
            f.write(b'corrupt')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'unknown'
        wb.save(os.path.join(source_folder, '未知银行_流水.xlsx'))
        wb.close()

        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        result = bankcheck.run_pipeline(source_folder, script_dir)

        assert len(result.processed_files) == 1
        assert len(result.error_files) == 1
        assert len(result.unprocessed_files) == 1

        new_folder = source_folder + '＋检验版'
        remaining = bankcheck.scan_excel_files(new_folder)
        remaining_names = [os.path.basename(f) for f in remaining]
        assert '北京银行_坏.xlsx' in remaining_names
        assert '未知银行_流水.xlsx' in remaining_names
        assert '北京银行_正常.xlsx' not in remaining_names

    def test_bank_processors_registry(self):
        assert '北京银行' in bankcheck.BANK_PROCESSORS
        assert '东亚银行' in bankcheck.BANK_PROCESSORS
        assert bankcheck.BANK_PROCESSORS['北京银行'] == bankcheck.process_beijing_bank
        assert bankcheck.BANK_PROCESSORS['东亚银行'] == bankcheck.process_east_asia_bank

    def test_sample_files(self):
        samples_dir = os.path.join(os.path.dirname(__file__), '..', 'samples')
        if not os.path.isdir(samples_dir):
            pytest.skip('samples directory not found')

        script_dir = os.path.join(os.path.dirname(__file__), '..', 'backend')
        lookup_file = bankcheck.find_lookup_file(script_dir)
        if not lookup_file:
            pytest.skip('主体查找表 not found in backend directory')

        beijing_sample = os.path.join(samples_dir, '北京银行_示例流水.xlsx')
        east_asia_sample = os.path.join(samples_dir, '东亚银行_示例流水.xlsx')

        if os.path.exists(beijing_sample):
            rows = bankcheck.process_beijing_bank(beijing_sample, lookup_file)
            assert len(rows) > 0
            assert rows[0]['银行'] == '北京银行'

        if os.path.exists(east_asia_sample):
            rows = bankcheck.process_east_asia_bank(east_asia_sample, lookup_file)
            assert len(rows) > 0
            assert rows[0]['银行'] == '东亚银行'


class TestFormatResultMessage:
    def test_folder_empty(self):
        result = bankcheck.ProcessingResult(folder_empty=True)
        msg = bankcheck.format_result_message(result)
        assert msg == '文件夹中未发现任何 Excel 文件。'

    def test_with_records(self):
        result = bankcheck.ProcessingResult(
            all_rows=[{'银行': '北京银行'}, {'银行': '东亚银行'}],
            processed_files=['/path/a.xlsx', '/path/b.xlsx'],
            output_path='/path/to/银行流水总表.xlsx',
        )
        msg = bankcheck.format_result_message(result)
        assert '处理完成！' in msg
        assert '已处理文件数：2' in msg
        assert '提取记录数：2' in msg
        assert '/path/to/银行流水总表.xlsx' in msg

    def test_no_records(self):
        result = bankcheck.ProcessingResult()
        msg = bankcheck.format_result_message(result)
        assert '未提取到任何银行流水记录。' in msg

    def test_unprocessed_files(self):
        result = bankcheck.ProcessingResult(
            all_rows=[{'银行': '北京银行'}],
            processed_files=['/path/a.xlsx'],
            unprocessed_files=['/path/未知银行_流水.xlsx'],
            output_path='/path/to/银行流水总表.xlsx',
        )
        msg = bankcheck.format_result_message(result)
        assert '无法识别的文件（1 个' in msg
        assert '未知银行_流水.xlsx' in msg

    def test_error_files(self):
        result = bankcheck.ProcessingResult(
            error_files=[('/path/北京银行_坏.xlsx', 'Bad file')],
        )
        msg = bankcheck.format_result_message(result)
        assert '处理出错的文件（1 个' in msg
        assert '北京银行_坏.xlsx' in msg
        assert 'Bad file' in msg

    def test_error_files_message_says_preserved(self):
        result = bankcheck.ProcessingResult(
            error_files=[('/path/北京银行_坏.xlsx', 'File is corrupt')],
        )
        msg = bankcheck.format_result_message(result)
        assert '处理出错的文件' in msg
        assert '已保留' in msg
        assert '北京银行_坏.xlsx' in msg

    def test_unprocessed_and_error(self):
        result = bankcheck.ProcessingResult(
            unprocessed_files=['/path/未知.xlsx'],
            error_files=[('/path/北京银行_坏.xlsx', 'Error')],
        )
        msg = bankcheck.format_result_message(result)
        assert '无法识别的文件' in msg
        assert '处理出错的文件' in msg


class TestDeleteProcessedFiles:

    def _create_files(self, tmp_dir):
        f_success = os.path.join(tmp_dir, '北京银行_正常.xlsx')
        f_error = os.path.join(tmp_dir, '北京银行_坏.xlsx')
        f_unprocessed = os.path.join(tmp_dir, '未知.xlsx')
        for f in [f_success, f_error, f_unprocessed]:
            with open(f, 'w') as fp:
                fp.write('test')
        return f_success, f_error, f_unprocessed

    def test_strategy_keep_unprocessed_deletes_only_success_files(self, tmp_dir):
        f_success, f_error, f_unprocessed = self._create_files(tmp_dir)
        excel_files = [f_success, f_error, f_unprocessed]
        processed_files = [f_success]
        error_files = [(f_error, 'parse error')]
        unprocessed_files = [f_unprocessed]

        bankcheck.delete_processed_files(
            excel_files, processed_files, error_files, unprocessed_files,
            strategy='keep_unprocessed'
        )

        assert not os.path.exists(f_success)
        assert os.path.exists(f_error)
        assert os.path.exists(f_unprocessed)

    def test_strategy_keep_unprocessed_is_default(self, tmp_dir):
        f_success, f_error, f_unprocessed = self._create_files(tmp_dir)
        excel_files = [f_success, f_error, f_unprocessed]
        processed_files = [f_success]
        error_files = [(f_error, 'parse error')]
        unprocessed_files = [f_unprocessed]

        bankcheck.delete_processed_files(
            excel_files, processed_files, error_files, unprocessed_files
        )

        assert not os.path.exists(f_success)
        assert os.path.exists(f_error)
        assert os.path.exists(f_unprocessed)

    def test_strategy_keep_all_preserves_all_files(self, tmp_dir):
        f_success, f_error, f_unprocessed = self._create_files(tmp_dir)
        excel_files = [f_success, f_error, f_unprocessed]
        processed_files = [f_success]
        error_files = [(f_error, 'parse error')]
        unprocessed_files = [f_unprocessed]

        bankcheck.delete_processed_files(
            excel_files, processed_files, error_files, unprocessed_files,
            strategy='keep_all'
        )

        assert os.path.exists(f_success)
        assert os.path.exists(f_error)
        assert os.path.exists(f_unprocessed)

    def test_strategy_delete_all_removes_everything(self, tmp_dir):
        f_success, f_error, f_unprocessed = self._create_files(tmp_dir)
        excel_files = [f_success, f_error, f_unprocessed]
        processed_files = [f_success]
        error_files = [(f_error, 'parse error')]
        unprocessed_files = [f_unprocessed]

        bankcheck.delete_processed_files(
            excel_files, processed_files, error_files, unprocessed_files,
            strategy='delete_all'
        )

        assert not os.path.exists(f_success)
        assert not os.path.exists(f_error)
        assert not os.path.exists(f_unprocessed)

    def test_strategy_move_to_archive_moves_success_files(self, tmp_dir):
        f_success, f_error, f_unprocessed = self._create_files(tmp_dir)
        excel_files = [f_success, f_error, f_unprocessed]
        processed_files = [f_success]
        error_files = [(f_error, 'parse error')]
        unprocessed_files = [f_unprocessed]

        bankcheck.delete_processed_files(
            excel_files, processed_files, error_files, unprocessed_files,
            strategy='move_to_archive'
        )

        assert not os.path.exists(f_success)
        assert os.path.exists(f_error)
        assert os.path.exists(f_unprocessed)

        archive_dir = os.path.join(tmp_dir, '已处理归档')
        assert os.path.isdir(archive_dir)
        archived_files = os.listdir(archive_dir)
        assert len(archived_files) == 1
        assert '北京银行_正常.xlsx' in archived_files
        assert os.path.exists(os.path.join(archive_dir, '北京银行_正常.xlsx'))

    def test_strategy_move_to_archive_custom_dir_name(self, tmp_dir):
        f_success, f_error, f_unprocessed = self._create_files(tmp_dir)
        excel_files = [f_success, f_error, f_unprocessed]
        processed_files = [f_success]
        error_files = [(f_error, 'parse error')]
        unprocessed_files = [f_unprocessed]

        bankcheck.delete_processed_files(
            excel_files, processed_files, error_files, unprocessed_files,
            strategy='move_to_archive',
            archive_dir_name='processed_backup'
        )

        archive_dir = os.path.join(tmp_dir, 'processed_backup')
        assert os.path.isdir(archive_dir)
        assert len(os.listdir(archive_dir)) == 1

    def test_strategy_move_to_archive_handles_name_conflicts(self, tmp_dir):
        f_success1 = os.path.join(tmp_dir, '北京银行_正常.xlsx')
        f_success2 = os.path.join(tmp_dir, '北京银行_正常_dup.xlsx')
        for f in [f_success1, f_success2]:
            with open(f, 'w') as fp:
                fp.write('test')

        archive_dir = os.path.join(tmp_dir, '已处理归档')
        os.makedirs(archive_dir, exist_ok=True)
        with open(os.path.join(archive_dir, '北京银行_正常.xlsx'), 'w') as fp:
            fp.write('existing')

        excel_files = [f_success1, f_success2]
        processed_files = [f_success1, f_success2]
        error_files = []
        unprocessed_files = []

        bankcheck.delete_processed_files(
            excel_files, processed_files, error_files, unprocessed_files,
            strategy='move_to_archive'
        )

        archived = sorted(os.listdir(archive_dir))
        assert len(archived) >= 3
        assert '北京银行_正常.xlsx' in archived
        assert '北京银行_正常_1.xlsx' in archived

    def test_strategy_move_to_archive_no_processed_files(self, tmp_dir):
        f_error = os.path.join(tmp_dir, '北京银行_坏.xlsx')
        f_unprocessed = os.path.join(tmp_dir, '未知.xlsx')
        for f in [f_error, f_unprocessed]:
            with open(f, 'w') as fp:
                fp.write('test')

        excel_files = [f_error, f_unprocessed]
        processed_files = []
        error_files = [(f_error, 'parse error')]
        unprocessed_files = [f_unprocessed]

        bankcheck.delete_processed_files(
            excel_files, processed_files, error_files, unprocessed_files,
            strategy='move_to_archive'
        )

        assert os.path.exists(f_error)
        assert os.path.exists(f_unprocessed)

    def test_unknown_strategy_falls_back_to_keep_unprocessed(self, tmp_dir):
        f_success, f_error, f_unprocessed = self._create_files(tmp_dir)
        excel_files = [f_success, f_error, f_unprocessed]
        processed_files = [f_success]
        error_files = [(f_error, 'parse error')]
        unprocessed_files = [f_unprocessed]

        bankcheck.delete_processed_files(
            excel_files, processed_files, error_files, unprocessed_files,
            strategy='invalid_strategy_xyz'
        )

        assert not os.path.exists(f_success)
        assert os.path.exists(f_error)
        assert os.path.exists(f_unprocessed)

    def test_empty_file_list_all_strategies(self, tmp_dir):
        for strategy in ['keep_all', 'keep_unprocessed', 'delete_all', 'move_to_archive']:
            bankcheck.delete_processed_files(
                [], [], [], [], strategy=strategy
            )

    def test_nonexistent_file_handled_gracefully(self, tmp_dir):
        nonexistent = os.path.join(tmp_dir, 'missing.xlsx')
        for strategy in ['keep_unprocessed', 'delete_all']:
            bankcheck.delete_processed_files(
                [nonexistent], [nonexistent], [], [], strategy=strategy
            )

    def test_strategy_keep_all_with_empty_processed(self, tmp_dir):
        f_error, f_unprocessed = (
            os.path.join(tmp_dir, '坏.xlsx'),
            os.path.join(tmp_dir, '未知.xlsx'),
        )
        for f in [f_error, f_unprocessed]:
            with open(f, 'w') as fp:
                fp.write('test')

        bankcheck.delete_processed_files(
            [f_error, f_unprocessed], [], [(f_error, 'err')], [f_unprocessed],
            strategy='keep_all'
        )

        assert os.path.exists(f_error)
        assert os.path.exists(f_unprocessed)

    def test_strategy_move_to_archive_preserves_error_and_unprocessed(self, tmp_dir):
        f_success, f_error, f_unprocessed = self._create_files(tmp_dir)
        excel_files = [f_success, f_error, f_unprocessed]
        processed_files = [f_success]
        error_files = [(f_error, 'parse error')]
        unprocessed_files = [f_unprocessed]

        bankcheck.delete_processed_files(
            excel_files, processed_files, error_files, unprocessed_files,
            strategy='move_to_archive'
        )

        assert os.path.exists(f_error), 'error file should remain in original location'
        assert os.path.exists(f_unprocessed), 'unprocessed file should remain in original location'
        archive_dir = os.path.join(tmp_dir, '已处理归档')
        archived = os.listdir(archive_dir)
        assert os.path.basename(f_error) not in archived
        assert os.path.basename(f_unprocessed) not in archived

    def test_multiple_success_files_all_moved_to_archive(self, tmp_dir):
        files = []
        for i in range(3):
            f = os.path.join(tmp_dir, f'银行{i}_流水.xlsx')
            with open(f, 'w') as fp:
                fp.write(f'data{i}')
            files.append(f)

        bankcheck.delete_processed_files(
            files, files, [], [], strategy='move_to_archive'
        )

        archive_dir = os.path.join(tmp_dir, '已处理归档')
        archived = os.listdir(archive_dir)
        assert len(archived) == 3
        for f in files:
            assert not os.path.exists(f)
            assert os.path.basename(f) in archived


class TestKeepStrategiesConfig:

    def test_keep_strategies_dict_has_expected_keys(self):
        expected_keys = {'keep_unprocessed', 'keep_all', 'delete_all', 'move_to_archive'}
        assert set(bankcheck.KEEP_STRATEGIES.keys()) == expected_keys

    def test_keep_strategies_descriptions_are_non_empty(self):
        for key, desc in bankcheck.KEEP_STRATEGIES.items():
            assert isinstance(desc, str)
            assert len(desc.strip()) > 0, f'Description for {key} should not be empty'

    def test_default_keep_strategy_is_keep_unprocessed(self):
        import inspect
        sig = inspect.signature(bankcheck.run_pipeline)
        assert sig.parameters['keep_strategy'].default == 'keep_unprocessed'

        sig2 = inspect.signature(bankcheck.delete_processed_files)
        assert sig2.parameters['strategy'].default == 'keep_unprocessed'

        sig3 = inspect.signature(bankcheck.run_pipeline_with_options)
        assert sig3.parameters['keep_strategy'].default == 'keep_unprocessed'


class TestCliAskKeepStrategy:

    def test_default_choice_is_keep_unprocessed(self, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: '')
        result = bankcheck.cli_ask_keep_strategy()
        assert result == 'keep_unprocessed'

    def test_choice_1_keep_unprocessed(self, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: '1')
        result = bankcheck.cli_ask_keep_strategy()
        assert result == 'keep_unprocessed'

    def test_choice_2_keep_all(self, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: '2')
        result = bankcheck.cli_ask_keep_strategy()
        assert result == 'keep_all'

    def test_choice_3_delete_all(self, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: '3')
        result = bankcheck.cli_ask_keep_strategy()
        assert result == 'delete_all'

    def test_choice_4_move_to_archive(self, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: '4')
        result = bankcheck.cli_ask_keep_strategy()
        assert result == 'move_to_archive'

    def test_invalid_choice_falls_back_to_default(self, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: '999')
        result = bankcheck.cli_ask_keep_strategy()
        assert result == 'keep_unprocessed'

    def test_non_numeric_choice_falls_back_to_default(self, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: 'abc')
        result = bankcheck.cli_ask_keep_strategy()
        assert result == 'keep_unprocessed'


class TestRunPipelineKeepStrategyIntegration:

    def _setup_folder(self, tmp_dir, script_dir):
        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)

        _create_beijing_bank_excel = bankcheck.__dict__.get('_create_beijing_bank_excel')
        if _create_beijing_bank_excel is None:
            from conftest import _create_beijing_bank_excel

        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '未知银行数据'
        wb.save(os.path.join(source_folder, '未知银行_流水.xlsx'))
        wb.close()

        from conftest import _create_lookup_table
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))
        return source_folder

    def test_pipeline_keep_all_strategy_preserves_all_files(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        result = bankcheck.run_pipeline(source, script_dir, keep_strategy='keep_all')

        assert len(result.processed_files) == 1
        assert len(result.unprocessed_files) == 1

        new_folder = source + '＋检验版'
        remaining = bankcheck.scan_excel_files(new_folder)
        assert len(remaining) == 2
        names = [os.path.basename(f) for f in remaining]
        assert '北京银行_流水.xlsx' in names
        assert '未知银行_流水.xlsx' in names

    def test_pipeline_move_to_archive_strategy_moves_success_files(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        result = bankcheck.run_pipeline(source, script_dir, keep_strategy='move_to_archive')

        assert len(result.processed_files) == 1
        assert len(result.unprocessed_files) == 1

        new_folder = source + '＋检验版'
        success_path = os.path.join(new_folder, '北京银行_流水.xlsx')
        unprocessed_path = os.path.join(new_folder, '未知银行_流水.xlsx')
        assert not os.path.exists(success_path), '成功文件应从原位置移除'
        assert os.path.exists(unprocessed_path), '未处理文件应保留在原位置'

        archive_dir = os.path.join(new_folder, '已处理归档')
        assert os.path.isdir(archive_dir)
        archived = os.listdir(archive_dir)
        assert '北京银行_流水.xlsx' in archived
        assert os.path.exists(os.path.join(archive_dir, '北京银行_流水.xlsx'))

    def test_pipeline_delete_all_strategy_removes_everything(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        result = bankcheck.run_pipeline(source, script_dir, keep_strategy='delete_all')

        assert len(result.processed_files) == 1
        assert len(result.unprocessed_files) == 1

        new_folder = source + '＋检验版'
        remaining = bankcheck.scan_excel_files(new_folder)
        assert len(remaining) == 0

    def test_pipeline_default_keep_unprocessed_deletes_only_success(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        result = bankcheck.run_pipeline(source, script_dir)

        new_folder = source + '＋检验版'
        remaining = bankcheck.scan_excel_files(new_folder)
        remaining_names = [os.path.basename(f) for f in remaining]
        assert len(remaining) == 1
        assert '未知银行_流水.xlsx' in remaining_names
        assert '北京银行_流水.xlsx' not in remaining_names


class TestInterestFeeCheckIntegration:
    """利息与手续费核对集成测试"""

    def _create_beijing_bank_with_interest_fee(self, path):
        """创建包含利息和手续费交易的北京银行Excel"""
        rows = [
            [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A公司', '622001234', '工商银行', '转账', '001', '采购付款', None, None, None, 'BJ20240105001'],
            [2, '2024-01-10', 'CNY', None, 80000, 1580000, '客户B公司', '622005678', '建设银行', '转账', '002', '销售收款', None, None, None, 'BJ20240110002'],
            [3, '2024-01-15', 'CNY', None, 125.50, 1580125.50, '北京银行', '622009999', '北京银行', '结息', '003', '2024年第一季度存款利息', None, None, None, 'BJ20240115003'],
            [4, '2024-01-20', 'CNY', 50, None, 1580075.50, '北京银行', '622009999', '北京银行', '手续费', '004', '跨行转账手续费', None, None, None, 'BJ20240120004'],
            [5, '2024-02-05', 'CNY', 10, None, 1580065.50, '北京银行', '622009999', '北京银行', '管理费', '005', '月度账户管理费', None, None, None, 'BJ20240205005'],
            [6, '2024-02-10', 'CNY', 30, None, 1580035.50, '北京银行', '622009999', '北京银行', '手续费', '006', '网银服务费', None, None, None, 'BJ20240210006'],
            [7, '2024-03-15', 'CNY', None, 150.00, 1580185.50, '北京银行', '622009999', '北京银行', '结息', '007', '2024年第二季度存款利息', None, None, None, 'BJ20240315007'],
            [8, '2024-03-20', 'CNY', 100, None, 1580085.50, '北京银行', '622009999', '北京银行', '手续费', '008', '大额转账手续费', None, None, None, 'BJ20240320008'],
        ]
        _create_beijing_bank_excel(path, rows=rows)
        return path

    def test_interest_fee_check_in_pipeline(self, tmp_dir):
        """测试主流程中自动生成利息手续费核对报告"""
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)

        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)
        self._create_beijing_bank_with_interest_fee(
            os.path.join(source_folder, '北京银行_流水.xlsx')
        )
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        result = bankcheck.run_pipeline(source_folder, script_dir)

        assert result.interest_fee_check_path is not None
        assert os.path.exists(result.interest_fee_check_path)
        assert '利息手续费核对报告' in os.path.basename(result.interest_fee_check_path)

        wb = openpyxl.load_workbook(result.interest_fee_check_path)
        assert '核对总览' in wb.sheetnames
        assert '交易明细' in wb.sheetnames
        assert '期间汇总' in wb.sheetnames
        assert '异常清单' in wb.sheetnames

        ws_detail = wb['交易明细']
        assert ws_detail.max_row >= 6
        wb.close()

    def test_interest_fee_check_standalone(self, tmp_dir):
        """测试独立运行利息手续费核对"""
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)

        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)
        excel_path = self._create_beijing_bank_with_interest_fee(
            os.path.join(source_folder, '北京银行_流水.xlsx')
        )
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        result = bankcheck.run_pipeline(source_folder, script_dir)
        total_path = result.output_path

        report_path = bankcheck.generate_interest_fee_check_from_total(
            total_path, output_dir=tmp_dir, period_type='month'
        )

        assert report_path is not None
        assert os.path.exists(report_path)

        wb = openpyxl.load_workbook(report_path)
        ws_overview = wb['核对总览']
        overview_data = {}
        for row in ws_overview.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] is not None:
                overview_data[row[0]] = row[1]

        assert overview_data.get('利息手续费交易总数') == 6
        assert overview_data.get('利息类交易数') == 2
        assert overview_data.get('手续费类交易数') == 4
        assert overview_data.get('利息总金额(元)') == pytest.approx(275.50)
        assert overview_data.get('手续费总金额(元)') == pytest.approx(190.0)
        wb.close()

    def test_interest_fee_check_quarterly(self, tmp_dir):
        """测试按季度汇总的利息手续费核对"""
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)

        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)
        excel_path = self._create_beijing_bank_with_interest_fee(
            os.path.join(source_folder, '北京银行_流水.xlsx')
        )
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        result = bankcheck.run_pipeline(source_folder, script_dir)
        total_path = result.output_path

        report_path = bankcheck.generate_interest_fee_check_from_total(
            total_path, output_dir=tmp_dir, period_type='quarter'
        )

        assert report_path is not None
        assert os.path.exists(report_path)

        wb = openpyxl.load_workbook(report_path)
        ws_period = wb['期间汇总']
        periods = set()
        for row in ws_period.iter_rows(min_row=2, values_only=True):
            if row[0]:
                periods.add(row[0])

        assert '2024Q1' in periods
        wb.close()

    def test_interest_fee_check_no_matching_transactions(self, tmp_dir):
        """测试无利息手续费交易的场景"""
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)

        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        result = bankcheck.run_pipeline(source_folder, script_dir)
        total_path = result.output_path

        report_path = bankcheck.generate_interest_fee_check_from_total(
            total_path, output_dir=tmp_dir, period_type='month'
        )

        assert report_path is None

    def test_interest_fee_check_multiple_banks(self, tmp_dir):
        """测试多银行场景下的利息手续费核对"""
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)

        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)

        beijing_rows = [
            [1, '2024-01-15', 'CNY', None, 125.50, 1500125.50, '北京银行', '622001', '北京银行', '结息', '001', '季度存款利息', None, None, None, 'BJ001'],
            [2, '2024-01-20', 'CNY', 50, None, 1500075.50, '北京银行', '622001', '北京银行', '手续费', '002', '转账手续费', None, None, None, 'BJ002'],
        ]
        _create_beijing_bank_excel(
            os.path.join(source_folder, '北京银行_流水.xlsx'), rows=beijing_rows
        )

        east_asia_rows = [
            ['2024-01-10', '09:30:00', 'CNY', 200, None, 499800, '手续费', 0, 0, '转账', 'EA001', '支付手续费'],
            ['2024-01-25', '14:15:00', 'CNY', None, 300.00, 500100, 0, 300, 0, '结息', 'EA002', '存款利息'],
        ]
        _create_east_asia_bank_excel(
            os.path.join(source_folder, '东亚银行_流水.xlsx'), rows=east_asia_rows
        )

        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        result = bankcheck.run_pipeline(source_folder, script_dir)
        total_path = result.output_path

        report_path = bankcheck.generate_interest_fee_check_from_total(
            total_path, output_dir=tmp_dir, period_type='month'
        )

        assert report_path is not None
        assert os.path.exists(report_path)

        wb = openpyxl.load_workbook(report_path)
        ws_detail = wb['交易明细']
        banks = set()
        for row in ws_detail.iter_rows(min_row=2, values_only=True):
            if row[2]:
                banks.add(row[2])

        assert '北京银行' in banks
        assert '东亚银行' in banks
        wb.close()


class TestHolidayCheckIntegration:
    """非工作日交易标记集成测试"""

    def _create_beijing_bank_with_weekend_holiday(self, path):
        rows = [
            [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A公司', '622001234', '工商银行', '转账', '001', '采购付款', None, None, None, 'BJ20240105001'],
            [2, '2024-01-06', 'CNY', 20000, None, 1480000, '供应商B公司', '622005678', '建设银行', '转账', '002', '周末紧急付款', None, None, None, 'BJ20240106002'],
            [3, '2024-01-01', 'CNY', None, 100000, 1580000, '客户C公司', '622009999', '北京银行', '转账', '003', '元旦收款', None, None, None, 'BJ20240101003'],
            [4, '2024-02-10', 'CNY', 30000, None, 1550000, '供应商D公司', '622001111', '农业银行', '转账', '004', '春节期间付款', None, None, None, 'BJ20240210004'],
            [5, '2024-02-04', 'CNY', None, 50000, 1600000, '客户E公司', '622002222', '工商银行', '转账', '005', '调休工作日收款', None, None, None, 'BJ20240204005'],
        ]
        _create_beijing_bank_excel(path, rows=rows)
        return path

    def test_holiday_check_in_pipeline(self, tmp_dir):
        """测试主流程中自动生成非工作日交易标记报告"""
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)

        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)
        self._create_beijing_bank_with_weekend_holiday(
            os.path.join(source_folder, '北京银行_流水.xlsx')
        )
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        result = bankcheck.run_pipeline(source_folder, script_dir)

        assert result.holiday_check_path is not None
        assert os.path.exists(result.holiday_check_path)
        assert '非工作日交易标记报告' in os.path.basename(result.holiday_check_path)

        wb = openpyxl.load_workbook(result.holiday_check_path)
        assert '标记总览' in wb.sheetnames
        assert '非工作日交易明细' in wb.sheetnames
        wb.close()

    def test_holiday_check_from_total(self, tmp_dir):
        """测试从总表文件生成非工作日交易标记报告"""
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)

        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        result = bankcheck.run_pipeline(source_folder, script_dir)
        total_path = result.output_path

        report_path = bankcheck.generate_holiday_check_from_total(
            total_path, output_dir=tmp_dir
        )

        assert report_path is not None
        assert os.path.exists(report_path)

    def test_holiday_tags_in_summary_table(self, tmp_dir):
        """测试总表中包含非工作日标签列"""
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)

        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)
        self._create_beijing_bank_with_weekend_holiday(
            os.path.join(source_folder, '北京银行_流水.xlsx')
        )
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))

        result = bankcheck.run_pipeline(source_folder, script_dir)

        assert result.output_path is not None
        df = pd.read_excel(result.output_path, engine='openpyxl')
        assert '非工作日标签' in df.columns
        assert '节假日名称' in df.columns


class TestBuildCliParser:

    def test_parser_has_process_subcommand(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['process', '/tmp/test_folder'])
        assert args.command == 'process'
        assert args.folder == '/tmp/test_folder'

    def test_parser_has_validate_lookup_subcommand(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['validate-lookup'])
        assert args.command == 'validate-lookup'

    def test_parser_has_version_subcommand(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['version'])
        assert args.command == 'version'

    def test_process_no_incremental_flag(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['process', '/tmp/test', '--no-incremental'])
        assert args.no_incremental is True

    def test_process_default_incremental(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['process', '/tmp/test'])
        assert args.no_incremental is False

    def test_process_keep_strategy(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['process', '/tmp/test', '--keep-strategy', 'keep_all'])
        assert args.keep_strategy == 'keep_all'

    def test_process_default_keep_strategy(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['process', '/tmp/test'])
        assert args.keep_strategy == 'keep_unprocessed'

    def test_process_output_dir(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['process', '/tmp/test', '--output-dir', '/tmp/out'])
        assert args.output_dir == '/tmp/out'

    def test_process_preset(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['process', '/tmp/test', '--preset', 'my-preset'])
        assert args.preset == 'my-preset'

    def test_process_enabled_banks(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['process', '/tmp/test', '--enabled-banks', '北京银行', '东亚银行'])
        assert args.enabled_banks == ['北京银行', '东亚银行']

    def test_process_date_filters(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args([
            'process', '/tmp/test',
            '--start-date', '2024-01-01',
            '--end-date', '2024-12-31',
        ])
        assert args.start_date == '2024-01-01'
        assert args.end_date == '2024-12-31'

    def test_validate_lookup_json_flag(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['validate-lookup', '--json'])
        assert args.json is True

    def test_validate_lookup_strict_flag(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['validate-lookup', '--strict'])
        assert args.strict is True

    def test_validate_lookup_file_option(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['validate-lookup', '--lookup-file', '/path/to/lookup.xlsx'])
        assert args.lookup_file == '/path/to/lookup.xlsx'

    def test_no_command_returns_none(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args([])
        assert args.command is None

    def test_legacy_flags_preserved(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['--watch-dir', '/tmp/test', '--once'])
        assert args.watch_dir == '/tmp/test'
        assert args.once is True

    def test_legacy_scheduler_flag(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['--scheduler'])
        assert args.scheduler is True

    def test_legacy_list_jobs_flag(self):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['--list-jobs'])
        assert args.list_jobs is True


class TestCmdProcess:

    def _setup_folder(self, tmp_dir, script_dir):
        source_folder = os.path.join(tmp_dir, '流水文件夹')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))
        _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))
        return source_folder

    def test_process_runs_pipeline(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['process', source])

        original_get_script_dir = bankcheck.get_script_dir
        bankcheck.get_script_dir = lambda: script_dir
        try:
            ret = bankcheck._cmd_process(args)
        finally:
            bankcheck.get_script_dir = original_get_script_dir

        assert ret == 0

    def test_process_nonexistent_dir(self, tmp_dir):
        nonexistent = os.path.join(tmp_dir, 'nonexistent_folder')

        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['process', nonexistent])

        original_get_script_dir = bankcheck.get_script_dir
        bankcheck.get_script_dir = lambda: tmp_dir
        try:
            ret = bankcheck._cmd_process(args)
        finally:
            bankcheck.get_script_dir = original_get_script_dir

        assert ret == 1

    def test_process_with_no_incremental(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['process', source, '--no-incremental'])

        original_get_script_dir = bankcheck.get_script_dir
        bankcheck.get_script_dir = lambda: script_dir
        try:
            ret = bankcheck._cmd_process(args)
        finally:
            bankcheck.get_script_dir = original_get_script_dir

        assert ret == 0

    def test_process_with_keep_all(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        source = self._setup_folder(tmp_dir, script_dir)

        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['process', source, '--keep-strategy', 'keep_all'])

        original_get_script_dir = bankcheck.get_script_dir
        bankcheck.get_script_dir = lambda: script_dir
        try:
            ret = bankcheck._cmd_process(args)
        finally:
            bankcheck.get_script_dir = original_get_script_dir

        assert ret == 0


class TestCmdValidateLookup:

    def _setup_lookup(self, tmp_dir, mappings=None):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        lookup_path = os.path.join(script_dir, '主体查找表.xlsx')
        _create_lookup_table(lookup_path, mappings)
        return script_dir, lookup_path

    def test_valid_lookup_returns_zero(self, tmp_dir):
        script_dir, lookup_path = self._setup_lookup(tmp_dir)

        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['validate-lookup', '--lookup-file', lookup_path])

        original_get_script_dir = bankcheck.get_script_dir
        bankcheck.get_script_dir = lambda: script_dir
        try:
            ret = bankcheck._cmd_validate_lookup(args)
        finally:
            bankcheck.get_script_dir = original_get_script_dir

        assert ret == 0

    def test_missing_lookup_without_strict(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)

        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['validate-lookup'])

        original_get_script_dir = bankcheck.get_script_dir
        bankcheck.get_script_dir = lambda: script_dir
        try:
            ret = bankcheck._cmd_validate_lookup(args)
        finally:
            bankcheck.get_script_dir = original_get_script_dir

        assert ret == 0

    def test_missing_lookup_with_strict(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)

        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['validate-lookup', '--strict'])

        original_get_script_dir = bankcheck.get_script_dir
        bankcheck.get_script_dir = lambda: script_dir
        try:
            ret = bankcheck._cmd_validate_lookup(args)
        finally:
            bankcheck.get_script_dir = original_get_script_dir

        assert ret == 1

    def test_json_output(self, tmp_dir, capsys):
        script_dir, lookup_path = self._setup_lookup(tmp_dir)

        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['validate-lookup', '--lookup-file', lookup_path, '--json'])

        original_get_script_dir = bankcheck.get_script_dir
        bankcheck.get_script_dir = lambda: script_dir
        try:
            ret = bankcheck._cmd_validate_lookup(args)
        finally:
            bankcheck.get_script_dir = original_get_script_dir

        assert ret == 0
        captured = capsys.readouterr()
        result = json.loads(captured.out)
        assert 'valid' in result
        assert 'file' in result
        assert 'total_entries' in result
        assert 'unique_accounts' in result
        assert result['valid'] is True

    def test_empty_lookup_json(self, tmp_dir, capsys):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)
        empty_lookup = os.path.join(script_dir, '主体查找表.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '主体映射'
        ws['A1'] = '主体名称'
        ws['B1'] = '银行账号'
        wb.save(empty_lookup)
        wb.close()

        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['validate-lookup', '--lookup-file', empty_lookup, '--json'])

        original_get_script_dir = bankcheck.get_script_dir
        bankcheck.get_script_dir = lambda: script_dir
        try:
            ret = bankcheck._cmd_validate_lookup(args)
        finally:
            bankcheck.get_script_dir = original_get_script_dir

        captured = capsys.readouterr()
        result = json.loads(captured.out)
        assert result['valid'] is False
        assert len(result['issues']) > 0

    def test_duplicate_account_mapping_detected(self, tmp_dir, capsys):
        mappings = [
            ('公司A', '1234567890'),
            ('公司B', '1234567890'),
        ]
        script_dir, lookup_path = self._setup_lookup(tmp_dir, mappings)

        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['validate-lookup', '--lookup-file', lookup_path, '--json'])

        original_get_script_dir = bankcheck.get_script_dir
        bankcheck.get_script_dir = lambda: script_dir
        try:
            ret = bankcheck._cmd_validate_lookup(args)
        finally:
            bankcheck.get_script_dir = original_get_script_dir

        captured = capsys.readouterr()
        result = json.loads(captured.out)
        assert result['valid'] is False
        assert len(result['duplicate_accounts']) > 0

    def test_missing_lookup_json_output(self, tmp_dir, capsys):
        script_dir = os.path.join(tmp_dir, 'script')
        os.makedirs(script_dir, exist_ok=True)

        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['validate-lookup', '--json'])

        original_get_script_dir = bankcheck.get_script_dir
        bankcheck.get_script_dir = lambda: script_dir
        try:
            ret = bankcheck._cmd_validate_lookup(args)
        finally:
            bankcheck.get_script_dir = original_get_script_dir

        captured = capsys.readouterr()
        result = json.loads(captured.out)
        assert result['valid'] is False
        assert '未找到主体查找表文件' in result['issues']


class TestCmdVersion:

    def test_version_returns_zero(self, capsys):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['version'])

        ret = bankcheck._cmd_version(args)
        assert ret == 0

    def test_version_prints_banner(self, capsys):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['version'])

        bankcheck._cmd_version(args)
        captured = capsys.readouterr()
        assert '银行流水检验工具' in captured.out
        assert '版本' in captured.out

    def test_version_prints_python_version(self, capsys):
        parser = bankcheck.build_cli_parser()
        args = parser.parse_args(['version'])

        bankcheck._cmd_version(args)
        captured = capsys.readouterr()
        assert 'Python' in captured.out
