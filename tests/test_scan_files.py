import os

import pytest

from bankcheck import (
    scan_excel_files,
    scan_bank_files,
    is_supported_bank_file,
    get_extension_format,
    detect_file_format,
    read_csv_as_workbook,
    open_workbook_compat,
)


class TestScanExcelFiles:
    def test_finds_xlsx_files(self, tmp_dir):
        open(os.path.join(tmp_dir, 'test.xlsx'), 'w').close()
        result = scan_excel_files(tmp_dir)
        assert len(result) == 1
        assert result[0].endswith('test.xlsx')

    def test_finds_xls_files(self, tmp_dir):
        open(os.path.join(tmp_dir, 'test.xls'), 'w').close()
        result = scan_excel_files(tmp_dir)
        assert len(result) == 1
        assert result[0].endswith('test.xls')

    def test_finds_xlsm_files(self, tmp_dir):
        open(os.path.join(tmp_dir, 'test.xlsm'), 'w').close()
        result = scan_excel_files(tmp_dir)
        assert len(result) == 1
        assert result[0].endswith('test.xlsm')

    def test_finds_csv_files(self, tmp_dir):
        with open(os.path.join(tmp_dir, 'test.csv'), 'w') as f:
            f.write('col1,col2\nval1,val2\n')
        result = scan_excel_files(tmp_dir)
        assert len(result) == 1
        assert result[0].endswith('test.csv')

    def test_finds_all_supported_formats(self, tmp_dir):
        open(os.path.join(tmp_dir, 'a.xlsx'), 'w').close()
        open(os.path.join(tmp_dir, 'b.xls'), 'w').close()
        open(os.path.join(tmp_dir, 'c.xlsm'), 'w').close()
        with open(os.path.join(tmp_dir, 'd.csv'), 'w') as f:
            f.write('a,b\n1,2\n')
        result = scan_excel_files(tmp_dir)
        assert len(result) == 4

    def test_excludes_temp_files(self, tmp_dir):
        open(os.path.join(tmp_dir, '~$temp.xlsx'), 'w').close()
        open(os.path.join(tmp_dir, 'real.xlsx'), 'w').close()
        result = scan_excel_files(tmp_dir)
        assert len(result) == 1
        assert 'real.xlsx' in result[0]

    def test_excludes_temp_csv_files(self, tmp_dir):
        open(os.path.join(tmp_dir, '~$temp.csv'), 'w').close()
        with open(os.path.join(tmp_dir, 'real.csv'), 'w') as f:
            f.write('a,b\n1,2\n')
        result = scan_excel_files(tmp_dir)
        assert len(result) == 1
        assert 'real.csv' in result[0]

    def test_excludes_non_supported_files(self, tmp_dir):
        open(os.path.join(tmp_dir, 'notes.txt'), 'w').close()
        open(os.path.join(tmp_dir, 'report.pdf'), 'w').close()
        open(os.path.join(tmp_dir, 'data.doc'), 'w').close()
        result = scan_excel_files(tmp_dir)
        assert len(result) == 0

    def test_recursive_scan(self, tmp_dir):
        sub = os.path.join(tmp_dir, 'subdir')
        os.makedirs(sub)
        open(os.path.join(tmp_dir, 'root.xlsx'), 'w').close()
        open(os.path.join(sub, 'nested.xlsx'), 'w').close()
        result = scan_excel_files(tmp_dir)
        assert len(result) == 2

    def test_recursive_scan_with_csv(self, tmp_dir):
        sub = os.path.join(tmp_dir, 'subdir')
        os.makedirs(sub)
        open(os.path.join(tmp_dir, 'root.xlsx'), 'w').close()
        with open(os.path.join(sub, 'nested.csv'), 'w') as f:
            f.write('a,b\n1,2\n')
        result = scan_excel_files(tmp_dir)
        assert len(result) == 2

    def test_empty_folder(self, tmp_dir):
        result = scan_excel_files(tmp_dir)
        assert len(result) == 0

    def test_case_insensitive_extension(self, tmp_dir):
        open(os.path.join(tmp_dir, 'test.XLSX'), 'w').close()
        open(os.path.join(tmp_dir, 'test2.Xls'), 'w').close()
        result = scan_excel_files(tmp_dir)
        assert len(result) == 2

    def test_case_insensitive_csv_extension(self, tmp_dir):
        with open(os.path.join(tmp_dir, 'test.CSV'), 'w') as f:
            f.write('a,b\n1,2\n')
        with open(os.path.join(tmp_dir, 'test2.Csv'), 'w') as f:
            f.write('a,b\n1,2\n')
        result = scan_excel_files(tmp_dir)
        assert len(result) == 2

    def test_case_insensitive_xlsm_extension(self, tmp_dir):
        open(os.path.join(tmp_dir, 'test.XLSM'), 'w').close()
        open(os.path.join(tmp_dir, 'test2.Xlsm'), 'w').close()
        result = scan_excel_files(tmp_dir)
        assert len(result) == 2


class TestScanBankFiles:
    def test_scan_bank_files_alias(self, tmp_dir):
        open(os.path.join(tmp_dir, 'test.xlsx'), 'w').close()
        with open(os.path.join(tmp_dir, 'test.csv'), 'w') as f:
            f.write('a,b\n1,2\n')
        result1 = scan_excel_files(tmp_dir)
        result2 = scan_bank_files(tmp_dir)
        assert sorted(result1) == sorted(result2)
        assert len(result2) == 2


class TestIsSupportedBankFile:
    def test_supported_formats(self):
        assert is_supported_bank_file('test.xlsx') is True
        assert is_supported_bank_file('test.xls') is True
        assert is_supported_bank_file('test.xlsm') is True
        assert is_supported_bank_file('test.csv') is True

    def test_unsupported_formats(self):
        assert is_supported_bank_file('test.txt') is False
        assert is_supported_bank_file('test.pdf') is False
        assert is_supported_bank_file('test.doc') is False
        assert is_supported_bank_file('test') is False

    def test_case_insensitive(self):
        assert is_supported_bank_file('test.XLSX') is True
        assert is_supported_bank_file('test.CSV') is True
        assert is_supported_bank_file('test.XLSM') is True

    def test_with_path(self):
        assert is_supported_bank_file('/path/to/test.xlsx') is True
        assert is_supported_bank_file('/path/to/test.csv') is True


class TestGetExtensionFormat:
    def test_xlsx_extension(self):
        assert get_extension_format('test.xlsx') == 'xlsx'

    def test_xlsm_extension(self):
        assert get_extension_format('test.xlsm') == 'xlsx'

    def test_xls_extension(self):
        assert get_extension_format('test.xls') == 'xls'

    def test_csv_extension(self):
        assert get_extension_format('test.csv') == 'csv'

    def test_unknown_extension(self):
        assert get_extension_format('test.txt') == 'unknown'

    def test_case_insensitive(self):
        assert get_extension_format('test.XLSX') == 'xlsx'
        assert get_extension_format('test.CSV') == 'csv'
        assert get_extension_format('test.XLSM') == 'xlsx'


class TestDetectFileFormat:
    def test_detect_csv_format(self, tmp_dir):
        csv_path = os.path.join(tmp_dir, 'test.csv')
        with open(csv_path, 'w', encoding='utf-8') as f:
            f.write('日期,金额,余额\n2024-01-01,1000,5000\n')
        fmt = detect_file_format(csv_path)
        assert fmt == 'csv'

    def test_detect_xlsx_format(self, tmp_dir):
        import openpyxl
        xlsx_path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        wb.save(xlsx_path)
        wb.close()
        fmt = detect_file_format(xlsx_path)
        assert fmt == 'xlsx'

    def test_detect_xlsm_extension_returns_xlsx(self, tmp_dir):
        xlsm_path = os.path.join(tmp_dir, 'test.xlsm')
        import zipfile
        with zipfile.ZipFile(xlsm_path, 'w') as zf:
            zf.writestr('[Content_Types].xml', '<Types/>')
        fmt = detect_file_format(xlsm_path)
        assert fmt == 'xlsx'


class TestReadCsvAsWorkbook:
    def test_read_simple_csv(self, tmp_dir):
        csv_path = os.path.join(tmp_dir, 'test.csv')
        with open(csv_path, 'w', encoding='utf-8') as f:
            f.write('col1,col2,col3\n')
            f.write('val1,val2,val3\n')
            f.write('val4,val5,val6\n')

        wb = read_csv_as_workbook(csv_path)
        assert wb is not None
        assert len(wb.worksheets) == 1
        ws = wb.active
        assert ws.max_row == 3
        assert ws.max_column == 3
        assert ws.cell(1, 1).value == 'col1'
        assert ws.cell(2, 2).value == 'val2'
        assert ws.cell(3, 3).value == 'val6'

    def test_read_csv_with_empty_values(self, tmp_dir):
        csv_path = os.path.join(tmp_dir, 'test.csv')
        with open(csv_path, 'w', encoding='utf-8') as f:
            f.write('a,b,c\n')
            f.write('1,,3\n')
            f.write(',5,\n')

        wb = read_csv_as_workbook(csv_path)
        ws = wb.active
        assert ws.cell(2, 2).value is None
        assert ws.cell(3, 1).value is None
        assert ws.cell(3, 3).value is None

    def test_read_csv_gbk_encoding(self, tmp_dir):
        csv_path = os.path.join(tmp_dir, 'test_gbk.csv')
        with open(csv_path, 'w', encoding='gbk') as f:
            f.write('日期,金额,摘要\n')
            f.write('2024-01-01,1000.00,测试\n')

        wb = read_csv_as_workbook(csv_path)
        ws = wb.active
        assert ws.cell(1, 1).value == '日期'
        assert ws.cell(2, 3).value == '测试'

    def test_read_csv_with_semicolon_delimiter(self, tmp_dir):
        csv_path = os.path.join(tmp_dir, 'test_semicolon.csv')
        with open(csv_path, 'w', encoding='utf-8') as f:
            f.write('col1;col2;col3\n')
            f.write('val1;val2;val3\n')

        wb = read_csv_as_workbook(csv_path)
        ws = wb.active
        assert ws.max_column == 3
        assert ws.cell(1, 1).value == 'col1'
        assert ws.cell(2, 2).value == 'val2'

    def test_csv_workbook_iter_rows(self, tmp_dir):
        csv_path = os.path.join(tmp_dir, 'test.csv')
        with open(csv_path, 'w', encoding='utf-8') as f:
            f.write('a,b\n')
            f.write('1,2\n')
            f.write('3,4\n')

        wb = read_csv_as_workbook(csv_path)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 3
        assert rows[0] == ('a', 'b')
        assert rows[1] == ('1', '2')
        assert rows[2] == ('3', '4')

    def test_csv_workbook_close(self, tmp_dir):
        csv_path = os.path.join(tmp_dir, 'test.csv')
        with open(csv_path, 'w', encoding='utf-8') as f:
            f.write('a,b\n1,2\n')
        wb = read_csv_as_workbook(csv_path)
        wb.close()


class TestOpenWorkbookCompat:
    def test_open_csv_with_compat(self, tmp_dir):
        csv_path = os.path.join(tmp_dir, 'test.csv')
        with open(csv_path, 'w', encoding='utf-8') as f:
            f.write('col1,col2\n')
            f.write('val1,val2\n')

        wb, tmp_path = open_workbook_compat(csv_path)
        try:
            assert wb is not None
            assert tmp_path is None
            ws = wb.active
            assert ws.cell(1, 1).value == 'col1'
        finally:
            wb.close()

    def test_open_xlsx_with_compat(self, tmp_dir):
        import openpyxl
        xlsx_path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'test'
        wb.save(xlsx_path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(xlsx_path)
        try:
            assert wb2 is not None
            assert wb2.active['A1'].value == 'test'
        finally:
            wb2.close()
