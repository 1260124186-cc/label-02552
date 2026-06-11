"""
会计期间自动归属模块单元测试
"""
import os
import sys
import tempfile
import shutil
from datetime import datetime

import openpyxl
import pandas as pd
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='accounting_period_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _make_record(trade_date, payment=None, receipt=None, **kwargs):
    rec = {
        '唯一id': kwargs.get('唯一id', 'TEST'),
        '银行': kwargs.get('银行', '测试银行'),
        '银行账号': kwargs.get('银行账号', '123456'),
        '主体': kwargs.get('主体', '测试主体'),
        '交易日期': trade_date,
        '付款': payment,
        '收款': receipt,
        '摘要': kwargs.get('摘要', '测试'),
        '对方户名': kwargs.get('对方户名', ''),
        '余额': kwargs.get('余额', 0),
        '交易流水号': kwargs.get('交易流水号', 'T001'),
    }
    return rec


def _monthly_config(cutoff_day=25, **kwargs):
    return bankcheck.AccountingPeriodConfig(
        period_type=kwargs.get('period_type', 'monthly'),
        cutoff_day=cutoff_day,
        fiscal_year_start_month=kwargs.get('fiscal_year_start_month', 1),
        period_name_format=kwargs.get('period_name_format', 'YYYY-MM'),
    )


def _quarterly_config(cutoff_day=25, **kwargs):
    return bankcheck.AccountingPeriodConfig(
        period_type='quarterly',
        cutoff_day=cutoff_day,
        fiscal_year_start_month=kwargs.get('fiscal_year_start_month', 1),
        period_name_format=kwargs.get('period_name_format', 'YYYY-QN'),
    )


class TestAccountingPeriodConfig:
    """测试 AccountingPeriodConfig 数据类"""

    def test_default_values(self):
        cfg = bankcheck.AccountingPeriodConfig()
        assert cfg.period_type == 'monthly'
        assert cfg.cutoff_day == 25
        assert cfg.fiscal_year_start_month == 1
        assert cfg.period_name_format == 'YYYY-MM'

    def test_custom_values(self):
        cfg = bankcheck.AccountingPeriodConfig(
            period_type='quarterly', cutoff_day=20,
            fiscal_year_start_month=4, period_name_format='YYYY-QN',
        )
        assert cfg.period_type == 'quarterly'
        assert cfg.cutoff_day == 20
        assert cfg.fiscal_year_start_month == 4
        assert cfg.period_name_format == 'YYYY-QN'


class TestDetermineMonthlyPeriod:
    """测试 _determine_monthly_period 函数"""

    def test_before_cutoff(self):
        assert bankcheck._determine_monthly_period(2024, 1, 10, 25) == (2024, 1)
        assert bankcheck._determine_monthly_period(2024, 6, 25, 25) == (2024, 6)

    def test_after_cutoff(self):
        assert bankcheck._determine_monthly_period(2024, 1, 26, 25) == (2024, 2)
        assert bankcheck._determine_monthly_period(2024, 6, 30, 25) == (2024, 7)

    def test_december_rollover(self):
        assert bankcheck._determine_monthly_period(2024, 12, 26, 25) == (2025, 1)
        assert bankcheck._determine_monthly_period(2024, 12, 25, 25) == (2024, 12)

    def test_cutoff_day_31_february(self):
        assert bankcheck._determine_monthly_period(2024, 2, 29, 31) == (2024, 2)
        assert bankcheck._determine_monthly_period(2024, 2, 28, 28) == (2024, 2)

    def test_cutoff_day_30_february(self):
        assert bankcheck._determine_monthly_period(2024, 2, 28, 30) == (2024, 2)
        assert bankcheck._determine_monthly_period(2024, 2, 29, 30) == (2024, 2)

    def test_cutoff_day_1(self):
        assert bankcheck._determine_monthly_period(2024, 1, 1, 1) == (2024, 1)
        assert bankcheck._determine_monthly_period(2024, 1, 2, 1) == (2024, 2)

    def test_cutoff_day_28_february_non_leap(self):
        assert bankcheck._determine_monthly_period(2023, 2, 28, 28) == (2023, 2)
        assert bankcheck._determine_monthly_period(2023, 2, 28, 29) == (2023, 2)


class TestDetermineQuarterlyPeriod:
    """测试 _determine_quarterly_period 函数"""

    def test_q1_months(self):
        y, qi = bankcheck._determine_quarterly_period(2024, 1, 10, 25, 1)
        assert y == 2024 and qi == 0
        y, qi = bankcheck._determine_quarterly_period(2024, 2, 10, 25, 1)
        assert y == 2024 and qi == 0
        y, qi = bankcheck._determine_quarterly_period(2024, 3, 10, 25, 1)
        assert y == 2024 and qi == 0

    def test_q1_to_q2_after_cutoff(self):
        y, qi = bankcheck._determine_quarterly_period(2024, 3, 26, 25, 1)
        assert y == 2024 and qi == 1

    def test_q4_to_next_year(self):
        y, qi = bankcheck._determine_quarterly_period(2024, 12, 26, 25, 1)
        assert y == 2025 and qi == 0

    def test_non_last_month_of_quarter_ignores_cutoff(self):
        y, qi = bankcheck._determine_quarterly_period(2024, 1, 26, 25, 1)
        assert y == 2024 and qi == 0
        y, qi = bankcheck._determine_quarterly_period(2024, 2, 28, 25, 1)
        assert y == 2024 and qi == 0

    def test_fiscal_year_start_april(self):
        y, qi = bankcheck._determine_quarterly_period(2024, 4, 10, 25, 4)
        assert y == 2024 and qi == 0
        y, qi = bankcheck._determine_quarterly_period(2024, 5, 10, 25, 4)
        assert y == 2024 and qi == 0
        y, qi = bankcheck._determine_quarterly_period(2024, 6, 10, 25, 4)
        assert y == 2024 and qi == 0

    def test_fiscal_year_start_april_q1_cutoff(self):
        y, qi = bankcheck._determine_quarterly_period(2024, 6, 26, 25, 4)
        assert y == 2024 and qi == 1

    def test_fiscal_year_start_april_q4_rollover(self):
        y, qi = bankcheck._determine_quarterly_period(2024, 3, 26, 25, 4)
        assert y == 2024 and qi == 0


class TestQuarterLabel:
    """测试 _quarter_label 函数"""

    def test_yyyy_qn_format(self):
        assert bankcheck._quarter_label(2024, 0, 1, 'YYYY-QN') == '2024-Q1'
        assert bankcheck._quarter_label(2024, 3, 1, 'YYYY-QN') == '2024-Q4'

    def test_detailed_format(self):
        label = bankcheck._quarter_label(2024, 0, 1, 'YYYY-MM')
        assert '2024' in label
        assert 'Q1' in label

    def test_fiscal_year_start_april(self):
        label = bankcheck._quarter_label(2024, 0, 4, 'YYYY-QN')
        assert label == '2024-Q1'


class TestAssignAccountingPeriod:
    """测试 assign_accounting_period 核心函数"""

    def test_empty_records(self):
        enriched, summary = bankcheck.assign_accounting_period([], _monthly_config())
        assert enriched == []
        assert summary == {}

    def test_monthly_before_cutoff(self):
        records = [_make_record('2024-01-10', payment=-100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _monthly_config(cutoff_day=25))
        assert enriched[0]['会计期间'] == '2024-01'

    def test_monthly_after_cutoff(self):
        records = [_make_record('2024-01-26', payment=-100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _monthly_config(cutoff_day=25))
        assert enriched[0]['会计期间'] == '2024-02'

    def test_monthly_on_cutoff_day(self):
        records = [_make_record('2024-01-25', payment=-100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _monthly_config(cutoff_day=25))
        assert enriched[0]['会计期间'] == '2024-01'

    def test_monthly_december_rollover(self):
        records = [_make_record('2024-12-26', payment=-100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _monthly_config(cutoff_day=25))
        assert enriched[0]['会计期间'] == '2025-01'

    def test_multiple_periods(self):
        records = [
            _make_record('2024-01-10', payment=-100),
            _make_record('2024-01-26', payment=-200),
            _make_record('2024-02-05', receipt=300),
        ]
        enriched, summary = bankcheck.assign_accounting_period(records, _monthly_config(cutoff_day=25))
        assert enriched[0]['会计期间'] == '2024-01'
        assert enriched[1]['会计期间'] == '2024-02'
        assert enriched[2]['会计期间'] == '2024-02'
        assert '2024-01' in summary
        assert '2024-02' in summary
        assert len(summary['2024-01']) == 1
        assert len(summary['2024-02']) == 2

    def test_quarterly_before_cutoff(self):
        records = [_make_record('2024-03-10', payment=-100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _quarterly_config(cutoff_day=25))
        assert enriched[0]['会计期间'] == '2024-Q1'

    def test_quarterly_after_cutoff_in_last_month(self):
        records = [_make_record('2024-03-26', payment=-100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _quarterly_config(cutoff_day=25))
        assert enriched[0]['会计期间'] == '2024-Q2'

    def test_quarterly_non_last_month_ignores_cutoff(self):
        records = [_make_record('2024-01-26', payment=-100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _quarterly_config(cutoff_day=25))
        assert enriched[0]['会计期间'] == '2024-Q1'

    def test_quarterly_year_rollover(self):
        records = [_make_record('2024-12-26', payment=-100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _quarterly_config(cutoff_day=25))
        assert enriched[0]['会计期间'] == '2025-Q1'

    def test_invalid_date_goes_to_unknown(self):
        records = [_make_record('invalid-date', payment=-100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _monthly_config())
        assert enriched[0]['会计期间'] == '未知期间'

    def test_none_date_goes_to_unknown(self):
        records = [_make_record(None, payment=-100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _monthly_config())
        assert enriched[0]['会计期间'] == '未知期间'

    def test_original_record_not_modified(self):
        records = [_make_record('2024-01-10', payment=-100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _monthly_config())
        assert '会计期间' not in records[0]
        assert '会计期间' in enriched[0]

    def test_datetime_object(self):
        dt = datetime(2024, 1, 10)
        records = [_make_record(dt, payment=-100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _monthly_config(cutoff_day=25))
        assert enriched[0]['会计期间'] == '2024-01'

    def test_cutoff_day_1(self):
        records = [_make_record('2024-01-01', receipt=100)]
        enriched, summary = bankcheck.assign_accounting_period(records, _monthly_config(cutoff_day=1))
        assert enriched[0]['会计期间'] == '2024-01'
        records2 = [_make_record('2024-01-02', receipt=100)]
        enriched2, _ = bankcheck.assign_accounting_period(records2, _monthly_config(cutoff_day=1))
        assert enriched2[0]['会计期间'] == '2024-02'

    def test_period_name_format_yyyy_qn_with_monthly(self):
        records = [_make_record('2024-01-10', payment=-100)]
        cfg = _monthly_config(period_name_format='YYYY-QN')
        enriched, summary = bankcheck.assign_accounting_period(records, cfg)
        assert enriched[0]['会计期间'] == '2024-Q1'


class TestComputePeriodAggregates:
    """测试 _compute_period_aggregates 函数"""

    def test_basic_aggregation(self):
        period_summary = {
            '2024-01': [
                _make_record('2024-01-10', payment=-1000, receipt=None),
                _make_record('2024-01-15', payment=None, receipt=2000),
            ],
        }
        result = bankcheck._compute_period_aggregates(period_summary)
        assert result['2024-01']['收入总额'] == 2000.0
        assert result['2024-01']['支出总额'] == 1000.0
        assert result['2024-01']['净额'] == 1000.0
        assert result['2024-01']['交易笔数'] == 2
        assert result['2024-01']['收入笔数'] == 1
        assert result['2024-01']['支出笔数'] == 1

    def test_zero_amount_records(self):
        period_summary = {
            '2024-01': [
                _make_record('2024-01-10', payment=0, receipt=0),
            ],
        }
        result = bankcheck._compute_period_aggregates(period_summary)
        assert result['2024-01']['收入总额'] == 0.0
        assert result['2024-01']['支出总额'] == 0.0
        assert result['2024-01']['收入笔数'] == 0
        assert result['2024-01']['支出笔数'] == 0

    def test_multiple_periods(self):
        period_summary = {
            '2024-01': [_make_record('2024-01-10', payment=-500)],
            '2024-02': [_make_record('2024-02-10', receipt=800)],
        }
        result = bankcheck._compute_period_aggregates(period_summary)
        assert len(result) == 2
        assert result['2024-01']['支出总额'] == 500.0
        assert result['2024-02']['收入总额'] == 800.0


class TestExportAccountingPeriodSummary:
    """测试 Excel 导出功能"""

    def test_export_creates_file(self, tmp_dir):
        records = [
            _make_record('2024-01-10', payment=-1000),
            _make_record('2024-01-20', receipt=2000),
        ]
        config = _monthly_config(cutoff_day=25)
        enriched, period_summary = bankcheck.assign_accounting_period(records, config)
        output_path = os.path.join(tmp_dir, '会计期间总表.xlsx')

        result = bankcheck.export_accounting_period_summary(
            enriched, period_summary, output_path)

        assert result == output_path
        assert os.path.exists(output_path)
        assert os.path.getsize(output_path) > 0

    def test_export_has_summary_sheet(self, tmp_dir):
        records = [
            _make_record('2024-01-10', payment=-1000),
            _make_record('2024-02-15', receipt=2000),
        ]
        config = _monthly_config(cutoff_day=25)
        enriched, period_summary = bankcheck.assign_accounting_period(records, config)
        output_path = os.path.join(tmp_dir, '会计期间总表.xlsx')
        bankcheck.export_accounting_period_summary(enriched, period_summary, output_path)

        wb = openpyxl.load_workbook(output_path)
        assert '期间汇总' in wb.sheetnames
        wb.close()

    def test_export_has_period_sheets(self, tmp_dir):
        records = [
            _make_record('2024-01-10', payment=-1000),
            _make_record('2024-02-15', receipt=2000),
        ]
        config = _monthly_config(cutoff_day=25)
        enriched, period_summary = bankcheck.assign_accounting_period(records, config)
        output_path = os.path.join(tmp_dir, '会计期间总表.xlsx')
        bankcheck.export_accounting_period_summary(enriched, period_summary, output_path)

        wb = openpyxl.load_workbook(output_path)
        sheet_names = wb.sheetnames
        assert '2024-01' in sheet_names
        assert '2024-02' in sheet_names
        wb.close()

    def test_export_period_sheet_content(self, tmp_dir):
        records = [
            _make_record('2024-01-10', payment=-1000, receipt=None),
            _make_record('2024-01-20', payment=None, receipt=2000),
        ]
        config = _monthly_config(cutoff_day=25)
        enriched, period_summary = bankcheck.assign_accounting_period(records, config)
        output_path = os.path.join(tmp_dir, '会计期间总表.xlsx')
        bankcheck.export_accounting_period_summary(enriched, period_summary, output_path)

        df = pd.read_excel(output_path, sheet_name='2024-01')
        assert '会计期间' in df.columns
        data_rows = df.dropna(subset=['会计期间'])
        data_rows = data_rows[~data_rows['会计期间'].astype(str).str.contains('期间汇总')]
        assert len(data_rows) == 2

    def test_export_summary_sheet_content(self, tmp_dir):
        records = [
            _make_record('2024-01-10', payment=-1000),
            _make_record('2024-01-20', receipt=2000),
            _make_record('2024-02-05', receipt=500),
        ]
        config = _monthly_config(cutoff_day=25)
        enriched, period_summary = bankcheck.assign_accounting_period(records, config)
        output_path = os.path.join(tmp_dir, '会计期间总表.xlsx')
        bankcheck.export_accounting_period_summary(enriched, period_summary, output_path)

        df = pd.read_excel(output_path, sheet_name='期间汇总')
        assert '会计期间' in df.columns
        assert '收入总额' in df.columns
        assert '支出总额' in df.columns
        assert '净额' in df.columns

    def test_export_with_source_info(self, tmp_dir):
        records = [_make_record('2024-01-10', payment=-1000)]
        config = _monthly_config(cutoff_day=25)
        enriched, period_summary = bankcheck.assign_accounting_period(records, config)
        output_path = os.path.join(tmp_dir, '会计期间总表.xlsx')
        source_info = {'数据来源': '单元测试', '记录数': 1}
        bankcheck.export_accounting_period_summary(
            enriched, period_summary, output_path, source_info)

        wb = openpyxl.load_workbook(output_path)
        ws = wb['期间汇总']
        assert ws.cell(row=1, column=1).value == '数据源信息'
        wb.close()


class TestGenerateAccountingPeriodReport:
    """测试从记录直接生成报告"""

    def test_generate_from_records(self, tmp_dir):
        records = [
            _make_record('2024-01-10', payment=-5000),
            _make_record('2024-02-15', receipt=8000),
        ]
        result_path = bankcheck.generate_accounting_period_report(records, tmp_dir)
        assert result_path is not None
        assert os.path.exists(result_path)
        assert '会计期间总表' in os.path.basename(result_path)

    def test_generate_empty_records(self, tmp_dir):
        result_path = bankcheck.generate_accounting_period_report([], tmp_dir)
        assert result_path is None

    def test_generate_with_custom_config(self, tmp_dir):
        records = [
            _make_record('2024-03-26', payment=-1000),
        ]
        config = _quarterly_config(cutoff_day=25)
        result_path = bankcheck.generate_accounting_period_report(
            records, tmp_dir, config=config)
        assert result_path is not None
        wb = openpyxl.load_workbook(result_path)
        assert '期间汇总' in wb.sheetnames
        wb.close()

    def test_generate_with_source_info(self, tmp_dir):
        records = [_make_record('2024-01-10', payment=-1000)]
        source_info = {'来源': '测试'}
        result_path = bankcheck.generate_accounting_period_report(
            records, tmp_dir, source_info=source_info)
        assert result_path is not None


class TestAccountingPeriodManager:
    """测试 AccountingPeriodManager 配置管理器"""

    def test_singleton_pattern(self):
        bankcheck.AccountingPeriodManager._instance = None
        bankcheck._ap_manager_singleton = None
        m1 = bankcheck.get_accounting_period_manager()
        m2 = bankcheck.get_accounting_period_manager()
        assert m1 is m2
        bankcheck.AccountingPeriodManager._instance = None
        bankcheck._ap_manager_singleton = None

    def test_load_default_config(self, tmp_dir):
        import yaml
        config_path = os.path.join(tmp_dir, 'test_rules.yaml')
        with open(config_path, 'w', encoding='utf-8') as f:
            yaml.dump({'banks': []}, f)
        bankcheck.AccountingPeriodManager._instance = None
        bankcheck._ap_manager_singleton = None
        manager = bankcheck.AccountingPeriodManager(config_path)
        config = manager.get_config()
        assert config.period_type == 'monthly'
        assert config.cutoff_day == 25
        bankcheck.AccountingPeriodManager._instance = None
        bankcheck._ap_manager_singleton = None

    def test_load_custom_config(self, tmp_dir):
        import yaml
        config_path = os.path.join(tmp_dir, 'test_rules.yaml')
        custom = {
            'accounting_period': {
                'period_type': 'quarterly',
                'cutoff_day': 20,
                'fiscal_year_start_month': 4,
                'period_name_format': 'YYYY-QN',
            },
            'banks': [],
        }
        with open(config_path, 'w', encoding='utf-8') as f:
            yaml.dump(custom, f)
        bankcheck.AccountingPeriodManager._instance = None
        bankcheck._ap_manager_singleton = None
        manager = bankcheck.AccountingPeriodManager(config_path)
        config = manager.get_config()
        assert config.period_type == 'quarterly'
        assert config.cutoff_day == 20
        assert config.fiscal_year_start_month == 4
        assert config.period_name_format == 'YYYY-QN'
        bankcheck.AccountingPeriodManager._instance = None
        bankcheck._ap_manager_singleton = None

    def test_cutoff_day_clamped(self, tmp_dir):
        import yaml
        config_path = os.path.join(tmp_dir, 'test_rules.yaml')
        custom = {
            'accounting_period': {
                'cutoff_day': 50,
            },
            'banks': [],
        }
        with open(config_path, 'w', encoding='utf-8') as f:
            yaml.dump(custom, f)
        bankcheck.AccountingPeriodManager._instance = None
        bankcheck._ap_manager_singleton = None
        manager = bankcheck.AccountingPeriodManager(config_path)
        config = manager.get_config()
        assert config.cutoff_day == 31
        bankcheck.AccountingPeriodManager._instance = None
        bankcheck._ap_manager_singleton = None

    def test_cutoff_day_clamped_below_one(self, tmp_dir):
        import yaml
        config_path = os.path.join(tmp_dir, 'test_rules.yaml')
        custom = {
            'accounting_period': {
                'cutoff_day': -5,
            },
            'banks': [],
        }
        with open(config_path, 'w', encoding='utf-8') as f:
            yaml.dump(custom, f)
        bankcheck.AccountingPeriodManager._instance = None
        bankcheck._ap_manager_singleton = None
        manager = bankcheck.AccountingPeriodManager(config_path)
        config = manager.get_config()
        assert config.cutoff_day == 1
        bankcheck.AccountingPeriodManager._instance = None
        bankcheck._ap_manager_singleton = None

    def test_missing_config_file(self, tmp_dir):
        bankcheck.AccountingPeriodManager._instance = None
        bankcheck._ap_manager_singleton = None
        manager = bankcheck.AccountingPeriodManager('/nonexistent/path.yaml')
        config = manager.get_config()
        assert config.period_type == 'monthly'
        bankcheck.AccountingPeriodManager._instance = None
        bankcheck._ap_manager_singleton = None


class TestEdgeCases:
    """测试边界情况和复杂场景"""

    def test_february_29_leap_year(self):
        records = [_make_record('2024-02-29', payment=-100)]
        config = _monthly_config(cutoff_day=28)
        enriched, _ = bankcheck.assign_accounting_period(records, config)
        assert enriched[0]['会计期间'] == '2024-03'

    def test_february_28_non_leap(self):
        records = [_make_record('2023-02-28', payment=-100)]
        config = _monthly_config(cutoff_day=28)
        enriched, _ = bankcheck.assign_accounting_period(records, config)
        assert enriched[0]['会计期间'] == '2023-02'

    def test_february_28_non_leap_after_cutoff(self):
        records = [_make_record('2023-02-28', payment=-100)]
        config = _monthly_config(cutoff_day=27)
        enriched, _ = bankcheck.assign_accounting_period(records, config)
        assert enriched[0]['会计期间'] == '2023-03'

    def test_mixed_valid_invalid_dates(self):
        records = [
            _make_record('2024-01-10', payment=-100),
            _make_record(None, payment=-200),
            _make_record('2024-02-05', receipt=300),
        ]
        config = _monthly_config(cutoff_day=25)
        enriched, summary = bankcheck.assign_accounting_period(records, config)
        assert len(enriched) == 3
        assert '未知期间' in summary

    def test_all_same_period(self):
        records = [
            _make_record('2024-01-05', payment=-100),
            _make_record('2024-01-10', receipt=200),
            _make_record('2024-01-20', payment=-50),
        ]
        config = _monthly_config(cutoff_day=25)
        enriched, summary = bankcheck.assign_accounting_period(records, config)
        assert len(summary) == 1
        assert '2024-01' in summary
        assert len(summary['2024-01']) == 3

    def test_cross_year_boundary(self):
        records = [
            _make_record('2024-12-25', payment=-100),
            _make_record('2024-12-26', receipt=200),
        ]
        config = _monthly_config(cutoff_day=25)
        enriched, summary = bankcheck.assign_accounting_period(records, config)
        assert enriched[0]['会计期间'] == '2024-12'
        assert enriched[1]['会计期间'] == '2025-01'

    def test_quarterly_cross_year(self):
        records = [
            _make_record('2024-12-25', payment=-100),
            _make_record('2024-12-26', receipt=200),
        ]
        config = _quarterly_config(cutoff_day=25)
        enriched, summary = bankcheck.assign_accounting_period(records, config)
        assert enriched[0]['会计期间'] == '2024-Q4'
        assert enriched[1]['会计期间'] == '2025-Q1'

    def test_fiscal_year_start_april_quarterly(self):
        records = [
            _make_record('2024-03-25', payment=-100),
            _make_record('2024-03-26', receipt=200),
            _make_record('2024-04-01', payment=-50),
        ]
        config = _quarterly_config(cutoff_day=25, fiscal_year_start_month=4)
        enriched, summary = bankcheck.assign_accounting_period(records, config)
        assert enriched[0]['会计期间'] == '2024-Q4'
        assert enriched[1]['会计期间'] == '2024-Q1'
        assert enriched[2]['会计期间'] == '2024-Q1'

    def test_slash_date_format(self):
        records = [_make_record('2024/01/10', payment=-100)]
        config = _monthly_config(cutoff_day=25)
        enriched, _ = bankcheck.assign_accounting_period(records, config)
        assert enriched[0]['会计期间'] == '2024-01'

    def test_pandas_timestamp(self):
        ts = pd.Timestamp('2024-06-15')
        records = [_make_record(ts, payment=-100)]
        config = _monthly_config(cutoff_day=25)
        enriched, _ = bankcheck.assign_accounting_period(records, config)
        assert enriched[0]['会计期间'] == '2024-06'

    def test_large_dataset(self):
        records = []
        for month in range(1, 13):
            for day in [5, 15, 25, 28]:
                records.append(_make_record(f'2024-{month:02d}-{day:02d}', payment=-100))
        config = _monthly_config(cutoff_day=25)
        enriched, summary = bankcheck.assign_accounting_period(records, config)
        assert len(enriched) == 48
        assert len(summary) >= 12

    def test_extra_fields_in_records(self, tmp_dir):
        records = [
            {**_make_record('2024-01-10', payment=-100), '黑白名单标签': '白名单'},
        ]
        config = _monthly_config(cutoff_day=25)
        enriched, period_summary = bankcheck.assign_accounting_period(records, config)
        output_path = os.path.join(tmp_dir, 'extra_fields.xlsx')
        bankcheck.export_accounting_period_summary(enriched, period_summary, output_path)

        df = pd.read_excel(output_path, sheet_name='2024-01')
        assert '黑白名单标签' in df.columns


class TestProcessingResultIntegration:
    """测试 ProcessingResult 中的 accounting_period_path 字段"""

    def test_processing_result_has_field(self):
        result = bankcheck.ProcessingResult()
        assert result.accounting_period_path is None

    def test_processing_result_with_path(self):
        result = bankcheck.ProcessingResult(accounting_period_path='/tmp/test.xlsx')
        assert result.accounting_period_path == '/tmp/test.xlsx'
