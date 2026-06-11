import os
import sys
import shutil
import tempfile

import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import onboarding
import bankcheck


@pytest.fixture(autouse=True)
def setup_and_teardown():
    onboarding.reset_first_run_marker()
    yield
    onboarding.reset_first_run_marker()


@pytest.fixture
def temp_script_dir():
    d = tempfile.mkdtemp(prefix='bankcheck_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


class TestFirstRunDetection:
    def test_is_first_run_initial(self):
        assert onboarding.is_first_run() is True

    def test_mark_first_run_complete(self):
        assert onboarding.is_first_run() is True
        assert onboarding.mark_first_run_complete() is True
        assert onboarding.is_first_run() is False

    def test_reset_first_run_marker(self):
        onboarding.mark_first_run_complete()
        assert onboarding.is_first_run() is False
        assert onboarding.reset_first_run_marker() is True
        assert onboarding.is_first_run() is True


class TestDemoLookupTable:
    def test_create_demo_lookup_table(self, temp_script_dir):
        lookup_path = os.path.join(temp_script_dir, '主体查找表.xlsx')
        result = onboarding.create_demo_lookup_table(lookup_path)

        assert result == lookup_path
        assert os.path.exists(lookup_path)

        import openpyxl
        wb = openpyxl.load_workbook(lookup_path)
        ws = wb.active

        assert ws['A1'].value == '主体名称'
        assert ws['B1'].value == '银行账号'
        assert ws['C1'].value == '优先级'
        assert ws['D1'].value == '备注'

        assert ws['A2'].value == '北京XX科技有限公司'
        assert ws['B2'].value == '01090312345678901'
        assert ws['C2'].value == 1

        assert ws['A3'].value == '上海YY贸易有限公司'
        assert ws['B3'].value == '38812345678'
        assert ws['C3'].value == 0

        wb.close()


class TestPrepareDemoEnvironment:
    def test_prepare_demo_environment_success(self, temp_script_dir):
        samples_dir = os.path.join(os.path.dirname(temp_script_dir), 'samples')
        os.makedirs(samples_dir, exist_ok=True)

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '测试'
        wb.save(os.path.join(samples_dir, '北京银行_示例.xlsx'))
        wb.close()

        demo_folder, demo_lookup_path, temp_dir = onboarding.prepare_demo_environment(temp_script_dir)

        assert demo_folder is not None
        assert demo_lookup_path is not None
        assert temp_dir != ''
        assert os.path.exists(demo_folder)
        assert os.path.exists(demo_lookup_path)
        assert len(os.listdir(demo_folder)) >= 1

        shutil.rmtree(samples_dir, ignore_errors=True)
        shutil.rmtree(temp_dir, ignore_errors=True)

    def test_prepare_demo_environment_no_samples(self, temp_script_dir):
        temp_dir = tempfile.mkdtemp(prefix='no_samples_test_')
        test_script_dir = os.path.join(temp_dir, 'backend')
        os.makedirs(test_script_dir, exist_ok=True)

        demo_folder, demo_lookup_path, result_temp_dir = onboarding.prepare_demo_environment(test_script_dir)

        assert demo_folder is None
        assert demo_lookup_path is None
        assert result_temp_dir == ''

        shutil.rmtree(temp_dir, ignore_errors=True)


class TestRunDemoMode:
    def test_run_demo_mode_integration(self, temp_script_dir):
        samples_dir = os.path.join(os.path.dirname(temp_script_dir), 'samples')
        os.makedirs(samples_dir, exist_ok=True)

        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '交易明细'
        ws['A1'] = '北京银行交易明细'
        ws['B2'] = '01090312345678901'
        headers = ['序号', '交易日期', '币种', '支出金额', '收入金额', '余额',
                   '对方户名', '对方账号', '对方行名', '凭证种类', '凭证号码',
                   '摘要', '备注1', '备注2', '备注3', '交易流水号']
        for c, h in enumerate(headers, 1):
            ws.cell(row=3, column=c, value=h)
        ws.cell(row=4, column=1, value=1)
        ws.cell(row=4, column=2, value='2024-01-05')
        ws.cell(row=4, column=4, value=50000)
        ws.cell(row=4, column=6, value=1500000)
        ws.cell(row=4, column=7, value='供应商A公司')
        ws.cell(row=4, column=12, value='采购付款')
        ws.cell(row=4, column=16, value='BJ20240105001')
        wb.save(os.path.join(samples_dir, '北京银行_示例.xlsx'))
        wb.close()

        wb2 = openpyxl.Workbook()
        ws2 = wb2.active
        ws2.title = '交易明细'
        ws2['A1'] = '账号'
        ws2['B1'] = '38812345678'
        ws2['A2'] = '东亚银行交易明细'
        headers2 = ['交易日期', '交易时间', '币种', '支出金额', '收入金额',
                    '手续费', '利息', '税费', '余额', '交易类型', '交易流水号', '交易描述/对方']
        for c, h in enumerate(headers2, 1):
            ws2.cell(row=4, column=c, value=h)
        ws2.cell(row=5, column=1, value='2024-01-03')
        ws2.cell(row=5, column=4, value=20000)
        ws2.cell(row=5, column=9, value=480000)
        ws2.cell(row=5, column=11, value='EA20240103001')
        ws2.cell(row=5, column=12, value='向 张三 付款-货款')
        wb2.save(os.path.join(samples_dir, '东亚银行_示例.xlsx'))
        wb2.close()

        result = onboarding.run_demo_mode(temp_script_dir)

        assert result is True

        shutil.rmtree(samples_dir, ignore_errors=True)


class TestWritableDir:
    def test_get_writable_dir(self):
        d = onboarding.get_writable_dir()
        assert os.path.exists(d)
        assert os.access(d, os.W_OK)

    def test_get_user_data_dir(self):
        d = onboarding.get_user_data_dir()
        assert 'bankcheck' in d.lower()
