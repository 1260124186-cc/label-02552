import os

import openpyxl
import pytest

from bankcheck import identify_bank
from conftest import _create_beijing_bank_excel, _create_east_asia_bank_excel


class TestIdentifyBank:

    def test_beijing_bank_prefix(self):
        assert identify_bank('/some/path/北京银行_流水.xlsx') == '北京银行'

    def test_beijing_bank_prefix_with_spaces(self):
        assert identify_bank('/some/path/北京银行流水.xlsx') == '北京银行'

    def test_east_asia_bank_prefix(self):
        assert identify_bank('/some/path/东亚银行_流水.xlsx') == '东亚银行'

    def test_east_asia_bank_prefix_with_spaces(self):
        assert identify_bank('/some/path/东亚银行流水.xlsx') == '东亚银行'

    def test_unknown_bank(self):
        assert identify_bank('/some/path/招商银行_流水.xlsx') is None

    def test_no_bank_name(self):
        assert identify_bank('/some/path/流水文件.xlsx') is None

    def test_bank_name_in_middle(self):
        assert identify_bank('/some/path/2024年北京银行流水.xlsx') == '北京银行'

    def test_empty_filename(self):
        assert identify_bank('/some/path/.xlsx') is None

    def test_beijing_bank_not_east_asia(self):
        result = identify_bank('/some/path/北京银行东亚银行_流水.xlsx')
        assert result == '北京银行'

    def test_case_sensitivity(self):
        assert identify_bank('/some/path/东亚银行.xlsx') == '东亚银行'

    def test_bank_name_at_end(self):
        assert identify_bank('/some/path/2024年交易明细_北京银行.xlsx') == '北京银行'

    def test_bank_name_with_underscore(self):
        assert identify_bank('/some/path/2024_北京银行_流水.xlsx') == '北京银行'

    def test_bank_name_with_dash(self):
        assert identify_bank('/some/path/2024-北京银行-流水.xlsx') == '北京银行'

    def test_bank_name_with_dot(self):
        assert identify_bank('/some/path/2024.北京银行.流水.xlsx') == '北京银行'

    def test_bank_name_fullwidth_parentheses(self):
        assert identify_bank('/some/path/交易明细（北京银行）.xlsx') == '北京银行'

    def test_bank_name_halfwidth_parentheses(self):
        assert identify_bank('/some/path/交易明细(北京银行).xlsx') == '北京银行'

    def test_fullwidth_spaces(self):
        assert identify_bank('/some/path/2024　北京银行　流水.xlsx') == '北京银行'

    def test_fullwidth_dash(self):
        assert identify_bank('/some/path/2024－北京银行－流水.xlsx') == '北京银行'

    def test_fullwidth_underscore_like(self):
        assert identify_bank('/some/path/2024＿北京银行＿流水.xlsx') == '北京银行'

    def test_mixed_width_symbols(self):
        assert identify_bank('/some/path/2024_（北京银行）-流水.xlsx') == '北京银行'

    def test_east_asia_in_middle(self):
        assert identify_bank('/some/path/2024年度东亚银行明细.xlsx') == '东亚银行'

    def test_east_asia_fullwidth(self):
        assert identify_bank('/some/path/【东亚银行】流水.xlsx') == '东亚银行'

    def test_content_beijing_bank_b2(self, tmp_dir):
        filepath = os.path.join(tmp_dir, 'unknown_file.xlsx')
        _create_beijing_bank_excel(filepath)
        assert identify_bank(filepath) == '北京银行'

    def test_content_east_asia_b1(self, tmp_dir):
        filepath = os.path.join(tmp_dir, 'unknown_file.xlsx')
        _create_east_asia_bank_excel(filepath)
        assert identify_bank(filepath) == '东亚银行'

    def test_content_filename_has_priority(self, tmp_dir):
        filepath = os.path.join(tmp_dir, '东亚银行_流水.xlsx')
        _create_beijing_bank_excel(filepath)
        assert identify_bank(filepath) == '东亚银行'

    def test_content_filename_middle_has_priority(self, tmp_dir):
        filepath = os.path.join(tmp_dir, '2024_东亚银行_明细.xlsx')
        _create_beijing_bank_excel(filepath)
        assert identify_bank(filepath) == '东亚银行'

    def test_content_non_excel_file(self, tmp_dir):
        filepath = os.path.join(tmp_dir, 'test.txt')
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write('北京银行')
        assert identify_bank(filepath) is None

    def test_content_not_existing_file(self):
        assert identify_bank('/nonexistent/path/北京银行.xlsx') == '北京银行'

    def test_content_empty_excel(self, tmp_dir):
        filepath = os.path.join(tmp_dir, 'empty.xlsx')
        wb = openpyxl.Workbook()
        wb.save(filepath)
        wb.close()
        assert identify_bank(filepath) is None

    def test_content_b1_no_account(self, tmp_dir):
        filepath = os.path.join(tmp_dir, 'no_account.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['B1'] = '这不是账号'
        ws['B2'] = '也不是账号'
        wb.save(filepath)
        wb.close()
        assert identify_bank(filepath) is None

    def test_content_short_account(self, tmp_dir):
        filepath = os.path.join(tmp_dir, 'short_account.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['B2'] = '12345'
        wb.save(filepath)
        wb.close()
        assert identify_bank(filepath) is None

    def test_content_valid_account_length(self, tmp_dir):
        filepath = os.path.join(tmp_dir, 'valid_account.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['B2'] = '123456'
        wb.save(filepath)
        wb.close()
        assert identify_bank(filepath) == '北京银行'

    def test_prefix_matches_first(self):
        assert identify_bank('/some/path/北京银行东亚.xlsx') == '北京银行'

    def test_beijing_with_mixed_separators(self):
        assert identify_bank('/some/path/2024（北京·银行）流水.xlsx') == '北京银行'
