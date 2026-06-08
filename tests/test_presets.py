# -*- coding: utf-8 -*-
"""
任务配置预设功能测试
"""

import os
import json
import tempfile
import shutil
from datetime import datetime

import pytest
import openpyxl

import sys
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck


@pytest.fixture
def temp_script_dir():
    tmpdir = tempfile.mkdtemp()
    original_dir = os.getcwd()
    os.chdir(tmpdir)

    bankcheck.setup_logging()

    yield tmpdir

    os.chdir(original_dir)
    shutil.rmtree(tmpdir, ignore_errors=True)


def test_preset_config_file_creation(temp_script_dir):
    config_path = bankcheck.get_preset_config_path(temp_script_dir)
    assert not os.path.exists(config_path)

    config = bankcheck.load_preset_config(temp_script_dir)
    assert os.path.exists(config_path)
    assert 'presets' in config
    assert 'settings' in config
    assert len(config['presets']) == 0


def test_save_and_load_preset(temp_script_dir):
    preset_data = {
        'name': '测试预设',
        'description': '这是一个测试预设',
        'output_dir': '/tmp/output',
        'start_date': '2024-01-01',
        'end_date': '2024-12-31',
        'keep_strategy': 'keep_all',
        'enabled_banks': ['北京银行'],
        'incremental': False,
    }

    preset_id = bankcheck.save_preset(preset_data, temp_script_dir)
    assert preset_id is not None
    assert preset_id.startswith('PRESET')

    loaded = bankcheck.load_preset(preset_id, temp_script_dir)
    assert loaded is not None
    assert loaded['name'] == '测试预设'
    assert loaded['description'] == '这是一个测试预设'
    assert loaded['output_dir'] == '/tmp/output'
    assert loaded['start_date'] == '2024-01-01'
    assert loaded['end_date'] == '2024-12-31'
    assert loaded['keep_strategy'] == 'keep_all'
    assert loaded['enabled_banks'] == ['北京银行']
    assert loaded['incremental'] is False
    assert 'created_at' in loaded
    assert 'updated_at' in loaded


def test_list_presets(temp_script_dir):
    presets = bankcheck.list_presets(temp_script_dir)
    assert len(presets) == 0

    for i in range(3):
        bankcheck.save_preset({
            'name': f'预设{i}',
            'enabled_banks': ['北京银行', '东亚银行'],
        }, temp_script_dir)

    presets = bankcheck.list_presets(temp_script_dir)
    assert len(presets) == 3


def test_update_preset(temp_script_dir):
    import time
    preset_id = bankcheck.save_preset({
        'name': '原始名称',
        'enabled_banks': ['北京银行'],
    }, temp_script_dir)

    loaded = bankcheck.load_preset(preset_id, temp_script_dir)
    original_created = loaded['created_at']
    original_updated = loaded['updated_at']

    time.sleep(1.1)

    updated_id = bankcheck.save_preset({
        'preset_id': preset_id,
        'name': '更新后的名称',
        'enabled_banks': ['东亚银行'],
        'incremental': False,
    }, temp_script_dir)

    assert updated_id == preset_id

    updated = bankcheck.load_preset(preset_id, temp_script_dir)
    assert updated['name'] == '更新后的名称'
    assert updated['enabled_banks'] == ['东亚银行']
    assert updated['incremental'] is False
    assert updated['created_at'] == original_created
    assert updated['updated_at'] >= original_updated


def test_delete_preset(temp_script_dir):
    preset_id = bankcheck.save_preset({
        'name': '要删除的预设',
    }, temp_script_dir)

    assert bankcheck.load_preset(preset_id, temp_script_dir) is not None

    result = bankcheck.delete_preset(preset_id, temp_script_dir)
    assert result is True
    assert bankcheck.load_preset(preset_id, temp_script_dir) is None

    result = bankcheck.delete_preset(preset_id, temp_script_dir)
    assert result is False


def test_default_preset(temp_script_dir):
    preset_id = bankcheck.save_preset({
        'name': '默认预设',
    }, temp_script_dir)

    assert bankcheck.get_default_preset(temp_script_dir) is None

    bankcheck.set_default_preset(preset_id, temp_script_dir)
    default = bankcheck.get_default_preset(temp_script_dir)
    assert default is not None
    assert default['preset_id'] == preset_id
    assert default['name'] == '默认预设'

    bankcheck.delete_preset(preset_id, temp_script_dir)
    assert bankcheck.get_default_preset(temp_script_dir) is None


def test_delete_default_preset_clears_setting(temp_script_dir):
    preset_id = bankcheck.save_preset({'name': '默认'}, temp_script_dir)
    bankcheck.set_default_preset(preset_id, temp_script_dir)

    config = bankcheck.load_preset_config(temp_script_dir)
    assert config['settings']['default_preset'] == preset_id

    bankcheck.delete_preset(preset_id, temp_script_dir)

    config = bankcheck.load_preset_config(temp_script_dir)
    assert config['settings']['default_preset'] == ''


def test_keep_strategies():
    assert 'keep_unprocessed' in bankcheck.KEEP_STRATEGIES
    assert 'keep_all' in bankcheck.KEEP_STRATEGIES
    assert 'delete_all' in bankcheck.KEEP_STRATEGIES
    assert bankcheck.KEEP_STRATEGIES['keep_unprocessed'] == '仅保留未处理文件'


def test_save_preset_default_values(temp_script_dir):
    preset_id = bankcheck.save_preset({
        'name': '测试默认值',
    }, temp_script_dir)

    preset = bankcheck.load_preset(preset_id, temp_script_dir)
    assert preset['enabled_banks'] == bankcheck.BANK_PREFIXES
    assert preset['keep_strategy'] == 'keep_unprocessed'
    assert preset['incremental'] is True


def test_preset_id_generation(temp_script_dir):
    ids = set()
    for _ in range(10):
        pid = bankcheck.save_preset({'name': f'预设{_}'}, temp_script_dir)
        assert pid not in ids
        ids.add(pid)
        assert pid.startswith('PRESET')
        assert len(pid) > 20


def test_preset_config_json_format(temp_script_dir):
    bankcheck.save_preset({
        'name': 'JSON测试',
        'description': '测试JSON格式',
        'enabled_banks': ['北京银行'],
    }, temp_script_dir)

    config_path = bankcheck.get_preset_config_path(temp_script_dir)
    with open(config_path, 'r', encoding='utf-8') as f:
        data = json.load(f)

    assert 'presets' in data
    assert len(data['presets']) == 1
    preset = data['presets'][0]
    assert preset['name'] == 'JSON测试'
    assert preset['description'] == '测试JSON格式'


def test_run_pipeline_with_options_bank_filtering(temp_script_dir):
    sample_dir = os.path.join(temp_script_dir, 'samples')
    os.makedirs(sample_dir, exist_ok=True)

    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['B2'] = '1234567890'
    ws.append(['', '2024-01-15', '测试摘要', 1000, '', '', 'TRX001'])
    test_file = os.path.join(sample_dir, '北京银行_测试.xlsx')
    wb.save(test_file)

    result = bankcheck.run_pipeline_with_options(
        folder=sample_dir,
        script_dir=temp_script_dir,
        enabled_banks=['东亚银行'],
        incremental=False,
    )

    assert len(result.processed_files) == 0
    assert len(result.unprocessed_files) == 1
    assert '北京银行_测试.xlsx' in result.unprocessed_files[0]


def test_run_pipeline_with_options_date_filtering(temp_script_dir):
    sample_dir = os.path.join(temp_script_dir, 'samples')
    os.makedirs(sample_dir, exist_ok=True)

    import openpyxl

    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1['B2'] = '1234567890'

    for _ in range(3):
        ws1.append([])

    row_data_1 = [''] * 20
    row_data_1[1] = '2024-01-15'
    row_data_1[3] = 1000
    row_data_1[5] = 5000
    row_data_1[6] = '对方公司1'
    row_data_1[11] = '1月交易'
    row_data_1[15] = 'TRX001'
    ws1.append(row_data_1)

    row_data_2 = [''] * 20
    row_data_2[1] = '2024-03-15'
    row_data_2[3] = 2000
    row_data_2[5] = 7000
    row_data_2[6] = '对方公司2'
    row_data_2[11] = '3月交易'
    row_data_2[15] = 'TRX002'
    ws1.append(row_data_2)

    test_file1 = os.path.join(sample_dir, '北京银行_测试1.xlsx')
    wb1.save(test_file1)

    result = bankcheck.run_pipeline_with_options(
        folder=sample_dir,
        script_dir=temp_script_dir,
        incremental=False,
        start_date='2024-02-01',
        end_date='2024-04-01',
    )

    assert len(result.all_rows) == 1
    assert result.all_rows[0]['摘要'] == '3月交易'


def test_generate_preset_id():
    id1 = bankcheck._generate_preset_id()
    id2 = bankcheck._generate_preset_id()
    assert id1 != id2
    assert id1.startswith('PRESET')
    assert len(id1) > 20


def test_preset_dataclass():
    preset = bankcheck.TaskPreset(
        preset_id='TEST123',
        name='测试',
        description='描述',
        output_dir='/tmp',
        start_date='2024-01-01',
        end_date='2024-12-31',
        keep_strategy='keep_all',
        enabled_banks=['北京银行'],
        incremental=True,
        created_at='2024-01-01 00:00:00',
        updated_at='2024-01-01 00:00:00',
    )
    assert preset.preset_id == 'TEST123'
    assert preset.name == '测试'
    assert preset.enabled_banks == ['北京银行']


def test_output_dir_in_preset():
    """测试预设中的output_dir字段"""
    preset_data = {
        'name': '测试输出目录',
        'output_dir': '/custom/output/path',
        'enabled_banks': ['北京银行'],
        'keep_strategy': 'keep_unprocessed',
        'incremental': True,
    }

    preset_id = bankcheck.save_preset(preset_data, script_dir=None)
    loaded = bankcheck.load_preset(preset_id, script_dir=None)
    assert loaded['output_dir'] == '/custom/output/path'


def test_get_summary_table_path_with_output_dir(tmp_path):
    """测试get_summary_table_path支持自定义输出目录"""
    script_dir = str(tmp_path)
    output_dir = str(tmp_path / 'custom_output')

    default_path = bankcheck.get_summary_table_path(script_dir)
    assert default_path == os.path.join(script_dir, '银行流水总表.xlsx')

    custom_path = bankcheck.get_summary_table_path(script_dir, output_dir)
    assert custom_path == os.path.join(output_dir, '银行流水总表.xlsx')


def test_run_pipeline_with_options_output_dir(tmp_path):
    """测试run_pipeline_with_options使用自定义输出目录"""
    script_dir = str(tmp_path)
    output_dir = str(tmp_path / 'custom_output')
    bank_dir = tmp_path / '北京银行测试'
    bank_dir.mkdir()

    wb = openpyxl.Workbook()
    ws = wb.active
    for row_idx in range(1, 4):
        ws.cell(row=row_idx, column=1, value=f'header{row_idx}')
    ws.cell(row=4, column=2, value='6222021234567890')
    ws.cell(row=4, column=4, value=1000.0)
    ws.cell(row=4, column=6, value=5000.0)
    ws.cell(row=4, column=7, value='张三')
    ws.cell(row=4, column=12, value='测试交易')
    ws.cell(row=4, column=16, value='TXN001')
    wb.save(bank_dir / '北京银行测试.xlsx')

    result = bankcheck.run_pipeline_with_options(
        folder=str(bank_dir),
        script_dir=script_dir,
        incremental=False,
        output_dir=output_dir,
    )

    expected_output = os.path.join(output_dir, '银行流水总表.xlsx')
    assert result.output_path == expected_output
    assert os.path.exists(expected_output)
    assert len(result.all_rows) == 1


def test_apply_preset_with_output_dir(tmp_path):
    """测试应用预设时使用自定义输出目录"""
    script_dir = str(tmp_path)
    output_dir = str(tmp_path / 'preset_output')
    bank_dir = tmp_path / '北京银行测试'
    bank_dir.mkdir()

    wb = openpyxl.Workbook()
    ws = wb.active
    for row_idx in range(1, 4):
        ws.cell(row=row_idx, column=1, value=f'header{row_idx}')
    ws.cell(row=4, column=2, value='6222021234567890')
    ws.cell(row=4, column=4, value=1000.0)
    ws.cell(row=4, column=6, value=5000.0)
    ws.cell(row=4, column=7, value='张三')
    ws.cell(row=4, column=12, value='测试交易')
    ws.cell(row=4, column=16, value='TXN001')
    wb.save(bank_dir / '北京银行测试.xlsx')

    preset_data = {
        'name': '测试输出目录预设',
        'output_dir': output_dir,
        'enabled_banks': ['北京银行'],
        'keep_strategy': 'keep_unprocessed',
        'incremental': False,
        'start_date': '',
        'end_date': '',
    }
    preset_id = bankcheck.save_preset(preset_data, script_dir=script_dir)
    preset = bankcheck.load_preset(preset_id, script_dir=script_dir)

    result = bankcheck.apply_preset_to_pipeline(
        preset=preset,
        folder=str(bank_dir),
        script_dir=script_dir,
    )

    expected_output = os.path.join(output_dir, '银行流水总表.xlsx')
    assert result.output_path == expected_output
    assert os.path.exists(expected_output)


def test_preset_output_dir_empty_defaults_to_script_dir(tmp_path):
    """测试预设output_dir为空时使用script_dir"""
    script_dir = str(tmp_path)
    bank_dir = tmp_path / '北京银行测试'
    bank_dir.mkdir()

    wb = openpyxl.Workbook()
    ws = wb.active
    for row_idx in range(1, 4):
        ws.cell(row=row_idx, column=1, value=f'header{row_idx}')
    ws.cell(row=4, column=2, value='6222021234567890')
    ws.cell(row=4, column=4, value=1000.0)
    ws.cell(row=4, column=6, value=5000.0)
    ws.cell(row=4, column=7, value='张三')
    ws.cell(row=4, column=12, value='测试交易')
    ws.cell(row=4, column=16, value='TXN001')
    wb.save(bank_dir / '北京银行测试.xlsx')

    preset_data = {
        'name': '测试空输出目录',
        'output_dir': '',
        'enabled_banks': ['北京银行'],
        'keep_strategy': 'keep_unprocessed',
        'incremental': False,
    }
    preset_id = bankcheck.save_preset(preset_data, script_dir=script_dir)
    preset = bankcheck.load_preset(preset_id, script_dir=script_dir)

    result = bankcheck.apply_preset_to_pipeline(
        preset=preset,
        folder=str(bank_dir),
        script_dir=script_dir,
    )

    expected_output = os.path.join(script_dir, '银行流水总表.xlsx')
    assert result.output_path == expected_output
    assert os.path.exists(expected_output)
