"""
期末余额与银行对账单核对模块单元测试

覆盖场景：
1. 文本解析 - 金额、账号、日期从文本中正确提取
2. 期末余额提取 - 从交易记录中正确提取各账号期末余额
3. 手动输入余额 - 手动输入对账单余额功能正常
4. 余额比对逻辑 - 一致、差异、缺失等场景正确处理
5. 差异说明生成 - 根据差异情况智能生成说明
6. 容差设置 - 不同容差下的比对结果正确
7. 账号标准化 - 账号标准化后能正确匹配
8. 导出功能 - Excel报告正确生成
9. 边界场景 - 空数据、部分匹配、多账号混合
10. 数据类 - 数据类实例创建和属性访问正常
"""
import os
import sys
import tempfile
import shutil
from datetime import datetime

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck
from bankcheck import (
    _parse_amount_from_text,
    _parse_account_from_text,
    _parse_date_from_text,
    extract_end_balances_from_total,
    manual_input_balance,
    reconcile_balances,
    export_balance_reconciliation_result,
    generate_balance_reconciliation_from_total,
    BankStatementBalance,
    AccountEndBalance,
    BalanceReconciliationRecord,
    BalanceReconciliationResult,
    BALANCE_RECONCILIATION_FILENAME,
)


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='balance_recon_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _record(unique_id, account, date, payment=None, receipt=None, balance=0.0,
            bank='北京银行', subject='北京XX科技有限公司',
            summary='', counterpart='', txn_id=''):
    """构造标准化交易记录"""
    return {
        '唯一id': unique_id,
        '银行': bank,
        '银行账号': account,
        '主体': subject,
        '交易日期': date,
        '付款': payment,
        '收款': receipt,
        '摘要': summary,
        '对方户名': counterpart,
        '余额': balance,
        '交易流水号': txn_id or unique_id,
    }


class TestTextParsing:
    """文本解析测试"""

    def test_parse_amount_with_thousand_separator(self):
        """解析带千分位的金额"""
        assert _parse_amount_from_text('1,234,567.89') == pytest.approx(1234567.89)

    def test_parse_amount_with_currency_symbol(self):
        """解析带货币符号的金额"""
        assert _parse_amount_from_text('¥12,345.67') == pytest.approx(12345.67)
        assert _parse_amount_from_text('$1,234.56') == pytest.approx(1234.56)

    def test_parse_amount_negative(self):
        """解析负数金额"""
        assert _parse_amount_from_text('-1,234.56') == pytest.approx(-1234.56)

    def test_parse_amount_with_chinese_thousand(self):
        """解析中文千分位金额"""
        assert _parse_amount_from_text('1，234，567.89') == pytest.approx(1234567.89)

    def test_parse_amount_plain_number(self):
        """解析普通数字金额"""
        assert _parse_amount_from_text('12345.67') == pytest.approx(12345.67)
        assert _parse_amount_from_text('12345') == pytest.approx(12345.0)

    def test_parse_amount_from_sentence(self):
        """从句子中解析金额"""
        assert _parse_amount_from_text('期末余额：12,345.67元') == pytest.approx(12345.67)
        assert _parse_amount_from_text('余额 98,765.43 元整') == pytest.approx(98765.43)

    def test_parse_amount_invalid(self):
        """无效金额解析返回None"""
        assert _parse_amount_from_text('') is None
        assert _parse_amount_from_text(None) is None
        assert _parse_amount_from_text('abcdef') is None

    def test_parse_account_with_colon(self):
        """解析带冒号前缀的账号"""
        assert _parse_account_from_text('账号：6222021234567890123') == '6222021234567890123'
        assert _parse_account_from_text('账户: 01090312345678901') == '01090312345678901'

    def test_parse_account_with_separators(self):
        """解析带分隔符的账号"""
        assert _parse_account_from_text('卡号：6222 0212 3456 7890 123') == '6222021234567890123'
        assert _parse_account_from_text('账号 6222-0212-3456-7890-123') == '6222021234567890123'

    def test_parse_account_english_prefix(self):
        """解析英文前缀的账号"""
        assert _parse_account_from_text('Account No.: 123456789012') == '123456789012'
        assert _parse_account_from_text('A/C No. 9876543210987654321') == '9876543210987654321'

    def test_parse_account_plain(self):
        """解析纯数字账号（10位以上）"""
        assert _parse_account_from_text('6222021234567890123') == '6222021234567890123'

    def test_parse_account_too_short(self):
        """账号太短（<10位）返回None"""
        assert _parse_account_from_text('123456789') is None

    def test_parse_account_invalid(self):
        """无效账号解析返回None"""
        assert _parse_account_from_text('') is None
        assert _parse_account_from_text(None) is None

    def test_parse_date_chinese_format(self):
        """解析中文日期格式"""
        assert _parse_date_from_text('2024年12月31日') == '2024-12-31'
        assert _parse_date_from_text('2024/12/31') == '2024-12-31'
        assert _parse_date_from_text('2024-12-31') == '2024-12-31'

    def test_parse_date_compact_format(self):
        """解析紧凑日期格式"""
        assert _parse_date_from_text('20241231') == '2024-12-31'

    def test_parse_date_mm_dd_yyyy(self):
        """解析月/日/年格式"""
        assert _parse_date_from_text('12/31/2024') == '2024-12-31'

    def test_parse_date_in_sentence(self):
        """从句子中解析日期"""
        assert _parse_date_from_text('对账单日期：2024年12月31日') == '2024-12-31'
        assert _parse_date_from_text('期末日期 2024-12-31') == '2024-12-31'

    def test_parse_date_invalid(self):
        """无效日期解析返回None"""
        assert _parse_date_from_text('') is None
        assert _parse_date_from_text(None) is None
        assert _parse_date_from_text('invalid date') is None


class TestDataClasses:
    """数据类测试"""

    def test_bank_statement_balance_creation(self):
        """创建银行对账单余额记录"""
        record = BankStatementBalance(
            bank_account='6222021234567890123',
            bank_name='北京银行',
            subject='北京XX科技',
            statement_balance=12345.67,
            statement_date='2024-12-31',
            source_type='manual',
            source_file='test.pdf',
        )
        assert record.bank_account == '6222021234567890123'
        assert record.statement_balance == pytest.approx(12345.67)
        assert record.statement_date == '2024-12-31'

    def test_account_end_balance_creation(self):
        """创建总表期末余额记录"""
        record = AccountEndBalance(
            bank_account='6222021234567890123',
            bank_name='北京银行',
            subject='北京XX科技',
            end_balance=12345.67,
            end_date='2024-12-31',
            transaction_count=100,
            last_transaction_id='TXN001',
        )
        assert record.bank_account == '6222021234567890123'
        assert record.end_balance == pytest.approx(12345.67)
        assert record.transaction_count == 100

    def test_balance_reconciliation_record_creation(self):
        """创建比对记录"""
        record = BalanceReconciliationRecord(
            bank_account='6222021234567890123',
            bank_name='北京银行',
            subject='北京XX科技',
            total_balance=12345.67,
            statement_balance=12345.67,
            diff_amount=0.0,
            status='matched',
            diff_note='余额一致',
        )
        assert record.status == 'matched'
        assert record.diff_amount == pytest.approx(0.0)

    def test_balance_reconciliation_result_creation(self):
        """创建比对结果汇总"""
        result = BalanceReconciliationResult(
            total_accounts=10,
            matched_accounts=8,
            diff_accounts=1,
            missing_statement=1,
            missing_total=0,
            total_diff_amount=500.0,
        )
        assert result.total_accounts == 10
        assert result.matched_accounts == 8
        assert result.match_rate == '80.0%'


class TestExtractEndBalances:
    """期末余额提取测试"""

    def test_extract_single_account(self):
        """提取单个账号期末余额"""
        records = [
            _record('T001', '62220001', '2024-01-05', receipt=10000, balance=10000.0),
            _record('T002', '62220001', '2024-01-10', payment=-2000, balance=8000.0),
            _record('T003', '62220001', '2024-01-15', receipt=5000, balance=13000.0),
        ]
        result = extract_end_balances_from_total(records)
        assert len(result) == 1
        assert '62220001' in result
        assert result['62220001'].end_balance == pytest.approx(13000.0)
        assert result['62220001'].end_date == '2024-01-15'
        assert result['62220001'].transaction_count == 3

    def test_extract_multiple_accounts(self):
        """提取多个账号期末余额"""
        records = [
            _record('T001', '62220001', '2024-01-05', balance=10000.0),
            _record('T002', '62220002', '2024-01-10', balance=20000.0),
            _record('T003', '62220001', '2024-01-15', balance=15000.0),
            _record('T004', '62220002', '2024-01-20', balance=25000.0),
        ]
        result = extract_end_balances_from_total(records)
        assert len(result) == 2
        assert result['62220001'].end_balance == pytest.approx(15000.0)
        assert result['62220002'].end_balance == pytest.approx(25000.0)

    def test_extract_same_date_different_txn_id(self):
        """相同日期不同交易流水号取最后一条"""
        records = [
            _record('T001', '62220001', '2024-01-15', balance=10000.0, txn_id='TXN001'),
            _record('T002', '62220001', '2024-01-15', balance=15000.0, txn_id='TXN002'),
            _record('T003', '62220001', '2024-01-15', balance=12000.0, txn_id='TXN003'),
        ]
        result = extract_end_balances_from_total(records)
        assert result['62220001'].end_balance == pytest.approx(12000.0)
        assert result['62220001'].last_transaction_id == 'TXN003'

    def test_extract_empty_records(self):
        """空记录返回空字典"""
        result = extract_end_balances_from_total([])
        assert result == {}

    def test_extract_empty_account(self):
        """空账号记录被跳过"""
        records = [
            _record('T001', '', '2024-01-15', balance=10000.0),
            _record('T002', '62220001', '2024-01-15', balance=15000.0),
        ]
        result = extract_end_balances_from_total(records)
        assert len(result) == 1
        assert '62220001' in result

    def test_extract_none_balance(self):
        """最后一条记录无有效余额则跳过"""
        records = [
            _record('T001', '62220001', '2024-01-15', balance=None),
        ]
        result = extract_end_balances_from_total(records)
        assert len(result) == 0


class TestManualInputBalance:
    """手动输入余额测试"""

    def test_manual_input_basic(self):
        """基本手动输入"""
        result = manual_input_balance('62220001', 12345.67)
        assert result.bank_account == '62220001'
        assert result.statement_balance == pytest.approx(12345.67)
        assert result.source_type == 'manual'
        assert result.source_file == 'manual_input'

    def test_manual_input_full(self):
        """完整信息手动输入"""
        result = manual_input_balance(
            account='62220001',
            balance=12345.67,
            bank_name='北京银行',
            subject='北京XX科技',
            statement_date='2024-12-31',
        )
        assert result.bank_name == '北京银行'
        assert result.subject == '北京XX科技'
        assert result.statement_date == '2024-12-31'

    def test_manual_input_convert_to_float(self):
        """余额自动转换为float"""
        result = manual_input_balance('62220001', '12345.67')
        assert isinstance(result.statement_balance, float)
        assert result.statement_balance == pytest.approx(12345.67)


class TestReconcileBalances:
    """余额比对逻辑测试"""

    def test_reconcile_perfect_match(self):
        """完全匹配场景"""
        total_balances = {
            '62220001': AccountEndBalance(
                bank_account='62220001',
                bank_name='北京银行',
                subject='北京XX科技',
                end_balance=12345.67,
                end_date='2024-12-31',
                transaction_count=100,
            ),
        }
        statement_balances = {
            '62220001': BankStatementBalance(
                bank_account='62220001',
                bank_name='北京银行',
                subject='北京XX科技',
                statement_balance=12345.67,
                statement_date='2024-12-31',
                source_type='pdf',
            ),
        }
        result = reconcile_balances(total_balances, statement_balances)
        assert result.total_accounts == 1
        assert result.matched_accounts == 1
        assert result.diff_accounts == 0
        assert result.records[0].status == 'matched'
        assert '余额一致' in result.records[0].diff_note

    def test_reconcile_within_tolerance(self):
        """容差范围内视为一致"""
        total_balances = {
            '62220001': AccountEndBalance(
                bank_account='62220001', end_balance=12345.67,
                bank_name='北京银行', subject='测试公司',
            ),
        }
        statement_balances = {
            '62220001': BankStatementBalance(
                bank_account='62220001', statement_balance=12345.67,
            ),
        }
        result = reconcile_balances(total_balances, statement_balances, tolerance=0.01)
        assert result.matched_accounts == 1

    def test_reconcile_exceed_tolerance(self):
        """超出容差视为差异"""
        total_balances = {
            '62220001': AccountEndBalance(
                bank_account='62220001', end_balance=12345.67,
                bank_name='北京银行', subject='测试公司',
            ),
        }
        statement_balances = {
            '62220001': BankStatementBalance(
                bank_account='62220001', statement_balance=12300.00,
            ),
        }
        result = reconcile_balances(total_balances, statement_balances)
        assert result.diff_accounts == 1
        assert result.records[0].status == 'diff'
        assert result.records[0].diff_amount == pytest.approx(45.67)

    def test_reconcile_missing_statement(self):
        """缺少对账单"""
        total_balances = {
            '62220001': AccountEndBalance(
                bank_account='62220001', end_balance=12345.67,
                bank_name='北京银行', subject='测试公司',
            ),
        }
        statement_balances = {}
        result = reconcile_balances(total_balances, statement_balances)
        assert result.missing_statement == 1
        assert result.records[0].status == 'missing_statement'
        assert '缺少该账号的银行对账单' in result.records[0].diff_note

    def test_reconcile_missing_total(self):
        """缺少总表记录"""
        total_balances = {}
        statement_balances = {
            '62220001': BankStatementBalance(
                bank_account='62220001', statement_balance=12345.67,
            ),
        }
        result = reconcile_balances(total_balances, statement_balances)
        assert result.missing_total == 1
        assert result.records[0].status == 'missing_total'
        assert '总表中无此账号记录' in result.records[0].diff_note

    def test_reconcile_multiple_accounts_mixed(self):
        """多账号混合场景"""
        total_balances = {
            'ACC001': AccountEndBalance(
                bank_account='ACC001', end_balance=10000.00,
                bank_name='银行A', subject='公司A',
            ),
            'ACC002': AccountEndBalance(
                bank_account='ACC002', end_balance=20000.00,
                bank_name='银行B', subject='公司B',
            ),
            'ACC003': AccountEndBalance(
                bank_account='ACC003', end_balance=30000.00,
                bank_name='银行C', subject='公司C',
            ),
        }
        statement_balances = {
            'ACC001': BankStatementBalance(
                bank_account='ACC001', statement_balance=10000.00,
            ),
            'ACC002': BankStatementBalance(
                bank_account='ACC002', statement_balance=20500.00,
            ),
            'ACC004': BankStatementBalance(
                bank_account='ACC004', statement_balance=40000.00,
            ),
        }
        result = reconcile_balances(total_balances, statement_balances)
        assert result.total_accounts == 4
        assert result.matched_accounts == 1  # ACC001
        assert result.diff_accounts == 1     # ACC002
        assert result.missing_statement == 1  # ACC003
        assert result.missing_total == 1      # ACC004
        assert result.total_diff_amount == pytest.approx(500.0)

    def test_reconcile_large_diff_note(self):
        """大额差异生成特别说明"""
        total_balances = {
            '62220001': AccountEndBalance(
                bank_account='62220001', end_balance=100000.00,
                bank_name='北京银行', subject='测试公司',
                end_date='2024-12-31',
            ),
        }
        statement_balances = {
            '62220001': BankStatementBalance(
                bank_account='62220001', statement_balance=50000.00,
                statement_date='2024-12-31',
            ),
        }
        result = reconcile_balances(total_balances, statement_balances)
        assert '差异金额较大' in result.records[0].diff_note
        assert '逐笔核对交易明细' in result.records[0].diff_note

    def test_reconcile_date_mismatch_note(self):
        """日期不一致生成说明"""
        total_balances = {
            '62220001': AccountEndBalance(
                bank_account='62220001', end_balance=12345.67,
                bank_name='北京银行', subject='测试公司',
                end_date='2024-12-31',
            ),
        }
        statement_balances = {
            '62220001': BankStatementBalance(
                bank_account='62220001', statement_balance=12000.00,
                statement_date='2024-11-30',
            ),
        }
        result = reconcile_balances(total_balances, statement_balances)
        assert '日期不一致' in result.records[0].diff_note
        assert '2024-12-31' in result.records[0].diff_note
        assert '2024-11-30' in result.records[0].diff_note

    def test_reconcile_empty_data(self):
        """空数据比对"""
        result = reconcile_balances({}, {})
        assert result.total_accounts == 0
        assert result.check_summary.get('status') == '无数据'

    def test_reconcile_custom_tolerance(self):
        """自定义容差"""
        total_balances = {
            '62220001': AccountEndBalance(
                bank_account='62220001', end_balance=12345.67,
                bank_name='北京银行', subject='测试公司',
            ),
        }
        statement_balances = {
            '62220001': BankStatementBalance(
                bank_account='62220001', statement_balance=12345.00,
            ),
        }
        result = reconcile_balances(total_balances, statement_balances, tolerance=1.0)
        assert result.matched_accounts == 1

        result2 = reconcile_balances(total_balances, statement_balances, tolerance=0.5)
        assert result2.diff_accounts == 1

    def test_reconcile_check_summary(self):
        """核对摘要字段正确"""
        total_balances = {
            'ACC001': AccountEndBalance(
                bank_account='ACC001', end_balance=100.00,
                bank_name='银行A', subject='公司A',
            ),
        }
        statement_balances = {
            'ACC001': BankStatementBalance(
                bank_account='ACC001', statement_balance=100.00,
            ),
        }
        result = reconcile_balances(total_balances, statement_balances)
        assert result.check_summary['total_accounts'] == 1
        assert result.check_summary['matched_accounts'] == 1
        assert result.check_summary['match_rate'] == '100.0%'
        assert result.check_summary['tolerance'] == 0.01


class TestExportFunction:
    """导出功能测试"""

    def test_export_reconciliation_result(self, tmp_dir):
        """导出核对结果到Excel"""
        total_balances = {
            'ACC001': AccountEndBalance(
                bank_account='ACC001', bank_name='北京银行', subject='公司A',
                end_balance=10000.00, end_date='2024-12-31', transaction_count=50,
            ),
            'ACC002': AccountEndBalance(
                bank_account='ACC002', bank_name='工商银行', subject='公司B',
                end_balance=20000.00, end_date='2024-12-31', transaction_count=30,
            ),
            'ACC003': AccountEndBalance(
                bank_account='ACC003', bank_name='建设银行', subject='公司C',
                end_balance=30000.00, end_date='2024-12-31', transaction_count=20,
            ),
        }
        statement_balances = {
            'ACC001': BankStatementBalance(
                bank_account='ACC001', statement_balance=10000.00,
                statement_date='2024-12-31', source_file='acc001.pdf',
            ),
            'ACC002': BankStatementBalance(
                bank_account='ACC002', statement_balance=20500.00,
                statement_date='2024-12-31', source_file='acc002.pdf',
            ),
            'ACC004': BankStatementBalance(
                bank_account='ACC004', statement_balance=40000.00,
                statement_date='2024-12-31', source_file='acc004.pdf',
            ),
        }
        recon_result = reconcile_balances(total_balances, statement_balances)

        output_path = os.path.join(tmp_dir, 'test_reconciliation_report.xlsx')
        source_info = {'数据来源': '测试数据', '记录数': 100}

        result_path = export_balance_reconciliation_result(recon_result, output_path, source_info)

        assert os.path.exists(result_path)
        assert os.path.getsize(result_path) > 0

        wb = openpyxl.load_workbook(result_path)
        assert '核对总览' in wb.sheetnames
        assert '比对明细' in wb.sheetnames
        assert '差异账号清单' in wb.sheetnames
        assert '待完善清单' in wb.sheetnames

        ws_overview = wb['核对总览']
        assert ws_overview.cell(row=1, column=1).value == '核对项'
        assert ws_overview.cell(row=1, column=2).value == '数值'

        ws_detail = wb['比对明细']
        assert ws_detail.max_row == 5  # 1 header + 4 records

        ws_diff = wb['差异账号清单']
        assert ws_diff.max_row == 2  # 1 header + 1 diff record

        ws_pending = wb['待完善清单']
        assert ws_pending.max_row == 3  # 1 header + 1 missing statement + 1 missing total

        wb.close()

    def test_export_empty_result(self, tmp_dir):
        """导出空结果"""
        recon_result = reconcile_balances({}, {})
        output_path = os.path.join(tmp_dir, 'test_empty_report.xlsx')
        result_path = export_balance_reconciliation_result(recon_result, output_path)

        assert os.path.exists(result_path)
        wb = openpyxl.load_workbook(result_path)
        assert '核对总览' in wb.sheetnames
        wb.close()

    def test_generate_from_total_records(self, tmp_dir):
        """从交易记录生成报告"""
        records = [
            _record('T001', '62220001', '2024-01-05', balance=10000.0),
            _record('T002', '62220001', '2024-01-10', balance=15000.0),
            _record('T003', '62220002', '2024-01-15', balance=20000.0),
        ]
        statement_balances = {
            '62220001': BankStatementBalance(
                bank_account='62220001', statement_balance=15000.00,
                statement_date='2024-01-10',
            ),
            '62220002': BankStatementBalance(
                bank_account='62220002', statement_balance=20000.00,
                statement_date='2024-01-15',
            ),
        }

        result_path = generate_balance_reconciliation_from_total(
            records,
            statement_balances=statement_balances,
            output_dir=tmp_dir,
        )

        assert result_path is not None
        assert os.path.exists(result_path)
        assert '期末余额与银行对账单核对报告' in os.path.basename(result_path)

    def test_generate_from_total_empty_records(self, tmp_dir):
        """空记录生成报告返回None"""
        result_path = generate_balance_reconciliation_from_total(
            [],
            statement_balances={},
            output_dir=tmp_dir,
        )
        assert result_path is None


class TestEdgeCases:
    """边界场景测试"""

    def test_negative_balance(self):
        """负数余额比对"""
        total_balances = {
            'ACC001': AccountEndBalance(
                bank_account='ACC001', end_balance=-5000.00,
                bank_name='银行A', subject='公司A',
            ),
        }
        statement_balances = {
            'ACC001': BankStatementBalance(
                bank_account='ACC001', statement_balance=-5000.00,
            ),
        }
        result = reconcile_balances(total_balances, statement_balances)
        assert result.matched_accounts == 1
        assert result.records[0].diff_amount == pytest.approx(0.0)

    def test_zero_balance(self):
        """零余额比对"""
        total_balances = {
            'ACC001': AccountEndBalance(
                bank_account='ACC001', end_balance=0.0,
                bank_name='银行A', subject='公司A',
            ),
        }
        statement_balances = {
            'ACC001': BankStatementBalance(
                bank_account='ACC001', statement_balance=0.0,
            ),
        }
        result = reconcile_balances(total_balances, statement_balances)
        assert result.matched_accounts == 1

    def test_very_large_balance(self):
        """大额余额比对"""
        total_balances = {
            'ACC001': AccountEndBalance(
                bank_account='ACC001', end_balance=999999999.99,
                bank_name='银行A', subject='公司A',
            ),
        }
        statement_balances = {
            'ACC001': BankStatementBalance(
                bank_account='ACC001', statement_balance=999999999.99,
            ),
        }
        result = reconcile_balances(total_balances, statement_balances)
        assert result.matched_accounts == 1

    def test_medium_diff_note(self):
        """中等差异说明"""
        total_balances = {
            'ACC001': AccountEndBalance(
                bank_account='ACC001', end_balance=5000.00,
                bank_name='银行A', subject='公司A',
            ),
        }
        statement_balances = {
            'ACC001': BankStatementBalance(
                bank_account='ACC001', statement_balance=3000.00,
            ),
        }
        result = reconcile_balances(total_balances, statement_balances)
        assert '差异金额中等' in result.records[0].diff_note
        assert '未达账项' in result.records[0].diff_note

    def test_filename_constant(self):
        """文件名常量正确"""
        assert BALANCE_RECONCILIATION_FILENAME == '期末余额与银行对账单核对报告.xlsx'
