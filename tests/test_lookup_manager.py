import os
import sys
import shutil
import tempfile

import pytest
import openpyxl

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import lookup_manager as lm


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='lookup_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


@pytest.fixture
def sample_lookup_file(tmp_dir):
    path = os.path.join(tmp_dir, '主体查找表.xlsx')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '主体映射'
    ws.cell(row=1, column=1, value='主体名称')
    ws.cell(row=1, column=2, value='银行账号')
    ws.cell(row=2, column=1, value='北京XX科技有限公司')
    ws.cell(row=2, column=2, value='01090312345678901')
    ws.cell(row=3, column=1, value='上海YY贸易有限公司')
    ws.cell(row=3, column=2, value='38812345678')
    ws.cell(row=4, column=1, value='广州ZZ实业公司')
    ws.cell(row=4, column=2, value='6222021234567890123')
    wb.save(path)
    wb.close()
    return path


class TestFindLookupFile:
    def test_find_exact_xlsx(self, tmp_dir):
        path = os.path.join(tmp_dir, '主体查找表.xlsx')
        openpyxl.Workbook().save(path)
        result = lm.find_lookup_file(tmp_dir)
        assert result == path

    def test_find_exact_xls(self, tmp_dir):
        path = os.path.join(tmp_dir, '主体查找表.xls')
        open(path, 'a').close()
        result = lm.find_lookup_file(tmp_dir)
        assert result == path

    def test_not_found(self, tmp_dir):
        result = lm.find_lookup_file(tmp_dir)
        assert result is None


class TestReadLookupEntries:
    def test_read_entries(self, sample_lookup_file):
        entries = lm.read_lookup_entries(sample_lookup_file)
        assert len(entries) == 3
        assert entries[0].subject == '北京XX科技有限公司'
        assert entries[0].account == '01090312345678901'
        assert entries[1].subject == '上海YY贸易有限公司'
        assert entries[1].account == '38812345678'

    def test_read_nonexistent_file(self):
        entries = lm.read_lookup_entries('/nonexistent/path.xlsx')
        assert len(entries) == 0

    def test_read_empty_file(self, tmp_dir):
        path = os.path.join(tmp_dir, 'empty.xlsx')
        wb = openpyxl.Workbook()
        wb.save(path)
        wb.close()
        entries = lm.read_lookup_entries(path)
        assert len(entries) == 0

    def test_read_with_header_only(self, tmp_dir):
        path = os.path.join(tmp_dir, 'header_only.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='主体名称')
        ws.cell(row=1, column=2, value='银行账号')
        wb.save(path)
        wb.close()
        entries = lm.read_lookup_entries(path)
        assert len(entries) == 0

    def test_read_normalizes_account(self, tmp_dir):
        path = os.path.join(tmp_dir, 'normalize_test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='主体名称')
        ws.cell(row=1, column=2, value='银行账号')
        ws.cell(row=2, column=1, value='测试公司')
        ws.cell(row=2, column=2, value=12345.0)
        ws.cell(row=3, column=1, value='测试公司2')
        ws.cell(row=3, column=2, value='  67890  ')
        wb.save(path)
        wb.close()
        entries = lm.read_lookup_entries(path)
        assert len(entries) == 2
        assert entries[0].account == '12345'
        assert entries[1].account == '67890'


class TestSaveLookupEntries:
    def test_save_entries(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test_save.xlsx')
        entries = [
            lm.LookupEntry(subject='公司A', account='111'),
            lm.LookupEntry(subject='公司B', account='222'),
        ]
        success, _ = lm.save_lookup_entries(entries, path)
        assert success is True
        assert os.path.exists(path)

        saved_entries = lm.read_lookup_entries(path)
        assert len(saved_entries) == 2
        assert saved_entries[0].subject == '公司A'
        assert saved_entries[0].account == '111'

    def test_save_overwrites_existing(self, sample_lookup_file):
        original_entries = lm.read_lookup_entries(sample_lookup_file)
        assert len(original_entries) == 3

        new_entries = [
            lm.LookupEntry(subject='新公司', account='999'),
        ]
        success, _ = lm.save_lookup_entries(new_entries, sample_lookup_file)
        assert success is True

        saved_entries = lm.read_lookup_entries(sample_lookup_file)
        assert len(saved_entries) == 1
        assert saved_entries[0].subject == '新公司'


class TestGetEntryByAccount:
    def test_get_found(self, sample_lookup_file):
        entry = lm.get_entry_by_account('01090312345678901', sample_lookup_file)
        assert entry is not None
        assert entry.subject == '北京XX科技有限公司'

    def test_get_not_found(self, sample_lookup_file):
        entry = lm.get_entry_by_account('999999999', sample_lookup_file)
        assert entry is None

    def test_get_with_normalization(self, sample_lookup_file):
        entry = lm.get_entry_by_account(1090312345678901, sample_lookup_file)
        assert entry is not None
        assert entry.subject == '北京XX科技有限公司'

    def test_get_with_leading_zero(self, sample_lookup_file):
        entry = lm.get_entry_by_account('001090312345678901', sample_lookup_file)
        assert entry is not None
        assert entry.subject == '北京XX科技有限公司'


class TestAddEntry:
    def test_add_success(self, sample_lookup_file):
        success, msg = lm.add_entry('新公司', '999999999', sample_lookup_file)
        assert success is True
        assert '添加成功' in msg

        entries = lm.read_lookup_entries(sample_lookup_file)
        assert len(entries) == 4
        assert any(e.account == '999999999' for e in entries)

    def test_add_duplicate(self, sample_lookup_file):
        success, msg = lm.add_entry('重复公司', '01090312345678901', sample_lookup_file)
        assert success is False
        assert '已存在' in msg

        entries = lm.read_lookup_entries(sample_lookup_file)
        assert len(entries) == 3

    def test_add_empty_subject(self, sample_lookup_file):
        success, msg = lm.add_entry('', '12345', sample_lookup_file)
        assert success is False
        assert '不能为空' in msg

    def test_add_empty_account(self, sample_lookup_file):
        success, msg = lm.add_entry('测试公司', '', sample_lookup_file)
        assert success is False
        assert '不能为空' in msg

    def test_add_normalizes_account(self, sample_lookup_file):
        success, msg = lm.add_entry('测试公司', '  88888  ', sample_lookup_file)
        assert success is True

        entry = lm.get_entry_by_account('88888', sample_lookup_file)
        assert entry is not None
        assert entry.account == '88888'


class TestUpdateEntry:
    def test_update_success(self, sample_lookup_file):
        success, msg = lm.update_entry(
            '01090312345678901',
            '更新后的公司名',
            '01090312345678901',
            sample_lookup_file
        )
        assert success is True
        assert '更新成功' in msg

        entry = lm.get_entry_by_account('01090312345678901', sample_lookup_file)
        assert entry.subject == '更新后的公司名'

    def test_update_account(self, sample_lookup_file):
        success, msg = lm.update_entry(
            '01090312345678901',
            '北京XX科技有限公司',
            '111111111111',
            sample_lookup_file
        )
        assert success is True

        old_entry = lm.get_entry_by_account('01090312345678901', sample_lookup_file)
        assert old_entry is None

        new_entry = lm.get_entry_by_account('111111111111', sample_lookup_file)
        assert new_entry is not None
        assert new_entry.subject == '北京XX科技有限公司'

    def test_update_not_found(self, sample_lookup_file):
        success, msg = lm.update_entry(
            '999999999',
            '测试公司',
            '999999999',
            sample_lookup_file
        )
        assert success is False
        assert '未找到' in msg

    def test_update_conflict(self, sample_lookup_file):
        success, msg = lm.update_entry(
            '01090312345678901',
            '测试公司',
            '38812345678',
            sample_lookup_file
        )
        assert success is False
        assert '已被使用' in msg


class TestDeleteEntry:
    def test_delete_success(self, sample_lookup_file):
        success, msg = lm.delete_entry('01090312345678901', sample_lookup_file)
        assert success is True
        assert '成功删除' in msg

        entries = lm.read_lookup_entries(sample_lookup_file)
        assert len(entries) == 2
        assert all(e.account != '01090312345678901' for e in entries)

    def test_delete_not_found(self, sample_lookup_file):
        success, msg = lm.delete_entry('999999999', sample_lookup_file)
        assert success is False
        assert '未找到' in msg

    def test_delete_with_normalization(self, sample_lookup_file):
        success, msg = lm.delete_entry(1090312345678901, sample_lookup_file)
        assert success is True

        entries = lm.read_lookup_entries(sample_lookup_file)
        assert len(entries) == 2


class TestSearchEntries:
    def test_search_by_subject(self, sample_lookup_file):
        results = lm.search_entries('北京', sample_lookup_file)
        assert len(results) == 1
        assert results[0].subject == '北京XX科技有限公司'

    def test_search_by_account(self, sample_lookup_file):
        results = lm.search_entries('38812', sample_lookup_file)
        assert len(results) == 1
        assert results[0].account == '38812345678'

    def test_search_empty_keyword(self, sample_lookup_file):
        results = lm.search_entries('', sample_lookup_file)
        assert len(results) == 3

    def test_search_no_results(self, sample_lookup_file):
        results = lm.search_entries('不存在的关键词', sample_lookup_file)
        assert len(results) == 0

    def test_search_case_insensitive(self, sample_lookup_file):
        results = lm.search_entries('beijing', sample_lookup_file)
        assert len(results) == 0

        results = lm.search_entries('北京', sample_lookup_file)
        assert len(results) == 1


class TestDuplicateEntries:
    def test_no_duplicates(self, sample_lookup_file):
        duplicates = lm.get_duplicate_entries(sample_lookup_file)
        assert len(duplicates) == 0

    def test_with_duplicates(self, tmp_dir):
        path = os.path.join(tmp_dir, 'dup.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='主体名称')
        ws.cell(row=1, column=2, value='银行账号')
        ws.cell(row=2, column=1, value='公司A')
        ws.cell(row=2, column=2, value='12345')
        ws.cell(row=3, column=1, value='公司B')
        ws.cell(row=3, column=2, value='12345')
        ws.cell(row=4, column=1, value='公司C')
        ws.cell(row=4, column=2, value='67890')
        wb.save(path)
        wb.close()

        duplicates = lm.get_duplicate_entries(path)
        assert len(duplicates) == 1
        assert duplicates[0]['account'] == '12345'
        assert duplicates[0]['count'] == 2
        assert '公司A' in duplicates[0]['subjects']
        assert '公司B' in duplicates[0]['subjects']


class TestImportExport:
    def test_export_excel(self, sample_lookup_file, tmp_dir):
        export_path = os.path.join(tmp_dir, 'export.xlsx')
        success, msg = lm.export_to_excel(export_path, sample_lookup_file)
        assert success is True
        assert os.path.exists(export_path)

        exported_entries = lm.read_lookup_entries(export_path)
        original_entries = lm.read_lookup_entries(sample_lookup_file)
        assert len(exported_entries) == len(original_entries)

    def test_import_merge(self, sample_lookup_file, tmp_dir):
        import_path = os.path.join(tmp_dir, 'import.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='主体名称')
        ws.cell(row=1, column=2, value='银行账号')
        ws.cell(row=2, column=1, value='新公司')
        ws.cell(row=2, column=2, value='99999')
        ws.cell(row=3, column=1, value='更新后的北京公司')
        ws.cell(row=3, column=2, value='01090312345678901')
        wb.save(import_path)
        wb.close()

        success, msg, stats = lm.import_from_excel(
            import_path, overwrite=False, lookup_file=sample_lookup_file
        )
        assert success is True
        assert stats['imported'] == 1
        assert stats['updated'] == 1
        assert stats['skipped'] == 0

        entries = lm.read_lookup_entries(sample_lookup_file)
        assert len(entries) == 4

        entry = lm.get_entry_by_account('01090312345678901', sample_lookup_file)
        assert entry.subject == '更新后的北京公司'

    def test_import_overwrite(self, sample_lookup_file, tmp_dir):
        import_path = os.path.join(tmp_dir, 'import.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='主体名称')
        ws.cell(row=1, column=2, value='银行账号')
        ws.cell(row=2, column=1, value='全新公司')
        ws.cell(row=2, column=2, value='11111')
        wb.save(import_path)
        wb.close()

        success, msg, stats = lm.import_from_excel(
            import_path, overwrite=True, lookup_file=sample_lookup_file
        )
        assert success is True
        assert stats['imported'] == 1

        entries = lm.read_lookup_entries(sample_lookup_file)
        assert len(entries) == 1
        assert entries[0].subject == '全新公司'

    def test_import_nonexistent_file(self):
        success, msg, stats = lm.import_from_excel('/nonexistent/file.xlsx')
        assert success is False
        assert '不存在' in msg


class TestAccountNormalization:
    def test_normalize_none(self):
        assert lm._normalize_account_str(None) == ''

    def test_normalize_int(self):
        assert lm._normalize_account_str(12345) == '12345'

    def test_normalize_float_whole(self):
        assert lm._normalize_account_str(12345.0) == '12345'

    def test_normalize_float_with_decimals(self):
        assert lm._normalize_account_str(123.45) == '123.45'

    def test_normalize_string_with_spaces(self):
        assert lm._normalize_account_str('  12345  ') == '12345'

    def test_normalize_string_dot_zero(self):
        assert lm._normalize_account_str('38812345678.0') == '38812345678'

    def test_account_key_leading_zero(self):
        assert lm._account_key('01090312345678901') == '1090312345678901'

    def test_account_key_int(self):
        assert lm._account_key(1090312345678901) == '1090312345678901'

    def test_account_key_equality(self):
        assert lm._account_key('01090312345678901') == lm._account_key(1090312345678901)


def _create_lookup_with_extra(path, entries=None):
    """创建带扩展字段和优先级的查找表"""
    if entries is None:
        entries = [
            {'subject': '北京XX科技有限公司', 'account': '01090312345678901',
             'priority': 5, 'extra_fields': {'部门': '技术部', '项目编号': 'PRJ001'}},
            {'subject': '上海YY贸易有限公司', 'account': '38812345678',
             'priority': 0, 'extra_fields': {'部门': '市场部', '项目编号': 'PRJ002'}},
        ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '主体映射'

    all_extra_keys = set()
    for entry in entries:
        if entry.get('extra_fields'):
            all_extra_keys.update(entry['extra_fields'].keys())
    sorted_extra_keys = sorted(all_extra_keys)

    has_priority = any(e.get('priority', 0) != 0 for e in entries)

    ws.cell(row=1, column=1, value='主体名称')
    ws.cell(row=1, column=2, value='银行账号')
    col_idx = 3
    if has_priority:
        ws.cell(row=1, column=col_idx, value='优先级')
        col_idx += 1
    for key in sorted_extra_keys:
        ws.cell(row=1, column=col_idx, value=key)
        col_idx += 1

    for i, entry in enumerate(entries, start=2):
        ws.cell(row=i, column=1, value=entry.get('subject', ''))
        ws.cell(row=i, column=2, value=entry.get('account', ''))
        col_idx = 3
        if has_priority:
            ws.cell(row=i, column=col_idx, value=entry.get('priority', 0))
            col_idx += 1
        for key in sorted_extra_keys:
            val = entry.get('extra_fields', {}).get(key, '')
            ws.cell(row=i, column=col_idx, value=val)
            col_idx += 1

    wb.save(path)
    wb.close()
    return path


class TestLookupEntry:
    """测试 LookupEntry 数据类"""

    def test_default_values(self):
        entry = lm.LookupEntry(subject='测试公司', account='12345')
        assert entry.subject == '测试公司'
        assert entry.account == '12345'
        assert entry.priority == 0
        assert entry.extra_fields == {}
        assert entry.row_id is None

    def test_with_priority(self):
        entry = lm.LookupEntry(subject='测试公司', account='12345', priority=10)
        assert entry.priority == 10

    def test_with_extra_fields(self):
        extra = {'部门': '技术部', '项目': 'PRJ001'}
        entry = lm.LookupEntry(subject='测试公司', account='12345', extra_fields=extra)
        assert entry.extra_fields == extra

    def test_none_extra_fields(self):
        entry = lm.LookupEntry(subject='测试公司', account='12345', extra_fields=None)
        assert entry.extra_fields == {}


class TestGetEntriesByAccount:
    """测试 get_entries_by_account - 返回所有匹配条目"""

    def test_single_match(self, sample_lookup_file):
        """单个匹配"""
        entries = lm.get_entries_by_account('01090312345678901', sample_lookup_file)
        assert len(entries) == 1
        assert entries[0].subject == '北京XX科技有限公司'

    def test_multiple_matches_with_priority(self, tmp_dir):
        """多个匹配，按优先级降序"""
        path = os.path.join(tmp_dir, 'multi.xlsx')
        _create_lookup_with_extra(path, [
            {'subject': '低优先级', 'account': '12345', 'priority': 1},
            {'subject': '高优先级', 'account': '12345', 'priority': 10},
            {'subject': '中优先级', 'account': '12345', 'priority': 5},
        ])
        entries = lm.get_entries_by_account('12345', path)
        assert len(entries) == 3
        assert entries[0].priority == 10
        assert entries[1].priority == 5
        assert entries[2].priority == 1

    def test_no_match(self, sample_lookup_file):
        """没有匹配"""
        entries = lm.get_entries_by_account('999999999', sample_lookup_file)
        assert len(entries) == 0

    def test_with_normalization(self, sample_lookup_file):
        """账号标准化"""
        entries = lm.get_entries_by_account(1090312345678901, sample_lookup_file)
        assert len(entries) == 1


class TestFuzzyMatchEntries:
    """测试 fuzzy_match_entries - 模糊匹配"""

    def test_exact_match(self, sample_lookup_file):
        """精确匹配"""
        results = lm.fuzzy_match_entries('01090312345678901', sample_lookup_file)
        assert len(results) >= 1
        assert results[0].subject == '北京XX科技有限公司'

    def test_fuzzy_close_match(self, sample_lookup_file):
        """接近的模糊匹配"""
        results = lm.fuzzy_match_entries('0109031234567890', sample_lookup_file, threshold=0.8)
        assert len(results) >= 1
        assert results[0].subject == '北京XX科技有限公司'

    def test_fuzzy_no_match(self, sample_lookup_file):
        """不匹配"""
        results = lm.fuzzy_match_entries('999999999999', sample_lookup_file, threshold=0.9)
        assert len(results) == 0

    def test_fuzzy_threshold(self, sample_lookup_file):
        """阈值过滤"""
        results_high = lm.fuzzy_match_entries('0109031234567890', sample_lookup_file, threshold=0.99)
        results_low = lm.fuzzy_match_entries('0109031234567890', sample_lookup_file, threshold=0.5)
        assert len(results_high) <= len(results_low)

    def test_fuzzy_sorted_by_similarity(self, tmp_dir):
        """按相似度降序排列"""
        path = os.path.join(tmp_dir, 'fuzzy.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='主体名称')
        ws.cell(row=1, column=2, value='银行账号')
        ws.cell(row=2, column=1, value='公司A')
        ws.cell(row=2, column=2, value='1234567890')
        ws.cell(row=3, column=1, value='公司B')
        ws.cell(row=3, column=2, value='1234500000')
        wb.save(path)
        wb.close()

        results = lm.fuzzy_match_entries('1234567899', path, threshold=0.5)
        assert len(results) == 2
        assert results[0].subject == '公司A'

    def test_empty_account(self, sample_lookup_file):
        """空账号"""
        results = lm.fuzzy_match_entries('', sample_lookup_file)
        assert len(results) == 0

    def test_none_account(self, sample_lookup_file):
        """None 账号"""
        results = lm.fuzzy_match_entries(None, sample_lookup_file)
        assert len(results) == 0


class TestGetSubjectInfo:
    """测试 get_subject_info - 获取完整主体信息"""

    def test_exact_match_info(self, sample_lookup_file):
        """精确匹配的信息"""
        info = lm.get_subject_info('01090312345678901', sample_lookup_file)
        assert info['matched'] is True
        assert info['fuzzy_matched'] is False
        assert info['subject'] == '北京XX科技有限公司'
        assert info['account'] == '01090312345678901'
        assert info['priority'] == 0
        assert info['extra_fields'] == {}
        assert info['similarity'] == 1.0

    def test_fuzzy_match_info(self, sample_lookup_file):
        """模糊匹配的信息"""
        info = lm.get_subject_info('0109031234567890', sample_lookup_file, use_fuzzy=True)
        assert info['matched'] is True
        assert info['fuzzy_matched'] is True
        assert info['subject'] == '北京XX科技有限公司'
        assert info['similarity'] < 1.0
        assert info['similarity'] > 0

    def test_no_match_info(self, sample_lookup_file):
        """未匹配的信息"""
        info = lm.get_subject_info('999999999999', sample_lookup_file)
        assert info['matched'] is False
        assert info['subject'] == ''
        assert info['account'] == '999999999999'
        assert info['priority'] == 0
        assert info['extra_fields'] == {}
        assert info['similarity'] == 0.0

    def test_with_extra_fields(self, tmp_dir):
        """带扩展字段的信息"""
        path = os.path.join(tmp_dir, 'extra.xlsx')
        _create_lookup_with_extra(path, [
            {'subject': '测试公司', 'account': '12345', 'priority': 5,
             'extra_fields': {'部门': '技术部', '项目编号': 'PRJ001'}},
        ])
        info = lm.get_subject_info('12345', path)
        assert info['matched'] is True
        assert info['priority'] == 5
        assert info['extra_fields']['部门'] == '技术部'
        assert info['extra_fields']['项目编号'] == 'PRJ001'

    def test_multiple_subjects_returns_highest_priority(self, tmp_dir):
        """多主体时返回最高优先级的"""
        path = os.path.join(tmp_dir, 'multi_prio.xlsx')
        _create_lookup_with_extra(path, [
            {'subject': '主体A', 'account': '12345', 'priority': 1},
            {'subject': '主体B', 'account': '12345', 'priority': 10},
        ])
        info = lm.get_subject_info('12345', path)
        assert info['subject'] == '主体B'
        assert info['priority'] == 10

    def test_none_lookup_file(self):
        """None 查找表"""
        info = lm.get_subject_info('12345', None)
        assert info['matched'] is False

    def test_nonexistent_lookup_file(self):
        """不存在的查找表"""
        info = lm.get_subject_info('12345', '/nonexistent/path.xlsx')
        assert info['matched'] is False


class TestExtraFieldsRead:
    """测试读取扩展字段"""

    def test_read_extra_fields(self, tmp_dir):
        """读取扩展字段"""
        path = os.path.join(tmp_dir, 'extra.xlsx')
        _create_lookup_with_extra(path, [
            {'subject': '北京XX科技有限公司', 'account': '01090312345678901',
             'priority': 0, 'extra_fields': {'部门': '技术部', '项目编号': 'PRJ001'}},
        ])
        entries = lm.read_lookup_entries(path)
        assert len(entries) == 1
        assert entries[0].extra_fields['部门'] == '技术部'
        assert entries[0].extra_fields['项目编号'] == 'PRJ001'

    def test_read_with_priority(self, tmp_dir):
        """读取优先级"""
        path = os.path.join(tmp_dir, 'prio.xlsx')
        _create_lookup_with_extra(path, [
            {'subject': '测试公司', 'account': '12345', 'priority': 10,
             'extra_fields': {}},
        ])
        entries = lm.read_lookup_entries(path)
        assert len(entries) == 1
        assert entries[0].priority == 10

    def test_read_no_extra_fields(self, sample_lookup_file):
        """没有扩展字段时"""
        entries = lm.read_lookup_entries(sample_lookup_file)
        assert all(len(e.extra_fields) == 0 for e in entries)

    def test_read_partial_extra_fields(self, tmp_dir):
        """部分行有扩展字段"""
        path = os.path.join(tmp_dir, 'partial.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='主体名称')
        ws.cell(row=1, column=2, value='银行账号')
        ws.cell(row=1, column=3, value='部门')
        ws.cell(row=2, column=1, value='公司A')
        ws.cell(row=2, column=2, value='111')
        ws.cell(row=2, column=3, value='技术部')
        ws.cell(row=3, column=1, value='公司B')
        ws.cell(row=3, column=2, value='222')
        wb.save(path)
        wb.close()

        entries = lm.read_lookup_entries(path)
        assert len(entries) == 2
        assert entries[0].extra_fields['部门'] == '技术部'
        assert entries[1].extra_fields.get('部门', '') == ''


class TestExtraFieldsSave:
    """测试保存扩展字段"""

    def test_save_with_extra_fields(self, tmp_dir):
        """保存带扩展字段的条目"""
        path = os.path.join(tmp_dir, 'save_extra.xlsx')
        entries = [
            lm.LookupEntry(subject='公司A', account='111', priority=5,
                          extra_fields={'部门': '技术部', '项目': 'PRJ001'}),
            lm.LookupEntry(subject='公司B', account='222',
                          extra_fields={'部门': '市场部'}),
        ]
        success, _ = lm.save_lookup_entries(entries, path)
        assert success is True

        saved = lm.read_lookup_entries(path)
        assert len(saved) == 2
        assert saved[0].priority == 5
        assert saved[0].extra_fields['部门'] == '技术部'
        assert saved[0].extra_fields['项目'] == 'PRJ001'
        assert saved[1].extra_fields['部门'] == '市场部'

    def test_save_roundtrip_preserves_data(self, tmp_dir):
        """保存后重新读取，数据一致"""
        path = os.path.join(tmp_dir, 'roundtrip.xlsx')
        original = [
            lm.LookupEntry(subject='公司A', account='111', priority=10,
                          extra_fields={'部门': '技术部', '项目编号': 'PRJ001', '负责人': '张三'}),
        ]
        lm.save_lookup_entries(original, path)
        loaded = lm.read_lookup_entries(path)

        assert len(loaded) == 1
        assert loaded[0].subject == original[0].subject
        assert loaded[0].account == original[0].account
        assert loaded[0].priority == original[0].priority
        assert loaded[0].extra_fields == original[0].extra_fields


class TestSearchEntriesWithExtra:
    """测试搜索扩展字段"""

    def test_search_in_extra_fields(self, tmp_dir):
        """搜索扩展字段中的内容"""
        path = os.path.join(tmp_dir, 'search_extra.xlsx')
        _create_lookup_with_extra(path, [
            {'subject': '公司A', 'account': '111', 'priority': 0,
             'extra_fields': {'部门': '技术部', '项目编号': 'PRJ001'}},
            {'subject': '公司B', 'account': '222', 'priority': 0,
             'extra_fields': {'部门': '市场部', '项目编号': 'PRJ002'}},
        ])
        results = lm.search_entries('技术部', path)
        assert len(results) == 1
        assert results[0].subject == '公司A'

    def test_search_project_number(self, tmp_dir):
        """搜索项目编号"""
        path = os.path.join(tmp_dir, 'search_prj.xlsx')
        _create_lookup_with_extra(path, [
            {'subject': '公司A', 'account': '111', 'priority': 0,
             'extra_fields': {'项目编号': 'PRJ001'}},
            {'subject': '公司B', 'account': '222', 'priority': 0,
             'extra_fields': {'项目编号': 'PRJ002'}},
        ])
        results = lm.search_entries('PRJ001', path)
        assert len(results) == 1
        assert results[0].subject == '公司A'

    def test_search_still_finds_subject(self, tmp_dir):
        """仍然可以搜索主体名称"""
        path = os.path.join(tmp_dir, 'search_subject.xlsx')
        _create_lookup_with_extra(path, [
            {'subject': '北京科技公司', 'account': '111', 'priority': 0,
             'extra_fields': {'部门': '技术部'}},
        ])
        results = lm.search_entries('北京', path)
        assert len(results) == 1


class TestImportExportWithExtra:
    """测试导入导出扩展字段和优先级"""

    def test_export_preserves_extra(self, tmp_dir):
        """导出时保留扩展字段"""
        src_path = os.path.join(tmp_dir, 'src.xlsx')
        _create_lookup_with_extra(src_path, [
            {'subject': '公司A', 'account': '111', 'priority': 5,
             'extra_fields': {'部门': '技术部', '项目': 'PRJ001'}},
        ])
        export_path = os.path.join(tmp_dir, 'export.xlsx')
        success, msg = lm.export_to_excel(export_path, src_path)
        assert success is True

        exported = lm.read_lookup_entries(export_path)
        assert len(exported) == 1
        assert exported[0].priority == 5
        assert exported[0].extra_fields['部门'] == '技术部'
        assert exported[0].extra_fields['项目'] == 'PRJ001'

    def test_import_with_extra_fields(self, tmp_dir):
        """导入带扩展字段的文件"""
        import_path = os.path.join(tmp_dir, 'import.xlsx')
        _create_lookup_with_extra(import_path, [
            {'subject': '新公司', 'account': '999', 'priority': 3,
             'extra_fields': {'部门': '财务部', '项目编号': 'PRJ999'}},
        ])

        dest_path = os.path.join(tmp_dir, 'dest.xlsx')
        success, msg, stats = lm.import_from_excel(
            import_path, overwrite=True, lookup_file=dest_path
        )
        assert success is True
        assert stats['imported'] == 1

        entries = lm.read_lookup_entries(dest_path)
        assert len(entries) == 1
        assert entries[0].subject == '新公司'
        assert entries[0].priority == 3
        assert entries[0].extra_fields['部门'] == '财务部'
        assert entries[0].extra_fields['项目编号'] == 'PRJ999'

    def test_import_merge_updates_extra(self, tmp_dir):
        """导入合并时更新扩展字段"""
        base_path = os.path.join(tmp_dir, 'base.xlsx')
        _create_lookup_with_extra(base_path, [
            {'subject': '公司A', 'account': '111', 'priority': 1,
             'extra_fields': {'部门': '旧部门'}},
        ])

        import_path = os.path.join(tmp_dir, 'import.xlsx')
        _create_lookup_with_extra(import_path, [
            {'subject': '公司A', 'account': '111', 'priority': 10,
             'extra_fields': {'部门': '新部门', '项目': '新项目'}},
        ])

        success, msg, stats = lm.import_from_excel(
            import_path, overwrite=False, lookup_file=base_path
        )
        assert success is True
        assert stats['updated'] == 1

        entries = lm.read_lookup_entries(base_path)
        assert len(entries) == 1
        assert entries[0].priority == 10
        assert entries[0].extra_fields['部门'] == '新部门'
        assert entries[0].extra_fields['项目'] == '新项目'

