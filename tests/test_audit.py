import os
import sys
import shutil
import tempfile
import sqlite3
import json

import openpyxl
import pandas as pd
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck
from conftest import _create_beijing_bank_excel, _create_lookup_table


@pytest.fixture
def audit_script_dir(tmp_dir):
    """创建带审计数据库的脚本目录"""
    script_dir = os.path.join(tmp_dir, 'script_audit')
    os.makedirs(script_dir, exist_ok=True)
    _create_lookup_table(os.path.join(script_dir, '主体查找表.xlsx'))
    db_path = bankcheck.get_audit_db_path(script_dir)
    bankcheck.init_audit_db(db_path)
    return script_dir


class TestAuditDBInit:
    def test_init_audit_db_creates_tables(self, audit_script_dir):
        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
        tables = {row[0] for row in cursor.fetchall()}

        assert 'users' in tables
        assert 'audit_logs' in tables
        assert 'config_changes' in tables

        cursor.execute("PRAGMA table_info(users)")
        user_cols = {row[1] for row in cursor.fetchall()}
        assert 'username' in user_cols
        assert 'role' in user_cols
        assert 'created_at' in user_cols

        cursor.execute("PRAGMA table_info(audit_logs)")
        audit_cols = {row[1] for row in cursor.fetchall()}
        assert 'audit_id' in audit_cols
        assert 'username' in audit_cols
        assert 'operation_type' in audit_cols
        assert 'input_directory' in audit_cols
        assert 'output_path' in audit_cols
        assert 'status' in audit_cols
        assert 'started_at' in audit_cols

        cursor.execute("PRAGMA table_info(config_changes)")
        config_cols = {row[1] for row in cursor.fetchall()}
        assert 'change_id' in config_cols
        assert 'config_type' in config_cols
        assert 'old_value' in config_cols
        assert 'new_value' in config_cols

        conn.close()

    def test_init_audit_db_creates_indexes(self, audit_script_dir):
        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='index'")
        indexes = {row[0] for row in cursor.fetchall()}
        assert 'idx_audit_logs_username' in indexes
        assert 'idx_audit_logs_operation' in indexes
        assert 'idx_audit_logs_started_at' in indexes
        conn.close()

    def test_init_audit_db_idempotent(self, audit_script_dir):
        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        bankcheck.init_audit_db(db_path)
        bankcheck.init_audit_db(db_path)
        bankcheck.init_audit_db(db_path)


class TestGetCurrentUser:
    def test_get_current_user_returns_string(self):
        user = bankcheck.get_current_user()
        assert isinstance(user, str)
        assert len(user) > 0

    def test_get_current_user_uses_env_variable(self, monkeypatch):
        monkeypatch.setenv('BANKCHECK_USER', 'test_user_123')
        assert bankcheck.get_current_user() == 'test_user_123'
        monkeypatch.delenv('BANKCHECK_USER', raising=False)

    def test_get_current_user_env_empty_falls_back(self, monkeypatch):
        monkeypatch.setenv('BANKCHECK_USER', '')
        user = bankcheck.get_current_user()
        assert user != ''
        monkeypatch.delenv('BANKCHECK_USER', raising=False)


class TestEnsureUser:
    def test_ensure_user_creates_new_user(self, audit_script_dir):
        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        user_id = bankcheck._ensure_user('new_test_user', db_path)

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT username, is_active FROM users WHERE id = ?', (user_id,))
        row = cursor.fetchone()
        conn.close()

        assert row is not None
        assert row[0] == 'new_test_user'
        assert row[1] == 1

    def test_ensure_user_returns_existing_user_id(self, audit_script_dir):
        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        user_id1 = bankcheck._ensure_user('repeat_user', db_path)
        user_id2 = bankcheck._ensure_user('repeat_user', db_path)

        assert user_id1 == user_id2

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM users WHERE username = ?', ('repeat_user',))
        count = cursor.fetchone()[0]
        conn.close()

        assert count == 1

    def test_ensure_user_updates_last_login(self, audit_script_dir):
        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        user_id = bankcheck._ensure_user('login_test_user', db_path)

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT last_login FROM users WHERE id = ?', (user_id,))
        first_login = cursor.fetchone()[0]
        conn.close()

        import time
        time.sleep(1.1)

        user_id2 = bankcheck._ensure_user('login_test_user', db_path)
        assert user_id == user_id2

        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT last_login FROM users WHERE id = ?', (user_id,))
        second_login = cursor.fetchone()[0]
        conn.close()

        assert first_login != second_login


class TestComputeFileHash:
    def test_compute_file_hash_returns_sha256(self, tmp_dir):
        test_file = os.path.join(tmp_dir, 'test_hash.xlsx')
        wb = openpyxl.Workbook()
        wb.active['A1'] = 'test content'
        wb.save(test_file)
        wb.close()

        h = bankcheck.compute_file_hash(test_file)
        assert h is not None
        assert isinstance(h, str)
        assert len(h) == 64

        h2 = bankcheck.compute_file_hash(test_file)
        assert h == h2

    def test_compute_file_hash_different_content_different_hash(self, tmp_dir):
        f1 = os.path.join(tmp_dir, 'f1.xlsx')
        f2 = os.path.join(tmp_dir, 'f2.xlsx')

        wb = openpyxl.Workbook()
        wb.active['A1'] = 'content1'
        wb.save(f1)
        wb.close()

        wb = openpyxl.Workbook()
        wb.active['A1'] = 'content2'
        wb.save(f2)
        wb.close()

        assert bankcheck.compute_file_hash(f1) != bankcheck.compute_file_hash(f2)

    def test_compute_file_hash_nonexistent_returns_none(self, tmp_dir):
        h = bankcheck.compute_file_hash(os.path.join(tmp_dir, 'nonexistent.xlsx'))
        assert h is None


class TestComputeConfigSnapshot:
    def test_compute_config_snapshot_returns_json(self, audit_script_dir):
        snapshot = bankcheck.compute_config_snapshot(audit_script_dir)
        assert isinstance(snapshot, str)
        data = json.loads(snapshot)
        assert 'supported_banks' in data
        assert 'bank_processors' in data
        assert 'lookup_file' in data
        assert 'lookup_file_hash' in data
        assert 'timestamp' in data
        assert '北京银行' in data['supported_banks']
        assert '东亚银行' in data['supported_banks']


class TestAuditLogger:
    def test_audit_logger_creates_record_on_init(self, audit_script_dir):
        with bankcheck.AuditLogger('pipeline', audit_script_dir, username='audit_test_user') as audit:
            assert audit.audit_id.startswith('AUD')
            assert audit.operation_type == 'pipeline'
            assert audit.username == 'audit_test_user'
            assert audit.record.status == 'running'
            assert audit.record.started_at is not None

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT status, operation_type, username FROM audit_logs WHERE audit_id = ?',
                       (audit.audit_id,))
        row = cursor.fetchone()
        conn.close()

        assert row is not None
        assert row[0] == 'success'
        assert row[1] == 'pipeline'
        assert row[2] == 'audit_test_user'

    def test_audit_logger_records_input(self, audit_script_dir):
        test_folder = '/test/input/folder'
        with bankcheck.AuditLogger('pipeline', audit_script_dir, username='input_test') as audit:
            audit.record_input(test_folder)

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT input_directory FROM audit_logs WHERE audit_id = ?',
                       (audit.audit_id,))
        row = cursor.fetchone()
        conn.close()

        assert row[0] == test_folder

    def test_audit_logger_records_processing_result(self, tmp_dir, audit_script_dir):
        source_folder = os.path.join(tmp_dir, '流水')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))

        with bankcheck.AuditLogger('pipeline', audit_script_dir, username='result_test') as audit:
            audit.record_input(source_folder)
            result = bankcheck.run_pipeline(source_folder, audit_script_dir)
            audit.record_result(result)

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        cursor.execute('SELECT * FROM audit_logs WHERE audit_id = ?', (audit.audit_id,))
        row = cursor.fetchone()
        conn.close()

        assert row['processed_files'] == 1
        assert row['extracted_records'] == 2
        assert row['unprocessed_files'] == 0
        assert row['error_files'] == 0
        assert row['status'] == 'success'
        assert row['output_path'] is not None
        assert row['duration_ms'] is not None
        assert row['duration_ms'] >= 0
        assert row['lookup_missing'] == 0

    def test_audit_logger_records_failure(self, audit_script_dir):
        try:
            with bankcheck.AuditLogger('pipeline', audit_script_dir, username='fail_test') as audit:
                raise ValueError('test error')
        except ValueError:
            pass

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT status, error_message FROM audit_logs WHERE username = ?',
                       ('fail_test',))
        row = cursor.fetchone()
        conn.close()

        assert row[0] == 'failed'
        assert 'test error' in row[1]
        assert 'ValueError' in row[1]

    def test_audit_logger_records_diff_result(self, tmp_dir, audit_script_dir):
        old_df = pd.DataFrame([{
            '唯一id': 'id1', '银行': '北京银行', '银行账号': '123', '主体': '测试',
            '交易日期': '2024-01-01', '付款': -100, '收款': None,
            '摘要': 'test', '对方户名': 'test', '余额': 900, '交易流水号': 't1'
        }])
        new_df = pd.DataFrame([{
            '唯一id': 'id1', '银行': '北京银行', '银行账号': '123', '主体': '测试',
            '交易日期': '2024-01-01', '付款': -100, '收款': None,
            '摘要': 'updated', '对方户名': 'test', '余额': 900, '交易流水号': 't1'
        }])

        old_path = os.path.join(tmp_dir, 'old.xlsx')
        new_path = os.path.join(tmp_dir, 'new.xlsx')
        old_df.to_excel(old_path, index=False, engine='openpyxl')
        new_df.to_excel(new_path, index=False, engine='openpyxl')

        with bankcheck.AuditLogger('diff', audit_script_dir, username='diff_test') as audit:
            audit.record_input(f'old:{old_path} | new:{new_path}')
            diff_result = bankcheck.run_diff(old_path, new_path, tmp_dir)
            audit.record_result(diff_result)

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT operation_type, extracted_records, output_path FROM audit_logs WHERE audit_id = ?',
                       (audit.audit_id,))
        row = cursor.fetchone()
        conn.close()

        assert row[0] == 'diff'
        assert row[1] == 1
        assert row[2] is not None

    def test_audit_logger_records_config_snapshot(self, audit_script_dir):
        with bankcheck.AuditLogger('pipeline', audit_script_dir, username='snapshot_test') as audit:
            pass

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT config_snapshot FROM audit_logs WHERE audit_id = ?',
                       (audit.audit_id,))
        snapshot_json = cursor.fetchone()[0]
        conn.close()

        assert snapshot_json is not None
        snapshot = json.loads(snapshot_json)
        assert 'supported_banks' in snapshot
        assert 'lookup_file_hash' in snapshot

    def test_audit_logger_computes_output_hash(self, tmp_dir, audit_script_dir):
        source_folder = os.path.join(tmp_dir, '流水')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))

        with bankcheck.AuditLogger('pipeline', audit_script_dir, username='hash_test') as audit:
            audit.record_input(source_folder)
            result = bankcheck.run_pipeline(source_folder, audit_script_dir)
            audit.record_result(result)

        expected_hash = bankcheck.compute_file_hash(result.output_path)

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT output_hash FROM audit_logs WHERE audit_id = ?',
                       (audit.audit_id,))
        row = cursor.fetchone()
        conn.close()

        assert row[0] == expected_hash
        assert len(row[0]) == 64


class TestRecordConfigChange:
    def test_record_config_change_creates_record(self, audit_script_dir):
        change_id = bankcheck.record_config_change(
            'lookup_table', '主体查找表.xlsx',
            'old_mapping', 'new_mapping',
            change_reason='添加新主体',
            script_dir=audit_script_dir,
            username='config_user'
        )

        assert change_id.startswith('CFG')

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('''
            SELECT username, config_type, config_name, old_value, new_value, change_reason
            FROM config_changes WHERE change_id = ?
        ''', (change_id,))
        row = cursor.fetchone()
        conn.close()

        assert row is not None
        assert row[0] == 'config_user'
        assert row[1] == 'lookup_table'
        assert row[2] == '主体查找表.xlsx'
        assert row[3] == 'old_mapping'
        assert row[4] == 'new_mapping'
        assert row[5] == '添加新主体'

    def test_record_config_change_dict_values(self, audit_script_dir):
        old_dict = {'account1': '主体A'}
        new_dict = {'account1': '主体A', 'account2': '主体B'}

        change_id = bankcheck.record_config_change(
            'lookup_table', 'mappings',
            old_dict, new_dict,
            script_dir=audit_script_dir,
            username='dict_user'
        )

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT old_value, new_value, old_hash, new_hash FROM config_changes WHERE change_id = ?',
                       (change_id,))
        row = cursor.fetchone()
        conn.close()

        assert '主体A' in row[0]
        assert '主体B' in row[1]
        assert row[2] != row[3]
        assert len(row[2]) == 64


class TestQueryAuditLogs:
    def test_query_audit_logs_by_username(self, audit_script_dir):
        for i in range(3):
            with bankcheck.AuditLogger('pipeline', audit_script_dir, username=f'user_{i}'):
                pass

        logs = bankcheck.query_audit_logs(script_dir=audit_script_dir, username='user_1')
        assert len(logs) == 1
        assert logs[0]['username'] == 'user_1'

    def test_query_audit_logs_by_operation_type(self, audit_script_dir):
        with bankcheck.AuditLogger('pipeline', audit_script_dir, username='op_user'):
            pass
        with bankcheck.AuditLogger('diff', audit_script_dir, username='op_user'):
            pass

        logs = bankcheck.query_audit_logs(script_dir=audit_script_dir, operation_type='diff')
        assert len(logs) >= 1
        assert all(log['operation_type'] == 'diff' for log in logs)

    def test_query_audit_logs_by_status(self, audit_script_dir):
        with bankcheck.AuditLogger('pipeline', audit_script_dir, username='status_user'):
            pass

        logs = bankcheck.query_audit_logs(script_dir=audit_script_dir, status='success')
        assert len(logs) >= 1
        assert all(log['status'] == 'success' for log in logs)

    def test_query_audit_logs_limit(self, audit_script_dir):
        for i in range(10):
            with bankcheck.AuditLogger('pipeline', audit_script_dir, username=f'limit_user_{i}'):
                pass

        logs = bankcheck.query_audit_logs(script_dir=audit_script_dir, limit=3)
        assert len(logs) == 3

    def test_query_audit_logs_date_range(self, audit_script_dir):
        from datetime import datetime, timedelta

        with bankcheck.AuditLogger('pipeline', audit_script_dir, username='date_user'):
            pass

        today = datetime.now().strftime('%Y-%m-%d')
        tomorrow = (datetime.now() + timedelta(days=1)).strftime('%Y-%m-%d')

        logs = bankcheck.query_audit_logs(
            script_dir=audit_script_dir,
            start_date=today,
            end_date=tomorrow
        )
        assert len(logs) >= 1

    def test_query_audit_logs_empty_db(self, tmp_dir):
        empty_script_dir = os.path.join(tmp_dir, 'empty_script')
        os.makedirs(empty_script_dir, exist_ok=True)

        logs = bankcheck.query_audit_logs(script_dir=empty_script_dir)
        assert logs == []


class TestQueryConfigChanges:
    def test_query_config_changes_by_type(self, audit_script_dir):
        bankcheck.record_config_change('lookup_table', 't1', 'old', 'new',
                                       script_dir=audit_script_dir, username='q1')
        bankcheck.record_config_change('bank_config', 't2', 'old', 'new',
                                       script_dir=audit_script_dir, username='q1')

        changes = bankcheck.query_config_changes(script_dir=audit_script_dir, config_type='lookup_table')
        assert len(changes) >= 1
        assert all(c['config_type'] == 'lookup_table' for c in changes)

    def test_query_config_changes_by_username(self, audit_script_dir):
        bankcheck.record_config_change('lookup_table', 't1', 'old', 'new',
                                       script_dir=audit_script_dir, username='specific_user')

        changes = bankcheck.query_config_changes(script_dir=audit_script_dir, username='specific_user')
        assert len(changes) == 1
        assert changes[0]['username'] == 'specific_user'


class TestExportAuditLogs:
    def test_export_audit_logs_creates_excel(self, tmp_dir, audit_script_dir):
        with bankcheck.AuditLogger('pipeline', audit_script_dir, username='export_user') as audit:
            pass

        output_path = os.path.join(tmp_dir, 'audit_export.xlsx')
        result = bankcheck.export_audit_logs(output_path, script_dir=audit_script_dir)

        assert result == output_path
        assert os.path.exists(output_path)

        df = pd.read_excel(output_path, engine='openpyxl')
        assert len(df) >= 1
        assert 'audit_id' in df.columns
        assert 'username' in df.columns
        assert 'operation_type' in df.columns
        assert 'status' in df.columns
        assert 'input_directory' in df.columns
        assert 'output_path' in df.columns
        assert 'started_at' in df.columns
        assert 'duration_ms' in df.columns


class TestExportConfigChanges:
    def test_export_config_changes_creates_excel(self, tmp_dir, audit_script_dir):
        bankcheck.record_config_change('lookup_table', 'test.xlsx', 'old', 'new',
                                       change_reason='test',
                                       script_dir=audit_script_dir, username='export_cfg')

        output_path = os.path.join(tmp_dir, 'config_export.xlsx')
        result = bankcheck.export_config_changes(output_path, script_dir=audit_script_dir)

        assert result == output_path
        assert os.path.exists(output_path)

        df = pd.read_excel(output_path, engine='openpyxl')
        assert len(df) >= 1
        assert 'change_id' in df.columns
        assert 'username' in df.columns
        assert 'config_type' in df.columns
        assert 'old_value' in df.columns
        assert 'new_value' in df.columns
        assert 'change_reason' in df.columns


class TestIntegrationWithPipeline:
    def test_full_pipeline_creates_audit_record(self, tmp_dir, audit_script_dir):
        source_folder = os.path.join(tmp_dir, '流水')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))

        result = bankcheck.run_pipeline(source_folder, audit_script_dir)

        with bankcheck.AuditLogger('pipeline', audit_script_dir, username='integration_user') as audit:
            audit.record_input(source_folder)
            audit.record_result(result)

        logs = bankcheck.query_audit_logs(script_dir=audit_script_dir, username='integration_user')
        assert len(logs) == 1
        log = logs[0]
        assert log['processed_files'] == 1
        assert log['extracted_records'] == 2
        assert log['status'] == 'success'
        assert log['output_hash'] is not None
        assert log['config_snapshot'] is not None


class TestGetClientInfo:
    def test_get_client_info_returns_dict(self):
        info = bankcheck.get_client_info()
        assert isinstance(info, dict)
        assert 'hostname' in info
        assert 'ip' in info


class TestReadLookupTableContent:
    def test_read_lookup_table_content_returns_structured_data(self, audit_script_dir):
        lookup_file = os.path.join(audit_script_dir, '主体查找表.xlsx')
        content = bankcheck.read_lookup_table_content(lookup_file)

        assert content is not None
        assert 'file_path' in content
        assert 'file_hash' in content
        assert 'headers' in content
        assert 'rows' in content
        assert 'account_map' in content
        assert content['file_path'] == lookup_file
        assert len(content['file_hash']) == 64
        assert '主体名称' in content['headers']
        assert '银行账号' in content['headers']
        assert len(content['rows']) >= 2
        assert len(content['account_map']) >= 2

    def test_read_lookup_table_content_nonexistent_file(self, audit_script_dir):
        content = bankcheck.read_lookup_table_content('/nonexistent/file.xlsx')
        assert content is None

    def test_read_lookup_table_content_none(self, audit_script_dir):
        content = bankcheck.read_lookup_table_content(None)
        assert content is None

    def test_read_lookup_table_account_mapping(self, audit_script_dir):
        lookup_file = os.path.join(audit_script_dir, '主体查找表.xlsx')
        content = bankcheck.read_lookup_table_content(lookup_file)

        account_key = bankcheck._account_key('01090312345678901')
        assert account_key in content['account_map']
        row = content['account_map'][account_key]
        assert row['主体名称'] == '北京XX科技有限公司'


class TestLookupSnapshots:
    def test_save_lookup_snapshot_creates_record(self, audit_script_dir):
        lookup_file = os.path.join(audit_script_dir, '主体查找表.xlsx')
        content = bankcheck.read_lookup_table_content(lookup_file)

        snapshot_id = bankcheck._save_lookup_snapshot(content, audit_script_dir, username='snapshot_user')

        assert snapshot_id.startswith('SNP')

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT snapshot_id, lookup_file, created_by FROM lookup_snapshots WHERE snapshot_id = ?',
                       (snapshot_id,))
        row = cursor.fetchone()
        conn.close()

        assert row is not None
        assert row[0] == snapshot_id
        assert row[1] == lookup_file
        assert row[2] == 'snapshot_user'

    def test_get_last_lookup_snapshot_returns_latest(self, audit_script_dir):
        lookup_file = os.path.join(audit_script_dir, '主体查找表.xlsx')
        content = bankcheck.read_lookup_table_content(lookup_file)

        bankcheck._save_lookup_snapshot(content, audit_script_dir, username='user1')
        import time
        time.sleep(0.1)
        bankcheck._save_lookup_snapshot(content, audit_script_dir, username='user2')

        snapshot = bankcheck._get_last_lookup_snapshot(lookup_file, audit_script_dir)

        assert snapshot is not None
        assert snapshot['created_by'] == 'user2'

    def test_get_last_lookup_snapshot_no_snapshot(self, tmp_dir):
        empty_script_dir = os.path.join(tmp_dir, 'empty')
        os.makedirs(empty_script_dir, exist_ok=True)
        snapshot = bankcheck._get_last_lookup_snapshot('/some/file.xlsx', empty_script_dir)
        assert snapshot is None


class TestDiffLookupSnapshots:
    def test_diff_detects_add(self):
        old_content = {
            'rows': [{'row_num': 2, '银行账号': '111', '主体名称': '公司A'}],
            'account_map': {'111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A'}}
        }
        new_content = {
            'rows': [
                {'row_num': 2, '银行账号': '111', '主体名称': '公司A'},
                {'row_num': 3, '银行账号': '222', '主体名称': '公司B'}
            ],
            'account_map': {
                '111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A'},
                '222': {'row_num': 3, '银行账号': '222', '主体名称': '公司B'}
            }
        }

        changes = bankcheck._diff_lookup_snapshots(old_content, new_content)

        assert len(changes) == 1
        assert changes[0]['change_type'] == 'add'
        assert changes[0]['account'] == '222'
        assert '公司B' in changes[0]['description']

    def test_diff_detects_remove(self):
        old_content = {
            'rows': [
                {'row_num': 2, '银行账号': '111', '主体名称': '公司A'},
                {'row_num': 3, '银行账号': '222', '主体名称': '公司B'}
            ],
            'account_map': {
                '111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A'},
                '222': {'row_num': 3, '银行账号': '222', '主体名称': '公司B'}
            }
        }
        new_content = {
            'rows': [{'row_num': 2, '银行账号': '111', '主体名称': '公司A'}],
            'account_map': {'111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A'}}
        }

        changes = bankcheck._diff_lookup_snapshots(old_content, new_content)

        assert len(changes) == 1
        assert changes[0]['change_type'] == 'remove'
        assert changes[0]['account'] == '222'

    def test_diff_detects_modify(self):
        old_content = {
            'rows': [{'row_num': 2, '银行账号': '111', '主体名称': '公司A'}],
            'account_map': {'111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A'}}
        }
        new_content = {
            'rows': [{'row_num': 2, '银行账号': '111', '主体名称': '公司A（更名）'}],
            'account_map': {'111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A（更名）'}}
        }

        changes = bankcheck._diff_lookup_snapshots(old_content, new_content)

        assert len(changes) == 1
        assert changes[0]['change_type'] == 'modify'
        assert changes[0]['account'] == '111'
        assert changes[0]['field'] == '主体名称'
        assert changes[0]['old_value'] == '公司A'
        assert changes[0]['new_value'] == '公司A（更名）'

    def test_diff_handles_none_old_content(self):
        new_content = {
            'rows': [{'row_num': 2, '银行账号': '111', '主体名称': '公司A'}],
            'account_map': {'111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A'}}
        }
        changes = bankcheck._diff_lookup_snapshots(None, new_content)
        assert len(changes) == 1
        assert changes[0]['change_type'] == 'add'

    def test_diff_handles_none_new_content(self):
        old_content = {
            'rows': [{'row_num': 2, '银行账号': '111', '主体名称': '公司A'}],
            'account_map': {'111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A'}}
        }
        changes = bankcheck._diff_lookup_snapshots(old_content, None)
        assert len(changes) == 1
        assert changes[0]['change_type'] == 'remove'

    def test_diff_both_none_returns_empty(self):
        changes = bankcheck._diff_lookup_snapshots(None, None)
        assert changes == []

    def test_diff_no_changes_returns_empty(self):
        content = {
            'rows': [{'row_num': 2, '银行账号': '111', '主体名称': '公司A'}],
            'account_map': {'111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A'}}
        }
        changes = bankcheck._diff_lookup_snapshots(content, content)
        assert changes == []

    def test_diff_whitespace_difference(self):
        old_content = {
            'rows': [{'row_num': 2, '银行账号': '111', '主体名称': '公司A '}],
            'account_map': {'111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A '}}
        }
        new_content = {
            'rows': [{'row_num': 2, '银行账号': '111', '主体名称': '公司A'}],
            'account_map': {'111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A'}}
        }
        changes = bankcheck._diff_lookup_snapshots(old_content, new_content)
        assert len(changes) == 0

    def test_diff_combined_changes(self):
        old_content = {
            'rows': [
                {'row_num': 2, '银行账号': '111', '主体名称': '公司A'},
                {'row_num': 3, '银行账号': '222', '主体名称': '公司B'},
                {'row_num': 4, '银行账号': '333', '主体名称': '公司C'},
                {'row_num': 5, '银行账号': '555', '主体名称': '公司E'}
            ],
            'account_map': {
                '111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A'},
                '222': {'row_num': 3, '银行账号': '222', '主体名称': '公司B'},
                '333': {'row_num': 4, '银行账号': '333', '主体名称': '公司C'},
                '555': {'row_num': 5, '银行账号': '555', '主体名称': '公司E'}
            }
        }
        new_content = {
            'rows': [
                {'row_num': 2, '银行账号': '111', '主体名称': '公司A（更名）'},
                {'row_num': 3, '银行账号': '222', '主体名称': '公司B'},
                {'row_num': 4, '银行账号': '444', '主体名称': '公司D'}
            ],
            'account_map': {
                '111': {'row_num': 2, '银行账号': '111', '主体名称': '公司A（更名）'},
                '222': {'row_num': 3, '银行账号': '222', '主体名称': '公司B'},
                '444': {'row_num': 4, '银行账号': '444', '主体名称': '公司D'}
            }
        }

        changes = bankcheck._diff_lookup_snapshots(old_content, new_content)

        add_count = sum(1 for c in changes if c['change_type'] == 'add')
        remove_count = sum(1 for c in changes if c['change_type'] == 'remove')
        modify_count = sum(1 for c in changes if c['change_type'] == 'modify')

        assert add_count == 1
        assert remove_count == 2
        assert modify_count == 1
        assert len(changes) == 4


class TestDetectAndRecordLookupChange:
    def test_first_run_saves_snapshot(self, audit_script_dir):
        result = bankcheck.detect_and_record_lookup_change(audit_script_dir, username='first_user')

        assert result.has_changes is False
        assert result.new_snapshot is not None

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM lookup_snapshots')
        count = cursor.fetchone()[0]
        conn.close()

        assert count == 1

    def test_no_changes_returns_false(self, audit_script_dir):
        bankcheck.detect_and_record_lookup_change(audit_script_dir, username='user1')
        result = bankcheck.detect_and_record_lookup_change(audit_script_dir, username='user2')

        assert result.has_changes is False

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM lookup_snapshots')
        snapshot_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM config_changes')
        change_count = cursor.fetchone()[0]
        conn.close()

        assert snapshot_count == 1
        assert change_count == 0

    def test_detects_change_and_records(self, audit_script_dir):
        bankcheck.detect_and_record_lookup_change(audit_script_dir, username='user1')

        lookup_file = os.path.join(audit_script_dir, '主体查找表.xlsx')
        wb = openpyxl.load_workbook(lookup_file)
        ws = wb.active
        ws.append(['新公司', '9999999999999999999'])
        wb.save(lookup_file)
        wb.close()

        result = bankcheck.detect_and_record_lookup_change(audit_script_dir, username='user2')

        assert result.has_changes is True
        assert result.change_id is not None
        assert result.change_id.startswith('CFG')
        assert len(result.change_details) >= 1
        assert any(c['change_type'] == 'add' for c in result.change_details)

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM lookup_snapshots')
        snapshot_count = cursor.fetchone()[0]
        cursor.execute('SELECT COUNT(*) FROM config_changes')
        change_count = cursor.fetchone()[0]
        conn.close()

        assert snapshot_count == 2
        assert change_count == 1

    def test_no_lookup_file_returns_no_changes(self, tmp_dir):
        empty_script_dir = os.path.join(tmp_dir, 'no_lookup')
        os.makedirs(empty_script_dir, exist_ok=True)

        result = bankcheck.detect_and_record_lookup_change(empty_script_dir, username='test')

        assert result.has_changes is False
        assert result.change_id is None

    def test_manual_record_lookup_change(self, audit_script_dir):
        bankcheck.detect_and_record_lookup_change(audit_script_dir, username='user1')

        lookup_file = os.path.join(audit_script_dir, '主体查找表.xlsx')
        wb = openpyxl.load_workbook(lookup_file)
        ws = wb.active
        ws.append(['手动新增公司', '8888888888888888888'])
        wb.save(lookup_file)
        wb.close()

        result = bankcheck.manual_record_lookup_change(
            audit_script_dir,
            username='manual_user',
            change_reason='新增客户主体'
        )

        assert result.has_changes is True
        assert result.change_id is not None

        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute('SELECT change_reason FROM config_changes WHERE change_id = ?',
                       (result.change_id,))
        reason = cursor.fetchone()[0]
        conn.close()

        assert reason == '新增客户主体'


class TestRecordConfigChangeWithDetails:
    def test_record_change_with_details(self, audit_script_dir):
        details = [
            {'change_type': 'add', 'account': '111', 'description': '新增公司A'},
            {'change_type': 'modify', 'account': '222', 'field': '主体名称', 'old_value': 'X', 'new_value': 'Y'}
        ]

        change_id = bankcheck.record_config_change(
            'lookup_table', '主体查找表.xlsx',
            'old_data', 'new_data',
            change_reason='测试变更',
            script_dir=audit_script_dir,
            username='detail_user',
            change_details=details
        )

        changes = bankcheck.query_config_changes(script_dir=audit_script_dir, limit=1)
        assert len(changes) == 1
        change = changes[0]

        assert change['change_id'] == change_id
        assert isinstance(change['change_details'], list)
        assert len(change['change_details']) == 2
        assert change['change_details'][0]['change_type'] == 'add'

    def test_query_without_parse_details_returns_json_string(self, audit_script_dir):
        details = [{'change_type': 'add', 'account': '111'}]

        bankcheck.record_config_change(
            'lookup_table', 'test.xlsx', 'old', 'new',
            script_dir=audit_script_dir, username='test',
            change_details=details
        )

        changes = bankcheck.query_config_changes(
            script_dir=audit_script_dir,
            parse_details=False,
            limit=1
        )
        assert isinstance(changes[0]['change_details'], str)


class TestExportConfigChangesWithDetails:
    def test_export_config_changes_without_expand(self, tmp_dir, audit_script_dir):
        details = [
            {'change_type': 'add', 'account': '111', 'description': '新增'},
            {'change_type': 'modify', 'account': '222', 'field': '名称', 'old_value': 'A', 'new_value': 'B'}
        ]
        bankcheck.record_config_change(
            'lookup_table', 'test.xlsx', 'old', 'new',
            script_dir=audit_script_dir, username='export_user',
            change_details=details
        )

        output_path = os.path.join(tmp_dir, 'config_export.xlsx')
        result = bankcheck.export_config_changes(output_path, script_dir=audit_script_dir)

        assert result == output_path
        df = pd.read_excel(output_path, engine='openpyxl')
        assert len(df) == 1
        assert 'change_details' in df.columns

    def test_export_config_changes_with_expand(self, tmp_dir, audit_script_dir):
        details = [
            {'change_type': 'add', 'account': '111', 'description': '新增'},
            {'change_type': 'modify', 'account': '222', 'field': '名称', 'old_value': 'A', 'new_value': 'B'}
        ]
        bankcheck.record_config_change(
            'lookup_table', 'test.xlsx', 'old', 'new',
            script_dir=audit_script_dir, username='export_user',
            change_details=details
        )

        output_path = os.path.join(tmp_dir, 'config_expand.xlsx')
        result = bankcheck.export_config_changes(
            output_path,
            script_dir=audit_script_dir,
            expand_details=True
        )

        assert result == output_path
        df = pd.read_excel(output_path, engine='openpyxl')
        assert len(df) == 2
        assert 'change_type' in df.columns
        assert 'account' in df.columns
        assert 'field' in df.columns
        assert 'description' in df.columns
        assert df.iloc[0]['change_type'] == 'add'
        assert df.iloc[1]['change_type'] == 'modify'


class TestConfigChangesTable:
    def test_config_changes_has_change_details_column(self, audit_script_dir):
        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("PRAGMA table_info(config_changes)")
        columns = {row[1] for row in cursor.fetchall()}
        conn.close()
        assert 'change_details' in columns

    def test_lookup_snapshots_table_exists(self, audit_script_dir):
        db_path = bankcheck.get_audit_db_path(audit_script_dir)
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='lookup_snapshots'")
        table = cursor.fetchone()
        conn.close()
        assert table is not None


class TestIntegrationLookupChangeDetection:
    def test_full_pipeline_with_lookup_change(self, tmp_dir, audit_script_dir):
        bankcheck.detect_and_record_lookup_change(audit_script_dir, username='init_user')

        lookup_file = os.path.join(audit_script_dir, '主体查找表.xlsx')
        wb = openpyxl.load_workbook(lookup_file)
        ws = wb.active
        ws.append(['新主体公司', '7777777777777777777'])
        wb.save(lookup_file)
        wb.close()

        source_folder = os.path.join(tmp_dir, '流水')
        os.makedirs(source_folder, exist_ok=True)
        _create_beijing_bank_excel(os.path.join(source_folder, '北京银行_流水.xlsx'))

        change_result = bankcheck.detect_and_record_lookup_change(
            audit_script_dir,
            username='operator_user',
            change_reason='新增客户主体'
        )

        result = bankcheck.run_pipeline(source_folder, audit_script_dir)

        changes = bankcheck.query_config_changes(
            script_dir=audit_script_dir,
            config_type='lookup_table',
            limit=1
        )

        assert len(changes) >= 1
        change = changes[0]
        assert change['username'] == 'operator_user'
        assert change['change_reason'] == '新增客户主体'
        assert isinstance(change['change_details'], list)
        assert len(change['change_details']) >= 1
        assert any(c['account'] == '7777777777777777777' for c in change['change_details'])


class TestWhoChangedWhat:
    def test_query_who_changed_what(self, audit_script_dir):
        bankcheck.detect_and_record_lookup_change(audit_script_dir, username='user1')

        lookup_file = os.path.join(audit_script_dir, '主体查找表.xlsx')
        wb = openpyxl.load_workbook(lookup_file)
        ws = wb.active
        ws.append(['变更测试公司', '6666666666666666666'])
        wb.save(lookup_file)
        wb.close()

        result = bankcheck.detect_and_record_lookup_change(
            audit_script_dir,
            username='operator_zhangsan',
            change_reason='客户主体更新'
        )

        changes = bankcheck.query_config_changes(
            script_dir=audit_script_dir,
            username='operator_zhangsan'
        )

        assert len(changes) == 1
        c = changes[0]

        assert c['username'] == 'operator_zhangsan'
        assert c['config_type'] == 'lookup_table'
        assert c['change_reason'] == '客户主体更新'
        assert isinstance(c['change_details'], list)
        assert len(c['change_details']) == 1

        detail = c['change_details'][0]
        assert detail['change_type'] == 'add'
        assert detail['account'] == '6666666666666666666'
        assert '变更测试公司' in detail['description']
        assert detail['new_value'] is not None
        assert '变更测试公司' in str(detail['new_value'])

