import os
import sys
import tempfile

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck
import database
import lookup_manager
import batch_manager


@pytest.fixture(autouse=True)
def setup_logging():
    bankcheck.setup_logging()


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='path_test_')
    yield d
    import shutil
    shutil.rmtree(d, ignore_errors=True)


class TestBankcheckPathConsistency:
    def test_all_modules_have_same_functions(self):
        modules = [bankcheck, database, lookup_manager, batch_manager]
        required_funcs = [
            'get_script_dir',
            'get_program_dir',
            'is_writable',
            'get_user_data_dir',
            'get_writable_dir',
            'get_output_dir',
        ]
        for module in modules:
            for func_name in required_funcs:
                assert hasattr(module, func_name), f"{module.__name__} missing {func_name}"
                assert callable(getattr(module, func_name)), f"{module.__name__}.{func_name} is not callable"

    def test_get_program_dir_consistency(self):
        bankcheck_dir = bankcheck.get_program_dir()
        database_dir = database.get_program_dir()
        lookup_dir = lookup_manager.get_program_dir()
        batch_dir = batch_manager.get_program_dir()
        assert bankcheck_dir == database_dir == lookup_dir == batch_dir

    def test_get_user_data_dir_consistency(self):
        bankcheck_dir = bankcheck.get_user_data_dir()
        database_dir = database.get_user_data_dir()
        lookup_dir = lookup_manager.get_user_data_dir()
        batch_dir = batch_manager.get_user_data_dir()
        assert bankcheck_dir == database_dir == lookup_dir == batch_dir

    def test_is_writable_consistency(self, tmp_dir):
        assert bankcheck.is_writable(tmp_dir) is True
        assert database.is_writable(tmp_dir) is True
        assert lookup_manager.is_writable(tmp_dir) is True
        assert batch_manager.is_writable(tmp_dir) is True

        fake_dir = '/nonexistent/path/that/should/not/exist'
        assert bankcheck.is_writable(fake_dir) is False
        assert database.is_writable(fake_dir) is False
        assert lookup_manager.is_writable(fake_dir) is False
        assert batch_manager.is_writable(fake_dir) is False

    def test_get_writable_dir_returns_writable(self):
        modules = [bankcheck, database, lookup_manager, batch_manager]
        for module in modules:
            result = module.get_writable_dir()
            assert os.path.isabs(result)
            assert os.path.isdir(result)
            assert module.is_writable(result) is True

    def test_get_output_dir_creates_subdir(self, tmp_dir, monkeypatch):
        def mock_get_writable_dir():
            return tmp_dir

        modules = [bankcheck, database, lookup_manager, batch_manager]
        for module in modules:
            monkeypatch.setattr(f'{module.__name__}.get_writable_dir', mock_get_writable_dir)
            result = module.get_output_dir('test_subdir')
            expected = os.path.join(tmp_dir, 'test_subdir')
            assert result == expected
            assert os.path.isdir(result)


class TestDatabasePathStrategy:
    def test_sqlite_db_path_in_output_dir(self, tmp_dir, monkeypatch):
        def mock_get_output_dir(subdir=None):
            return tmp_dir
        monkeypatch.setattr('database.get_output_dir', mock_get_output_dir)

        backend = database.SQLiteBackend()
        assert backend.db_path == os.path.join(tmp_dir, 'transactions.db')

    def test_load_database_config_prefers_output_dir(self, tmp_dir, monkeypatch):
        program_dir = os.path.join(tmp_dir, 'program')
        output_dir = os.path.join(tmp_dir, 'output')
        os.makedirs(program_dir)
        os.makedirs(output_dir)

        def mock_get_program_dir():
            return program_dir
        def mock_get_output_dir(subdir=None):
            return output_dir

        monkeypatch.setattr('database.get_program_dir', mock_get_program_dir)
        monkeypatch.setattr('database.get_output_dir', mock_get_output_dir)

        output_config = os.path.join(output_dir, 'database_config.json')
        with open(output_config, 'w') as f:
            f.write('{"backend": "sqlite", "auto_persist": false}')

        program_config = os.path.join(program_dir, 'database_config.json')
        with open(program_config, 'w') as f:
            f.write('{"backend": "sqlite", "auto_persist": true}')

        config = database.load_database_config()
        assert config['auto_persist'] is False

    def test_save_database_config_to_output_dir(self, tmp_dir, monkeypatch):
        output_dir = os.path.join(tmp_dir, 'output')
        os.makedirs(output_dir)

        def mock_get_output_dir(subdir=None):
            return output_dir
        monkeypatch.setattr('database.get_output_dir', mock_get_output_dir)

        config = {'backend': 'sqlite', 'auto_persist': True}
        saved_path = database.save_database_config(config)
        assert saved_path == os.path.join(output_dir, 'database_config.json')
        assert os.path.exists(saved_path)

    def test_load_database_config_with_script_dir_only(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'custom')
        other_dir = os.path.join(tmp_dir, 'other')
        os.makedirs(script_dir)
        os.makedirs(other_dir)

        script_config = os.path.join(script_dir, 'database_config.json')
        other_config = os.path.join(other_dir, 'database_config.json')

        with open(script_config, 'w') as f:
            f.write('{"backend": "sqlite", "auto_persist": false}')
        with open(other_config, 'w') as f:
            f.write('{"backend": "sqlite", "auto_persist": true}')

        config = database.load_database_config(script_dir=script_dir)
        assert config['auto_persist'] is False

    def test_save_database_config_with_script_dir(self, tmp_dir):
        script_dir = os.path.join(tmp_dir, 'custom')
        os.makedirs(script_dir)

        config = {'backend': 'sqlite', 'auto_persist': False}
        saved_path = database.save_database_config(config, script_dir=script_dir)
        assert saved_path == os.path.join(script_dir, 'database_config.json')
        assert os.path.exists(saved_path)

        loaded_config = database.load_database_config(script_dir=script_dir)
        assert loaded_config['auto_persist'] is False


class TestLookupManagerPathStrategy:
    def test_find_lookup_file_prefers_output_dir(self, tmp_dir, monkeypatch):
        output_dir = os.path.join(tmp_dir, 'output')
        program_dir = os.path.join(tmp_dir, 'program')
        os.makedirs(output_dir)
        os.makedirs(program_dir)

        output_lookup = os.path.join(output_dir, '主体查找表.xlsx')
        program_lookup = os.path.join(program_dir, '主体查找表.xlsx')

        wb = openpyxl.Workbook()
        wb.save(output_lookup)
        wb.save(program_lookup)
        wb.close()

        def mock_get_output_dir(subdir=None):
            return output_dir
        def mock_get_program_dir():
            return program_dir

        monkeypatch.setattr('lookup_manager.get_output_dir', mock_get_output_dir)
        monkeypatch.setattr('lookup_manager.get_program_dir', mock_get_program_dir)

        result = lookup_manager.find_lookup_file()
        assert result == output_lookup

    def test_get_lookup_file_path_defaults_to_output_dir(self, tmp_dir, monkeypatch):
        output_dir = os.path.join(tmp_dir, 'output')
        os.makedirs(output_dir)

        def mock_get_output_dir(subdir=None):
            return output_dir
        def mock_find_lookup_file(script_dir=None):
            return None

        monkeypatch.setattr('lookup_manager.get_output_dir', mock_get_output_dir)
        monkeypatch.setattr('lookup_manager.find_lookup_file', mock_find_lookup_file)

        result = lookup_manager.get_lookup_file_path()
        assert result == os.path.join(output_dir, '主体查找表.xlsx')


class TestBatchManagerPathStrategy:
    def test_batch_manager_uses_output_dir_by_default(self, tmp_dir, monkeypatch):
        output_dir = os.path.join(tmp_dir, 'output')
        os.makedirs(output_dir)

        def mock_is_writable(path):
            return False
        def mock_get_output_dir(subdir=None):
            return output_dir

        monkeypatch.setattr('batch_manager.is_writable', mock_is_writable)
        monkeypatch.setattr('batch_manager.get_output_dir', mock_get_output_dir)

        manager = batch_manager.BatchManager()
        assert manager.script_dir == output_dir
        assert manager.db_path == os.path.join(output_dir, 'batch_history.db')
        assert manager.history_root == os.path.join(output_dir, 'history')

    def test_batch_manager_uses_custom_writable_dir(self, tmp_dir, monkeypatch):
        custom_dir = os.path.join(tmp_dir, 'custom')
        os.makedirs(custom_dir)

        def mock_is_writable(path):
            return path == custom_dir

        monkeypatch.setattr('batch_manager.is_writable', mock_is_writable)

        manager = batch_manager.BatchManager(script_dir=custom_dir)
        assert manager.script_dir == custom_dir

    def test_batch_manager_fallback_when_custom_not_writable(self, tmp_dir, monkeypatch):
        output_dir = os.path.join(tmp_dir, 'output')
        custom_dir = os.path.join(tmp_dir, 'custom')
        os.makedirs(output_dir)
        os.makedirs(custom_dir)

        def mock_is_writable(path):
            return False
        def mock_get_output_dir(subdir=None):
            return output_dir

        monkeypatch.setattr('batch_manager.is_writable', mock_is_writable)
        monkeypatch.setattr('batch_manager.get_output_dir', mock_get_output_dir)

        manager = batch_manager.BatchManager(script_dir=custom_dir)
        assert manager.script_dir == output_dir


class TestProtectedDirectoryScenario:
    def test_full_scenario_program_dir_protected(self, tmp_dir, monkeypatch):
        program_dir = os.path.join(tmp_dir, 'Program Files', 'bankcheck')
        user_dir = os.path.join(tmp_dir, 'Users', 'test', 'AppData', 'Roaming', 'bankcheck')
        os.makedirs(program_dir)

        program_lookup = os.path.join(program_dir, '主体查找表.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '主体名称'
        ws['B1'] = '银行账号'
        ws['A2'] = '测试公司'
        ws['B2'] = '123456789'
        wb.save(program_lookup)
        wb.close()

        def mock_is_writable(path):
            if path == program_dir:
                return False
            return True

        def mock_get_program_dir():
            return program_dir

        def mock_get_user_data_dir():
            return user_dir

        def mock_get_output_dir(subdir=None):
            base = user_dir
            if subdir:
                base = os.path.join(base, subdir)
            os.makedirs(base, exist_ok=True)
            return base

        modules = [bankcheck, database, lookup_manager, batch_manager]
        for module in modules:
            monkeypatch.setattr(f'{module.__name__}.is_writable', mock_is_writable)
            monkeypatch.setattr(f'{module.__name__}.get_program_dir', mock_get_program_dir)
            monkeypatch.setattr(f'{module.__name__}.get_user_data_dir', mock_get_user_data_dir)
            monkeypatch.setattr(f'{module.__name__}.get_output_dir', mock_get_output_dir)

        assert bankcheck.get_writable_dir() == user_dir
        assert os.path.exists(user_dir)

        lookup_path = bankcheck.find_lookup_file()
        assert lookup_path == os.path.join(user_dir, '主体查找表.xlsx')
        assert os.path.exists(lookup_path)

        summary_path = bankcheck.get_summary_table_path()
        assert summary_path == os.path.join(user_dir, '银行流水总表.xlsx')

        log_dir = bankcheck.get_output_dir()
        log_file = os.path.join(log_dir, 'bankcheck.log')
        assert bankcheck.is_writable(os.path.dirname(log_file)) is True

        audit_db_path = bankcheck.get_audit_db_path()
        assert audit_db_path == os.path.join(user_dir, 'audit_log.db')

        db_backend = database.SQLiteBackend()
        assert db_backend.db_path == os.path.join(user_dir, 'transactions.db')

        lookup_mgr_path = lookup_manager.get_lookup_file_path()
        assert lookup_mgr_path == os.path.join(user_dir, '主体查找表.xlsx')

        batch_mgr = batch_manager.BatchManager()
        assert batch_mgr.script_dir == user_dir
        assert batch_mgr.db_path == os.path.join(user_dir, 'batch_history.db')
