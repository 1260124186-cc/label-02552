import os
import sys
import tempfile

import openpyxl
import pytest

import bankcheck
from bankcheck import (
    open_workbook_compat,
    cleanup_temp_file,
    iter_sheet_rows,
    iter_sheet_records,
    estimate_row_count,
    DEFAULT_CHUNK_SIZE,
)


class TestOpenWorkbookCompatXlsx:
    def test_open_xlsx(self, tmp_dir):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'test'
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb.save(path)
        wb.close()

        result_wb, tmp_path = open_workbook_compat(path)
        assert tmp_path is None
        assert result_wb.active['A1'].value == 'test'
        result_wb.close()

    def test_no_temp_file_for_xlsx(self, tmp_dir):
        wb = openpyxl.Workbook()
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb.save(path)
        wb.close()

        _, tmp_path = open_workbook_compat(path)
        assert tmp_path is None


class TestOpenWorkbookCompatReadOnly:
    def test_read_only_default_false(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        wb.active['A1'] = 'hello'
        wb.save(path)
        wb.close()

        result_wb, tmp_path = open_workbook_compat(path)
        assert result_wb.read_only is False
        result_wb.close()

    def test_read_only_true(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        wb.active['A1'] = 'hello'
        wb.active['B1'] = 'world'
        wb.save(path)
        wb.close()

        result_wb, tmp_path = open_workbook_compat(path, read_only=True)
        assert result_wb.read_only is True
        result_wb.close()

    def test_read_only_can_read_values(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'name'
        ws['B1'] = 'value'
        ws['A2'] = 'test'
        ws['B2'] = 123
        wb.save(path)
        wb.close()

        result_wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws_result = result_wb.active
        rows = list(ws_result.iter_rows(values_only=True))
        assert len(rows) == 2
        assert rows[0] == ('name', 'value')
        assert rows[1] == ('test', 123)
        result_wb.close()

    def test_read_only_no_temp_file(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        wb.save(path)
        wb.close()

        _, tmp_path = open_workbook_compat(path, read_only=True)
        assert tmp_path is None

    def test_read_only_multi_sheet(self, tmp_dir):
        path = os.path.join(tmp_dir, 'multi.xlsx')
        wb = openpyxl.Workbook()
        wb.active.title = 'Sheet1'
        wb.active['A1'] = 's1'
        ws2 = wb.create_sheet('Sheet2')
        ws2['A1'] = 's2'
        wb.save(path)
        wb.close()

        result_wb, tmp_path = open_workbook_compat(path, read_only=True)
        assert len(result_wb.worksheets) == 2
        assert result_wb['Sheet1'] is not None
        assert result_wb['Sheet2'] is not None
        result_wb.close()

    def test_read_only_large_sheet(self, tmp_dir):
        path = os.path.join(tmp_dir, 'large.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 101):
            ws.cell(row=i, column=1, value=f'row_{i}')
            ws.cell(row=i, column=2, value=i * 10)
        wb.save(path)
        wb.close()

        result_wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws_result = result_wb.active
        row_count = 0
        for row in ws_result.iter_rows(values_only=True):
            row_count += 1
        assert row_count == 100
        result_wb.close()


class TestIterSheetRows:
    def _create_test_sheet(self, tmp_dir, num_rows=100, num_cols=5):
        path = os.path.join(tmp_dir, 'iter_test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        for r in range(1, num_rows + 1):
            for c in range(1, num_cols + 1):
                ws.cell(row=r, column=c, value=f'r{r}c{c}')
        wb.save(path)
        wb.close()
        return path

    def test_iter_all_rows_single_chunk(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=50)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws, chunk_size=100))
        assert len(chunks) == 1
        assert len(chunks[0]) == 50
        wb.close()

    def test_iter_all_rows_multiple_chunks(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=100)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws, chunk_size=30))
        assert len(chunks) == 4
        assert len(chunks[0]) == 30
        assert len(chunks[1]) == 30
        assert len(chunks[2]) == 30
        assert len(chunks[3]) == 10
        wb.close()

    def test_iter_with_start_row(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=10)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws, start_row=5, chunk_size=100))
        assert len(chunks) == 1
        assert len(chunks[0]) == 6
        assert chunks[0][0][0] == 'r5c1'
        wb.close()

    def test_iter_with_end_row(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=20)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws, end_row=5, chunk_size=100))
        assert len(chunks) == 1
        assert len(chunks[0]) == 5
        assert chunks[0][-1][0] == 'r5c1'
        wb.close()

    def test_iter_with_start_and_end_row(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=20)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws, start_row=3, end_row=7, chunk_size=100))
        assert len(chunks) == 1
        assert len(chunks[0]) == 5
        assert chunks[0][0][0] == 'r3c1'
        assert chunks[0][-1][0] == 'r7c1'
        wb.close()

    def test_iter_values_only_true(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=3)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws, chunk_size=10, values_only=True))
        assert isinstance(chunks[0][0][0], str)
        assert chunks[0][0][0] == 'r1c1'
        wb.close()

    def test_iter_values_only_false(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=3)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws, chunk_size=10, values_only=False))
        assert hasattr(chunks[0][0][0], 'value')
        assert chunks[0][0][0].value == 'r1c1'
        wb.close()

    def test_iter_start_row_gt_end_row(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=10)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws, start_row=10, end_row=5, chunk_size=100))
        assert len(chunks) == 0
        wb.close()

    def test_iter_default_chunk_size(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=5)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws))
        assert len(chunks) == 1
        assert len(chunks[0]) == 5
        wb.close()

    def test_iter_chunk_size_one(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=5)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws, chunk_size=1))
        assert len(chunks) == 5
        for chunk in chunks:
            assert len(chunk) == 1
        wb.close()

    def test_iter_empty_sheet(self, tmp_dir):
        path = os.path.join(tmp_dir, 'empty.xlsx')
        wb = openpyxl.Workbook()
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb2.active
        chunks = list(iter_sheet_rows(ws))
        assert len(chunks) == 0
        wb2.close()

    def test_iter_chunks_have_correct_order(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=10)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        all_rows = []
        for chunk in iter_sheet_rows(ws, chunk_size=3):
            all_rows.extend(chunk)

        assert len(all_rows) == 10
        for i, row in enumerate(all_rows, 1):
            assert row[0] == f'r{i}c1'
        wb.close()

    def test_iter_without_read_only(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=10)
        wb, tmp_path = open_workbook_compat(path, read_only=False)
        ws = wb.active

        chunks = list(iter_sheet_rows(ws, chunk_size=5))
        assert len(chunks) == 2
        assert len(chunks[0]) == 5
        assert len(chunks[1]) == 5
        wb.close()


class TestIterSheetRecords:
    def _create_test_sheet(self, tmp_dir, num_rows=50):
        path = os.path.join(tmp_dir, 'records_test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = 'name'
        ws['B1'] = 'age'
        ws['C1'] = 'city'
        for i in range(2, num_rows + 2):
            ws.cell(row=i, column=1, value=f'person_{i-1}')
            ws.cell(row=i, column=2, value=20 + (i % 30))
            ws.cell(row=i, column=3, value=f'city_{i % 5}')
        wb.save(path)
        wb.close()
        return path

    def test_iter_records_basic(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=10)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        columns_map = {
            'name': 1,
            'age': 2,
            'city': 3,
        }

        chunks = list(iter_sheet_records(ws, columns_map, start_row=2, chunk_size=100))
        assert len(chunks) == 1
        assert len(chunks[0]) == 10
        assert 'name' in chunks[0][0]
        assert 'age' in chunks[0][0]
        assert 'city' in chunks[0][0]
        wb.close()

    def test_iter_records_multiple_chunks(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=25)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        columns_map = {'name': 1, 'age': 2, 'city': 3}
        chunks = list(iter_sheet_records(ws, columns_map, start_row=2, chunk_size=10))
        assert len(chunks) == 3
        assert len(chunks[0]) == 10
        assert len(chunks[1]) == 10
        assert len(chunks[2]) == 5
        wb.close()

    def test_iter_records_values_correct(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=3)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        columns_map = {'name': 1, 'age': 2, 'city': 3}
        chunks = list(iter_sheet_records(ws, columns_map, start_row=2, chunk_size=100))
        records = chunks[0]

        assert records[0]['name'] == 'person_1'
        assert records[1]['name'] == 'person_2'
        assert records[2]['name'] == 'person_3'
        assert records[0]['age'] == 22
        assert records[0]['city'] == 'city_2'
        wb.close()

    def test_iter_records_subset_columns(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=5)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        columns_map = {'name': 1, 'city': 3}
        chunks = list(iter_sheet_records(ws, columns_map, start_row=2, chunk_size=100))
        record = chunks[0][0]
        assert 'name' in record
        assert 'city' in record
        assert 'age' not in record
        wb.close()

    def test_iter_records_with_end_row(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=20)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        columns_map = {'name': 1}
        chunks = list(iter_sheet_records(ws, columns_map, start_row=2, end_row=6, chunk_size=100))
        assert len(chunks) == 1
        assert len(chunks[0]) == 5
        wb.close()

    def test_iter_records_sparse_columns(self, tmp_dir):
        path = os.path.join(tmp_dir, 'sparse.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='col1')
        ws.cell(row=1, column=5, value='col5')
        ws.cell(row=1, column=10, value='col10')
        ws.cell(row=2, column=1, value='a')
        ws.cell(row=2, column=5, value='b')
        ws.cell(row=2, column=10, value='c')
        wb.save(path)
        wb.close()

        wb2, tmp_path = open_workbook_compat(path, read_only=True)
        ws2 = wb2.active

        columns_map = {'col1': 1, 'col5': 5, 'col10': 10}
        chunks = list(iter_sheet_records(ws2, columns_map, start_row=2, chunk_size=100))
        record = chunks[0][0]
        assert record['col1'] == 'a'
        assert record['col5'] == 'b'
        assert record['col10'] == 'c'
        wb2.close()

    def test_iter_records_chunk_order_preserved(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=20)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        columns_map = {'name': 1}
        all_records = []
        for chunk in iter_sheet_records(ws, columns_map, start_row=2, chunk_size=7):
            all_records.extend(chunk)

        assert len(all_records) == 20
        for i, rec in enumerate(all_records, 1):
            assert rec['name'] == f'person_{i}'
        wb.close()

    def test_iter_records_empty_result(self, tmp_dir):
        path = self._create_test_sheet(tmp_dir, num_rows=0)
        wb, tmp_path = open_workbook_compat(path, read_only=True)
        ws = wb.active

        columns_map = {'name': 1}
        chunks = list(iter_sheet_records(ws, columns_map, start_row=2, chunk_size=100))
        assert len(chunks) == 0
        wb.close()


class TestEstimateRowCount:
    def test_estimate_row_count_small(self, tmp_dir):
        path = os.path.join(tmp_dir, 'count_test.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        for i in range(1, 51):
            ws.cell(row=i, column=1, value=i)
        wb.save(path)
        wb.close()

        count = estimate_row_count(path)
        assert count == 50

    def test_estimate_row_count_specific_sheet(self, tmp_dir):
        path = os.path.join(tmp_dir, 'multi_sheet.xlsx')
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Sheet1'
        for i in range(1, 21):
            ws1.cell(row=i, column=1, value=i)
        ws2 = wb.create_sheet('Sheet2')
        for i in range(1, 101):
            ws2.cell(row=i, column=1, value=i)
        wb.save(path)
        wb.close()

        count1 = estimate_row_count(path, sheet_name='Sheet1')
        count2 = estimate_row_count(path, sheet_name='Sheet2')
        assert count1 == 20
        assert count2 == 100

    def test_estimate_row_count_first_sheet_default(self, tmp_dir):
        path = os.path.join(tmp_dir, 'multi_sheet.xlsx')
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'First'
        for i in range(1, 11):
            ws1.cell(row=i, column=1, value=i)
        ws2 = wb.create_sheet('Second')
        for i in range(1, 101):
            ws2.cell(row=i, column=1, value=i)
        wb.save(path)
        wb.close()

        count = estimate_row_count(path)
        assert count == 10

    def test_estimate_row_count_returns_int(self, tmp_dir):
        path = os.path.join(tmp_dir, 'test.xlsx')
        wb = openpyxl.Workbook()
        wb.active['A1'] = 'test'
        wb.save(path)
        wb.close()

        count = estimate_row_count(path)
        assert isinstance(count, int)

    def test_estimate_row_count_nonexistent_file(self, tmp_dir):
        count = estimate_row_count('/nonexistent/path/file.xlsx')
        assert count is None

    def test_estimate_row_count_uses_read_only(self, tmp_dir, monkeypatch):
        path = os.path.join(tmp_dir, 'spy.xlsx')
        wb = openpyxl.Workbook()
        wb.active['A1'] = 'test'
        wb.save(path)
        wb.close()

        original_open = open_workbook_compat
        call_args = []

        def spy_open(filepath, read_only=False):
            call_args.append({'filepath': filepath, 'read_only': read_only})
            return original_open(filepath, read_only=read_only)

        monkeypatch.setattr('bankcheck.open_workbook_compat', spy_open)
        estimate_row_count(path)
        assert len(call_args) == 1
        assert call_args[0]['read_only'] is True


class TestDefaultChunkSize:
    def test_default_chunk_size_value(self):
        assert isinstance(DEFAULT_CHUNK_SIZE, int)
        assert DEFAULT_CHUNK_SIZE > 0
        assert DEFAULT_CHUNK_SIZE == 1000


class TestCleanupTempFile:
    def test_cleanup_existing_file(self, tmp_dir):
        tmp_path = os.path.join(tmp_dir, 'temp_file.xlsx')
        with open(tmp_path, 'w') as f:
            f.write('test')
        assert os.path.exists(tmp_path)
        cleanup_temp_file(tmp_path)
        assert not os.path.exists(tmp_path)

    def test_cleanup_nonexistent_file(self):
        cleanup_temp_file('/nonexistent/file.xlsx')

    def test_cleanup_none(self):
        cleanup_temp_file(None)


class TestGetScriptDir:
    def test_returns_directory(self):
        result = bankcheck.get_script_dir()
        assert os.path.isdir(result)

    def test_returns_absolute_path(self):
        result = bankcheck.get_script_dir()
        assert os.path.isabs(result)


class TestGetProgramDir:
    def test_returns_same_as_script_dir(self):
        assert bankcheck.get_program_dir() == bankcheck.get_script_dir()

    def test_returns_absolute_path(self):
        assert os.path.isabs(bankcheck.get_program_dir())


class TestIsWritable:
    def test_writable_directory(self, tmp_dir):
        assert bankcheck.is_writable(tmp_dir) is True

    def test_nonexistent_directory(self):
        assert bankcheck.is_writable('/nonexistent/path/that/should/not/exist') is False

    def test_file_instead_of_directory(self, tmp_dir):
        test_file = os.path.join(tmp_dir, 'test.txt')
        with open(test_file, 'w') as f:
            f.write('test')
        assert bankcheck.is_writable(test_file) is False

    def test_readonly_directory(self, tmp_dir, monkeypatch):
        def mock_open(*args, **kwargs):
            raise OSError('Permission denied')
        monkeypatch.setattr('builtins.open', mock_open)
        assert bankcheck.is_writable(tmp_dir) is False


class TestGetUserDataDir:
    def test_returns_absolute_path(self):
        result = bankcheck.get_user_data_dir()
        assert os.path.isabs(result)

    def test_contains_app_name(self):
        result = bankcheck.get_user_data_dir()
        assert 'bankcheck' in result

    def test_windows_appdata(self, monkeypatch):
        monkeypatch.setattr('sys.platform', 'win32')
        monkeypatch.setenv('APPDATA', 'C:\\Users\\test\\AppData\\Roaming')
        result = bankcheck.get_user_data_dir()
        assert result == os.path.join('C:\\Users\\test\\AppData\\Roaming', 'bankcheck')

    def test_macos_library(self, monkeypatch):
        monkeypatch.setattr('sys.platform', 'darwin')
        monkeypatch.setenv('HOME', '/Users/test')
        result = bankcheck.get_user_data_dir()
        assert result == '/Users/test/Library/Application Support/bankcheck'

    def test_linux_home(self, monkeypatch):
        monkeypatch.setattr('sys.platform', 'linux')
        monkeypatch.setenv('HOME', '/home/test')
        result = bankcheck.get_user_data_dir()
        assert result == '/home/test/.bankcheck'


class TestGetWritableDir:
    def test_returns_absolute_path(self):
        result = bankcheck.get_writable_dir()
        assert os.path.isabs(result)

    def test_returns_writable_directory(self):
        result = bankcheck.get_writable_dir()
        assert bankcheck.is_writable(result) is True

    def test_fallback_to_user_data_when_program_not_writable(self, tmp_dir, monkeypatch):
        program_dir = os.path.join(tmp_dir, 'program')
        user_data_dir = os.path.join(tmp_dir, 'user_data')
        os.makedirs(program_dir)
        os.makedirs(user_data_dir)

        def mock_is_writable(path):
            if path == program_dir:
                return False
            return True

        def mock_get_program_dir():
            return program_dir

        def mock_get_user_data_dir():
            return user_data_dir

        monkeypatch.setattr('bankcheck.is_writable', mock_is_writable)
        monkeypatch.setattr('bankcheck.get_program_dir', mock_get_program_dir)
        monkeypatch.setattr('bankcheck.get_user_data_dir', mock_get_user_data_dir)

        result = bankcheck.get_writable_dir()
        assert result == user_data_dir


class TestGetOutputDir:
    def test_returns_absolute_path(self):
        result = bankcheck.get_output_dir()
        assert os.path.isabs(result)

    def test_creates_directory(self, tmp_dir, monkeypatch):
        test_dir = os.path.join(tmp_dir, 'test_output')
        assert not os.path.exists(test_dir)
        def mock_get_writable_dir():
            return tmp_dir
        monkeypatch.setattr('bankcheck.get_writable_dir', mock_get_writable_dir)
        result = bankcheck.get_output_dir('test_output')
        assert os.path.exists(result)
        assert result == test_dir

    def test_with_subdir(self, tmp_dir, monkeypatch):
        def mock_get_writable_dir():
            return tmp_dir
        monkeypatch.setattr('bankcheck.get_writable_dir', mock_get_writable_dir)
        result = bankcheck.get_output_dir('logs')
        assert result == os.path.join(tmp_dir, 'logs')
        assert os.path.isdir(result)

    def test_without_subdir(self, tmp_dir, monkeypatch):
        def mock_get_writable_dir():
            return tmp_dir
        monkeypatch.setattr('bankcheck.get_writable_dir', mock_get_writable_dir)
        result = bankcheck.get_output_dir()
        assert result == tmp_dir


class TestGetSummaryTablePath:
    def test_default_path(self, tmp_dir, monkeypatch):
        def mock_get_output_dir():
            return tmp_dir
        monkeypatch.setattr('bankcheck.get_output_dir', mock_get_output_dir)
        result = bankcheck.get_summary_table_path()
        assert result == os.path.join(tmp_dir, '银行流水总表.xlsx')

    def test_custom_output_dir(self, tmp_dir):
        custom_dir = os.path.join(tmp_dir, 'custom')
        os.makedirs(custom_dir)
        result = bankcheck.get_summary_table_path(output_dir=custom_dir)
        assert result == os.path.join(custom_dir, '银行流水总表.xlsx')


class TestGetAuditDbPath:
    def test_default_path(self, tmp_dir, monkeypatch):
        def mock_get_output_dir():
            return tmp_dir
        monkeypatch.setattr('bankcheck.get_output_dir', mock_get_output_dir)
        result = bankcheck.get_audit_db_path()
        assert result == os.path.join(tmp_dir, 'audit_log.db')

    def test_custom_writable_dir(self, tmp_dir):
        result = bankcheck.get_audit_db_path(script_dir=tmp_dir)
        assert result == os.path.join(tmp_dir, 'audit_log.db')

    def test_fallback_when_script_dir_not_writable(self, tmp_dir, monkeypatch):
        def mock_is_writable(path):
            return False
        monkeypatch.setattr('bankcheck.is_writable', mock_is_writable)
        def mock_get_output_dir():
            return tmp_dir
        monkeypatch.setattr('bankcheck.get_output_dir', mock_get_output_dir)
        result = bankcheck.get_audit_db_path(script_dir='/readonly/path')
        assert result == os.path.join(tmp_dir, 'audit_log.db')


class TestPyInstallerFrozenMode:
    def test_get_script_dir_frozen(self, monkeypatch, tmp_dir):
        fake_exe = os.path.join(tmp_dir, 'bankcheck.exe')
        monkeypatch.setattr('sys.frozen', True, raising=False)
        monkeypatch.setattr('sys.executable', fake_exe)
        result = bankcheck.get_script_dir()
        assert result == tmp_dir

    def test_get_script_dir_not_frozen(self, monkeypatch):
        if hasattr(sys, 'frozen'):
            monkeypatch.delattr('sys', 'frozen')
        result = bankcheck.get_script_dir()
        expected = os.path.dirname(os.path.abspath(bankcheck.__file__))
        assert result == expected


class TestFindLookupFileSmart:
    def test_find_in_output_dir_first(self, tmp_dir, monkeypatch):
        output_dir = os.path.join(tmp_dir, 'output')
        program_dir = os.path.join(tmp_dir, 'program')
        os.makedirs(output_dir)
        os.makedirs(program_dir)

        output_lookup = os.path.join(output_dir, '主体查找表.xlsx')
        program_lookup = os.path.join(program_dir, '主体查找表.xlsx')

        wb = openpyxl.Workbook()
        wb.save(output_lookup)
        wb.save(program_lookup)
        wb.close()

        def mock_get_output_dir():
            return output_dir
        def mock_get_program_dir():
            return program_dir

        monkeypatch.setattr('bankcheck.get_output_dir', mock_get_output_dir)
        monkeypatch.setattr('bankcheck.get_program_dir', mock_get_program_dir)

        result = bankcheck.find_lookup_file()
        assert result == output_lookup

    def test_copy_from_program_to_output(self, tmp_dir, monkeypatch):
        output_dir = os.path.join(tmp_dir, 'output')
        program_dir = os.path.join(tmp_dir, 'program')
        os.makedirs(output_dir)
        os.makedirs(program_dir)

        program_lookup = os.path.join(program_dir, '主体查找表.xlsx')
        wb = openpyxl.Workbook()
        wb.save(program_lookup)
        wb.close()

        def mock_get_output_dir():
            return output_dir
        def mock_get_program_dir():
            return program_dir

        monkeypatch.setattr('bankcheck.get_output_dir', mock_get_output_dir)
        monkeypatch.setattr('bankcheck.get_program_dir', mock_get_program_dir)

        result = bankcheck.find_lookup_file()
        expected_copied = os.path.join(output_dir, '主体查找表.xlsx')
        assert result == expected_copied
        assert os.path.exists(expected_copied)

    def test_no_lookup_file_returns_none(self, tmp_dir, monkeypatch):
        output_dir = os.path.join(tmp_dir, 'output')
        program_dir = os.path.join(tmp_dir, 'program')
        os.makedirs(output_dir)
        os.makedirs(program_dir)

        def mock_get_output_dir():
            return output_dir
        def mock_get_program_dir():
            return program_dir

        monkeypatch.setattr('bankcheck.get_output_dir', mock_get_output_dir)
        monkeypatch.setattr('bankcheck.get_program_dir', mock_get_program_dir)

        result = bankcheck.find_lookup_file()
        assert result is None

    def test_custom_script_dir_only_searches_that_dir(self, tmp_dir, monkeypatch):
        output_dir = os.path.join(tmp_dir, 'output')
        program_dir = os.path.join(tmp_dir, 'program')
        custom_dir = os.path.join(tmp_dir, 'custom')
        other_dir = os.path.join(tmp_dir, 'other')
        os.makedirs(output_dir)
        os.makedirs(program_dir)
        os.makedirs(custom_dir)
        os.makedirs(other_dir)

        other_lookup = os.path.join(other_dir, '主体查找表.xlsx')
        wb = openpyxl.Workbook()
        wb.save(other_lookup)
        wb.close()

        def mock_get_output_dir():
            return output_dir
        def mock_get_program_dir():
            return program_dir

        monkeypatch.setattr('bankcheck.get_output_dir', mock_get_output_dir)
        monkeypatch.setattr('bankcheck.get_program_dir', mock_get_program_dir)

        result = bankcheck.find_lookup_file(script_dir=custom_dir)
        assert result is None

        custom_lookup = os.path.join(custom_dir, '主体查找表.xlsx')
        wb = openpyxl.Workbook()
        wb.save(custom_lookup)
        wb.close()

        result = bankcheck.find_lookup_file(script_dir=custom_dir)
        assert result == custom_lookup

    def test_custom_script_dir_with_lookup_only_in_other_dir(self, tmp_dir, monkeypatch):
        output_dir = os.path.join(tmp_dir, 'output')
        program_dir = os.path.join(tmp_dir, 'program')
        custom_dir = os.path.join(tmp_dir, 'custom')
        os.makedirs(output_dir)
        os.makedirs(program_dir)
        os.makedirs(custom_dir)

        program_lookup = os.path.join(program_dir, '主体查找表.xlsx')
        wb = openpyxl.Workbook()
        wb.save(program_lookup)
        wb.close()

        def mock_get_output_dir():
            return output_dir
        def mock_get_program_dir():
            return program_dir

        monkeypatch.setattr('bankcheck.get_output_dir', mock_get_output_dir)
        monkeypatch.setattr('bankcheck.get_program_dir', mock_get_program_dir)

        result = bankcheck.find_lookup_file(script_dir=custom_dir)
        assert result is None

        result_no_arg = bankcheck.find_lookup_file()
        assert result_no_arg == os.path.join(output_dir, '主体查找表.xlsx')


class TestSetupLogging:
    def test_returns_logger(self):
        logger = bankcheck.setup_logging()
        assert logger is not None
        assert logger.name == 'bankcheck'

    def test_logger_has_handlers(self):
        logger = bankcheck.setup_logging()
        assert len(logger.handlers) >= 1

    def test_get_logger(self):
        logger = bankcheck.get_logger()
        assert logger.name == 'bankcheck'


class TestCLIFunctions:
    def test_cli_showinfo(self, capsys):
        bankcheck.cli_showinfo('Title', 'Message')
        captured = capsys.readouterr()
        assert 'Title' in captured.out
        assert 'Message' in captured.out

    def test_cli_showwarning(self, capsys):
        bankcheck.cli_showwarning('Title', 'Message')
        captured = capsys.readouterr()
        assert 'Title' in captured.out
        assert 'Message' in captured.out

    def test_cli_askdirectory_valid(self, tmp_dir, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: tmp_dir)
        result = bankcheck.cli_askdirectory()
        assert result == tmp_dir

    def test_cli_askdirectory_invalid(self, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: '/nonexistent/path')
        result = bankcheck.cli_askdirectory()
        assert result == ''

    def test_cli_askdirectory_empty(self, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: '')
        result = bankcheck.cli_askdirectory()
        assert result == ''

    def test_cli_askdirectory_returns_absolute_path(self, tmp_dir, monkeypatch):
        rel_path = os.path.relpath(tmp_dir, os.getcwd())
        monkeypatch.setattr('builtins.input', lambda _: rel_path)
        result = bankcheck.cli_askdirectory()
        assert os.path.isabs(result)
        assert os.path.isdir(result)

    def test_cli_askdirectory_with_quotes(self, tmp_dir, monkeypatch):
        quoted = f'"{tmp_dir}"'
        monkeypatch.setattr('builtins.input', lambda _: quoted)
        result = bankcheck.cli_askdirectory()
        assert result == os.path.abspath(tmp_dir)

    def test_cli_askdirectory_default_path_valid(self, tmp_dir):
        result = bankcheck.cli_askdirectory(default_path=tmp_dir)
        assert result == os.path.abspath(tmp_dir)

    def test_cli_askdirectory_default_path_invalid_then_input(self, tmp_dir, monkeypatch):
        inputs = [tmp_dir]
        monkeypatch.setattr('builtins.input', lambda _: inputs.pop(0))
        result = bankcheck.cli_askdirectory(default_path='/nonexistent/path')
        assert result == os.path.abspath(tmp_dir)

    def test_cli_askdirectory_quit_with_q(self, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: 'q')
        result = bankcheck.cli_askdirectory()
        assert result == ''

    def test_cli_askdirectory_max_retries(self, monkeypatch):
        call_count = [0]
        def mock_input(_):
            call_count[0] += 1
            return '/nonexistent/path'
        monkeypatch.setattr('builtins.input', mock_input)
        result = bankcheck.cli_askdirectory(max_retries=2)
        assert result == ''
        assert call_count[0] == 2

    def test_cli_askfile_valid(self, tmp_dir, monkeypatch):
        test_file = os.path.join(tmp_dir, 'test.txt')
        with open(test_file, 'w') as f:
            f.write('test')
        monkeypatch.setattr('builtins.input', lambda _: test_file)
        result = bankcheck.cli_askfile()
        assert result == os.path.abspath(test_file)

    def test_cli_askfile_invalid(self, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: '/nonexistent/file.txt')
        result = bankcheck.cli_askfile()
        assert result == ''

    def test_cli_askfile_default_path_valid(self, tmp_dir):
        test_file = os.path.join(tmp_dir, 'test.txt')
        with open(test_file, 'w') as f:
            f.write('test')
        result = bankcheck.cli_askfile(default_path=test_file)
        assert result == os.path.abspath(test_file)

    def test_cli_askfile_quit_with_q(self, monkeypatch):
        monkeypatch.setattr('builtins.input', lambda _: 'q')
        result = bankcheck.cli_askfile()
        assert result == ''


class TestNormalizePath:
    def test_empty_path(self):
        assert bankcheck._normalize_path('') == ''
        assert bankcheck._normalize_path(None) == ''

    def test_strip_quotes(self):
        assert bankcheck._normalize_path('"/some/path"') == '/some/path'
        assert bankcheck._normalize_path("'/some/path'") == '/some/path'

    def test_expand_user(self, monkeypatch):
        monkeypatch.setenv('HOME', '/home/testuser')
        result = bankcheck._normalize_path('~/docs')
        assert result == os.path.abspath('/home/testuser/docs')

    def test_relative_to_absolute(self):
        result = bankcheck._normalize_path('relative/path')
        assert os.path.isabs(result)

    def test_normpath(self):
        result = bankcheck._normalize_path('/a/b/../c')
        assert result == '/a/c'


class TestCliDefaults:
    def test_set_and_get_default_dir(self, tmp_dir):
        bankcheck.set_cli_default_dir(tmp_dir)
        assert bankcheck.get_cli_default_dir() == os.path.abspath(tmp_dir)
        bankcheck.set_cli_default_dir(None)
        assert bankcheck.get_cli_default_dir() is None

    def test_set_and_get_default_file(self, tmp_dir):
        test_file = os.path.join(tmp_dir, 'test.txt')
        with open(test_file, 'w') as f:
            f.write('test')
        bankcheck.set_cli_default_file(test_file)
        assert bankcheck.get_cli_default_file() == os.path.abspath(test_file)
        bankcheck.set_cli_default_file(None)
        assert bankcheck.get_cli_default_file() is None

    def test_cli_askdirectory_uses_global_default(self, tmp_dir):
        bankcheck.set_cli_default_dir(tmp_dir)
        result = bankcheck.cli_askdirectory()
        assert result == os.path.abspath(tmp_dir)
        bankcheck.set_cli_default_dir(None)

    def test_cli_askfile_uses_global_default(self, tmp_dir):
        test_file = os.path.join(tmp_dir, 'test.txt')
        with open(test_file, 'w') as f:
            f.write('test')
        bankcheck.set_cli_default_file(test_file)
        result = bankcheck.cli_askfile()
        assert result == os.path.abspath(test_file)
        bankcheck.set_cli_default_file(None)

    def test_explicit_default_overrides_global(self, tmp_dir):
        dir1 = os.path.join(tmp_dir, 'dir1')
        dir2 = os.path.join(tmp_dir, 'dir2')
        os.makedirs(dir1)
        os.makedirs(dir2)
        bankcheck.set_cli_default_dir(dir1)
        result = bankcheck.cli_askdirectory(default_path=dir2)
        assert result == os.path.abspath(dir2)
        bankcheck.set_cli_default_dir(None)
