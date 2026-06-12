"""
利息与手续费专项核对模块单元测试

覆盖场景：
1. 关键词筛选 - 利息、手续费等关键词正确识别
2. 交易分类 - 利息收入、利息支出、各类手续费正确分类
3. 期间汇总 - 按月/季/年正确汇总
4. 费率对比 - 与预期费率对比正确计算偏差
5. 历史均值对比 - 与历史均值对比正确计算偏差
6. 环比变动 - 期间环比变动正确计算
7. 异常检测 - 超出阈值的异常正确标记
8. 导出功能 - Excel报告正确生成
9. 边界场景 - 空数据、无匹配交易、单条记录
10. 多主体多银行混合场景
"""
import os
import sys
import tempfile
import shutil
from datetime import datetime

import openpyxl
import pandas as pd
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck
from bankcheck import (
    _classify_transaction,
    _get_period_key,
    _calculate_expected_amount,
    filter_interest_fee_transactions,
    summarize_by_period,
    check_interest_fee,
    export_interest_fee_check_result,
    generate_interest_fee_check_from_records,
    InterestFeeCheckResult,
    InterestFeeTransaction,
    PeriodSummary,
    INTEREST_FEE_KEYWORDS,
    TRANSACTION_TYPE_INTEREST,
    TRANSACTION_TYPE_FEE,
    ANOMALY_THRESHOLD,
)


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='interest_fee_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _record(unique_id, account, date, payment=None, receipt=None, balance=0.0,
            bank='北京银行', subject='北京XX科技有限公司',
            summary='', counterpart='', txn_id=''):
    """构造标准化交易记录"""
    return {
        '唯一id': unique_id,
        '银行': bank,
        '银行账号': account,
        '主体': subject,
        '交易日期': date,
        '付款': payment,
        '收款': receipt,
        '摘要': summary,
        '对方户名': counterpart,
        '余额': balance,
        '交易流水号': txn_id or unique_id,
    }


class TestTransactionClassification:
    """交易分类测试"""

    def test_interest_income_classification(self):
        """利息收入正确分类"""
        trans_type, category, kw = _classify_transaction('2024年第一季度存款利息')
        assert trans_type == TRANSACTION_TYPE_INTEREST
        assert category == '利息收入'
        assert kw == '利息'

    def test_interest_expense_classification(self):
        """利息支出正确分类"""
        trans_type, category, kw = _classify_transaction('2024年3月贷款利息支出')
        assert trans_type == TRANSACTION_TYPE_INTEREST
        assert category == '利息支出'
        assert kw == '利息'

    def test_general_interest_classification(self):
        """普通利息默认分类为利息收入"""
        trans_type, category, kw = _classify_transaction('结息')
        assert trans_type == TRANSACTION_TYPE_INTEREST
        assert category == '利息收入'

    def test_transfer_fee_classification(self):
        """转账手续费正确分类"""
        trans_type, category, kw = _classify_transaction('跨行转账手续费')
        assert trans_type == TRANSACTION_TYPE_FEE
        assert category == '转账手续费'
        assert kw == '手续费'

    def test_withdrawal_fee_classification(self):
        """提现手续费正确分类"""
        trans_type, category, kw = _classify_transaction('ATM提现手续费')
        assert trans_type == TRANSACTION_TYPE_FEE
        assert category == '提现手续费'

    def test_payment_fee_classification(self):
        """支付手续费正确分类"""
        trans_type, category, kw = _classify_transaction('快捷支付手续费')
        assert trans_type == TRANSACTION_TYPE_FEE
        assert category == '支付手续费'

    def test_account_management_fee_classification(self):
        """账户管理费正确分类"""
        trans_type, category, kw = _classify_transaction('季度账户管理费')
        assert trans_type == TRANSACTION_TYPE_FEE
        assert category == '账户管理费'

    def test_annual_fee_classification(self):
        """年费正确分类"""
        trans_type, category, kw = _classify_transaction('银行卡年费')
        assert trans_type == TRANSACTION_TYPE_FEE
        assert category == '账户管理费'

    def test_online_banking_fee_classification(self):
        """网银服务费正确分类"""
        trans_type, category, kw = _classify_transaction('网银服务费')
        assert trans_type == TRANSACTION_TYPE_FEE
        assert category == '网银服务费'

    def test_sms_fee_classification(self):
        """短信服务费正确分类"""
        trans_type, category, kw = _classify_transaction('短信通知服务费')
        assert trans_type == TRANSACTION_TYPE_FEE
        assert category == '短信服务费'

    def test_other_fee_classification(self):
        """其他手续费正确分类"""
        trans_type, category, kw = _classify_transaction('银行工本费')
        assert trans_type == TRANSACTION_TYPE_FEE
        assert category == '其他手续费'

    def test_english_keyword_classification(self):
        """英文关键词正确分类"""
        trans_type, category, kw = _classify_transaction('Monthly account fee')
        assert trans_type == TRANSACTION_TYPE_FEE
        assert category != ''
        assert kw.lower() in ['fee', 'interest']

    def test_non_matching_summary(self):
        """非利息手续费摘要返回空"""
        trans_type, category, kw = _classify_transaction('采购货款')
        assert trans_type == ''
        assert category == ''
        assert kw == ''

    def test_empty_summary(self):
        """空摘要返回空"""
        trans_type, category, kw = _classify_transaction('')
        assert trans_type == ''
        assert category == ''
        assert kw == ''

    def test_none_summary(self):
        """None摘要返回空"""
        trans_type, category, kw = _classify_transaction(None)
        assert trans_type == ''
        assert category == ''
        assert kw == ''


class TestPeriodKey:
    """期间键值生成测试"""

    def test_month_period(self):
        """按月生成期间键"""
        dt = datetime(2024, 3, 15)
        assert _get_period_key(dt, 'month') == '2024-03'

    def test_quarter_period_q1(self):
        """按季度生成期间键 - Q1"""
        dt = datetime(2024, 2, 15)
        assert _get_period_key(dt, 'quarter') == '2024Q1'

    def test_quarter_period_q3(self):
        """按季度生成期间键 - Q3"""
        dt = datetime(2024, 8, 15)
        assert _get_period_key(dt, 'quarter') == '2024Q3'

    def test_year_period(self):
        """按年生成期间键"""
        dt = datetime(2024, 6, 15)
        assert _get_period_key(dt, 'year') == '2024'

    def test_week_period(self):
        """按周生成期间键"""
        dt = datetime(2024, 1, 1)
        period = _get_period_key(dt, 'week')
        assert period.startswith('2024-W')

    def test_none_date(self):
        """None日期返回未知"""
        assert _get_period_key(None, 'month') == '未知'

    def test_default_period_type(self):
        """默认期间类型为月"""
        dt = datetime(2024, 3, 15)
        assert _get_period_key(dt) == '2024-03'


class TestExpectedAmountCalculation:
    """预期金额计算测试"""

    def test_amount_based_expected(self):
        """基于交易金额的费率计算"""
        expected = _calculate_expected_amount('转账手续费', 10000, 'month')
        assert expected == pytest.approx(50.0)

    def test_amount_based_expected_with_min(self):
        """最小费用限制"""
        expected = _calculate_expected_amount('转账手续费', 100, 'month')
        assert expected == pytest.approx(1.0)

    def test_amount_based_expected_with_max(self):
        """最大费用限制"""
        expected = _calculate_expected_amount('转账手续费', 1000000, 'month')
        assert expected == pytest.approx(50.0)

    def test_monthly_fixed_fee(self):
        """月度固定费用"""
        expected = _calculate_expected_amount('账户管理费', 0, 'month')
        assert expected == pytest.approx(10.0)

    def test_yearly_fixed_fee_year_period(self):
        """年度固定费用（按年汇总）"""
        expected = _calculate_expected_amount('年费', 0, 'year')
        assert expected == pytest.approx(100.0)

    def test_yearly_fixed_fee_month_period(self):
        """年度固定费用（按月汇总时分摊）"""
        expected = _calculate_expected_amount('年费', 0, 'month')
        assert expected == pytest.approx(100.0 / 12)

    def test_unknown_category(self):
        """未知类别返回0"""
        expected = _calculate_expected_amount('未知类别', 1000, 'month')
        assert expected == pytest.approx(0.0)


class TestFilterTransactions:
    """交易筛选测试"""

    def test_filter_mixed_transactions(self):
        """筛选混合交易，只保留利息手续费"""
        records = [
            _record('T001', '62220001', '2024-01-05', receipt=50000, summary='销售收款'),
            _record('T002', '62220001', '2024-01-10', payment=-1000, summary='采购付款'),
            _record('T003', '62220001', '2024-01-15', receipt=125.50, summary='2024年第一季度存款利息'),
            _record('T004', '62220001', '2024-01-20', payment=-50, summary='跨行转账手续费'),
            _record('T005', '62220001', '2024-01-25', receipt=80000, summary='客户回款'),
        ]
        result = filter_interest_fee_transactions(records)
        assert len(result) == 2
        assert result[0].fee_category == '利息收入'
        assert result[0].amount == pytest.approx(125.50)
        assert result[1].fee_category == '转账手续费'
        assert result[1].amount == pytest.approx(50.0)

    def test_filter_no_matching_transactions(self):
        """无匹配交易返回空列表"""
        records = [
            _record('T001', '62220001', '2024-01-05', receipt=50000, summary='销售收款'),
            _record('T002', '62220001', '2024-01-10', payment=-1000, summary='采购付款'),
        ]
        result = filter_interest_fee_transactions(records)
        assert len(result) == 0

    def test_filter_empty_records(self):
        """空记录列表返回空"""
        result = filter_interest_fee_transactions([])
        assert len(result) == 0

    def test_filter_amount_calculation_receipt(self):
        """收款金额正确提取"""
        records = [
            _record('T001', '62220001', '2024-01-15', receipt=250.75, summary='季度结息'),
        ]
        result = filter_interest_fee_transactions(records)
        assert len(result) == 1
        assert result[0].amount == pytest.approx(250.75)

    def test_filter_amount_calculation_payment(self):
        """付款金额取绝对值"""
        records = [
            _record('T001', '62220001', '2024-01-20', payment=-25.50, summary='转账手续费'),
        ]
        result = filter_interest_fee_transactions(records)
        assert len(result) == 1
        assert result[0].amount == pytest.approx(25.50)


class TestSummarizeByPeriod:
    """期间汇总测试"""

    def test_monthly_summary_single_category(self):
        """按月汇总单一类别"""
        transactions = [
            InterestFeeTransaction(
                transaction_id='T001',
                trade_date=datetime(2024, 1, 5),
                bank='北京银行',
                bank_account='62220001',
                subject='北京XX科技',
                summary='转账手续费',
                amount=50.0,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='转账手续费',
                matched_keyword='手续费',
            ),
            InterestFeeTransaction(
                transaction_id='T002',
                trade_date=datetime(2024, 1, 15),
                bank='北京银行',
                bank_account='62220001',
                subject='北京XX科技',
                summary='转账手续费',
                amount=30.0,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='转账手续费',
                matched_keyword='手续费',
            ),
            InterestFeeTransaction(
                transaction_id='T003',
                trade_date=datetime(2024, 2, 10),
                bank='北京银行',
                bank_account='62220001',
                subject='北京XX科技',
                summary='转账手续费',
                amount=40.0,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='转账手续费',
                matched_keyword='手续费',
            ),
        ]
        summaries = summarize_by_period(transactions, 'month')
        assert len(summaries) == 2

        jan_summary = [s for s in summaries if s.period == '2024-01'][0]
        assert jan_summary.transaction_count == 2
        assert jan_summary.total_amount == pytest.approx(80.0)
        assert jan_summary.avg_amount == pytest.approx(40.0)
        assert jan_summary.max_amount == pytest.approx(50.0)
        assert jan_summary.min_amount == pytest.approx(30.0)

        feb_summary = [s for s in summaries if s.period == '2024-02'][0]
        assert feb_summary.transaction_count == 1
        assert feb_summary.total_amount == pytest.approx(40.0)

    def test_quarterly_summary(self):
        """按季度汇总"""
        transactions = [
            InterestFeeTransaction(
                transaction_id='T001',
                trade_date=datetime(2024, 1, 5),
                bank='北京银行',
                bank_account='62220001',
                subject='北京XX科技',
                summary='账户管理费',
                amount=10.0,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='账户管理费',
                matched_keyword='管理费',
            ),
            InterestFeeTransaction(
                transaction_id='T002',
                trade_date=datetime(2024, 2, 5),
                bank='北京银行',
                bank_account='62220001',
                subject='北京XX科技',
                summary='账户管理费',
                amount=10.0,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='账户管理费',
                matched_keyword='管理费',
            ),
            InterestFeeTransaction(
                transaction_id='T003',
                trade_date=datetime(2024, 3, 5),
                bank='北京银行',
                bank_account='62220001',
                subject='北京XX科技',
                summary='账户管理费',
                amount=10.0,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='账户管理费',
                matched_keyword='管理费',
            ),
        ]
        summaries = summarize_by_period(transactions, 'quarter')
        assert len(summaries) == 1
        assert summaries[0].period == '2024Q1'
        assert summaries[0].transaction_count == 3
        assert summaries[0].total_amount == pytest.approx(30.0)

    def test_multiple_subjects_and_banks(self):
        """多主体多银行独立汇总"""
        transactions = [
            InterestFeeTransaction(
                transaction_id='T001',
                trade_date=datetime(2024, 1, 5),
                bank='北京银行',
                bank_account='62220001',
                subject='公司A',
                summary='转账手续费',
                amount=50.0,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='转账手续费',
                matched_keyword='手续费',
            ),
            InterestFeeTransaction(
                transaction_id='T002',
                trade_date=datetime(2024, 1, 5),
                bank='工商银行',
                bank_account='62220002',
                subject='公司B',
                summary='转账手续费',
                amount=100.0,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='转账手续费',
                matched_keyword='手续费',
            ),
        ]
        summaries = summarize_by_period(transactions, 'month')
        assert len(summaries) == 2

    def test_empty_transactions(self):
        """空交易列表返回空汇总"""
        summaries = summarize_by_period([], 'month')
        assert len(summaries) == 0


class TestAnomalyDetection:
    """异常检测测试"""

    def test_deviation_from_expected_rate(self):
        """与预期费率偏差超出阈值标记异常"""
        transactions = []
        for i in range(5):
            for month in range(1, 4):
                txn = InterestFeeTransaction(
                    transaction_id=f'T{i:03d}{month:02d}',
                    trade_date=datetime(2024, month, 5),
                    bank='北京银行',
                    bank_account='62220001',
                    subject='北京XX科技',
                    summary='转账手续费',
                    amount=5.0 if month < 3 else 100.0,
                    counterpart='',
                    transaction_type=TRANSACTION_TYPE_FEE,
                    fee_category='转账手续费',
                    matched_keyword='手续费',
                )
                transactions.append(txn)

        summaries = summarize_by_period(transactions, 'month')
        march_summary = [s for s in summaries if s.period == '2024-03'][0]
        assert march_summary.is_anomaly == True
        assert any('预期费率' in r for r in march_summary.anomaly_reasons)

    def test_deviation_from_historical_mean(self):
        """与历史均值偏差超出阈值标记异常"""
        transactions = []
        amounts = [10.0, 12.0, 9.0, 11.0, 10.5, 50.0]
        for i, amount in enumerate(amounts):
            txn = InterestFeeTransaction(
                transaction_id=f'T{i:03d}',
                trade_date=datetime(2024, i + 1, 5),
                bank='北京银行',
                bank_account='62220001',
                subject='北京XX科技',
                summary='账户管理费',
                amount=amount,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='账户管理费',
                matched_keyword='管理费',
            )
            transactions.append(txn)

        summaries = summarize_by_period(transactions, 'month')
        june_summary = [s for s in summaries if s.period == '2024-06'][0]
        assert june_summary.is_anomaly == True
        assert any('历史均值' in r for r in june_summary.anomaly_reasons)

    def test_period_over_period_change(self):
        """环比变动超出阈值标记异常"""
        transactions = []
        for month in range(1, 4):
            amount = 10.0 if month < 3 else 30.0
            txn = InterestFeeTransaction(
                transaction_id=f'T{month:02d}',
                trade_date=datetime(2024, month, 5),
                bank='北京银行',
                bank_account='62220001',
                subject='北京XX科技',
                summary='账户管理费',
                amount=amount,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='账户管理费',
                matched_keyword='管理费',
            )
            transactions.append(txn)

        summaries = summarize_by_period(transactions, 'month')
        march_summary = [s for s in summaries if s.period == '2024-03'][0]
        assert march_summary.is_anomaly == True
        assert any('环比' in r for r in march_summary.anomaly_reasons)

    def test_no_anomaly_normal_fluctuation(self):
        """正常波动不标记异常"""
        transactions = []
        amounts = [10.0, 10.5, 9.8, 10.2, 10.1, 9.9]
        for i, amount in enumerate(amounts):
            txn = InterestFeeTransaction(
                transaction_id=f'T{i:03d}',
                trade_date=datetime(2024, i + 1, 5),
                bank='北京银行',
                bank_account='62220001',
                subject='北京XX科技',
                summary='账户管理费',
                amount=amount,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='账户管理费',
                matched_keyword='管理费',
            )
            transactions.append(txn)

        summaries = summarize_by_period(transactions, 'month')
        for s in summaries:
            assert s.is_anomaly == False

    def test_single_period_no_historical_comparison(self):
        """单一期间无历史对比，不标记历史均值异常"""
        transactions = [
            InterestFeeTransaction(
                transaction_id='T001',
                trade_date=datetime(2024, 1, 5),
                bank='北京银行',
                bank_account='62220001',
                subject='北京XX科技',
                summary='账户管理费',
                amount=10.0,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='账户管理费',
                matched_keyword='管理费',
            ),
        ]
        summaries = summarize_by_period(transactions, 'month')
        assert len(summaries) == 1
        assert summaries[0].is_anomaly == False


class TestCheckInterestFee:
    """核心核对功能测试"""

    def test_full_check_workflow(self):
        """完整核对流程"""
        records = [
            _record('T001', '62220001', '2024-01-05', receipt=50000, summary='销售收款'),
            _record('T002', '62220001', '2024-01-15', receipt=125.50, summary='季度存款利息'),
            _record('T003', '62220001', '2024-01-20', payment=-50, summary='跨行转账手续费'),
            _record('T004', '62220001', '2024-02-10', payment=-30, summary='转账手续费'),
            _record('T005', '62220001', '2024-02-15', receipt=200.0, summary='贷款利息支出'),
            _record('T006', '62220001', '2024-03-05', payment=-10, summary='账户管理费'),
        ]
        result = check_interest_fee(records, period_type='month')

        assert result.total_records == 5
        assert result.interest_records == 2
        assert result.fee_records == 3
        assert result.total_interest_amount == pytest.approx(325.50)
        assert result.total_fee_amount == pytest.approx(90.0)
        assert len(result.period_summaries) >= 3

    def test_check_empty_records(self):
        """空记录核对"""
        result = check_interest_fee([], period_type='month')
        assert result.total_records == 0
        assert result.check_summary.get('status') == '无数据'

    def test_check_no_matching_records(self):
        """无匹配记录核对"""
        records = [
            _record('T001', '62220001', '2024-01-05', receipt=50000, summary='销售收款'),
            _record('T002', '62220001', '2024-01-10', payment=-1000, summary='采购付款'),
        ]
        result = check_interest_fee(records, period_type='month')
        assert result.total_records == 0
        assert len(result.filtered_transactions) == 0

    def test_check_quarterly_period(self):
        """按季度核对"""
        records = [
            _record('T001', '62220001', '2024-01-15', receipt=125.50, summary='季度存款利息'),
            _record('T002', '62220001', '2024-02-20', payment=-50, summary='转账手续费'),
            _record('T003', '62220001', '2024-03-10', payment=-30, summary='转账手续费'),
        ]
        result = check_interest_fee(records, period_type='quarter')
        assert result.total_records == 3
        assert len(result.period_summaries) >= 1
        assert result.period_summaries[0].period == '2024Q1'


class TestExportFunction:
    """导出功能测试"""

    def test_export_result_to_excel(self, tmp_dir):
        """导出核对结果到Excel"""
        records = [
            _record('T001', '62220001', '2024-01-15', receipt=125.50, summary='季度存款利息'),
            _record('T002', '62220001', '2024-01-20', payment=-50, summary='跨行转账手续费'),
            _record('T003', '62220001', '2024-02-10', payment=-30, summary='转账手续费'),
        ]
        check_result = check_interest_fee(records, period_type='month')

        output_path = os.path.join(tmp_dir, 'test_interest_fee_report.xlsx')
        source_info = {'数据来源': '测试数据', '记录数': len(records)}

        result_path = export_interest_fee_check_result(check_result, output_path, source_info)

        assert os.path.exists(result_path)
        assert os.path.getsize(result_path) > 0

        wb = openpyxl.load_workbook(result_path)
        assert '核对总览' in wb.sheetnames
        assert '交易明细' in wb.sheetnames
        assert '期间汇总' in wb.sheetnames
        assert '异常清单' in wb.sheetnames

        ws_overview = wb['核对总览']
        assert ws_overview.cell(row=1, column=1).value == '核对项'
        assert ws_overview.cell(row=1, column=2).value == '数值'

        ws_detail = wb['交易明细']
        assert ws_detail.max_row == 4

        ws_period = wb['期间汇总']
        assert ws_period.max_row >= 3

        wb.close()

    def test_generate_from_records(self, tmp_dir):
        """从记录列表生成报告"""
        records = [
            _record('T001', '62220001', '2024-01-15', receipt=125.50, summary='季度存款利息'),
            _record('T002', '62220001', '2024-01-20', payment=-50, summary='跨行转账手续费'),
        ]

        result_path = generate_interest_fee_check_from_records(
            records, output_dir=tmp_dir, period_type='month'
        )

        assert result_path is not None
        assert os.path.exists(result_path)
        assert '利息手续费核对报告' in os.path.basename(result_path)

    def test_generate_from_records_empty(self, tmp_dir):
        """空记录生成报告返回None"""
        result_path = generate_interest_fee_check_from_records(
            [], output_dir=tmp_dir, period_type='month'
        )
        assert result_path is None

    def test_export_with_anomalies(self, tmp_dir):
        """导出包含异常的报告"""
        transactions = []
        amounts = [10.0, 10.0, 10.0, 10.0, 10.0, 100.0]
        for i, amount in enumerate(amounts):
            txn = InterestFeeTransaction(
                transaction_id=f'T{i:03d}',
                trade_date=datetime(2024, i + 1, 5),
                bank='北京银行',
                bank_account='62220001',
                subject='北京XX科技',
                summary='账户管理费',
                amount=amount,
                counterpart='',
                transaction_type=TRANSACTION_TYPE_FEE,
                fee_category='账户管理费',
                matched_keyword='管理费',
            )
            transactions.append(txn)

        summaries = summarize_by_period(transactions, 'month')
        check_result = InterestFeeCheckResult(
            total_records=len(transactions),
            fee_records=len(transactions),
            total_fee_amount=sum(t.amount for t in transactions),
            filtered_transactions=transactions,
            period_summaries=summaries,
            anomaly_summaries=[s for s in summaries if s.is_anomaly],
            anomaly_count=sum(1 for s in summaries if s.is_anomaly),
            check_summary={'test': 'data'},
        )

        output_path = os.path.join(tmp_dir, 'test_anomaly_report.xlsx')
        result_path = export_interest_fee_check_result(check_result, output_path)

        assert os.path.exists(result_path)

        wb = openpyxl.load_workbook(result_path)
        ws_anomaly = wb['异常清单']
        assert ws_anomaly.max_row >= 2
        wb.close()


class TestMultipleSubjectsBanks:
    """多主体多银行场景测试"""

    def test_multiple_subjects_separate_summary(self):
        """不同主体独立汇总"""
        records = [
            _record('T001', '62220001', '2024-01-15', receipt=125.50,
                    summary='季度存款利息', subject='公司A', bank='北京银行'),
            _record('T002', '62220002', '2024-01-15', receipt=250.00,
                    summary='季度存款利息', subject='公司B', bank='工商银行'),
        ]
        result = check_interest_fee(records, period_type='month')
        assert len(result.period_summaries) == 2
        subjects = {s.subject for s in result.period_summaries}
        assert subjects == {'公司A', '公司B'}

    def test_multiple_banks_same_subject(self):
        """同一主体不同银行独立汇总"""
        records = [
            _record('T001', '62220001', '2024-01-15', payment=-50,
                    summary='转账手续费', subject='公司A', bank='北京银行'),
            _record('T002', '62220002', '2024-01-15', payment=-30,
                    summary='转账手续费', subject='公司A', bank='工商银行'),
        ]
        result = check_interest_fee(records, period_type='month')
        assert len(result.period_summaries) == 2
        banks = {s.bank for s in result.period_summaries}
        assert banks == {'北京银行', '工商银行'}


class TestEdgeCases:
    """边界场景测试"""

    def test_none_date_transaction(self):
        """交易日期为None的情况"""
        records = [
            _record('T001', '62220001', None, receipt=125.50, summary='季度存款利息'),
        ]
        result = check_interest_fee(records, period_type='month')
        assert result.total_records == 1
        unknown_summaries = [s for s in result.period_summaries if s.period == '未知']
        assert len(unknown_summaries) == 1

    def test_mixed_valid_invalid_dates(self):
        """混合有效和无效日期"""
        records = [
            _record('T001', '62220001', '2024-01-15', receipt=125.50, summary='季度存款利息'),
            _record('T002', '62220001', None, payment=-50, summary='转账手续费'),
        ]
        result = check_interest_fee(records, period_type='month')
        assert result.total_records == 2
        periods = {s.period for s in result.period_summaries}
        assert '2024-01' in periods
        assert '未知' in periods

    def test_keyword_case_insensitive(self):
        """关键词大小写不敏感"""
        records = [
            _record('T001', '62220001', '2024-01-15', receipt=125.50, summary='INTEREST INCOME'),
            _record('T002', '62220001', '2024-01-20', payment=-50, summary='Transaction FEE'),
        ]
        result = filter_interest_fee_transactions(records)
        assert len(result) == 2

    def test_very_large_amount(self):
        """处理大额交易"""
        records = [
            _record('T001', '62220001', '2024-01-15', receipt=999999999.99, summary='大额利息'),
        ]
        result = check_interest_fee(records, period_type='month')
        assert result.total_records == 1
        assert result.total_interest_amount == pytest.approx(999999999.99)

    def test_zero_amount(self):
        """零金额交易"""
        records = [
            _record('T001', '62220001', '2024-01-15', receipt=0, summary='免手续费'),
        ]
        result = filter_interest_fee_transactions(records)
        assert len(result) == 1
        assert result[0].amount == pytest.approx(0.0)
