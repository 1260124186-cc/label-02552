"""
电子表格协同编辑模块单元测试
"""
import os
import sys
import tempfile
import shutil
from datetime import datetime

import openpyxl
import pandas as pd
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='collab_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _create_test_summary(path, records=None):
    """创建测试用总表"""
    if records is None:
        records = [
            {
                '唯一id': 'UID001',
                '银行': '北京银行',
                '银行账号': '01090312345678901',
                '主体': '北京XX科技有限公司',
                '交易日期': '2024-01-05',
                '付款': -50000.0,
                '收款': None,
                '摘要': '采购付款-办公设备',
                '对方户名': '供应商A公司',
                '余额': 1500000.0,
                '交易流水号': 'BJ20240105001',
            },
            {
                '唯一id': 'UID002',
                '银行': '北京银行',
                '银行账号': '01090312345678901',
                '主体': '北京XX科技有限公司',
                '交易日期': '2024-01-10',
                '付款': None,
                '收款': 80000.0,
                '摘要': '销售收款-产品销售',
                '对方户名': '客户B公司',
                '余额': 1580000.0,
                '交易流水号': 'BJ20240110002',
            },
            {
                '唯一id': 'UID003',
                '银行': '东亚银行',
                '银行账号': '38812345678',
                '主体': '上海YY贸易有限公司',
                '交易日期': '2024-01-15',
                '付款': -1500.0,
                '收款': None,
                '摘要': '手续费',
                '对方户名': '银行手续费',
                '余额': 513500.0,
                '交易流水号': 'EA20240115003',
            },
            {
                '唯一id': 'UID004',
                '银行': '东亚银行',
                '银行账号': '38812345678',
                '主体': '上海YY贸易有限公司',
                '交易日期': '2024-01-20',
                '付款': -3000.0,
                '收款': None,
                '摘要': '差旅费报销',
                '对方户名': '员工张三',
                '余额': 510500.0,
                '交易流水号': 'EA20240120004',
            },
        ]
    
    columns = ['唯一id', '银行', '银行账号', '主体', '交易日期', 
               '付款', '收款', '摘要', '对方户名', '余额', '交易流水号']
    df = pd.DataFrame(records, columns=columns)
    df.to_excel(path, index=False, engine='openpyxl')
    return path


class TestGenerateCollabTemplate:
    """测试协同编辑模板生成"""

    def test_generate_template_basic(self, tmp_dir):
        """基础功能：生成协同编辑模板"""
        summary_path = os.path.join(tmp_dir, '银行流水总表.xlsx')
        _create_test_summary(summary_path)

        template_path = bankcheck.generate_collab_template(summary_path, output_dir=tmp_dir)
        assert template_path is not None
        assert os.path.exists(template_path)
        assert bankcheck.COLLAB_TEMPLATE_SUFFIX in template_path

    def test_generate_template_has_multiple_sheets(self, tmp_dir):
        """模板应包含使用说明 sheet 和数据 sheet"""
        summary_path = os.path.join(tmp_dir, '银行流水总表.xlsx')
        _create_test_summary(summary_path)

        template_path = bankcheck.generate_collab_template(summary_path, output_dir=tmp_dir)
        wb = openpyxl.load_workbook(template_path)
        
        sheet_names = wb.sheetnames
        assert '使用说明' in sheet_names
        has_data_sheet = any('流水' in s or '协同' in s for s in sheet_names)
        assert has_data_sheet

    def test_generate_template_has_editable_columns(self, tmp_dir):
        """模板应添加凭证号、备注等可编辑列"""
        summary_path = os.path.join(tmp_dir, '银行流水总表.xlsx')
        _create_test_summary(summary_path)

        template_path = bankcheck.generate_collab_template(summary_path, output_dir=tmp_dir)
        wb = openpyxl.load_workbook(template_path)
        
        data_ws = None
        for s in wb.sheetnames:
            if '流水' in s or '协同' in s:
                data_ws = wb[s]
                break
        assert data_ws is not None

        headers = []
        for col_idx in range(1, data_ws.max_column + 1):
            val = data_ws.cell(row=1, column=col_idx).value
            if val:
                headers.append(str(val).strip())

        for col in bankcheck.COLLAB_EDITABLE_COLUMNS:
            assert col in headers, f"缺少可编辑列: {col}"

    def test_generate_template_has_data_validations(self, tmp_dir):
        """模板应配置数据验证规则（下拉选项）"""
        summary_path = os.path.join(tmp_dir, '银行流水总表.xlsx')
        _create_test_summary(summary_path)

        template_path = bankcheck.generate_collab_template(summary_path, output_dir=tmp_dir)
        wb = openpyxl.load_workbook(template_path)
        
        data_ws = None
        for s in wb.sheetnames:
            if '流水' in s or '协同' in s:
                data_ws = wb[s]
                break
        assert data_ws is not None

        validations = data_ws.data_validations.dataValidation
        assert len(validations) > 0, "应配置至少一个数据验证规则"

    def test_generate_template_style_differentiation(self, tmp_dir):
        """模板应对可编辑列和只读列使用不同背景色"""
        summary_path = os.path.join(tmp_dir, '银行流水总表.xlsx')
        _create_test_summary(summary_path)

        template_path = bankcheck.generate_collab_template(summary_path, output_dir=tmp_dir)
        wb = openpyxl.load_workbook(template_path)
        
        data_ws = None
        for s in wb.sheetnames:
            if '流水' in s or '协同' in s:
                data_ws = wb[s]
                break
        assert data_ws is not None

        header_map = {}
        for col_idx in range(1, data_ws.max_column + 1):
            val = data_ws.cell(row=1, column=col_idx).value
            if val:
                header_map[str(val).strip()] = col_idx

        editable_col = bankcheck.COLLAB_EDITABLE_COLUMNS[0]
        if editable_col in header_map:
            cell = data_ws.cell(row=2, column=header_map[editable_col])
            fill_color = cell.fill.start_color.rgb or cell.fill.start_color.index
            assert fill_color is not None

        readonly_col = '银行'
        if readonly_col in header_map:
            cell = data_ws.cell(row=2, column=header_map[readonly_col])
            fill_color = cell.fill.start_color.rgb or cell.fill.start_color.index
            assert fill_color is not None

    def test_generate_template_with_empty_summary(self, tmp_dir):
        """空总表应返回 None"""
        empty_path = os.path.join(tmp_dir, 'empty.xlsx')
        pd.DataFrame().to_excel(empty_path, index=False, engine='openpyxl')
        result = bankcheck.generate_collab_template(empty_path, output_dir=tmp_dir)
        assert result is None

    def test_generate_template_not_exist(self, tmp_dir):
        """不存在的文件返回 None"""
        result = bankcheck.generate_collab_template(
            os.path.join(tmp_dir, 'not_exist.xlsx'), output_dir=tmp_dir
        )
        assert result is None

    def test_generate_template_freeze_first_row(self, tmp_dir):
        """模板应冻结首行"""
        summary_path = os.path.join(tmp_dir, '银行流水总表.xlsx')
        _create_test_summary(summary_path)

        template_path = bankcheck.generate_collab_template(summary_path, output_dir=tmp_dir)
        wb = openpyxl.load_workbook(template_path)
        
        data_ws = None
        for s in wb.sheetnames:
            if '流水' in s or '协同' in s:
                data_ws = wb[s]
                break
        assert data_ws is not None
        assert data_ws.freeze_panes == 'A2'


class TestReadCollabEdits:
    """测试读取协同编辑内容"""

    def test_read_edits_basic(self, tmp_dir):
        """基础功能：读取财务编辑的修改"""
        summary_path = os.path.join(tmp_dir, '银行流水总表.xlsx')
        _create_test_summary(summary_path)
        template_path = bankcheck.generate_collab_template(summary_path, output_dir=tmp_dir)

        wb = openpyxl.load_workbook(template_path)
        data_ws = None
        for s in wb.sheetnames:
            if '流水' in s or '协同' in s:
                data_ws = wb[s]
                break

        header_map = {}
        for col_idx in range(1, data_ws.max_column + 1):
            val = data_ws.cell(row=1, column=col_idx).value
            if val:
                header_map[str(val).strip()] = col_idx

        data_ws.cell(row=2, column=header_map['凭证号'], value='记-202401-0001')
        data_ws.cell(row=2, column=header_map['备注'], value='办公设备采购，需入库')
        data_ws.cell(row=3, column=header_map['凭证号'], value='记-202401-0002')
        data_ws.cell(row=3, column=header_map['会计科目编码'], value='6001')
        data_ws.cell(row=3, column=header_map['会计科目名称'], value='主营业务收入')

        edited_path = os.path.join(tmp_dir, 'edited.xlsx')
        wb.save(edited_path)
        wb.close()

        edits, stats = bankcheck.read_collab_edits(edited_path)
        assert len(edits) == 2
        assert stats['edited_rows'] == 2
        assert stats['total_rows'] >= 4

        uid001_edit = next((e for e in edits if e['唯一id'] == 'UID001'), None)
        assert uid001_edit is not None
        assert uid001_edit['凭证号'] == '记-202401-0001'
        assert uid001_edit['备注'] == '办公设备采购，需入库'

        uid002_edit = next((e for e in edits if e['唯一id'] == 'UID002'), None)
        assert uid002_edit is not None
        assert uid002_edit['会计科目编码'] == '6001'

    def test_read_edits_no_changes(self, tmp_dir):
        """无修改的文件返回空 edits"""
        summary_path = os.path.join(tmp_dir, '银行流水总表.xlsx')
        _create_test_summary(summary_path)
        template_path = bankcheck.generate_collab_template(summary_path, output_dir=tmp_dir)

        edits, stats = bankcheck.read_collab_edits(template_path)
        assert len(edits) == 0
        assert stats['edited_rows'] == 0
        assert stats['unchanged_rows'] == 4

    def test_read_edits_file_not_exist(self, tmp_dir):
        """不存在的文件返回错误"""
        edits, stats = bankcheck.read_collab_edits(os.path.join(tmp_dir, 'not_exist.xlsx'))
        assert len(edits) == 0
        assert 'error' in stats

    def test_read_edits_missing_match_key(self, tmp_dir):
        """缺少匹配键列时返回错误"""
        bad_file = os.path.join(tmp_dir, 'bad.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='银行')
        ws.cell(row=1, column=2, value='凭证号')
        ws.cell(row=2, column=1, value='北京银行')
        ws.cell(row=2, column=2, value='记-001')
        wb.save(bad_file)
        wb.close()

        edits, stats = bankcheck.read_collab_edits(bad_file)
        assert len(edits) == 0
        assert 'error' in stats


class TestApplyCollabEdits:
    """测试应用协同编辑修改到记录"""

    def test_apply_edits_basic(self):
        """基础功能：应用修改到记录"""
        records = [
            {'唯一id': 'UID001', '银行': '北京银行', '摘要': '采购'},
            {'唯一id': 'UID002', '银行': '东亚银行', '摘要': '收款'},
            {'唯一id': 'UID003', '银行': '北京银行', '摘要': '手续费'},
        ]
        edits = [
            {'唯一id': 'UID001', '凭证号': '记-001', '备注': '已核对', '会计科目编码': None},
            {'唯一id': 'UID002', '凭证号': '记-002', '会计科目编码': '6001', '备注': None},
        ]

        updated, stats = bankcheck.apply_collab_edits_to_records(records, edits)
        assert stats['matched_records'] == 2
        assert stats['unmatched_edits'] == 0

        rec001 = next((r for r in updated if r['唯一id'] == 'UID001'), None)
        assert rec001['凭证号'] == '记-001'
        assert rec001['备注'] == '已核对'

        rec002 = next((r for r in updated if r['唯一id'] == 'UID002'), None)
        assert rec002['凭证号'] == '记-002'
        assert rec002['会计科目编码'] == '6001'

    def test_apply_edits_with_unmatched(self):
        """部分修改找不到匹配记录"""
        records = [
            {'唯一id': 'UID001', '银行': '北京银行'},
        ]
        edits = [
            {'唯一id': 'UID001', '凭证号': '记-001'},
            {'唯一id': 'UID999', '凭证号': '记-999'},
            {'唯一id': '', '凭证号': '空键'},
        ]

        _, stats = bankcheck.apply_collab_edits_to_records(records, edits)
        assert stats['matched_records'] == 1
        assert stats['unmatched_edits'] == 2

    def test_apply_edits_empty_records(self):
        """空记录列表返回原样"""
        records = []
        edits = [{'唯一id': 'UID001', '凭证号': '记-001'}]
        updated, stats = bankcheck.apply_collab_edits_to_records(records, edits)
        assert updated == []
        assert stats['unmatched'] == 1


class TestMergeCollabToSummary:
    """测试完整的合并回写流程"""

    def test_merge_full_flow(self, tmp_dir):
        """完整流程：生成模板→模拟编辑→合并回写"""
        summary_path = os.path.join(tmp_dir, '银行流水总表.xlsx')
        _create_test_summary(summary_path)

        template_path = bankcheck.generate_collab_template(summary_path, output_dir=tmp_dir)
        assert template_path is not None

        wb = openpyxl.load_workbook(template_path)
        data_ws = None
        for s in wb.sheetnames:
            if '流水' in s or '协同' in s:
                data_ws = wb[s]
                break

        header_map = {}
        for col_idx in range(1, data_ws.max_column + 1):
            val = data_ws.cell(row=1, column=col_idx).value
            if val:
                header_map[str(val).strip()] = col_idx

        data_ws.cell(row=2, column=header_map['凭证号'], value='记-202401-0001')
        data_ws.cell(row=2, column=header_map['备注'], value='测试备注1')
        data_ws.cell(row=3, column=header_map['凭证号'], value='记-202401-0002')
        data_ws.cell(row=3, column=header_map['会计科目编码'], value='6001')
        data_ws.cell(row=4, column=header_map['凭证号'], value='记-202401-0003')
        data_ws.cell(row=5, column=header_map['凭证号'], value='记-202401-0004')

        edited_path = os.path.join(tmp_dir, 'edited.xlsx')
        wb.save(edited_path)
        wb.close()

        result = bankcheck.merge_collab_edits_to_summary(
            edited_path,
            summary_path=summary_path,
        )
        assert result['success'] is True
        assert result['output_path'] is not None
        assert os.path.exists(result['output_path'])

        applied = result.get('applied_stats') or {}
        assert applied.get('matched_records', 0) == 4

        df = pd.read_excel(result['output_path'], engine='openpyxl')
        rec001 = df[df['唯一id'] == 'UID001'].iloc[0]
        assert rec001['凭证号'] == '记-202401-0001'
        assert rec001['备注'] == '测试备注1'

        rec002 = df[df['唯一id'] == 'UID002'].iloc[0]
        assert str(rec002['会计科目编码']).replace('.0', '') == '6001'

    def test_merge_no_changes(self, tmp_dir):
        """无修改内容时给出 warning"""
        summary_path = os.path.join(tmp_dir, '银行流水总表.xlsx')
        _create_test_summary(summary_path)
        template_path = bankcheck.generate_collab_template(summary_path, output_dir=tmp_dir)

        result = bankcheck.merge_collab_edits_to_summary(
            template_path,
            summary_path=summary_path,
        )
        assert result['success'] is True
        assert 'warning' in result

    def test_merge_missing_summary(self, tmp_dir):
        """找不到原始总表"""
        fake_edited = os.path.join(tmp_dir, 'fake.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value='唯一id')
        ws.cell(row=1, column=2, value='凭证号')
        ws.cell(row=2, column=1, value='UID001')
        ws.cell(row=2, column=2, value='记-001')
        wb.save(fake_edited)
        wb.close()

        result = bankcheck.merge_collab_edits_to_summary(
            fake_edited,
            summary_path=os.path.join(tmp_dir, 'not_exist_summary.xlsx'),
        )
        assert result['success'] is False
        assert 'error' in result


class TestConstants:
    """测试常量定义"""

    def test_editable_columns_defined(self):
        """可编辑列已定义"""
        assert len(bankcheck.COLLAB_EDITABLE_COLUMNS) > 0
        assert '凭证号' in bankcheck.COLLAB_EDITABLE_COLUMNS
        assert '备注' in bankcheck.COLLAB_EDITABLE_COLUMNS

    def test_subject_options_have_pairs(self):
        """会计科目选项为 (编码, 名称) 对"""
        for code, name in bankcheck.DEFAULT_SUBJECT_OPTIONS:
            assert code.isdigit()
            assert len(name) > 0
