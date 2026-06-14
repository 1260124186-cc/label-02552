import os
import sys
import shutil
import tempfile
from unittest.mock import patch, MagicMock

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

import bankcheck
from bankcheck import (
    BatchProcessingResult,
    FolderProcessingItem,
    ProcessingResult,
    format_batch_result_message,
    cli_ask_directories,
    run_batch_pipeline_flow,
)


@pytest.fixture(autouse=True)
def init_logging():
    bankcheck.setup_logging()


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='bankcheck_batch_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


def _create_bank_excel(path, bank_name='北京银行', account='01090312345678901'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '交易明细'
    ws['A1'] = '北京银行交易明细'
    ws['B2'] = account
    headers = ['序号', '交易日期', '币种', '支出金额', '收入金额', '余额',
               '对方户名', '对方账号', '对方行名', '凭证种类', '凭证号码',
               '摘要', '备注1', '备注2', '备注3', '交易流水号']
    for c, h in enumerate(headers, 1):
        ws.cell(row=3, column=c, value=h)
    rows = [
        [1, '2024-01-05', 'CNY', 50000, None, 1500000, '供应商A公司',
         '622001234', '工商银行', '转账', '001', '采购付款',
         None, None, None, 'BJ20240105001'],
        [2, '2024-01-10', 'CNY', None, 80000, 1580000, '客户B公司',
         '622005678', '建设银行', '转账', '002', '销售收款',
         None, None, None, 'BJ20240110002'],
    ]
    for i, row_data in enumerate(rows):
        for j, val in enumerate(row_data):
            ws.cell(row=4 + i, column=j + 1, value=val)
    wb.save(path)
    wb.close()


def _create_test_folder_tree(base_dir, subfolder_names):
    folders = []
    for name in subfolder_names:
        folder = os.path.join(base_dir, name)
        os.makedirs(folder, exist_ok=True)
        _create_bank_excel(os.path.join(folder, f'{name}.xlsx'))
        folders.append(folder)
    return folders


class TestFolderProcessingItem:
    def test_default_values(self):
        item = FolderProcessingItem(folder='/tmp/test')
        assert item.folder == '/tmp/test'
        assert item.result is None
        assert item.status == 'pending'
        assert item.error_message is None

    def test_with_result(self):
        result = ProcessingResult(all_rows=[{'a': 1}])
        item = FolderProcessingItem(folder='/tmp/test', result=result, status='success')
        assert item.status == 'success'
        assert len(item.result.all_rows) == 1


class TestBatchProcessingResult:
    def test_aggregate_empty(self):
        br = BatchProcessingResult()
        br.aggregate()
        assert br.total_folders == 0
        assert br.success_count == 0
        assert br.error_count == 0
        assert br.empty_count == 0

    def test_aggregate_success_items(self):
        r1 = ProcessingResult(all_rows=[{'a': 1}, {'b': 2}], new_record_count=2, processed_files=['f1'])
        r2 = ProcessingResult(all_rows=[{'c': 3}], new_record_count=1, processed_files=['f2'])
        br = BatchProcessingResult(items=[
            FolderProcessingItem(folder='/a', result=r1, status='success'),
            FolderProcessingItem(folder='/b', result=r2, status='success'),
        ])
        br.aggregate()
        assert br.total_folders == 2
        assert br.success_count == 2
        assert br.total_records == 3
        assert br.total_new_records == 3
        assert br.total_processed_files == 2

    def test_aggregate_mixed_statuses(self):
        r1 = ProcessingResult(all_rows=[], new_record_count=0, processed_files=[])
        r2 = ProcessingResult(all_rows=[{'a': 1}], new_record_count=1, processed_files=['f1'])
        br = BatchProcessingResult(items=[
            FolderProcessingItem(folder='/a', result=r1, status='empty'),
            FolderProcessingItem(folder='/b', result=r2, status='success'),
            FolderProcessingItem(folder='/c', result=None, status='error', error_message='fail'),
        ])
        br.aggregate()
        assert br.total_folders == 3
        assert br.success_count == 1
        assert br.empty_count == 1
        assert br.error_count == 1
        assert br.total_records == 1
        assert br.total_new_records == 1

    def test_aggregate_with_none_result(self):
        br = BatchProcessingResult(items=[
            FolderProcessingItem(folder='/a', result=None, status='error', error_message='boom'),
        ])
        br.aggregate()
        assert br.error_count == 1
        assert br.total_records == 0


class TestFormatBatchResultMessage:
    def test_empty_items(self):
        br = BatchProcessingResult()
        msg = format_batch_result_message(br)
        assert '未选择任何文件夹' in msg

    def test_success_message(self):
        r = ProcessingResult(all_rows=[{'a': 1}], new_record_count=1, processed_files=['f1'])
        br = BatchProcessingResult(items=[
            FolderProcessingItem(folder='/path/to/项目A', result=r, status='success'),
        ])
        br.aggregate()
        msg = format_batch_result_message(br)
        assert '批量处理完成' in msg
        assert '成功' in msg
        assert '项目A' in msg

    def test_dry_run_banner(self):
        r = ProcessingResult(all_rows=[{'a': 1}], new_record_count=1, processed_files=['f1'])
        br = BatchProcessingResult(
            dry_run=True,
            items=[FolderProcessingItem(folder='/p/a', result=r, status='success')],
        )
        br.aggregate()
        msg = format_batch_result_message(br)
        assert '试运行模式' in msg

    def test_error_item_message(self):
        br = BatchProcessingResult(items=[
            FolderProcessingItem(folder='/p/bad', status='error', error_message='读取失败'),
        ])
        br.aggregate()
        msg = format_batch_result_message(br)
        assert '失败' in msg
        assert '读取失败' in msg

    def test_incremental_mode_label(self):
        r = ProcessingResult(all_rows=[])
        br = BatchProcessingResult(
            incremental_mode=True,
            items=[FolderProcessingItem(folder='/p/x', result=r, status='empty')],
        )
        br.aggregate()
        msg = format_batch_result_message(br)
        assert '增量合并' in msg


class TestCliAskDirectories:
    def test_single_valid_path(self, tmp_dir):
        with patch('builtins.input', side_effect=[tmp_dir, '']):
            result = cli_ask_directories(use_history=False)
        assert result == [tmp_dir]

    def test_multiple_semicolon_paths(self, tmp_dir):
        dir_a = os.path.join(tmp_dir, 'a')
        dir_b = os.path.join(tmp_dir, 'b')
        os.makedirs(dir_a)
        os.makedirs(dir_b)
        with patch('builtins.input', side_effect=[f'{dir_a};{dir_b}', '']):
            result = cli_ask_directories(use_history=False)
        assert len(result) == 2

    def test_empty_input_returns_empty(self):
        with patch('builtins.input', return_value=''):
            result = cli_ask_directories(use_history=False)
        assert result == []

    def test_invalid_path_skipped(self, tmp_dir):
        with patch('builtins.input', side_effect=[f'{tmp_dir};/nonexistent/path', '']):
            result = cli_ask_directories(use_history=False)
        assert result == [tmp_dir]

    def test_continued_add(self, tmp_dir):
        dir_a = os.path.join(tmp_dir, 'a')
        dir_b = os.path.join(tmp_dir, 'b')
        os.makedirs(dir_a)
        os.makedirs(dir_b)
        with patch('builtins.input', side_effect=[dir_a, dir_b, '']):
            result = cli_ask_directories(use_history=False)
        assert len(result) == 2

    def test_quoted_path_stripped(self, tmp_dir):
        with patch('builtins.input', side_effect=[f'"{tmp_dir}"', '']):
            result = cli_ask_directories(use_history=False)
        assert result == [tmp_dir]


class TestRunBatchPipelineFlow:
    def test_empty_folders_returns_none(self, tmp_dir):
        with patch.object(bankcheck, 'ask_directories', return_value=[]):
            result = run_batch_pipeline_flow(tmp_dir)
        assert result is None

    def test_single_folder_success(self, tmp_dir):
        script_dir = tmp_dir
        folders = _create_test_folder_tree(tmp_dir, ['batch1'])
        lookup_path = os.path.join(script_dir, '主体查找表.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '主体名称'
        ws['B1'] = '银行账号'
        ws.cell(row=2, column=1, value='北京XX科技有限公司')
        ws.cell(row=2, column=2, value='01090312345678901')
        wb.save(lookup_path)
        wb.close()

        mock_audit = MagicMock()
        mock_audit.__enter__ = MagicMock(return_value=mock_audit)
        mock_audit.__exit__ = MagicMock(return_value=False)

        with patch.object(bankcheck, 'ask_incremental_mode', return_value=False), \
             patch.object(bankcheck, 'ask_dry_run_mode', return_value=True), \
             patch.object(bankcheck, 'ask_commit_changes', return_value=False), \
             patch.object(bankcheck, 'AuditLogger', return_value=mock_audit), \
             patch.object(bankcheck, 'show_info') as mock_info:
            result = run_batch_pipeline_flow(script_dir, folders=folders)
            mock_info.assert_called()

        assert isinstance(result, BatchProcessingResult)
        assert result.total_folders == 1
        assert result.success_count == 1

    def test_two_folders_sequential(self, tmp_dir):
        script_dir = tmp_dir
        folders = _create_test_folder_tree(tmp_dir, ['proj_a', 'proj_b'])
        lookup_path = os.path.join(script_dir, '主体查找表.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '主体名称'
        ws['B1'] = '银行账号'
        ws.cell(row=2, column=1, value='北京XX科技有限公司')
        ws.cell(row=2, column=2, value='01090312345678901')
        wb.save(lookup_path)
        wb.close()

        mock_audit = MagicMock()
        mock_audit.__enter__ = MagicMock(return_value=mock_audit)
        mock_audit.__exit__ = MagicMock(return_value=False)

        with patch.object(bankcheck, 'ask_incremental_mode', return_value=False), \
             patch.object(bankcheck, 'ask_dry_run_mode', return_value=True), \
             patch.object(bankcheck, 'ask_commit_changes', return_value=False), \
             patch.object(bankcheck, 'AuditLogger', return_value=mock_audit), \
             patch.object(bankcheck, 'show_info'):
            result = run_batch_pipeline_flow(script_dir, folders=folders)

        assert result.total_folders == 2
        assert result.success_count == 2

    def test_folder_with_error(self, tmp_dir):
        script_dir = tmp_dir
        good_folder = os.path.join(tmp_dir, 'good')
        os.makedirs(good_folder)
        _create_bank_excel(os.path.join(good_folder, 'good.xlsx'))

        bad_folder = os.path.join(tmp_dir, 'nonexistent_folder_xyz')
        folders = [good_folder, bad_folder]

        lookup_path = os.path.join(script_dir, '主体查找表.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '主体名称'
        ws['B1'] = '银行账号'
        ws.cell(row=2, column=1, value='北京XX科技有限公司')
        ws.cell(row=2, column=2, value='01090312345678901')
        wb.save(lookup_path)
        wb.close()

        mock_audit = MagicMock()
        mock_audit.__enter__ = MagicMock(return_value=mock_audit)
        mock_audit.__exit__ = MagicMock(return_value=False)

        with patch.object(bankcheck, 'ask_incremental_mode', return_value=False), \
             patch.object(bankcheck, 'ask_dry_run_mode', return_value=True), \
             patch.object(bankcheck, 'ask_commit_changes', return_value=False), \
             patch.object(bankcheck, 'AuditLogger', return_value=mock_audit), \
             patch.object(bankcheck, 'show_info'):
            result = run_batch_pipeline_flow(script_dir, folders=folders)

        assert result.total_folders == 2

    def test_dry_run_mode_flag(self, tmp_dir):
        script_dir = tmp_dir
        folders = _create_test_folder_tree(tmp_dir, ['dry1'])
        lookup_path = os.path.join(script_dir, '主体查找表.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '主体名称'
        ws['B1'] = '银行账号'
        ws.cell(row=2, column=1, value='北京XX科技有限公司')
        ws.cell(row=2, column=2, value='01090312345678901')
        wb.save(lookup_path)
        wb.close()

        mock_audit = MagicMock()
        mock_audit.__enter__ = MagicMock(return_value=mock_audit)
        mock_audit.__exit__ = MagicMock(return_value=False)

        with patch.object(bankcheck, 'ask_incremental_mode', return_value=True), \
             patch.object(bankcheck, 'ask_dry_run_mode', return_value=True), \
             patch.object(bankcheck, 'ask_commit_changes', return_value=False), \
             patch.object(bankcheck, 'AuditLogger', return_value=mock_audit), \
             patch.object(bankcheck, 'show_info'):
            result = run_batch_pipeline_flow(script_dir, folders=folders)

        assert result.dry_run is True
        assert result.incremental_mode is True

    def test_incremental_mode_flag(self, tmp_dir):
        script_dir = tmp_dir
        folders = _create_test_folder_tree(tmp_dir, ['inc1'])
        lookup_path = os.path.join(script_dir, '主体查找表.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '主体名称'
        ws['B1'] = '银行账号'
        ws.cell(row=2, column=1, value='北京XX科技有限公司')
        ws.cell(row=2, column=2, value='01090312345678901')
        wb.save(lookup_path)
        wb.close()

        mock_audit = MagicMock()
        mock_audit.__enter__ = MagicMock(return_value=mock_audit)
        mock_audit.__exit__ = MagicMock(return_value=False)

        with patch.object(bankcheck, 'ask_incremental_mode', return_value=False), \
             patch.object(bankcheck, 'ask_dry_run_mode', return_value=True), \
             patch.object(bankcheck, 'ask_commit_changes', return_value=False), \
             patch.object(bankcheck, 'AuditLogger', return_value=mock_audit), \
             patch.object(bankcheck, 'show_info'):
            result = run_batch_pipeline_flow(script_dir, folders=folders)

        assert result.incremental_mode is False


class TestBatchProcessingResultAggregate:
    def test_aggregate_duplicate_records(self):
        r = ProcessingResult(
            all_rows=[{'a': 1}],
            new_record_count=5,
            duplicate_record_count=3,
            processed_files=['f1'],
        )
        br = BatchProcessingResult(items=[
            FolderProcessingItem(folder='/x', result=r, status='success'),
        ])
        br.aggregate()
        assert br.total_duplicate_records == 3

    def test_aggregate_multiple_results(self):
        r1 = ProcessingResult(all_rows=[{'a': 1}], new_record_count=1, processed_files=['f1'])
        r2 = ProcessingResult(all_rows=[{'b': 2}, {'c': 3}], new_record_count=2, processed_files=['f2', 'f3'])
        br = BatchProcessingResult(items=[
            FolderProcessingItem(folder='/a', result=r1, status='success'),
            FolderProcessingItem(folder='/b', result=r2, status='success'),
        ])
        br.aggregate()
        assert br.total_records == 3
        assert br.total_new_records == 3
        assert br.total_processed_files == 3
