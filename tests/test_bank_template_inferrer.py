import os
import sys
import shutil
import tempfile

import openpyxl
import pytest

sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'backend'))

from bank_template_inferrer import (
    scan_workbook,
    confirm_and_save,
    _normalize_text,
    _looks_like_account,
    _looks_like_date,
    _looks_like_amount,
    _score_header_match,
    _is_header_row,
    HEADER_KEYWORDS,
)
from conftest import (
    _create_beijing_bank_excel,
    _create_east_asia_bank_excel,
    BANK_TEST_CONFIGS,
)


@pytest.fixture
def tmp_dir():
    d = tempfile.mkdtemp(prefix='bank_inferrer_test_')
    yield d
    shutil.rmtree(d, ignore_errors=True)


class TestNormalizeText:
    def test_none(self):
        assert _normalize_text(None) == ''

    def test_number(self):
        assert _normalize_text(123) == '123'

    def test_strip(self):
        assert _normalize_text('  hello  ') == 'hello'

    def test_fullwidth_digits(self):
        result = _normalize_text('１２３')
        assert '123' in result


class TestLooksLikeAccount:
    def test_valid_account(self):
        assert _looks_like_account('01090312345678901') is True

    def test_short_account(self):
        assert _looks_like_account('12345') is False

    def test_empty(self):
        assert _looks_like_account('') is False

    def test_account_with_spaces(self):
        assert _looks_like_account('6222 0000 1234 5678') is True

    def test_letters_mixed(self):
        assert _looks_like_account('abc123') is False


class TestLooksLikeDate:
    def test_iso_date(self):
        assert _looks_like_date('2024-01-05') is True

    def test_chinese_date(self):
        assert _looks_like_date('2024年01月05日') is True

    def test_slash_date(self):
        assert _looks_like_date('2024/01/05') is True

    def test_compact_date(self):
        assert _looks_like_date('20240105') is True

    def test_not_date(self):
        assert _looks_like_date('hello') is False

    def test_empty(self):
        assert _looks_like_date('') is False


class TestLooksLikeAmount:
    def test_integer(self):
        assert _looks_like_amount('50000') is True

    def test_float(self):
        assert _looks_like_amount('1234.56') is True

    def test_with_comma(self):
        assert _looks_like_amount('1,234,567.89') is True

    def test_not_amount(self):
        assert _looks_like_amount('abc') is False


class TestScoreHeaderMatch:
    def test_exact_match(self):
        score = _score_header_match('交易日期', HEADER_KEYWORDS['trade_date'])
        assert score == 1.0

    def test_partial_match(self):
        score = _score_header_match('交易日期时间', HEADER_KEYWORDS['trade_date'])
        assert score >= 0.7

    def test_no_match(self):
        score = _score_header_match('序号', HEADER_KEYWORDS['trade_date'])
        assert score < 0.5

    def test_empty(self):
        score = _score_header_match('', HEADER_KEYWORDS['trade_date'])
        assert score == 0.0


class TestIsHeaderRow:
    def test_valid_header_row(self):
        values = ['序号', '交易日期', '支出金额', '收入金额', '余额']
        assert _is_header_row(values, min_matches=2) is True

    def test_data_row(self):
        values = ['1', '2024-01-05', '50000', '', '1500000']
        assert _is_header_row(values, min_matches=2) is False

    def test_empty_row(self):
        values = ['', '', '', '', '']
        assert _is_header_row(values, min_matches=2) is False


class TestScanWorkbook:
    def test_beijing_bank(self, tmp_dir):
        filepath = os.path.join(tmp_dir, '北京银行_流水.xlsx')
        _create_beijing_bank_excel(filepath)
        result = scan_workbook(filepath)
        assert result['success'] is True
        assert result['header_row'] == 3
        assert result['start_row'] == 4
        assert result['account_cell'] == 'B2'
        assert 'trade_date' in result['column_map']
        assert result['column_map']['trade_date'] == 2
        assert 'payment' in result['column_map']
        assert 'receipt' in result['column_map']
        assert result['confidence'] > 0.5

    def test_east_asia_bank(self, tmp_dir):
        filepath = os.path.join(tmp_dir, '东亚银行_流水.xlsx')
        _create_east_asia_bank_excel(filepath)
        result = scan_workbook(filepath)
        assert result['success'] is True
        assert result['header_row'] == 4
        assert result['start_row'] == 5
        assert result['account_cell'] == 'B1'
        assert 'trade_date' in result['column_map']
        assert 'payment' in result['column_map']
        assert 'receipt' in result['column_map']

    def test_nonexistent_file(self):
        result = scan_workbook('/nonexistent/file.xlsx')
        assert result['success'] is False
        assert '不存在' in result['error']

    def test_empty_workbook(self, tmp_dir):
        filepath = os.path.join(tmp_dir, 'empty.xlsx')
        wb = openpyxl.Workbook()
        wb.save(filepath)
        wb.close()
        result = scan_workbook(filepath)
        assert result['success'] is False
        assert '表头行' in result['error']

    def test_custom_bank_format(self, tmp_dir):
        filepath = os.path.join(tmp_dir, '自定义银行_流水.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '交易明细'
        ws['A1'] = '某银行交易明细'
        ws['B2'] = '6225880123456789'
        ws['A3'] = '查询期间：2024-01-01 至 2024-12-31'
        headers = ['交易日期', '收入金额', '支出金额', '余额', '摘要', '对方户名']
        for c, h in enumerate(headers, 1):
            ws.cell(row=4, column=c, value=h)
        ws.cell(row=5, column=1, value='2024-03-15')
        ws.cell(row=5, column=2, value=100000)
        ws.cell(row=5, column=3, value=None)
        ws.cell(row=5, column=4, value=500000)
        ws.cell(row=5, column=5, value='转账收入')
        ws.cell(row=5, column=6, value='某公司')
        wb.save(filepath)
        wb.close()

        result = scan_workbook(filepath)
        assert result['success'] is True
        assert result['header_row'] == 4
        assert result['start_row'] == 5
        assert result['account_cell'] == 'B2'
        assert result['column_map']['trade_date'] == 1
        assert result['column_map']['receipt'] == 2
        assert result['column_map']['payment'] == 3
        assert result['column_map']['balance'] == 4
        assert result['column_map']['summary'] == 5
        assert result['column_map']['counterpart'] == 6

    def test_unmatched_headers_reported(self, tmp_dir):
        filepath = os.path.join(tmp_dir, 'extra_cols.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '账号'
        ws['B1'] = '6222000000000001'
        headers = ['交易日期', '收入金额', '支出金额', '余额', '自定义列A', '自定义列B']
        for c, h in enumerate(headers, 1):
            ws.cell(row=2, column=c, value=h)
        wb.save(filepath)
        wb.close()

        result = scan_workbook(filepath)
        assert result['success'] is True
        assert '自定义列A' in result['unmatched_headers']
        assert '自定义列B' in result['unmatched_headers']

    def test_confidence_high_when_all_matched(self, tmp_dir):
        filepath = os.path.join(tmp_dir, '北京银行_流水.xlsx')
        _create_beijing_bank_excel(filepath)
        result = scan_workbook(filepath)
        assert result['confidence'] >= 0.7

    def test_confidence_low_when_few_matched(self, tmp_dir):
        filepath = os.path.join(tmp_dir, 'minimal.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '交易日期'
        ws['B1'] = '备注'
        ws['A2'] = '2024-01-01'
        ws['B2'] = '测试'
        wb.save(filepath)
        wb.close()

        result = scan_workbook(filepath)
        assert result['success'] is True
        assert result['confidence'] < 0.7

    def test_no_account_warning(self, tmp_dir):
        filepath = os.path.join(tmp_dir, 'no_account.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ['交易日期', '收入金额', '支出金额', '余额']
        for c, h in enumerate(headers, 1):
            ws.cell(row=1, column=c, value=h)
        ws.cell(row=2, column=1, value='2024-01-01')
        wb.save(filepath)
        wb.close()

        result = scan_workbook(filepath)
        assert result['success'] is True
        assert any('账号' in w for w in result['warnings'])

    def test_account_label_detection(self, tmp_dir):
        filepath = os.path.join(tmp_dir, 'labeled_account.xlsx')
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = '账号'
        ws['B1'] = '6225880123456789'
        headers = ['交易日期', '收入金额', '支出金额', '余额']
        for c, h in enumerate(headers, 1):
            ws.cell(row=2, column=c, value=h)
        wb.save(filepath)
        wb.close()

        result = scan_workbook(filepath)
        assert result['success'] is True
        assert result['account_cell'] == 'B1'


class TestConfirmAndSave:
    def test_missing_bank_name(self):
        inferred = {
            'success': True,
            'column_map': {'trade_date': 1},
            'expected_headers': {'trade_date': '交易日期'},
            'account_cell': 'A1',
            'start_row': 2,
            'header_row': 1,
        }
        ok, msg = confirm_and_save(inferred, '')
        assert ok is False
        assert '银行名称' in msg

    def test_missing_trade_date(self):
        inferred = {
            'success': True,
            'column_map': {},
            'expected_headers': {},
            'account_cell': 'A1',
            'start_row': 2,
            'header_row': 1,
        }
        ok, msg = confirm_and_save(inferred, '测试银行')
        assert ok is False
        assert '交易日期' in msg

    def test_overrides_applied(self, tmp_dir):
        inferred = {
            'success': True,
            'column_map': {'trade_date': 1, 'payment': 2},
            'expected_headers': {'trade_date': '日期', 'payment': '支出'},
            'account_cell': 'A1',
            'start_row': 2,
            'header_row': 1,
        }
        overrides = {
            'column_map': {'trade_date': 2, 'receipt': 3},
            'expected_headers': {'trade_date': '交易日期', 'receipt': '收入'},
            'account_cell': 'B2',
            'start_row': 3,
            'payment_sign': 'positive',
            'enabled': True,
            'header_validation': 'warn',
            'multi_account': False,
            'skip_sheets': [],
        }
        config_path = os.path.join(tmp_dir, 'test_bank_rules.yaml')
        import yaml
        with open(config_path, 'w', encoding='utf-8') as f:
            yaml.safe_dump({'banks': []}, f, allow_unicode=True, default_flow_style=False)

        from bankcheck import BankRuleConfig
        original_instance = BankRuleConfig._instance
        BankRuleConfig._instance = None
        try:
            config = BankRuleConfig(config_path)
            from bankcheck import get_bank_config
            import bankcheck
            original_get = bankcheck.get_bank_config
            bankcheck.get_bank_config = lambda: config

            ok, msg = confirm_and_save(inferred, '测试银行', overrides=overrides, config_path=config_path)
            assert ok is True

            with open(config_path, 'r', encoding='utf-8') as f:
                saved = yaml.safe_load(f)
            banks = saved.get('banks', [])
            found = [b for b in banks if b.get('bank_name') == '测试银行']
            assert len(found) == 1
            assert found[0]['columns']['trade_date'] == 2
            assert found[0]['columns']['receipt'] == 3
            assert found[0]['account_cell'] == 'B2'
            assert found[0]['start_row'] == 3
            assert found[0]['payment_sign'] == 'positive'

            bankcheck.get_bank_config = original_get
        finally:
            BankRuleConfig._instance = original_instance
