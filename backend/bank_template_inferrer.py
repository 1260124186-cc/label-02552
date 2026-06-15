# -*- coding: utf-8 -*-
"""
银行模板自动推断模块
====================

用户上传一份未知格式银行流水样本后，程序通过表头关键词与列位置启发式
推荐列映射草案，经确认后生成可复用的银行配置，降低新银行接入成本。

流程：
  1. 读取 Excel 文件，扫描每一行寻找表头行
  2. 通过预定义的表头关键词字典匹配列到标准字段
  3. 检测银行账号所在单元格
  4. 推断数据起始行
  5. 输出列映射草案，前端可展示并允许用户修改
  6. 用户确认后调用 BankRuleConfig.save_rule 写入 YAML
"""

import os
import re
import logging
from typing import Dict, List, Optional, Tuple, Any

import openpyxl

try:
    from bankcheck import (
        open_workbook_compat,
        cleanup_temp_file,
        BankRuleConfig,
        get_bank_config,
        get_cell_ref,
        col_letter_to_index,
        _normalize_width,
    )
except ImportError:
    from bankcheck import (
        open_workbook_compat,
        cleanup_temp_file,
    )

    try:
        from bankcheck import BankRuleConfig, get_bank_config
    except ImportError:
        BankRuleConfig = None
        get_bank_config = None

    try:
        from bankcheck import get_cell_ref, col_letter_to_index, _normalize_width
    except ImportError:
        get_cell_ref = None
        col_letter_to_index = None
        _normalize_width = None

logger = logging.getLogger('bankcheck.inferrer')

HEADER_KEYWORDS: Dict[str, List[str]] = {
    'trade_date': [
        '交易日期', '交易时间日期', '记账日期', '日期', '交易日期时间',
        '发生日期', '业务日期', '入账日期', '账务日期', '交易发生日期',
        'transdate', 'transactiondate', 'date', '交易日期/时间',
    ],
    'payment': [
        '支出金额', '支出', '借方金额', '付款金额', '付款', '借方发生额',
        '支出(元)', '借方', '付款(元)', '转出金额', '转出',
        'debit', 'payment', 'withdrawal', '支出金额(元)',
    ],
    'receipt': [
        '收入金额', '收入', '贷方金额', '收款金额', '收款', '贷方发生额',
        '收入(元)', '贷方', '收款(元)', '转入金额', '转入',
        'credit', 'receipt', 'deposit', '收入金额(元)',
    ],
    'balance': [
        '余额', '账户余额', '当前余额', '可用余额', '实时余额',
        'balance', '账面余额', '结余',
    ],
    'counterpart': [
        '对方户名', '对方名称', '交易对方', '对方账户名', '对方',
        '对方姓名', '交易对方名称', '收付款人', '付款人', '收款人',
        'counterpart', '对方信息', '交易描述/对方', '对方账户名称',
    ],
    'summary': [
        '摘要', '交易摘要', '用途', '备注', '交易备注', '交易用途',
        '摘要/备注', '说明', '交易说明', '摘要说明',
        'summary', 'description', 'remark',
    ],
    'transaction_id': [
        '交易流水号', '流水号', '交易号', '凭证号码', '交易编号',
        '交易序号', '业务流水号', '交易参考号', '参考号',
        'transactionid', 'reference', 'refno',
    ],
}

_REQUIRED_FIELDS = ['trade_date']

_ACCOUNT_PATTERNS = [
    re.compile(r'^\d{6,30}$'),
    re.compile(r'^\d{4}\s?\d{4}\s?\d{4}\s?\d{4}'),
]

_ACCOUNT_LABEL_PATTERNS = [
    re.compile(r'账号|账户|卡号|account', re.IGNORECASE),
]


def _normalize_text(val: Any) -> str:
    if val is None:
        return ''
    s = str(val).strip()
    try:
        s = _normalize_width(s)
    except (TypeError, AttributeError):
        pass
    return s


def _looks_like_account(text: str) -> bool:
    text = text.replace(' ', '').replace('\u3000', '')
    if not text:
        return False
    for pat in _ACCOUNT_PATTERNS:
        if pat.match(text):
            return True
    return False


def _looks_like_date(text: str) -> bool:
    if not text:
        return False
    date_patterns = [
        r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}',
        r'^\d{4}年\d{1,2}月\d{1,2}日',
        r'^\d{8}$',
    ]
    for pat in date_patterns:
        if re.match(pat, text):
            return True
    return False


def _looks_like_amount(text: str) -> bool:
    if not text:
        return False
    cleaned = text.replace(',', '').replace('，', '').replace(' ', '')
    try:
        float(cleaned)
        return True
    except (ValueError, TypeError):
        return False


def _score_header_match(header_text: str, keywords: List[str]) -> float:
    if not header_text:
        return 0.0
    normalized = _normalize_text(header_text).lower()
    best = 0.0
    for kw in keywords:
        kw_lower = kw.lower()
        if normalized == kw_lower:
            best = max(best, 1.0)
        elif normalized.startswith(kw_lower) or normalized.endswith(kw_lower):
            best = max(best, 0.9)
        elif kw_lower in normalized:
            best = max(best, 0.7)
        elif normalized in kw_lower:
            best = max(best, 0.5)
    return best


def _is_header_row(values: List[str], min_matches: int = 2) -> bool:
    matches = 0
    for val in values:
        txt = _normalize_text(val)
        if not txt:
            continue
        for field_key, keywords in HEADER_KEYWORDS.items():
            if _score_header_match(txt, keywords) >= 0.7:
                matches += 1
                break
    return matches >= min_matches


def scan_workbook(filepath: str, sheet_name: Optional[str] = None,
                  max_scan_rows: int = 30, max_scan_cols: int = 30) -> Dict[str, Any]:
    """
    扫描 Excel 文件，自动推断银行流水格式。

    Returns:
        包含推断结果的字典：
        - success: 是否成功
        - error: 错误信息
        - sheet_names: 工作表名列表
        - used_sheet: 使用的工作表名
        - header_row: 推断的表头行号（1-based）
        - start_row: 推断的数据起始行号（1-based）
        - account_cell: 推断的账号单元格引用（如 B2）
        - account_value: 账号值
        - column_map: 列映射 {field_key: col_index(1-based)}
        - expected_headers: 列映射对应的表头文本 {field_key: header_text}
        - header_values: 表头行全部值
        - confidence: 整体置信度 0-1
        - unmatched_headers: 未匹配的表头文本列表
        - warnings: 警告列表
    """
    result: Dict[str, Any] = {
        'success': False,
        'error': '',
        'sheet_names': [],
        'used_sheet': '',
        'header_row': None,
        'start_row': None,
        'account_cell': None,
        'account_value': '',
        'column_map': {},
        'expected_headers': {},
        'header_values': [],
        'confidence': 0.0,
        'unmatched_headers': [],
        'warnings': [],
    }

    if not os.path.isfile(filepath):
        result['error'] = f'文件不存在: {filepath}'
        return result

    wb = None
    tmp_path = None
    try:
        wb, tmp_path = open_workbook_compat(filepath)

        sheet_names = [ws.title for ws in wb.worksheets]
        result['sheet_names'] = sheet_names

        if sheet_name and sheet_name in sheet_names:
            ws = wb[sheet_name]
        else:
            ws = wb.active
        result['used_sheet'] = ws.title

        scan_rows = min(ws.max_row, max_scan_rows)
        scan_cols = min(ws.max_column, max_scan_cols)

        header_row_idx = None
        for row_idx in range(1, scan_rows + 1):
            row_values = []
            for col_idx in range(1, scan_cols + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                row_values.append(_normalize_text(val))
            if _is_header_row(row_values, min_matches=2):
                header_row_idx = row_idx
                break

        if header_row_idx is None:
            result['error'] = '未检测到有效的表头行，请确认文件格式'
            result['warnings'].append('无法自动检测表头行，请手动指定')
            return result

        result['header_row'] = header_row_idx
        result['start_row'] = header_row_idx + 1

        header_values = []
        for col_idx in range(1, scan_cols + 1):
            val = ws.cell(row=header_row_idx, column=col_idx).value
            header_values.append(_normalize_text(val))
        result['header_values'] = header_values

        column_map: Dict[str, int] = {}
        expected_headers: Dict[str, str] = {}
        field_scores: Dict[str, List[Tuple[int, float, str]]] = {}

        for col_idx, header_text in enumerate(header_values, start=1):
            if not header_text:
                continue
            for field_key, keywords in HEADER_KEYWORDS.items():
                score = _score_header_match(header_text, keywords)
                if score >= 0.5:
                    if field_key not in field_scores:
                        field_scores[field_key] = []
                    field_scores[field_key].append((col_idx, score, header_text))

        for field_key, candidates in field_scores.items():
            candidates.sort(key=lambda x: x[1], reverse=True)
            best_col, best_score, best_header = candidates[0]
            if best_score >= 0.7:
                column_map[field_key] = best_col
                expected_headers[field_key] = best_header

        for field_key in _REQUIRED_FIELDS:
            if field_key not in column_map:
                for col_idx, header_text in enumerate(header_values, start=1):
                    if not header_text:
                        continue
                    txt = _normalize_text(header_text)
                    if field_key == 'trade_date' and _looks_like_date(txt):
                        result['warnings'].append(
                            f'必填字段「交易日期」未通过关键词匹配，尝试按日期格式辅助识别到第{col_idx}列'
                        )
                        column_map[field_key] = col_idx
                        expected_headers[field_key] = header_text
                        break

        result['column_map'] = column_map
        result['expected_headers'] = expected_headers

        matched_cols = set(column_map.values())
        result['unmatched_headers'] = [
            hv for i, hv in enumerate(header_values, start=1)
            if hv and i not in matched_cols
        ]

        account_cell = None
        account_value = ''
        for row_idx in range(1, min(header_row_idx, scan_rows + 1)):
            for col_idx in range(1, scan_cols + 1):
                val = ws.cell(row=row_idx, column=col_idx).value
                txt = _normalize_text(val)
                if _looks_like_account(txt):
                    account_cell = get_cell_ref(col_idx, row_idx)
                    account_value = txt
                    break
            if account_cell:
                break

        if not account_cell:
            for row_idx in range(1, min(header_row_idx, scan_rows + 1)):
                for col_idx in range(1, scan_cols + 1):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    txt = _normalize_text(val)
                    if not txt:
                        continue
                    label_on_left = False
                    if col_idx > 1:
                        left_val = ws.cell(row=row_idx, column=col_idx - 1).value
                        left_txt = _normalize_text(left_val)
                        for pat in _ACCOUNT_LABEL_PATTERNS:
                            if pat.search(left_txt):
                                label_on_left = True
                                break
                    if label_on_left and _looks_like_account(txt):
                        account_cell = get_cell_ref(col_idx, row_idx)
                        account_value = txt
                        break
                if account_cell:
                    break

        result['account_cell'] = account_cell
        result['account_value'] = account_value

        if not account_cell:
            result['warnings'].append('未自动检测到银行账号单元格，请手动指定')

        total_fields = len(HEADER_KEYWORDS)
        matched_fields = len(column_map)
        field_ratio = matched_fields / total_fields if total_fields else 0

        has_account = 1.0 if account_cell else 0.0
        has_required = 1.0 if column_map.get('trade_date') else 0.0

        confidence = (field_ratio * 0.5 + has_account * 0.25 + has_required * 0.25)
        result['confidence'] = round(min(confidence, 1.0), 2)

        if matched_fields < 3:
            result['warnings'].append(f'仅匹配到 {matched_fields} 个字段，置信度较低，请仔细检查映射')
        if 'payment' not in column_map and 'receipt' not in column_map:
            result['warnings'].append('未匹配到收入/支出金额列，请手动映射')

        result['success'] = True
        return result

    except Exception as e:
        logger.error('扫描文件失败: %s', e, exc_info=True)
        result['error'] = str(e)
        return result
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass
        if tmp_path:
            cleanup_temp_file(tmp_path)


def confirm_and_save(inferred: Dict[str, Any], bank_name: str,
                     overrides: Optional[Dict[str, Any]] = None,
                     config_path: Optional[str] = None) -> Tuple[bool, str]:
    """
    用户确认推断草案后，生成可复用的银行配置并写入 YAML。

    Args:
        inferred: scan_workbook 返回的推断结果
        bank_name: 银行名称
        overrides: 用户修改的覆盖值（如修改列映射、账号单元格等）
        config_path: 配置文件路径（默认使用 BankRuleConfig 默认路径）

    Returns:
        (success, message)
    """
    if not bank_name or not bank_name.strip():
        return False, '银行名称不能为空'

    column_map = dict(inferred.get('column_map', {}))
    expected_headers = dict(inferred.get('expected_headers', {}))
    account_cell = inferred.get('account_cell', 'A1') or 'A1'
    start_row = inferred.get('start_row', 2)
    header_row = inferred.get('header_row', 1)

    if overrides:
        if 'column_map' in overrides:
            column_map = {k: int(v) for k, v in overrides['column_map'].items() if v}
        if 'expected_headers' in overrides:
            expected_headers = overrides['expected_headers']
        if 'account_cell' in overrides:
            account_cell = overrides['account_cell']
        if 'start_row' in overrides:
            start_row = int(overrides['start_row'])
        if 'header_row' in overrides:
            header_row = int(overrides['header_row'])

    if 'trade_date' not in column_map:
        return False, '缺少必填字段「交易日期」的列映射'

    try:
        col_letter_to_index(account_cell[:1] if account_cell else 'A')
    except Exception:
        pass

    rule_data = {
        'bank_name': bank_name.strip(),
        'account_cell': account_cell,
        'start_row': int(start_row),
        'columns': column_map,
        'payment_sign': (overrides or {}).get('payment_sign', 'negative'),
        'enabled': bool((overrides or {}).get('enabled', True)),
        'expected_headers': expected_headers,
        'header_validation': (overrides or {}).get('header_validation', 'warn'),
        'multi_account': bool((overrides or {}).get('multi_account', False)),
        'skip_sheets': (overrides or {}).get('skip_sheets', []),
    }

    try:
        config = get_bank_config()
        if config_path:
            config._config_path = config_path
        ok, _ = config.save_rule(rule_data)
        if ok:
            logger.info('银行模板推断结果已保存: %s', bank_name)
            return True, f'银行规则「{bank_name}」保存成功'
        return False, '保存失败，请检查日志'
    except Exception as e:
        logger.error('保存银行规则失败: %s', e, exc_info=True)
        return False, f'保存失败: {e}'
