# -*- coding: utf-8 -*-
"""
银行流水检验工具 - bankcheck.py
功能：
  1. 用户选择一个文件夹
  2. 在同路径下复制一份，命名为"原名＋检验版"
  3. 对复制后的文件夹递归扫描所有 Excel 文件（.xlsx 和 .xls）
  4. 根据文件名前缀判断银行类型，提取指定列
  5. 生成唯一ID，合并输出总表
  6. 删除已处理的 Excel 文件，保留无法识别银行的文件

银行解析规则通过 bank_rules.yaml 配置文件管理，支持热更新。
业务人员可通过修改配置文件增删银行规则，无需修改 Python 代码。
"""

import os
import re
import sys
import shutil
import uuid
import logging
import tempfile
import sqlite3
import json
import getpass
import hashlib
import unicodedata
from collections import defaultdict
from dataclasses import dataclass, field, asdict
from datetime import datetime
from typing import List, Optional, Tuple, Dict, Any

import openpyxl
import pandas as pd
import yaml

try:
    from i18n import t, set_language, get_language, get_available_languages, init_i18n
    HAS_I18N = True
except ImportError:
    HAS_I18N = False
    def t(key, **kwargs):
        return key
    def set_language(lang):
        return False
    def get_language():
        return 'zh_CN'
    def get_available_languages():
        return {'zh_CN': '简体中文'}
    def init_i18n(lang=None):
        return None

try:
    from build_info import (
        get_version, get_build_time, get_build_info, get_build_platform, format_version_banner,
    )
    HAS_BUILD_INFO = True
except ImportError:
    HAS_BUILD_INFO = False
    def get_version():
        return "1.0.0"
    def get_build_time():
        return "unknown"
    def get_build_platform():
        import platform as _platform
        return _platform.system()
    def get_build_info():
        return {'version': get_version(), 'build_time': get_build_time(), 'platform': get_build_platform()}
    def format_version_banner(app_name="银行流水检验工具"):
        v = get_version()
        bt = get_build_time()
        bp = get_build_platform()
        pd = {"Windows": "Windows", "Darwin": "macOS", "Linux": "Linux"}.get(bp, bp)
        line = "=" * 48
        return f"{line}\n  {app_name}\n  版本: v{v}\n  构建时间: {bt}\n  构建平台: {pd}\n{line}"

try:
    from pii_classifier import (
        setup_pii_aware_logging, PIILogFilter, build_safe_log_context,
        mask_value, _mask_bank_account, _mask_subject_name,
    )
    HAS_PII_CLASSIFIER = True
except ImportError:
    HAS_PII_CLASSIFIER = False
    mask_value = None
    _mask_bank_account = None
    _mask_subject_name = None

try:
    from file_encryption import (
        encrypt_output_files as _encrypt_output_files,
        validate_password_strength as _validate_password_strength,
        BatchEncryptionResult as _BatchEncryptionResult,
        EncryptionResult as _EncryptionResult,
        save_encryption_record as _save_encryption_record,
        is_encrypted_file as _is_encrypted_file,
        get_encryption_info as _get_encryption_info,
        HAS_MSOFFCRYPTO as _HAS_MSOFFCRYPTO,
    )
    HAS_FILE_ENCRYPTION = True
except ImportError:
    HAS_FILE_ENCRYPTION = False
    _encrypt_output_files = None
    _validate_password_strength = None
    _BatchEncryptionResult = None
    _EncryptionResult = None
    _save_encryption_record = None
    _is_encrypted_file = None
    _get_encryption_info = None
    _HAS_MSOFFCRYPTO = False

try:
    import database as db_module
    HAS_DATABASE = True
except ImportError:
    HAS_DATABASE = False
    db_module = None

try:
    import perf_profiler
    HAS_PERF_PROFILER = True
except ImportError:
    HAS_PERF_PROFILER = False

try:
    import batch_manager as batch_module
    HAS_BATCH_MANAGER = True
except ImportError:
    HAS_BATCH_MANAGER = False
    batch_module = None

try:
    import onboarding
    HAS_ONBOARDING = True
except ImportError:
    HAS_ONBOARDING = False
    onboarding = None

try:
    from cryptography.hazmat.primitives.asymmetric import rsa, padding
    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.backends import default_backend
    from cryptography.exceptions import InvalidSignature
    HAS_CRYPTOGRAPHY = True
except ImportError:
    HAS_CRYPTOGRAPHY = False
    rsa = None
    padding = None
    hashes = None
    serialization = None
    default_backend = None
    InvalidSignature = Exception

# ──────────────────────────────────────────────
# tkinter 兼容：尝试导入并安全测试，失败则回退命令行模式
# ──────────────────────────────────────────────

def _verify_tkinter():
    import subprocess
    import sys
    try:
        # 在子进程中测试 tk.Tk()。这样如果底层动态库版本不匹配发生 Abort trap: 6 崩溃，
        # 只会使子进程报错退出（返回非 0 状态码），不会导致主程序崩溃。
        code = "import tkinter; tkinter.Tk().destroy()"
        result = subprocess.run([sys.executable, '-c', code], capture_output=True, text=True, timeout=5)
        return result.returncode == 0
    except Exception:
        return False

HAS_TKINTER = False
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    if _verify_tkinter():
        HAS_TKINTER = True
    else:
        tk = None
except (ImportError, ModuleNotFoundError):
    # 在无桌面或缺少 tk 的环境中，回退到命令行模式
    tk = None


# ──────────────────────────────────────────────
# 最近使用文件夹历史记录
# ──────────────────────────────────────────────

RECENT_FOLDERS_FILENAME = 'recent_folders.json'
MAX_RECENT_FOLDERS = 10


def get_recent_folders_path(script_dir=None):
    """获取最近文件夹历史记录文件路径"""
    if script_dir is None:
        script_dir = get_script_dir()
    return os.path.join(script_dir, RECENT_FOLDERS_FILENAME)


def load_recent_folders(script_dir=None):
    """加载最近使用的文件夹列表"""
    path = get_recent_folders_path(script_dir)
    if not os.path.exists(path):
        return []
    try:
        with open(path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            folders = data.get('folders', [])
            return [f for f in folders if os.path.isdir(f)]
    except Exception:
        return []


def save_recent_folders(folders, script_dir=None):
    """保存最近使用的文件夹列表"""
    path = get_recent_folders_path(script_dir)
    try:
        with open(path, 'w', encoding='utf-8') as f:
            json.dump({'folders': folders[:MAX_RECENT_FOLDERS]}, f, ensure_ascii=False, indent=2)
    except Exception as e:
        logger = get_logger()
        logger.warning('保存最近文件夹历史失败: %s', e)


def add_to_recent_folders(folder, script_dir=None):
    """将文件夹添加到最近使用列表（去重，最新的在前）"""
    if not folder or not os.path.isdir(folder):
        return
    folders = load_recent_folders(script_dir)
    if folder in folders:
        folders.remove(folder)
    folders.insert(0, folder)
    save_recent_folders(folders, script_dir)


def get_last_used_folder(script_dir=None):
    """获取最近使用的文件夹（第一个）"""
    folders = load_recent_folders(script_dir)
    return folders[0] if folders else None


def cli_askdirectory(title=None, use_history=True):
    """命令行模式下让用户输入文件夹路径，支持最近使用历史"""
    if title is None:
        title = t('gui.select_folder')
    print(f'\n{title}')

    if use_history:
        recent = load_recent_folders()
        if recent:
            print(t('cli.recent_folders'))
            for i, folder in enumerate(recent[:MAX_RECENT_FOLDERS], 1):
                print(f'  {i}) {folder}')
            print(t('cli.select_recent_hint'))

    path = input(t('cli.enter_folder_path')).strip().strip('"').strip("'")

    if use_history and path and path.isdigit():
        idx = int(path) - 1
        recent = load_recent_folders()
        if 0 <= idx < len(recent):
            path = recent[idx]

    if path and os.path.isdir(path):
        add_to_recent_folders(path)
        return path
    return ''


def cli_showinfo(title, message):
    """命令行模式下打印信息"""
    print(f'\n[{title}] {message}')


def cli_showwarning(title, message):
    """命令行模式下打印警告"""
    print(f'\n[{t("gui.warning")} - {title}] {message}')


def cli_askfile(title=None):
    """命令行模式下让用户输入文件路径"""
    if title is None:
        title = t('gui.select_file')
    print(f'\n{title}')
    path = input(t('cli.enter_file_path')).strip().strip('"').strip("'")
    if path and os.path.isfile(path):
        return path
    return ''


def gui_askdirectory(title=None, initialdir=None, use_history=True, show_recent_dialog=True):
    """GUI 模式选择文件夹，支持最近使用历史

    Args:
        title: 对话框标题
        initialdir: 初始打开目录，默认使用最近使用的文件夹
        use_history: 是否使用历史记录
        show_recent_dialog: 是否显示最近文件夹快速选择对话框
    """
    if title is None:
        title = t('gui.select_folder')

    if use_history and initialdir is None:
        initialdir = get_last_used_folder()

    if show_recent_dialog and use_history:
        recent = load_recent_folders()
        if recent:
            folder = _gui_show_recent_folders_dialog(title, recent, initialdir)
            if folder is not None:
                if folder:
                    add_to_recent_folders(folder)
                return folder

    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title=title, initialdir=initialdir or '')
    if not folder:
        messagebox.showinfo(t('gui.info'), t('gui.no_folder_selected'))
    else:
        if use_history:
            add_to_recent_folders(folder)
    root.destroy()
    return folder


def _gui_show_recent_folders_dialog(title, recent_folders, initialdir=None):
    """显示最近使用文件夹快速选择对话框

    Returns:
        str: 选择的文件夹路径，空字符串表示取消，None 表示用户选择"浏览其他"
    """
    result = {'folder': None}

    try:
        root = tk.Tk()
        root.title(title)
        root.geometry('560x480')
        root.resizable(False, False)

        tk.Label(root, text=t('gui.recent_folders_title'),
                 font=('Arial', 14, 'bold')).pack(pady=(15, 5))
        tk.Label(root, text=t('gui.recent_folders_subtitle'),
                 font=('Arial', 10), fg='#666').pack(pady=(0, 10))

        list_frame = tk.Frame(root)
        list_frame.pack(fill=tk.BOTH, expand=True, padx=20, pady=5)

        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        listbox = tk.Listbox(list_frame, font=('Arial', 11),
                            yscrollcommand=scrollbar.set, activestyle='dotbox')
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=listbox.yview)

        for folder in recent_folders:
            display_name = os.path.basename(folder)
            listbox.insert(tk.END, f'  {display_name}')
            listbox.insert(tk.END, f'      {folder}')
            listbox.insert(tk.END, '')

        def on_select(event):
            idx = listbox.curselection()
            if idx:
                folder_idx = idx[0] // 3
                if folder_idx < len(recent_folders):
                    result['folder'] = recent_folders[folder_idx]
                    root.destroy()

        listbox.bind('<Double-1>', on_select)

        btn_frame = tk.Frame(root)
        btn_frame.pack(fill=tk.X, padx=20, pady=15)

        def select_from_list():
            idx = listbox.curselection()
            if not idx:
                messagebox.showinfo(t('gui.info'), t('gui.please_select_recent'))
                return
            folder_idx = idx[0] // 3
            if folder_idx < len(recent_folders):
                result['folder'] = recent_folders[folder_idx]
                root.destroy()

        def browse_other():
            root.destroy()
            result['folder'] = None

        def cancel():
            result['folder'] = ''
            root.destroy()

        tk.Button(btn_frame, text=t('gui.use_selected'),
                  width=12, command=select_from_list,
                  bg='#4CAF50', fg='white', font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text=t('gui.browse_other'),
                  width=12, command=browse_other,
                  bg='#2196F3', fg='white', font=('Arial', 10, 'bold')).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text=t('gui.cancel'),
                  width=10, command=cancel,
                  bg='#f44336', fg='white', font=('Arial', 10, 'bold')).pack(side=tk.RIGHT, padx=5)

        root.protocol('WM_DELETE_WINDOW', cancel)
        root.mainloop()

        return result['folder']

    except Exception as e:
        logger = get_logger()
        logger.warning('显示最近文件夹对话框失败: %s', e)
        return None


def gui_showinfo(title, message):
    """GUI 模式弹出信息"""
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(title, message)
    root.destroy()


def gui_showwarning(title, message):
    """GUI 模式弹出警告"""
    root = tk.Tk()
    root.withdraw()
    messagebox.showwarning(title, message)
    root.destroy()


def gui_askfile(title=None):
    """GUI 模式选择文件"""
    if title is None:
        title = t('gui.select_summary_file')
    root = tk.Tk()
    root.withdraw()
    filepath = filedialog.askopenfilename(
        title=title,
        filetypes=[(t('gui.excel_files'), '*.xlsx *.xls'), (t('gui.all_files'), '*.*')],
    )
    if not filepath:
        messagebox.showinfo(t('gui.info'), t('gui.no_file_selected'))
    root.destroy()
    return filepath


def cli_askmode():
    """命令行模式下让用户选择运行模式"""
    print('\n' + t('cli.select_mode'))
    print(t('cli.option_pipeline'))
    print(t('cli.option_diff'))
    print(t('cli.option_monitor'))
    print(t('cli.option_scheduler'))
    print(t('cli.option_export'))
    print(t('cli.option_db_query'))
    print(t('cli.option_db_stats'))
    print(t('cli.option_batch_history'))
    print(t('cli.option_preset'))
    print(t('cli.option_subject_summary'))
    print(t('cli.option_balance_check'))
    print(t('cli.option_duplicate_check'))
    print(t('cli.option_interest_fee_check'))
    print(t('cli.option_balance_reconciliation'))
    print('  15) 非工作日交易标记')
    choice = input(t('cli.enter_choice')).strip()
    if choice == '2':
        return 'diff'
    elif choice == '3':
        return 'monitor'
    elif choice == '4':
        return 'scheduler'
    elif choice == '5':
        return 'export'
    elif choice == '6':
        return 'db_query'
    elif choice == '7':
        return 'db_stats'
    elif choice == '8':
        return 'batch_history'
    elif choice == '9':
        return 'preset'
    elif choice == '10':
        return 'subject_summary'
    elif choice == '11':
        return 'balance_check'
    elif choice == '12':
        return 'duplicate_check'
    elif choice == '13':
        return 'interest_fee_check'
    elif choice == '14':
        return 'balance_reconciliation'
    elif choice == '15':
        return 'holiday_check'
    return 'pipeline'


def gui_askmode():
    """GUI 模式下让用户选择运行模式"""
    if tk is None:
        return cli_askmode()

    try:
        result = _gui_askmode_full()
        if result is not None:
            return result
    except Exception:
        pass

    try:
        root = tk.Tk()
        root.withdraw()
        choice = messagebox.askyesnocancel(
            t('gui.select_mode'),
            t('gui.mode_selection_hint'),
        )
        root.destroy()
        if choice is None:
            return 'export'
        return 'pipeline' if choice else 'diff'
    except Exception:
        return cli_askmode()


def _gui_askmode_full():
    """完整的GUI模式选择界面"""
    if not hasattr(tk, 'Tk') or not callable(tk.Tk):
        raise RuntimeError('Tk not available')

    root = tk.Tk()
    if not hasattr(root, 'mainloop') or not callable(root.mainloop):
        root.destroy()
        raise RuntimeError('Mock Tk detected')

    root.title(t('gui.mode_window_title'))
    root.geometry('480x730')
    root.resizable(False, False)

    result = {'mode': None}

    def select_mode(mode):
        result['mode'] = mode
        root.destroy()

    tk.Label(root, text=t('gui.select_mode'), font=('Arial', 16, 'bold')).pack(pady=20)

    button_frame = tk.Frame(root)
    button_frame.pack(pady=10)

    modes = [
        (t('modes.pipeline_name'), 'pipeline', t('modes.pipeline_desc'), '#4CAF50'),
        (t('modes.diff_name'), 'diff', t('modes.diff_desc'), '#2196F3'),
        (t('modes.monitor_name'), 'monitor', t('modes.monitor_desc'), '#FF9800'),
        (t('modes.scheduler_name'), 'scheduler', t('modes.scheduler_desc'), '#9C27B0'),
        (t('modes.export_name'), 'export', t('modes.export_desc'), '#607D8B'),
        (t('modes.subject_summary_name'), 'subject_summary', t('modes.subject_summary_desc'), '#3F51B5'),
        (t('modes.balance_check_name'), 'balance_check', t('modes.balance_check_desc'), '#8BC34A'),
        (t('modes.duplicate_check_name'), 'duplicate_check', t('modes.duplicate_check_desc'), '#F44336'),
        (t('modes.interest_fee_check_name'), 'interest_fee_check', t('modes.interest_fee_check_desc'), '#009688'),
        (t('modes.balance_reconciliation_name'), 'balance_reconciliation', t('modes.balance_reconciliation_desc'), '#9C27B0'),
        (t('modes.db_query_name'), 'db_query', t('modes.db_query_desc'), '#00BCD4'),
        (t('modes.db_stats_name'), 'db_stats', t('modes.db_stats_desc'), '#795548'),
        (t('modes.batch_history_name'), 'batch_history', t('modes.batch_history_desc'), '#E91E63'),
        (t('modes.preset_name'), 'preset', t('modes.preset_desc'), '#FF5722'),
    ]

    for i, (name, mode, desc, color) in enumerate(modes):
        row = i // 3
        col = i % 3
        btn = tk.Button(
            button_frame,
            text=name,
            width=14,
            height=3,
            bg=color,
            fg='white',
            font=('Arial', 11, 'bold'),
            command=lambda m=mode: select_mode(m),
        )
        btn.grid(row=row, column=col, padx=8, pady=8)
        tk.Label(button_frame, text=desc, font=('Arial', 8), fg='#666').grid(
            row=row * 2 + 1, column=col, padx=8, pady=(0, 8)
        )

    tk.Button(
        root,
        text=t('gui.exit'),
        width=12,
        command=lambda: select_mode(None),
        bg='#f44336',
        fg='white',
        font=('Arial', 10, 'bold'),
    ).pack(pady=20)

    root.mainloop()
    return result['mode']


def _ask_monitor_or_scheduler():
    """二级菜单：选择监控或调度"""
    if tk is not None:
        try:
            root = tk.Tk()
            root.withdraw()
            choice = messagebox.askyesnocancel(
                '选择功能',
                '是 = 监控面板：运行监控与告警管理\n\n否 = 定时调度：定时批处理调度管理\n\n取消 = 返回主菜单',
            )
            root.destroy()
            if choice is None:
                return 'monitor'
            return 'monitor' if choice else 'scheduler'
        except Exception:
            pass

    print('\n请选择功能：')
    print('  1) 监控面板：运行监控与告警管理')
    print('  2) 定时调度：定时批处理调度管理')
    print('  3) 返回主菜单')
    choice = input('请输入选项: ').strip()
    if choice == '1':
        return 'monitor'
    elif choice == '2':
        return 'scheduler'
    return 'monitor'


def cli_ask_monitor_export_menu():
    """二级菜单：选择监控/调度/财务导出"""
    print('\n请选择功能：')
    print('  1) 监控面板：运行监控与告警管理')
    print('  2) 定时调度：定时批处理调度管理')
    print('  3) 财务导出：按用友/金蝶等模板导出凭证或日记账')
    print('  4) 返回主菜单')
    choice = input('请输入选项: ').strip()
    if choice == '1':
        return 'monitor'
    elif choice == '2':
        return 'scheduler'
    elif choice == '3':
        return 'export'
    return 'monitor'


# 根据 tkinter 是否可用，选择交互方式
if HAS_TKINTER:
    ask_directory = gui_askdirectory
    show_info = gui_showinfo
    show_warning = gui_showwarning
    ask_file = gui_askfile
    ask_mode = gui_askmode
else:
    ask_directory = cli_askdirectory
    show_info = cli_showinfo
    show_warning = cli_showwarning
    ask_file = cli_askfile
    ask_mode = cli_askmode


# ──────────────────────────────────────────────
# 日志配置
# ──────────────────────────────────────────────

def get_script_dir():
    """获取脚本（或打包后的 exe）所在目录"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


LOG_ROTATION_WHEN = 'midnight'
LOG_ROTATION_INTERVAL = 1
LOG_BACKUP_COUNT = 0
LOG_MAX_BYTES = 50 * 1024 * 1024
LOG_RETENTION_DAYS = 30
LOG_DIR_NAME = 'logs'

_CURRENT_BATCH_LOG_FILE = None


def get_log_dir(script_dir: Optional[str] = None) -> str:
    if script_dir is None:
        script_dir = get_script_dir()
    log_dir = os.path.join(script_dir, LOG_DIR_NAME)
    os.makedirs(log_dir, exist_ok=True)
    return log_dir


def generate_batch_log_filename(prefix: str = 'bankcheck') -> str:
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    return f'{prefix}_{timestamp}.log'


def get_current_log_file() -> Optional[str]:
    return _CURRENT_BATCH_LOG_FILE


def cleanup_expired_logs(log_dir: str,
                         retention_days: int = LOG_RETENTION_DAYS,
                         prefix: str = 'bankcheck') -> int:
    import time
    import re
    if not os.path.isdir(log_dir) or retention_days <= 0:
        return 0

    cutoff_time = time.time() - (retention_days * 86400)
    removed_count = 0
    pattern = re.compile(rf'^{re.escape(prefix)}_\d{{8}}_\d{{6}}(\.\d+)?\.log(?:\.\d+)?$')

    try:
        for filename in os.listdir(log_dir):
            filepath = os.path.join(log_dir, filename)
            if not os.path.isfile(filepath):
                continue
            if not pattern.match(filename):
                continue
            try:
                mtime = os.path.getmtime(filepath)
                if mtime < cutoff_time:
                    os.remove(filepath)
                    removed_count += 1
            except OSError:
                continue
    except OSError:
        pass

    return removed_count


def find_latest_log_file_in_dir(log_dir: str,
                                prefix: str = 'bankcheck') -> Optional[str]:
    import re
    if not os.path.isdir(log_dir):
        return None

    pattern = re.compile(rf'^{re.escape(prefix)}_\d{{8}}_\d{{6}}\.log$')
    candidates = []
    for filename in os.listdir(log_dir):
        if pattern.match(filename):
            filepath = os.path.join(log_dir, filename)
            try:
                mtime = os.path.getmtime(filepath)
                candidates.append((mtime, filepath))
            except OSError:
                continue

    if not candidates:
        plain_log = os.path.join(os.path.dirname(log_dir), f'{prefix}.log')
        if os.path.isfile(plain_log):
            return plain_log
        return None

    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def setup_logging():
    """
    初始化日志系统（带 PII 脱敏）。
    - 控制台输出 INFO 级别及以上日志（严格脱敏）
    - 日志文件按运行批次切分，格式: bankcheck_YYYYMMDD_HHMMSS.log
      文件保存在脚本/exe 所在目录的 logs/ 子目录下
    - 支持按大小轮转（默认 50MB）和按保留天数清理（默认 30 天）
    - 所有 handler 均附加 PIILogFilter，确保敏感字段不被落盘
    """
    global _CURRENT_BATCH_LOG_FILE
    log_dir = get_log_dir()
    log_filename = generate_batch_log_filename('bankcheck')
    log_file = os.path.join(log_dir, log_filename)
    _CURRENT_BATCH_LOG_FILE = log_file

    try:
        removed = cleanup_expired_logs(log_dir, retention_days=LOG_RETENTION_DAYS)
    except Exception:
        removed = 0

    if HAS_PII_CLASSIFIER:
        logger = setup_pii_aware_logging(
            logger_name='bankcheck',
            log_file=log_file,
            console_level=logging.INFO,
            file_level=logging.DEBUG,
            rotation_when=LOG_ROTATION_WHEN,
            rotation_interval=LOG_ROTATION_INTERVAL,
            backup_count=LOG_BACKUP_COUNT,
            max_bytes=LOG_MAX_BYTES,
        )
    else:
        from logging.handlers import TimedRotatingFileHandler, RotatingFileHandler
        logger = logging.getLogger('bankcheck')
        logger.setLevel(logging.DEBUG)

        if logger.handlers:
            return logger

        console_handler = logging.StreamHandler(sys.stdout)
        console_handler.setLevel(logging.INFO)
        console_fmt = logging.Formatter(
            '[%(asctime)s] %(levelname)s - %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S',
        )
        console_handler.setFormatter(console_fmt)

        if LOG_MAX_BYTES and LOG_MAX_BYTES > 0:
            file_handler = RotatingFileHandler(
                log_file,
                maxBytes=LOG_MAX_BYTES,
                backupCount=LOG_BACKUP_COUNT,
                encoding='utf-8',
            )
        else:
            file_handler = TimedRotatingFileHandler(
                log_file,
                when=LOG_ROTATION_WHEN,
                interval=LOG_ROTATION_INTERVAL,
                backupCount=LOG_BACKUP_COUNT,
                encoding='utf-8',
            )
            file_handler.suffix = '%Y%m%d'
        file_handler.setLevel(logging.DEBUG)
        file_fmt = logging.Formatter(
            '[%(asctime)s] %(levelname)s [%(funcName)s] %(message)s',
            datefmt='%Y-%m-%d %H:%M:%S',
        )
        file_handler.setFormatter(file_fmt)

        logger.addHandler(console_handler)
        logger.addHandler(file_handler)

    logger.info('日志系统初始化完成，日志文件: %s', log_file)
    if removed > 0:
        logger.info('已清理 %d 个超过 %d 天保留期的历史日志文件', removed, LOG_RETENTION_DAYS)
    if not HAS_TKINTER:
        logger.info('未检测到 tkinter，将使用命令行交互模式')
    return logger


def get_logger():
    """获取名为 'bankcheck' 的全局 logger"""
    return logging.getLogger('bankcheck')


# ──────────────────────────────────────────────
# 文件格式检测：基于 Magic Bytes
# ──────────────────────────────────────────────

XLSX_MAGIC_BYTES = b'PK\x03\x04'
XLS_MAGIC_BYTES = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'


def detect_excel_format(filepath):
    """
    通过文件头 Magic Bytes 检测 Excel 文件的真实格式。

    Args:
        filepath: Excel 文件路径

    Returns:
        str: 'xlsx' 或 'xls' 或 'unknown'
    """
    logger = get_logger()
    try:
        with open(filepath, 'rb') as f:
            header = f.read(8)

        if len(header) < 4:
            logger.debug('文件 %s 头部数据不足，无法检测格式', filepath)
            return 'unknown'

        if header.startswith(XLSX_MAGIC_BYTES):
            logger.debug('文件 %s Magic Bytes 检测为 xlsx 格式 (ZIP/OOXML)', filepath)
            return 'xlsx'

        if header.startswith(XLS_MAGIC_BYTES):
            logger.debug('文件 %s Magic Bytes 检测为 xls 格式 (OLE/BIFF)', filepath)
            return 'xls'

        logger.debug('文件 %s 未识别的 Magic Bytes: %s', filepath, header[:4].hex())
        return 'unknown'
    except Exception as e:
        logger.warning('检测文件格式失败 %s: %s', filepath, e)
        return 'unknown'


def get_extension_format(filepath):
    """
    根据文件扩展名判断 Excel 格式。

    Args:
        filepath: Excel 文件路径

    Returns:
        str: 'xlsx' 或 'xls' 或 'unknown'
    """
    lower_path = filepath.lower()
    if lower_path.endswith('.xlsx') or lower_path.endswith('.xlsm'):
        return 'xlsx'
    if lower_path.endswith('.xls') and not lower_path.endswith('.xlsx'):
        return 'xls'
    return 'unknown'


# ──────────────────────────────────────────────
# .xls 兼容：将 .xls 转换为 .xlsx
# ──────────────────────────────────────────────

def convert_xls_to_xlsx(xls_path):
    """
    将 .xls 文件转换为 .xlsx 文件（临时文件），返回临时 .xlsx 路径。
    使用 xlrd 读取 .xls，再用 openpyxl 写入 .xlsx。
    """
    logger = get_logger()
    try:
        import xlrd
    except ImportError:
        logger.error('处理 .xls 文件需要 xlrd 库，请运行: pip install xlrd')
        raise ImportError('缺少 xlrd 库，无法处理 .xls 文件。请运行: pip install xlrd')

    logger.info('正在将 .xls 文件转换为 .xlsx: %s', xls_path)

    xls_book = xlrd.open_workbook(xls_path)
    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    for sheet_name in xls_book.sheet_names():
        xls_sheet = xls_book.sheet_by_name(sheet_name)
        ws = wb.create_sheet(title=sheet_name)
        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                cell_value = xls_sheet.cell_value(row_idx, col_idx)
                cell_type = xls_sheet.cell_type(row_idx, col_idx)
                # xlrd 日期类型转换
                if cell_type == xlrd.XL_CELL_DATE:
                    try:
                        date_tuple = xlrd.xldate_as_tuple(cell_value, xls_book.datemode)
                        cell_value = datetime(*date_tuple)
                    except Exception:
                        pass
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_value)

    # 保存为临时 .xlsx 文件（使用安全的随机文件名）
    fd, tmp_path = tempfile.mkstemp(prefix='bankcheck_', suffix='.xlsx')
    os.close(fd)
    wb.save(tmp_path)
    wb.close()
    xls_book.release_resources()

    logger.info('.xls 转换完成: %s -> %s', xls_path, tmp_path)
    return tmp_path


def open_workbook_compat(filepath):
    """
    兼容打开 .xlsx 和 .xls 文件，统一返回 (openpyxl.Workbook, 临时文件路径或None)。
    如果是 .xls 文件，先转换为 .xlsx 再打开。
    支持基于 Magic Bytes 的自动格式检测，当扩展名与实际格式不一致时按真实格式处理。
    调用方负责在使用完毕后清理临时文件。
    """
    logger = get_logger()
    tmp_path = None

    ext_format = get_extension_format(filepath)
    magic_format = detect_excel_format(filepath)

    actual_format = magic_format if magic_format != 'unknown' else ext_format

    if magic_format != 'unknown' and ext_format != 'unknown' and magic_format != ext_format:
        logger.warning(
            '文件「%s」扩展名与实际格式不一致：扩展名为 %s，实际为 %s，将按 %s 格式解析',
            filepath, ext_format, magic_format, actual_format,
        )

    _profiler_ctx = None
    if HAS_PERF_PROFILER:
        _profiler_ctx = perf_profiler.get_profiler().measure_file_open(
            filepath, is_xls_convert=(actual_format == 'xls')
        )
        _profiler_ctx.__enter__()

    try:
        if actual_format == 'xls':
            tmp_path = convert_xls_to_xlsx(filepath)
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
        else:
            if ext_format == 'xlsx' or ext_format == 'unknown':
                wb = openpyxl.load_workbook(filepath, data_only=True)
            else:
                with open(filepath, 'rb') as f:
                    wb = openpyxl.load_workbook(f, data_only=True)
    finally:
        if _profiler_ctx is not None:
            _profiler_ctx.__exit__(None, None, None)

    return wb, tmp_path


def cleanup_temp_file(tmp_path):
    """清理临时转换的 .xlsx 文件"""
    if tmp_path and os.path.exists(tmp_path):
        try:
            os.remove(tmp_path)
        except OSError:
            pass


# ──────────────────────────────────────────────
# 文件扫描
# ──────────────────────────────────────────────

def scan_excel_files(folder):
    """递归扫描文件夹中的所有 Excel 文件（.xlsx 和 .xls），排除临时文件"""
    logger = get_logger()
    excel_files = []
    for root, _dirs, files in os.walk(folder):
        for f in files:
            if f.startswith('~$'):
                continue
            if f.lower().endswith(('.xlsx', '.xls')):
                full_path = os.path.join(root, f)
                excel_files.append(full_path)
                logger.debug('发现 Excel 文件: %s', full_path)
    logger.info('共扫描到 %d 个 Excel 文件', len(excel_files))
    return excel_files


# ──────────────────────────────────────────────
# 银行规则配置模块
# ──────────────────────────────────────────────

BANK_RULES_CONFIG_FILE = 'bank_rules.yaml'


class HeaderValidationError(Exception):
    pass


@dataclass
class BankRule:
    bank_name: str
    account_cell: str
    start_row: int
    columns: Dict[str, int]
    payment_sign: str = 'negative'
    enabled: bool = True
    skip_sheets: List[str] = field(default_factory=list)
    expected_headers: Dict[str, List[str]] = field(default_factory=dict)
    header_validation: str = 'warn'
    multi_account: bool = False


class BankRuleConfig:
    """银行规则配置管理器 - 单例模式"""
    _instance = None
    _config_path = None
    _rules: Dict[str, BankRule] = field(default_factory=dict)
    _last_modified: float = 0.0

    def __new__(cls, config_path=None):
        if cls._instance is None:
            cls._instance = super(BankRuleConfig, cls).__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self, config_path=None):
        if self._initialized:
            return
        self._initialized = True
        if config_path is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            config_path = os.path.join(script_dir, BANK_RULES_CONFIG_FILE)
        self._config_path = config_path
        self._rules = {}
        self._last_modified = 0.0
        self.load_config()

    def load_config(self):
        """加载配置文件，支持热更新"""
        logger = get_logger()
        if not os.path.exists(self._config_path):
            logger.error('银行规则配置文件不存在: %s', self._config_path)
            return False

        try:
            current_mtime = os.path.getmtime(self._config_path)
            if current_mtime == self._last_modified and self._rules:
                return True

            with open(self._config_path, 'r', encoding='utf-8') as f:
                config_data = yaml.safe_load(f)

            self._rules = {}
            banks_config = config_data.get('banks', [])
            for bank_config in banks_config:
                if not bank_config.get('enabled', True):
                    continue
                raw_headers = bank_config.get('expected_headers', {})
                normalized_headers = {}
                for col_key, header_names in raw_headers.items():
                    if isinstance(header_names, str):
                        normalized_headers[col_key] = [header_names]
                    elif isinstance(header_names, list):
                        normalized_headers[col_key] = header_names
                rule = BankRule(
                    bank_name=bank_config['bank_name'],
                    account_cell=bank_config['account_cell'],
                    start_row=bank_config['start_row'],
                    columns=bank_config['columns'],
                    payment_sign=bank_config.get('payment_sign', 'negative'),
                    enabled=bank_config.get('enabled', True),
                    skip_sheets=bank_config.get('skip_sheets', []),
                    expected_headers=normalized_headers,
                    header_validation=bank_config.get('header_validation', 'warn'),
                    multi_account=bank_config.get('multi_account', False),
                )
                self._rules[rule.bank_name] = rule

            self._last_modified = current_mtime
            logger.info('已加载 %d 个银行规则配置', len(self._rules))
            return True
        except Exception as e:
            logger.error('加载银行规则配置失败: %s', e, exc_info=True)
            return False

    def get_rule(self, bank_name: str) -> Optional[BankRule]:
        """根据银行名称获取规则，自动检查配置更新"""
        self.load_config()
        return self._rules.get(bank_name)

    def get_all_bank_names(self) -> List[str]:
        """获取所有已启用的银行名称列表，自动检查配置更新"""
        self.load_config()
        return list(self._rules.keys())

    def get_config_path(self) -> str:
        """获取配置文件路径"""
        return self._config_path

    def get_all_rules(self) -> Dict[str, BankRule]:
        """获取所有银行规则（含未启用），返回副本"""
        self.load_config()
        return dict(self._rules)

    def list_rules_detailed(self) -> List[Dict[str, Any]]:
        """获取所有银行规则的详细信息列表"""
        self.load_config()
        result = []
        for name, rule in self._rules.items():
            result.append({
                'bank_name': rule.bank_name,
                'account_cell': rule.account_cell,
                'start_row': rule.start_row,
                'columns': dict(rule.columns),
                'payment_sign': rule.payment_sign,
                'enabled': rule.enabled,
                'skip_sheets': list(rule.skip_sheets),
                'expected_headers': {k: list(v) for k, v in rule.expected_headers.items()},
                'header_validation': rule.header_validation,
                'multi_account': rule.multi_account,
            })
        return result

    def save_rule(self, rule_data: Dict[str, Any]) -> bool:
        """
        保存（新增或更新）一条银行规则到 YAML 配置文件。

        Args:
            rule_data: 包含银行规则字段的字典

        Returns:
            是否保存成功
        """
        logger = get_logger()
        try:
            if not os.path.exists(self._config_path):
                config_data = {'banks': []}
            else:
                with open(self._config_path, 'r', encoding='utf-8') as f:
                    config_data = yaml.safe_load(f) or {}
                if 'banks' not in config_data:
                    config_data['banks'] = []

            bank_name = rule_data.get('bank_name', '').strip()
            if not bank_name:
                logger.error('保存银行规则失败：银行名称不能为空')
                return False

            expected_headers = rule_data.get('expected_headers', {})
            normalized_headers = {}
            for col_key, header_names in expected_headers.items():
                if isinstance(header_names, str):
                    normalized_headers[col_key] = [header_names]
                elif isinstance(header_names, list):
                    normalized_headers[col_key] = [h for h in header_names if h]

            new_entry = {
                'bank_name': bank_name,
                'account_cell': rule_data.get('account_cell', 'A1'),
                'start_row': int(rule_data.get('start_row', 1)),
                'payment_sign': rule_data.get('payment_sign', 'negative'),
                'enabled': bool(rule_data.get('enabled', True)),
                'columns': rule_data.get('columns', {}),
            }
            if rule_data.get('skip_sheets'):
                new_entry['skip_sheets'] = rule_data['skip_sheets']
            if normalized_headers:
                new_entry['expected_headers'] = normalized_headers
            if rule_data.get('header_validation'):
                new_entry['header_validation'] = rule_data['header_validation']
            if rule_data.get('multi_account'):
                new_entry['multi_account'] = bool(rule_data['multi_account'])

            replaced = False
            for i, bank_cfg in enumerate(config_data['banks']):
                if bank_cfg.get('bank_name') == bank_name:
                    config_data['banks'][i] = new_entry
                    replaced = True
                    break
            if not replaced:
                config_data['banks'].append(new_entry)

            with open(self._config_path, 'w', encoding='utf-8') as f:
                yaml.safe_dump(config_data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)

            self._last_modified = 0.0
            self.load_config()
            reload_bank_processors()

            logger.info('银行规则「%s」已保存到配置文件', bank_name)
            return True
        except Exception as e:
            logger.error('保存银行规则失败: %s', e, exc_info=True)
            return False

    def delete_rule(self, bank_name: str) -> bool:
        """
        从 YAML 配置文件中删除指定银行规则。

        Args:
            bank_name: 银行名称

        Returns:
            是否删除成功
        """
        logger = get_logger()
        try:
            if not os.path.exists(self._config_path):
                return False
            with open(self._config_path, 'r', encoding='utf-8') as f:
                config_data = yaml.safe_load(f) or {}
            if 'banks' not in config_data:
                return False
            original_len = len(config_data['banks'])
            config_data['banks'] = [b for b in config_data['banks'] if b.get('bank_name') != bank_name]
            if len(config_data['banks']) == original_len:
                return False
            with open(self._config_path, 'w', encoding='utf-8') as f:
                yaml.safe_dump(config_data, f, allow_unicode=True, default_flow_style=False, sort_keys=False)
            self._last_modified = 0.0
            self.load_config()
            reload_bank_processors()
            logger.info('银行规则「%s」已从配置文件中删除', bank_name)
            return True
        except Exception as e:
            logger.error('删除银行规则失败: %s', e, exc_info=True)
            return False


def col_letter_to_index(letter: str) -> int:
    """
    将 Excel 列字母（如 A, B, ..., Z, AA, AB）转换为 1-based 列号。

    Args:
        letter: 列字母（大小写不敏感）

    Returns:
        1-based 列号
    """
    letter = letter.upper().strip()
    if not letter or not letter.isalpha():
        raise ValueError(f'无效的列字母: {letter}')
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result


def col_index_to_letter(index: int) -> str:
    """
    将 1-based 列号转换为 Excel 列字母。

    Args:
        index: 1-based 列号

    Returns:
        列字母（大写）
    """
    if index < 1:
        raise ValueError(f'无效的列号: {index}')
    result = ''
    while index > 0:
        index, remainder = divmod(index - 1, 26)
        result = chr(ord('A') + remainder) + result
    return result


def parse_cell_ref(cell_ref: str) -> tuple:
    """
    解析 Excel 单元格引用（如 'B2'）为 (1-based_col_index, 1-based_row_index)。

    Args:
        cell_ref: 单元格引用字符串，如 'B2'、'AA10'

    Returns:
        (col_index, row_index) 元组，均为 1-based
    """
    if not cell_ref:
        raise ValueError('单元格引用不能为空')
    match = re.match(r'^([A-Za-z]+)(\d+)$', cell_ref.strip())
    if not match:
        raise ValueError(f'无效的单元格引用: {cell_ref}')
    col_letter = match.group(1)
    row_str = match.group(2)
    col_index = col_letter_to_index(col_letter)
    row_index = int(row_str)
    if row_index < 1:
        raise ValueError(f'行号必须 >= 1: {row_index}')
    return col_index, row_index


def get_cell_ref(col_index: int, row_index: int) -> str:
    """
    根据列号和行号生成 Excel 单元格引用。

    Args:
        col_index: 1-based 列号
        row_index: 1-based 行号

    Returns:
        单元格引用字符串，如 'B2'
    """
    return f"{col_index_to_letter(col_index)}{row_index}"


def read_excel_preview(filepath: str, sheet_name: Optional[str] = None,
                       max_rows: int = 50, max_cols: int = 30) -> Dict[str, Any]:
    """
    读取 Excel 文件预览数据，用于向导中交互式选择。

    Args:
        filepath: Excel 文件路径
        sheet_name: 工作表名称，None 时取第一个工作表
        max_rows: 最多读取的行数
        max_cols: 最多读取的列数

    Returns:
        包含以下字段的字典：
        - sheet_names: 所有工作表名称列表
        - current_sheet: 当前工作表名称
        - data: 二维列表形式的单元格数据 [row_index][col_index]，索引从 0 开始
        - cell_refs: 对应的单元格引用二维列表
        - max_row: 实际总行数
        - max_col: 实际总列数
    """
    logger = get_logger()
    wb, tmp_path = open_workbook_compat(filepath)
    try:
        sheet_names = [ws.title for ws in wb.worksheets]
        if sheet_name:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            sheet_name = ws.title

        actual_max_row = min(ws.max_row, max_rows)
        actual_max_col = min(ws.max_column, max_cols)

        data = []
        cell_refs = []
        for row_idx in range(1, actual_max_row + 1):
            row_data = []
            row_refs = []
            for col_idx in range(1, actual_max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                val = cell.value
                if val is not None:
                    row_data.append(str(val))
                else:
                    row_data.append('')
                row_refs.append(get_cell_ref(col_idx, row_idx))
            data.append(row_data)
            cell_refs.append(row_refs)

        result = {
            'sheet_names': sheet_names,
            'current_sheet': sheet_name,
            'data': data,
            'cell_refs': cell_refs,
            'max_row': actual_max_row,
            'max_col': actual_max_col,
            'total_rows': ws.max_row,
            'total_cols': ws.max_column,
        }
        wb.close()
        logger.info('已读取 Excel 预览: %s, 工作表=%s, %d行 x %d列',
                    os.path.basename(filepath), sheet_name, actual_max_row, actual_max_col)
        return result
    finally:
        cleanup_temp_file(tmp_path)


def preview_extraction(filepath: str, rule_data: Dict[str, Any],
                       sheet_name: Optional[str] = None,
                       max_preview_rows: int = 10) -> Dict[str, Any]:
    """
    根据给定的规则配置预览提取结果，不写入任何文件。

    Args:
        filepath: Excel 文件路径
        rule_data: 规则配置字典（含 bank_name, account_cell, start_row, columns 等）
        sheet_name: 指定工作表，None 时取第一个
        max_preview_rows: 最多预览的记录行数

    Returns:
        包含预览结果的字典：
        - success: 是否成功
        - account: 提取到的账号
        - header_values: 表头行各字段对应的实际值
        - records: 提取到的记录列表（限制条数）
        - total_records: 总记录数
        - error: 错误信息（如有）
    """
    logger = get_logger()
    try:
        columns = rule_data.get('columns', {})
        if not columns:
            return {'success': False, 'error': '未配置任何列映射', 'records': [], 'total_records': 0}

        required = ['trade_date']
        for f in required:
            if f not in columns:
                return {'success': False,
                        'error': f'缺少必填字段列映射: {f}',
                        'records': [], 'total_records': 0}

        temp_rule = BankRule(
            bank_name=rule_data.get('bank_name', '预览银行'),
            account_cell=rule_data.get('account_cell', 'A1'),
            start_row=int(rule_data.get('start_row', 1)),
            columns={k: int(v) for k, v in columns.items()},
            payment_sign=rule_data.get('payment_sign', 'negative'),
            enabled=True,
            skip_sheets=rule_data.get('skip_sheets', []),
            expected_headers=rule_data.get('expected_headers', {}),
            header_validation='off',
            multi_account=bool(rule_data.get('multi_account', False)),
        )
        parser = GenericBankParser(temp_rule)

        wb, tmp_path = open_workbook_compat(filepath)
        try:
            if sheet_name:
                ws_list = [wb[sheet_name]]
            else:
                ws_list = wb.worksheets

            all_records = []
            account_val = None
            header_values = {}

            for ws in ws_list:
                if ws.title in (temp_rule.skip_sheets or []):
                    continue
                try:
                    account_cell = ws[temp_rule.account_cell]
                    if account_cell.value is not None:
                        account_val = str(account_cell.value).strip()
                except Exception:
                    pass

                header_row_idx = temp_rule.start_row - 1
                if header_row_idx >= 1:
                    for col_key, col_idx in temp_rule.columns.items():
                        try:
                            hv = ws.cell(row=header_row_idx, column=col_idx).value
                            if hv is not None:
                                header_values[col_key] = str(hv).strip()
                        except Exception:
                            pass

                sheet_records = parser._parse_sheet(ws, filepath, ws.title, {})
                all_records.extend(sheet_records)

            wb.close()

            preview_records = all_records[:max_preview_rows]
            for r in preview_records:
                r.pop('唯一id', None)

            logger.info('预览提取完成: %d 条记录（预览前 %d 条）',
                        len(all_records), len(preview_records))

            return {
                'success': True,
                'account': account_val or '',
                'header_values': header_values,
                'records': preview_records,
                'total_records': len(all_records),
            }
        finally:
            cleanup_temp_file(tmp_path)
    except Exception as e:
        logger.error('预览提取失败: %s', e, exc_info=True)
        return {
            'success': False,
            'error': str(e),
            'records': [],
            'total_records': 0,
        }


class GenericBankParser:
    """通用银行流水解析器 - 根据配置规则动态解析 Excel"""

    def __init__(self, rule: BankRule):
        self.rule = rule
        self.logger = get_logger()

    def validate_headers(self, ws, filepath: str, sheet_name: str) -> List[str]:
        """
        校验工作表表头与预期是否一致，返回不匹配的字段列表。

        表头行取 start_row - 1，逐列检查 columns 中每个字段对应的单元格文本
        是否在 expected_headers 指定的候选名称中。

        Args:
            ws: openpyxl Worksheet 对象
            filepath: 文件路径（用于日志）
            sheet_name: 工作表名称（用于日志）

        Returns:
            不匹配的字段名列表（空列表表示全部匹配或未配置 expected_headers）
        """
        expected = self.rule.expected_headers
        if not expected or self.rule.header_validation == 'off':
            return []

        header_row = self.rule.start_row - 1
        if header_row < 1:
            header_row = 1

        mismatches = []
        for col_key, acceptable_names in expected.items():
            col_idx = self.rule.columns.get(col_key)
            if col_idx is None:
                continue

            actual_value = ws.cell(row=header_row, column=col_idx).value
            actual_text = str(actual_value).strip() if actual_value is not None else ''

            if not actual_text:
                mismatches.append(col_key)
                self.logger.warning(
                    '%s文件工作表「%s」表头校验：字段「%s」第 %d 列第 %d 行单元格为空'
                    '，预期为「%s」',
                    self.rule.bank_name, sheet_name, col_key,
                    col_idx, header_row, '」或「'.join(acceptable_names))
                continue

            if actual_text not in acceptable_names:
                mismatches.append(col_key)
                self.logger.warning(
                    '%s文件工作表「%s」表头校验：字段「%s」第 %d 列实际表头为「%s」'
                    '，预期为「%s」',
                    self.rule.bank_name, sheet_name, col_key,
                    col_idx, actual_text, '」或「'.join(acceptable_names))

        if not mismatches:
            self.logger.info(
                '%s文件工作表「%s」表头校验通过，%d 个字段均匹配',
                self.rule.bank_name, sheet_name, len(expected))
        else:
            self.logger.warning(
                '%s文件工作表「%s」表头校验未通过：%d/%d 个字段不匹配（%s）',
                self.rule.bank_name, sheet_name,
                len(mismatches), len(expected), ', '.join(mismatches))

        return mismatches

    def _detect_account_blocks(self, ws):
        """
        扫描工作表中是否包含多个账号区块。

        在账号所在列中查找所有看起来像银行账号的值（6位及以上纯数字），
        并根据配置中的 account_cell 与 start_row 偏移量，推算每个区块的
        表头行和数据起始行。

        Returns:
            列表，每项为字典：
            {
                'account': 账号字符串,
                'account_row': 账号所在行号,
                'header_row': 表头行号,
                'data_start_row': 数据起始行号,
                'data_end_row': 数据结束行号（含）,
            }
            仅一个账号时返回单元素列表；无账号时返回空列表。
        """
        cell_match = re.match(r'^([A-Z]+)(\d+)$', self.rule.account_cell.upper())
        if not cell_match:
            return []

        col_letter = cell_match.group(1)
        original_row = int(cell_match.group(2))

        col_idx = 0
        for ch in col_letter:
            col_idx = col_idx * 26 + (ord(ch) - ord('A') + 1)

        header_offset = (self.rule.start_row - 1) - original_row
        data_offset = self.rule.start_row - original_row

        account_blocks = []
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is None:
                continue
            cell_str = _normalize_width(str(cell_value).strip())
            if re.match(r'^\d{6,}$', cell_str):
                header_row = row_idx + header_offset
                data_start_row = row_idx + data_offset
                if header_row < 1:
                    header_row = 1
                if data_start_row < 1:
                    data_start_row = 1
                account_blocks.append({
                    'account': cell_str,
                    'account_row': row_idx,
                    'header_row': header_row,
                    'data_start_row': data_start_row,
                })

        for i, block in enumerate(account_blocks):
            if i + 1 < len(account_blocks):
                next_account_row = account_blocks[i + 1]['account_row']
                block['data_end_row'] = min(
                    next_account_row - 1,
                    ws.max_row,
                )
            else:
                block['data_end_row'] = ws.max_row

        return account_blocks

    def _validate_headers_at_row(self, ws, filepath, sheet_name, header_row):
        """
        在指定行校验表头，返回不匹配的字段列表。
        """
        expected = self.rule.expected_headers
        if not expected or self.rule.header_validation == 'off':
            return []

        if header_row < 1:
            header_row = 1

        mismatches = []
        for col_key, acceptable_names in expected.items():
            col_idx = self.rule.columns.get(col_key)
            if col_idx is None:
                continue

            actual_value = ws.cell(row=header_row, column=col_idx).value
            actual_text = str(actual_value).strip() if actual_value is not None else ''

            if not actual_text:
                mismatches.append(col_key)
                continue

            if actual_text not in acceptable_names:
                mismatches.append(col_key)

        return mismatches

    def _parse_segment(self, ws, filepath, sheet_name, lookup_source,
                       account_value, data_start_row, data_end_row):
        """
        解析工作表中指定行范围的数据段，用于多账号场景。

        Args:
            ws: openpyxl Worksheet 对象
            filepath: 文件路径
            sheet_name: 工作表名称
            lookup_source: 查找表
            account_value: 当前段的银行账号
            data_start_row: 数据起始行号
            data_end_row: 数据结束行号（含）

        Returns:
            当前段的记录列表
        """
        subject_info = get_subject_info(account_value, lookup_source)
        subject = subject_info.get('subject', '')
        extra_fields = subject_info.get('extra_fields', {})

        rows = []
        columns = self.rule.columns

        for row_idx in range(data_start_row, data_end_row + 1):
            trade_date = ws.cell(row=row_idx, column=columns['trade_date']).value
            if trade_date is None:
                continue

            payment_val = ws.cell(row=row_idx, column=columns['payment']).value
            if is_numeric(payment_val):
                payment = to_float(payment_val)
                if self.rule.payment_sign == 'negative':
                    payment = -abs(payment)
            else:
                payment = None

            receipt_val = ws.cell(row=row_idx, column=columns['receipt']).value
            receipt = to_float(receipt_val) if is_numeric(receipt_val) else None

            summary = ws.cell(row=row_idx, column=columns['summary']).value
            counterpart = ws.cell(row=row_idx, column=columns['counterpart']).value
            balance = ws.cell(row=row_idx, column=columns['balance']).value
            transaction_id = ws.cell(row=row_idx, column=columns['transaction_id']).value

            record = {
                '唯一id': generate_unique_id(),
                '银行': self.rule.bank_name,
                '银行账号': account_value,
                '主体': subject,
                '交易日期': trade_date,
                '付款': payment,
                '收款': receipt,
                '摘要': summary,
                '对方户名': counterpart,
                '余额': balance,
                '交易流水号': transaction_id,
            }
            for key, val in extra_fields.items():
                record[key] = val

            rows.append(record)

        if rows:
            self.logger.info(
                '%s文件工作表「%s」账号「%s」提取 %d 条记录',
                self.rule.bank_name, sheet_name,
                _mask_bank_account(account_value) if HAS_PII_CLASSIFIER else account_value,
                len(rows))
        else:
            self.logger.info(
                '%s文件工作表「%s」账号「%s」未提取到记录',
                self.rule.bank_name, sheet_name,
                _mask_bank_account(account_value) if HAS_PII_CLASSIFIER else account_value)

        return rows

    def _parse_sheet_multi_account(self, ws, filepath, sheet_name, lookup_source):
        """
        多账号单文件拆分解析：自动检测同一工作表中的多个账号区块，
        按账号切段并分别匹配主体后返回记录列表。
        """
        blocks = self._detect_account_blocks(ws)
        if not blocks:
            self.logger.info(
                '%s文件工作表「%s」未检测到账号区块，跳过',
                self.rule.bank_name, sheet_name)
            return []

        if len(blocks) == 1:
            self.logger.info(
                '%s文件工作表「%s」仅检测到 1 个账号区块，按单账号处理',
                self.rule.bank_name, sheet_name)

        all_rows = []
        for block in blocks:
            account_value = block['account']
            header_row = block['header_row']
            data_start_row = block['data_start_row']
            data_end_row = block['data_end_row']

            mismatches = self._validate_headers_at_row(
                ws, filepath, sheet_name, header_row)
            if mismatches and self.rule.header_validation == 'strict':
                detail = ', '.join(mismatches)
                raise HeaderValidationError(
                    f'{self.rule.bank_name}文件工作表「{sheet_name}」'
                    f'账号「{account_value}」表头校验失败，不匹配字段: {detail}')
            elif mismatches:
                self.logger.warning(
                    '%s文件工作表「%s」账号「%s」表头校验未通过：%d 个字段不匹配（%s），继续处理',
                    self.rule.bank_name, sheet_name, account_value,
                    len(mismatches), ', '.join(mismatches))

            segment_rows = self._parse_segment(
                ws, filepath, sheet_name, lookup_source,
                account_value, data_start_row, data_end_row)
            all_rows.extend(segment_rows)

        self.logger.info(
            '%s文件工作表「%s」多账号拆分完成，共 %d 个账号区块，提取 %d 条记录',
            self.rule.bank_name, sheet_name, len(blocks), len(all_rows))

        return all_rows

    def _parse_sheet(self, ws, filepath: str, sheet_name: str,
                     lookup_source) -> List[Dict[str, Any]]:
        if self.rule.multi_account:
            return self._parse_sheet_multi_account(ws, filepath, sheet_name, lookup_source)

        mismatches = self.validate_headers(ws, filepath, sheet_name)

        if mismatches and self.rule.header_validation == 'strict':
            detail = ', '.join(mismatches)
            raise HeaderValidationError(
                f'{self.rule.bank_name}文件工作表「{sheet_name}」'
                f'表头校验失败，不匹配字段: {detail}')

        bank_account = ws[self.rule.account_cell].value
        if bank_account is None:
            self.logger.warning(
                '文件「%s」工作表「%s」%s 单元格为空，银行账号缺失',
                filepath, sheet_name, self.rule.account_cell)

        subject_info = get_subject_info(bank_account, lookup_source)
        subject = subject_info.get('subject', '')
        extra_fields = subject_info.get('extra_fields', {})

        _traversal_start = None
        if HAS_PERF_PROFILER:
            import time as _time
            _traversal_start = _time.perf_counter()

        rows = []
        columns = self.rule.columns
        start_row = self.rule.start_row

        for row_idx in range(start_row, ws.max_row + 1):
            trade_date = ws.cell(row=row_idx, column=columns['trade_date']).value
            if trade_date is None:
                continue

            payment_val = ws.cell(row=row_idx, column=columns['payment']).value
            if is_numeric(payment_val):
                payment = to_float(payment_val)
                if self.rule.payment_sign == 'negative':
                    payment = -abs(payment)
            else:
                payment = None

            receipt_val = ws.cell(row=row_idx, column=columns['receipt']).value
            receipt = to_float(receipt_val) if is_numeric(receipt_val) else None

            summary = ws.cell(row=row_idx, column=columns['summary']).value
            counterpart = ws.cell(row=row_idx, column=columns['counterpart']).value
            balance = ws.cell(row=row_idx, column=columns['balance']).value
            transaction_id = ws.cell(row=row_idx, column=columns['transaction_id']).value

            record = {
                '唯一id': generate_unique_id(),
                '银行': self.rule.bank_name,
                '银行账号': bank_account,
                '主体': subject,
                '交易日期': trade_date,
                '付款': payment,
                '收款': receipt,
                '摘要': summary,
                '对方户名': counterpart,
                '余额': balance,
                '交易流水号': transaction_id,
            }
            for key, val in extra_fields.items():
                record[key] = val

            rows.append(record)

        if _traversal_start is not None:
            import time as _time
            _duration_ms = (_time.perf_counter() - _traversal_start) * 1000
            perf_profiler.get_profiler().record_row_traversal(
                filepath, sheet_name,
                ws.max_row - start_row + 1, _duration_ms,
                self.rule.bank_name, len(rows),
            )

        if rows:
            self.logger.info(
                '%s文件工作表「%s」提取 %d 条记录',
                self.rule.bank_name, sheet_name, len(rows))
        else:
            self.logger.info(
                '%s文件工作表「%s」未提取到记录，跳过',
                self.rule.bank_name, sheet_name)

        return rows

    def parse(self, filepath: str, lookup_source) -> List[Dict[str, Any]]:
        """
        根据配置规则解析银行流水 Excel 文件，支持多工作表遍历。

        当流水数据分布在多个 Sheet 时，按同一银行规则分别提取并合并到总表。

        Args:
            filepath: Excel 文件路径
            lookup_source: 查找表文件路径(str) 或 load_lookup_table() 返回的预加载 dict

        Returns:
            解析后的记录列表（所有工作表合并结果）
        """
        self.logger.info('开始处理%s文件: %s', self.rule.bank_name, filepath)

        wb, tmp_path = open_workbook_compat(filepath)
        try:
            all_rows = []
            skip_sheets = self.rule.skip_sheets or []

            for ws in wb.worksheets:
                if ws.title in skip_sheets:
                    self.logger.info(
                        '%s文件工作表「%s」在 skip_sheets 中，跳过',
                        self.rule.bank_name, ws.title)
                    continue

                sheet_rows = self._parse_sheet(ws, filepath, ws.title, lookup_source)
                all_rows.extend(sheet_rows)

            wb.close()
            self.logger.info('%s文件处理完成，共提取 %d 条记录',
                             self.rule.bank_name, len(all_rows))
            return all_rows
        finally:
            cleanup_temp_file(tmp_path)


def _create_bank_processor(bank_name: str):
    """创建基于配置的银行处理器函数"""
    def processor(filepath, lookup_source):
        config = BankRuleConfig()
        rule = config.get_rule(bank_name)
        if rule is None:
            logger = get_logger()
            logger.error('未找到银行「%s」的解析规则', bank_name)
            return []
        parser = GenericBankParser(rule)
        return parser.parse(filepath, lookup_source)
    return processor


_bank_config_singleton = None


def get_bank_config():
    """获取银行规则配置单例"""
    global _bank_config_singleton
    if _bank_config_singleton is None:
        _bank_config_singleton = BankRuleConfig()
    return _bank_config_singleton


# ──────────────────────────────────────────────
# 银行识别
# ──────────────────────────────────────────────

def _get_bank_prefixes():
    """从配置获取银行前缀列表"""
    config = get_bank_config()
    return config.get_all_bank_names()


BANK_PREFIXES = _get_bank_prefixes()


_SEP_PATTERN = r'[\s\-_·.．（）()\[\]【】]'


def _normalize_width(s: str) -> str:
    """将字符串中的全角字符转换为半角字符，统一符号宽度"""
    if not s:
        return s
    result = []
    for ch in s:
        code = ord(ch)
        if code == 0x3000:
            result.append(' ')
        elif 0xFF01 <= code <= 0xFF5E:
            result.append(chr(code - 0xFEE0))
        else:
            result.append(unicodedata.normalize('NFKC', ch))
    return ''.join(result)


def _strip_separators(s: str) -> str:
    """移除字符串中的分隔符（空格、横线、下划线、括号等）"""
    return re.sub(_SEP_PATTERN, '', s)


def _build_bank_regex(bank_name: str) -> re.Pattern:
    """
    为银行名构建正则模式，支持：
    - 字符间存在任意数量的分隔符（空格、横线、下划线、点、括号等）
    - 全角/半角符号差异
    """
    normalized = _normalize_width(bank_name)
    stripped = _strip_separators(normalized)
    chars = list(stripped)
    if not chars:
        return re.compile(r'(?!x)x')
    sep = _SEP_PATTERN + '*'
    pattern = sep.join(re.escape(c) for c in chars)
    return re.compile(pattern, re.IGNORECASE)


def _match_bank_in_filename(basename: str, bank_name: str) -> bool:
    """在文件名中任意位置匹配银行名，忽略全角/半角差异和分隔符"""
    normalized_name = _normalize_width(basename)
    pattern = _build_bank_regex(bank_name)
    return pattern.search(normalized_name) is not None


def _safe_get_cell_value(ws, cell_ref: str):
    """安全获取单元格值，不存在则返回 None"""
    try:
        cell = ws[cell_ref]
        return cell.value
    except (KeyError, IndexError):
        return None


def _identify_bank_by_content(filepath: str) -> Optional[str]:
    """
    根据 Excel 内容辅助识别银行：
    从各银行配置读取 account_cell，逐一检查对应单元格是否包含账号格式内容。
    当且仅当唯一匹配一个银行时返回结果，否则返回 None。
    """
    logger = get_logger()
    if not os.path.isfile(filepath):
        return None
    if not filepath.lower().endswith(('.xlsx', '.xls')):
        return None

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        try:
            ws = wb.active
            if ws is None:
                return None

            def _cell_to_str(val):
                if val is None:
                    return ''
                return _normalize_width(str(val).strip())

            looks_like_account = lambda s: bool(s) and bool(re.match(r'^\d{6,}$', s))

            config = get_bank_config()
            bank_names = config.get_all_bank_names()

            cell_values = {}
            matched = []
            for bank_name in bank_names:
                rule = config.get_rule(bank_name)
                if not rule:
                    continue
                cell = rule.account_cell.upper()
                if cell not in cell_values:
                    raw = _safe_get_cell_value(ws, cell)
                    cell_values[cell] = _cell_to_str(raw)
                if looks_like_account(cell_values[cell]):
                    matched.append(bank_name)

            if len(matched) == 1:
                logger.info('文件「%s」通过内容特征识别为: %s', os.path.basename(filepath), matched[0])
                return matched[0]
            elif len(matched) > 1:
                logger.warning('文件「%s」内容匹配多个银行: %s，无法确定', os.path.basename(filepath), matched)
        finally:
            wb.close()
    except Exception as e:
        logger.debug('内容辅助识别失败「%s」: %s', os.path.basename(filepath), e)

    return None


def identify_bank(filepath):
    """
    识别银行类型，返回银行名称或 None。

    识别策略（按优先级）：
    1. 文件名前缀精确匹配（兼容性保留）
    2. 文件名任意位置包含银行名的正则匹配（忽略全角/半角差异）
    3. 基于 Excel 内容的 B1/B2 账号单元格特征辅助识别
    """
    logger = get_logger()
    basename = os.path.basename(filepath)

    for prefix in BANK_PREFIXES:
        if basename.startswith(prefix):
            logger.info('文件「%s」通过前缀匹配识别为: %s', basename, prefix)
            return prefix

    for bank_name in BANK_PREFIXES:
        if _match_bank_in_filename(basename, bank_name):
            logger.info('文件「%s」通过正则匹配识别为: %s', basename, bank_name)
            return bank_name

    content_match = _identify_bank_by_content(filepath)
    if content_match:
        return content_match

    logger.warning('文件「%s」无法识别银行类型', basename)
    return None


# ──────────────────────────────────────────────
# 主体查找
# ──────────────────────────────────────────────

# 主体查找表的推荐文件名（优先精确匹配）
LOOKUP_FILE_NAMES = ['主体查找表.xlsx', '主体查找表.xls']


def find_lookup_file(script_dir):
    """
    在脚本所在目录下查找主体查找表 Excel 文件。

    查找策略（按优先级）：
    1. 优先按文件名精确匹配 "主体查找表.xlsx" 或 "主体查找表.xls"
    2. 若未精确匹配到，回退到查找目录下唯一的 Excel 文件（排除输出总表和临时文件）
    """
    logger = get_logger()

    # ── 策略 1：按文件名精确匹配 ──
    for name in LOOKUP_FILE_NAMES:
        candidate = os.path.join(script_dir, name)
        if os.path.isfile(candidate):
            logger.info('精确匹配到主体查找表: %s', candidate)
            return candidate

    # ── 策略 2：回退到唯一 Excel 文件 ──
    excel_exts = ('.xlsx', '.xls')
    exclude_names = {'银行流水总表.xlsx', '银行流水总表.xls'}
    excel_files = []
    for f in os.listdir(script_dir):
        if f.startswith('~$'):
            continue
        if f in exclude_names:
            continue
        if f.lower().endswith(excel_exts):
            excel_files.append(os.path.join(script_dir, f))

    if len(excel_files) == 1:
        logger.info('找到主体查找表（唯一 Excel 文件）: %s', excel_files[0])
        return excel_files[0]
    elif len(excel_files) == 0:
        logger.warning('程序目录下未找到任何 Excel 文件作为主体查找表')
    else:
        logger.warning(
            '程序目录下存在 %d 个 Excel 文件，无法确定唯一查找表: %s。'
            '建议将查找表文件命名为 "主体查找表.xlsx"',
            len(excel_files),
            [os.path.basename(f) for f in excel_files],
        )
    return None


def _normalize_account_str(value):
    if value is None:
        return ''
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value:
            return ''
        if value == int(value):
            return str(int(value))
        return str(value)
    s = str(value).strip()
    if '.' in s:
        try:
            f = float(s)
            if f == int(f):
                return str(int(f))
        except (ValueError, TypeError, OverflowError):
            pass
    return s


def _account_key(value):
    normalized = _normalize_account_str(value)
    return normalized.lstrip('0') or '0'


def load_lookup_table(lookup_file):
    """
    一次性加载主体查找表到内存字典，避免每条记录重复打开 Excel。

    Args:
        lookup_file: 查找表文件路径

    Returns:
        dict，包含以下键：
            - by_account: {normalized_account_key: [entry, ...]}，每个账号对应条目列表（按优先级降序）
            - all_entries: [entry, ...]，所有条目列表，用于模糊匹配
            - extra_field_names: [str, ...]，扩展字段名列表（已排序）
            - _source_file: 原始文件路径，用于调试
        如果 lookup_file 为空或不存在，返回空结构（by_account 为空 dict 等）
    """
    logger = get_logger()

    empty_result = {
        'by_account': {},
        'all_entries': [],
        'extra_field_names': [],
        '_source_file': lookup_file,
    }

    if not lookup_file or not os.path.exists(lookup_file):
        logger.warning('主体查找表不存在或未指定，返回空查找表结构')
        return empty_result

    tmp_path = None
    try:
        wb, tmp_path = open_workbook_compat(lookup_file)
        ws = wb.active

        header_map = _detect_lookup_header_columns(ws)

        subject_col = _get_lookup_col_index(header_map, ['主体名称', '主体', 'subject', 'Subject'])
        account_col = _get_lookup_col_index(header_map, ['银行账号', '账号', 'account', 'Account'])
        priority_col = _get_lookup_col_index(header_map, ['优先级', 'priority', 'Priority'])

        if subject_col is None:
            subject_col = 1
        if account_col is None:
            account_col = 2

        extra_col_names = sorted([
            name for name in header_map
            if name not in {'主体名称', '主体', 'subject', 'Subject',
                           '银行账号', '账号', 'account', 'Account',
                           '优先级', 'priority', 'Priority'}
        ])

        all_entries = []
        for row_idx in range(2, ws.max_row + 1):
            account_val = ws.cell(row=row_idx, column=account_col).value
            if account_val is None:
                continue

            subject = ws.cell(row=row_idx, column=subject_col).value or ''

            priority = 0
            if priority_col is not None:
                priority_val = ws.cell(row=row_idx, column=priority_col).value
                if priority_val is not None:
                    try:
                        priority = int(priority_val)
                    except (ValueError, TypeError):
                        priority = 0

            extra_fields = {}
            for col_name in extra_col_names:
                col_idx = header_map[col_name]
                val = ws.cell(row=row_idx, column=col_idx).value
                extra_fields[col_name] = str(val).strip() if val is not None else ''

            all_entries.append({
                'subject': str(subject).strip() if subject else '',
                'account_raw': account_val,
                'account_norm': _normalize_account_str(account_val),
                'account_key': _account_key(account_val),
                'priority': priority,
                'extra_fields': extra_fields,
            })

        wb.close()
        cleanup_temp_file(tmp_path)
        tmp_path = None

        by_account = {}
        for entry in all_entries:
            key = entry['account_key']
            if key not in by_account:
                by_account[key] = []
            by_account[key].append(entry)

        for key in by_account:
            by_account[key].sort(key=lambda x: x['priority'], reverse=True)

        logger.info('查找表加载完成：共 %d 条条目，%d 个唯一账号，%d 个扩展字段',
                    len(all_entries), len(by_account), len(extra_col_names))

        return {
            'by_account': by_account,
            'all_entries': all_entries,
            'extra_field_names': extra_col_names,
            '_source_file': lookup_file,
        }

    except Exception as e:
        logger.error('加载主体查找表「%s」时发生错误: %s', lookup_file, e, exc_info=True)
        return empty_result
    finally:
        cleanup_temp_file(tmp_path)


def _resolve_lookup(lookup_source):
    """
    内部工具：将 lookup_source 统一解析为 load_lookup_table 的结果结构。
    支持传入文件路径（str）或已加载的查找表 dict。

    Args:
        lookup_source: 文件路径(str) 或 load_lookup_table 返回的 dict

    Returns:
        dict: load_lookup_table 返回的结构
    """
    if isinstance(lookup_source, dict) and 'by_account' in lookup_source:
        return lookup_source
    return load_lookup_table(lookup_source)


def get_subject(bank_account, lookup_source):
    """
    根据银行账号在查找表中找到对应的主体。
    同一账号对应多个主体时，返回优先级最高的。

    Args:
        bank_account: 银行账号
        lookup_source: 查找表文件路径(str) 或 load_lookup_table() 返回的预加载 dict
    """
    info = get_subject_info(bank_account, lookup_source)
    return info.get('subject', '')


def get_subject_info(bank_account, lookup_source, use_fuzzy=False, fuzzy_threshold=0.6):
    """
    根据银行账号获取主体信息（包含扩展字段、优先级等）。

    Args:
        bank_account: 银行账号
        lookup_source: 查找表文件路径(str) 或 load_lookup_table() 返回的预加载 dict
        use_fuzzy: 是否启用模糊匹配
        fuzzy_threshold: 模糊匹配相似度阈值

    Returns:
        字典，包含 subject、priority、extra_fields、matched、fuzzy_matched、similarity 等
    """
    logger = get_logger()

    _lookup_start = None
    if HAS_PERF_PROFILER:
        import time as _time
        _lookup_start = _time.perf_counter()

    result = {
        'subject': '',
        'account': bank_account,
        'priority': 0,
        'extra_fields': {},
        'matched': False,
        'fuzzy_matched': False,
        'similarity': 0.0,
    }

    if bank_account is None:
        logger.warning('银行账号为空，无法查找主体')
        if _lookup_start is not None:
            import time as _time
            _duration_ms = (_time.perf_counter() - _lookup_start) * 1000
            perf_profiler.get_profiler().record_lookup_hit(
                str(bank_account), _duration_ms, hit=False
            )
        return result

    lookup = _resolve_lookup(lookup_source)
    if not lookup.get('by_account'):
        source_file = lookup.get('_source_file', '')
        if source_file:
            logger.warning('主体查找表「%s」为空或不存在，银行账号「%s」的主体将为空',
                           source_file, bank_account)
        else:
            logger.warning('主体查找表不存在或未指定，银行账号「%s」的主体将为空', bank_account)
        if _lookup_start is not None:
            import time as _time
            _duration_ms = (_time.perf_counter() - _lookup_start) * 1000
            perf_profiler.get_profiler().record_lookup_hit(
                str(bank_account), _duration_ms, hit=False
            )
        return result

    target_key = _account_key(bank_account)

    exact_entries = lookup['by_account'].get(target_key, [])
    if exact_entries:
        best = exact_entries[0]
        result['subject'] = best['subject']
        result['priority'] = best['priority']
        result['extra_fields'] = dict(best['extra_fields'])
        result['matched'] = True
        result['similarity'] = 1.0
        logger.debug('银行账号「%s」匹配到主体: %s（优先级: %d）',
                     bank_account, best['subject'], best['priority'])
        if _lookup_start is not None:
            import time as _time
            _duration_ms = (_time.perf_counter() - _lookup_start) * 1000
            perf_profiler.get_profiler().record_lookup_hit(
                str(bank_account), _duration_ms, hit=True
            )
        return result

    if use_fuzzy:
        target_norm = _normalize_account_str(bank_account)
        fuzzy_matches = []
        for entry in lookup['all_entries']:
            entry_norm = entry['account_norm']
            if not entry_norm:
                continue
            sim = _calculate_string_similarity(target_norm, entry_norm)
            if sim >= fuzzy_threshold:
                fuzzy_matches.append({
                    'subject': entry['subject'],
                    'priority': entry['priority'],
                    'extra_fields': dict(entry['extra_fields']),
                    'similarity': sim,
                })

        if fuzzy_matches:
            fuzzy_matches.sort(key=lambda x: (x['similarity'], x['priority']), reverse=True)
            best = fuzzy_matches[0]
            result['subject'] = best['subject']
            result['priority'] = best['priority']
            result['extra_fields'] = best['extra_fields']
            result['matched'] = True
            result['fuzzy_matched'] = True
            result['similarity'] = best['similarity']
            logger.debug('银行账号「%s」模糊匹配到主体: %s（相似度: %.2f）',
                         bank_account, best['subject'], best['similarity'])
            if _lookup_start is not None:
                import time as _time
                _duration_ms = (_time.perf_counter() - _lookup_start) * 1000
                perf_profiler.get_profiler().record_lookup_hit(
                    str(bank_account), _duration_ms,
                    hit=True, fuzzy=True, similarity=best['similarity']
                )
            return result

    logger.warning('银行账号「%s」在查找表中未找到对应主体', bank_account)
    if _lookup_start is not None:
        import time as _time
        _duration_ms = (_time.perf_counter() - _lookup_start) * 1000
        perf_profiler.get_profiler().record_lookup_hit(
            str(bank_account), _duration_ms, hit=False
        )
    return result


def _detect_lookup_header_columns(ws) -> Dict[str, int]:
    """检测查找表表头列"""
    header_map = {}
    for col_idx in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col_idx).value
        if cell_value is None:
            continue
        col_name = str(cell_value).strip()
        if col_name:
            header_map[col_name] = col_idx
    return header_map


def _get_lookup_col_index(header_map: Dict[str, int], candidates: List[str]) -> Optional[int]:
    """根据候选列名列表获取列索引"""
    for name in candidates:
        if name in header_map:
            return header_map[name]
    return None


def _calculate_string_similarity(s1: str, s2: str) -> float:
    """计算两个字符串的相似度（基于编辑距离）"""
    if s1 == s2:
        return 1.0
    if not s1 or not s2:
        return 0.0

    len1, len2 = len(s1), len(s2)
    if len1 == 0 or len2 == 0:
        return 0.0

    dp = [[0] * (len2 + 1) for _ in range(len1 + 1)]
    for i in range(len1 + 1):
        dp[i][0] = i
    for j in range(len2 + 1):
        dp[0][j] = j

    for i in range(1, len1 + 1):
        for j in range(1, len2 + 1):
            if s1[i - 1] == s2[j - 1]:
                dp[i][j] = dp[i - 1][j - 1]
            else:
                dp[i][j] = min(dp[i - 1][j] + 1,
                               dp[i][j - 1] + 1,
                               dp[i - 1][j - 1] + 1)

    max_len = max(len1, len2)
    if max_len == 0:
        return 0.0
    return 1.0 - dp[len1][len2] / max_len


def get_lookup_extra_fields(lookup_source) -> List[str]:
    """
    获取查找表中的扩展字段名称列表。

    Args:
        lookup_source: 查找表文件路径(str) 或 load_lookup_table() 返回的预加载 dict

    Returns:
        扩展字段名称列表
    """
    if isinstance(lookup_source, dict) and 'extra_field_names' in lookup_source:
        return list(lookup_source['extra_field_names'])

    if not lookup_source or not isinstance(lookup_source, str):
        return []

    if not os.path.exists(lookup_source):
        return []

    lookup = load_lookup_table(lookup_source)
    return list(lookup['extra_field_names'])


# ──────────────────────────────────────────────
# 唯一 ID
# ──────────────────────────────────────────────

def generate_unique_id():
    """生成唯一 ID：当前时间戳 + UUID"""
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S%f')
    uid = uuid.uuid4().hex
    return f"{timestamp}{uid}"


# ──────────────────────────────────────────────
# 数值工具
# ──────────────────────────────────────────────

def is_numeric(value):
    """判断值是否为数字（int / float / 可转换的字符串）"""
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return True
    try:
        float(str(value).strip())
        return True
    except (ValueError, TypeError):
        return False


def to_float(value):
    """安全地将值转为 float"""
    if isinstance(value, (int, float)):
        return float(value)
    try:
        s = str(value).strip()
        s = s.replace(',', '')
        return float(s)
    except (ValueError, TypeError):
        return None


# ──────────────────────────────────────────────
# 银行处理器（基于配置动态生成）
# ──────────────────────────────────────────────

def _build_bank_processors():
    """从配置文件动态构建银行处理器注册表"""
    config = get_bank_config()
    bank_names = config.get_all_bank_names()
    processors = {}
    for bank_name in bank_names:
        processors[bank_name] = _create_bank_processor(bank_name)
    return processors


# 银行处理器注册表（从配置动态生成）
BANK_PROCESSORS = _build_bank_processors()


def reload_bank_processors():
    """重新加载银行配置并重建处理器注册表（热更新）"""
    global BANK_PROCESSORS, BANK_PREFIXES, _bank_config_singleton
    _bank_config_singleton = None
    config = get_bank_config()
    config.load_config()
    BANK_PREFIXES = _get_bank_prefixes()
    BANK_PROCESSORS = _build_bank_processors()
    logger = get_logger()
    logger.info('已重新加载 %d 个银行处理器', len(BANK_PROCESSORS))


# 向后兼容：保留旧函数名作为别名，确保现有代码和测试正常运行
# 直接从 BANK_PROCESSORS 中获取，确保是同一个函数实例
process_beijing_bank = BANK_PROCESSORS.get('北京银行', _create_bank_processor('北京银行'))
process_east_asia_bank = BANK_PROCESSORS.get('东亚银行', _create_bank_processor('东亚银行'))
process_icbc_bank = BANK_PROCESSORS.get('工商银行', _create_bank_processor('工商银行'))
process_ccb_bank = BANK_PROCESSORS.get('建设银行', _create_bank_processor('建设银行'))
process_cmb_bank = BANK_PROCESSORS.get('招商银行', _create_bank_processor('招商银行'))
process_industrial_commercial_bank = process_icbc_bank
process_construction_bank = process_ccb_bank
process_merchants_bank = process_cmb_bank


# ──────────────────────────────────────────────
# 主流程
# ──────────────────────────────────────────────

@dataclass
class ProcessingResult:
    all_rows: List[dict] = field(default_factory=list)
    processed_files: List[str] = field(default_factory=list)
    unprocessed_files: List[str] = field(default_factory=list)
    error_files: List[Tuple[str, str]] = field(default_factory=list)
    output_path: Optional[str] = None
    masked_output_path: Optional[str] = None
    subject_summary_path: Optional[str] = None
    balance_check_path: Optional[str] = None
    duplicate_check_path: Optional[str] = None
    interest_fee_check_path: Optional[str] = None
    holiday_check_path: Optional[str] = None
    accounting_period_path: Optional[str] = None
    perf_report_path: Optional[str] = None
    collab_template_path: Optional[str] = None
    pending_list_path: Optional[str] = None
    lookup_missing: bool = False
    folder_empty: bool = False
    incremental_mode: bool = False
    existing_record_count: int = 0
    new_record_count: int = 0
    duplicate_record_count: int = 0
    db_inserted_count: int = 0
    db_duplicate_count: int = 0
    output_hash: Optional[str] = None
    signature_id: Optional[str] = None
    signature_info: Optional[Dict[str, Any]] = None
    encryption_result: Optional[Any] = None
    encrypted_files: List[str] = field(default_factory=list)
    dry_run: bool = False
    pending_deletion_files: List[str] = field(default_factory=list)
    pending_keep_set: Set[str] = field(default_factory=set)
    pending_all_files: List[str] = field(default_factory=list)
    pending_final_rows: List[dict] = field(default_factory=list)
    pending_existing_records: List[dict] = field(default_factory=list)
    pending_incremental_rows: List[dict] = field(default_factory=list)
    pending_script_dir: Optional[str] = None
    pending_output_dir: Optional[str] = None
    pending_lookup_source: Any = None
    pending_enable_signature: bool = False
    pending_signature_password: Optional[str] = None
    pending_auto_generate_key: bool = True
    pending_enable_encryption: bool = False
    pending_encryption_password: Optional[str] = None
    pending_encryption_mode: str = 'excel_password'
    pending_batch_id: Optional[str] = None
    pending_input_folder: Optional[str] = None
    pending_cp_tag_summary: Dict[str, Any] = field(default_factory=dict)
    pending_holiday_tag_summary: Dict[str, Any] = field(default_factory=dict)
    pending_internal_transfer_summary: Dict[str, Any] = field(default_factory=dict)
    pending_internal_transfer_result: Any = None
    pending_keep_strategy: str = 'keep_unprocessed'
    pending_archive_dir_name: str = '已处理归档'
    changes_committed: bool = False


# ──────────────────────────────────────────────
# 增量合并模块
# ──────────────────────────────────────────────

SUMMARY_TABLE_FILENAME = '银行流水总表.xlsx'
SUMMARY_TABLE_MASKED_FILENAME = '银行流水总表_脱敏版.xlsx'

MASKED_FIELDS = ['银行账号', '对方户名', '对方账号']


def get_summary_table_path(script_dir, output_dir=None):
    """获取历史总表文件路径"""
    base_dir = output_dir or script_dir
    return os.path.join(base_dir, SUMMARY_TABLE_FILENAME)


def get_masked_summary_table_path(script_dir, output_dir=None):
    """获取脱敏版总表文件路径"""
    base_dir = output_dir or script_dir
    return os.path.join(base_dir, SUMMARY_TABLE_MASKED_FILENAME)


def mask_record(record, fields=None):
    """
    对单条记录进行敏感信息脱敏。

    Args:
        record: 记录字典
        fields: 需要脱敏的字段列表，默认为 MASKED_FIELDS

    Returns:
        dict: 脱敏后的记录
    """
    if fields is None:
        fields = MASKED_FIELDS

    masked = dict(record)

    if not HAS_PII_CLASSIFIER:
        return masked

    for field in fields:
        if field in masked and masked[field] is not None:
            value = str(masked[field])
            if field in ('银行账号', '对方账号', 'account'):
                masked[field] = _mask_bank_account(value)
            elif field in ('对方户名', '主体', '主体名称', 'counterparty', 'subject'):
                masked[field] = _mask_subject_name(value)
            else:
                masked[field] = mask_value(field, value)

    return masked


def mask_records(records, fields=None):
    """
    对记录列表进行批量脱敏。

    Args:
        records: 记录列表
        fields: 需要脱敏的字段列表

    Returns:
        list: 脱敏后的记录列表
    """
    if not records:
        return []
    return [mask_record(r, fields) for r in records]


def export_masked_summary(records, script_dir, output_dir=None, lookup_source=None, columns=None):
    """
    导出脱敏版总表。

    Args:
        records: 记录列表
        script_dir: 脚本目录
        output_dir: 输出目录
        lookup_source: 查找表来源
        columns: 列名列表，如为 None 则自动获取

    Returns:
        str: 脱敏版总表文件路径，无记录时返回 None
    """
    logger = get_logger()

    if not records:
        logger.warning('无记录可导出脱敏版')
        return None

    if columns is None:
        columns = get_summary_columns(records, lookup_source)

    masked_records = mask_records(records)
    df = pd.DataFrame(masked_records, columns=columns)

    output_path = get_masked_summary_table_path(script_dir, output_dir)

    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    df.to_excel(output_path, index=False, engine='openpyxl')

    logger.info('脱敏版总表输出完成: %s（共 %d 条记录）', output_path, len(records))
    return output_path


# 异常标记列名
ANOMALY_FLAG_COLUMN = '异常标记'
ANOMALY_DETAIL_COLUMN = '异常详情'

STANDARD_COLUMNS = [
    '唯一id', '银行', '银行账号', '主体', '交易日期',
    '付款', '收款', '摘要', '对方户名', '余额', '交易流水号',
    ANOMALY_FLAG_COLUMN, ANOMALY_DETAIL_COLUMN,
]


def get_summary_columns(records=None, lookup_source=None):
    """
    获取总表列名列表，包含标准列和扩展字段列。

    Args:
        records: 记录列表，用于从中提取扩展字段（可选）
        lookup_source: 查找表文件路径(str) 或 load_lookup_table() 返回的预加载 dict（可选）

    Returns:
        列名列表
    """
    columns = list(STANDARD_COLUMNS)

    extra_fields = set()

    if lookup_source:
        lookup_extra = get_lookup_extra_fields(lookup_source)
        extra_fields.update(lookup_extra)

    if records:
        for record in records:
            for key in record.keys():
                if key not in columns:
                    extra_fields.add(key)

    for field in sorted(extra_fields):
        if field not in columns:
            columns.append(field)

    return columns


def load_existing_keys(summary_path):
    """
    读取历史总表，提取所有记录的匹配键集合。
    用于快速检测新记录是否已存在。

    Args:
        summary_path: 历史总表文件路径

    Returns:
        tuple: (existing_keys_set, existing_records_list)
            - existing_keys_set: 匹配键集合，用于 O(1) 复杂度的存在性检测
            - existing_records_list: 历史记录列表，用于合并输出
    """
    logger = get_logger()

    if not summary_path or not os.path.exists(summary_path):
        logger.info('未找到历史总表，将以全量模式运行')
        return set(), []

    try:
        df = pd.read_excel(summary_path, engine='openpyxl')
        if df.empty:
            logger.info('历史总表为空，将以全量模式运行')
            return set(), []

        required_cols = ['银行账号', '交易流水号', '交易日期', '付款', '收款']
        for col in required_cols:
            if col not in df.columns:
                logger.warning('历史总表缺少必要列「%s」，无法进行增量检测，将以全量模式运行', col)
                return set(), []

        existing_keys = set()
        existing_records = []

        for _, row in df.iterrows():
            row_dict = row.to_dict()
            key = _make_match_key(row_dict)
            existing_keys.add(key)
            existing_records.append(row_dict)

        logger.info('已加载历史总表，共 %d 条记录，%d 个唯一匹配键',
                    len(existing_records), len(existing_keys))
        return existing_keys, existing_records

    except Exception as e:
        logger.error('读取历史总表失败: %s，将以全量模式运行', e, exc_info=True)
        return set(), []


def filter_incremental_records(new_rows, existing_keys):
    """
    过滤新提取的记录，只返回增量（未在历史总表中出现过的）记录。

    Args:
        new_rows: 新提取的记录列表
        existing_keys: 历史总表的匹配键集合

    Returns:
        tuple: (incremental_rows, duplicate_count)
            - incremental_rows: 增量记录列表
            - duplicate_count: 检测到的重复记录数量
    """
    logger = get_logger()

    if not existing_keys:
        logger.info('无历史记录，所有 %d 条记录均为新增', len(new_rows))
        return new_rows, 0

    incremental_rows = []
    duplicate_count = 0

    for row in new_rows:
        key = _make_match_key(row)
        if key in existing_keys:
            duplicate_count += 1
            logger.debug('检测到重复记录，匹配键: %s', key)
        else:
            incremental_rows.append(row)
            existing_keys.add(key)

    logger.info('增量过滤完成: 新记录 %d 条，重复 %d 条，实际新增 %d 条',
                len(new_rows), duplicate_count, len(incremental_rows))
    return incremental_rows, duplicate_count


def merge_and_export_summary(existing_records, incremental_rows, script_dir, output_dir=None, lookup_source=None):
    """
    合并历史记录与增量记录，并输出到总表。

    Args:
        existing_records: 历史记录列表
        incremental_rows: 新增记录列表
        script_dir: 脚本目录
        output_dir: 输出目录，默认为script_dir
        lookup_source: 查找表文件路径(str) 或 load_lookup_table() 返回的预加载 dict

    Returns:
        str: 输出文件路径
    """
    logger = get_logger()

    merged_records = existing_records + incremental_rows

    if not merged_records:
        logger.warning('无任何记录可输出')
        return None

    columns = get_summary_columns(merged_records, lookup_source)
    df = pd.DataFrame(merged_records, columns=columns)
    output_path = get_summary_table_path(script_dir, output_dir)

    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    df.to_excel(output_path, index=False, engine='openpyxl')

    logger.info('总表输出完成: %s（历史 %d 条 + 新增 %d 条 = 共 %d 条）',
                output_path, len(existing_records), len(incremental_rows), len(merged_records))
    return output_path


def cli_ask_incremental_mode():
    """命令行模式下询问用户是否启用增量合并"""
    print('\n请选择运行模式：')
    print('  1) 增量合并（推荐）：检测历史记录，仅追加新增数据')
    print('  2) 全量覆盖：重新生成总表，覆盖历史数据')
    choice = input('请输入选项（直接回车默认为 1 增量模式）: ').strip()
    return choice != '2'


def cli_ask_keep_strategy():
    """命令行模式下询问用户文件保留策略"""
    print('\n请选择文件保留策略：')
    strategies = [
        ('keep_unprocessed', '1) 仅保留未处理文件（默认）：删除已处理成功的文件，保留错误与未识别文件'),
        ('keep_all', '2) 保留所有文件：不删除或移动任何文件'),
        ('delete_all', '3) 删除所有文件：无论处理状态，删除所有源文件'),
        ('move_to_archive', '4) 归档已处理文件：将成功处理的文件移动到「已处理归档」子目录'),
    ]
    for _, desc in strategies:
        print(f'  {desc}')
    choice = input('请输入选项（直接回车默认为 1）: ').strip()
    mapping = {'1': 'keep_unprocessed', '2': 'keep_all', '3': 'delete_all', '4': 'move_to_archive'}
    if choice in mapping:
        return mapping[choice]
    if choice in [k for k, _ in strategies]:
        return choice
    return 'keep_unprocessed'


def gui_ask_incremental_mode():
    """GUI 模式下询问用户是否启用增量合并"""
    root = tk.Tk()
    root.withdraw()
    choice = messagebox.askyesnocancel(
        '选择运行模式',
        '是 = 增量合并（推荐）：检测历史记录，仅追加新增数据\n\n'
        '否 = 全量覆盖：重新生成总表，覆盖历史数据\n\n'
        '取消 = 返回主菜单',
    )
    root.destroy()
    if choice is None:
        return None
    return choice


if HAS_TKINTER:
    ask_incremental_mode = gui_ask_incremental_mode
else:
    ask_incremental_mode = cli_ask_incremental_mode


def run_pipeline(folder, script_dir, incremental=True, batch_id=None,
                 enable_signature=False, signature_password=None, auto_generate_key=True,
                 enable_encryption=False, encryption_password=None, encryption_mode='excel_password',
                 dry_run=False, keep_strategy='keep_unprocessed', archive_dir_name='已处理归档'):
    logger = get_logger()
    if dry_run:
        logger.info('===== 试运行模式已启用（不执行删除与写盘操作）=====')

    _profiler = None
    if HAS_PERF_PROFILER:
        perf_profiler.reset_profiler()
        _profiler = perf_profiler.get_profiler()
        _profiler.start()

    _phase_lookup_start = None
    if _profiler is not None:
        _phase_lookup_start = __import__('time').perf_counter()

    lookup_file = find_lookup_file(script_dir)
    lookup_missing = lookup_file is None
    if lookup_missing:
        logger.warning('未找到主体查找表，"主体"列将为空')
        lookup_data = load_lookup_table(None)
    else:
        logger.info('正在预加载主体查找表...')
        lookup_data = load_lookup_table(lookup_file)
        logger.info('主体查找表预加载完成')

    if _profiler is not None and _phase_lookup_start is not None:
        _lookup_dur = (__import__('time').perf_counter() - _phase_lookup_start) * 1000
        _profiler.record_phase('lookup_preload', _lookup_dur, '查找表预加载')

    existing_keys = set()
    existing_records = []
    actual_incremental = False
    duplicate_count = 0
    new_record_count = 0

    if incremental:
        summary_path = get_summary_table_path(script_dir)
        existing_keys, existing_records = load_existing_keys(summary_path)
        actual_incremental = len(existing_records) > 0
        if actual_incremental:
            logger.info('===== 增量合并模式已启用 =====')
        else:
            logger.info('无历史数据，将以全量模式运行')

    _phase_copy_start = None
    if _profiler is not None:
        _phase_copy_start = __import__('time').perf_counter()

    folder_name = os.path.basename(folder.rstrip('/\\'))
    parent_dir = os.path.dirname(folder.rstrip('/\\'))
    new_folder = os.path.join(parent_dir, f"{folder_name}＋检验版")

    if os.path.exists(new_folder):
        logger.info('＋检验版文件夹已存在，先删除: %s', new_folder)
        shutil.rmtree(new_folder)
    shutil.copytree(folder, new_folder)
    logger.info('已复制文件夹为＋检验版: %s', new_folder)

    if _profiler is not None and _phase_copy_start is not None:
        _copy_dur = (__import__('time').perf_counter() - _phase_copy_start) * 1000
        _profiler.record_phase('folder_copy', _copy_dur, '复制文件夹')

    excel_files = scan_excel_files(new_folder)
    try:
        from pdf_bank_parser import scan_pdf_files, is_pdf_file, process_pdf_file
        pdf_files = scan_pdf_files(new_folder)
    except ImportError:
        pdf_files = []
    all_files = excel_files + pdf_files
    if not all_files:
        logger.warning('检验版文件夹中未发现任何 Excel 或 PDF 文件')
        if _profiler is not None:
            _profiler.stop()
        return ProcessingResult(
            lookup_missing=lookup_missing,
            folder_empty=True,
            incremental_mode=actual_incremental,
            existing_record_count=len(existing_records),
            dry_run=dry_run,
            pending_script_dir=script_dir if dry_run else None,
            pending_input_folder=folder if dry_run else None,
        )

    _phase_process_start = None
    if _profiler is not None:
        _phase_process_start = __import__('time').perf_counter()

    all_rows = []
    processed_files = []
    unprocessed_files = []
    error_files = []

    for filepath in all_files:
        file_is_pdf = is_pdf_file(filepath) if 'is_pdf_file' in dir() else filepath.lower().endswith('.pdf')
        if file_is_pdf:
            try:
                rows = process_pdf_file(filepath, lookup_data)
                if rows:
                    all_rows.extend(rows)
                    processed_files.append(filepath)
                    logger.info('成功处理 PDF 文件: %s（%d 条记录）', filepath, len(rows))
                else:
                    unprocessed_files.append(filepath)
                    logger.warning('PDF 文件未解析出有效记录: %s', filepath)
            except Exception as e:
                error_files.append((filepath, str(e)))
                logger.error('处理 PDF 文件「%s」时发生错误: %s', filepath, e, exc_info=True)
            continue

        bank = identify_bank(filepath)
        if bank and bank in BANK_PROCESSORS:
            try:
                processor = BANK_PROCESSORS[bank]
                rows = processor(filepath, lookup_data)
                all_rows.extend(rows)
                processed_files.append(filepath)
                logger.info('成功处理文件: %s（%d 条记录）', filepath, len(rows))
            except Exception as e:
                error_files.append((filepath, str(e)))
                logger.error('处理文件「%s」时发生错误: %s', filepath, e, exc_info=True)
        else:
            unprocessed_files.append(filepath)

    if _profiler is not None and _phase_process_start is not None:
        _process_dur = (__import__('time').perf_counter() - _phase_process_start) * 1000
        _profiler.record_phase('file_processing', _process_dur,
                               f'处理 {len(all_files)} 个文件')

    error_file_paths = {f for f, _ in error_files}
    keep_set = set(unprocessed_files) | error_file_paths
    pending_deletion_files = [f for f in all_files if f not in keep_set]

    if dry_run:
        logger.info('[试运行] 跳过删除文件操作，待删除文件 %d 个', len(pending_deletion_files))
    else:
        delete_processed_files(
            all_files, processed_files, error_files, unprocessed_files,
            strategy=keep_strategy, archive_dir_name=archive_dir_name
        )

    output_path = None
    final_rows = []
    incremental_rows = []
    _cp_tag_summary = {}
    _holiday_tag_summary = {}
    _it_summary = {}
    _it_result = None

    if all_rows:
        if actual_incremental:
            incremental_rows, duplicate_count = filter_incremental_records(all_rows, existing_keys)
            new_record_count = len(incremental_rows)
            if dry_run:
                logger.info('[试运行] 跳过总表写盘，模式=增量合并，历史 %d 条 + 新增 %d 条',
                            len(existing_records), len(incremental_rows))
                final_rows = existing_records + incremental_rows
            else:
                output_path = merge_and_export_summary(
                    existing_records, incremental_rows, script_dir, lookup_source=lookup_data
                )
                final_rows = existing_records + incremental_rows
        else:
            if dry_run:
                logger.info('[试运行] 跳过总表写盘，模式=全量覆盖，共 %d 条记录', len(all_rows))
                final_rows = all_rows
            else:
                columns = get_summary_columns(all_rows, lookup_data)
                df = pd.DataFrame(all_rows, columns=columns)
                output_path = get_summary_table_path(script_dir)
                df.to_excel(output_path, index=False, engine='openpyxl')
                logger.info('总表输出完成: %s（共 %d 条记录）', output_path, len(all_rows))
                final_rows = all_rows
            new_record_count = len(all_rows)
    else:
        logger.warning('未提取到任何银行流水记录')
        if existing_records:
            if dry_run:
                logger.info('[试运行] 跳过总表写盘，仅使用历史记录 %d 条', len(existing_records))
                final_rows = existing_records
            else:
                output_path = merge_and_export_summary(
                    existing_records, [], script_dir, lookup_source=lookup_data
                )
                final_rows = existing_records

    if final_rows:
        final_rows, _anomaly_summary = apply_amount_anomaly_detection(final_rows)
        if _anomaly_summary.get('anomaly_count', 0) > 0:
            logger.info('金额异常检测: 总记录 %d, 异常 %d (%.2f%%)',
                        _anomaly_summary.get('total_records', 0),
                        _anomaly_summary.get('anomaly_count', 0),
                        _anomaly_summary.get('anomaly_rate', 0) * 100)
            if output_path and not dry_run:
                base_columns = get_summary_columns(final_rows, lookup_data)
                _anomaly_columns = base_columns
                pd.DataFrame(final_rows, columns=_anomaly_columns).to_excel(
                    output_path, index=False, engine='openpyxl')
                logger.info('已将金额异常检测结果回写到总表: %s', output_path)
            elif dry_run:
                logger.info('[试运行] 跳过金额异常检测结果回写总表')

    if final_rows:
        final_rows, _cp_tag_summary = apply_counterparty_rules(final_rows, script_dir)
        if _cp_tag_summary.get('tagged_count', 0) > 0:
            logger.info('对方户名黑白名单打标: 总记录 %d, 命中 %d (黑名单 %d, 白名单 %d)',
                        _cp_tag_summary.get('total_records', 0),
                        _cp_tag_summary.get('tagged_count', 0),
                        _cp_tag_summary.get('blacklist_hits', 0),
                        _cp_tag_summary.get('whitelist_hits', 0))
            if output_path and not dry_run:
                base_columns = get_summary_columns(final_rows, lookup_data)
                cp_extra_cols = ['黑白名单标签', '命中规则名称', '命中关键词']
                _cp_columns = base_columns + [
                    col for col in cp_extra_cols if col not in base_columns
                ]
                pd.DataFrame(final_rows, columns=_cp_columns).to_excel(
                    output_path, index=False, engine='openpyxl')
                logger.info('已将黑白名单打标结果回写到总表: %s', output_path)
            elif dry_run:
                logger.info('[试运行] 跳过黑白名单打标结果回写总表')

    if final_rows:
        final_rows, _holiday_tag_summary = apply_holiday_tags(final_rows)
        if _holiday_tag_summary.get('tagged_count', 0) > 0:
            logger.info('非工作日交易打标: 总记录 %d, 非工作日 %d (周末 %d, 节假日 %d)',
                        _holiday_tag_summary.get('total_records', 0),
                        _holiday_tag_summary.get('tagged_count', 0),
                        _holiday_tag_summary.get('weekend_count', 0),
                        _holiday_tag_summary.get('holiday_count', 0))
            if output_path and not dry_run:
                base_columns = get_summary_columns(final_rows, lookup_data)
                holiday_extra_cols = ['非工作日标签', '节假日名称']
                _holiday_columns = base_columns + [
                    col for col in holiday_extra_cols if col not in base_columns
                ]
                pd.DataFrame(final_rows, columns=_holiday_columns).to_excel(
                    output_path, index=False, engine='openpyxl')
                logger.info('已将非工作日打标结果回写到总表: %s', output_path)
            elif dry_run:
                logger.info('[试运行] 跳过非工作日打标结果回写总表')

    masked_output_path = None
    if final_rows and not dry_run:
        try:
            output_dir = os.path.dirname(output_path) or script_dir
            _masked_columns = get_summary_columns(final_rows, lookup_data)
            if _cp_tag_summary and _cp_tag_summary.get('tagged_count', 0) > 0:
                cp_extra_cols = ['黑白名单标签', '命中规则名称', '命中关键词']
                _masked_columns = _masked_columns + [
                    col for col in cp_extra_cols if col not in _masked_columns
                ]
            else:
                cp_extra_cols = ['黑白名单标签', '命中规则名称', '命中关键词']
                _masked_columns = [col for col in _masked_columns if col not in cp_extra_cols]
            masked_output_path = export_masked_summary(
                final_rows, script_dir, output_dir=output_dir,
                lookup_source=lookup_data, columns=_masked_columns
            )
        except Exception as e:
            logger.error('生成脱敏版总表失败: %s', e, exc_info=True)
            masked_output_path = None
    elif final_rows and dry_run:
        logger.info('[试运行] 跳过生成脱敏版总表')

    db_inserted = 0
    db_duplicates = 0
    if HAS_DATABASE and final_rows and not dry_run:
        try:
            if batch_id is None:
                batch_id = f"BATCH{datetime.now().strftime('%Y%m%d%H%M%S')}"
            db_inserted, db_duplicates = db_module.persist_transactions(
                final_rows,
                batch_id=batch_id,
                deduplicate=True,
                script_dir=script_dir,
            )
            logger.info(
                '数据库持久化完成: 批次 %s, 插入 %d 条, 去重跳过 %d 条',
                batch_id, db_inserted, db_duplicates,
            )
        except Exception as e:
            logger.error('数据库持久化失败: %s', e, exc_info=True)
    elif HAS_DATABASE and final_rows and dry_run:
        logger.info('[试运行] 跳过数据库持久化操作')

    internal_transfer_path = None
    if final_rows:
        try:
            final_rows, _it_summary, _it_result = identify_and_tag_internal_transfers(
                final_rows,
            )
            if _it_summary.get('match_pairs', 0) > 0:
                logger.info(
                    '跨账号内部划转识别: 总记录 %d, 识别 %d 对 (划出 %d + 划入 %d), '
                    '涉及 %d 主体 %d 银行, 划转总金额 %.2f 元',
                    _it_summary.get('total_records', 0),
                    _it_summary.get('match_pairs', 0),
                    _it_summary.get('marked_out_count', 0),
                    _it_summary.get('marked_in_count', 0),
                    len(_it_summary.get('involved_subjects', [])),
                    len(_it_summary.get('involved_banks', [])),
                    _it_summary.get('total_transfer_amount', 0.0),
                )
                if output_path and not dry_run:
                    base_columns = get_summary_columns(final_rows, lookup_data)
                    it_extra_cols = list(INTERNAL_TRANSFER_EXTRA_COLUMNS)
                    _it_columns = base_columns + [
                        col for col in it_extra_cols if col not in base_columns
                    ]
                    pd.DataFrame(final_rows, columns=_it_columns).to_excel(
                        output_path, index=False, engine='openpyxl')
                    logger.info('已将内部划转标记回写到总表: %s', output_path)
                elif dry_run:
                    logger.info('[试运行] 跳过内部划转标记回写总表')

                _it_out_dir = script_dir
                if output_path:
                    _it_out_dir = os.path.dirname(output_path) or _it_out_dir
                if output_dir:
                    _it_out_dir = output_dir or _it_out_dir
                _it_src_info = {
                    '数据来源': '主流程自动生成',
                    '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                    '记录数': len(final_rows),
                    '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                    '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                }
                if _it_result.match_pairs > 0:
                    _it_ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                    _it_out_path = os.path.join(_it_out_dir, f'内部划转识别报告_{_it_ts}.xlsx')
                    internal_transfer_path = export_internal_transfer_report(
                        _it_result, _it_out_path, _it_src_info,
                    )
                    if internal_transfer_path:
                        logger.info('内部划转识别报告已自动生成: %s', internal_transfer_path)
        except Exception as e:
            logger.error('内部划转识别处理失败: %s', e, exc_info=True)
            internal_transfer_path = None

    subject_summary_path = None
    balance_check_path = None
    if final_rows:
        try:
            output_dir = script_dir
            if output_path:
                output_dir = os.path.dirname(output_path) or script_dir
            source_info = {
                '数据来源': '主流程自动生成',
                '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                '记录数': len(final_rows),
                '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            subject_summary_path = generate_subject_summary_from_records(
                final_rows, output_dir, source_info
            )
            if subject_summary_path:
                logger.info('主体维度汇总分析已自动生成: %s', subject_summary_path)
        except Exception as e:
            logger.error('自动生成主体汇总分析失败: %s', e, exc_info=True)
            subject_summary_path = None

        try:
            output_dir = script_dir
            if output_path:
                output_dir = os.path.dirname(output_path) or script_dir
            source_info = {
                '数据来源': '主流程自动生成',
                '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                '记录数': len(final_rows),
                '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            balance_check_path = generate_balance_check_from_records(
                final_rows, output_dir, source_info
            )
            if balance_check_path:
                logger.info('余额连续性校验报告已自动生成: %s', balance_check_path)
        except Exception as e:
            logger.error('自动生成余额连续性校验报告失败: %s', e, exc_info=True)
            balance_check_path = None

    duplicate_check_path = None
    if final_rows:
        try:
            output_dir = script_dir
            if output_path:
                output_dir = os.path.dirname(output_path) or script_dir
            source_info = {
                '数据来源': '主流程自动生成',
                '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                '记录数': len(final_rows),
                '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            duplicate_check_path = generate_duplicate_check_from_records(
                final_rows, output_dir, source_info
            )
            if duplicate_check_path:
                logger.info('重复交易检测报告已自动生成: %s', duplicate_check_path)
        except Exception as e:
            logger.error('自动生成重复交易检测报告失败: %s', e, exc_info=True)
            duplicate_check_path = None

    interest_fee_check_path = None
    if final_rows:
        try:
            output_dir = script_dir
            if output_path:
                output_dir = os.path.dirname(output_path) or script_dir
            source_info = {
                '数据来源': '主流程自动生成',
                '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                '记录数': len(final_rows),
                '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            interest_fee_check_path = generate_interest_fee_check_from_records(
                final_rows, output_dir, source_info
            )
            if interest_fee_check_path:
                logger.info(t('success.interest_fee_check_generated', path=interest_fee_check_path))
        except Exception as e:
            logger.error('自动生成利息手续费核对报告失败: %s', e, exc_info=True)
            interest_fee_check_path = None

    holiday_check_path = None
    if final_rows:
        try:
            output_dir = script_dir
            if output_path:
                output_dir = os.path.dirname(output_path) or script_dir
            source_info = {
                '数据来源': '主流程自动生成',
                '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                '记录数': len(final_rows),
                '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            holiday_check_path = generate_holiday_check_from_records(
                final_rows, output_dir, source_info
            )
            if holiday_check_path:
                logger.info('非工作日交易标记报告已自动生成: %s', holiday_check_path)
        except Exception as e:
            logger.error('自动生成非工作日交易标记报告失败: %s', e, exc_info=True)
            holiday_check_path = None

    accounting_period_path = None
    if final_rows:
        try:
            output_dir = script_dir
            if output_path:
                output_dir = os.path.dirname(output_path) or script_dir
            source_info = {
                '数据来源': '主流程自动生成',
                '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                '记录数': len(final_rows),
                '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            accounting_period_path = generate_accounting_period_report(
                final_rows, output_dir, source_info
            )
            if accounting_period_path:
                logger.info('会计期间总表已自动生成: %s', accounting_period_path)
        except Exception as e:
            logger.error('自动生成会计期间总表失败: %s', e, exc_info=True)
            accounting_period_path = None

    perf_report_path = None
    if _profiler is not None:
        _profiler.stop()
        try:
            output_dir = script_dir
            if output_path:
                output_dir = os.path.dirname(output_path) or script_dir
            perf_report_path = _profiler.save_report(output_dir)
            logger.info('性能剖析报告已生成: %s', perf_report_path)
        except Exception as e:
            logger.error('生成性能剖析报告失败: %s', e, exc_info=True)

    collab_template_path = None
    if output_path and final_rows and not dry_run:
        try:
            collab_output_dir = os.path.dirname(output_path) or script_dir
            collab_template_path = generate_collab_template(
                output_path,
                output_dir=collab_output_dir,
                lookup_source=lookup_data,
            )
            if collab_template_path:
                logger.info('财务协同编辑模板已自动生成: %s', collab_template_path)
        except Exception as e:
            logger.error('自动生成协同编辑模板失败: %s', e, exc_info=True)
            collab_template_path = None
    elif dry_run and final_rows:
        logger.info('[试运行] 跳过生成财务协同编辑模板')

    output_hash = None
    signature_id = None
    signature_info = None

    if output_path and enable_signature and HAS_CRYPTOGRAPHY and not dry_run:
        try:
            output_hash = compute_file_hash(output_path)

            if auto_generate_key and not has_signing_key(script_dir):
                ensure_signing_key(script_dir, auto_generate=True, password=signature_password)

            signature_info = sign_output_file(
                output_path=output_path,
                script_dir=script_dir,
                input_directory=folder,
                password=signature_password,
                extra_data={
                    'record_count': len(final_rows),
                    'incremental_mode': actual_incremental,
                    'processed_files': len(processed_files),
                }
            )

            if signature_info:
                signature_id = save_signature_record(
                    signature_info,
                    script_dir=script_dir
                )
                signature_info['signature_id'] = signature_id
                logger.info('总表数字签名完成: %s', signature_id)
            else:
                logger.warning('数字签名未生成（可能缺少签名密钥）')
        except Exception as e:
            logger.error('数字签名失败: %s', e, exc_info=True)
            signature_info = None
    elif enable_signature and dry_run:
        logger.info('[试运行] 跳过数字签名操作')

    encryption_result = None
    encrypted_files = []

    if enable_encryption and encryption_password and HAS_FILE_ENCRYPTION and not dry_run:
        try:
            output_dir = script_dir
            if output_path:
                output_dir = os.path.dirname(output_path) or script_dir

            files_to_encrypt = []
            for fp in [output_path, subject_summary_path, balance_check_path,
                       duplicate_check_path, interest_fee_check_path, holiday_check_path, accounting_period_path]:
                if fp and os.path.isfile(fp):
                    files_to_encrypt.append(fp)

            if files_to_encrypt:
                encryption_result = _encrypt_output_files(
                    files_to_encrypt,
                    password=encryption_password,
                    mode=encryption_mode,
                    output_dir=output_dir,
                )
                encrypted_files = [
                    r.encrypted_path for r in encryption_result.results
                    if r.success and r.encrypted_path
                ]
                _save_encryption_record(encryption_result, script_dir=script_dir)
                logger.info('输出文件加密完成: 模式=%s, 成功=%d, 失败=%d',
                            encryption_mode,
                            encryption_result.success_count,
                            encryption_result.failure_count)
        except Exception as e:
            logger.error('输出文件加密失败: %s', e, exc_info=True)
            encryption_result = None
    elif enable_encryption and not HAS_FILE_ENCRYPTION:
        logger.warning('file_encryption 模块不可用，无法加密输出文件')
    elif enable_encryption and not encryption_password:
        logger.warning('已启用加密但未提供加密密码，跳过加密')
    elif enable_encryption and dry_run:
        logger.info('[试运行] 跳过输出文件加密操作')

    pending_list_path = None
    if (unprocessed_files or error_files) and not dry_run:
        pending_list_path = generate_pending_list_xlsx(
            unprocessed_files, error_files, script_dir
        )
    elif (unprocessed_files or error_files) and dry_run:
        logger.info('[试运行] 跳过待处理清单生成')

    return ProcessingResult(
        all_rows=final_rows,
        processed_files=processed_files,
        unprocessed_files=unprocessed_files,
        error_files=error_files,
        output_path=output_path,
        masked_output_path=masked_output_path,
        subject_summary_path=subject_summary_path,
        balance_check_path=balance_check_path,
        duplicate_check_path=duplicate_check_path,
        interest_fee_check_path=interest_fee_check_path,
        holiday_check_path=holiday_check_path,
        accounting_period_path=accounting_period_path,
        perf_report_path=perf_report_path,
        collab_template_path=collab_template_path,
        pending_list_path=pending_list_path,
        lookup_missing=lookup_missing,
        incremental_mode=actual_incremental,
        existing_record_count=len(existing_records),
        new_record_count=new_record_count,
        duplicate_record_count=duplicate_count,
        db_inserted_count=db_inserted,
        db_duplicate_count=db_duplicates,
        output_hash=output_hash,
        signature_id=signature_id,
        signature_info=signature_info,
        encryption_result=encryption_result,
        encrypted_files=encrypted_files,
        dry_run=dry_run,
        pending_deletion_files=pending_deletion_files,
        pending_keep_set=keep_set,
        pending_all_files=list(all_files),
        pending_final_rows=final_rows,
        pending_existing_records=existing_records,
        pending_incremental_rows=incremental_rows,
        pending_script_dir=script_dir,
        pending_lookup_source=lookup_data,
        pending_enable_signature=enable_signature,
        pending_signature_password=signature_password,
        pending_auto_generate_key=auto_generate_key,
        pending_enable_encryption=enable_encryption,
        pending_encryption_password=encryption_password,
        pending_encryption_mode=encryption_mode,
        pending_batch_id=batch_id,
        pending_input_folder=folder,
        pending_cp_tag_summary=_cp_tag_summary,
        pending_holiday_tag_summary=_holiday_tag_summary,
        pending_internal_transfer_summary=_it_summary,
        pending_internal_transfer_result=_it_result,
        pending_keep_strategy=keep_strategy,
        pending_archive_dir_name=archive_dir_name,
    )


def format_result_message(result):
    if result.folder_empty:
        return '文件夹中未发现任何 Excel 文件。'

    dry_run_banner = ''
    if result.dry_run:
        dry_run_banner = '【试运行模式】仅生成报告，未执行删除与写盘操作\n\n'

    if result.all_rows:
        if result.incremental_mode:
            msg = (
                f'{dry_run_banner}增量合并处理完成！\n\n'
                f'运行模式：增量合并\n'
                f'已处理文件数：{len(result.processed_files)}\n'
                f'历史总记录数：{result.existing_record_count}\n'
                f'本次新提取记录数：{result.new_record_count + result.duplicate_record_count}\n'
                f'├─ 重复记录（已跳过）：{result.duplicate_record_count}\n'
                f'└─ 新增记录（{result.dry_run and "待追加" or "已追加"}）：{result.new_record_count}\n'
                f'总表当前总记录数：{len(result.all_rows)}\n'
                f'总表路径：{result.output_path or "(试运行未写盘)"}'
            )
        else:
            msg = (
                f'{dry_run_banner}处理完成！\n\n'
                f'运行模式：全量覆盖\n'
                f'已处理文件数：{len(result.processed_files)}\n'
                f'提取记录数：{len(result.all_rows)}\n'
                f'总表路径：{result.output_path or "(试运行未写盘)"}'
            )

        if result.dry_run and result.pending_deletion_files:
            msg += f'\n\n待删除文件（{len(result.pending_deletion_files)} 个）：'
            for f in result.pending_deletion_files[:10]:
                msg += f'\n  └─ {os.path.basename(f)}'
            if len(result.pending_deletion_files) > 10:
                msg += f'\n  └─ ... 等共 {len(result.pending_deletion_files)} 个'

        if result.masked_output_path:
            msg += f'\n脱敏版总表：{result.masked_output_path}'
            msg += '（银行账号、对方户名已脱敏，可用于对外分享）'
        elif result.dry_run:
            msg += '\n脱敏版总表：(试运行未生成)'

        if HAS_DATABASE and (result.db_inserted_count > 0 or result.db_duplicate_count > 0):
            msg += (
                f'\n\n数据库持久化：\n'
                f'├─ 新增入库：{result.db_inserted_count} 条\n'
                f'└─ 重复跳过：{result.db_duplicate_count} 条'
            )
        elif HAS_DATABASE and result.dry_run:
            msg += '\n\n数据库持久化：(试运行未执行)'

        if result.subject_summary_path:
            msg += f'\n\n主体汇总分析：{result.subject_summary_path}'

        if result.balance_check_path:
            msg += f'\n\n余额连续性校验：{result.balance_check_path}'

        if result.duplicate_check_path:
            msg += f'\n\n重复交易检测：{result.duplicate_check_path}'

        if result.interest_fee_check_path:
            msg += f'\n\n利息手续费核对：{result.interest_fee_check_path}'

        if result.accounting_period_path:
            msg += f'\n\n会计期间总表：{result.accounting_period_path}'

        if result.perf_report_path:
            msg += f'\n\n性能剖析报告：{result.perf_report_path}'

        if result.collab_template_path:
            msg += f'\n\n财务协同编辑模板：{result.collab_template_path}'
            msg += '\n（黄色列为可编辑区，灰色列为只读区，编辑后请运行"回写合并"功能）'
        elif result.dry_run:
            msg += '\n财务协同编辑模板：(试运行未生成)'

        if result.encrypted_files:
            mode_label = 'Excel密码保护' if result.encryption_result and result.encryption_result.mode == 'excel_password' else 'AES-256加密'
            msg += f'\n\n文件加密（{mode_label}）：{len(result.encrypted_files)} 个文件'
            for ef in result.encrypted_files:
                msg += f'\n  └─ {os.path.basename(ef)}'
        elif result.dry_run and result.pending_enable_encryption:
            msg += '\n文件加密：(试运行未执行)'

        if result.dry_run and not result.changes_committed:
            msg += '\n\n⚠️  试运行提示：请检查以上报告是否符合预期。'
            msg += '\n确认无误后，请调用 commit_pipeline_changes(result) 执行正式写盘。'
    else:
        if result.incremental_mode and result.existing_record_count > 0:
            msg = (
                f'{dry_run_banner}本次未提取到任何新增银行流水记录。\n\n'
                f'运行模式：增量合并\n'
                f'历史记录保留：{result.existing_record_count} 条\n'
                f'总表路径：{result.output_path or "(试运行未写盘)"}'
            )
        else:
            msg = f'{dry_run_banner}未提取到任何银行流水记录。'

    if result.unprocessed_files:
        names = '\n  '.join(os.path.basename(f) for f in result.unprocessed_files)
        msg += f'\n\n无法识别的文件（{len(result.unprocessed_files)} 个，已保留）：\n  {names}'
    if result.error_files:
        err_info = '\n  '.join(f'{os.path.basename(f)}: {e}' for f, e in result.error_files)
        msg += f'\n\n处理出错的文件（{len(result.error_files)} 个，已保留）：\n  {err_info}'

    if result.pending_list_path:
        msg += f'\n\n待处理清单：{result.pending_list_path}'
        msg += '\n（请根据清单修正文件后重新导入）'
    elif result.dry_run and (result.unprocessed_files or result.error_files):
        msg += '\n\n待处理清单：(试运行未生成)'

    return msg


def generate_pending_list_xlsx(unprocessed_files, error_files, script_dir,
                                output_dir=None):
    """
    生成待处理清单.xlsx，列出无法识别和处理出错的文件。

    清单包含三列：
    - 文件路径：待处理文件的完整路径
    - 识别结果：无法识别银行类型 / 处理出错
    - 错误信息：具体的错误描述（无法识别的文件此项为空）

    Args:
        unprocessed_files: 无法识别银行类型的文件路径列表
        error_files: 处理出错的文件列表，每个元素为 (文件路径, 错误信息) 元组
        script_dir: 脚本目录，用于默认输出路径
        output_dir: 可选的输出目录，不指定时使用 script_dir

    Returns:
        生成的待处理清单文件路径，如果没有待处理文件则返回 None
    """
    logger = get_logger()

    if not unprocessed_files and not error_files:
        logger.info('没有待处理的文件，无需生成待处理清单')
        return None

    rows = []
    for filepath in unprocessed_files:
        rows.append({
            '文件路径': filepath,
            '识别结果': '无法识别银行类型',
            '错误信息': '',
        })
    for filepath, error_msg in error_files:
        rows.append({
            '文件路径': filepath,
            '识别结果': '处理出错',
            '错误信息': error_msg,
        })

    columns = ['文件路径', '识别结果', '错误信息']
    df = pd.DataFrame(rows, columns=columns)

    base_dir = output_dir or script_dir
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'待处理清单_{timestamp}.xlsx'
    output_path = os.path.join(base_dir, filename)

    try:
        df.to_excel(output_path, index=False, engine='openpyxl')
        logger.info('待处理清单已生成: %s（共 %d 条记录）', output_path, len(rows))
        return output_path
    except Exception as e:
        logger.error('生成待处理清单失败: %s', e, exc_info=True)
        return None


def delete_processed_files(excel_files, processed_files, error_files, unprocessed_files,
                           strategy='keep_unprocessed', archive_dir_name='已处理归档'):
    """
    根据策略处理已处理成功的 Excel 文件。

    策略说明:
    - keep_unprocessed: 仅保留未处理/有错误的文件，删除已处理成功的文件（默认）
    - keep_all: 保留所有文件，不做任何删除或移动
    - delete_all: 删除所有文件（无论处理状态）
    - move_to_archive: 将已处理成功的文件移动到归档子目录，保留未处理/有错误的文件

    Args:
        excel_files: 所有待处理的 Excel 文件列表
        processed_files: 已成功处理的文件列表
        error_files: 处理出错的文件列表 [(path, error_msg), ...]
        unprocessed_files: 未识别/未处理的文件列表
        strategy: 文件保留策略，默认为 'keep_unprocessed'
        archive_dir_name: 归档目录名（仅 move_to_archive 策略使用），默认为 '已处理归档'
    """
    logger = get_logger()

    if not excel_files:
        return

    if strategy not in KEEP_STRATEGIES:
        logger.warning('未知的保留策略「%s」，回退到默认 keep_unprocessed', strategy)
        strategy = 'keep_unprocessed'

    if strategy == 'keep_all':
        logger.debug('保留策略 keep_all：跳过所有文件操作')
        return

    error_file_paths = {f for f, _ in error_files}
    processed_set = set(processed_files)
    keep_set = set(unprocessed_files) | error_file_paths

    def _get_common_parent_dir(files):
        dirs = [os.path.dirname(os.path.abspath(f)) for f in files if f]
        if not dirs:
            return None
        if len(set(dirs)) == 1:
            return dirs[0]
        return os.path.commonpath(dirs)

    if strategy == 'move_to_archive':
        working_dir = _get_common_parent_dir(excel_files)
        if not working_dir:
            logger.error('无法确定归档目录的父路径，跳过归档操作')
            return
        archive_dir = os.path.join(working_dir, archive_dir_name)
        try:
            os.makedirs(archive_dir, exist_ok=True)
        except OSError as e:
            logger.error('创建归档目录失败「%s」: %s', archive_dir, e)
            return

        for filepath in excel_files:
            if filepath in processed_set and filepath not in error_file_paths:
                try:
                    if not os.path.exists(filepath):
                        logger.debug('归档目标不存在，跳过: %s', filepath)
                        continue
                    base_name = os.path.basename(filepath)
                    target_path = os.path.join(archive_dir, base_name)
                    counter = 1
                    while os.path.exists(target_path):
                        name, ext = os.path.splitext(base_name)
                        target_path = os.path.join(archive_dir, f'{name}_{counter}{ext}')
                        counter += 1
                    import shutil as _shutil
                    _shutil.move(filepath, target_path)
                    logger.debug('已归档文件: %s -> %s', filepath, target_path)
                except OSError as e:
                    logger.error('归档文件「%s」失败: %s', filepath, e)
        return

    if strategy == 'keep_unprocessed':
        delete_set = processed_set - error_file_paths
    elif strategy == 'delete_all':
        delete_set = set(excel_files)
    else:
        delete_set = processed_set - error_file_paths

    for filepath in excel_files:
        if filepath in delete_set:
            try:
                if os.path.exists(filepath):
                    os.remove(filepath)
                    logger.debug('已删除文件: %s', filepath)
            except OSError as e:
                logger.error('删除文件「%s」失败: %s', filepath, e)


def commit_pipeline_changes(result: ProcessingResult) -> ProcessingResult:
    """
    执行试运行结果的正式写盘操作。

    接收试运行（dry_run=True）模式下生成的 ProcessingResult，
    执行所有被推迟的破坏性写操作：
    1. 删除已处理成功的 Excel 源文件
    2. 写入/覆盖银行流水总表
    3. 回写黑白名单打标结果
    4. 回写非工作日打标结果
    5. 回写内部划转标记结果
    6. 生成脱敏版总表
    7. 数据库持久化
    8. 生成协同编辑模板
    9. 数字签名
    10. 文件加密

    Args:
        result: 试运行模式生成的 ProcessingResult 对象

    Returns:
        ProcessingResult: 更新后的结果对象，包含实际写盘后的路径和状态
    """
    logger = get_logger()

    if not result.dry_run:
        logger.warning('commit_pipeline_changes: 结果不是试运行模式（dry_run=False），无需提交')
        return result

    if result.changes_committed:
        logger.warning('commit_pipeline_changes: 变更已被提交过，跳过重复提交')
        return result

    if result.folder_empty:
        logger.warning('commit_pipeline_changes: 试运行结果为空文件夹，无需提交')
        result.changes_committed = True
        return result

    logger.info('===== 开始提交试运行变更（正式写盘）=====')

    script_dir = result.pending_script_dir
    lookup_data = result.pending_lookup_source
    final_rows = result.pending_final_rows
    all_files = result.pending_all_files
    keep_set = result.pending_keep_set
    existing_records = result.pending_existing_records
    incremental_rows = result.pending_incremental_rows
    incremental_mode = result.incremental_mode
    folder = result.pending_input_folder
    output_dir = result.pending_output_dir

    _cp_tag_summary = result.pending_cp_tag_summary or {}
    _holiday_tag_summary = result.pending_holiday_tag_summary or {}
    _it_summary = result.pending_internal_transfer_summary or {}
    _it_result = result.pending_internal_transfer_result

    if not script_dir:
        raise ValueError('试运行结果缺少必要字段：pending_script_dir')

    if result.pending_deletion_files:
        logger.info('[提交] 执行文件删除操作，待删除 %d 个文件', len(result.pending_deletion_files))
        delete_processed_files(
            all_files,
            result.processed_files,
            result.error_files,
            result.unprocessed_files,
            strategy=result.pending_keep_strategy,
            archive_dir_name=result.pending_archive_dir_name,
        )
    else:
        logger.info('[提交] 无待删除文件')

    output_path = None
    if final_rows:
        if incremental_mode:
            logger.info('[提交] 执行增量合并写总表：历史 %d 条 + 新增 %d 条',
                        len(existing_records), len(incremental_rows))
            output_path = merge_and_export_summary(
                existing_records, incremental_rows, script_dir,
                output_dir=output_dir, lookup_source=lookup_data
            )
        else:
            logger.info('[提交] 执行全量覆盖写总表：共 %d 条记录', len(final_rows))
            columns = get_summary_columns(final_rows, lookup_data)
            df = pd.DataFrame(final_rows, columns=columns)
            output_path = get_summary_table_path(script_dir, output_dir)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            df.to_excel(output_path, index=False, engine='openpyxl')
            logger.info('[提交] 总表输出完成: %s', output_path)
    elif existing_records:
        logger.info('[提交] 无新增记录，仅基于历史记录写总表')
        output_path = merge_and_export_summary(
            existing_records, [], script_dir,
            output_dir=output_dir, lookup_source=lookup_data
        )

    result.output_path = output_path

    if output_path and _cp_tag_summary.get('tagged_count', 0) > 0:
        logger.info('[提交] 回写黑白名单打标结果到总表')
        base_columns = get_summary_columns(final_rows, lookup_data)
        cp_extra_cols = ['黑白名单标签', '命中规则名称', '命中关键词']
        _cp_columns = base_columns + [
            col for col in cp_extra_cols if col not in base_columns
        ]
        pd.DataFrame(final_rows, columns=_cp_columns).to_excel(
            output_path, index=False, engine='openpyxl')
        logger.info('[提交] 黑白名单打标结果已回写')

    if output_path and _holiday_tag_summary.get('tagged_count', 0) > 0:
        logger.info('[提交] 回写非工作日打标结果到总表')
        base_columns = get_summary_columns(final_rows, lookup_data)
        holiday_extra_cols = ['非工作日标签', '节假日名称']
        _holiday_columns = base_columns + [
            col for col in holiday_extra_cols if col not in base_columns
        ]
        pd.DataFrame(final_rows, columns=_holiday_columns).to_excel(
            output_path, index=False, engine='openpyxl')
        logger.info('[提交] 非工作日打标结果已回写')

    if output_path and _it_summary.get('match_pairs', 0) > 0:
        logger.info('[提交] 回写内部划转标记结果到总表')
        base_columns = get_summary_columns(final_rows, lookup_data)
        it_extra_cols = list(INTERNAL_TRANSFER_EXTRA_COLUMNS)
        _it_columns = base_columns + [
            col for col in it_extra_cols if col not in base_columns
        ]
        pd.DataFrame(final_rows, columns=_it_columns).to_excel(
            output_path, index=False, engine='openpyxl')
        logger.info('[提交] 内部划转标记结果已回写')

    masked_output_path = None
    if final_rows and output_path:
        try:
            logger.info('[提交] 生成脱敏版总表')
            _output_dir = os.path.dirname(output_path) or script_dir
            _masked_columns = get_summary_columns(final_rows, lookup_data)
            if _cp_tag_summary and _cp_tag_summary.get('tagged_count', 0) > 0:
                cp_extra_cols = ['黑白名单标签', '命中规则名称', '命中关键词']
                _masked_columns = _masked_columns + [
                    col for col in cp_extra_cols if col not in _masked_columns
                ]
            else:
                cp_extra_cols = ['黑白名单标签', '命中规则名称', '命中关键词']
                _masked_columns = [col for col in _masked_columns if col not in cp_extra_cols]
            masked_output_path = export_masked_summary(
                final_rows, script_dir, output_dir=_output_dir,
                lookup_source=lookup_data, columns=_masked_columns
            )
        except Exception as e:
            logger.error('[提交] 生成脱敏版总表失败: %s', e, exc_info=True)
    result.masked_output_path = masked_output_path

    db_inserted = 0
    db_duplicates = 0
    if HAS_DATABASE and final_rows:
        try:
            batch_id = result.pending_batch_id
            if batch_id is None:
                batch_id = f"BATCH{datetime.now().strftime('%Y%m%d%H%M%S')}"
            logger.info('[提交] 执行数据库持久化')
            db_inserted, db_duplicates = db_module.persist_transactions(
                final_rows,
                batch_id=batch_id,
                deduplicate=True,
                script_dir=script_dir,
            )
            logger.info('[提交] 数据库持久化完成: 插入 %d 条, 去重跳过 %d 条',
                        db_inserted, db_duplicates)
        except Exception as e:
            logger.error('[提交] 数据库持久化失败: %s', e, exc_info=True)
    result.db_inserted_count = db_inserted
    result.db_duplicate_count = db_duplicates

    collab_template_path = None
    if output_path and final_rows:
        try:
            logger.info('[提交] 生成财务协同编辑模板')
            collab_output_dir = os.path.dirname(output_path) or script_dir
            collab_template_path = generate_collab_template(
                output_path,
                output_dir=collab_output_dir,
                lookup_source=lookup_data,
            )
        except Exception as e:
            logger.error('[提交] 生成协同编辑模板失败: %s', e, exc_info=True)
    result.collab_template_path = collab_template_path

    output_hash = None
    signature_id = None
    signature_info = None
    if output_path and result.pending_enable_signature and HAS_CRYPTOGRAPHY:
        try:
            logger.info('[提交] 执行数字签名')
            output_hash = compute_file_hash(output_path)
            if result.pending_auto_generate_key and not has_signing_key(script_dir):
                ensure_signing_key(script_dir, auto_generate=True,
                                   password=result.pending_signature_password)
            signature_info = sign_output_file(
                output_path=output_path,
                script_dir=script_dir,
                input_directory=folder,
                password=result.pending_signature_password,
                extra_data={
                    'record_count': len(final_rows),
                    'incremental_mode': incremental_mode,
                    'processed_files': len(result.processed_files),
                }
            )
            if signature_info:
                signature_id = save_signature_record(
                    signature_info, script_dir=script_dir
                )
                signature_info['signature_id'] = signature_id
                logger.info('[提交] 数字签名完成: %s', signature_id)
        except Exception as e:
            logger.error('[提交] 数字签名失败: %s', e, exc_info=True)
    result.output_hash = output_hash
    result.signature_id = signature_id
    result.signature_info = signature_info

    encryption_result = None
    encrypted_files = []
    if (result.pending_enable_encryption and result.pending_encryption_password
            and HAS_FILE_ENCRYPTION):
        try:
            logger.info('[提交] 执行输出文件加密')
            enc_output_dir = script_dir
            if output_path:
                enc_output_dir = os.path.dirname(output_path) or script_dir

            files_to_encrypt = []
            for fp in [output_path, result.subject_summary_path, result.balance_check_path,
                       result.duplicate_check_path, result.interest_fee_check_path,
                       result.holiday_check_path, result.accounting_period_path]:
                if fp and os.path.isfile(fp):
                    files_to_encrypt.append(fp)

            if files_to_encrypt:
                encryption_result = _encrypt_output_files(
                    files_to_encrypt,
                    password=result.pending_encryption_password,
                    mode=result.pending_encryption_mode,
                    output_dir=enc_output_dir,
                )
                encrypted_files = [
                    r.encrypted_path for r in encryption_result.results
                    if r.success and r.encrypted_path
                ]
                _save_encryption_record(encryption_result, script_dir=script_dir)
                logger.info('[提交] 文件加密完成: 成功 %d, 失败 %d',
                            encryption_result.success_count,
                            encryption_result.failure_count)
        except Exception as e:
            logger.error('[提交] 输出文件加密失败: %s', e, exc_info=True)
    result.encryption_result = encryption_result
    result.encrypted_files = encrypted_files

    if result.unprocessed_files or result.error_files:
        pending_list_path = generate_pending_list_xlsx(
            result.unprocessed_files, result.error_files, script_dir
        )
        result.pending_list_path = pending_list_path

    result.dry_run = False
    result.changes_committed = True
    logger.info('===== 试运行变更提交完成 =====')
    return result


# ──────────────────────────────────────────────
# 流水文件变更对比
# ──────────────────────────────────────────────

DIFF_COLUMNS = [
    '银行', '银行账号', '主体', '交易日期',
    '付款', '收款', '摘要', '对方户名', '余额', '交易流水号',
]

AMOUNT_FIELDS = ['付款', '收款', '余额']

# ──────────────────────────────────────────────
# 金额异常检测配置
# ──────────────────────────────────────────────

class AmountAnomalyType:
    """金额异常类型常量"""
    LARGE_AMOUNT = 'large_amount'
    ZERO_WITH_COUNTERPARTY = 'zero_with_counterparty'
    NEGATIVE_RECEIPT = 'negative_receipt'
    POSITIVE_PAYMENT = 'positive_payment'
    BALANCE_NEGATIVE = 'balance_negative'
    BOTH_PAYMENT_AND_RECEIPT = 'both_payment_and_receipt'
    NEITHER_PAYMENT_NOR_RECEIPT = 'neither_payment_nor_receipt'

    LABELS = {
        LARGE_AMOUNT: '单笔金额超过阈值',
        ZERO_WITH_COUNTERPARTY: '金额为0但有对方户名',
        NEGATIVE_RECEIPT: '收款金额为负数',
        POSITIVE_PAYMENT: '付款金额为正数',
        BALANCE_NEGATIVE: '余额为负数',
        BOTH_PAYMENT_AND_RECEIPT: '付款和收款同时有值',
        NEITHER_PAYMENT_NOR_RECEIPT: '付款和收款均无值',
    }

    RISK_LEVELS = {
        LARGE_AMOUNT: 'high',
        ZERO_WITH_COUNTERPARTY: 'medium',
        NEGATIVE_RECEIPT: 'high',
        POSITIVE_PAYMENT: 'high',
        BALANCE_NEGATIVE: 'medium',
        BOTH_PAYMENT_AND_RECEIPT: 'medium',
        NEITHER_PAYMENT_NOR_RECEIPT: 'low',
    }


@dataclass
class AmountAnomalyConfig:
    """金额异常检测配置"""
    single_amount_threshold: float = 500000.0
    enable_zero_with_counterparty: bool = True
    enable_negative_receipt: bool = True
    enable_positive_payment: bool = True
    enable_negative_balance: bool = True
    enable_both_amounts: bool = True
    enable_no_amounts: bool = True

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

    @classmethod
    def default(cls) -> 'AmountAnomalyConfig':
        return cls()


def detect_amount_anomalies(
    record: Dict[str, Any],
    config: Optional[AmountAnomalyConfig] = None
) -> Tuple[bool, List[str], List[str]]:
    """
    检测单条记录的金额异常。

    Args:
        record: 交易记录字典
        config: 异常检测配置，为空则使用默认配置

    Returns:
        Tuple[bool, List[str], List[str]]: (是否有异常, 异常类型列表, 异常描述列表)
    """
    if config is None:
        config = AmountAnomalyConfig.default()

    anomalies = []
    descriptions = []

    payment = to_float(record.get('付款'))
    receipt = to_float(record.get('收款'))
    balance = to_float(record.get('余额'))
    counterparty = record.get('对方户名')

    has_payment = payment is not None and payment != 0
    has_receipt = receipt is not None and receipt != 0

    # 1. 单笔金额超过阈值
    if config.single_amount_threshold and config.single_amount_threshold > 0:
        amount = None
        if has_payment:
            amount = abs(payment)
        elif has_receipt:
            amount = receipt
        if amount is not None and amount >= config.single_amount_threshold:
            anomalies.append(AmountAnomalyType.LARGE_AMOUNT)
            descriptions.append(
                f'{AmountAnomalyType.LABELS[AmountAnomalyType.LARGE_AMOUNT]}: '
                f'{amount:,.2f} 元，阈值 {config.single_amount_threshold:,.2f} 元'
            )

    # 2. 金额为0但有对方户名
    if config.enable_zero_with_counterparty and counterparty is not None:
        cp_str = str(counterparty).strip()
        if cp_str:
            both_zero = (
                (payment is not None and payment == 0) and
                (receipt is not None and receipt == 0)
            )
            both_none = payment is None and receipt is None
            one_zero_one_none = (
                (payment == 0 and receipt is None) or
                (payment is None and receipt == 0)
            )
            if both_zero or both_none or one_zero_one_none:
                anomalies.append(AmountAnomalyType.ZERO_WITH_COUNTERPARTY)
                descriptions.append(
                    f'{AmountAnomalyType.LABELS[AmountAnomalyType.ZERO_WITH_COUNTERPARTY]}: '
                    f'对方户名「{cp_str}」'
                )

    # 3. 收款金额为负数
    if config.enable_negative_receipt and has_receipt and receipt < 0:
        anomalies.append(AmountAnomalyType.NEGATIVE_RECEIPT)
        descriptions.append(
            f'{AmountAnomalyType.LABELS[AmountAnomalyType.NEGATIVE_RECEIPT]}: '
            f'{receipt:,.2f} 元'
        )

    # 4. 付款金额为正数（付款约定为负数）
    if config.enable_positive_payment and has_payment and payment > 0:
        anomalies.append(AmountAnomalyType.POSITIVE_PAYMENT)
        descriptions.append(
            f'{AmountAnomalyType.LABELS[AmountAnomalyType.POSITIVE_PAYMENT]}: '
            f'{payment:,.2f} 元'
        )

    # 5. 余额为负数
    if config.enable_negative_balance and balance is not None and balance < 0:
        anomalies.append(AmountAnomalyType.BALANCE_NEGATIVE)
        descriptions.append(
            f'{AmountAnomalyType.LABELS[AmountAnomalyType.BALANCE_NEGATIVE]}: '
            f'{balance:,.2f} 元'
        )

    # 6. 付款和收款同时有非零值
    if config.enable_both_amounts and has_payment and has_receipt:
        anomalies.append(AmountAnomalyType.BOTH_PAYMENT_AND_RECEIPT)
        descriptions.append(
            f'{AmountAnomalyType.LABELS[AmountAnomalyType.BOTH_PAYMENT_AND_RECEIPT]}: '
            f'付款 {payment:,.2f} 元，收款 {receipt:,.2f} 元'
        )

    # 7. 付款和收款均无值或均为0
    if config.enable_no_amounts and not has_payment and not has_receipt:
        anomalies.append(AmountAnomalyType.NEITHER_PAYMENT_NOR_RECEIPT)
        descriptions.append(
            AmountAnomalyType.LABELS[AmountAnomalyType.NEITHER_PAYMENT_NOR_RECEIPT]
        )

    has_anomaly = len(anomalies) > 0
    return has_anomaly, anomalies, descriptions


def apply_amount_anomaly_detection(
    records: List[Dict[str, Any]],
    config: Optional[AmountAnomalyConfig] = None
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    对记录列表应用金额异常检测，添加异常标记列。

    Args:
        records: 交易记录列表
        config: 异常检测配置

    Returns:
        Tuple[List[Dict[str, Any]], Dict[str, Any]]: (处理后的记录列表, 统计摘要)
    """
    logger = get_logger()

    if config is None:
        config = AmountAnomalyConfig.default()

    if not records:
        logger.info('无记录可进行金额异常检测')
        return records, {
            'total_records': 0,
            'anomaly_count': 0,
            'anomaly_types': {},
        }

    threshold_str = f'{config.single_amount_threshold:,.2f}'
    logger.info('开始金额异常检测，单笔阈值: %s 元', threshold_str)

    anomaly_type_counts: Dict[str, int] = defaultdict(int)
    anomaly_count = 0

    for record in records:
        has_anomaly, anomalies, descriptions = detect_amount_anomalies(record, config)

        if has_anomaly:
            anomaly_count += 1
            flag_parts = []
            for anomaly_type in anomalies:
                anomaly_type_counts[anomaly_type] += 1
                risk_level = AmountAnomalyType.RISK_LEVELS.get(anomaly_type, 'medium')
                risk_label = {'high': '高', 'medium': '中', 'low': '低'}.get(risk_level, '中')
                type_label = AmountAnomalyType.LABELS.get(anomaly_type, anomaly_type)
                flag_parts.append(f'[{risk_label}]{type_label}')

            record[ANOMALY_FLAG_COLUMN] = '; '.join(flag_parts)
            record[ANOMALY_DETAIL_COLUMN] = '; '.join(descriptions)
        else:
            record[ANOMALY_FLAG_COLUMN] = ''
            record[ANOMALY_DETAIL_COLUMN] = ''

    summary = {
        'total_records': len(records),
        'anomaly_count': anomaly_count,
        'anomaly_rate': anomaly_count / len(records) if records else 0,
        'anomaly_types': dict(anomaly_type_counts),
        'anomaly_type_labels': {
            k: {'count': v, 'label': AmountAnomalyType.LABELS.get(k, k)}
            for k, v in anomaly_type_counts.items()
        },
        'config': config.to_dict(),
    }

    logger.info(
        '金额异常检测完成：共 %d 条记录，%d 条异常 (%.2f%%)',
        len(records), anomaly_count,
        summary['anomaly_rate'] * 100
    )
    for anomaly_type, count in anomaly_type_counts.items():
        label = AmountAnomalyType.LABELS.get(anomaly_type, anomaly_type)
        logger.info('  - %s: %d 条', label, count)

    return records, summary


def _make_match_key(row):
    transaction_id = row.get('交易流水号')
    bank_account = _account_key(row.get('银行账号'))
    if transaction_id is not None and str(transaction_id).strip():
        tid = str(transaction_id).strip()
        return f"{bank_account}::{tid}"
    payment = row.get('付款')
    receipt = row.get('收款')
    trade_date = row.get('交易日期')
    p = '' if payment is None else str(payment)
    r = '' if receipt is None else str(receipt)
    d = '' if trade_date is None else str(trade_date)
    return f"{bank_account}::{d}::{p}::{r}"


def _is_nan_or_none(value):
    if value is None:
        return True
    if isinstance(value, float) and value != value:
        return True
    return False


def _values_equal(val_a, val_b, field_name):
    if field_name in AMOUNT_FIELDS:
        fa = to_float(val_a)
        fb = to_float(val_b)
        if _is_nan_or_none(fa) and _is_nan_or_none(fb):
            return True
        if _is_nan_or_none(fa) or _is_nan_or_none(fb):
            return False
        return abs(fa - fb) < 0.005
    sa = '' if _is_nan_or_none(val_a) else str(val_a).strip()
    sb = '' if _is_nan_or_none(val_b) else str(val_b).strip()
    return sa == sb


@dataclass
class DiffRecord:
    change_type: str
    match_key: str
    old_row: Optional[dict] = None
    new_row: Optional[dict] = None
    changed_fields: Optional[List[str]] = None

    def to_flat_dict(self):
        source = self.new_row if self.new_row else self.old_row
        result = {'变更类型': self.change_type}
        for col in DIFF_COLUMNS:
            result[col] = source.get(col) if source else None
        if self.change_type == '变更' and self.changed_fields:
            parts = []
            for f in self.changed_fields:
                old_v = self.old_row.get(f) if self.old_row else None
                new_v = self.new_row.get(f) if self.new_row else None
                parts.append(f"{f}: {old_v} → {new_v}")
            result['变更明细'] = '; '.join(parts)
        else:
            result['变更明细'] = ''
        if self.change_type == '变更':
            for f in (self.changed_fields or []):
                if f in AMOUNT_FIELDS:
                    old_v = self.old_row.get(f) if self.old_row else None
                    new_v = self.new_row.get(f) if self.new_row else None
                    result[f'{f}(旧)'] = old_v
                    result[f'{f}(新)'] = new_v
        return result


@dataclass
class DiffResult:
    records: List[DiffRecord] = field(default_factory=list)
    added_count: int = 0
    deleted_count: int = 0
    changed_count: int = 0
    unchanged_count: int = 0
    output_path: Optional[str] = None


def diff_transactions(old_df, new_df):
    logger = get_logger()

    old_keyed = {}
    for _, row in old_df.iterrows():
        key = _make_match_key(row)
        old_keyed[key] = row.to_dict()

    new_keyed = {}
    for _, row in new_df.iterrows():
        key = _make_match_key(row)
        new_keyed[key] = row.to_dict()

    all_keys = list(dict.fromkeys(list(old_keyed.keys()) + list(new_keyed.keys())))

    records = []
    added = deleted = changed = unchanged = 0

    for key in all_keys:
        in_old = key in old_keyed
        in_new = key in new_keyed

        if in_old and not in_new:
            records.append(DiffRecord(
                change_type='删除', match_key=key, old_row=old_keyed[key],
            ))
            deleted += 1
        elif in_new and not in_old:
            records.append(DiffRecord(
                change_type='新增', match_key=key, new_row=new_keyed[key],
            ))
            added += 1
        else:
            old_r = old_keyed[key]
            new_r = new_keyed[key]
            changed_fields = []
            for col in DIFF_COLUMNS:
                if not _values_equal(old_r.get(col), new_r.get(col), col):
                    changed_fields.append(col)
            if changed_fields:
                records.append(DiffRecord(
                    change_type='变更', match_key=key,
                    old_row=old_r, new_row=new_r,
                    changed_fields=changed_fields,
                ))
                changed += 1
            else:
                records.append(DiffRecord(
                    change_type='未变更', match_key=key,
                    old_row=old_r, new_row=new_r,
                ))
                unchanged += 1

    logger.info(
        '变更对比完成: 新增 %d, 删除 %d, 变更 %d, 未变更 %d',
        added, deleted, changed, unchanged,
    )
    return DiffResult(
        records=records,
        added_count=added,
        deleted_count=deleted,
        changed_count=changed,
        unchanged_count=unchanged,
    )


DIFF_HIGHLIGHT_COLORS = {
    '新增': 'C6EFCE',
    '删除': 'FFC7CE',
    '变更': 'FFEB9C',
}


def export_diff_result(diff_result, output_path):
    logger = get_logger()

    flat_rows = [r.to_flat_dict() for r in diff_result.records]
    if not flat_rows:
        logger.warning('无可对比结果，不生成输出文件')
        return None

    has_changes = any(r['变更明细'] for r in flat_rows)
    extra_cols = []
    if has_changes:
        extra_cols.append('变更明细')
    amount_change_cols = []
    for f in AMOUNT_FIELDS:
        col_old = f'{f}(旧)'
        col_new = f'{f}(新)'
        if any(col_old in r for r in flat_rows):
            amount_change_cols.extend([col_old, col_new])
    extra_cols.extend(amount_change_cols)

    columns = ['变更类型'] + DIFF_COLUMNS + extra_cols

    df = pd.DataFrame(flat_rows)
    for col in columns:
        if col not in df.columns:
            df[col] = None
    df = df[columns]

    df.to_excel(output_path, index=False, engine='openpyxl')

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    type_col_idx = 1
    for row_idx in range(2, ws.max_row + 1):
        cell_type = ws.cell(row=row_idx, column=type_col_idx)
        change_type = str(cell_type.value).strip() if cell_type.value else ''
        fill_color = DIFF_HIGHLIGHT_COLORS.get(change_type)
        if fill_color:
            fill = openpyxl.styles.PatternFill(
                start_color=fill_color, end_color=fill_color, fill_type='solid',
            )
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    wb.save(output_path)
    wb.close()

    logger.info('变更对比结果已输出: %s', output_path)
    return output_path


def run_diff(old_path, new_path, output_dir=None):
    logger = get_logger()
    logger.info('========== 流水变更对比开始 ==========')
    logger.info('旧批次文件: %s', old_path)
    logger.info('新批次文件: %s', new_path)

    if not os.path.exists(old_path):
        logger.error('旧批次总表文件不存在: %s', old_path)
        raise FileNotFoundError(f'旧批次总表文件不存在: {old_path}')
    if not os.path.exists(new_path):
        logger.error('新批次总表文件不存在: %s', new_path)
        raise FileNotFoundError(f'新批次总表文件不存在: {new_path}')

    old_df = pd.read_excel(old_path, engine='openpyxl')
    new_df = pd.read_excel(new_path, engine='openpyxl')

    required_cols = ['银行账号', '交易流水号']
    for col in required_cols:
        if col not in old_df.columns:
            logger.warning('旧批次总表缺少列: %s', col)
        if col not in new_df.columns:
            logger.warning('新批次总表缺少列: %s', col)

    diff_result = diff_transactions(old_df, new_df)

    if output_dir is None:
        output_dir = get_script_dir()
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f'流水变更对比_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, output_filename)

    export_diff_result(diff_result, output_path)
    diff_result.output_path = output_path

    logger.info('========== 流水变更对比结束 ==========')
    return diff_result


def format_diff_message(diff_result):
    if not diff_result.records:
        return '两份总表无数据可对比。'

    total = len(diff_result.records)
    msg = (
        f'变更对比完成！\n\n'
        f'总记录数：{total}\n'
        f'新增交易：{diff_result.added_count}\n'
        f'删除交易：{diff_result.deleted_count}\n'
        f'金额/内容变更：{diff_result.changed_count}\n'
        f'未变更：{diff_result.unchanged_count}'
    )
    if diff_result.output_path:
        msg += f'\n\n对比结果文件：{diff_result.output_path}'
    return msg


def ask_dry_run_mode():
    """询问用户是否启用试运行模式"""
    print('\n请选择执行模式：')
    print('  1) 试运行模式（推荐）：仅生成统计与异常报告，不删除文件、不写总表')
    print('  2) 正式写盘模式：直接执行删除与写盘操作')
    choice = input('请输入选项（直接回车默认为 1 试运行模式）: ').strip()
    if choice == '':
        return True
    if choice == '1':
        return True
    if choice == '2':
        return False
    print('无效选项，默认使用试运行模式')
    return True


def ask_commit_changes():
    """询问用户是否确认提交试运行变更"""
    print('\n' + '=' * 60)
    print('试运行完成，已生成统计与异常报告，请检查报告内容。')
    print('=' * 60)
    print('\n是否确认执行以下正式操作？')
    print('  - 删除已处理成功的源文件（按保留策略）')
    print('  - 写入/覆盖银行流水总表')
    print('  - 生成脱敏版总表')
    print('  - 执行数据库持久化（如已配置）')
    print('  - 执行数字签名和文件加密（如已启用）')
    print()
    choice = input('请确认提交 (y/N): ').strip().lower()
    return choice in ('y', 'yes')


def run_pipeline_flow(script_dir):
    """主流程：处理银行流水文件夹，输出总表"""
    logger = get_logger()

    folder = ask_directory('请选择银行流水文件夹')
    if not folder:
        show_info('提示', '未选择文件夹，程序退出。')
        logger.info('用户未选择文件夹，程序退出')
        return

    logger.info('用户选择文件夹: %s', folder)

    incremental = ask_incremental_mode()
    if incremental is None:
        logger.info('用户取消增量模式选择，返回主菜单')
        return
    logger.info('用户选择运行模式: %s', '增量合并' if incremental else '全量覆盖')

    dry_run = ask_dry_run_mode()
    logger.info('用户选择执行模式: %s', '试运行（不写盘）' if dry_run else '正式写盘')

    result = run_pipeline(folder, script_dir, incremental=incremental, dry_run=dry_run)

    if result.lookup_missing:
        show_warning(
            '警告',
            '在程序所在目录下未找到主体查找表文件，\n"主体"列将为空。\n'
            '建议将查找表文件命名为"主体查找表.xlsx"并放在程序所在目录下。'
        )

    msg = format_result_message(result)

    if result.dry_run and not result.folder_empty and result.all_rows:
        print('\n' + msg)
        if ask_commit_changes():
            logger.info('用户确认提交试运行变更')
            result = commit_pipeline_changes(result)
            msg = format_result_message(result)
            show_info('提交完成', msg)
        else:
            logger.info('用户取消提交试运行变更')
            msg += '\n\n⚠️  已取消提交，未执行任何删除与写盘操作。'
            show_info('已取消', msg)
    else:
        show_info('完成' if result.all_rows else '提示', msg)


def run_diff_flow(script_dir):
    """变更对比流程：选择两次总表，输出对比结果"""
    logger = get_logger()

    old_path = ask_file('请选择【旧批次】银行流水总表')
    if not old_path:
        show_info('提示', '未选择旧批次文件，程序退出。')
        logger.info('用户未选择旧批次文件，程序退出')
        return
    logger.info('用户选择旧批次文件: %s', old_path)

    new_path = ask_file('请选择【新批次】银行流水总表')
    if not new_path:
        show_info('提示', '未选择新批次文件，程序退出。')
        logger.info('用户未选择新批次文件，程序退出')
        return
    logger.info('用户选择新批次文件: %s', new_path)

    try:
        diff_result = run_diff(old_path, new_path, script_dir)
        msg = format_diff_message(diff_result)

        has_changes = (
            diff_result.added_count > 0
            or diff_result.deleted_count > 0
            or diff_result.changed_count > 0
        )
        title = '对比完成（发现差异' if has_changes else '对比完成（无差异）'
        show_info(title, msg)
    except FileNotFoundError as e:
        show_warning('错误', str(e))
        logger.error('对比失败: %s', e)


# ──────────────────────────────────────────────
# 多用户与操作审计模块
# ──────────────────────────────────────────────

AUDIT_DB_FILENAME = 'audit_log.db'


def get_audit_db_path(script_dir=None):
    """获取审计数据库文件路径"""
    if script_dir is None:
        script_dir = get_script_dir()
    return os.path.join(script_dir, AUDIT_DB_FILENAME)


def _migrate_audit_db_columns(db_path):
    """迁移审计数据库表结构，添加缺失的列"""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    try:
        cursor.execute("PRAGMA table_info(audit_logs)")
        existing_columns = {row[1] for row in cursor.fetchall()}

        new_columns = [
            ('signature_id', 'TEXT'),
            ('digital_signature', 'TEXT'),
            ('signature_algorithm', 'TEXT'),
            ('signed_at', 'TEXT'),
        ]

        for col_name, col_type in new_columns:
            if col_name not in existing_columns:
                cursor.execute(f'ALTER TABLE audit_logs ADD COLUMN {col_name} {col_type}')
                logger = get_logger()
                logger.info('审计数据库已添加列: %s', col_name)

        conn.commit()
    except Exception as e:
        logger = get_logger()
        logger.warning('审计数据库迁移失败: %s', e)
        conn.rollback()
    finally:
        conn.close()


def init_audit_db(db_path=None):
    """
    初始化审计数据库，创建所需表结构。
    包含四张核心表：
    - users: 用户信息表
    - audit_logs: 操作审计主表
    - config_changes: 配置变更历史表
    - lookup_snapshots: 查找表快照表（用于自动检测变更）
    """
    if db_path is None:
        db_path = get_audit_db_path()

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            display_name TEXT,
            role TEXT DEFAULT 'operator',
            created_at TEXT NOT NULL,
            last_login TEXT,
            is_active INTEGER DEFAULT 1
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS audit_logs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            audit_id TEXT NOT NULL UNIQUE,
            user_id INTEGER,
            username TEXT NOT NULL,
            operation_type TEXT NOT NULL,
            input_directory TEXT,
            output_path TEXT,
            processed_files INTEGER DEFAULT 0,
            extracted_records INTEGER DEFAULT 0,
            unprocessed_files INTEGER DEFAULT 0,
            error_files INTEGER DEFAULT 0,
            lookup_file TEXT,
            lookup_missing INTEGER DEFAULT 0,
            status TEXT NOT NULL,
            error_message TEXT,
            config_snapshot TEXT,
            input_files_hash TEXT,
            output_hash TEXT,
            started_at TEXT NOT NULL,
            completed_at TEXT,
            duration_ms INTEGER,
            client_ip TEXT,
            client_hostname TEXT,
            signature_id TEXT,
            digital_signature TEXT,
            signature_algorithm TEXT,
            signed_at TEXT,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')

    _migrate_audit_db_columns(db_path)

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS config_changes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            change_id TEXT NOT NULL UNIQUE,
            user_id INTEGER,
            username TEXT NOT NULL,
            config_type TEXT NOT NULL,
            config_name TEXT NOT NULL,
            old_value TEXT,
            new_value TEXT,
            old_hash TEXT,
            new_hash TEXT,
            change_reason TEXT,
            changed_at TEXT NOT NULL,
            change_details TEXT,
            FOREIGN KEY (user_id) REFERENCES users (id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS lookup_snapshots (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            snapshot_id TEXT NOT NULL UNIQUE,
            lookup_file TEXT NOT NULL,
            file_hash TEXT NOT NULL,
            content_json TEXT NOT NULL,
            created_at TEXT NOT NULL,
            created_by TEXT NOT NULL
        )
    ''')

    cursor.execute('CREATE INDEX IF NOT EXISTS idx_audit_logs_username ON audit_logs(username)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_audit_logs_operation ON audit_logs(operation_type)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_audit_logs_started_at ON audit_logs(started_at)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_config_changes_config ON config_changes(config_type, config_name)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_lookup_snapshots_file ON lookup_snapshots(lookup_file)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_lookup_snapshots_created ON lookup_snapshots(created_at)')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS alert_rules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            rule_id TEXT NOT NULL UNIQUE,
            rule_name TEXT NOT NULL,
            rule_type TEXT NOT NULL,
            metric TEXT NOT NULL,
            operator TEXT NOT NULL,
            threshold REAL NOT NULL,
            window_minutes INTEGER DEFAULT 60,
            severity TEXT DEFAULT 'warning',
            enabled INTEGER DEFAULT 1,
            description TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            created_by TEXT NOT NULL
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS alerts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            alert_id TEXT NOT NULL UNIQUE,
            rule_id TEXT NOT NULL,
            rule_name TEXT NOT NULL,
            severity TEXT NOT NULL,
            alert_type TEXT NOT NULL,
            metric TEXT NOT NULL,
            current_value REAL,
            threshold REAL,
            operator TEXT,
            window_minutes INTEGER,
            description TEXT,
            triggered_at TEXT NOT NULL,
            resolved_at TEXT,
            status TEXT DEFAULT 'active',
            acknowledged INTEGER DEFAULT 0,
            acknowledged_by TEXT,
            acknowledged_at TEXT,
            audit_log_id TEXT,
            details TEXT,
            FOREIGN KEY (rule_id) REFERENCES alert_rules (rule_id)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS file_processing_details (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            detail_id TEXT NOT NULL UNIQUE,
            audit_id TEXT NOT NULL,
            file_path TEXT NOT NULL,
            file_name TEXT NOT NULL,
            bank TEXT,
            file_size INTEGER,
            record_count INTEGER DEFAULT 0,
            status TEXT NOT NULL,
            error_message TEXT,
            processing_duration_ms INTEGER,
            started_at TEXT NOT NULL,
            completed_at TEXT,
            FOREIGN KEY (audit_id) REFERENCES audit_logs (audit_id)
        )
    ''')

    cursor.execute('CREATE INDEX IF NOT EXISTS idx_alert_rules_type ON alert_rules(rule_type)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_alert_rules_enabled ON alert_rules(enabled)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_alerts_status ON alerts(status)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_alerts_severity ON alerts(severity)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_alerts_triggered ON alerts(triggered_at)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_alerts_rule ON alerts(rule_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_file_details_audit ON file_processing_details(audit_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_file_details_status ON file_processing_details(status)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_file_details_bank ON file_processing_details(bank)')

    conn.commit()
    conn.close()

    logger = get_logger()
    logger.info('审计数据库初始化完成: %s', db_path)


def get_current_user():
    """
    获取当前操作用户。
    优先级：环境变量 BANKCHECK_USER > 系统登录用户 > 'unknown'
    """
    env_user = os.environ.get('BANKCHECK_USER', '').strip()
    if env_user:
        return env_user
    try:
        return getpass.getuser()
    except Exception:
        return 'unknown'


def get_client_info():
    """获取客户端信息（主机名、IP等）"""
    import socket
    hostname = ''
    ip = ''
    try:
        hostname = socket.gethostname()
        ip = socket.gethostbyname(hostname)
    except Exception:
        pass
    return {'hostname': hostname, 'ip': ip}


def _ensure_user(username, db_path=None):
    """确保用户存在于数据库中，不存在则创建"""
    if db_path is None:
        db_path = get_audit_db_path()

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute('SELECT id FROM users WHERE username = ?', (username,))
    row = cursor.fetchone()

    if row:
        user_id = row[0]
        cursor.execute(
            'UPDATE users SET last_login = ? WHERE id = ?',
            (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), user_id)
        )
    else:
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        cursor.execute(
            'INSERT INTO users (username, display_name, created_at, last_login) VALUES (?, ?, ?, ?)',
            (username, username, now, now)
        )
        user_id = cursor.lastrowid

    conn.commit()
    conn.close()
    return user_id


def compute_file_hash(filepath):
    """计算文件的 SHA256 哈希值，用于完整性校验"""
    if not filepath or not os.path.exists(filepath):
        return None
    sha256 = hashlib.sha256()
    try:
        with open(filepath, 'rb') as f:
            for chunk in iter(lambda: f.read(8192), b''):
                sha256.update(chunk)
        return sha256.hexdigest()
    except Exception:
        return None


def compute_config_snapshot(script_dir):
    """
    生成当前配置快照，包含：
    - 支持的银行列表
    - 银行处理器配置
    - 查找表文件信息
    """
    lookup_file = find_lookup_file(script_dir)
    snapshot = {
        'supported_banks': BANK_PREFIXES,
        'bank_processors': list(BANK_PROCESSORS.keys()),
        'lookup_file': lookup_file,
        'lookup_file_hash': compute_file_hash(lookup_file) if lookup_file else None,
        'diff_columns': DIFF_COLUMNS,
        'amount_fields': AMOUNT_FIELDS,
        'timestamp': datetime.now().isoformat(),
    }
    return json.dumps(snapshot, ensure_ascii=False)


@dataclass
class AuditRecord:
    """审计记录数据类"""
    audit_id: str
    username: str
    operation_type: str
    input_directory: Optional[str] = None
    output_path: Optional[str] = None
    processed_files: int = 0
    extracted_records: int = 0
    unprocessed_files: int = 0
    error_files: int = 0
    lookup_file: Optional[str] = None
    lookup_missing: bool = False
    status: str = 'running'
    error_message: Optional[str] = None
    config_snapshot: Optional[str] = None
    input_files_hash: Optional[str] = None
    output_hash: Optional[str] = None
    started_at: Optional[str] = None
    completed_at: Optional[str] = None
    duration_ms: Optional[int] = None
    client_ip: Optional[str] = None
    client_hostname: Optional[str] = None
    signature_id: Optional[str] = None
    digital_signature: Optional[str] = None
    signature_algorithm: Optional[str] = None
    signed_at: Optional[str] = None


class AuditLogger:
    """
    审计日志核心类，提供上下文管理器支持。
    使用方式：
        with AuditLogger('pipeline', script_dir) as audit:
            audit.record_input(folder)
            ... 执行操作 ...
            audit.record_result(result)
    """

    def __init__(self, operation_type, script_dir=None, username=None):
        self.script_dir = script_dir or get_script_dir()
        self.db_path = get_audit_db_path(self.script_dir)
        init_audit_db(self.db_path)

        self.username = username or get_current_user()
        self.user_id = _ensure_user(self.username, self.db_path)

        self.audit_id = f"AUD{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:8].upper()}"
        self.operation_type = operation_type

        client_info = get_client_info()

        self.record = AuditRecord(
            audit_id=self.audit_id,
            username=self.username,
            operation_type=operation_type,
            status='running',
            started_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f'),
            client_ip=client_info.get('ip'),
            client_hostname=client_info.get('hostname'),
            config_snapshot=compute_config_snapshot(self.script_dir),
        )

        self.logger = get_logger()
        self._start_time = datetime.now()
        self._save_record()
        self.logger.info('审计记录已创建 [%s] 操作: %s, 用户: %s',
                         self.audit_id, operation_type, self.username)

    def _save_record(self):
        """保存审计记录到数据库"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()

        cursor.execute('SELECT id FROM audit_logs WHERE audit_id = ?', (self.audit_id,))
        exists = cursor.fetchone()

        if exists:
            cursor.execute('''
                UPDATE audit_logs SET
                    input_directory = ?, output_path = ?,
                    processed_files = ?, extracted_records = ?,
                    unprocessed_files = ?, error_files = ?,
                    lookup_file = ?, lookup_missing = ?,
                    status = ?, error_message = ?,
                    config_snapshot = ?, input_files_hash = ?,
                    output_hash = ?, started_at = ?,
                    completed_at = ?, duration_ms = ?,
                    client_ip = ?, client_hostname = ?,
                    signature_id = ?, digital_signature = ?,
                    signature_algorithm = ?, signed_at = ?
                WHERE audit_id = ?
            ''', (
                self.record.input_directory, self.record.output_path,
                self.record.processed_files, self.record.extracted_records,
                self.record.unprocessed_files, self.record.error_files,
                self.record.lookup_file, 1 if self.record.lookup_missing else 0,
                self.record.status, self.record.error_message,
                self.record.config_snapshot, self.record.input_files_hash,
                self.record.output_hash, self.record.started_at,
                self.record.completed_at, self.record.duration_ms,
                self.record.client_ip, self.record.client_hostname,
                self.record.signature_id, self.record.digital_signature,
                self.record.signature_algorithm, self.record.signed_at,
                self.audit_id,
            ))
        else:
            cursor.execute('''
                INSERT INTO audit_logs (
                    audit_id, user_id, username, operation_type,
                    input_directory, output_path,
                    processed_files, extracted_records,
                    unprocessed_files, error_files,
                    lookup_file, lookup_missing,
                    status, error_message,
                    config_snapshot, input_files_hash,
                    output_hash, started_at,
                    completed_at, duration_ms,
                    client_ip, client_hostname,
                    signature_id, digital_signature,
                    signature_algorithm, signed_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (
                self.audit_id, self.user_id, self.username, self.operation_type,
                self.record.input_directory, self.record.output_path,
                self.record.processed_files, self.record.extracted_records,
                self.record.unprocessed_files, self.record.error_files,
                self.record.lookup_file, 1 if self.record.lookup_missing else 0,
                self.record.status, self.record.error_message,
                self.record.config_snapshot, self.record.input_files_hash,
                self.record.output_hash, self.record.started_at,
                self.record.completed_at, self.record.duration_ms,
                self.record.client_ip, self.record.client_hostname,
                self.record.signature_id, self.record.digital_signature,
                self.record.signature_algorithm, self.record.signed_at,
            ))

        conn.commit()
        conn.close()

    def record_signature(self, signature_info):
        """
        记录数字签名信息。

        Args:
            signature_info: 签名信息字典（来自 sign_output_file）
        """
        if not signature_info:
            return

        signature_id = save_signature_record(
            signature_info,
            audit_id=self.audit_id,
            script_dir=self.script_dir,
            db_path=self.db_path
        )

        self.record.signature_id = signature_id
        self.record.digital_signature = signature_info.get('signature')
        self.record.signature_algorithm = signature_info.get('algorithm')
        self.record.signed_at = signature_info.get('signed_at')
        self._save_record()

        self.logger.info('审计记录 [%s] 已记录数字签名: %s', self.audit_id, signature_id)

    def record_input(self, input_directory, lookup_file=None):
        """记录输入信息"""
        self.record.input_directory = input_directory
        if lookup_file:
            self.record.lookup_file = lookup_file
        elif self.record.input_directory:
            self.record.lookup_file = find_lookup_file(self.script_dir)
        self._save_record()
        self.logger.debug('审计记录 [%s] 已记录输入目录: %s', self.audit_id, input_directory)

    def set_extra_info(self, extra_info):
        """记录额外信息到config_snapshot中"""
        try:
            if self.record.config_snapshot:
                snapshot = json.loads(self.record.config_snapshot)
            else:
                snapshot = {}
            snapshot.update(extra_info)
            self.record.config_snapshot = json.dumps(snapshot, ensure_ascii=False)
            self._save_record()
            self.logger.debug('审计记录 [%s] 已添加额外信息: %s', self.audit_id, extra_info)
        except Exception as e:
            self.logger.warning('添加额外信息失败: %s', e)

    def record_result(self, result):
        """
        根据 ProcessingResult 或 DiffResult 记录处理结果
        """
        if isinstance(result, ProcessingResult):
            self.record.processed_files = len(result.processed_files)
            self.record.extracted_records = len(result.all_rows)
            self.record.unprocessed_files = len(result.unprocessed_files)
            self.record.error_files = len(result.error_files)
            self.record.output_path = result.output_path
            self.record.lookup_missing = result.lookup_missing
            if result.output_path:
                self.record.output_hash = compute_file_hash(result.output_path)
            if result.signature_info:
                self.record_signature(result.signature_info)
        elif isinstance(result, DiffResult):
            self.record.extracted_records = result.added_count + result.deleted_count + result.changed_count
            self.record.output_path = result.output_path
            if result.output_path:
                self.record.output_hash = compute_file_hash(result.output_path)

        self._save_record()

    def record_success(self):
        """标记操作成功完成"""
        self.record.status = 'success'
        self.record.completed_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')
        duration = (datetime.now() - self._start_time).total_seconds() * 1000
        self.record.duration_ms = int(duration)
        self._save_record()
        self.logger.info('审计记录 [%s] 操作成功完成，耗时 %d ms', self.audit_id, self.record.duration_ms)

    def record_failure(self, error_message):
        """标记操作失败"""
        self.record.status = 'failed'
        self.record.error_message = str(error_message)
        self.record.completed_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')
        duration = (datetime.now() - self._start_time).total_seconds() * 1000
        self.record.duration_ms = int(duration)
        self._save_record()
        self.logger.error('审计记录 [%s] 操作失败: %s', self.audit_id, error_message)

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if exc_type is not None:
            self.record_failure(f'{exc_type.__name__}: {exc_val}')
        elif self.record.status == 'running':
            self.record_success()
        return False


def record_config_change(config_type, config_name, old_value, new_value,
                         change_reason='', script_dir=None, username=None,
                         change_details=None):
    """
    记录配置变更历史。
    适用于查找表更新、银行配置调整等场景。

    Args:
        config_type: 配置类型（如 'lookup_table', 'bank_config'）
        config_name: 配置名称（如 '主体查找表.xlsx'）
        old_value: 变更前的值
        new_value: 变更后的值
        change_reason: 变更原因说明
        script_dir: 脚本目录
        username: 操作用户（默认自动获取）
        change_details: 详细变更内容（字典或列表，描述具体变更项）

    Returns:
        change_id: 变更记录ID
    """
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)
    init_audit_db(db_path)

    username = username or get_current_user()
    user_id = _ensure_user(username, db_path)

    change_id = f"CFG{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"
    changed_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')

    old_hash = hashlib.sha256(str(old_value or '').encode('utf-8')).hexdigest()
    new_hash = hashlib.sha256(str(new_value or '').encode('utf-8')).hexdigest()

    old_str = json.dumps(old_value, ensure_ascii=False) if isinstance(old_value, (dict, list)) else str(old_value)
    new_str = json.dumps(new_value, ensure_ascii=False) if isinstance(new_value, (dict, list)) else str(new_value)

    details_str = None
    if change_details is not None:
        details_str = json.dumps(change_details, ensure_ascii=False)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO config_changes (
            change_id, user_id, username, config_type, config_name,
            old_value, new_value, old_hash, new_hash, change_reason, changed_at,
            change_details
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        change_id, user_id, username, config_type, config_name,
        old_str, new_str, old_hash, new_hash, change_reason, changed_at,
        details_str
    ))
    conn.commit()
    conn.close()

    logger = get_logger()
    if change_details:
        change_count = len(change_details) if isinstance(change_details, list) else 1
        logger.info('配置变更已记录 [%s] %s.%s 由 %s 修改，%d 项变更',
                    change_id, config_type, config_name, username, change_count)
    else:
        logger.info('配置变更已记录 [%s] %s.%s 由 %s 修改',
                    change_id, config_type, config_name, username)

    return change_id


def read_lookup_table_content(lookup_file):
    """
    读取查找表的完整内容（单元格级），返回结构化数据用于比对。

    返回格式：
    {
        'file_path': '...',
        'file_hash': '...',
        'sheet_name': '...',
        'headers': ['主体名称', '银行账号'],
        'rows': [
            {'row_num': 2, '主体名称': 'XX公司', '银行账号': '12345'},
            {'row_num': 3, '主体名称': 'YY公司', '银行账号': '67890'}
        ],
        'account_map': {'12345': {'row_num': 2, '主体名称': 'XX公司'}}
    }
    """
    logger = get_logger()
    if not lookup_file or not os.path.exists(lookup_file):
        return None

    file_hash = compute_file_hash(lookup_file)
    tmp_path = None
    try:
        wb, tmp_path = open_workbook_compat(lookup_file)
        ws = wb.active

        headers = []
        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=1, column=col).value
            headers.append(str(val) if val is not None else f'列{col}')

        rows = []
        account_map = {}

        for row_num in range(2, ws.max_row + 1):
            row_data = {'row_num': row_num}
            for col_idx, header in enumerate(headers):
                val = ws.cell(row=row_num, column=col_idx + 1).value
                row_data[header] = val

            account = row_data.get('银行账号') or row_data.get('账号')
            if account:
                account_key = _account_key(account)
                account_map[account_key] = row_data

            has_content = any(v is not None and str(v).strip() for v in row_data.values() if v != row_data['row_num'])
            if has_content:
                rows.append(row_data)

        wb.close()
        cleanup_temp_file(tmp_path)

        result = {
            'file_path': lookup_file,
            'file_hash': file_hash,
            'sheet_name': ws.title,
            'headers': headers,
            'rows': rows,
            'account_map': account_map,
        }

        logger.debug('读取查找表内容完成: %s (%d 条记录)', lookup_file, len(rows))
        return result

    except Exception as e:
        logger.error('读取查找表内容失败: %s', e)
        cleanup_temp_file(tmp_path)
        return None


def _save_lookup_snapshot(lookup_content, script_dir=None, username=None):
    """保存查找表快照到数据库"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)
    init_audit_db(db_path)

    username = username or get_current_user()
    _ensure_user(username, db_path)

    snapshot_id = f"SNP{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"
    created_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')

    content_json = json.dumps(lookup_content, ensure_ascii=False)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO lookup_snapshots (
            snapshot_id, lookup_file, file_hash, content_json,
            created_at, created_by
        ) VALUES (?, ?, ?, ?, ?, ?)
    ''', (
        snapshot_id,
        lookup_content.get('file_path', ''),
        lookup_content.get('file_hash', ''),
        content_json,
        created_at,
        username
    ))
    conn.commit()
    conn.close()

    logger = get_logger()
    logger.debug('查找表快照已保存: %s', snapshot_id)
    return snapshot_id


def _get_last_lookup_snapshot(lookup_file, script_dir=None):
    """获取指定查找表的最新快照"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)

    if not os.path.exists(db_path):
        return None

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('''
        SELECT * FROM lookup_snapshots
        WHERE lookup_file = ?
        ORDER BY created_at DESC
        LIMIT 1
    ''', (lookup_file,))
    row = cursor.fetchone()
    conn.close()

    if row:
        return {
            'snapshot_id': row['snapshot_id'],
            'lookup_file': row['lookup_file'],
            'file_hash': row['file_hash'],
            'content': json.loads(row['content_json']),
            'created_at': row['created_at'],
            'created_by': row['created_by'],
        }
    return None


def _diff_lookup_snapshots(old_content, new_content):
    """
    比对两个查找表快照，生成详细变更列表。

    返回格式：
    [
        {
            'change_type': 'add' | 'remove' | 'modify',
            'account': '银行账号',
            'row_num': 行号,
            'field': '变更字段',
            'old_value': '旧值',
            'new_value': '新值',
            'description': '可读描述'
        },
        ...
    ]
    """
    changes = []

    if old_content is None and new_content is None:
        return changes

    if old_content is None:
        for row in (new_content or {}).get('rows', []):
            account = row.get('银行账号') or row.get('账号') or ''
            changes.append({
                'change_type': 'add',
                'account': str(account),
                'row_num': row.get('row_num'),
                'field': None,
                'old_value': None,
                'new_value': {k: v for k, v in row.items() if k != 'row_num'},
                'description': f'新增账号「{account}」: {row.get("主体名称") or row.get("主体") or "未命名主体"}'
            })
        return changes

    if new_content is None:
        for row in old_content.get('rows', []):
            account = row.get('银行账号') or row.get('账号') or ''
            changes.append({
                'change_type': 'remove',
                'account': str(account),
                'row_num': row.get('row_num'),
                'field': None,
                'old_value': {k: v for k, v in row.items() if k != 'row_num'},
                'new_value': None,
                'description': f'删除账号「{account}」: {row.get("主体名称") or row.get("主体") or "未命名主体"}'
            })
        return changes

    old_map = old_content.get('account_map', {})
    new_map = new_content.get('account_map', {})

    all_keys = set(list(old_map.keys()) + list(new_map.keys()))

    for key in all_keys:
        old_row = old_map.get(key)
        new_row = new_map.get(key)

        if old_row is None and new_row is not None:
            account = new_row.get('银行账号') or new_row.get('账号') or key
            changes.append({
                'change_type': 'add',
                'account': str(account),
                'row_num': new_row.get('row_num'),
                'field': None,
                'old_value': None,
                'new_value': {k: v for k, v in new_row.items() if k != 'row_num'},
                'description': f'新增账号「{account}」: {new_row.get("主体名称") or new_row.get("主体") or "未命名主体"}'
            })
        elif old_row is not None and new_row is None:
            account = old_row.get('银行账号') or old_row.get('账号') or key
            changes.append({
                'change_type': 'remove',
                'account': str(account),
                'row_num': old_row.get('row_num'),
                'field': None,
                'old_value': {k: v for k, v in old_row.items() if k != 'row_num'},
                'new_value': None,
                'description': f'删除账号「{account}」: {old_row.get("主体名称") or old_row.get("主体") or "未命名主体"}'
            })
        else:
            for field in set(list(old_row.keys()) + list(new_row.keys())):
                if field == 'row_num':
                    continue
                old_val = old_row.get(field)
                new_val = new_row.get(field)

                old_str = '' if old_val is None else str(old_val).strip()
                new_str = '' if new_val is None else str(new_val).strip()

                if old_str != new_str:
                    account = old_row.get('银行账号') or old_row.get('账号') or key
                    changes.append({
                        'change_type': 'modify',
                        'account': str(account),
                        'row_num': new_row.get('row_num'),
                        'field': field,
                        'old_value': old_val,
                        'new_value': new_val,
                        'description': f'账号「{account}」的「{field}」变更: {old_val} → {new_val}'
                    })

    return changes


@dataclass
class ChangeDetectionResult:
    """变更检测结果"""
    has_changes: bool = False
    change_id: Optional[str] = None
    change_details: List[Dict[str, Any]] = field(default_factory=list)
    old_snapshot: Optional[Dict] = None
    new_snapshot: Optional[Dict] = None
    old_content: Optional[Dict] = None
    new_content: Optional[Dict] = None


def detect_and_record_lookup_change(script_dir=None, username=None,
                                    change_reason='自动检测到查找表变更'):
    """
    自动检测查找表是否发生变更，如果有变更则记录到配置变更历史。

    工作流程：
    1. 读取当前查找表内容
    2. 获取数据库中保存的最新快照
    3. 比对两者差异（文件哈希 + 内容比对）
    4. 如果有变更，记录详细变更并保存新快照
    5. 返回变更检测结果

    Args:
        script_dir: 脚本目录
        username: 操作用户（默认自动获取）
        change_reason: 变更原因说明

    Returns:
        ChangeDetectionResult: 变更检测结果
    """
    if script_dir is None:
        script_dir = get_script_dir()
    logger = get_logger()

    lookup_file = find_lookup_file(script_dir)
    if lookup_file is None:
        logger.debug('未找到查找表，跳过变更检测')
        return ChangeDetectionResult(has_changes=False)

    username = username or get_current_user()

    current_content = read_lookup_table_content(lookup_file)
    if current_content is None:
        logger.warning('无法读取当前查找表内容，跳过变更检测')
        return ChangeDetectionResult(has_changes=False)

    last_snapshot = _get_last_lookup_snapshot(lookup_file, script_dir)
    last_content = last_snapshot.get('content') if last_snapshot else None

    current_hash = current_content.get('file_hash')
    last_hash = last_content.get('file_hash') if last_content else None

    if last_content is None:
        logger.info('首次运行，保存查找表初始快照')
        _save_lookup_snapshot(current_content, script_dir, username)
        return ChangeDetectionResult(
            has_changes=False,
            new_snapshot={'snapshot_id': 'initial'},
            new_content=current_content
        )

    if current_hash == last_hash:
        logger.debug('查找表文件哈希未变化，无变更')
        return ChangeDetectionResult(
            has_changes=False,
            old_snapshot=last_snapshot,
            old_content=last_content,
            new_content=current_content
        )

    changes = _diff_lookup_snapshots(last_content, current_content)

    if not changes:
        logger.info('查找表哈希变化但内容无差异，可能是格式或元数据变更')
        _save_lookup_snapshot(current_content, script_dir, username)
        return ChangeDetectionResult(
            has_changes=False,
            old_snapshot=last_snapshot,
            old_content=last_content,
            new_content=current_content
        )

    change_id = record_config_change(
        config_type='lookup_table',
        config_name=os.path.basename(lookup_file),
        old_value=last_content,
        new_value=current_content,
        change_reason=change_reason,
        script_dir=script_dir,
        username=username,
        change_details=changes
    )

    _save_lookup_snapshot(current_content, script_dir, username)

    add_count = sum(1 for c in changes if c['change_type'] == 'add')
    remove_count = sum(1 for c in changes if c['change_type'] == 'remove')
    modify_count = sum(1 for c in changes if c['change_type'] == 'modify')

    logger.warning(
        '检测到查找表变更 [%s]: 新增 %d, 删除 %d, 修改 %d',
        change_id, add_count, remove_count, modify_count
    )

    return ChangeDetectionResult(
        has_changes=True,
        change_id=change_id,
        change_details=changes,
        old_snapshot=last_snapshot,
        new_content=current_content,
        old_content=last_content
    )


def manual_record_lookup_change(script_dir=None, username=None, change_reason=''):
    """
    手动触发查找表变更记录。
    用于用户明确修改了查找表后需要立即记录的场景。

    Args:
        script_dir: 脚本目录
        username: 操作用户
        change_reason: 变更原因

    Returns:
        ChangeDetectionResult: 变更检测结果
    """
    return detect_and_record_lookup_change(
        script_dir=script_dir,
        username=username,
        change_reason=change_reason or '手动记录查找表变更'
    )


# ──────────────────────────────────────────────
# 电子签章与数字签名模块
# ──────────────────────────────────────────────

SIGNATURE_KEY_FILENAME = 'bankcheck_signing_key.pem'
SIGNATURE_PUB_FILENAME = 'bankcheck_signing_key.pub.pem'
SIGNATURE_ALGORITHM = 'RSA-SHA256'


def get_signature_key_path(script_dir=None, public=False):
    """获取签名密钥文件路径"""
    if script_dir is None:
        script_dir = get_script_dir()
    filename = SIGNATURE_PUB_FILENAME if public else SIGNATURE_KEY_FILENAME
    return os.path.join(script_dir, filename)


def generate_signing_key_pair(script_dir=None, key_size=2048, password=None):
    """
    生成 RSA 密钥对用于数字签名。

    Args:
        script_dir: 脚本目录，用于存储密钥
        key_size: 密钥长度，默认 2048 位
        password: 私钥密码（可选）

    Returns:
        Tuple[str, str]: (私钥路径, 公钥路径)
    """
    if not HAS_CRYPTOGRAPHY:
        raise ImportError('cryptography 库未安装，无法使用数字签名功能')

    script_dir = script_dir or get_script_dir()
    private_key_path = get_signature_key_path(script_dir, public=False)
    public_key_path = get_signature_key_path(script_dir, public=True)

    private_key = rsa.generate_private_key(
        public_exponent=65537,
        key_size=key_size,
        backend=default_backend()
    )

    encryption_algorithm = serialization.NoEncryption()
    if password:
        encryption_algorithm = serialization.BestAvailableEncryption(password.encode('utf-8'))

    private_pem = private_key.private_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PrivateFormat.PKCS8,
        encryption_algorithm=encryption_algorithm
    )

    public_key = private_key.public_key()
    public_pem = public_key.public_bytes(
        encoding=serialization.Encoding.PEM,
        format=serialization.PublicFormat.SubjectPublicKeyInfo
    )

    with open(private_key_path, 'wb') as f:
        f.write(private_pem)

    with open(public_key_path, 'wb') as f:
        f.write(public_pem)

    os.chmod(private_key_path, 0o600)
    os.chmod(public_key_path, 0o644)

    logger = get_logger()
    logger.info('RSA 密钥对已生成: 私钥=%s, 公钥=%s', private_key_path, public_key_path)
    return private_key_path, public_key_path


def load_private_key(script_dir=None, password=None):
    """
    加载私钥用于签名。

    Args:
        script_dir: 脚本目录
        password: 私钥密码（如果有）

    Returns:
        private_key: RSA 私钥对象，失败返回 None
    """
    if not HAS_CRYPTOGRAPHY:
        return None

    private_key_path = get_signature_key_path(script_dir, public=False)
    if not os.path.exists(private_key_path):
        return None

    try:
        with open(private_key_path, 'rb') as f:
            key_data = f.read()

        password_bytes = password.encode('utf-8') if password else None
        private_key = serialization.load_pem_private_key(
            key_data,
            password=password_bytes,
            backend=default_backend()
        )
        return private_key
    except Exception as e:
        logger = get_logger()
        logger.error('加载私钥失败: %s', e)
        return None


def load_public_key(script_dir=None, public_key_path=None):
    """
    加载公钥用于验证签名。

    Args:
        script_dir: 脚本目录（用于查找默认公钥）
        public_key_path: 公钥文件路径（优先使用）

    Returns:
        public_key: RSA 公钥对象，失败返回 None
    """
    if not HAS_CRYPTOGRAPHY:
        return None

    if public_key_path is None:
        public_key_path = get_signature_key_path(script_dir, public=True)

    if not os.path.exists(public_key_path):
        return None

    try:
        with open(public_key_path, 'rb') as f:
            key_data = f.read()

        public_key = serialization.load_pem_public_key(
            key_data,
            backend=default_backend()
        )
        return public_key
    except Exception as e:
        logger = get_logger()
        logger.error('加载公钥失败: %s', e)
        return None


def build_signature_payload(file_hash, username, operation_time, input_directory, extra_data=None):
    """
    构建签名载荷数据。

    Args:
        file_hash: 文件 SHA256 哈希值
        username: 操作用户
        operation_time: 操作时间
        input_directory: 输入目录
        extra_data: 额外数据（可选）

    Returns:
        bytes: 签名载荷字节串
    """
    payload = {
        'file_hash': file_hash,
        'username': username,
        'operation_time': operation_time,
        'input_directory': input_directory,
        'algorithm': SIGNATURE_ALGORITHM,
    }
    if extra_data:
        payload['extra_data'] = extra_data

    payload_str = json.dumps(payload, sort_keys=True, ensure_ascii=False)
    return payload_str.encode('utf-8')


def sign_data(private_key, data_bytes):
    """
    使用私钥对数据进行数字签名。

    Args:
        private_key: RSA 私钥对象
        data_bytes: 待签名的数据字节串

    Returns:
        str: Base64 编码的签名字符串，失败返回 None
    """
    if not HAS_CRYPTOGRAPHY or private_key is None:
        return None

    try:
        import base64
        signature = private_key.sign(
            data_bytes,
            padding.PSS(
                mgf=padding.MGF1(hashes.SHA256()),
                salt_length=padding.PSS.MAX_LENGTH
            ),
            hashes.SHA256()
        )
        return base64.b64encode(signature).decode('ascii')
    except Exception as e:
        logger = get_logger()
        logger.error('数字签名失败: %s', e)
        return None


def verify_signature(public_key, signature_b64, data_bytes):
    """
    使用公钥验证数字签名。

    Args:
        public_key: RSA 公钥对象
        signature_b64: Base64 编码的签名字符串
        data_bytes: 原始数据字节串

    Returns:
        bool: 签名是否有效
    """
    if not HAS_CRYPTOGRAPHY or public_key is None or signature_b64 is None:
        return False

    try:
        import base64
        signature = base64.b64decode(signature_b64)
        public_key.verify(
            signature,
            data_bytes,
            padding.PSS(
                mgf=padding.MGF1(hashes.SHA256()),
                salt_length=padding.PSS.MAX_LENGTH
            ),
            hashes.SHA256()
        )
        return True
    except InvalidSignature:
        return False
    except Exception as e:
        logger = get_logger()
        logger.error('签名验证失败: %s', e)
        return False


def sign_output_file(output_path, script_dir=None, username=None, input_directory=None,
                     private_key=None, password=None, extra_data=None):
    """
    为输出文件生成数字签名。

    Args:
        output_path: 输出文件路径
        script_dir: 脚本目录
        username: 操作用户（默认自动获取）
        input_directory: 输入目录
        private_key: 私钥对象（可选，优先使用）
        password: 私钥密码（可选）
        extra_data: 额外数据（可选）

    Returns:
        Dict: 签名结果信息，包含 signature、payload 等
    """
    if not HAS_CRYPTOGRAPHY:
        return None

    if not output_path or not os.path.exists(output_path):
        return None

    file_hash = compute_file_hash(output_path)
    if not file_hash:
        return None

    username = username or get_current_user()
    operation_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')

    payload_bytes = build_signature_payload(
        file_hash=file_hash,
        username=username,
        operation_time=operation_time,
        input_directory=input_directory,
        extra_data=extra_data
    )

    if private_key is None:
        private_key = load_private_key(script_dir, password)

    if private_key is None:
        logger = get_logger()
        logger.warning('未找到签名私钥，跳过数字签名')
        return None

    signature_b64 = sign_data(private_key, payload_bytes)
    if not signature_b64:
        return None

    return {
        'file_path': output_path,
        'file_hash': file_hash,
        'username': username,
        'operation_time': operation_time,
        'input_directory': input_directory,
        'algorithm': SIGNATURE_ALGORITHM,
        'signature': signature_b64,
        'payload': payload_bytes.decode('utf-8'),
        'signed_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f'),
    }


def verify_output_file_signature(output_path, signature_info, script_dir=None, public_key=None):
    """
    验证输出文件的数字签名。

    Args:
        output_path: 输出文件路径
        signature_info: 签名信息字典（来自 sign_output_file 的返回值）
        script_dir: 脚本目录
        public_key: 公钥对象（可选，优先使用）

    Returns:
        bool: 签名是否有效
    """
    if not HAS_CRYPTOGRAPHY or not signature_info:
        return False

    current_hash = compute_file_hash(output_path)
    if not current_hash or current_hash != signature_info.get('file_hash'):
        return False

    payload_bytes = build_signature_payload(
        file_hash=signature_info.get('file_hash'),
        username=signature_info.get('username'),
        operation_time=signature_info.get('operation_time'),
        input_directory=signature_info.get('input_directory'),
        extra_data=json.loads(signature_info.get('payload', '{}')).get('extra_data')
    )

    if public_key is None:
        public_key = load_public_key(script_dir)

    return verify_signature(
        public_key=public_key,
        signature_b64=signature_info.get('signature'),
        data_bytes=payload_bytes
    )


def has_signing_key(script_dir=None):
    """检查是否存在签名密钥"""
    private_path = get_signature_key_path(script_dir, public=False)
    return os.path.exists(private_path)


def ensure_signing_key(script_dir=None, auto_generate=True, password=None):
    """
    确保签名密钥存在，不存在则自动生成。

    Args:
        script_dir: 脚本目录
        auto_generate: 是否自动生成
        password: 私钥密码（可选）

    Returns:
        bool: 密钥是否可用
    """
    if has_signing_key(script_dir):
        return True

    if not auto_generate or not HAS_CRYPTOGRAPHY:
        return False

    try:
        generate_signing_key_pair(script_dir, password=password)
        return True
    except Exception as e:
        logger = get_logger()
        logger.error('自动生成签名密钥失败: %s', e)
        return False


@dataclass
class SignatureRecord:
    """数字签名记录数据类"""
    signature_id: str
    audit_id: str
    file_path: str
    file_hash: str
    username: str
    operation_time: str
    input_directory: Optional[str] = None
    algorithm: str = SIGNATURE_ALGORITHM
    signature: Optional[str] = None
    payload: Optional[str] = None
    signed_at: Optional[str] = None
    public_key_path: Optional[str] = None
    is_verified: bool = False
    verified_at: Optional[str] = None


def _ensure_signature_tables(db_path):
    """确保数字签名相关表存在"""
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS digital_signatures (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            signature_id TEXT NOT NULL UNIQUE,
            audit_id TEXT,
            file_path TEXT NOT NULL,
            file_hash TEXT NOT NULL,
            username TEXT NOT NULL,
            operation_time TEXT NOT NULL,
            input_directory TEXT,
            algorithm TEXT NOT NULL,
            signature TEXT NOT NULL,
            payload TEXT,
            signed_at TEXT NOT NULL,
            public_key_path TEXT,
            is_verified INTEGER DEFAULT 0,
            verified_at TEXT,
            FOREIGN KEY (audit_id) REFERENCES audit_logs (audit_id)
        )
    ''')

    cursor.execute('CREATE INDEX IF NOT EXISTS idx_signatures_audit ON digital_signatures(audit_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_signatures_file ON digital_signatures(file_path)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_signatures_hash ON digital_signatures(file_hash)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_signatures_signed ON digital_signatures(signed_at)')

    conn.commit()
    conn.close()


def save_signature_record(signature_info, audit_id=None, script_dir=None, db_path=None):
    """
    保存或更新数字签名记录到数据库。

    如果 signature_info 中已有 signature_id，则更新该记录（主要用于添加 audit_id）；
    否则插入新记录。

    Args:
        signature_info: 签名信息字典（来自 sign_output_file）
        audit_id: 关联的审计记录 ID
        script_dir: 脚本目录
        db_path: 数据库路径（可选）

    Returns:
        str: signature_id，失败返回 None
    """
    if not signature_info:
        return None

    if db_path is None:
        db_path = get_audit_db_path(script_dir)

    init_audit_db(db_path)
    _ensure_signature_tables(db_path)

    existing_signature_id = signature_info.get('signature_id')
    signed_at = signature_info.get('signed_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f'))
    public_key_path = get_signature_key_path(script_dir, public=True)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    if existing_signature_id:
        cursor.execute('SELECT signature_id FROM digital_signatures WHERE signature_id = ?', (existing_signature_id,))
        exists = cursor.fetchone()
        if exists:
            cursor.execute('''
                UPDATE digital_signatures SET
                    audit_id = ?,
                    file_path = ?,
                    file_hash = ?,
                    username = ?,
                    operation_time = ?,
                    input_directory = ?,
                    algorithm = ?,
                    signature = ?,
                    payload = ?,
                    signed_at = ?,
                    public_key_path = ?
                WHERE signature_id = ?
            ''', (
                audit_id, signature_info.get('file_path'),
                signature_info.get('file_hash'), signature_info.get('username'),
                signature_info.get('operation_time'), signature_info.get('input_directory'),
                signature_info.get('algorithm'), signature_info.get('signature'),
                signature_info.get('payload'), signed_at, public_key_path,
                existing_signature_id
            ))
            conn.commit()
            conn.close()

            logger = get_logger()
            logger.info('数字签名记录已更新 [%s] 关联审计: %s', existing_signature_id, audit_id)
            return existing_signature_id

    signature_id = existing_signature_id or f"SIG{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:8].upper()}"

    cursor.execute('''
        INSERT INTO digital_signatures (
            signature_id, audit_id, file_path, file_hash, username,
            operation_time, input_directory, algorithm, signature, payload,
            signed_at, public_key_path
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        signature_id, audit_id, signature_info.get('file_path'),
        signature_info.get('file_hash'), signature_info.get('username'),
        signature_info.get('operation_time'), signature_info.get('input_directory'),
        signature_info.get('algorithm'), signature_info.get('signature'),
        signature_info.get('payload'), signed_at, public_key_path
    ))

    conn.commit()
    conn.close()

    logger = get_logger()
    logger.info('数字签名记录已保存 [%s] 文件: %s', signature_id, signature_info.get('file_path'))
    return signature_id


def query_signatures(script_dir=None, audit_id=None, signature_id=None, file_path=None,
                    file_hash=None, username=None, limit=100):
    """
    查询数字签名记录。

    Args:
        script_dir: 脚本目录
        audit_id: 按审计 ID 过滤
        signature_id: 按签名 ID 过滤
        file_path: 按文件路径过滤
        file_hash: 按文件哈希过滤
        username: 按用户名过滤
        limit: 返回记录数限制

    Returns:
        List[Dict]: 签名记录列表
    """
    db_path = get_audit_db_path(script_dir)
    if not os.path.exists(db_path):
        return []

    _ensure_signature_tables(db_path)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    query = 'SELECT * FROM digital_signatures WHERE 1=1'
    params = []

    if audit_id:
        query += ' AND audit_id = ?'
        params.append(audit_id)
    if signature_id:
        query += ' AND signature_id = ?'
        params.append(signature_id)
    if file_path:
        query += ' AND file_path = ?'
        params.append(file_path)
    if file_hash:
        query += ' AND file_hash = ?'
        params.append(file_hash)
    if username:
        query += ' AND username = ?'
        params.append(username)

    query += ' ORDER BY signed_at DESC LIMIT ?'
    params.append(limit)

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()

    results = []
    for row in rows:
        result = dict(row)
        result['is_verified'] = bool(result.get('is_verified', 0))
        results.append(result)

    return results


def verify_and_update_signature(signature_id, script_dir=None, public_key=None):
    """
    验证签名并更新验证状态。

    Args:
        signature_id: 签名记录 ID
        script_dir: 脚本目录
        public_key: 公钥对象（可选）

    Returns:
        bool: 签名是否有效
    """
    db_path = get_audit_db_path(script_dir)
    if not os.path.exists(db_path):
        return False

    _ensure_signature_tables(db_path)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute('SELECT * FROM digital_signatures WHERE signature_id = ?', (signature_id,))
    row = cursor.fetchone()
    if not row:
        conn.close()
        return False

    signature_info = dict(row)
    conn.close()

    is_valid = verify_output_file_signature(
        output_path=signature_info.get('file_path'),
        signature_info=signature_info,
        script_dir=script_dir,
        public_key=public_key
    )

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE digital_signatures SET is_verified = ?, verified_at = ?
        WHERE signature_id = ?
    ''', (
        1 if is_valid else 0,
        datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f') if is_valid else None,
        signature_id
    ))
    conn.commit()
    conn.close()

    return is_valid


def export_signature_manifest(output_path, script_dir=None, **query_kwargs):
    """
    导出签名清单到 JSON 文件。

    Args:
        output_path: 输出 JSON 文件路径
        script_dir: 脚本目录
        query_kwargs: 查询参数（同 query_signatures）

    Returns:
        str: 输出文件路径，失败返回 None
    """
    signatures = query_signatures(script_dir=script_dir, **query_kwargs)
    if not signatures:
        return None

    manifest = {
        'exported_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'signature_count': len(signatures),
        'algorithm': SIGNATURE_ALGORITHM,
        'public_key_path': get_signature_key_path(script_dir, public=True),
        'signatures': signatures
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)

    return output_path


def query_audit_logs(script_dir=None, username=None, operation_type=None,
                     start_date=None, end_date=None, status=None, limit=100,
                     has_signature=None, output_hash=None, signature_id=None):
    """
    查询审计日志记录。

    Args:
        script_dir: 脚本目录
        username: 按用户名过滤
        operation_type: 按操作类型过滤
        start_date: 开始日期 (YYYY-MM-DD)
        end_date: 结束日期 (YYYY-MM-DD)
        status: 按状态过滤
        limit: 返回记录数限制
        has_signature: 是否有数字签名（True/False/None）
        output_hash: 按输出文件哈希过滤
        signature_id: 按签名 ID 过滤

    Returns:
        List[Dict] 审计记录列表
    """
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)

    if not os.path.exists(db_path):
        return []

    _migrate_audit_db_columns(db_path)

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    query = 'SELECT * FROM audit_logs WHERE 1=1'
    params = []

    if username:
        query += ' AND username = ?'
        params.append(username)
    if operation_type:
        query += ' AND operation_type = ?'
        params.append(operation_type)
    if start_date:
        query += ' AND date(started_at) >= date(?)'
        params.append(start_date)
    if end_date:
        query += ' AND date(started_at) <= date(?)'
        params.append(end_date)
    if status:
        query += ' AND status = ?'
        params.append(status)
    if has_signature is not None:
        if has_signature:
            query += ' AND digital_signature IS NOT NULL'
        else:
            query += ' AND digital_signature IS NULL'
    if output_hash:
        query += ' AND output_hash = ?'
        params.append(output_hash)
    if signature_id:
        query += ' AND signature_id = ?'
        params.append(signature_id)

    query += ' ORDER BY started_at DESC LIMIT ?'
    params.append(limit)

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()

    results = []
    for row in rows:
        record = dict(row)
        record['lookup_missing'] = bool(record.get('lookup_missing', 0))
        results.append(record)

    return results


def query_config_changes(script_dir=None, config_type=None, config_name=None,
                         username=None, limit=100, parse_details=True):
    """
    查询配置变更历史。

    Args:
        script_dir: 脚本目录
        config_type: 按配置类型过滤
        config_name: 按配置名称过滤
        username: 按用户名过滤
        limit: 返回记录数限制
        parse_details: 是否解析 change_details JSON 字段

    Returns:
        List[Dict] 配置变更记录列表，包含 change_details 解析结果
    """
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)

    if not os.path.exists(db_path):
        return []

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    query = 'SELECT * FROM config_changes WHERE 1=1'
    params = []

    if config_type:
        query += ' AND config_type = ?'
        params.append(config_type)
    if config_name:
        query += ' AND config_name = ?'
        params.append(config_name)
    if username:
        query += ' AND username = ?'
        params.append(username)

    query += ' ORDER BY changed_at DESC LIMIT ?'
    params.append(limit)

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()

    result = []
    for row in rows:
        record = dict(row)
        if parse_details and record.get('change_details'):
            try:
                record['change_details'] = json.loads(record['change_details'])
            except (json.JSONDecodeError, TypeError):
                pass
        result.append(record)

    return result


def export_audit_logs(output_path, script_dir=None, **query_kwargs):
    """
    导出审计日志到 Excel 文件。
    """
    records = query_audit_logs(script_dir=script_dir, **query_kwargs)
    if not records:
        return None

    df = pd.DataFrame(records)
    columns_order = [
        'audit_id', 'username', 'operation_type', 'status',
        'input_directory', 'output_path',
        'processed_files', 'extracted_records',
        'unprocessed_files', 'error_files',
        'lookup_file', 'lookup_missing',
        'started_at', 'completed_at', 'duration_ms',
        'client_ip', 'client_hostname',
        'error_message', 'config_snapshot',
        'input_files_hash', 'output_hash',
        'signature_id', 'digital_signature',
        'signature_algorithm', 'signed_at',
        'id'
    ]
    for col in columns_order:
        if col not in df.columns:
            df[col] = None
    df = df[columns_order]

    df.to_excel(output_path, index=False, engine='openpyxl')
    logger = get_logger()
    logger.info('审计日志已导出到: %s (共 %d 条记录)', output_path, len(records))
    return output_path


def export_signatures(output_path, script_dir=None, **query_kwargs):
    """
    导出数字签名记录到 Excel 文件。

    Args:
        output_path: 输出文件路径
        script_dir: 脚本目录
        **query_kwargs: 查询参数（同 query_signatures）

    Returns:
        输出文件路径，失败返回 None
    """
    signatures = query_signatures(script_dir=script_dir, **query_kwargs)
    if not signatures:
        return None

    df = pd.DataFrame(signatures)
    columns_order = [
        'signature_id', 'audit_id', 'file_path', 'file_hash',
        'username', 'operation_time', 'input_directory',
        'algorithm', 'signature', 'payload',
        'signed_at', 'public_key_path',
        'is_verified', 'verified_at', 'id'
    ]
    for col in columns_order:
        if col not in df.columns:
            df[col] = None
    df = df[columns_order]

    df.to_excel(output_path, index=False, engine='openpyxl')
    logger = get_logger()
    logger.info('数字签名记录已导出到: %s (共 %d 条记录)', output_path, len(signatures))
    return output_path


def verify_file_by_path(file_path, script_dir=None, public_key=None):
    """
    根据文件路径验证文件完整性和数字签名。

    Args:
        file_path: 要验证的文件路径
        script_dir: 脚本目录
        public_key: 公钥对象（可选）

    Returns:
        Dict: 验证结果，包含 integrity_valid（完整性）、signature_valid（签名有效性）等信息
    """
    result = {
        'file_path': file_path,
        'exists': False,
        'integrity_valid': False,
        'signature_valid': False,
        'signature_record': None,
        'error': None,
    }

    if not file_path or not os.path.exists(file_path):
        result['error'] = '文件不存在'
        return result

    result['exists'] = True

    try:
        current_hash = compute_file_hash(file_path)
        signatures = query_signatures(script_dir=script_dir, file_path=file_path, limit=10)

        if not signatures:
            result['error'] = '未找到该文件的签名记录'
            return result

        latest_signature = signatures[0]
        result['signature_record'] = latest_signature

        result['integrity_valid'] = current_hash == latest_signature.get('file_hash')
        if not result['integrity_valid']:
            result['error'] = '文件内容已被篡改'
            return result

        result['signature_valid'] = verify_output_file_signature(
            output_path=file_path,
            signature_info=latest_signature,
            script_dir=script_dir,
            public_key=public_key
        )

        if not result['signature_valid']:
            result['error'] = '数字签名验证失败'

        return result

    except Exception as e:
        result['error'] = f'验证过程出错: {e}'
        logger = get_logger()
        logger.error('文件验证失败: %s', e, exc_info=True)
        return result


def export_config_changes(output_path, script_dir=None, expand_details=False, **query_kwargs):
    """
    导出配置变更历史到 Excel 文件。

    Args:
        output_path: 输出文件路径
        script_dir: 脚本目录
        expand_details: 是否展开详细变更记录为多行
        **query_kwargs: 查询参数

    Returns:
        输出文件路径
    """
    records = query_config_changes(script_dir=script_dir, **query_kwargs)
    if not records:
        return None

    if expand_details:
        expanded_rows = []
        for record in records:
            details = record.get('change_details')
            if isinstance(details, list) and details:
                for idx, detail in enumerate(details, 1):
                    row = record.copy()
                    row['detail_index'] = idx
                    row['change_type'] = detail.get('change_type', '')
                    row['account'] = detail.get('account', '')
                    row['row_num'] = detail.get('row_num', '')
                    row['field'] = detail.get('field', '')
                    row['old_value_detail'] = json.dumps(detail.get('old_value'), ensure_ascii=False) if isinstance(detail.get('old_value'), (dict, list)) else str(detail.get('old_value', ''))
                    row['new_value_detail'] = json.dumps(detail.get('new_value'), ensure_ascii=False) if isinstance(detail.get('new_value'), (dict, list)) else str(detail.get('new_value', ''))
                    row['description'] = detail.get('description', '')
                    expanded_rows.append(row)
            else:
                row = record.copy()
                row['detail_index'] = None
                row['change_type'] = ''
                row['account'] = ''
                row['row_num'] = ''
                row['field'] = ''
                row['old_value_detail'] = ''
                row['new_value_detail'] = ''
                row['description'] = ''
                expanded_rows.append(row)

        df = pd.DataFrame(expanded_rows)
        columns_order = [
            'change_id', 'username', 'config_type', 'config_name',
            'detail_index', 'change_type', 'account', 'row_num', 'field',
            'old_value_detail', 'new_value_detail', 'description',
            'change_reason', 'old_hash', 'new_hash', 'changed_at',
            'old_value', 'new_value', 'change_details', 'user_id', 'id'
        ]
    else:
        df = pd.DataFrame(records)
        if 'change_details' in df.columns:
            df['change_details'] = df['change_details'].apply(
                lambda x: json.dumps(x, ensure_ascii=False) if isinstance(x, (dict, list)) else str(x) if x is not None else ''
            )
        columns_order = [
            'change_id', 'username', 'config_type', 'config_name',
            'old_value', 'new_value', 'change_reason',
            'old_hash', 'new_hash', 'changed_at', 'change_details', 'user_id', 'id'
        ]

    for col in columns_order:
        if col not in df.columns:
            df[col] = None
    df = df[columns_order]

    df.to_excel(output_path, index=False, engine='openpyxl')
    logger = get_logger()
    logger.info('配置变更历史已导出到: %s (共 %d 条记录)', output_path, len(records))
    return output_path


# ──────────────────────────────────────────────
# 财务软件对接导出模块
# ──────────────────────────────────────────────

@dataclass
class StandardTransaction:
    """标准化交易记录，作为总表到各财务软件模板的中间格式"""
    transaction_date: Optional[datetime] = None
    voucher_date: Optional[datetime] = None
    voucher_number: str = ''
    summary: str = ''
    subject_code: str = ''
    subject_name: str = ''
    debit_amount: float = 0.0
    credit_amount: float = 0.0
    bank_account: str = ''
    bank_name: str = ''
    entity: str = ''
    counterparty: str = ''
    transaction_id: str = ''
    balance: float = 0.0
    direction: str = ''
    department: str = ''
    personnel: str = ''
    customer: str = ''
    supplier: str = ''
    project: str = ''
    attachment_count: int = 1
    prepared_by: str = ''
    reviewed_by: str = ''
    posted_by: str = ''
    remark: str = ''


@dataclass
class StandardVoucher:
    """标准化凭证，包含多条借贷分录"""
    voucher_date: Optional[datetime] = None
    voucher_number: str = ''
    voucher_type: str = '记'
    attachment_count: int = 1
    prepared_by: str = ''
    reviewed_by: str = ''
    posted_by: str = ''
    entries: List[StandardTransaction] = field(default_factory=list)
    source_transaction: Optional[dict] = None


FINANCIAL_EXPORT_TEMPLATES = {
    'yonyou_voucher': {
        'name': '用友凭证导入模板',
        'description': '用友U8/U9/NC系列财务软件凭证导入格式',
        'file_suffix': '_用友凭证导入',
    },
    'kingdee_voucher': {
        'name': '金蝶凭证导入模板',
        'description': '金蝶K3/KIS/EAS系列财务软件凭证导入格式',
        'file_suffix': '_金蝶凭证导入',
    },
    'bank_journal': {
        'name': '银行日记账模板',
        'description': '标准银行日记账格式，可直接导入财务软件',
        'file_suffix': '_银行日记账',
    },
}

DEFAULT_ACCOUNT_MAPPING = {
    'cash': {'code': '1001', 'name': '库存现金'},
    'bank_deposit': {'code': '1002', 'name': '银行存款'},
    'accounts_receivable': {'code': '1122', 'name': '应收账款'},
    'accounts_payable': {'code': '2202', 'name': '应付账款'},
    'operating_revenue': {'code': '6001', 'name': '主营业务收入'},
    'operating_cost': {'code': '6401', 'name': '主营业务成本'},
    'management_expense': {'code': '6602', 'name': '管理费用'},
    'sales_expense': {'code': '6601', 'name': '销售费用'},
    'financial_expense': {'code': '6603', 'name': '财务费用'},
}


def _normalize_date(value):
    """规范化日期为 datetime 对象"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        for fmt in ['%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d', '%Y-%m-%d %H:%M:%S', '%Y/%m/%d %H:%M:%S']:
            try:
                return datetime.strptime(value.strip(), fmt)
            except ValueError:
                continue
        try:
            return pd.to_datetime(value).to_pydatetime()
        except Exception:
            pass
    return None


def _format_date_for_export(dt, format_str='%Y-%m-%d'):
    """格式化日期用于导出"""
    if dt is None:
        return ''
    return dt.strftime(format_str)


def _format_amount(amount):
    """格式化金额，保留2位小数"""
    if amount is None:
        return 0.0
    try:
        return round(float(amount), 2)
    except (ValueError, TypeError):
        return 0.0


def total_to_standard_transactions(total_records, account_mapping=None, operator=''):
    """
    将总表数据转换为标准化交易记录列表。

    标准化层是财务导出的核心，它将银行流水的每一条记录转换为：
    1. 借方分录（收款时）或贷方分录（付款时）- 银行存款科目
    2. 对应的对方科目分录（需要根据摘要/对方户名智能判断）

    Args:
        total_records: 总表记录列表，每条包含银行流水字段
        account_mapping: 科目映射字典，可选
        operator: 制单人名称

    Returns:
        List[StandardVoucher]: 标准化凭证列表
    """
    logger = get_logger()

    if account_mapping is None:
        account_mapping = DEFAULT_ACCOUNT_MAPPING

    bank_account_info = account_mapping['bank_deposit']
    vouchers = []

    for idx, record in enumerate(total_records, 1):
        payment = to_float(record.get('付款'))
        receipt = to_float(record.get('收款'))
        trade_date = _normalize_date(record.get('交易日期'))
        voucher_num = f"记-{datetime.now().strftime('%Y%m')}-{idx:04d}"

        bank_entry = StandardTransaction(
            transaction_date=trade_date,
            voucher_date=trade_date,
            voucher_number=voucher_num,
            summary=str(record.get('摘要', '')) or '银行流水',
            subject_code=bank_account_info['code'],
            subject_name=bank_account_info['name'],
            bank_account=str(record.get('银行账号', '')),
            bank_name=str(record.get('银行', '')),
            entity=str(record.get('主体', '')),
            counterparty=str(record.get('对方户名', '')),
            transaction_id=str(record.get('交易流水号', '')),
            balance=_format_amount(record.get('余额')),
            prepared_by=operator,
            attachment_count=1,
        )

        counter_entry = StandardTransaction(
            transaction_date=trade_date,
            voucher_date=trade_date,
            voucher_number=voucher_num,
            summary=str(record.get('摘要', '')) or '银行流水',
            bank_account=str(record.get('银行账号', '')),
            bank_name=str(record.get('银行', '')),
            entity=str(record.get('主体', '')),
            counterparty=str(record.get('对方户名', '')),
            transaction_id=str(record.get('交易流水号', '')),
            prepared_by=operator,
            attachment_count=1,
        )

        if receipt and receipt > 0:
            bank_entry.debit_amount = _format_amount(receipt)
            bank_entry.direction = '借'

            counter_entry.credit_amount = _format_amount(receipt)
            summary_lower = str(record.get('摘要', '')).lower()
            counterparty_lower = str(record.get('对方户名', '')).lower()

            if any(k in summary_lower for k in ['收入', '销售', '货款', '营收', '主营业务']):
                counter_entry.subject_code = account_mapping['operating_revenue']['code']
                counter_entry.subject_name = account_mapping['operating_revenue']['name']
            elif any(k in summary_lower or k in counterparty_lower for k in ['客户', '应收']):
                counter_entry.subject_code = account_mapping['accounts_receivable']['code']
                counter_entry.subject_name = account_mapping['accounts_receivable']['name']
            else:
                counter_entry.subject_code = account_mapping['accounts_receivable']['code']
                counter_entry.subject_name = account_mapping['accounts_receivable']['name']

        elif payment and payment < 0:
            payment_abs = abs(payment)
            bank_entry.credit_amount = _format_amount(payment_abs)
            bank_entry.direction = '贷'

            counter_entry.debit_amount = _format_amount(payment_abs)
            summary_lower = str(record.get('摘要', '')).lower()
            counterparty_lower = str(record.get('对方户名', '')).lower()

            if any(k in summary_lower for k in ['成本', '主营成本']):
                counter_entry.subject_code = account_mapping['operating_cost']['code']
                counter_entry.subject_name = account_mapping['operating_cost']['name']
            elif any(k in summary_lower for k in ['管理费用', '办公费', '差旅费', '招待费', '工资']):
                counter_entry.subject_code = account_mapping['management_expense']['code']
                counter_entry.subject_name = account_mapping['management_expense']['name']
            elif any(k in summary_lower for k in ['销售费用', '广告费', '推广费']):
                counter_entry.subject_code = account_mapping['sales_expense']['code']
                counter_entry.subject_name = account_mapping['sales_expense']['name']
            elif any(k in summary_lower for k in ['财务费用', '手续费', '利息']):
                counter_entry.subject_code = account_mapping['financial_expense']['code']
                counter_entry.subject_name = account_mapping['financial_expense']['name']
            elif any(k in summary_lower or k in counterparty_lower for k in ['供应商', '应付', '采购', '货款']):
                counter_entry.subject_code = account_mapping['accounts_payable']['code']
                counter_entry.subject_name = account_mapping['accounts_payable']['name']
            else:
                counter_entry.subject_code = account_mapping['accounts_payable']['code']
                counter_entry.subject_name = account_mapping['accounts_payable']['name']
        else:
            continue

        voucher = StandardVoucher(
            voucher_date=trade_date,
            voucher_number=voucher_num,
            voucher_type='记',
            attachment_count=1,
            prepared_by=operator,
            entries=[bank_entry, counter_entry],
            source_transaction=record,
        )
        vouchers.append(voucher)

    logger.info('总表 %d 条记录转换为 %d 张标准化凭证', len(total_records), len(vouchers))
    return vouchers


def load_total_table(total_path):
    """加载总表数据"""
    logger = get_logger()

    if not total_path or not os.path.exists(total_path):
        logger.error('总表文件不存在: %s', total_path)
        return []

    try:
        df = pd.read_excel(total_path, engine='openpyxl')
        records = df.to_dict('records')
        logger.info('加载总表成功，共 %d 条记录', len(records))
        return records
    except Exception as e:
        logger.error('加载总表失败: %s', e, exc_info=True)
        return []


def export_yonyou_voucher(vouchers, output_path):
    """
    导出用友凭证导入格式。

    用友U8标准凭证导入格式包含以下字段：
    凭证类别字、凭证编号、凭证日期、附单据数、制单人、审核人、记账人、
    摘要、科目编码、科目名称、借方金额、贷方金额、
    部门编码、部门名称、个人编码、个人名称、客户编码、客户名称、
    供应商编码、供应商名称、项目编码、项目名称、银行账号、票据号
    """
    logger = get_logger()

    columns = [
        '凭证类别字', '凭证编号', '凭证日期', '附单据数',
        '制单人', '审核人', '记账人',
        '摘要', '科目编码', '科目名称',
        '借方金额', '贷方金额',
        '部门编码', '部门名称',
        '个人编码', '个人名称',
        '客户编码', '客户名称',
        '供应商编码', '供应商名称',
        '项目编码', '项目名称',
        '银行账号', '票据号',
    ]

    rows = []
    for voucher in vouchers:
        for entry in voucher.entries:
            rows.append({
                '凭证类别字': voucher.voucher_type,
                '凭证编号': voucher.voucher_number,
                '凭证日期': _format_date_for_export(voucher.voucher_date),
                '附单据数': voucher.attachment_count,
                '制单人': voucher.prepared_by,
                '审核人': voucher.reviewed_by,
                '记账人': voucher.posted_by,
                '摘要': entry.summary,
                '科目编码': str(entry.subject_code) if entry.subject_code is not None else '',
                '科目名称': entry.subject_name,
                '借方金额': entry.debit_amount,
                '贷方金额': entry.credit_amount,
                '部门编码': '',
                '部门名称': entry.department,
                '个人编码': '',
                '个人名称': entry.personnel,
                '客户编码': '',
                '客户名称': entry.customer,
                '供应商编码': '',
                '供应商名称': entry.supplier,
                '项目编码': '',
                '项目名称': entry.project,
                '银行账号': entry.bank_account,
                '票据号': entry.transaction_id,
            })

    if not rows:
        logger.warning('无凭证数据可导出')
        return None

    df = pd.DataFrame(rows, columns=columns)

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='凭证导入', index=False)

            ws = writer.sheets['凭证导入']
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        logger.info('用友凭证导出成功: %s (共 %d 张凭证，%d 条分录)',
                    output_path, len(vouchers), len(rows))
        return output_path
    except Exception as e:
        logger.error('导出用友凭证失败: %s', e, exc_info=True)
        return None


def export_kingdee_voucher(vouchers, output_path):
    """
    导出金蝶凭证导入格式。

    金蝶K3标准凭证导入格式包含以下字段：
    凭证字号、凭证日期、附件数、制单人、审核人、过账人、
    摘要、科目代码、科目名称、借方金额、贷方金额、
    核算项目类别、核算项目代码、核算项目名称、
    币别、汇率、原币金额、结算方式、结算号、业务日期
    """
    logger = get_logger()

    columns = [
        '凭证字号', '凭证日期', '附件数',
        '制单人', '审核人', '过账人',
        '摘要', '科目代码', '科目名称',
        '借方金额', '贷方金额',
        '核算项目类别', '核算项目代码', '核算项目名称',
        '币别', '汇率', '原币金额',
        '结算方式', '结算号', '业务日期',
    ]

    rows = []
    for voucher in vouchers:
        for entry in voucher.entries:
            rows.append({
                '凭证字号': voucher.voucher_number,
                '凭证日期': _format_date_for_export(voucher.voucher_date),
                '附件数': voucher.attachment_count,
                '制单人': voucher.prepared_by,
                '审核人': voucher.reviewed_by,
                '过账人': voucher.posted_by,
                '摘要': entry.summary,
                '科目代码': str(entry.subject_code) if entry.subject_code is not None else '',
                '科目名称': entry.subject_name,
                '借方金额': entry.debit_amount,
                '贷方金额': entry.credit_amount,
                '核算项目类别': '',
                '核算项目代码': '',
                '核算项目名称': '',
                '币别': '人民币',
                '汇率': 1.0,
                '原币金额': entry.debit_amount if entry.debit_amount else entry.credit_amount,
                '结算方式': '银行转账',
                '结算号': entry.transaction_id,
                '业务日期': _format_date_for_export(entry.transaction_date),
            })

    if not rows:
        logger.warning('无凭证数据可导出')
        return None

    df = pd.DataFrame(rows, columns=columns)

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='凭证导入', index=False)

            ws = writer.sheets['凭证导入']
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

        logger.info('金蝶凭证导出成功: %s (共 %d 张凭证，%d 条分录)',
                    output_path, len(vouchers), len(rows))
        return output_path
    except Exception as e:
        logger.error('导出金蝶凭证失败: %s', e, exc_info=True)
        return None


def export_bank_journal(vouchers, output_path):
    """
    导出银行日记账格式。

    标准银行日记账格式包含以下字段：
    日期、凭证号、摘要、对方科目、借方金额、贷方金额、方向、余额、
    银行账号、开户银行、核算主体、对方户名、交易流水号、备注
    """
    logger = get_logger()

    columns = [
        '日期', '凭证号', '摘要', '对方科目',
        '借方金额', '贷方金额', '方向', '余额',
        '银行账号', '开户银行', '核算主体',
        '对方户名', '交易流水号', '备注',
    ]

    rows = []
    for voucher in vouchers:
        bank_entry = None
        other_entry = None

        for entry in voucher.entries:
            if '银行' in entry.subject_name or entry.subject_code.startswith('1002'):
                bank_entry = entry
            else:
                other_entry = entry

        if bank_entry is None:
            continue

        rows.append({
            '日期': _format_date_for_export(bank_entry.transaction_date),
            '凭证号': voucher.voucher_number,
            '摘要': bank_entry.summary,
            '对方科目': other_entry.subject_name if other_entry else '',
            '借方金额': bank_entry.debit_amount,
            '贷方金额': bank_entry.credit_amount,
            '方向': bank_entry.direction,
            '余额': bank_entry.balance,
            '银行账号': bank_entry.bank_account,
            '开户银行': bank_entry.bank_name,
            '核算主体': bank_entry.entity,
            '对方户名': bank_entry.counterparty,
            '交易流水号': bank_entry.transaction_id,
            '备注': bank_entry.remark,
        })

    if not rows:
        logger.warning('无日记账数据可导出')
        return None

    df = pd.DataFrame(rows, columns=columns)

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='银行日记账', index=False)

            ws = writer.sheets['银行日记账']
            for col in ws.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)

            for row in ws.iter_rows(min_row=2):
                for cell in row:
                    if cell.column_letter in ['E', 'F', 'H']:
                        cell.number_format = '#,##0.00'

        logger.info('银行日记账导出成功: %s (共 %d 条记录)', output_path, len(rows))
        return output_path
    except Exception as e:
        logger.error('导出银行日记账失败: %s', e, exc_info=True)
        return None


def export_financial_template(total_path, template_type, output_dir=None,
                              account_mapping=None, operator=''):
    """
    标准化导出入口：从总表到财务软件模板的统一入口。

    这是在总表之上增加的一步标准化导出流程：
    总表数据 → 标准化转换 → 模板格式化 → 导出文件

    Args:
        total_path: 总表文件路径
        template_type: 模板类型 ('yonyou_voucher', 'kingdee_voucher', 'bank_journal')
        output_dir: 输出目录，默认为总表所在目录
        account_mapping: 科目映射字典，可选
        operator: 制单人名称

    Returns:
        dict: 导出结果，包含 output_path, voucher_count, entry_count 等信息
    """
    logger = get_logger()

    if template_type not in FINANCIAL_EXPORT_TEMPLATES:
        raise ValueError(f'不支持的导出模板类型: {template_type}')

    template_info = FINANCIAL_EXPORT_TEMPLATES[template_type]

    if output_dir is None:
        output_dir = os.path.dirname(total_path) or get_script_dir()

    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    base_name = os.path.splitext(os.path.basename(total_path))[0]
    output_filename = f'{base_name}{template_info["file_suffix"]}_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, output_filename)

    logger.info('开始财务导出: 模板=%s, 总表=%s, 输出=%s',
                template_info['name'], total_path, output_path)

    total_records = load_total_table(total_path)
    if not total_records:
        logger.error('总表无数据可导出')
        return {'success': False, 'error': '总表无数据', 'output_path': None}

    vouchers = total_to_standard_transactions(total_records, account_mapping, operator)
    if not vouchers:
        logger.error('无有效凭证可导出')
        return {'success': False, 'error': '无有效凭证', 'output_path': None}

    export_func = {
        'yonyou_voucher': export_yonyou_voucher,
        'kingdee_voucher': export_kingdee_voucher,
        'bank_journal': export_bank_journal,
    }[template_type]

    result = export_func(vouchers, output_path)

    if result:
        entry_count = sum(len(v.entries) for v in vouchers)
        return {
            'success': True,
            'output_path': result,
            'voucher_count': len(vouchers),
            'entry_count': entry_count,
            'template_type': template_type,
            'template_name': template_info['name'],
        }
    else:
        return {'success': False, 'error': '导出失败', 'output_path': None}


def cli_ask_export_template():
    """命令行模式下询问导出模板类型"""
    print('\n请选择导出模板：')
    for idx, (key, info) in enumerate(FINANCIAL_EXPORT_TEMPLATES.items(), 1):
        print(f'  {idx}) {info["name"]}')
        print(f'     {info["description"]}')

    choice = input('\n请输入选项（直接回车默认为 1 用友凭证）: ').strip()

    if not choice:
        return 'yonyou_voucher'

    if choice.isdigit():
        keys = list(FINANCIAL_EXPORT_TEMPLATES.keys())
        idx = int(choice) - 1
        if 0 <= idx < len(keys):
            return keys[idx]

    if choice in FINANCIAL_EXPORT_TEMPLATES:
        return choice

    return 'yonyou_voucher'


def run_export_flow(script_dir):
    """财务导出主流程"""
    logger = get_logger()
    logger.info('========== 财务软件对接导出开始 ==========')

    total_path = ask_file('请选择【银行流水总表】文件')
    if not total_path:
        show_info('提示', '未选择总表文件，程序退出。')
        logger.info('用户未选择总表文件，退出导出流程')
        return

    logger.info('用户选择总表文件: %s', total_path)

    template_type = cli_ask_export_template()
    template_info = FINANCIAL_EXPORT_TEMPLATES[template_type]
    logger.info('用户选择导出模板: %s', template_info['name'])

    operator = input('请输入制单人名称（可选，直接回车为空）: ').strip()

    output_dir = input(f'请输入输出目录（直接回车默认为总表所在目录）: ').strip()
    if not output_dir:
        output_dir = os.path.dirname(total_path) or script_dir

    result = export_financial_template(
        total_path=total_path,
        template_type=template_type,
        output_dir=output_dir,
        operator=operator,
    )

    if result['success']:
        msg = (
            f'导出完成！\n\n'
            f'导出模板：{result["template_name"]}\n'
            f'凭证张数：{result["voucher_count"]}\n'
            f'分录条数：{result["entry_count"]}\n'
            f'输出文件：{result["output_path"]}'
        )
        show_info('导出成功', msg)
        logger.info('财务导出完成: %s', result["output_path"])
    else:
        msg = f'导出失败：{result.get("error", "未知错误")}'
        show_warning('导出失败', msg)
        logger.error('财务导出失败: %s', result.get("error"))

    logger.info('========== 财务软件对接导出结束 ==========')


# ──────────────────────────────────────────────
# 文件处理详情记录
# ──────────────────────────────────────────────

def record_file_processing_detail(audit_id, file_path, file_name, bank=None,
                                  status='processing', error_message=None,
                                  record_count=0, processing_duration_ms=None,
                                  started_at=None, completed_at=None, script_dir=None):
    """记录单个文件的处理详情"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)
    init_audit_db(db_path)

    detail_id = f"FPD{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')

    file_size = None
    try:
        if os.path.exists(file_path):
            file_size = os.path.getsize(file_path)
    except Exception:
        pass

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO file_processing_details (
            detail_id, audit_id, file_path, file_name, bank, file_size,
            record_count, status, error_message, processing_duration_ms,
            started_at, completed_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        detail_id, audit_id, file_path, file_name, bank, file_size,
        record_count, status, error_message, processing_duration_ms,
        started_at or now, completed_at
    ))
    conn.commit()
    conn.close()
    return detail_id


def update_file_processing_detail(detail_id, status=None, error_message=None,
                                  record_count=None, processing_duration_ms=None,
                                  completed_at=None, script_dir=None):
    """更新文件处理详情"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    updates = []
    params = []
    if status is not None:
        updates.append('status = ?')
        params.append(status)
    if error_message is not None:
        updates.append('error_message = ?')
        params.append(error_message)
    if record_count is not None:
        updates.append('record_count = ?')
        params.append(record_count)
    if processing_duration_ms is not None:
        updates.append('processing_duration_ms = ?')
        params.append(processing_duration_ms)
    if completed_at is not None:
        updates.append('completed_at = ?')
        params.append(completed_at)

    if not updates:
        conn.close()
        return

    params.append(detail_id)
    query = f"UPDATE file_processing_details SET {', '.join(updates)} WHERE detail_id = ?"
    cursor.execute(query, params)
    conn.commit()
    conn.close()


# ──────────────────────────────────────────────
# 告警规则管理
# ──────────────────────────────────────────────

@dataclass
class AlertRule:
    rule_id: str
    rule_name: str
    rule_type: str
    metric: str
    operator: str
    threshold: float
    window_minutes: int = 60
    severity: str = 'warning'
    enabled: bool = True
    description: Optional[str] = None


DEFAULT_ALERT_RULES = [
    {
        'rule_name': '处理成功率过低',
        'rule_type': 'pipeline',
        'metric': 'success_rate',
        'operator': '<',
        'threshold': 80.0,
        'window_minutes': 60,
        'severity': 'critical',
        'description': '最近1小时处理成功率低于80%',
    },
    {
        'rule_name': '单次运行耗时过长',
        'rule_type': 'pipeline',
        'metric': 'duration_ms',
        'operator': '>',
        'threshold': 300000,
        'window_minutes': 1,
        'severity': 'warning',
        'description': '单次运行耗时超过5分钟',
    },
    {
        'rule_name': '错误文件过多',
        'rule_type': 'pipeline',
        'metric': 'error_files',
        'operator': '>',
        'threshold': 5,
        'window_minutes': 60,
        'severity': 'warning',
        'description': '最近1小时错误文件数超过5个',
    },
    {
        'rule_name': '无法识别文件过多',
        'rule_type': 'pipeline',
        'metric': 'unprocessed_files',
        'operator': '>',
        'threshold': 10,
        'window_minutes': 60,
        'severity': 'warning',
        'description': '最近1小时无法识别文件数超过10个',
    },
    {
        'rule_name': '连续运行失败',
        'rule_type': 'pipeline',
        'metric': 'consecutive_failures',
        'operator': '>=',
        'threshold': 3,
        'window_minutes': 120,
        'severity': 'critical',
        'description': '最近2小时内连续3次运行失败',
    },
    {
        'rule_name': '处理量突增',
        'rule_type': 'pipeline',
        'metric': 'processed_files',
        'operator': '>',
        'threshold': 100,
        'window_minutes': 60,
        'severity': 'info',
        'description': '最近1小时处理文件数超过100个',
    },
]


def init_default_alert_rules(script_dir=None, username=None):
    """初始化默认告警规则"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)
    init_audit_db(db_path)

    username = username or get_current_user()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')
    created_count = 0

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    for rule_def in DEFAULT_ALERT_RULES:
        cursor.execute('SELECT rule_id FROM alert_rules WHERE rule_name = ?',
                       (rule_def['rule_name'],))
        if cursor.fetchone():
            continue

        rule_id = f"RUL{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"
        cursor.execute('''
            INSERT INTO alert_rules (
                rule_id, rule_name, rule_type, metric, operator, threshold,
                window_minutes, severity, enabled, description, created_at, updated_at, created_by
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            rule_id, rule_def['rule_name'], rule_def['rule_type'], rule_def['metric'],
            rule_def['operator'], rule_def['threshold'], rule_def['window_minutes'],
            rule_def['severity'], 1, rule_def.get('description'),
            now, now, username
        ))
        created_count += 1

    conn.commit()
    conn.close()

    if created_count > 0:
        logger = get_logger()
        logger.info('已初始化 %d 条默认告警规则', created_count)
    return created_count


def query_alert_rules(script_dir=None, rule_type=None, enabled=None, limit=100):
    """查询告警规则"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)

    if not os.path.exists(db_path):
        return []

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    query = 'SELECT * FROM alert_rules WHERE 1=1'
    params = []

    if rule_type:
        query += ' AND rule_type = ?'
        params.append(rule_type)
    if enabled is not None:
        query += ' AND enabled = ?'
        params.append(1 if enabled else 0)

    query += ' ORDER BY created_at DESC LIMIT ?'
    params.append(limit)

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()

    return [dict(row) for row in rows]


def create_alert_rule(rule_name, rule_type, metric, operator, threshold,
                      window_minutes=60, severity='warning', description=None,
                      script_dir=None, username=None):
    """创建新的告警规则"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)
    init_audit_db(db_path)

    username = username or get_current_user()
    rule_id = f"RUL{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        INSERT INTO alert_rules (
            rule_id, rule_name, rule_type, metric, operator, threshold,
            window_minutes, severity, enabled, description, created_at, updated_at, created_by
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        rule_id, rule_name, rule_type, metric, operator, threshold,
        window_minutes, severity, 1, description, now, now, username
    ))
    conn.commit()
    conn.close()

    logger = get_logger()
    logger.info('已创建告警规则 [%s] %s', rule_id, rule_name)
    return rule_id


def toggle_alert_rule(rule_id, enabled, script_dir=None):
    """启用/禁用告警规则"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE alert_rules SET enabled = ?, updated_at = ? WHERE rule_id = ?
    ''', (1 if enabled else 0, datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f'), rule_id))
    conn.commit()
    conn.close()

    logger = get_logger()
    logger.info('告警规则 [%s] 已%s', rule_id, '启用' if enabled else '禁用')


def delete_alert_rule(rule_id, script_dir=None):
    """删除告警规则"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('DELETE FROM alert_rules WHERE rule_id = ?', (rule_id,))
    conn.commit()
    conn.close()

    logger = get_logger()
    logger.info('告警规则 [%s] 已删除', rule_id)


# ──────────────────────────────────────────────
# 告警检测引擎
# ──────────────────────────────────────────────

def _evaluate_condition(current_value, operator, threshold):
    """评估告警条件"""
    if current_value is None:
        return False

    ops = {
        '>': lambda a, b: a > b,
        '>=': lambda a, b: a >= b,
        '<': lambda a, b: a < b,
        '<=': lambda a, b: a <= b,
        '==': lambda a, b: a == b,
        '!=': lambda a, b: a != b,
    }

    func = ops.get(operator)
    if func is None:
        return False
    return func(current_value, threshold)


def _get_metric_value(metric, records):
    """从审计记录中计算指标值"""
    if not records:
        return None

    if metric == 'success_rate':
        total = len(records)
        success = sum(1 for r in records if r['status'] == 'success')
        return (success / total * 100) if total > 0 else 100.0

    if metric == 'duration_ms':
        latest = max(records, key=lambda r: r['started_at'] or '')
        return latest.get('duration_ms')

    if metric == 'error_files':
        return sum(r.get('error_files', 0) for r in records)

    if metric == 'unprocessed_files':
        return sum(r.get('unprocessed_files', 0) for r in records)

    if metric == 'processed_files':
        return sum(r.get('processed_files', 0) for r in records)

    if metric == 'consecutive_failures':
        count = 0
        for r in sorted(records, key=lambda x: x['started_at'] or '', reverse=True):
            if r['status'] == 'failed':
                count += 1
            else:
                break
        return count

    return None


def create_alert(rule, current_value, audit_log_id=None, details=None, script_dir=None):
    """创建告警记录"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)
    init_audit_db(db_path)

    alert_id = f"ALT{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:8].upper()}"
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')

    description = rule.get('description', '')
    if current_value is not None:
        description += f' 当前值: {current_value}, 阈值: {rule["operator"]}{rule["threshold"]}'

    details_json = json.dumps(details, ensure_ascii=False) if details else None

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute('''
        SELECT alert_id FROM alerts
        WHERE rule_id = ? AND status = 'active'
        ORDER BY triggered_at DESC LIMIT 1
    ''', (rule['rule_id'],))
    existing = cursor.fetchone()

    if existing:
        conn.close()
        return None

    cursor.execute('''
        INSERT INTO alerts (
            alert_id, rule_id, rule_name, severity, alert_type, metric,
            current_value, threshold, operator, window_minutes, description,
            triggered_at, status, audit_log_id, details
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
    ''', (
        alert_id, rule['rule_id'], rule['rule_name'], rule['severity'],
        rule['rule_type'], rule['metric'], current_value, rule['threshold'],
        rule['operator'], rule['window_minutes'], description,
        now, 'active', audit_log_id, details_json
    ))
    conn.commit()
    conn.close()

    logger = get_logger()
    logger.warning('告警触发 [%s] %s: %s', alert_id, rule['severity'].upper(),
                   rule['rule_name'])
    return alert_id


def run_alert_detection(script_dir=None, username=None):
    """运行告警检测，检查所有启用的规则"""
    if script_dir is None:
        script_dir = get_script_dir()
    logger = get_logger()

    init_default_alert_rules(script_dir, username)

    rules = query_alert_rules(script_dir=script_dir, enabled=True)
    if not rules:
        return []

    triggered_alerts = []

    for rule in rules:
        window_minutes = rule['window_minutes']
        from datetime import timedelta
        start_time = (datetime.now() - timedelta(minutes=window_minutes)).strftime('%Y-%m-%d %H:%M:%S')

        records = query_audit_logs(
            script_dir=script_dir,
            start_date=start_time.split(' ')[0],
            limit=1000
        )

        records_in_window = [
            r for r in records
            if r.get('started_at') and r['started_at'] >= start_time
        ]

        if not records_in_window:
            continue

        current_value = _get_metric_value(rule['metric'], records_in_window)
        if current_value is None:
            continue

        if _evaluate_condition(current_value, rule['operator'], rule['threshold']):
            latest_record = max(records_in_window, key=lambda r: r['started_at'] or '')
            alert_id = create_alert(
                rule, current_value,
                audit_log_id=latest_record.get('audit_id'),
                details={'records_analyzed': len(records_in_window)},
                script_dir=script_dir
            )
            if alert_id:
                triggered_alerts.append({
                    'alert_id': alert_id,
                    'rule': rule,
                    'current_value': current_value
                })

    if triggered_alerts:
        logger.warning('告警检测完成，触发 %d 条告警', len(triggered_alerts))
    else:
        logger.info('告警检测完成，无异常')

    return triggered_alerts


def query_alerts(script_dir=None, status=None, severity=None, limit=100):
    """查询告警记录"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)

    if not os.path.exists(db_path):
        return []

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    query = 'SELECT * FROM alerts WHERE 1=1'
    params = []

    if status:
        query += ' AND status = ?'
        params.append(status)
    if severity:
        query += ' AND severity = ?'
        params.append(severity)

    query += ' ORDER BY triggered_at DESC LIMIT ?'
    params.append(limit)

    cursor.execute(query, params)
    rows = cursor.fetchall()
    conn.close()

    result = []
    for row in rows:
        record = dict(row)
        if record.get('details'):
            try:
                record['details'] = json.loads(record['details'])
            except (json.JSONDecodeError, TypeError):
                pass
        result.append(record)
    return result


def acknowledge_alert(alert_id, script_dir=None, username=None):
    """确认告警"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)

    username = username or get_current_user()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE alerts SET acknowledged = 1, acknowledged_by = ?, acknowledged_at = ?
        WHERE alert_id = ?
    ''', (username, now, alert_id))
    conn.commit()
    conn.close()

    logger = get_logger()
    logger.info('告警 [%s] 已由 %s 确认', alert_id, username)


def resolve_alert(alert_id, script_dir=None, username=None):
    """解决告警"""
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)

    username = username or get_current_user()
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        UPDATE alerts SET status = 'resolved', resolved_at = ?,
                   acknowledged = 1, acknowledged_by = ?, acknowledged_at = ?
        WHERE alert_id = ?
    ''', (now, username, now, alert_id))
    conn.commit()
    conn.close()

    logger = get_logger()
    logger.info('告警 [%s] 已由 %s 标记为已解决', alert_id, username)


# ──────────────────────────────────────────────
# 监控统计分析
# ──────────────────────────────────────────────

@dataclass
class MonitorStats:
    total_runs: int = 0
    success_count: int = 0
    failed_count: int = 0
    success_rate: float = 0.0
    avg_duration_ms: float = 0.0
    max_duration_ms: int = 0
    min_duration_ms: int = 0
    total_processed_files: int = 0
    total_extracted_records: int = 0
    total_error_files: int = 0
    total_unprocessed_files: int = 0


def get_monitor_stats(script_dir=None, days=7):
    """获取指定天数内的监控统计数据"""
    from datetime import timedelta
    if script_dir is None:
        script_dir = get_script_dir()

    start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')
    records = query_audit_logs(
        script_dir=script_dir,
        start_date=start_date,
        limit=10000
    )

    if not records:
        return MonitorStats()

    durations = [r['duration_ms'] for r in records if r.get('duration_ms') is not None]

    stats = MonitorStats(
        total_runs=len(records),
        success_count=sum(1 for r in records if r['status'] == 'success'),
        failed_count=sum(1 for r in records if r['status'] == 'failed'),
        total_processed_files=sum(r.get('processed_files', 0) for r in records),
        total_extracted_records=sum(r.get('extracted_records', 0) for r in records),
        total_error_files=sum(r.get('error_files', 0) for r in records),
        total_unprocessed_files=sum(r.get('unprocessed_files', 0) for r in records),
    )

    if stats.total_runs > 0:
        stats.success_rate = (stats.success_count / stats.total_runs) * 100

    if durations:
        stats.avg_duration_ms = sum(durations) / len(durations)
        stats.max_duration_ms = max(durations)
        stats.min_duration_ms = min(durations)

    return stats


def get_daily_trend(script_dir=None, days=7):
    """获取每日运行趋势数据"""
    from datetime import timedelta
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)

    if not os.path.exists(db_path):
        return []

    start_date = (datetime.now() - timedelta(days=days - 1)).strftime('%Y-%m-%d')

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute('''
        SELECT
            date(started_at) as run_date,
            COUNT(*) as total_runs,
            SUM(CASE WHEN status = 'success' THEN 1 ELSE 0 END) as success_count,
            SUM(CASE WHEN status = 'failed' THEN 1 ELSE 0 END) as failed_count,
            AVG(duration_ms) as avg_duration_ms,
            SUM(processed_files) as total_processed_files,
            SUM(extracted_records) as total_extracted_records,
            SUM(error_files) as total_error_files,
            SUM(unprocessed_files) as total_unprocessed_files
        FROM audit_logs
        WHERE date(started_at) >= date(?)
        GROUP BY date(started_at)
        ORDER BY run_date DESC
    ''', (start_date,))

    rows = cursor.fetchall()
    conn.close()

    trend_data = []
    for row in rows:
        data = dict(row)
        total = data['total_runs'] or 0
        success = data['success_count'] or 0
        data['success_rate'] = (success / total * 100) if total > 0 else 0
        trend_data.append(data)

    return trend_data


def get_bank_distribution(script_dir=None, days=7):
    """获取各银行处理分布"""
    from datetime import timedelta
    if script_dir is None:
        script_dir = get_script_dir()
    db_path = get_audit_db_path(script_dir)

    if not os.path.exists(db_path):
        return []

    start_date = (datetime.now() - timedelta(days=days)).strftime('%Y-%m-%d')

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()

    cursor.execute('''
        SELECT
            fpd.bank,
            COUNT(*) as file_count,
            SUM(fpd.record_count) as total_records,
            AVG(fpd.processing_duration_ms) as avg_duration_ms,
            SUM(CASE WHEN fpd.status = 'success' THEN 1 ELSE 0 END) as success_count,
            SUM(CASE WHEN fpd.status = 'error' THEN 1 ELSE 0 END) as error_count
        FROM file_processing_details fpd
        JOIN audit_logs al ON fpd.audit_id = al.audit_id
        WHERE date(al.started_at) >= date(?) AND fpd.bank IS NOT NULL
        GROUP BY fpd.bank
        ORDER BY file_count DESC
    ''', (start_date,))

    rows = cursor.fetchall()
    conn.close()

    result = []
    for row in rows:
        data = dict(row)
        total = data['file_count'] or 0
        success = data['success_count'] or 0
        data['success_rate'] = (success / total * 100) if total > 0 else 0
        result.append(data)

    return result


def get_recent_runs(script_dir=None, limit=10):
    """获取最近运行记录列表"""
    records = query_audit_logs(script_dir=script_dir, limit=limit)
    return records


def get_active_alerts_summary(script_dir=None):
    """获取活跃告警摘要"""
    alerts = query_alerts(script_dir=script_dir, status='active', limit=100)
    summary = {
        'total': len(alerts),
        'critical': sum(1 for a in alerts if a['severity'] == 'critical'),
        'warning': sum(1 for a in alerts if a['severity'] == 'warning'),
        'info': sum(1 for a in alerts if a['severity'] == 'info'),
        'alerts': alerts
    }
    return summary


# ──────────────────────────────────────────────
# 运行监控面板
# ──────────────────────────────────────────────

def _format_duration(ms):
    """格式化毫秒为可读时间"""
    if ms is None:
        return 'N/A'
    seconds = ms / 1000
    if seconds < 60:
        return f'{seconds:.1f}秒'
    minutes = seconds / 60
    if minutes < 60:
        return f'{minutes:.1f}分钟'
    hours = minutes / 60
    return f'{hours:.2f}小时'


def _format_rate(rate):
    """格式化百分比"""
    if rate is None:
        return 'N/A'
    return f'{rate:.2f}%'


def _get_severity_color(severity):
    """获取告警严重度显示符号"""
    colors = {
        'critical': '🔴',
        'warning': '🟡',
        'info': '🔵',
    }
    return colors.get(severity, '⚪')


def _get_status_symbol(status):
    """获取状态显示符号"""
    symbols = {
        'success': '✅',
        'failed': '❌',
        'running': '⏳',
        'active': '🔴',
        'resolved': '✅',
        'acknowledged': '🟡',
    }
    return symbols.get(status, '⚪')


def render_monitor_dashboard(script_dir=None):
    """渲染监控面板主界面"""
    if script_dir is None:
        script_dir = get_script_dir()

    init_default_alert_rules(script_dir)

    stats = get_monitor_stats(script_dir, days=7)
    daily_trend = get_daily_trend(script_dir, days=7)
    recent_runs = get_recent_runs(script_dir, limit=5)
    alerts_summary = get_active_alerts_summary(script_dir)
    bank_dist = get_bank_distribution(script_dir, days=7)

    print('\n' + '=' * 80)
    print('📊 运行监控与告警面板'.center(70))
    print('=' * 80)
    print(f'统计周期: 最近7天')
    print()

    print('┌' + '─' * 78 + '┐')
    print('│  📈 运行概览' + ' ' * 65 + '│')
    print('├' + '─' * 78 + '┤')
    print(f'│  总运行次数: {stats.total_runs:<10}  |  成功: {stats.success_count:<10}  |  失败: {stats.failed_count:<10}  |  成功率: {_format_rate(stats.success_rate):>10} │')
    print(f'│  平均耗时: {_format_duration(stats.avg_duration_ms):<15} |  最大: {_format_duration(stats.max_duration_ms):<15} |  最小: {_format_duration(stats.min_duration_ms):<15} │')
    print(f'│  处理文件: {stats.total_processed_files:<10} |  提取记录: {stats.total_extracted_records:<10} |  错误文件: {stats.total_error_files:<10} |  未识别: {stats.total_unprocessed_files:<10} │')
    print('└' + '─' * 78 + '┘')
    print()

    if alerts_summary['total'] > 0:
        print('┌' + '─' * 78 + '┐')
        print(f'│  🚨 活跃告警 ({alerts_summary["total"]}条)' + ' ' * 56 + '│')
        print('├' + '─' * 78 + '┤')
        print(f'│  🔴 严重: {alerts_summary["critical"]:<5} | 🟡 警告: {alerts_summary["warning"]:<5} | 🔵 信息: {alerts_summary["info"]:<5}' + ' ' * 45 + '│')
        for alert in alerts_summary['alerts'][:3]:
            sev = _get_severity_color(alert['severity'])
            short_desc = alert.get('description', '')[:60]
            print(f'│  {sev} [{alert["alert_id"]}] {alert["rule_name"]}: {short_desc} │')
        if alerts_summary['total'] > 3:
            print(f'│  ... 还有 {alerts_summary["total"] - 3} 条告警，请查看详情' + ' ' * 38 + '│')
        print('└' + '─' * 78 + '┘')
        print()

    if daily_trend:
        print('┌' + '─' * 78 + '┐')
        print('│  📅 每日运行趋势' + ' ' * 63 + '│')
        print('├' + '─' * 78 + '┤')
        print(f'│  {"日期":<12} {"次数":<6} {"成功":<6} {"失败":<6} {"成功率":<9} {"处理":<6} {"错误":<6} {"未识别":<8} {"耗时":<10} │')
        print('│' + '─' * 78 + '│')
        for day in daily_trend[:7]:
            err = day.get('total_error_files', 0) or 0
            unproc = day.get('total_unprocessed_files', 0) or 0
            print(f'│  {day["run_date"]:<12} {day["total_runs"]:<6} {day["success_count"]:<6} {day["failed_count"]:<6} {_format_rate(day["success_rate"]):<9} {day["total_processed_files"]:<6} {err:<6} {unproc:<8} {_format_duration(day["avg_duration_ms"]):<10} │')
        print('└' + '─' * 78 + '┘')
        print()

    if bank_dist:
        print('┌' + '─' * 78 + '┐')
        print('│  🏦 银行处理分布' + ' ' * 63 + '│')
        print('├' + '─' * 78 + '┤')
        print(f'│  {"银行":<12} {"文件数":<8} {"记录数":<10} {"成功率":<10} {"平均耗时":<12} {"错误数":<8} │')
        print('│' + '─' * 78 + '│')
        for bank in bank_dist:
            name = bank.get('bank', '未知')[:10]
            print(f'│  {name:<12} {bank["file_count"]:<8} {bank["total_records"]:<10} {_format_rate(bank["success_rate"]):<10} {_format_duration(bank["avg_duration_ms"]):<12} {bank["error_count"]:<8} │')
        print('└' + '─' * 78 + '┘')
        print()

    if recent_runs:
        print('┌' + '─' * 78 + '┐')
        print('│  🕐 最近运行记录' + ' ' * 63 + '│')
        print('├' + '─' * 78 + '┤')
        print(f'│  {"审计ID":<14} {"时间":<20} {"类型":<10} {"状态":<8} {"耗时":<12} {"文件":<6} │')
        print('│' + '─' * 78 + '│')
        for run in recent_runs:
            status_sym = _get_status_symbol(run['status'])
            audit_id = run['audit_id'][:12]
            start_time = run.get('started_at', '')[:19]
            op_type = run.get('operation_type', '')[:8]
            duration = _format_duration(run.get('duration_ms'))
            files = run.get('processed_files', 0)
            print(f'│  {audit_id:<14} {start_time:<20} {op_type:<10} {status_sym} {run["status"]:<6} {duration:<12} {files:<6} │')
        print('└' + '─' * 78 + '┘')
    print()


def show_monitor_menu(script_dir=None):
    """显示监控面板菜单并处理用户选择"""
    if script_dir is None:
        script_dir = get_script_dir()
    logger = get_logger()

    while True:
        run_alert_detection(script_dir)
        render_monitor_dashboard(script_dir)

        print('\n请选择操作：')
        print('  1) 🔄 刷新面板')
        print('  2) 🚨 查看所有告警')
        print('  3) 📋 查看详细运行历史')
        print('  4) ⚙️  告警规则管理')
        print('  5) 📊 导出监控报告')
        print('  6) ✅ 确认/解决告警')
        print('  0) 返回主菜单')

        choice = input('\n请输入选项: ').strip()

        if choice == '0':
            break
        elif choice == '1':
            continue
        elif choice == '2':
            _show_all_alerts(script_dir)
        elif choice == '3':
            _show_run_history(script_dir)
        elif choice == '4':
            _show_alert_rules_menu(script_dir)
        elif choice == '5':
            _export_monitor_report(script_dir)
        elif choice == '6':
            _acknowledge_alert_menu(script_dir)
        else:
            print('无效选项，请重试')
            input('\n按回车继续...')


def _show_all_alerts(script_dir=None):
    """显示所有告警列表"""
    if script_dir is None:
        script_dir = get_script_dir()

    alerts = query_alerts(script_dir=script_dir, limit=50)

    print('\n' + '=' * 80)
    print('📋 告警记录列表'.center(70))
    print('=' * 80)
    print(f'{"告警ID":<16} {"时间":<20} {"严重度":<8} {"规则名称":<20} {"状态":<10}')
    print('-' * 80)

    for alert in alerts:
        sev = _get_severity_color(alert['severity'])
        status_sym = _get_status_symbol(alert['status'])
        alert_id = alert['alert_id'][:14]
        triggered = alert.get('triggered_at', '')[:19]
        rule_name = alert.get('rule_name', '')[:18]
        status = alert.get('status', '')
        print(f'{alert_id:<16} {triggered:<20} {sev} {alert["severity"]:<6} {rule_name:<20} {status_sym} {status:<8}')
        print(f'  → {alert.get("description", "")[:70]}')

    if not alerts:
        print('暂无告警记录')

    input('\n按回车继续...')


def _show_run_history(script_dir=None):
    """显示详细运行历史"""
    if script_dir is None:
        script_dir = get_script_dir()

    days_input = input('查看最近几天的记录 (默认7): ').strip()
    days = int(days_input) if days_input.isdigit() else 7

    records = query_audit_logs(script_dir=script_dir, limit=50)
    daily_trend = get_daily_trend(script_dir, days=days)

    print('\n' + '=' * 80)
    print(f'📊 运行历史记录 (最近{days}天)'.center(70))
    print('=' * 80)
    print(f'{"审计ID":<14} {"开始时间":<20} {"类型":<10} {"状态":<8} {"耗时":<10} {"处理":<6} {"错误":<6} {"未识别":<8} {"记录":<8}')
    print('-' * 80)

    for run in records:
        status_sym = _get_status_symbol(run['status'])
        audit_id = run['audit_id'][:12]
        start_time = run.get('started_at', '')[:19]
        op_type = run.get('operation_type', '')[:8]
        duration = _format_duration(run.get('duration_ms'))
        files = run.get('processed_files', 0)
        err = run.get('error_files', 0)
        unproc = run.get('unprocessed_files', 0)
        recs = run.get('extracted_records', 0)
        print(f'{audit_id:<14} {start_time:<20} {op_type:<10} {status_sym} {run["status"]:<6} {duration:<10} {files:<6} {err:<6} {unproc:<8} {recs:<8}')

    if daily_trend:
        print('\n' + '=' * 80)
        print('📈 处理量趋势图 (ASCII)'.center(70))
        print('=' * 80)
        max_files = max(d.get('total_processed_files', 0) for d in daily_trend) or 1
        for day in reversed(daily_trend):
            date = day['run_date']
            files = day.get('total_processed_files', 0)
            bar_len = int((files / max_files) * 40)
            bar = '█' * bar_len
            print(f'{date} | {bar:<40} | {files:>5} 文件')

        print('\n' + '=' * 80)
        print('🚨 错误文件趋势图 (ASCII)'.center(70))
        print('=' * 80)
        max_err = max((d.get('total_error_files', 0) or 0) for d in daily_trend) or 1
        for day in reversed(daily_trend):
            date = day['run_date']
            err = day.get('total_error_files', 0) or 0
            bar_len = int((err / max_err) * 40) if max_err > 0 else 0
            bar = '▓' * bar_len
            print(f'{date} | {bar:<40} | {err:>5} 错误')

        print('\n' + '=' * 80)
        print('⚠️  未识别文件趋势图 (ASCII)'.center(70))
        print('=' * 80)
        max_unproc = max((d.get('total_unprocessed_files', 0) or 0) for d in daily_trend) or 1
        for day in reversed(daily_trend):
            date = day['run_date']
            unproc = day.get('total_unprocessed_files', 0) or 0
            bar_len = int((unproc / max_unproc) * 40) if max_unproc > 0 else 0
            bar = '▒' * bar_len
            print(f'{date} | {bar:<40} | {unproc:>5} 未识别')

    input('\n按回车继续...')


def _show_alert_rules_menu(script_dir=None):
    """告警规则管理菜单"""
    if script_dir is None:
        script_dir = get_script_dir()

    while True:
        rules = query_alert_rules(script_dir=script_dir)

        print('\n' + '=' * 80)
        print('⚙️  告警规则管理'.center(70))
        print('=' * 80)
        print(f'{"规则ID":<12} {"名称":<20} {"指标":<16} {"条件":<16} {"严重度":<8} {"状态":<8}')
        print('-' * 80)

        for rule in rules:
            rid = rule['rule_id'][:10]
            name = rule['rule_name'][:18]
            metric = rule['metric'][:14]
            condition = f'{rule["operator"]}{rule["threshold"]}'
            enabled = '✅启用' if rule['enabled'] else '❌禁用'
            print(f'{rid:<12} {name:<20} {metric:<16} {condition:<16} {rule["severity"]:<8} {enabled:<8}')

        print('\n请选择操作：')
        print('  1) 启用/禁用规则')
        print('  2) 新增规则')
        print('  3) 删除规则')
        print('  4) 重置为默认规则')
        print('  0) 返回')

        choice = input('\n请输入选项: ').strip()

        if choice == '0':
            break
        elif choice == '1':
            rule_id = input('请输入规则ID: ').strip()
            rule = next((r for r in rules if r['rule_id'] == rule_id), None)
            if rule:
                new_state = not rule['enabled']
                toggle_alert_rule(rule_id, new_state, script_dir)
                print(f'规则已{"启用" if new_state else "禁用"}')
            else:
                print('未找到该规则')
        elif choice == '2':
            _create_new_rule_interactive(script_dir)
        elif choice == '3':
            rule_id = input('请输入要删除的规则ID: ').strip()
            confirm = input(f'确认删除规则 {rule_id}? (y/N): ').strip().lower()
            if confirm == 'y':
                delete_alert_rule(rule_id, script_dir)
                print('规则已删除')
        elif choice == '4':
            confirm = input('确认重置所有规则为默认? (y/N): ').strip().lower()
            if confirm == 'y':
                conn = sqlite3.connect(get_audit_db_path(script_dir))
                cursor = conn.cursor()
                cursor.execute('DELETE FROM alert_rules')
                conn.commit()
                conn.close()
                init_default_alert_rules(script_dir)
                print('已重置为默认规则')
        input('\n按回车继续...')


def _create_new_rule_interactive(script_dir=None):
    """交互式创建新告警规则"""
    print('\n--- 新建告警规则 ---')
    rule_name = input('规则名称: ').strip()
    if not rule_name:
        print('名称不能为空')
        return

    print('\n指标类型:')
    print('  success_rate - 成功率')
    print('  duration_ms - 运行耗时(ms)')
    print('  error_files - 错误文件数')
    print('  unprocessed_files - 未识别文件数')
    print('  processed_files - 处理文件数')
    print('  consecutive_failures - 连续失败次数')
    metric = input('请输入指标: ').strip()

    print('\n操作符: >, >=, <, <=, ==, !=')
    operator = input('请输入操作符: ').strip()

    threshold_input = input('阈值: ').strip()
    try:
        threshold = float(threshold_input)
    except ValueError:
        print('阈值必须是数字')
        return

    window_input = input('时间窗口(分钟，默认60): ').strip()
    window_minutes = int(window_input) if window_input.isdigit() else 60

    print('\n严重度: critical, warning, info')
    severity = input('严重度 (默认warning): ').strip() or 'warning'

    description = input('描述 (可选): ').strip() or None

    rule_id = create_alert_rule(
        rule_name=rule_name,
        rule_type='pipeline',
        metric=metric,
        operator=operator,
        threshold=threshold,
        window_minutes=window_minutes,
        severity=severity,
        description=description,
        script_dir=script_dir
    )
    print(f'规则已创建，ID: {rule_id}')


def _acknowledge_alert_menu(script_dir=None):
    """确认/解决告警菜单"""
    if script_dir is None:
        script_dir = get_script_dir()

    alerts = query_alerts(script_dir=script_dir, status='active', limit=50)

    if not alerts:
        print('\n没有活跃的告警需要处理')
        input('按回车继续...')
        return

    print('\n' + '=' * 80)
    print('✅ 告警处理'.center(70))
    print('=' * 80)
    print(f'{"告警ID":<16} {"时间":<20} {"严重度":<8} {"规则名称":<20}')
    print('-' * 80)

    for alert in alerts:
        sev = _get_severity_color(alert['severity'])
        alert_id = alert['alert_id'][:14]
        triggered = alert.get('triggered_at', '')[:19]
        rule_name = alert.get('rule_name', '')[:18]
        print(f'{alert_id:<16} {triggered:<20} {sev} {alert["severity"]:<6} {rule_name:<20}')
        print(f'  → {alert.get("description", "")[:70]}')

    alert_id = input('\n请输入要处理的告警ID: ').strip()
    if not alert_id:
        return

    alert = next((a for a in alerts if a['alert_id'] == alert_id), None)
    if not alert:
        print('未找到该告警')
        input('按回车继续...')
        return

    print('\n请选择操作：')
    print('  1) 确认告警')
    print('  2) 标记为已解决')
    action = input('请输入选项: ').strip()

    if action == '1':
        acknowledge_alert(alert_id, script_dir)
        print('告警已确认')
    elif action == '2':
        resolve_alert(alert_id, script_dir)
        print('告警已标记为已解决')

    input('按回车继续...')


def _export_monitor_report(script_dir=None):
    """导出监控报告到Excel"""
    if script_dir is None:
        script_dir = get_script_dir()
    logger = get_logger()

    days_input = input('导出最近几天的数据 (默认7): ').strip()
    days = int(days_input) if days_input.isdigit() else 7

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(script_dir, f'监控报告_{timestamp}.xlsx')

    stats = get_monitor_stats(script_dir, days=days)
    daily_trend = get_daily_trend(script_dir, days=days)
    bank_dist = get_bank_distribution(script_dir, days=days)
    alerts = query_alerts(script_dir=script_dir, limit=1000)
    runs = query_audit_logs(script_dir=script_dir, limit=1000)

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            summary_data = [{
                '统计项目': '总运行次数',
                '数值': stats.total_runs,
            }, {
                '统计项目': '成功次数',
                '数值': stats.success_count,
            }, {
                '统计项目': '失败次数',
                '数值': stats.failed_count,
            }, {
                '统计项目': '成功率(%)',
                '数值': round(stats.success_rate, 2),
            }, {
                '统计项目': '平均耗时(秒)',
                '数值': round(stats.avg_duration_ms / 1000, 2) if stats.avg_duration_ms else None,
            }, {
                '统计项目': '最大耗时(秒)',
                '数值': round(stats.max_duration_ms / 1000, 2) if stats.max_duration_ms else None,
            }, {
                '统计项目': '最小耗时(秒)',
                '数值': round(stats.min_duration_ms / 1000, 2) if stats.min_duration_ms else None,
            }, {
                '统计项目': '总处理文件数',
                '数值': stats.total_processed_files,
            }, {
                '统计项目': '总提取记录数',
                '数值': stats.total_extracted_records,
            }, {
                '统计项目': '总错误文件数',
                '数值': stats.total_error_files,
            }, {
                '统计项目': '总未识别文件数',
                '数值': stats.total_unprocessed_files,
            }, {
                '统计项目': '活跃告警数',
                '数值': sum(1 for a in alerts if a['status'] == 'active'),
            }, {
                '统计项目': '严重告警数',
                '数值': sum(1 for a in alerts if a['status'] == 'active' and a['severity'] == 'critical'),
            }]
            pd.DataFrame(summary_data).to_excel(writer, sheet_name='概览', index=False)

            if daily_trend:
                df_trend = pd.DataFrame(daily_trend)
                df_trend.to_excel(writer, sheet_name='每日趋势', index=False)

            if bank_dist:
                df_bank = pd.DataFrame(bank_dist)
                df_bank.to_excel(writer, sheet_name='银行分布', index=False)

            if runs:
                df_runs = pd.DataFrame(runs)
                df_runs.to_excel(writer, sheet_name='运行历史', index=False)

            if alerts:
                for a in alerts:
                    if isinstance(a.get('details'), (dict, list)):
                        a['details'] = json.dumps(a['details'], ensure_ascii=False)
                df_alerts = pd.DataFrame(alerts)
                df_alerts.to_excel(writer, sheet_name='告警记录', index=False)

        logger.info('监控报告已导出: %s', output_path)
        print(f'\n✅ 监控报告已导出到: {output_path}')
    except Exception as e:
        logger.error('导出监控报告失败: %s', e)
        print(f'\n❌ 导出失败: {e}')

    input('\n按回车继续...')


def run_monitor_flow(script_dir):
    """监控面板流程"""
    logger = get_logger()
    logger.info('========== 运行监控面板启动 ==========')

    init_default_alert_rules(script_dir)
    run_alert_detection(script_dir)
    show_monitor_menu(script_dir)

    logger.info('========== 运行监控面板关闭 ==========')


def run_pipeline_flow(script_dir):
    """主流程：处理银行流水文件夹，输出总表"""
    logger = get_logger()

    folder = ask_directory('请选择银行流水文件夹')
    if not folder:
        show_info('提示', '未选择文件夹，程序退出。')
        logger.info('用户未选择文件夹，程序退出')
        return

    logger.info('用户选择文件夹: %s', folder)

    change_result = detect_and_record_lookup_change(script_dir)
    if change_result.has_changes:
        add_count = sum(1 for c in change_result.change_details if c['change_type'] == 'add')
        remove_count = sum(1 for c in change_result.change_details if c['change_type'] == 'remove')
        modify_count = sum(1 for c in change_result.change_details if c['change_type'] == 'modify')
        warning_msg = (
            f'检测到查找表已变更！\n\n'
            f'变更编号: {change_result.change_id}\n'
            f'新增记录: {add_count} 条\n'
            f'删除记录: {remove_count} 条\n'
            f'修改记录: {modify_count} 条\n\n'
        )
        show_warning('查找表变更提醒', warning_msg)
        logger.warning('查找表变更已记录: %s', change_result.change_id)

    incremental = ask_incremental_mode()
    if incremental is None:
        logger.info('用户取消增量模式选择，返回主菜单')
        return
    logger.info('用户选择运行模式: %s', '增量合并' if incremental else '全量覆盖')

    dry_run = ask_dry_run_mode()
    logger.info('用户选择执行模式: %s', '试运行（不写盘）' if dry_run else '正式写盘')

    batch_id = None
    batch_manager = None
    if HAS_BATCH_MANAGER:
        try:
            batch_manager = batch_module.get_batch_manager(script_dir)
            operator = get_current_user()
            batch_info = batch_manager.start_batch(input_folder=folder, operator=operator)
            batch_id = batch_info.batch_id
            logger.info('批次管理已启用，批次号: %s', batch_id)
        except Exception as e:
            logger.error('批次创建失败: %s', e, exc_info=True)
            batch_manager = None

    with AuditLogger('pipeline', script_dir) as audit:
        audit.record_input(folder)

        result = run_pipeline(folder, script_dir, incremental=incremental,
                              batch_id=batch_id, dry_run=dry_run)

        if result.lookup_missing:
            show_warning(
                '警告',
                '在程序所在目录下未找到主体查找表文件，\n"主体"列将为空。\n'
                '建议将查找表文件命名为"主体查找表.xlsx"并放在程序所在目录下。'
            )

        msg = format_result_message(result)

        if result.dry_run and not result.folder_empty and result.all_rows:
            print('\n' + msg)
            if ask_commit_changes():
                logger.info('用户确认提交试运行变更')
                result = commit_pipeline_changes(result)
                msg = format_result_message(result)
            else:
                logger.info('用户取消提交试运行变更')
                msg += '\n\n⚠️  已取消提交，未执行任何删除与写盘操作。'

        audit.record_result(result)

        msg += f'\n\n审计编号: {audit.audit_id}'
        if change_result.change_id:
            msg += f'\n配置变更编号: {change_result.change_id}'

        if batch_manager and batch_id:
            try:
                log_file = get_current_log_file() or find_latest_log_file_in_dir(get_log_dir(script_dir))
                result_data = {
                    'total_records': len(result.all_rows),
                    'new_records': result.new_record_count,
                    'duplicate_records': result.duplicate_record_count,
                    'processed_files': result.processed_files,
                    'unprocessed_files': result.unprocessed_files,
                    'error_files': result.error_files,
                    'incremental_mode': result.incremental_mode,
                    'output_folder': folder,
                    'summary_table_path': result.output_path,
                    'log_file_path': log_file,
                    'audit_id': audit.audit_id,
                    'dry_run': result.dry_run,
                    'changes_committed': result.changes_committed,
                }
                status = 'success' if result.all_rows or result.existing_record_count > 0 else 'warning'
                if result.error_files:
                    status = 'warning'
                if result.dry_run and not result.changes_committed:
                    status = 'warning'
                batch_manager.finish_batch(batch_id, result_data, status=status)
                msg += f'\n批次号: {batch_id}'
                msg += f'\n归档目录: {batch_info.batch_dir}'
            except Exception as e:
                logger.error('批次归档失败: %s', e, exc_info=True)
                if batch_manager:
                    try:
                        batch_manager.finish_batch(batch_id, {}, status='failed', error_message=str(e))
                    except Exception:
                        pass

        title = '完成'
        if result.dry_run and not result.changes_committed:
            title = '试运行完成（未提交）'
        elif result.changes_committed:
            title = '提交完成'
        elif not result.all_rows:
            title = '提示'
        show_info(title, msg)


def run_diff_flow(script_dir):
    """变更对比流程：选择两次总表，输出对比结果"""
    logger = get_logger()

    old_path = ask_file('请选择【旧批次】银行流水总表')
    if not old_path:
        show_info('提示', '未选择旧批次文件，程序退出。')
        logger.info('用户未选择旧批次文件，程序退出')
        return
    logger.info('用户选择旧批次文件: %s', old_path)

    new_path = ask_file('请选择【新批次】银行流水总表')
    if not new_path:
        show_info('提示', '未选择新批次文件，程序退出。')
        logger.info('用户未选择新批次文件，程序退出')
        return
    logger.info('用户选择新批次文件: %s', new_path)

    change_result = detect_and_record_lookup_change(script_dir)
    if change_result.has_changes:
        add_count = sum(1 for c in change_result.change_details if c['change_type'] == 'add')
        remove_count = sum(1 for c in change_result.change_details if c['change_type'] == 'remove')
        modify_count = sum(1 for c in change_result.change_details if c['change_type'] == 'modify')
        warning_msg = (
            f'检测到查找表已变更！\n\n'
            f'变更编号: {change_result.change_id}\n'
            f'新增记录: {add_count} 条\n'
            f'删除记录: {remove_count} 条\n'
            f'修改记录: {modify_count} 条\n\n'
        )
        show_warning('查找表变更提醒', warning_msg)
        logger.warning('查找表变更已记录: %s', change_result.change_id)

    with AuditLogger('diff', script_dir) as audit:
        audit.record_input(f'旧:{old_path} | 新:{new_path}')

        try:
            diff_result = run_diff(old_path, new_path, script_dir)
            audit.record_result(diff_result)

            msg = format_diff_message(diff_result)

            has_changes = (
                diff_result.added_count > 0
                or diff_result.deleted_count > 0
                or diff_result.changed_count > 0
            )
            title = '对比完成（发现差异' if has_changes else '对比完成（无差异）'
            msg += f'\n\n审计编号: {audit.audit_id}'
            if change_result.change_id:
                msg += f'\n配置变更编号: {change_result.change_id}'
            show_info(title, msg)
        except FileNotFoundError as e:
            show_warning('错误', str(e))
            logger.error('对比失败: %s', e)


# ──────────────────────────────────────────────
# 定时批处理调度模块
# ──────────────────────────────────────────────

SCHEDULER_CONFIG_FILENAME = 'scheduler_config.json'
PROCESSED_FILES_DB_FILENAME = 'processed_files.db'
SCHEDULE_LOG_FILENAME = 'scheduler.log'


@dataclass
class ScheduleJobConfig:
    job_id: str
    name: str
    watch_directory: str
    cron_expression: str = '0 0 * * *'
    interval_minutes: Optional[int] = None
    schedule_type: str = 'cron'
    incremental: bool = True
    enabled: bool = True
    description: str = ''
    created_at: str = ''
    updated_at: str = ''


@dataclass
class ProcessedFileRecord:
    file_path: str
    file_hash: str
    file_size: int
    last_modified: float
    processed_at: str
    job_id: str
    record_count: int = 0
    status: str = 'success'


def get_scheduler_config_path(script_dir=None):
    if script_dir is None:
        script_dir = get_script_dir()
    return os.path.join(script_dir, SCHEDULER_CONFIG_FILENAME)


def get_processed_files_db_path(script_dir=None):
    if script_dir is None:
        script_dir = get_script_dir()
    return os.path.join(script_dir, PROCESSED_FILES_DB_FILENAME)


def get_scheduler_log_path(script_dir=None):
    if script_dir is None:
        script_dir = get_script_dir()
    return os.path.join(script_dir, SCHEDULE_LOG_FILENAME)


def init_processed_files_db(db_path=None):
    if db_path is None:
        db_path = get_processed_files_db_path()

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS processed_files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_path TEXT NOT NULL,
            file_hash TEXT NOT NULL,
            file_size INTEGER NOT NULL,
            last_modified REAL NOT NULL,
            processed_at TEXT NOT NULL,
            job_id TEXT NOT NULL,
            record_count INTEGER DEFAULT 0,
            status TEXT DEFAULT 'success',
            UNIQUE(file_path, file_hash)
        )
    ''')

    cursor.execute('''
        CREATE TABLE IF NOT EXISTS scheduler_runs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            run_id TEXT NOT NULL UNIQUE,
            job_id TEXT NOT NULL,
            job_name TEXT NOT NULL,
            started_at TEXT NOT NULL,
            completed_at TEXT,
            status TEXT NOT NULL,
            files_scanned INTEGER DEFAULT 0,
            files_new INTEGER DEFAULT 0,
            files_processed INTEGER DEFAULT 0,
            files_skipped INTEGER DEFAULT 0,
            files_error INTEGER DEFAULT 0,
            records_extracted INTEGER DEFAULT 0,
            error_message TEXT,
            output_path TEXT,
            duration_ms INTEGER
        )
    ''')

    cursor.execute('CREATE INDEX IF NOT EXISTS idx_processed_files_job ON processed_files(job_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_processed_files_path ON processed_files(file_path)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_scheduler_runs_job ON scheduler_runs(job_id)')
    cursor.execute('CREATE INDEX IF NOT EXISTS idx_scheduler_runs_started ON scheduler_runs(started_at)')

    conn.commit()
    conn.close()

    logger = get_logger()
    logger.debug('已处理文件数据库初始化完成: %s', db_path)


def load_scheduler_config(script_dir=None):
    config_path = get_scheduler_config_path(script_dir)
    logger = get_logger()

    if not os.path.exists(config_path):
        logger.info('调度配置文件不存在，将创建默认配置: %s', config_path)
        default_config = {
            'jobs': [],
            'settings': {
                'max_concurrent_jobs': 1,
                'check_interval_seconds': 30,
                'enable_alerts': True,
                'log_level': 'INFO'
            }
        }
        save_scheduler_config(default_config, script_dir)
        return default_config

    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        logger.info('已加载调度配置，共 %d 个任务', len(config.get('jobs', [])))
        return config
    except Exception as e:
        logger.error('加载调度配置失败: %s', e)
        return {'jobs': [], 'settings': {}}


def save_scheduler_config(config, script_dir=None):
    config_path = get_scheduler_config_path(script_dir)
    try:
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        logger = get_logger()
        logger.info('调度配置已保存: %s', config_path)
        return True
    except Exception as e:
        logger = get_logger()
        logger.error('保存调度配置失败: %s', e)
        return False


def add_schedule_job(job_config, script_dir=None):
    config = load_scheduler_config(script_dir)
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    if not job_config.get('job_id'):
        job_config['job_id'] = f"JOB{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"

    job_config['created_at'] = now
    job_config['updated_at'] = now
    job_config['enabled'] = job_config.get('enabled', True)

    config['jobs'].append(job_config)
    save_scheduler_config(config, script_dir)

    logger = get_logger()
    logger.info('已添加调度任务 [%s] %s', job_config['job_id'], job_config.get('name', ''))
    return job_config['job_id']


def update_schedule_job(job_id, updates, script_dir=None):
    config = load_scheduler_config(script_dir)
    logger = get_logger()

    for job in config['jobs']:
        if job['job_id'] == job_id:
            job.update(updates)
            job['updated_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            save_scheduler_config(config, script_dir)
            logger.info('已更新调度任务 [%s]', job_id)
            return True

    logger.warning('未找到调度任务 [%s]', job_id)
    return False


def remove_schedule_job(job_id, script_dir=None):
    config = load_scheduler_config(script_dir)
    logger = get_logger()

    config['jobs'] = [job for job in config['jobs'] if job['job_id'] != job_id]
    save_scheduler_config(config, script_dir)
    logger.info('已删除调度任务 [%s]', job_id)


def list_schedule_jobs(script_dir=None):
    config = load_scheduler_config(script_dir)
    return config.get('jobs', [])


def is_file_processed(file_path, job_id, db_path=None):
    if db_path is None:
        db_path = get_processed_files_db_path()

    file_hash = compute_file_hash(file_path)
    if not file_hash:
        return False, None

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT file_path, processed_at FROM processed_files
        WHERE file_path = ? AND file_hash = ? AND job_id = ?
    ''', (file_path, file_hash, job_id))
    result = cursor.fetchone()
    conn.close()

    return result is not None, file_hash


def mark_file_processed(file_path, file_hash, job_id, record_count=0, status='success', db_path=None):
    if db_path is None:
        db_path = get_processed_files_db_path()
    init_processed_files_db(db_path)

    file_size = os.path.getsize(file_path) if os.path.exists(file_path) else 0
    last_modified = os.path.getmtime(file_path) if os.path.exists(file_path) else 0
    processed_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT OR REPLACE INTO processed_files (
                file_path, file_hash, file_size, last_modified,
                processed_at, job_id, record_count, status
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (file_path, file_hash, file_size, last_modified,
              processed_at, job_id, record_count, status))
        conn.commit()
    finally:
        conn.close()

    logger = get_logger()
    logger.debug('已标记文件为已处理: %s (job=%s)', file_path, job_id)


def scan_new_files(watch_directory, job_id, script_dir=None):
    logger = get_logger()
    db_path = get_processed_files_db_path(script_dir)
    init_processed_files_db(db_path)

    all_files = scan_excel_files(watch_directory)
    new_files = []

    for file_path in all_files:
        processed, file_hash = is_file_processed(file_path, job_id, db_path)
        if not processed:
            new_files.append((file_path, file_hash))
            logger.debug('发现新文件: %s', file_path)

    logger.info('扫描目录 %s: 共 %d 个文件，新增 %d 个',
                watch_directory, len(all_files), len(new_files))
    return new_files


def record_scheduler_run(run_data, script_dir=None):
    db_path = get_processed_files_db_path(script_dir)
    init_processed_files_db(db_path)

    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO scheduler_runs (
                run_id, job_id, job_name, started_at, completed_at,
                status, files_scanned, files_new, files_processed,
                files_skipped, files_error, records_extracted,
                error_message, output_path, duration_ms
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            run_data['run_id'], run_data['job_id'], run_data['job_name'],
            run_data['started_at'], run_data.get('completed_at'),
            run_data['status'], run_data.get('files_scanned', 0),
            run_data.get('files_new', 0), run_data.get('files_processed', 0),
            run_data.get('files_skipped', 0), run_data.get('files_error', 0),
            run_data.get('records_extracted', 0), run_data.get('error_message'),
            run_data.get('output_path'), run_data.get('duration_ms')
        ))
        conn.commit()
    finally:
        conn.close()


def run_scheduled_pipeline(job_config, script_dir=None):
    if script_dir is None:
        script_dir = get_script_dir()

    logger = get_logger()
    job_id = job_config['job_id']
    job_name = job_config.get('name', job_id)
    watch_directory = job_config['watch_directory']
    incremental = job_config.get('incremental', True)

    logger.info('========== 定时任务启动 [%s] %s ==========', job_id, job_name)
    logger.info('监控目录: %s', watch_directory)
    logger.info('运行模式: %s', '增量合并' if incremental else '全量覆盖')

    run_id = f"SCH{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:8].upper()}"
    start_time = datetime.now()

    run_data = {
        'run_id': run_id,
        'job_id': job_id,
        'job_name': job_name,
        'started_at': start_time.strftime('%Y-%m-%d %H:%M:%S.%f'),
        'status': 'running',
    }

    try:
        change_result = detect_and_record_lookup_change(script_dir, username='scheduler')
        if change_result.has_changes:
            logger.warning('检测到查找表变更: %s', change_result.change_id)

        if not os.path.exists(watch_directory):
            raise FileNotFoundError(f'监控目录不存在: {watch_directory}')

        new_files = scan_new_files(watch_directory, job_id, script_dir)
        run_data['files_scanned'] = len(scan_excel_files(watch_directory))
        run_data['files_new'] = len(new_files)

        if not new_files:
            logger.info('没有新文件需要处理，任务结束')
            run_data['status'] = 'success'
            run_data['completed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')
            run_data['duration_ms'] = int((datetime.now() - start_time).total_seconds() * 1000)
            return run_data

        with AuditLogger('scheduled_pipeline', script_dir, username='scheduler') as audit:
            audit.record_input(watch_directory)

            result = run_pipeline(watch_directory, script_dir, incremental=incremental)
            audit.record_result(result)

            db_path = get_processed_files_db_path(script_dir)
            for file_path, file_hash in new_files:
                record_count = 0
                status = 'success'
                for proc_file in result.processed_files:
                    if os.path.basename(file_path) in proc_file:
                        record_count = len(result.all_rows)
                        break
                for err_file, _ in result.error_files:
                    if os.path.basename(file_path) in err_file:
                        status = 'error'
                        break

                mark_file_processed(file_path, file_hash, job_id, record_count, status, db_path)

            run_data['files_processed'] = len(result.processed_files)
            run_data['files_error'] = len(result.error_files)
            run_data['files_skipped'] = len(result.unprocessed_files)
            run_data['records_extracted'] = result.new_record_count
            run_data['output_path'] = result.output_path
            run_data['status'] = 'success' if not result.error_files else 'partial'

            if result.lookup_missing:
                logger.warning('未找到主体查找表，主体列为空')

            logger.info('定时任务完成: 新增文件 %d 个，处理 %d 个，错误 %d 个，提取记录 %d 条',
                        len(new_files), len(result.processed_files),
                        len(result.error_files), result.new_record_count)

            if result.output_path:
                logger.info('输出总表: %s', result.output_path)

    except Exception as e:
        logger.error('定时任务执行失败: %s', e, exc_info=True)
        run_data['status'] = 'failed'
        run_data['error_message'] = str(e)

    finally:
        run_data['completed_at'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')
        run_data['duration_ms'] = int((datetime.now() - start_time).total_seconds() * 1000)
        record_scheduler_run(run_data, script_dir)

        if run_data['status'] == 'failed' and load_scheduler_config(script_dir).get('settings', {}).get('enable_alerts', True):
            run_alert_detection(script_dir, username='scheduler')

    logger.info('========== 定时任务结束 [%s] 状态: %s 耗时: %dms ==========',
                job_id, run_data['status'], run_data['duration_ms'])

    return run_data


def _try_import_apscheduler():
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        from apscheduler.triggers.cron import CronTrigger
        from apscheduler.triggers.interval import IntervalTrigger
        return BackgroundScheduler, CronTrigger, IntervalTrigger
    except ImportError:
        logger = get_logger()
        logger.warning('未安装 APScheduler，将使用简单内置调度器。请运行: pip install APScheduler')
        return None, None, None


class SimpleScheduler:
    def __init__(self):
        self.jobs = []
        self.running = False
        self._thread = None
        self.logger = get_logger()

    def add_job(self, func, trigger, args=None, kwargs=None, id=None, name=None):
        job = {
            'id': id or f"JOB{uuid.uuid4().hex[:8].upper()}",
            'name': name or func.__name__,
            'func': func,
            'trigger': trigger,
            'args': args or [],
            'kwargs': kwargs or {},
            'last_run': None,
        }
        self.jobs.append(job)
        self.logger.info('已添加调度任务: %s (%s)', job['id'], job['name'])
        return job

    def _should_run(self, job, now):
        trigger = job['trigger']

        if job['last_run'] is None:
            return True

        if hasattr(trigger, 'get_interval'):
            interval_sec = trigger.get_interval().total_seconds()
            return (now - job['last_run']).total_seconds() >= interval_sec

        if hasattr(trigger, 'fields'):
            from datetime import datetime as dt
            last_run_date = job['last_run'].replace(second=0, microsecond=0)
            now_date = now.replace(second=0, microsecond=0)
            return now_date > last_run_date

        return False

    def _run_loop(self):
        import time
        while self.running:
            now = datetime.now()
            for job in self.jobs:
                if self._should_run(job, now):
                    try:
                        self.logger.info('触发定时任务: %s', job['name'])
                        job['last_run'] = now
                        job['func'](*job['args'], **job['kwargs'])
                    except Exception as e:
                        self.logger.error('任务执行错误 [%s]: %s', job['name'], e, exc_info=True)
            time.sleep(30)

    def start(self):
        import threading
        if self.running:
            return
        self.running = True
        self._thread = threading.Thread(target=self._run_loop, daemon=True)
        self._thread.start()
        self.logger.info('简单调度器已启动')

    def shutdown(self):
        self.running = False
        if self._thread:
            self._thread.join(timeout=5)
        self.logger.info('简单调度器已停止')


def get_scheduler():
    BackgroundScheduler, CronTrigger, IntervalTrigger = _try_import_apscheduler()

    if BackgroundScheduler:
        scheduler = BackgroundScheduler(timezone='Asia/Shanghai')
        scheduler_type = 'apscheduler'
    else:
        scheduler = SimpleScheduler()
        CronTrigger = None
        IntervalTrigger = None
        scheduler_type = 'simple'

    return scheduler, scheduler_type, CronTrigger, IntervalTrigger


def start_scheduler(script_dir=None):
    if script_dir is None:
        script_dir = get_script_dir()

    logger = get_logger()
    logger.info('========== 定时调度器启动 ==========')

    config = load_scheduler_config(script_dir)
    jobs = [j for j in config.get('jobs', []) if j.get('enabled', True)]

    if not jobs:
        logger.warning('没有启用的调度任务，请先配置任务')
        return None

    init_processed_files_db(get_processed_files_db_path(script_dir))

    scheduler, scheduler_type, CronTrigger, IntervalTrigger = get_scheduler()

    for job_config in jobs:
        job_id = job_config['job_id']
        job_name = job_config.get('name', job_id)
        schedule_type = job_config.get('schedule_type', 'cron')

        try:
            if scheduler_type == 'apscheduler':
                if schedule_type == 'interval' and job_config.get('interval_minutes'):
                    trigger = IntervalTrigger(minutes=job_config['interval_minutes'])
                else:
                    cron_expr = job_config.get('cron_expression', '0 0 * * *')
                    parts = cron_expr.split()
                    if len(parts) == 5:
                        minute, hour, day, month, day_of_week = parts
                        trigger = CronTrigger(
                            minute=minute, hour=hour, day=day,
                            month=month, day_of_week=day_of_week,
                            timezone='Asia/Shanghai'
                        )
                    else:
                        logger.warning('cron 表达式格式错误，使用默认每日0点: %s', cron_expr)
                        trigger = CronTrigger(hour=0, minute=0, timezone='Asia/Shanghai')

                scheduler.add_job(
                    run_scheduled_pipeline,
                    trigger=trigger,
                    args=[job_config, script_dir],
                    id=job_id,
                    name=job_name,
                    misfire_grace_time=3600,
                    coalesce=True,
                    max_instances=1,
                )
            else:
                class SimpleIntervalTrigger:
                    def __init__(self, minutes):
                        self._minutes = minutes
                    def get_interval(self):
                        from datetime import timedelta
                        return timedelta(minutes=self._minutes)

                class SimpleCronTrigger:
                    def __init__(self, cron_expr):
                        self.fields = cron_expr.split()

                if schedule_type == 'interval' and job_config.get('interval_minutes'):
                    trigger = SimpleIntervalTrigger(job_config['interval_minutes'])
                else:
                    trigger = SimpleCronTrigger(job_config.get('cron_expression', '0 0 * * *'))

                scheduler.add_job(
                    run_scheduled_pipeline,
                    trigger=trigger,
                    args=[job_config, script_dir],
                    id=job_id,
                    name=job_name,
                )

            logger.info('已注册任务 [%s] %s: 类型=%s, 配置=%s',
                        job_id, job_name, schedule_type,
                        job_config.get('cron_expression') or f"{job_config.get('interval_minutes')}分钟")

        except Exception as e:
            logger.error('注册任务失败 [%s]: %s', job_id, e)

    try:
        scheduler.start()
        logger.info('调度器已启动 (%s)，共 %d 个任务', scheduler_type, len(jobs))
        logger.info('按 Ctrl+C 停止调度器')

        if scheduler_type == 'apscheduler':
            import time
            try:
                while True:
                    time.sleep(60)
            except (KeyboardInterrupt, SystemExit):
                logger.info('收到停止信号')
                scheduler.shutdown()
        else:
            import time
            try:
                while scheduler.running:
                    time.sleep(60)
            except (KeyboardInterrupt, SystemExit):
                logger.info('收到停止信号')
                scheduler.shutdown()

    except Exception as e:
        logger.error('调度器运行错误: %s', e)
        scheduler.shutdown()

    logger.info('========== 定时调度器停止 ==========')
    return scheduler


def generate_cron_script(job_config, script_path, script_dir=None):
    if script_dir is None:
        script_dir = get_script_dir()

    logger = get_logger()
    job_id = job_config['job_id']
    cron_expr = job_config.get('cron_expression', '0 0 * * *')

    python_path = sys.executable
    script_abs_path = os.path.abspath(script_path)

    log_path = get_scheduler_log_path(script_dir)

    cron_line = f'{cron_expr} {python_path} {script_abs_path} --run-job {job_id} >> {log_path} 2>&1'

    script_content = f'''#!/bin/bash
# 银行流水检验工具 - 定时任务 cron 配置
# 任务ID: {job_id}
# 任务名称: {job_config.get('name', '')}
# 监控目录: {job_config.get('watch_directory', '')}
# 调度时间: {cron_expr}
# 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

# 添加到 crontab 的命令:
# crontab -e
# 然后添加以下行:

{cron_line}

# 或者运行以下命令自动添加:
# (crontab -l 2>/dev/null; echo "{cron_line}") | crontab -

# 查看当前 crontab:
# crontab -l

# 日志文件: {log_path}
'''

    output_path = os.path.join(script_dir, f'cron_job_{job_id}.sh')
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write(script_content)

    os.chmod(output_path, 0o755)
    logger.info('已生成 cron 脚本: %s', output_path)
    return output_path, cron_line


def generate_windows_task_script(job_config, script_path, script_dir=None):
    if script_dir is None:
        script_dir = get_script_dir()

    logger = get_logger()
    job_id = job_config['job_id']
    job_name = job_config.get('name', f'bankcheck_{job_id}')
    cron_expr = job_config.get('cron_expression', '0 0 * * *')

    parts = cron_expr.split()
    start_time = '00:00'
    if len(parts) >= 2:
        start_time = f'{parts[1].zfill(2)}:{parts[0].zfill(2)}'

    python_path = sys.executable
    script_abs_path = os.path.abspath(script_path)
    log_path = get_scheduler_log_path(script_dir)

    schtasks_cmd = (
        f'SchTasks /Create /SC DAILY /TN "{job_name}" '
        f'/TR "\\"{python_path}\\" \\"{script_abs_path}\\" --run-job {job_id}" '
        f'/ST {start_time} /RL HIGHEST /F'
    )

    ps_content = f'''# 银行流水检验工具 - Windows 任务计划配置
# 任务ID: {job_id}
# 任务名称: {job_name}
# 监控目录: {job_config.get('watch_directory', '')}
# 调度时间: {cron_expr} (每日 {start_time})
# 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

# 方法1: 以管理员身份运行此 PowerShell 脚本
$action = New-ScheduledTaskAction -Execute "{python_path}" -Argument "`"{script_abs_path}`" --run-job {job_id}"
$trigger = New-ScheduledTaskTrigger -Daily -At {start_time}
$settings = New-ScheduledTaskSettingsSet -StartWhenAvailable -AllowStartIfOnBatteries -DontStopIfGoingOnBatteries
Register-ScheduledTask -TaskName "{job_name}" -Action $action -Trigger $trigger -Settings $settings -Description "银行流水检验工具定时任务" -Force
Write-Host "任务已创建: {job_name}"

# 方法2: 在命令行(cmd)中运行以下命令:
# {schtasks_cmd}

# 查看任务:
# Get-ScheduledTask -TaskName "{job_name}"

# 删除任务:
# Unregister-ScheduledTask -TaskName "{job_name}" -Confirm:$false

# 日志文件: {log_path}
'''

    bat_content = f'''@echo off
REM 银行流水检验工具 - Windows 任务计划配置
REM 任务ID: {job_id}
REM 任务名称: {job_name}
REM 监控目录: {job_config.get('watch_directory', '')}
REM 调度时间: {cron_expr} (每日 {start_time})
REM 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

echo 创建 Windows 任务计划: {job_name}
{schtasks_cmd}

if %errorlevel%==0 (
    echo 任务创建成功！
    echo 查看任务: schtasks /Query /TN "{job_name}"
    echo 删除任务: schtasks /Delete /TN "{job_name}" /F
) else (
    echo 任务创建失败，请以管理员身份运行。
)

pause
'''

    ps_output = os.path.join(script_dir, f'task_job_{job_id}.ps1')
    bat_output = os.path.join(script_dir, f'task_job_{job_id}.bat')

    with open(ps_output, 'w', encoding='utf-8') as f:
        f.write(ps_content)

    with open(bat_output, 'w', encoding='gbk') as f:
        f.write(bat_content)

    logger.info('已生成 Windows 任务计划脚本: %s, %s', ps_output, bat_output)
    return ps_output, bat_output


def show_scheduler_menu(script_dir=None):
    if script_dir is None:
        script_dir = get_script_dir()
    logger = get_logger()

    while True:
        jobs = list_schedule_jobs(script_dir)

        print('\n' + '=' * 80)
        print('⏰ 定时批处理调度管理'.center(70))
        print('=' * 80)

        if jobs:
            print(f'{"ID":<14} {"名称":<20} {"类型":<10} {"调度":<20} {"状态":<8} {"目录":<30}')
            print('-' * 80)
            for job in jobs:
                status = '✅启用' if job.get('enabled', True) else '❌禁用'
                sched_type = job.get('schedule_type', 'cron')
                if sched_type == 'interval':
                    sched_display = f"每{job.get('interval_minutes', 60)}分钟"
                else:
                    sched_display = job.get('cron_expression', '0 0 * * *')
                watch_dir = job.get('watch_directory', '')[:28]
                print(f'{job["job_id"][:12]:<14} {job.get("name", "")[:18]:<20} {sched_type:<10} {sched_display[:18]:<20} {status:<8} {watch_dir:<30}')
        else:
            print('暂无调度任务，请先添加任务')

        print('\n请选择操作：')
        print('  1) ➕ 添加定时任务')
        print('  2) ✏️  编辑任务')
        print('  3) 🗑️  删除任务')
        print('  4) ▶️  启动调度器')
        print('  5) 🏃  立即运行指定任务')
        print('  6) 📜  查看任务运行历史')
        print('  7) 📄  生成系统任务脚本')
        print('  8) ⚙️  调度器设置')
        print('  0) 返回主菜单')

        choice = input('\n请输入选项: ').strip()

        if choice == '0':
            break
        elif choice == '1':
            _add_job_interactive(script_dir)
        elif choice == '2':
            _edit_job_interactive(script_dir)
        elif choice == '3':
            _remove_job_interactive(script_dir)
        elif choice == '4':
            start_scheduler(script_dir)
        elif choice == '5':
            _run_job_now_interactive(script_dir)
        elif choice == '6':
            _show_scheduler_history(script_dir)
        elif choice == '7':
            _generate_system_task_script(script_dir)
        elif choice == '8':
            _scheduler_settings_menu(script_dir)
        else:
            print('无效选项')
            input('按回车继续...')


def _add_job_interactive(script_dir=None):
    print('\n--- 添加定时任务 ---')

    name = input('任务名称: ').strip()
    if not name:
        print('名称不能为空')
        input('按回车继续...')
        return

    watch_directory = input('监控目录路径: ').strip().strip('"').strip("'")
    if not watch_directory or not os.path.isdir(watch_directory):
        print('目录不存在')
        input('按回车继续...')
        return

    print('\n调度类型:')
    print('  1) cron 表达式 (推荐)')
    print('  2) 固定间隔(分钟)')
    type_choice = input('请选择 (默认1): ').strip()

    job_config = {
        'name': name,
        'watch_directory': watch_directory,
        'enabled': True,
    }

    if type_choice == '2':
        interval_input = input('间隔分钟数 (默认60): ').strip()
        interval = int(interval_input) if interval_input.isdigit() else 60
        job_config['schedule_type'] = 'interval'
        job_config['interval_minutes'] = interval
    else:
        print('\ncron 表达式格式: 分 时 日 月 周')
        print('  示例:')
        print('    0 0 * * *    = 每天 00:00')
        print('    0 2 * * *    = 每天 02:00')
        print('    0 */6 * * *  = 每6小时')
        print('    0 8 * * 1-5  = 周一至周五 08:00')
        cron_expr = input('请输入 cron 表达式 (默认 0 0 * * *): ').strip() or '0 0 * * *'
        job_config['schedule_type'] = 'cron'
        job_config['cron_expression'] = cron_expr

    inc_choice = input('启用增量合并? (Y/n): ').strip().lower()
    job_config['incremental'] = inc_choice != 'n'

    job_config['description'] = input('任务描述 (可选): ').strip()

    job_id = add_schedule_job(job_config, script_dir)
    print(f'✅ 任务已添加，ID: {job_id}')
    input('按回车继续...')


def _edit_job_interactive(script_dir=None):
    jobs = list_schedule_jobs(script_dir)
    if not jobs:
        print('暂无任务')
        input('按回车继续...')
        return

    job_id = input('请输入要编辑的任务ID: ').strip()
    job = next((j for j in jobs if j['job_id'] == job_id), None)
    if not job:
        print('未找到该任务')
        input('按回车继续...')
        return

    print(f'\n当前配置:')
    print(f'  名称: {job.get("name", "")}')
    print(f'  目录: {job.get("watch_directory", "")}')
    print(f'  类型: {job.get("schedule_type", "cron")}')
    if job.get('schedule_type') == 'interval':
        print(f'  间隔: {job.get("interval_minutes", 60)} 分钟')
    else:
        print(f'  cron: {job.get("cron_expression", "")}')
    print(f'  增量: {"启用" if job.get("incremental", True) else "禁用"}')
    print(f'  状态: {"启用" if job.get("enabled", True) else "禁用"}')

    print(f'\n编辑 (留空保持当前值):')
    updates = {}

    new_name = input(f'新名称 [{job.get("name", "")}]: ').strip()
    if new_name:
        updates['name'] = new_name

    new_dir = input(f'新目录 [{job.get("watch_directory", "")}]: ').strip().strip('"').strip("'")
    if new_dir:
        if os.path.isdir(new_dir):
            updates['watch_directory'] = new_dir
        else:
            print('目录不存在，跳过')

    new_enabled = input(f'启用? (y/n) [{job.get("enabled", True)}]: ').strip().lower()
    if new_enabled in ['y', 'n']:
        updates['enabled'] = new_enabled == 'y'

    new_inc = input(f'增量合并? (y/n) [{job.get("incremental", True)}]: ').strip().lower()
    if new_inc in ['y', 'n']:
        updates['incremental'] = new_inc == 'y'

    if updates:
        update_schedule_job(job_id, updates, script_dir)
        print('✅ 任务已更新')
    else:
        print('未做修改')

    input('按回车继续...')


def _remove_job_interactive(script_dir=None):
    jobs = list_schedule_jobs(script_dir)
    if not jobs:
        print('暂无任务')
        input('按回车继续...')
        return

    job_id = input('请输入要删除的任务ID: ').strip()
    job = next((j for j in jobs if j['job_id'] == job_id), None)
    if not job:
        print('未找到该任务')
        input('按回车继续...')
        return

    confirm = input(f'确认删除任务 [{job.get("name", job_id)}]? (y/N): ').strip().lower()
    if confirm == 'y':
        remove_schedule_job(job_id, script_dir)
        print('✅ 任务已删除')
    else:
        print('已取消')

    input('按回车继续...')


def _run_job_now_interactive(script_dir=None):
    jobs = list_schedule_jobs(script_dir)
    if not jobs:
        print('暂无任务')
        input('按回车继续...')
        return

    job_id = input('请输入要立即运行的任务ID: ').strip()
    job = next((j for j in jobs if j['job_id'] == job_id), None)
    if not job:
        print('未找到该任务')
        input('按回车继续...')
        return

    print(f'🚀 立即运行任务: {job.get("name", job_id)}')
    result = run_scheduled_pipeline(job, script_dir)
    print(f'\n执行完成，状态: {result["status"]}')
    print(f'扫描文件: {result.get("files_scanned", 0)} 个')
    print(f'新增文件: {result.get("files_new", 0)} 个')
    print(f'处理文件: {result.get("files_processed", 0)} 个')
    print(f'提取记录: {result.get("records_extracted", 0)} 条')
    print(f'耗时: {result.get("duration_ms", 0)} ms')
    if result.get("output_path"):
        print(f'输出: {result["output_path"]}')

    input('按回车继续...')


def _show_scheduler_history(script_dir=None):
    if script_dir is None:
        script_dir = get_script_dir()

    db_path = get_processed_files_db_path(script_dir)
    if not os.path.exists(db_path):
        print('暂无运行历史')
        input('按回车继续...')
        return

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute('''
        SELECT * FROM scheduler_runs
        ORDER BY started_at DESC
        LIMIT 50
    ''')
    rows = cursor.fetchall()
    conn.close()

    print('\n' + '=' * 80)
    print('📜 调度任务运行历史'.center(70))
    print('=' * 80)
    print(f'{"运行ID":<14} {"任务":<18} {"开始时间":<20} {"状态":<10} {"文件":<8} {"记录":<8}')
    print('-' * 80)

    for row in rows:
        run_id = row['run_id'][:12]
        job_name = row['job_name'][:16]
        started = row['started_at'][:19]
        status_sym = _get_status_symbol(row['status'])
        files_new = row['files_new'] or 0
        records = row['records_extracted'] or 0
        print(f'{run_id:<14} {job_name:<18} {started:<20} {status_sym} {row["status"]:<8} {files_new:<8} {records:<8}')

    if not rows:
        print('暂无运行记录')

    input('\n按回车继续...')


def _generate_system_task_script(script_dir=None):
    jobs = list_schedule_jobs(script_dir)
    if not jobs:
        print('暂无任务')
        input('按回车继续...')
        return

    job_id = input('请输入任务ID: ').strip()
    job = next((j for j in jobs if j['job_id'] == job_id), None)
    if not job:
        print('未找到该任务')
        input('按回车继续...')
        return

    script_path = os.path.abspath(__file__)

    print('\n选择目标系统:')
    print('  1) Linux/macOS (cron)')
    print('  2) Windows (任务计划)')
    sys_choice = input('请选择 (默认1): ').strip()

    if sys_choice == '2':
        ps_output, bat_output = generate_windows_task_script(job, script_path, script_dir)
        print(f'\n✅ 已生成 Windows 任务计划脚本:')
        print(f'  PowerShell: {ps_output}')
        print(f'  批处理: {bat_output}')
    else:
        output_path, cron_line = generate_cron_script(job, script_path, script_dir)
        print(f'\n✅ 已生成 cron 脚本: {output_path}')
        print(f'\n手动添加到 crontab:')
        print(f'  {cron_line}')

    input('按回车继续...')


def _scheduler_settings_menu(script_dir=None):
    config = load_scheduler_config(script_dir)
    settings = config.get('settings', {})

    while True:
        print('\n--- 调度器设置 ---')
        print(f'  1) 告警通知: {"启用" if settings.get("enable_alerts", True) else "禁用"}')
        print(f'  2) 最大并发任务数: {settings.get("max_concurrent_jobs", 1)}')
        print(f'  3) 检查间隔(秒): {settings.get("check_interval_seconds", 30)}')
        print(f'  4) 日志级别: {settings.get("log_level", "INFO")}')
        print('  0) 返回')

        choice = input('\n请选择: ').strip()

        if choice == '0':
            break
        elif choice == '1':
            settings['enable_alerts'] = not settings.get('enable_alerts', True)
        elif choice == '2':
            val = input('最大并发任务数: ').strip()
            if val.isdigit():
                settings['max_concurrent_jobs'] = int(val)
        elif choice == '3':
            val = input('检查间隔(秒): ').strip()
            if val.isdigit():
                settings['check_interval_seconds'] = int(val)
        elif choice == '4':
            level = input('日志级别 (DEBUG/INFO/WARNING/ERROR): ').strip().upper()
            if level in ['DEBUG', 'INFO', 'WARNING', 'ERROR']:
                settings['log_level'] = level

        config['settings'] = settings
        save_scheduler_config(config, script_dir)
        print('✅ 设置已保存')


def run_scheduler_flow(script_dir):
    logger = get_logger()
    logger.info('========== 定时调度管理启动 ==========')
    init_default_alert_rules(script_dir)
    init_processed_files_db(get_processed_files_db_path(script_dir))
    show_scheduler_menu(script_dir)
    logger.info('========== 定时调度管理关闭 ==========')


def build_cli_parser():
    import argparse

    parser = argparse.ArgumentParser(
        description='银行流水检验工具',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='示例:\n'
               '  python bankcheck.py process /path/to/statement_folder\n'
               '  python bankcheck.py process /path/to/folder --no-incremental\n'
               '  python bankcheck.py validate-lookup\n'
               '  python bankcheck.py validate-lookup --lookup-file /path/to/主体查找表.xlsx\n'
               '  python bankcheck.py version\n',
    )

    subparsers = parser.add_subparsers(dest='command', help='可用子命令')

    process_parser = subparsers.add_parser(
        'process',
        help='处理银行流水文件夹，输出总表及检验报告',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='示例:\n'
               '  python bankcheck.py process /path/to/statement_folder\n'
               '  python bankcheck.py process /path/to/folder --no-incremental --keep-strategy keep_all\n'
               '  python bankcheck.py process /path/to/folder --output-dir /path/to/output\n'
               '  python bankcheck.py process /path/to/folder --preset my-preset-id\n',
    )
    process_parser.add_argument(
        'folder',
        type=str,
        help='银行流水文件夹路径',
    )
    process_parser.add_argument(
        '--no-incremental',
        action='store_true',
        default=False,
        help='禁用增量合并，使用全量覆盖模式',
    )
    process_parser.add_argument(
        '--keep-strategy',
        type=str,
        choices=list(KEEP_STRATEGIES.keys()),
        default='keep_unprocessed',
        help='文件保留策略 (默认: keep_unprocessed)',
    )
    process_parser.add_argument(
        '--archive-dir-name',
        type=str,
        default='已处理归档',
        help='归档子目录名称（仅 move_to_archive 策略使用，默认: 已处理归档）',
    )
    process_parser.add_argument(
        '--output-dir',
        type=str,
        metavar='DIR',
        default=None,
        help='指定输出目录（默认: 流水文件夹同级目录）',
    )
    process_parser.add_argument(
        '--preset',
        type=str,
        metavar='PRESET_ID',
        default=None,
        help='应用指定ID的预设配置',
    )
    process_parser.add_argument(
        '--enabled-banks',
        type=str,
        nargs='+',
        default=None,
        help='仅处理指定银行（如: 北京银行 东亚银行）',
    )
    process_parser.add_argument(
        '--start-date',
        type=str,
        default='',
        help='筛选起始日期 (YYYY-MM-DD)',
    )
    process_parser.add_argument(
        '--end-date',
        type=str,
        default='',
        help='筛选截止日期 (YYYY-MM-DD)',
    )
    process_parser.add_argument(
        '--batch-id',
        type=str,
        default=None,
        help='指定批次ID',
    )
    process_parser.add_argument(
        '--enable-signature',
        action='store_true',
        default=False,
        help='启用数字签名',
    )
    process_parser.add_argument(
        '--signature-password',
        type=str,
        default=None,
        help='签名密钥密码',
    )
    process_parser.add_argument(
        '--enable-encryption',
        action='store_true',
        default=False,
        help='启用输出文件加密',
    )
    process_parser.add_argument(
        '--encryption-password',
        type=str,
        default=None,
        help='加密密码',
    )
    process_parser.add_argument(
        '--encryption-mode',
        type=str,
        choices=['excel_password', 'aes'],
        default='excel_password',
        help='加密模式 (默认: excel_password)',
    )
    process_parser.add_argument(
        '--dry-run',
        action='store_true',
        default=False,
        help='试运行模式：仅生成统计与异常报告，不执行删除与写盘操作',
    )
    process_parser.add_argument(
        '--yes',
        '--commit',
        action='store_true',
        default=False,
        dest='auto_commit',
        help='试运行后自动确认提交，无需交互式确认（配合 --dry-run 使用）',
    )

    validate_parser = subparsers.add_parser(
        'validate-lookup',
        help='校验主体查找表的完整性与一致性',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='示例:\n'
               '  python bankcheck.py validate-lookup\n'
               '  python bankcheck.py validate-lookup --lookup-file /path/to/主体查找表.xlsx\n'
               '  python bankcheck.py validate-lookup --json\n',
    )
    validate_parser.add_argument(
        '--lookup-file',
        type=str,
        metavar='FILE',
        default=None,
        help='指定查找表文件路径（默认: 自动在程序目录下查找）',
    )
    validate_parser.add_argument(
        '--json',
        action='store_true',
        default=False,
        help='以 JSON 格式输出校验结果',
    )
    validate_parser.add_argument(
        '--strict',
        action='store_true',
        default=False,
        help='严格模式: 发现任何问题即返回非零退出码',
    )

    version_parser = subparsers.add_parser(
        'version',
        help='显示版本信息',
    )

    parser.add_argument('--scheduler', action='store_true', help='启动定时调度器')
    parser.add_argument('--scheduler-menu', action='store_true', help='打开调度管理菜单')
    parser.add_argument('--run-job', type=str, metavar='JOB_ID', help='立即运行指定ID的定时任务')
    parser.add_argument('--add-job', action='store_true', help='交互式添加定时任务')
    parser.add_argument('--list-jobs', action='store_true', help='列出所有定时任务')
    parser.add_argument('--watch-dir', type=str, metavar='DIR', help='单次运行: 监控目录路径')
    parser.add_argument('--once', action='store_true', help='单次运行后退出(配合--watch-dir使用)')
    parser.add_argument('--interval', type=int, metavar='MINUTES', help='单次运行模式下的间隔分钟数')
    parser.add_argument('--no-incremental', action='store_true', help='禁用增量合并，使用全量覆盖')
    parser.add_argument('--export', action='store_true', help='进入财务软件对接导出功能')
    parser.add_argument('--export-template', type=str, metavar='TEMPLATE',
                       help='指定导出模板类型: yonyou_voucher(用友), kingdee_voucher(金蝶), bank_journal(日记账)')
    parser.add_argument('--export-total', type=str, metavar='TOTAL_FILE',
                       help='指定总表文件路径用于导出')
    parser.add_argument('--export-output', type=str, metavar='OUTPUT_DIR',
                       help='指定导出文件输出目录')
    parser.add_argument('--export-operator', type=str, metavar='OPERATOR',
                       help='指定制单人名称')
    parser.add_argument('--preset-menu', action='store_true', help='打开预设管理菜单')
    parser.add_argument('--list-presets', action='store_true', help='列出所有预设')
    parser.add_argument('--apply-preset', type=str, metavar='PRESET_ID',
                       help='应用指定ID的预设，需配合--watch-dir使用')
    parser.add_argument('--save-preset', type=str, metavar='NAME',
                       help='保存当前配置为新预设')
    parser.add_argument('--subject-summary', action='store_true',
                       help='进入主体维度汇总分析功能')
    parser.add_argument('--summary-total', type=str, metavar='TOTAL_FILE',
                       help='指定总表文件直接生成主体维度汇总分析')
    parser.add_argument('--summary-output', type=str, metavar='OUTPUT_DIR',
                       help='指定主体汇总分析输出目录')
    parser.add_argument('--balance-check', action='store_true',
                       help='进入余额连续性校验功能')
    parser.add_argument('--balance-total', type=str, metavar='TOTAL_FILE',
                       help='指定总表文件直接生成余额连续性校验报告')
    parser.add_argument('--balance-output', type=str, metavar='OUTPUT_DIR',
                       help='指定余额连续性校验报告输出目录')
    parser.add_argument('--balance-tolerance', type=float, metavar='TOLERANCE',
                       help='指定余额校验容差（元），默认 0.01')
    parser.add_argument('--duplicate-check', action='store_true',
                       help='进入重复交易检测功能')
    parser.add_argument('--duplicate-total', type=str, metavar='TOTAL_FILE',
                       help='指定总表文件直接生成重复交易检测报告')
    parser.add_argument('--duplicate-output', type=str, metavar='OUTPUT_DIR',
                       help='指定重复交易检测报告输出目录')
    parser.add_argument('--interest-fee-check', action='store_true',
                       help='进入利息与手续费专项核对功能')
    parser.add_argument('--interest-fee-total', type=str, metavar='TOTAL_FILE',
                       help='指定总表文件直接生成利息手续费核对报告')
    parser.add_argument('--interest-fee-output', type=str, metavar='OUTPUT_DIR',
                       help='指定利息手续费核对报告输出目录')
    parser.add_argument('--interest-fee-period', type=str, metavar='PERIOD',
                       help='指定汇总期间类型: month(月), quarter(季), year(年), week(周)，默认 month')

    return parser


def _cmd_process(args):
    logger = get_logger()
    script_dir = get_script_dir()

    folder = os.path.abspath(args.folder)
    if not os.path.isdir(folder):
        logger.error('目录不存在: %s', folder)
        print(f'错误: 目录不存在 {folder}')
        return 1

    incremental = not args.no_incremental
    keep_strategy = args.keep_strategy
    archive_dir_name = args.archive_dir_name
    output_dir = args.output_dir
    enabled_banks = args.enabled_banks
    start_date = args.start_date
    end_date = args.end_date
    batch_id = args.batch_id
    dry_run = args.dry_run
    auto_commit = args.auto_commit

    if dry_run:
        print('\n🔬 试运行模式：仅生成统计与异常报告，不执行删除与写盘操作')
        if auto_commit:
            print('⚡ 自动提交模式：试运行完成后将自动执行写盘操作')

    if args.preset:
        preset = load_preset(args.preset, script_dir)
        if not preset:
            logger.error('未找到预设ID: %s', args.preset)
            print(f'错误: 未找到预设ID {args.preset}')
            return 1
        print(f'应用预设: {preset.get("name", "")} ({args.preset})')
        print(f'处理目录: {folder}')
        with AuditLogger('preset_pipeline', script_dir) as audit:
            audit.record_input(folder)
            audit.set_extra_info({'preset_id': args.preset, 'preset_name': preset.get('name', ''),
                                  'dry_run': dry_run, 'auto_commit': auto_commit})
            result = apply_preset_to_pipeline(preset, folder, script_dir, dry_run=dry_run)
            audit.record_result(result)
    else:
        logger.info('CLI process: 目录=%s, 增量=%s, 保留策略=%s, dry_run=%s',
                    folder, incremental, keep_strategy, dry_run)

        result = run_pipeline_with_options(
            folder=folder,
            script_dir=script_dir,
            incremental=incremental,
            enabled_banks=enabled_banks,
            keep_strategy=keep_strategy,
            start_date=start_date,
            end_date=end_date,
            batch_id=batch_id,
            output_dir=output_dir,
            enable_encryption=args.enable_encryption,
            encryption_password=args.encryption_password,
            encryption_mode=args.encryption_mode,
            dry_run=dry_run,
            archive_dir_name=archive_dir_name,
        )

    if result.folder_empty:
        print(f'\n⚠️  文件夹中未发现任何 Excel 文件')
        return 0

    committed = False
    if result.dry_run and result.all_rows:
        msg = format_result_message(result)
        print('\n' + msg)
        if auto_commit:
            print('\n⚡ 自动提交试运行变更...')
            result = commit_pipeline_changes(result)
            committed = True
        else:
            print('\n是否确认提交并执行正式写盘操作？')
            print('  - 删除已处理成功的源文件（按保留策略）')
            print('  - 写入/覆盖银行流水总表与脱敏版总表')
            print('  - 执行数据库持久化（如已配置）')
            print('  - 执行文件加密（如已启用）')
            choice = input('\n请确认提交 (y/N): ').strip().lower()
            if choice in ('y', 'yes'):
                print('\n📝 正在提交试运行变更...')
                result = commit_pipeline_changes(result)
                committed = True
            else:
                print('\n⏭️  已取消提交，未执行任何删除与写盘操作。')
                result.changes_committed = False

    if result.all_rows:
        if committed:
            print(f'\n✅ 提交完成！')
        elif result.dry_run and not result.changes_committed:
            print(f'\n⏭️  试运行完成（未提交）')
        else:
            print(f'\n✅ 处理完成！')
        print(f'   总记录数: {len(result.all_rows)}')
        print(f'   新增记录: {result.new_record_count}')
        print(f'   已处理文件: {len(result.processed_files)}')
        if result.output_path:
            print(f'   输出文件: {result.output_path}')
        elif result.dry_run and not committed:
            print(f'   输出文件: (试运行未写盘)')
        if result.masked_output_path:
            print(f'   脱敏版总表: {result.masked_output_path}')
        elif result.dry_run and not committed:
            print(f'   脱敏版总表: (试运行未生成)')
        if result.lookup_missing:
            print(f'   ⚠️  未找到主体查找表，"主体"列为空')
        if result.unprocessed_files:
            print(f'   无法识别的文件: {len(result.unprocessed_files)} 个')
        if result.error_files:
            print(f'   处理出错的文件: {len(result.error_files)} 个')
            for fpath, err in result.error_files:
                print(f'     - {os.path.basename(fpath)}: {err}')
        if result.balance_check_path:
            print(f'   余额校验报告: {result.balance_check_path}')
        if result.duplicate_check_path:
            print(f'   重复交易报告: {result.duplicate_check_path}')
        if result.interest_fee_check_path:
            print(f'   利息手续费报告: {result.interest_fee_check_path}')
        if result.holiday_check_path:
            print(f'   非工作日交易报告: {result.holiday_check_path}')
        if result.subject_summary_path:
            print(f'   主体汇总分析: {result.subject_summary_path}')
        if HAS_DATABASE and (result.db_inserted_count > 0 or result.db_duplicate_count > 0):
            print(f'   数据库入库: {result.db_inserted_count} 条新增, {result.db_duplicate_count} 条重复跳过')
        elif HAS_DATABASE and result.dry_run and not committed:
            print(f'   数据库入库: (试运行未执行)')
        return 0
    else:
        print(f'\n⚠️  未提取到任何银行流水记录')
        if result.unprocessed_files:
            print(f'   无法识别的文件: {len(result.unprocessed_files)} 个')
        if result.error_files:
            print(f'   处理出错的文件: {len(result.error_files)} 个')
            for fpath, err in result.error_files:
                print(f'     - {os.path.basename(fpath)}: {err}')
        return 0


def _cmd_validate_lookup(args):
    import json as _json

    script_dir = get_script_dir()
    logger = get_logger()

    lookup_file = args.lookup_file
    if lookup_file is None:
        lookup_file = find_lookup_file(script_dir)

    issues = []
    warnings = []

    if lookup_file is None or not os.path.isfile(lookup_file):
        msg = '未找到主体查找表文件'
        issues.append(msg)
        if args.json:
            print(_json.dumps({'valid': False, 'issues': issues, 'warnings': warnings},
                              ensure_ascii=False, indent=2))
        else:
            print(f'❌ {msg}')
        return 1 if args.strict else 0

    if not args.json:
        print(f'正在校验查找表: {lookup_file}')

    try:
        lookup_data = load_lookup_table(lookup_file)
    except Exception as e:
        msg = f'查找表加载失败: {e}'
        issues.append(msg)
        if args.json:
            print(_json.dumps({'valid': False, 'issues': issues, 'warnings': warnings, 'file': lookup_file},
                              ensure_ascii=False, indent=2))
        else:
            print(f'❌ {msg}')
        return 1 if args.strict else 0

    by_account = lookup_data.get('by_account', {})
    all_entries = lookup_data.get('all_entries', [])

    if not all_entries:
        msg = '查找表为空，没有任何条目'
        issues.append(msg)

    empty_subject_count = 0
    empty_account_count = 0
    duplicate_account_entries = []

    seen_account_keys = {}
    for entry in all_entries:
        subject = entry.get('subject', '').strip()
        account_raw = str(entry.get('account_raw', '')).strip()
        account_key = entry.get('account_key', '').strip()

        if not subject:
            empty_subject_count += 1
        if not account_raw:
            empty_account_count += 1

        if account_key:
            if account_key not in seen_account_keys:
                seen_account_keys[account_key] = []
            seen_account_keys[account_key].append(subject)

    if empty_subject_count > 0:
        msg = f'存在 {empty_subject_count} 条主体为空的记录'
        issues.append(msg)

    if empty_account_count > 0:
        msg = f'存在 {empty_account_count} 条账号为空的记录'
        issues.append(msg)

    for account_key, subjects in seen_account_keys.items():
        unique_subjects = list(set(s for s in subjects if s))
        if len(unique_subjects) > 1:
            duplicate_account_entries.append({
                'account': account_key,
                'subjects': unique_subjects,
                'count': len(subjects),
            })

    if duplicate_account_entries:
        msg = f'存在 {len(duplicate_account_entries)} 个账号映射到多个不同主体'
        issues.append(msg)
        for dup in duplicate_account_entries:
            detail = f'  账号 {dup["account"]} -> {", ".join(dup["subjects"])}'
            issues.append(detail)

    try:
        from lookup_manager import get_duplicate_entries as _get_dup_entries
        dup_entries = _get_dup_entries(lookup_file)
        if dup_entries and not duplicate_account_entries:
            msg = f'查找表存在 {len(dup_entries)} 组重复账号条目'
            warnings.append(msg)
    except Exception:
        pass

    file_size = os.path.getsize(lookup_file)
    if file_size == 0:
        msg = '查找表文件大小为 0 字节'
        issues.append(msg)

    is_valid = len([i for i in issues if not i.startswith('  ')]) == 0

    result = {
        'valid': is_valid,
        'file': lookup_file,
        'total_entries': len(all_entries),
        'unique_accounts': len(by_account),
        'issues': [i for i in issues if not i.startswith('  ')],
        'warnings': warnings,
        'duplicate_accounts': duplicate_account_entries,
    }

    if args.json:
        print(_json.dumps(result, ensure_ascii=False, indent=2))
    else:
        if is_valid:
            print(f'✅ 查找表校验通过')
        else:
            print(f'❌ 查找表校验未通过')
        print(f'   文件: {lookup_file}')
        print(f'   总条目数: {len(all_entries)}')
        print(f'   唯一账号数: {len(by_account)}')
        if issues:
            problem_issues = [i for i in issues if not i.startswith('  ')]
            detail_issues = [i for i in issues if i.startswith('  ')]
            if problem_issues:
                print(f'   问题 ({len(problem_issues)} 个):')
                for issue in problem_issues:
                    print(f'     - {issue}')
            if detail_issues:
                for issue in detail_issues:
                    print(f'     {issue}')
        if warnings:
            print(f'   警告 ({len(warnings)} 个):')
            for w in warnings:
                print(f'     - {w}')

    return 0 if is_valid or not args.strict else 1


def _cmd_version(args):
    print(format_version_banner())
    build_info = get_build_info()
    print(f'  Python: {sys.version.split()[0]}')
    print(f'  工作目录: {os.getcwd()}')
    return 0


def parse_args_and_run():
    parser = build_cli_parser()
    args = parser.parse_args()

    setup_logging()
    logger = get_logger()
    script_dir = get_script_dir()

    init_audit_db(get_audit_db_path(script_dir))
    init_default_alert_rules(script_dir)

    if args.command == 'process':
        sys.exit(_cmd_process(args))

    if args.command == 'validate-lookup':
        sys.exit(_cmd_validate_lookup(args))

    if args.command == 'version':
        sys.exit(_cmd_version(args))

    if args.list_jobs:
        jobs = list_schedule_jobs(script_dir)
        print(f'定时任务列表 (共 {len(jobs)} 个):')
        for job in jobs:
            status = '启用' if job.get('enabled', True) else '禁用'
            print(f'  [{job["job_id"]}] {job.get("name", "")} - {status}')
            print(f'    目录: {job.get("watch_directory", "")}')
            if job.get('schedule_type') == 'interval':
                print(f'    调度: 每 {job.get("interval_minutes", 60)} 分钟')
            else:
                print(f'    调度: {job.get("cron_expression", "")}')
        return True

    if args.list_presets:
        presets = list_presets(script_dir)
        print(f'预设列表 (共 {len(presets)} 个):')
        for preset in presets:
            banks = ', '.join(preset.get('enabled_banks', []))
            keep = KEEP_STRATEGIES.get(preset.get('keep_strategy'), '未知')
            print(f'  [{preset["preset_id"]}] {preset.get("name", "")}')
            print(f'    描述: {preset.get("description", "无")}')
            print(f'    银行: {banks}')
            print(f'    保留策略: {keep}')
            print(f'    增量: {"是" if preset.get("incremental", True) else "否"}')
            if preset.get('start_date') or preset.get('end_date'):
                print(f'    日期: {preset.get("start_date", "不限")} ~ {preset.get("end_date", "不限")}')
            print(f'    更新时间: {preset.get("updated_at", "")}')
        return True

    if args.apply_preset and args.watch_dir:
        preset = load_preset(args.apply_preset, script_dir)
        if not preset:
            logger.error('未找到预设ID: %s', args.apply_preset)
            print(f'错误: 未找到预设ID {args.apply_preset}')
            return True

        watch_dir = os.path.abspath(args.watch_dir)
        if not os.path.isdir(watch_dir):
            logger.error('目录不存在: %s', watch_dir)
            print(f'错误: 目录不存在 {watch_dir}')
            return True

        print(f'应用预设: {preset.get("name", "")} ({args.apply_preset})')
        print(f'处理目录: {watch_dir}')

        with AuditLogger('preset_pipeline', script_dir) as audit:
            audit.record_input(watch_dir)
            audit.set_extra_info({'preset_id': args.apply_preset, 'preset_name': preset.get('name', '')})
            result = apply_preset_to_pipeline(preset, watch_dir, script_dir)
            audit.record_result(result)

            if result.all_rows:
                print(f'\n✅ 处理完成！')
                print(f'   总记录数: {len(result.all_rows)}')
                print(f'   新增记录: {result.new_record_count}')
                print(f'   输出文件: {result.output_path}')
            else:
                print(f'\n⚠️  未提取到记录')
            return True

    if args.save_preset:
        preset_data = {
            'name': args.save_preset,
            'output_dir': args.export_output or '',
            'enabled_banks': list(BANK_PREFIXES),
            'keep_strategy': 'keep_unprocessed',
            'incremental': not args.no_incremental,
        }
        preset_id = save_preset(preset_data, script_dir)
        print(f'\n✅ 预设已保存，ID: {preset_id}')
        return True

    if args.add_job:
        _add_job_interactive(script_dir)
        return True

    if args.run_job:
        jobs = list_schedule_jobs(script_dir)
        job = next((j for j in jobs if j['job_id'] == args.run_job), None)
        if not job:
            logger.error('未找到任务ID: %s', args.run_job)
            return True
        run_scheduled_pipeline(job, script_dir)
        return True

    if args.watch_dir:
        watch_dir = os.path.abspath(args.watch_dir)
        if not os.path.isdir(watch_dir):
            logger.error('目录不存在: %s', watch_dir)
            return True

        job_config = {
            'job_id': 'ONETIME',
            'name': '单次运行任务',
            'watch_directory': watch_dir,
            'incremental': not args.no_incremental,
            'schedule_type': 'interval',
            'interval_minutes': args.interval or 60,
        }

        if args.once:
            logger.info('单次运行模式，目录: %s', watch_dir)
            run_scheduled_pipeline(job_config, script_dir)
            return True
        else:
            if not args.interval:
                logger.warning('未指定--interval，将使用默认60分钟间隔')

            logger.info('监控模式启动，目录: %s，间隔: %d 分钟，Ctrl+C 停止',
                        watch_dir, job_config['interval_minutes'])

            scheduler, scheduler_type, CronTrigger, IntervalTrigger = get_scheduler()

            if scheduler_type == 'apscheduler':
                trigger = IntervalTrigger(minutes=job_config['interval_minutes'])
                scheduler.add_job(
                    run_scheduled_pipeline,
                    trigger=trigger,
                    args=[job_config, script_dir],
                    id='watch_job',
                    name='监控目录任务',
                )
            else:
                class SimpleIntervalTrigger:
                    def __init__(self, minutes):
                        self._minutes = minutes
                    def get_interval(self):
                        from datetime import timedelta
                        return timedelta(minutes=self._minutes)
                scheduler.add_job(
                    run_scheduled_pipeline,
                    trigger=SimpleIntervalTrigger(job_config['interval_minutes']),
                    args=[job_config, script_dir],
                    id='watch_job',
                    name='监控目录任务',
                )

            try:
                run_scheduled_pipeline(job_config, script_dir)
                scheduler.start()
            except (KeyboardInterrupt, SystemExit):
                scheduler.shutdown()
            return True

    if args.scheduler:
        start_scheduler(script_dir)
        return True

    if args.scheduler_menu:
        run_scheduler_flow(script_dir)
        return True

    if args.preset_menu:
        run_preset_flow(script_dir)
        return True

    if args.export or args.export_template or args.export_total:
        if args.export_total and args.export_template:
            template_type = args.export_template
            if template_type not in FINANCIAL_EXPORT_TEMPLATES:
                logger.error('不支持的导出模板类型: %s', template_type)
                print(f'错误: 不支持的导出模板类型 "{template_type}"')
                print(f'支持的类型: {", ".join(FINANCIAL_EXPORT_TEMPLATES.keys())}')
                return True

            total_path = os.path.abspath(args.export_total)
            if not os.path.exists(total_path):
                logger.error('总表文件不存在: %s', total_path)
                print(f'错误: 总表文件不存在: {total_path}')
                return True

            result = export_financial_template(
                total_path=total_path,
                template_type=template_type,
                output_dir=args.export_output,
                operator=args.export_operator or '',
            )

            if result['success']:
                print(f'\n✅ 导出成功！')
                print(f'   模板: {result["template_name"]}')
                print(f'   凭证张数: {result["voucher_count"]}')
                print(f'   分录条数: {result["entry_count"]}')
                print(f'   输出文件: {result["output_path"]}\n')
            else:
                print(f'\n❌ 导出失败: {result.get("error", "未知错误")}\n')

            return True
        else:
            run_export_flow(script_dir)
            return True

    if args.subject_summary or args.summary_total:
        if args.summary_total:
            total_path = os.path.abspath(args.summary_total)
            if not os.path.exists(total_path):
                logger.error('总表文件不存在: %s', total_path)
                print(f'错误: 总表文件不存在: {total_path}')
                return True

            output_dir = None
            if args.summary_output:
                output_dir = os.path.abspath(args.summary_output)
                os.makedirs(output_dir, exist_ok=True)

            result_path = generate_subject_summary_from_total(total_path, output_dir)
            if result_path:
                print(f'\n✅ 主体维度汇总分析已生成！')
                print(f'   输出文件: {result_path}\n')
            else:
                print(f'\n❌ 生成失败，请检查总表文件是否有数据\n')
            return True
        else:
            run_subject_summary_flow(script_dir)
            return True

    if args.balance_check or args.balance_total:
        if args.balance_total:
            total_path = os.path.abspath(args.balance_total)
            if not os.path.exists(total_path):
                logger.error('总表文件不存在: %s', total_path)
                print(f'错误: 总表文件不存在: {total_path}')
                return True

            output_dir = None
            if args.balance_output:
                output_dir = os.path.abspath(args.balance_output)
                os.makedirs(output_dir, exist_ok=True)

            tolerance = args.balance_tolerance if args.balance_tolerance is not None else 0.01

            result_path = generate_balance_check_from_total(total_path, output_dir, tolerance)
            if result_path:
                print(f'\n✅ 余额连续性校验报告已生成！')
                print(f'   输出文件: {result_path}\n')
            else:
                print(f'\n❌ 生成失败，请检查总表文件是否有数据\n')
            return True
        else:
            run_balance_check_flow(script_dir)
            return True

    if args.duplicate_check or args.duplicate_total:
        if args.duplicate_total:
            total_path = os.path.abspath(args.duplicate_total)
            if not os.path.exists(total_path):
                logger.error('总表文件不存在: %s', total_path)
                print(f'错误: 总表文件不存在: {total_path}')
                return True

            output_dir = None
            if args.duplicate_output:
                output_dir = os.path.abspath(args.duplicate_output)
                os.makedirs(output_dir, exist_ok=True)

            result_path = generate_duplicate_check_from_total(total_path, output_dir)
            if result_path:
                print(f'\n✅ 重复交易检测报告已生成！')
                print(f'   输出文件: {result_path}\n')
            else:
                print(f'\n❌ 生成失败，请检查总表文件是否有数据\n')
            return True
        else:
            run_duplicate_check_flow(script_dir)
            return True

    if args.interest_fee_check or args.interest_fee_total:
        if args.interest_fee_total:
            total_path = os.path.abspath(args.interest_fee_total)
            if not os.path.exists(total_path):
                logger.error('总表文件不存在: %s', total_path)
                print(f'错误: 总表文件不存在: {total_path}')
                return True

            output_dir = None
            if args.interest_fee_output:
                output_dir = os.path.abspath(args.interest_fee_output)
                os.makedirs(output_dir, exist_ok=True)

            period_type = args.interest_fee_period if args.interest_fee_period else 'month'

            result_path = generate_interest_fee_check_from_total(total_path, output_dir, period_type)
            if result_path:
                print(f'\n✅ 利息手续费核对报告已生成！')
                print(f'   输出文件: {result_path}\n')
            else:
                print(f'\n❌ 生成失败，请检查总表文件是否有数据\n')
            return True
        else:
            run_interest_fee_check_flow(script_dir)
            return True

    return None


# ──────────────────────────────────────────────
# 对方户名黑名单/白名单匹配模块
# ──────────────────────────────────────────────

@dataclass
class CounterpartyRule:
    rule_id: str
    name: str
    rule_type: str
    keywords: List[str]
    match_mode: str = 'contains'
    category: str = ''
    severity: str = 'medium'
    enabled: bool = True
    description: Optional[str] = None
    created_at: str = ''
    updated_at: str = ''
    created_by: str = ''


class CounterpartyRuleConfig:

    def __init__(self, script_dir=None):
        self.script_dir = script_dir or get_script_dir()
        self.config_path = os.path.join(self.script_dir, 'counterparty_rules.json')
        self._rules: List[CounterpartyRule] = []
        self.load_config()

    def load_config(self):
        logger = get_logger()
        if os.path.exists(self.config_path):
            try:
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self._rules = [CounterpartyRule(**r) for r in data.get('rules', [])]
                logger.info('对方户名规则配置已加载: %d 条规则', len(self._rules))
            except Exception as e:
                logger.error('加载对方户名规则配置失败: %s', e)
                self._rules = []
        else:
            self._rules = []
            self.save_config()

    def save_config(self):
        logger = get_logger()
        try:
            data = {
                'rules': [vars(r) for r in self._rules],
            }
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.info('对方户名规则配置已保存: %s', self.config_path)
        except Exception as e:
            logger.error('保存对方户名规则配置失败: %s', e)

    def get_rules(self, rule_type=None, enabled=None) -> List[CounterpartyRule]:
        result = self._rules
        if rule_type is not None:
            result = [r for r in result if r.rule_type == rule_type]
        if enabled is not None:
            result = [r for r in result if r.enabled == enabled]
        return result

    def add_rule(self, rule: CounterpartyRule) -> str:
        if not rule.rule_id:
            rule.rule_id = f"CPR{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if not rule.created_at:
            rule.created_at = now
        rule.updated_at = now
        self._rules.append(rule)
        self.save_config()
        return rule.rule_id

    def update_rule(self, rule_id: str, updates: dict) -> bool:
        for i, r in enumerate(self._rules):
            if r.rule_id == rule_id:
                for k, v in updates.items():
                    if hasattr(r, k):
                        setattr(r, k, v)
                r.updated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.save_config()
                return True
        return False

    def delete_rule(self, rule_id: str) -> bool:
        original_len = len(self._rules)
        self._rules = [r for r in self._rules if r.rule_id != rule_id]
        if len(self._rules) < original_len:
            self.save_config()
            return True
        return False

    def toggle_rule(self, rule_id: str, enabled: bool) -> bool:
        for r in self._rules:
            if r.rule_id == rule_id:
                r.enabled = enabled
                r.updated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.save_config()
                return True
        return False


_counterparty_rule_config_instance = None


def get_counterparty_rule_config(script_dir=None) -> CounterpartyRuleConfig:
    global _counterparty_rule_config_instance
    if _counterparty_rule_config_instance is None:
        _counterparty_rule_config_instance = CounterpartyRuleConfig(script_dir)
    return _counterparty_rule_config_instance


def _match_counterparty(name: str, rule: CounterpartyRule) -> Optional[str]:
    if not name:
        return None
    import re
    name = str(name).strip()
    for kw in rule.keywords:
        kw = str(kw).strip()
        if not kw:
            continue
        if rule.match_mode == 'exact':
            if name == kw:
                return kw
        elif rule.match_mode == 'startswith':
            if name.startswith(kw):
                return kw
        elif rule.match_mode == 'endswith':
            if name.endswith(kw):
                return kw
        elif rule.match_mode == 'regex':
            try:
                if re.search(kw, name):
                    return kw
            except re.error:
                pass
        else:
            if kw in name:
                return kw
    return None


def apply_counterparty_rules(records: List[Dict], script_dir=None) -> Tuple[List[Dict], Dict[str, Any]]:
    logger = get_logger()
    config = get_counterparty_rule_config(script_dir)
    rules = config.get_rules(enabled=True)

    blacklist_hits = 0
    whitelist_hits = 0
    rule_hit_counts: Dict[str, int] = {}
    tagged_count = 0

    for rec in records:
        tags = []
        hit_names = []
        hit_kw = []
        counterparty = rec.get('对方户名', '')
        for rule in rules:
            matched_kw = _match_counterparty(counterparty, rule)
            if matched_kw:
                prefix = '黑名单' if rule.rule_type == 'blacklist' else '白名单'
                cat = f"-{rule.category}" if rule.category else ""
                label = f"{prefix}:{cat}{rule.name}"
                tags.append(label)
                hit_names.append(rule.name)
                hit_kw.append(matched_kw)
                rule_hit_counts[rule.name] = rule_hit_counts.get(rule.name, 0) + 1
                if rule.rule_type == 'blacklist':
                    blacklist_hits += 1
                else:
                    whitelist_hits += 1

        if tags:
            rec['黑白名单标签'] = ','.join(tags)
            rec['命中规则名称'] = ','.join(hit_names)
            rec['命中关键词'] = ','.join(hit_kw)
            tagged_count += 1
        else:
            rec['黑白名单标签'] = ''
            rec['命中规则名称'] = ''
            rec['命中关键词'] = ''

    summary = {
        'total_records': len(records),
        'tagged_count': tagged_count,
        'blacklist_hits': blacklist_hits,
        'whitelist_hits': whitelist_hits,
        'rule_hit_counts': rule_hit_counts,
    }
    logger.info(
        '对方户名规则匹配完成: 总记录 %d, 命中 %d, 黑名单 %d, 白名单 %d',
        summary['total_records'], summary['tagged_count'],
        summary['blacklist_hits'], summary['whitelist_hits'],
    )
    return records, summary


def export_counterparty_tags(records: List[Dict], output_path: str) -> str:
    logger = get_logger()
    tagged = [r for r in records if r.get('黑白名单标签')]
    if not tagged:
        logger.info('没有命中对方户名规则的记录，跳过导出')
        return ''

    import pandas as pd
    df = pd.DataFrame(tagged)

    cols = list(df.columns)
    priority_cols = ['黑白名单标签', '命中规则名称', '命中关键词']
    for pc in reversed(priority_cols):
        if pc in cols:
            cols.remove(pc)
            cols.insert(0, pc)
    df = df[cols]

    df.to_excel(output_path, index=False, engine='openpyxl')

    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    tag_col = None
    for col_idx in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col_idx).value == '黑白名单标签':
            tag_col = col_idx
            break

    if tag_col:
        red_fill = openpyxl.styles.PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        green_fill = openpyxl.styles.PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
        for row_idx in range(2, ws.max_row + 1):
            cell_val = str(ws.cell(row=row_idx, column=tag_col).value or '')
            if '黑名单' in cell_val:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = red_fill
            elif '白名单' in cell_val:
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = green_fill

    wb.save(output_path)
    wb.close()
    logger.info('对方户名标签导出完成: %s (%d 条记录)', output_path, len(tagged))
    return output_path


def add_counterparty_keyword_rule(name, rule_type, keywords, match_mode='contains',
                                   category='', severity='medium', description=None,
                                   script_dir=None, username=None) -> str:
    config = get_counterparty_rule_config(script_dir)
    rule = CounterpartyRule(
        rule_id='',
        name=name,
        rule_type=rule_type,
        keywords=keywords,
        match_mode=match_mode,
        category=category,
        severity=severity,
        enabled=True,
        description=description,
        created_at='',
        updated_at='',
        created_by=username or get_current_user(),
    )
    return config.add_rule(rule)


# ──────────────────────────────────────────────
# 凭证附件关联模块
# ──────────────────────────────────────────────

VOUCHER_ATTACHMENT_TYPES = ('发票', '回单', '其他')


def _verify_has_database():
    if not HAS_DATABASE:
        raise ImportError('database 模块不可用，凭证附件功能需要数据库支持')


def add_voucher_attachment(transaction_id: str,
                           attachment_path: str,
                           attachment_type: str = '其他',
                           remark: Optional[str] = None,
                           script_dir=None) -> int:
    """
    新增凭证附件映射

    Args:
        transaction_id: 交易流水号
        attachment_path: 附件文件路径（绝对路径）
        attachment_type: 附件类型（'发票'/'回单'/'其他'）
        remark: 备注
        script_dir: 脚本目录

    Returns:
        新记录 id
    """
    _verify_has_database()
    if not transaction_id or not str(transaction_id).strip():
        raise ValueError('交易流水号不能为空')
    if not attachment_path or not str(attachment_path).strip():
        raise ValueError('附件路径不能为空')
    if attachment_type and attachment_type not in VOUCHER_ATTACHMENT_TYPES:
        logger.warning('未知的附件类型 "%s"，建议使用: %s', attachment_type, VOUCHER_ATTACHMENT_TYPES)
    return db_module.add_voucher_attachment(
        transaction_id=str(transaction_id).strip(),
        attachment_path=str(attachment_path).strip(),
        attachment_type=str(attachment_type or '').strip(),
        remark=remark,
        script_dir=script_dir,
    )


def update_voucher_attachment(attachment_id: int,
                              attachment_path: Optional[str] = None,
                              attachment_type: Optional[str] = None,
                              remark: Optional[str] = None,
                              script_dir=None) -> bool:
    """更新凭证附件记录"""
    _verify_has_database()
    return db_module.update_voucher_attachment(
        attachment_id=attachment_id,
        attachment_path=attachment_path,
        attachment_type=attachment_type,
        remark=remark,
        script_dir=script_dir,
    )


def delete_voucher_attachment(attachment_id: int, script_dir=None) -> bool:
    """删除凭证附件记录"""
    _verify_has_database()
    return db_module.remove_voucher_attachment(
        attachment_id=attachment_id,
        script_dir=script_dir,
    )


def get_voucher_attachment(attachment_id: int, script_dir=None):
    """根据 id 获取凭证附件"""
    _verify_has_database()
    return db_module.get_voucher_attachment_by_id(
        attachment_id=attachment_id,
        script_dir=script_dir,
    )


def list_voucher_attachments(transaction_id: Optional[str] = None,
                             attachment_type: Optional[str] = None,
                             keyword: Optional[str] = None,
                             limit: Optional[int] = None,
                             offset: int = 0,
                             script_dir=None):
    """
    查询凭证附件列表

    Args:
        transaction_id: 按交易流水号筛选
        attachment_type: 按附件类型筛选
        keyword: 按备注/路径关键词搜索
        limit: 分页条数
        offset: 分页偏移
        script_dir: 脚本目录

    Returns:
        VoucherAttachmentQueryResult
    """
    _verify_has_database()
    return db_module.query_voucher_attachments(
        transaction_id=transaction_id,
        attachment_type=attachment_type,
        keyword=keyword,
        limit=limit,
        offset=offset,
        script_dir=script_dir,
    )


def open_voucher_attachment(attachment_id: Optional[int] = None,
                            attachment_path: Optional[str] = None,
                            script_dir=None) -> Tuple[bool, str]:
    """
    一键打开凭证附件（使用系统默认程序）

    Args:
        attachment_id: 附件记录 id（与 attachment_path 二选一）
        attachment_path: 直接指定附件路径（与 attachment_id 二选一）
        script_dir: 脚本目录

    Returns:
        (是否成功, 消息)
    """
    logger = get_logger()

    if attachment_path is None:
        if attachment_id is None:
            return False, '必须提供 attachment_id 或 attachment_path'
        _verify_has_database()
        att = db_module.get_voucher_attachment_by_id(attachment_id, script_dir=script_dir)
        if att is None:
            return False, f'未找到 id={attachment_id} 的凭证附件记录'
        attachment_path = att.附件路径

    if not attachment_path:
        return False, '附件路径为空'

    if not os.path.exists(attachment_path):
        return False, f'附件文件不存在: {attachment_path}'

    try:
        if sys.platform.startswith('darwin'):
            import subprocess
            subprocess.run(['open', attachment_path], check=False)
        elif os.name == 'nt':
            os.startfile(attachment_path)  # type: ignore[attr-defined]
        elif os.name == 'posix':
            import subprocess
            subprocess.run(['xdg-open', attachment_path], check=False)
        else:
            return False, f'不支持的操作系统: {sys.platform}'
        logger.info('已打开凭证附件: %s', attachment_path)
        return True, f'已打开: {attachment_path}'
    except Exception as e:
        logger.error('打开凭证附件失败: %s, %s', attachment_path, e)
        return False, f'打开失败: {str(e)}'


def get_voucher_attachments_for_transaction(transaction_id: str, script_dir=None):
    """
    获取某笔交易流水关联的所有附件

    Args:
        transaction_id: 交易流水号
        script_dir: 脚本目录

    Returns:
        List[VoucherAttachment]
    """
    result = list_voucher_attachments(transaction_id=transaction_id, script_dir=script_dir)
    return result.records


def open_voucher_attachments_for_transaction(transaction_id: str, script_dir=None) -> Tuple[int, List[str]]:
    """
    一键打开某笔交易流水关联的所有凭证附件

    Args:
        transaction_id: 交易流水号
        script_dir: 脚本目录

    Returns:
        (成功打开数量, 失败消息列表)
    """
    attachments = get_voucher_attachments_for_transaction(transaction_id, script_dir=script_dir)
    success_count = 0
    errors = []
    for att in attachments:
        ok, msg = open_voucher_attachment(attachment_path=att.附件路径)
        if ok:
            success_count += 1
        else:
            errors.append(msg)
    return success_count, errors


def main():
    script_dir = get_script_dir()

    print()
    print(format_version_banner())
    print()

    try:
        from self_check import self_check_and_exit_if_failed
        self_check_and_exit_if_failed(
            include_optional=False,
            script_dir=script_dir,
            verbose=False,
        )
    except ImportError:
        pass

    result = parse_args_and_run()
    if result is not None:
        return

    setup_logging()
    logger = get_logger()
    logger.info('========== 银行流水检验工具启动 ==========')
    logger.info('版本: v%s, 构建时间: %s', get_version(), get_build_time())

    if HAS_ONBOARDING and onboarding is not None:
        onboarding_result = onboarding.run_onboarding_flow(script_dir)
        if onboarding_result == 'exit':
            logger.info('用户在引导流程中选择退出程序')
            return

    init_audit_db(get_audit_db_path(script_dir))
    init_default_alert_rules(script_dir)
    run_alert_detection(script_dir)
    logger.info('当前操作用户: %s', get_current_user())

    mode = ask_mode()
    if mode is None:
        show_info('提示', '未选择模式，程序退出。')
        logger.info('用户未选择模式，程序退出')
        return
    logger.info('用户选择模式: %s', mode)

    if mode == 'pipeline':
        run_pipeline_flow(script_dir)
    elif mode == 'diff':
        run_diff_flow(script_dir)
    elif mode == 'monitor':
        run_monitor_flow(script_dir)
    elif mode == 'scheduler':
        run_scheduler_flow(script_dir)
    elif mode == 'export':
        run_export_flow(script_dir)
    elif mode == 'db_query':
        run_db_query_flow(script_dir)
    elif mode == 'db_stats':
        run_db_stats_flow(script_dir)
    elif mode == 'batch_history':
        run_batch_history_flow(script_dir)
    elif mode == 'preset':
        run_preset_flow(script_dir)
    elif mode == 'subject_summary':
        run_subject_summary_flow(script_dir)
    elif mode == 'balance_check':
        run_balance_check_flow(script_dir)
    elif mode == 'duplicate_check':
        run_duplicate_check_flow(script_dir)
    elif mode == 'interest_fee_check':
        run_interest_fee_check_flow(script_dir)
    elif mode == 'balance_reconciliation':
        run_balance_reconciliation_flow(script_dir)
    elif mode == 'holiday_check':
        run_holiday_check_flow(script_dir)

    logger.info('========== 银行流水检验工具运行结束 ==========')


# ──────────────────────────────────────────────
# 数据库查询与统计流程
# ──────────────────────────────────────────────

def run_db_query_flow(script_dir):
    """数据库查询流程：按主体/账号/时间范围查询流水记录"""
    logger = get_logger()

    if not HAS_DATABASE:
        show_warning('错误', '数据库模块不可用，请检查 database.py 文件是否存在。')
        logger.error('数据库模块不可用')
        return

    print('\n' + '=' * 60)
    print('数据库查询 - 支持多条件组合查询')
    print('=' * 60)
    print('提示：直接回车表示不使用该条件')
    print()

    subject = input('请输入主体名称（支持模糊匹配）: ').strip()
    account = input('请输入银行账号（支持模糊匹配）: ').strip()
    bank = input('请输入银行名称（如：北京银行、东亚银行）: ').strip()
    start_date = input('请输入开始日期（YYYY-MM-DD）: ').strip()
    end_date = input('请输入结束日期（YYYY-MM-DD）: ').strip()

    min_amount_str = input('请输入最小金额（单位：元）: ').strip()
    max_amount_str = input('请输入最大金额（单位：元）: ').strip()
    counterpart = input('请输入对方户名（支持模糊匹配）: ').strip()
    summary_keyword = input('请输入摘要关键词（支持模糊匹配）: ').strip()

    limit_str = input('请输入返回记录数上限（默认 1000）: ').strip()
    limit = int(limit_str) if limit_str.isdigit() else 1000

    min_amount = float(min_amount_str) if min_amount_str else None
    max_amount = float(max_amount_str) if max_amount_str else None

    print(f'\n正在查询数据库...')

    try:
        result = db_module.query_transactions(
            subject=subject if subject else None,
            account=account if account else None,
            bank=bank if bank else None,
            start_date=start_date if start_date else None,
            end_date=end_date if end_date else None,
            min_amount=min_amount,
            max_amount=max_amount,
            counterpart=counterpart if counterpart else None,
            summary_keyword=summary_keyword if summary_keyword else None,
            limit=limit,
            script_dir=script_dir,
        )

        print(f'\n查询完成！共找到 {result.total_count} 条记录')

        if result.records:
            print(f'\n返回前 {len(result.records)} 条记录：')
            print('-' * 100)
            print(f'{"序号":<6}{"日期":<20}{"主体":<20}{"银行":<10}{"金额(元)":<15}{"对方户名":<20}')
            print('-' * 100)

            for i, record in enumerate(result.records[:50], 1):
                amount = record.收款 if (record.收款 and record.收款 > 0) else (record.付款 or 0)
                counterpart = record.对方户名 or ''
                if len(counterpart) > 18:
                    counterpart = counterpart[:16] + '...'

                print(f'{i:<6}{str(record.交易日期 or "")[:19]:<20}'
                      f'{str(record.主体 or "")[:18]:<20}'
                      f'{str(record.银行 or "")[:8]:<10}'
                      f'{amount:>12,.2f}  '
                      f'{counterpart:<20}')

            if len(result.records) > 50:
                print(f'\n... 还有 {len(result.records) - 50} 条记录未显示')

        if result.summary:
            print(f'\n{"=" * 60}')
            print('查询结果汇总：')
            print('-' * 60)
            print(f'记录总数：{result.summary.get("记录总数", 0)}')
            print(f'付款笔数：{result.summary.get("付款笔数", 0)}')
            print(f'收款笔数：{result.summary.get("收款笔数", 0)}')
            print(f'付款总额：{result.summary.get("付款总额", 0):,.2f} 元')
            print(f'收款总额：{result.summary.get("收款总额", 0):,.2f} 元')
            print(f'净　　额：{result.summary.get("净额", 0):,.2f} 元')

            bank_dist = result.summary.get('银行分布', {})
            if bank_dist:
                print(f'\n银行分布：')
                for bank_name, cnt in bank_dist.items():
                    print(f'  {bank_name}: {cnt} 条')

            subject_dist = result.summary.get('主体分布', {})
            if subject_dist:
                print(f'\n主体分布（前 10）：')
                for i, (subj_name, cnt) in enumerate(list(subject_dist.items())[:10], 1):
                    print(f'  {i}. {subj_name}: {cnt} 条')

        export = input(f'\n是否导出查询结果到 Excel？(y/N): ').strip().lower()
        if export == 'y':
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join(script_dir, f'查询结果_{timestamp}.xlsx')
            try:
                result.to_excel(output_path)
                show_info('导出成功', f'查询结果已导出到：\n{output_path}')
                logger.info('查询结果已导出: %s', output_path)
            except Exception as e:
                show_warning('导出失败', f'导出 Excel 失败：{e}')
                logger.error('导出查询结果失败: %s', e)

    except Exception as e:
        show_warning('查询失败', f'数据库查询失败：{e}')
        logger.error('数据库查询失败: %s', e, exc_info=True)


def run_db_stats_flow(script_dir):
    """数据库统计流程：查看数据汇总统计信息"""
    logger = get_logger()

    if not HAS_DATABASE:
        show_warning('错误', '数据库模块不可用，请检查 database.py 文件是否存在。')
        logger.error('数据库模块不可用')
        return

    print('\n' + '=' * 60)
    print('数据库统计信息')
    print('=' * 60)

    try:
        stats = db_module.get_db_statistics(script_dir=script_dir)

        print(f'\n【总体概览】')
        print(f'  总记录数：{stats.get("总记录数", 0):,} 条')
        print(f'  导入批次：{stats.get("导入批次数量", 0)} 次')

        date_range = stats.get('日期范围', {})
        if date_range:
            print(f'  最早交易：{date_range.get("最早交易日期", "无")}')
            print(f'  最晚交易：{date_range.get("最晚交易日期", "无")}')

        by_bank = stats.get('按银行统计', [])
        if by_bank:
            print(f'\n【按银行统计】')
            print('-' * 80)
            print(f'{"银行":<15}{"记录数":<12}{"付款总额(元)":<20}{"收款总额(元)":<20}')
            print('-' * 80)
            for row in by_bank:
                bank_name = row.get('银行', '未知')
                cnt = row.get('cnt', 0)
                payment = row.get('total_payment', 0) or 0
                receipt = row.get('total_receipt', 0) or 0
                print(f'{bank_name:<15}{cnt:<12,}{payment:>18,.2f}  {receipt:>18,.2f}')

        by_subject = stats.get('按主体统计', [])
        if by_subject:
            print(f'\n【按主体统计（前 15）】')
            print('-' * 90)
            print(f'{"序号":<6}{"主体":<30}{"记录数":<12}{"付款总额(元)":<20}{"收款总额(元)":<20}')
            print('-' * 90)
            for i, row in enumerate(by_subject[:15], 1):
                subject_name = str(row.get('主体', '未知'))[:28]
                cnt = row.get('cnt', 0)
                payment = row.get('total_payment', 0) or 0
                receipt = row.get('total_receipt', 0) or 0
                print(f'{i:<6}{subject_name:<30}{cnt:<12,}'
                      f'{payment:>18,.2f}  {receipt:>18,.2f}')

        by_account = stats.get('按账号统计', [])
        if by_account:
            print(f'\n【按账号统计（前 15）】')
            print('-' * 80)
            print(f'{"序号":<6}{"银行账号":<25}{"主体":<20}{"银行":<12}{"记录数":<10}')
            print('-' * 80)
            for i, row in enumerate(by_account[:15], 1):
                account = str(row.get('银行账号', '未知'))[:22]
                subject = str(row.get('主体', '未知'))[:18]
                bank = str(row.get('银行', ''))[:10]
                cnt = row.get('cnt', 0)
                print(f'{i:<6}{account:<25}{subject:<20}{bank:<12}{cnt:<10,}')

        by_date = stats.get('近30天交易趋势', [])
        if by_date:
            print(f'\n【近 30 天交易趋势】')
            print('-' * 60)
            print(f'{"日期":<15}{"记录数":<10}{"付款(元)":<18}{"收款(元)":<18}')
            print('-' * 60)
            for row in reversed(by_date):
                dt = str(row.get('dt', ''))[:10]
                cnt = row.get('cnt', 0)
                payment = row.get('total_payment', 0) or 0
                receipt = row.get('total_receipt', 0) or 0
                print(f'{dt:<15}{cnt:<10}{payment:>16,.2f}  {receipt:>16,.2f}')

        export = input(f'\n是否导出统计信息到 Excel？(y/N): ').strip().lower()
        if export == 'y':
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = os.path.join(script_dir, f'数据库统计_{timestamp}.xlsx')
            try:
                with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                    if by_bank:
                        pd.DataFrame(by_bank).to_excel(writer, sheet_name='按银行统计', index=False)
                    if by_subject:
                        pd.DataFrame(by_subject).to_excel(writer, sheet_name='按主体统计', index=False)
                    if by_account:
                        pd.DataFrame(by_account).to_excel(writer, sheet_name='按账号统计', index=False)
                    if by_date:
                        pd.DataFrame(by_date).to_excel(writer, sheet_name='近30天趋势', index=False)

                    overview_df = pd.DataFrame([{
                        '总记录数': stats.get('总记录数', 0),
                        '导入批次数量': stats.get('导入批次数量', 0),
                        '最早交易日期': date_range.get('最早交易日期', ''),
                        '最晚交易日期': date_range.get('最晚交易日期', ''),
                    }])
                    overview_df.to_excel(writer, sheet_name='总体概览', index=False)

                show_info('导出成功', f'统计信息已导出到：\n{output_path}')
                logger.info('数据库统计已导出: %s', output_path)
            except Exception as e:
                show_warning('导出失败', f'导出 Excel 失败：{e}')
                logger.error('导出统计信息失败: %s', e)

    except Exception as e:
        show_warning('统计失败', f'获取统计信息失败：{e}')
        logger.error('获取数据库统计失败: %s', e, exc_info=True)


# ──────────────────────────────────────────────
# 历史批次与版本管理流程
# ──────────────────────────────────────────────

def run_batch_history_flow(script_dir):
    """批次管理流程：查看历史批次与版本回溯"""
    logger = get_logger()

    if not HAS_BATCH_MANAGER:
        show_warning('错误', '批次管理模块不可用，请检查 batch_manager.py 文件是否存在。')
        logger.error('批次管理模块不可用')
        return

    try:
        batch_manager = batch_module.get_batch_manager(script_dir)
    except Exception as e:
        show_warning('错误', f'批次管理器初始化失败：{e}')
        logger.error('批次管理器初始化失败: %s', e, exc_info=True)
        return

    logger.info('========== 批次管理面板启动 ==========')

    while True:
        stats = batch_manager.get_statistics()

        print('\n' + '=' * 60)
        print('批次管理面板')
        print('=' * 60)
        print(f'总批次数: {stats["total_batches"]}')
        print(f'  ├─ 成功: {stats["success_batches"]}')
        print(f'  ├─ 失败: {stats["failed_batches"]}')
        print(f'  └─ 运行中: {stats["running_batches"]}')
        print(f'累计处理记录: {stats["total_records"]:,} 条')
        print(f'累计新增记录: {stats["total_new_records"]:,} 条')
        print('-' * 60)
        print('  1) 查看最近批次列表')
        print('  2) 按条件查询批次')
        print('  3) 查看批次详情')
        print('  4) 恢复批次文件')
        print('  5) 删除批次')
        print('  0) 返回主菜单')
        print('-' * 60)

        choice = input('请选择操作（0-5）: ').strip()

        if choice == '0':
            break
        elif choice == '1':
            _show_recent_batches(batch_manager)
        elif choice == '2':
            _query_batches(batch_manager)
        elif choice == '3':
            _show_batch_detail(batch_manager)
        elif choice == '4':
            _restore_batch(batch_manager)
        elif choice == '5':
            _delete_batch(batch_manager)
        else:
            print('无效选项，请重新选择。')

    logger.info('========== 批次管理面板关闭 ==========')


def _display_batch_list(batches):
    if not batches:
        print('\n未找到符合条件的批次。')
        return

    print('\n' + '-' * 100)
    print(f'{"序号":<5}{"批次号":<24}{"状态":<10}{"开始时间":<20}'
          f'{"记录数":<10}{"新增":<10}{"模式":<8}')
    print('-' * 100)

    for i, b in enumerate(batches, 1):
        mode = '增量' if b.incremental_mode else '全量'
        status = f'{b.status}'
        if status == 'success':
            status = '✅ 成功'
        elif status == 'warning':
            status = '⚠️ 警告'
        elif status == 'failed':
            status = '❌ 失败'
        elif status == 'running':
            status = '⏳ 运行中'
        print(f'{i:<5}{b.batch_id:<24}{status:<10}{b.start_time:<20}'
              f'{b.total_records:<10,}{b.new_records:<10,}{mode:<8}')
    print('-' * 100)


def _show_recent_batches(batch_manager):
    limit_str = input('请输入显示数量（默认 20）: ').strip()
    limit = int(limit_str) if limit_str.isdigit() else 20

    batches = batch_manager.query_batches(limit=limit)
    _display_batch_list(batches)


def _query_batches(batch_manager):
    print('\n查询条件（直接回车表示不使用该条件）:')
    start_date = input('开始日期（YYYY-MM-DD）: ').strip() or None
    end_date = input('结束日期（YYYY-MM-DD）: ').strip() or None
    status = input('状态（success/warning/failed/running）: ').strip() or None
    operator = input('操作员: ').strip() or None
    min_records_str = input('最小记录数: ').strip()
    min_records = int(min_records_str) if min_records_str.isdigit() else None

    batches = batch_manager.query_batches(
        start_date=start_date,
        end_date=end_date,
        status=status,
        operator=operator,
        min_records=min_records,
        limit=100,
    )
    _display_batch_list(batches)


def _show_batch_detail(batch_manager):
    batch_id = input('请输入批次号: ').strip()
    if not batch_id:
        print('未输入批次号。')
        return

    detail = batch_manager.get_batch_detail(batch_id)
    if not detail:
        print(f'\n批次不存在: {batch_id}')
        return

    info = detail['batch_info']
    metadata = detail['metadata']
    files = detail['files']

    status = info.get('status', '')
    if status == 'success':
        status_icon = '✅'
    elif status == 'warning':
        status_icon = '⚠️'
    elif status == 'failed':
        status_icon = '❌'
    else:
        status_icon = '⏳'

    mode = '增量合并' if info.get('incremental_mode') else '全量覆盖'

    print('\n' + '=' * 60)
    print(f'批次详情 - {batch_id}')
    print('=' * 60)
    print(f'状态: {status_icon} {status}')
    print(f'开始时间: {info.get("start_time", "")}')
    print(f'结束时间: {info.get("end_time", "")}')
    print(f'操作员: {info.get("operator", "未指定")}')
    print(f'输入文件夹: {info.get("input_folder", "")}')
    print(f'运行模式: {mode}')
    print('-' * 40)
    print('📊 处理统计:')
    print(f'  总记录数: {info.get("total_records", 0):,}')
    print(f'  新增记录: {info.get("new_records", 0):,}')
    print(f'  重复记录: {info.get("duplicate_records", 0):,}')
    print(f'  处理文件: {info.get("processed_files", 0)} 个')
    print(f'  未识别文件: {info.get("unprocessed_files", 0)} 个')
    print(f'  出错文件: {info.get("error_files", 0)} 个')
    print('-' * 40)

    processed_files = metadata.get('processed_files', [])
    if processed_files:
        print('📄 已处理文件:')
        for f in processed_files:
            print(f'  - {os.path.basename(f)}')

    unprocessed_files = metadata.get('unprocessed_files', [])
    if unprocessed_files:
        print('\n⚠️  未识别文件:')
        for f in unprocessed_files:
            print(f'  - {os.path.basename(f)}')

    error_files = metadata.get('error_files', [])
    if error_files:
        print('\n❌ 出错文件:')
        for f, err in error_files:
            print(f'  - {os.path.basename(f)}: {err}')

    if info.get('error_message'):
        print(f'\n❌ 错误信息:\n{info["error_message"]}')

    print('-' * 40)
    print('📁 归档文件:')
    for name, path in files.items():
        print(f'  - {name}')
        print(f'    路径: {path}')

    print(f'\n📂 归档目录: {info.get("batch_dir", "")}')
    print('=' * 60)


def _restore_batch(batch_manager):
    batch_id = input('请输入要恢复的批次号: ').strip()
    if not batch_id:
        print('未输入批次号。')
        return

    target_dir = input('请输入恢复目录（直接回车使用默认目录）: ').strip() or None

    try:
        restored = batch_manager.restore_batch(batch_id, target_dir)
        print(f'\n✅ 批次 {batch_id} 恢复成功！')
        print(f'恢复目录: {os.path.dirname(next(iter(restored.values()))) if restored else target_dir}')
        print('\n恢复的文件:')
        for name in restored.keys():
            print(f'  - {name}')
    except ValueError as e:
        print(f'\n❌ {e}')
    except Exception as e:
        print(f'\n❌ 恢复失败: {e}')
        get_logger().error('批次恢复失败: %s', e, exc_info=True)


def _delete_batch(batch_manager):
    batch_id = input('请输入要删除的批次号: ').strip()
    if not batch_id:
        print('未输入批次号。')
        return

    confirm = input(f'确认要删除批次 {batch_id} 吗？此操作不可恢复！(yes/N): ').strip().lower()
    if confirm != 'yes':
        print('已取消删除。')
        return

    try:
        success = batch_manager.delete_batch(batch_id)
        if success:
            print(f'\n✅ 批次 {batch_id} 已删除。')
        else:
            print(f'\n❌ 批次不存在: {batch_id}')
    except Exception as e:
        print(f'\n❌ 删除失败: {e}')
        get_logger().error('批次删除失败: %s', e, exc_info=True)


# ──────────────────────────────────────────────
# 任务配置预设管理模块
# ──────────────────────────────────────────────

PRESET_CONFIG_FILENAME = 'task_presets.json'


@dataclass
class TaskPreset:
    preset_id: str
    name: str
    description: str = ''
    output_dir: str = ''
    start_date: str = ''
    end_date: str = ''
    keep_strategy: str = 'keep_unprocessed'
    enabled_banks: List[str] = field(default_factory=list)
    incremental: bool = True
    created_at: str = ''
    updated_at: str = ''


KEEP_STRATEGIES = {
    'keep_unprocessed': '仅保留未处理文件',
    'keep_all': '保留所有文件',
    'delete_all': '删除所有已处理文件',
    'move_to_archive': '归档已处理文件到「已处理归档」子目录',
}


def get_preset_config_path(script_dir=None):
    if script_dir is None:
        script_dir = get_script_dir()
    return os.path.join(script_dir, PRESET_CONFIG_FILENAME)


def load_preset_config(script_dir=None):
    config_path = get_preset_config_path(script_dir)
    logger = get_logger()

    if not os.path.exists(config_path):
        logger.info('预设配置文件不存在，将创建默认配置: %s', config_path)
        default_config = {
            'presets': [],
            'settings': {
                'default_preset': '',
            }
        }
        save_preset_config(default_config, script_dir)
        return default_config

    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            config = json.load(f)
        logger.info('已加载预设配置，共 %d 个预设', len(config.get('presets', [])))
        return config
    except Exception as e:
        logger.error('加载预设配置失败: %s', e)
        return {'presets': [], 'settings': {}}


def save_preset_config(config, script_dir=None):
    config_path = get_preset_config_path(script_dir)
    try:
        with open(config_path, 'w', encoding='utf-8') as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
        logger = get_logger()
        logger.info('预设配置已保存: %s', config_path)
        return True
    except Exception as e:
        logger = get_logger()
        logger.error('保存预设配置失败: %s', e)
        return False


def _generate_preset_id():
    return f"PRESET{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"


def save_preset(preset_data, script_dir=None):
    config = load_preset_config(script_dir)
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    preset_id = preset_data.get('preset_id') or _generate_preset_id()
    preset_data['preset_id'] = preset_id

    if not preset_data.get('name'):
        preset_data['name'] = f'预设_{datetime.now().strftime("%Y%m%d_%H%M%S")}'

    preset_data['enabled_banks'] = preset_data.get('enabled_banks') or list(BANK_PREFIXES)
    preset_data['keep_strategy'] = preset_data.get('keep_strategy') or 'keep_unprocessed'
    preset_data['incremental'] = preset_data.get('incremental', True)

    existing = None
    for i, p in enumerate(config['presets']):
        if p['preset_id'] == preset_id:
            existing = i
            break

    if existing is not None:
        preset_data['created_at'] = config['presets'][existing].get('created_at', now)
        preset_data['updated_at'] = now
        config['presets'][existing] = preset_data
        action = '更新'
    else:
        preset_data['created_at'] = now
        preset_data['updated_at'] = now
        config['presets'].append(preset_data)
        action = '添加'

    save_preset_config(config, script_dir)

    logger = get_logger()
    logger.info('已%s预设 [%s] %s', action, preset_id, preset_data.get('name', ''))
    return preset_id


def load_preset(preset_id, script_dir=None):
    config = load_preset_config(script_dir)
    for preset in config.get('presets', []):
        if preset['preset_id'] == preset_id:
            return preset
    return None


def delete_preset(preset_id, script_dir=None):
    config = load_preset_config(script_dir)
    logger = get_logger()

    original_count = len(config['presets'])
    config['presets'] = [p for p in config['presets'] if p['preset_id'] != preset_id]

    if len(config['presets']) == original_count:
        logger.warning('未找到预设 [%s]', preset_id)
        return False

    if config.get('settings', {}).get('default_preset') == preset_id:
        config['settings']['default_preset'] = ''

    save_preset_config(config, script_dir)
    logger.info('已删除预设 [%s]', preset_id)
    return True


def list_presets(script_dir=None):
    config = load_preset_config(script_dir)
    return config.get('presets', [])


def set_default_preset(preset_id, script_dir=None):
    config = load_preset_config(script_dir)
    if 'settings' not in config:
        config['settings'] = {}
    config['settings']['default_preset'] = preset_id
    save_preset_config(config, script_dir)


def get_default_preset(script_dir=None):
    config = load_preset_config(script_dir)
    default_id = config.get('settings', {}).get('default_preset', '')
    if default_id:
        return load_preset(default_id, script_dir)
    return None


def apply_preset_to_pipeline(preset, folder, script_dir, dry_run=False):
    logger = get_logger()

    if not preset:
        logger.warning('预设为空，使用默认配置运行')
        return run_pipeline(folder, script_dir, dry_run=dry_run)

    logger.info('应用预设 [%s] %s (dry_run=%s)', preset.get('preset_id'), preset.get('name', ''), dry_run)

    enabled_banks = preset.get('enabled_banks', BANK_PREFIXES)
    keep_strategy = preset.get('keep_strategy', 'keep_unprocessed')
    incremental = preset.get('incremental', True)
    start_date = preset.get('start_date', '')
    end_date = preset.get('end_date', '')
    output_dir = preset.get('output_dir', '') or None

    result = run_pipeline_with_options(
        folder=folder,
        script_dir=script_dir,
        incremental=incremental,
        enabled_banks=enabled_banks,
        keep_strategy=keep_strategy,
        start_date=start_date,
        end_date=end_date,
        output_dir=output_dir,
        dry_run=dry_run,
    )

    return result


def run_pipeline_with_options(folder, script_dir, incremental=True,
                              enabled_banks=None, keep_strategy='keep_unprocessed',
                              start_date='', end_date='', batch_id=None, output_dir=None,
                              enable_encryption=False, encryption_password=None, encryption_mode='excel_password',
                              dry_run=False, archive_dir_name='已处理归档'):
    logger = get_logger()
    if dry_run:
        logger.info('===== 试运行模式已启用（不执行删除与写盘操作）=====')

    if enabled_banks is None:
        enabled_banks = BANK_PREFIXES

    lookup_file = find_lookup_file(script_dir)
    lookup_missing = lookup_file is None
    if lookup_missing:
        logger.warning('未找到主体查找表，"主体"列将为空')
        lookup_data = load_lookup_table(None)
    else:
        logger.info('正在预加载主体查找表...')
        lookup_data = load_lookup_table(lookup_file)
        logger.info('主体查找表预加载完成')

    existing_keys = set()
    existing_records = []
    actual_incremental = False
    duplicate_count = 0
    new_record_count = 0

    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        logger.info('使用自定义输出目录: %s', output_dir)

    if incremental:
        summary_path = get_summary_table_path(script_dir, output_dir)
        existing_keys, existing_records = load_existing_keys(summary_path)
        actual_incremental = len(existing_records) > 0
        if actual_incremental:
            logger.info('===== 增量合并模式已启用 =====')
        else:
            logger.info('无历史数据，将以全量模式运行')

    folder_name = os.path.basename(folder.rstrip('/\\'))
    parent_dir = os.path.dirname(folder.rstrip('/\\'))
    new_folder = os.path.join(parent_dir, f"{folder_name}＋检验版")

    if os.path.exists(new_folder):
        logger.info('＋检验版文件夹已存在，先删除: %s', new_folder)
        shutil.rmtree(new_folder)
    shutil.copytree(folder, new_folder)
    logger.info('已复制文件夹为＋检验版: %s', new_folder)

    excel_files = scan_excel_files(new_folder)
    if not excel_files:
        logger.warning('检验版文件夹中未发现任何 Excel 文件')
        return ProcessingResult(
            lookup_missing=lookup_missing,
            folder_empty=True,
            incremental_mode=actual_incremental,
            existing_record_count=len(existing_records),
            dry_run=dry_run,
            pending_script_dir=script_dir if dry_run else None,
            pending_output_dir=output_dir if dry_run else None,
            pending_input_folder=folder if dry_run else None,
        )

    def _parse_date(value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value
        try:
            if isinstance(value, str):
                return datetime.strptime(value[:10], '%Y-%m-%d')
        except (ValueError, TypeError):
            pass
        return None

    start_dt = _parse_date(start_date)
    end_dt = _parse_date(end_date)

    if start_dt:
        logger.info('日期过滤 - 开始日期: %s', start_dt.strftime('%Y-%m-%d'))
    if end_dt:
        logger.info('日期过滤 - 结束日期: %s', end_dt.strftime('%Y-%m-%d'))

    def _is_date_in_range(trade_date):
        if trade_date is None:
            return True
        dt = _parse_date(trade_date)
        if dt is None:
            return True
        if start_dt and dt < start_dt:
            return False
        if end_dt and dt > end_dt:
            return False
        return True

    all_rows = []
    processed_files = []
    unprocessed_files = []
    error_files = []
    filtered_out_count = 0

    for filepath in excel_files:
        bank = identify_bank(filepath)
        if bank and bank in BANK_PROCESSORS and bank in enabled_banks:
            try:
                processor = BANK_PROCESSORS[bank]
                rows = processor(filepath, lookup_data)

                if start_dt or end_dt:
                    original_len = len(rows)
                    rows = [r for r in rows if _is_date_in_range(r.get('交易日期'))]
                    filtered_out_count += (original_len - len(rows))

                all_rows.extend(rows)
                processed_files.append(filepath)
                logger.info('成功处理文件: %s（%d 条记录）', filepath, len(rows))
            except Exception as e:
                error_files.append((filepath, str(e)))
                logger.error('处理文件「%s」时发生错误: %s', filepath, e, exc_info=True)
        else:
            unprocessed_files.append(filepath)
            if bank and bank not in enabled_banks:
                logger.info('文件「%s」所属银行「%s」不在启用列表中，跳过', filepath, bank)

    if filtered_out_count > 0:
        logger.info('日期过滤共排除 %d 条记录', filtered_out_count)

    error_file_paths = {f for f, _ in error_files}
    keep_set = set(unprocessed_files) | error_file_paths

    if keep_strategy == 'keep_all':
        keep_set = set(excel_files)
        logger.info('保留策略：保留所有文件')
    elif keep_strategy == 'delete_all':
        keep_set = set()
        logger.info('保留策略：删除所有已处理文件')
    elif keep_strategy == 'move_to_archive':
        logger.info('保留策略：归档已处理文件到「%s」子目录', archive_dir_name)
    else:
        logger.info('保留策略：仅保留未处理文件')

    pending_deletion_files = [f for f in excel_files if f not in keep_set] if keep_strategy != 'move_to_archive' else [
        f for f in processed_files if f not in error_file_paths
    ]
    if dry_run:
        logger.info('[试运行] 跳过删除文件操作，待删除文件 %d 个', len(pending_deletion_files))
    else:
        delete_processed_files(
            excel_files, processed_files, error_files, unprocessed_files,
            strategy=keep_strategy, archive_dir_name=archive_dir_name
        )

    output_path = None
    final_rows = []
    incremental_rows = []
    _cp_tag_summary = {}
    _it_summary = {}
    _it_result = None

    if all_rows:
        if actual_incremental:
            incremental_rows, duplicate_count = filter_incremental_records(all_rows, existing_keys)
            new_record_count = len(incremental_rows)
            if dry_run:
                logger.info('[试运行] 跳过总表写盘，模式=增量合并，历史 %d 条 + 新增 %d 条',
                            len(existing_records), len(incremental_rows))
                final_rows = existing_records + incremental_rows
            else:
                output_path = merge_and_export_summary(
                    existing_records, incremental_rows, script_dir, output_dir, lookup_source=lookup_data
                )
                final_rows = existing_records + incremental_rows
        else:
            if dry_run:
                logger.info('[试运行] 跳过总表写盘，模式=全量覆盖，共 %d 条记录', len(all_rows))
                final_rows = all_rows
            else:
                columns = get_summary_columns(all_rows, lookup_data)
                df = pd.DataFrame(all_rows, columns=columns)
                output_path = get_summary_table_path(script_dir, output_dir)
                if output_dir:
                    os.makedirs(output_dir, exist_ok=True)
                df.to_excel(output_path, index=False, engine='openpyxl')
                logger.info('总表输出完成: %s（共 %d 条记录）', output_path, len(all_rows))
                final_rows = all_rows
            new_record_count = len(all_rows)
    else:
        logger.warning('未提取到任何银行流水记录')
        if existing_records:
            if dry_run:
                logger.info('[试运行] 跳过总表写盘，仅使用历史记录 %d 条', len(existing_records))
                final_rows = existing_records
            else:
                output_path = merge_and_export_summary(
                    existing_records, [], script_dir, output_dir, lookup_source=lookup_data
                )
                final_rows = existing_records

    if final_rows:
        final_rows, _anomaly_summary = apply_amount_anomaly_detection(final_rows)
        if _anomaly_summary.get('anomaly_count', 0) > 0:
            logger.info('金额异常检测: 总记录 %d, 异常 %d (%.2f%%)',
                        _anomaly_summary.get('total_records', 0),
                        _anomaly_summary.get('anomaly_count', 0),
                        _anomaly_summary.get('anomaly_rate', 0) * 100)
            if output_path and not dry_run:
                base_columns = get_summary_columns(final_rows, lookup_data)
                _anomaly_columns = base_columns
                pd.DataFrame(final_rows, columns=_anomaly_columns).to_excel(
                    output_path, index=False, engine='openpyxl')
                logger.info('已将金额异常检测结果回写到总表: %s', output_path)
            elif dry_run:
                logger.info('[试运行] 跳过金额异常检测结果回写总表')

    if final_rows:
        final_rows, _cp_tag_summary = apply_counterparty_rules(final_rows, script_dir)
        if _cp_tag_summary.get('tagged_count', 0) > 0:
            logger.info('对方户名黑白名单打标: 总记录 %d, 命中 %d (黑名单 %d, 白名单 %d)',
                        _cp_tag_summary.get('total_records', 0),
                        _cp_tag_summary.get('tagged_count', 0),
                        _cp_tag_summary.get('blacklist_hits', 0),
                        _cp_tag_summary.get('whitelist_hits', 0))
            if output_path and not dry_run:
                base_columns = get_summary_columns(final_rows, lookup_data)
                cp_extra_cols = ['黑白名单标签', '命中规则名称', '命中关键词']
                _cp_columns = base_columns + [
                    col for col in cp_extra_cols if col not in base_columns
                ]
                pd.DataFrame(final_rows, columns=_cp_columns).to_excel(
                    output_path, index=False, engine='openpyxl')
                logger.info('已将黑白名单打标结果回写到总表: %s', output_path)
            elif dry_run:
                logger.info('[试运行] 跳过黑白名单打标结果回写总表')

    masked_output_path = None
    if final_rows and output_path and not dry_run:
        try:
            output_dir = os.path.dirname(output_path) or script_dir
            _masked_columns = get_summary_columns(final_rows, lookup_data)
            if _cp_tag_summary and _cp_tag_summary.get('tagged_count', 0) > 0:
                cp_extra_cols = ['黑白名单标签', '命中规则名称', '命中关键词']
                _masked_columns = _masked_columns + [
                    col for col in cp_extra_cols if col not in _masked_columns
                ]
            else:
                cp_extra_cols = ['黑白名单标签', '命中规则名称', '命中关键词']
                _masked_columns = [col for col in _masked_columns if col not in cp_extra_cols]
            masked_output_path = export_masked_summary(
                final_rows, script_dir, output_dir=output_dir,
                lookup_source=lookup_data, columns=_masked_columns
            )
        except Exception as e:
            logger.error('生成脱敏版总表失败: %s', e, exc_info=True)
            masked_output_path = None
    elif final_rows and dry_run:
        logger.info('[试运行] 跳过生成脱敏版总表')

    db_inserted = 0
    db_duplicates = 0
    if HAS_DATABASE and final_rows and not dry_run:
        try:
            if batch_id is None:
                batch_id = f"BATCH{datetime.now().strftime('%Y%m%d%H%M%S')}"
            db_inserted, db_duplicates = db_module.persist_transactions(
                final_rows,
                batch_id=batch_id,
                deduplicate=True,
                script_dir=script_dir,
            )
            logger.info(
                '数据库持久化完成: 批次 %s, 插入 %d 条, 去重跳过 %d 条',
                batch_id, db_inserted, db_duplicates,
            )
        except Exception as e:
            logger.error('数据库持久化失败: %s', e, exc_info=True)
    elif HAS_DATABASE and final_rows and dry_run:
        logger.info('[试运行] 跳过数据库持久化操作')

    internal_transfer_path = None
    if final_rows:
        try:
            final_rows, _it_summary, _it_result = identify_and_tag_internal_transfers(
                final_rows,
            )
            if _it_summary.get('match_pairs', 0) > 0:
                logger.info(
                    '跨账号内部划转识别: 总记录 %d, 识别 %d 对 (划出 %d + 划入 %d), '
                    '涉及 %d 主体 %d 银行, 划转总金额 %.2f 元',
                    _it_summary.get('total_records', 0),
                    _it_summary.get('match_pairs', 0),
                    _it_summary.get('marked_out_count', 0),
                    _it_summary.get('marked_in_count', 0),
                    len(_it_summary.get('involved_subjects', [])),
                    len(_it_summary.get('involved_banks', [])),
                    _it_summary.get('total_transfer_amount', 0.0),
                )
                if output_path and not dry_run:
                    base_columns = get_summary_columns(final_rows, lookup_data)
                    it_extra_cols = list(INTERNAL_TRANSFER_EXTRA_COLUMNS)
                    _it_columns = base_columns + [
                        col for col in it_extra_cols if col not in base_columns
                    ]
                    pd.DataFrame(final_rows, columns=_it_columns).to_excel(
                        output_path, index=False, engine='openpyxl')
                    logger.info('已将内部划转标记回写到总表: %s', output_path)
                elif dry_run:
                    logger.info('[试运行] 跳过内部划转标记回写总表')

                _it_out_dir = script_dir
                if output_path:
                    _it_out_dir = os.path.dirname(output_path) or _it_out_dir
                if output_dir:
                    _it_out_dir = output_dir or _it_out_dir
                _it_src_info = {
                    '数据来源': '主流程自动生成',
                    '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                    '记录数': len(final_rows),
                    '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                    '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                }
                if _it_result.match_pairs > 0:
                    _it_ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                    _it_out_path = os.path.join(_it_out_dir, f'内部划转识别报告_{_it_ts}.xlsx')
                    internal_transfer_path = export_internal_transfer_report(
                        _it_result, _it_out_path, _it_src_info,
                    )
                    if internal_transfer_path:
                        logger.info('内部划转识别报告已自动生成: %s', internal_transfer_path)
        except Exception as e:
            logger.error('内部划转识别处理失败: %s', e, exc_info=True)
            internal_transfer_path = None

    subject_summary_path = None
    balance_check_path = None
    if final_rows:
        try:
            output_dir_for_summary = output_dir or script_dir
            if output_path:
                output_dir_for_summary = os.path.dirname(output_path) or output_dir_for_summary
            source_info = {
                '数据来源': '主流程自动生成(预设)',
                '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                '记录数': len(final_rows),
                '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                '启用银行': ', '.join(enabled_banks) if enabled_banks else '全部',
                '日期范围': f'{start_date or "不限"} ~ {end_date or "不限"}',
                '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            subject_summary_path = generate_subject_summary_from_records(
                final_rows, output_dir_for_summary, source_info
            )
            if subject_summary_path:
                logger.info('主体维度汇总分析已自动生成: %s', subject_summary_path)
        except Exception as e:
            logger.error('自动生成主体汇总分析失败: %s', e, exc_info=True)
            subject_summary_path = None

        try:
            output_dir_for_check = output_dir or script_dir
            if output_path:
                output_dir_for_check = os.path.dirname(output_path) or output_dir_for_check
            source_info = {
                '数据来源': '主流程自动生成(预设)',
                '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                '记录数': len(final_rows),
                '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                '启用银行': ', '.join(enabled_banks) if enabled_banks else '全部',
                '日期范围': f'{start_date or "不限"} ~ {end_date or "不限"}',
                '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            balance_check_path = generate_balance_check_from_records(
                final_rows, output_dir_for_check, source_info
            )
            if balance_check_path:
                logger.info('余额连续性校验报告已自动生成: %s', balance_check_path)
        except Exception as e:
            logger.error('自动生成余额连续性校验报告失败: %s', e, exc_info=True)
            balance_check_path = None

    duplicate_check_path = None
    if final_rows:
        try:
            output_dir_for_check = output_dir or script_dir
            if output_path:
                output_dir_for_check = os.path.dirname(output_path) or output_dir_for_check
            source_info = {
                '数据来源': '主流程自动生成(预设)',
                '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                '记录数': len(final_rows),
                '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                '启用银行': ', '.join(enabled_banks) if enabled_banks else '全部',
                '日期范围': f'{start_date or "不限"} ~ {end_date or "不限"}',
                '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            duplicate_check_path = generate_duplicate_check_from_records(
                final_rows, output_dir_for_check, source_info
            )
            if duplicate_check_path:
                logger.info('重复交易检测报告已自动生成: %s', duplicate_check_path)
        except Exception as e:
            logger.error('自动生成重复交易检测报告失败: %s', e, exc_info=True)
            duplicate_check_path = None

    interest_fee_check_path_opt = None
    if final_rows:
        try:
            output_dir_for_check = output_dir or script_dir
            if output_path:
                output_dir_for_check = os.path.dirname(output_path) or output_dir_for_check
            source_info = {
                '数据来源': '主流程自动生成(预设)',
                '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                '记录数': len(final_rows),
                '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                '启用银行': ', '.join(enabled_banks) if enabled_banks else '全部',
                '日期范围': f'{start_date or "不限"} ~ {end_date or "不限"}',
                '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            interest_fee_check_path_opt = generate_interest_fee_check_from_records(
                final_rows, output_dir_for_check, source_info
            )
            if interest_fee_check_path_opt:
                logger.info('利息手续费核对报告已自动生成: %s', interest_fee_check_path_opt)
        except Exception as e:
            logger.error('自动生成利息手续费核对报告失败: %s', e, exc_info=True)
            interest_fee_check_path_opt = None

    holiday_check_path_opt = None
    if final_rows:
        try:
            output_dir_for_check = output_dir or script_dir
            if output_path:
                output_dir_for_check = os.path.dirname(output_path) or output_dir_for_check
            source_info = {
                '数据来源': '主流程自动生成(预设)',
                '总表文件': os.path.basename(output_path) if output_path else '内存数据',
                '记录数': len(final_rows),
                '运行模式': '增量合并' if actual_incremental else '全量覆盖',
                '启用银行': ', '.join(enabled_banks) if enabled_banks else '全部',
                '日期范围': f'{start_date or "不限"} ~ {end_date or "不限"}',
                '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            }
            holiday_check_path_opt = generate_holiday_check_from_records(
                final_rows, output_dir_for_check, source_info
            )
            if holiday_check_path_opt:
                logger.info('非工作日交易标记报告已自动生成: %s', holiday_check_path_opt)
        except Exception as e:
            logger.error('自动生成非工作日交易标记报告失败: %s', e, exc_info=True)
            holiday_check_path_opt = None

    encryption_result = None
    encrypted_files = []

    if enable_encryption and encryption_password and HAS_FILE_ENCRYPTION and not dry_run:
        try:
            enc_output_dir = output_dir or script_dir
            if output_path:
                enc_output_dir = os.path.dirname(output_path) or script_dir

            files_to_encrypt = []
            for fp in [output_path, subject_summary_path, balance_check_path]:
                if fp and os.path.isfile(fp):
                    files_to_encrypt.append(fp)

            if files_to_encrypt:
                encryption_result = _encrypt_output_files(
                    files_to_encrypt,
                    password=encryption_password,
                    mode=encryption_mode,
                    output_dir=enc_output_dir,
                )
                encrypted_files = [
                    r.encrypted_path for r in encryption_result.results
                    if r.success and r.encrypted_path
                ]
                _save_encryption_record(encryption_result, script_dir=script_dir)
                logger.info('输出文件加密完成: 模式=%s, 成功=%d, 失败=%d',
                            encryption_mode,
                            encryption_result.success_count,
                            encryption_result.failure_count)
        except Exception as e:
            logger.error('输出文件加密失败: %s', e, exc_info=True)
            encryption_result = None
    elif enable_encryption and dry_run:
        logger.info('[试运行] 跳过输出文件加密操作')

    pending_list_path = None
    if (unprocessed_files or error_files) and not dry_run:
        pending_list_path = generate_pending_list_xlsx(
            unprocessed_files, error_files, script_dir, output_dir
        )
    elif (unprocessed_files or error_files) and dry_run:
        logger.info('[试运行] 跳过待处理清单生成')

    return ProcessingResult(
        all_rows=final_rows,
        processed_files=processed_files,
        unprocessed_files=unprocessed_files,
        error_files=error_files,
        output_path=output_path,
        masked_output_path=masked_output_path,
        subject_summary_path=subject_summary_path,
        balance_check_path=balance_check_path,
        duplicate_check_path=duplicate_check_path,
        interest_fee_check_path=interest_fee_check_path_opt,
        holiday_check_path=holiday_check_path_opt,
        pending_list_path=pending_list_path,
        lookup_missing=lookup_missing,
        incremental_mode=actual_incremental,
        existing_record_count=len(existing_records),
        new_record_count=new_record_count,
        duplicate_record_count=duplicate_count,
        db_inserted_count=db_inserted,
        db_duplicate_count=db_duplicates,
        encryption_result=encryption_result,
        encrypted_files=encrypted_files,
        dry_run=dry_run,
        pending_deletion_files=pending_deletion_files,
        pending_keep_set=keep_set,
        pending_all_files=list(excel_files),
        pending_final_rows=final_rows,
        pending_existing_records=existing_records,
        pending_incremental_rows=incremental_rows,
        pending_script_dir=script_dir,
        pending_output_dir=output_dir,
        pending_lookup_source=lookup_data,
        pending_enable_encryption=enable_encryption,
        pending_encryption_password=encryption_password,
        pending_encryption_mode=encryption_mode,
        pending_batch_id=batch_id,
        pending_input_folder=folder,
        pending_cp_tag_summary=_cp_tag_summary,
        pending_internal_transfer_summary=_it_summary,
        pending_internal_transfer_result=_it_result,
        pending_keep_strategy=keep_strategy,
        pending_archive_dir_name=archive_dir_name,
    )


def run_preset_flow(script_dir):
    """预设管理流程：管理和应用任务配置预设"""
    logger = get_logger()

    if HAS_TKINTER and tk is not None:
        try:
            gui_preset_manager(script_dir)
            return
        except Exception as e:
            logger.warning('GUI预设管理启动失败，将使用命令行模式: %s', e)

    print('\n' + '=' * 60)
    print('任务配置预设管理')
    print('=' * 60)

    while True:
        presets = list_presets(script_dir)
        default_preset = get_default_preset(script_dir)

        print(f'\n当前预设数量: {len(presets)}')
        if default_preset:
            print(f'默认预设: [{default_preset["preset_id"]}] {default_preset.get("name", "")}')

        print('\n请选择操作：')
        print('  1) 列出所有预设')
        print('  2) 保存当前配置为新预设')
        print('  3) 加载并应用预设')
        print('  4) 删除预设')
        print('  5) 设为默认预设')
        print('  6) 查看预设详情')
        print('  0) 返回主菜单')

        choice = input('\n请输入选项: ').strip()

        if choice == '1':
            _list_presets_cli(presets)
        elif choice == '2':
            _save_preset_cli(script_dir)
        elif choice == '3':
            _apply_preset_cli(script_dir)
        elif choice == '4':
            _delete_preset_cli(script_dir)
        elif choice == '5':
            _set_default_preset_cli(script_dir)
        elif choice == '6':
            _show_preset_detail_cli(script_dir)
        elif choice == '0':
            break
        else:
            print('无效选项，请重新输入')


def gui_preset_manager(script_dir):
    """GUI预设管理窗口"""
    if tk is None:
        raise RuntimeError('Tkinter not available')

    logger = get_logger()
    logger.info('启动GUI预设管理')

    root = tk.Tk()
    root.title('任务配置预设管理')
    root.geometry('750x600')
    root.minsize(700, 550)

    presets_list = []
    selected_preset_id = tk.StringVar()
    detail_text = None

    def refresh_presets():
        nonlocal presets_list
        presets_list = list_presets(script_dir)
        listbox.delete(0, tk.END)
        default_id = get_default_preset(script_dir)
        default_id = default_id['preset_id'] if default_id else ''
        for p in presets_list:
            marker = '★ ' if p['preset_id'] == default_id else '  '
            listbox.insert(tk.END, f"{marker}{p['preset_id']} - {p.get('name', '未命名')}")

    def on_select(event):
        selection = listbox.curselection()
        if not selection:
            return
        idx = selection[0]
        preset = presets_list[idx]
        selected_preset_id.set(preset['preset_id'])
        show_preset_detail(preset)

    def show_preset_detail(preset):
        if detail_text is None:
            return
        detail_text.config(state=tk.NORMAL)
        detail_text.delete(1.0, tk.END)

        banks = ', '.join(preset.get('enabled_banks', []))
        keep = KEEP_STRATEGIES.get(preset.get('keep_strategy'), '未知')
        incremental = '是' if preset.get('incremental', True) else '否'

        detail = f"""预设名称: {preset.get('name', '')}
预设ID: {preset.get('preset_id', '')}
描述: {preset.get('description', '无')}
{'-' * 50}
输出目录: {preset.get('output_dir', '默认')}
开始日期: {preset.get('start_date', '不限制')}
结束日期: {preset.get('end_date', '不限制')}
保留策略: {keep}
启用银行: {banks}
增量合并: {incremental}
{'-' * 50}
创建时间: {preset.get('created_at', '')}
更新时间: {preset.get('updated_at', '')}
"""
        detail_text.insert(tk.END, detail)
        detail_text.config(state=tk.DISABLED)

    def add_preset():
        PresetEditorDialog(root, script_dir, on_saved=refresh_presets)

    def edit_preset():
        pid = selected_preset_id.get()
        if not pid:
            messagebox.showwarning('提示', '请先选择一个预设')
            return
        preset = load_preset(pid, script_dir)
        if preset:
            PresetEditorDialog(root, script_dir, preset=preset, on_saved=refresh_presets)

    def delete_selected():
        pid = selected_preset_id.get()
        if not pid:
            messagebox.showwarning('提示', '请先选择一个预设')
            return
        preset = load_preset(pid, script_dir)
        if not preset:
            return
        if messagebox.askyesno('确认删除', f'确定要删除预设「{preset.get("name", "")}」吗？'):
            if delete_preset(pid, script_dir):
                messagebox.showinfo('成功', '预设已删除')
                refresh_presets()
                selected_preset_id.set('')
                if detail_text:
                    detail_text.config(state=tk.NORMAL)
                    detail_text.delete(1.0, tk.END)
                    detail_text.config(state=tk.DISABLED)
            else:
                messagebox.showerror('错误', '删除失败')

    def set_default():
        pid = selected_preset_id.get()
        if not pid:
            messagebox.showwarning('提示', '请先选择一个预设')
            return
        set_default_preset(pid, script_dir)
        messagebox.showinfo('成功', '已设为默认预设')
        refresh_presets()

    def apply_preset():
        pid = selected_preset_id.get()
        if not pid:
            messagebox.showwarning('提示', '请先选择一个预设')
            return
        preset = load_preset(pid, script_dir)
        if not preset:
            return

        folder = gui_askdirectory(title='请选择银行流水文件夹')
        if not folder:
            return

        if not messagebox.askyesno('确认应用',
                                   f'即将应用预设「{preset.get("name", "")}」\n处理文件夹: {folder}\n\n是否继续？'):
            return

        root.destroy()

        with AuditLogger('preset_pipeline', script_dir) as audit:
            audit.record_input(folder)
            audit.set_extra_info({'preset_id': pid, 'preset_name': preset.get('name', '')})
            result = apply_preset_to_pipeline(preset, folder, script_dir)
            audit.record_result(result)

            msg = format_result_message(result)
            msg += f'\n\n审计编号: {audit.audit_id}'
            msg += f'\n预设: {preset.get("name", "")} ({pid})'
            show_info('完成' if result.all_rows else '提示', msg)

    main_frame = tk.Frame(root)
    main_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    left_frame = tk.Frame(main_frame)
    left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))

    tk.Label(left_frame, text='预设列表（★ 表示默认）', font=('Arial', 12, 'bold')).pack(anchor=tk.W)

    listbox_frame = tk.Frame(left_frame)
    listbox_frame.pack(fill=tk.BOTH, expand=True, pady=5)

    scrollbar = tk.Scrollbar(listbox_frame)
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    listbox = tk.Listbox(listbox_frame, font=('Arial', 11), yscrollcommand=scrollbar.set)
    listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
    scrollbar.config(command=listbox.yview)
    listbox.bind('<<ListboxSelect>>', on_select)

    btn_frame = tk.Frame(left_frame)
    btn_frame.pack(fill=tk.X, pady=5)

    tk.Button(btn_frame, text='新增', width=8, command=add_preset, bg='#4CAF50', fg='white').pack(side=tk.LEFT, padx=2)
    tk.Button(btn_frame, text='编辑', width=8, command=edit_preset, bg='#2196F3', fg='white').pack(side=tk.LEFT, padx=2)
    tk.Button(btn_frame, text='删除', width=8, command=delete_selected, bg='#f44336', fg='white').pack(side=tk.LEFT, padx=2)
    tk.Button(btn_frame, text='设为默认', width=10, command=set_default, bg='#FF9800', fg='white').pack(side=tk.LEFT, padx=2)

    right_frame = tk.Frame(main_frame)
    right_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True, padx=(5, 0))

    tk.Label(right_frame, text='预设详情', font=('Arial', 12, 'bold')).pack(anchor=tk.W)

    detail_text = tk.Text(right_frame, font=('Arial', 11), wrap=tk.WORD, height=15)
    detail_text.pack(fill=tk.BOTH, expand=True, pady=5)
    detail_text.config(state=tk.DISABLED)

    bottom_frame = tk.Frame(root)
    bottom_frame.pack(fill=tk.X, padx=10, pady=(0, 10))

    tk.Button(bottom_frame, text='应用预设处理文件', height=2, command=apply_preset,
              bg='#FF5722', fg='white', font=('Arial', 12, 'bold')).pack(fill=tk.X)
    tk.Button(bottom_frame, text='关闭', height=2, command=root.destroy,
              bg='#9E9E9E', fg='white', font=('Arial', 11)).pack(fill=tk.X, pady=(5, 0))

    refresh_presets()
    root.mainloop()


class PresetEditorDialog:
    """预设编辑器对话框"""

    def __init__(self, parent, script_dir, preset=None, on_saved=None):
        self.script_dir = script_dir
        self.preset = preset
        self.on_saved = on_saved

        self.dialog = tk.Toplevel(parent)
        self.dialog.title('编辑预设' if preset else '新增预设')
        self.dialog.geometry('500x560')
        self.dialog.transient(parent)
        self.dialog.grab_set()

        self._build_ui()

        if preset:
            self._load_preset(preset)

    def _build_ui(self):
        pad = {'padx': 10, 'pady': 5}

        tk.Label(self.dialog, text='预设名称:', font=('Arial', 11)).pack(anchor=tk.W, **pad)
        self.name_var = tk.StringVar()
        tk.Entry(self.dialog, textvariable=self.name_var, font=('Arial', 11)).pack(fill=tk.X, **pad)

        tk.Label(self.dialog, text='描述:', font=('Arial', 11)).pack(anchor=tk.W, **pad)
        self.desc_text = tk.Text(self.dialog, height=3, font=('Arial', 11))
        self.desc_text.pack(fill=tk.X, **pad)

        tk.Label(self.dialog, text='输出目录（可选）:', font=('Arial', 11)).pack(anchor=tk.W, **pad)
        output_frame = tk.Frame(self.dialog)
        output_frame.pack(fill=tk.X, **pad)
        self.output_var = tk.StringVar()
        tk.Entry(output_frame, textvariable=self.output_var, font=('Arial', 11)).pack(side=tk.LEFT, fill=tk.X, expand=True)

        def browse_output():
            d = gui_askdirectory(title='选择输出目录', show_recent_dialog=False)
            if d:
                self.output_var.set(d)

        tk.Button(output_frame, text='浏览...', command=browse_output).pack(side=tk.LEFT, padx=5)

        date_frame = tk.Frame(self.dialog)
        date_frame.pack(fill=tk.X, **pad)

        tk.Label(date_frame, text='开始日期:', font=('Arial', 11)).pack(side=tk.LEFT)
        self.start_date_var = tk.StringVar()
        tk.Entry(date_frame, textvariable=self.start_date_var, width=15, font=('Arial', 11)).pack(side=tk.LEFT, padx=5)
        tk.Label(date_frame, text='(YYYY-MM-DD)', font=('Arial', 9), fg='#666').pack(side=tk.LEFT)

        date_frame2 = tk.Frame(self.dialog)
        date_frame2.pack(fill=tk.X, **pad)
        tk.Label(date_frame2, text='结束日期:', font=('Arial', 11)).pack(side=tk.LEFT)
        self.end_date_var = tk.StringVar()
        tk.Entry(date_frame2, textvariable=self.end_date_var, width=15, font=('Arial', 11)).pack(side=tk.LEFT, padx=5)
        tk.Label(date_frame2, text='(YYYY-MM-DD)', font=('Arial', 9), fg='#666').pack(side=tk.LEFT)

        tk.Label(self.dialog, text='保留策略:', font=('Arial', 11)).pack(anchor=tk.W, **pad)
        self.keep_var = tk.StringVar(value='keep_unprocessed')
        keep_frame = tk.Frame(self.dialog)
        keep_frame.pack(anchor=tk.W, **pad)
        for key, desc in KEEP_STRATEGIES.items():
            tk.Radiobutton(keep_frame, text=desc, variable=self.keep_var, value=key, font=('Arial', 10)).pack(anchor=tk.W)

        tk.Label(self.dialog, text='启用银行:', font=('Arial', 11)).pack(anchor=tk.W, **pad)
        self.bank_vars = {}
        bank_frame = tk.Frame(self.dialog)
        bank_frame.pack(anchor=tk.W, **pad)
        for bank in BANK_PREFIXES:
            var = tk.BooleanVar(value=True)
            self.bank_vars[bank] = var
            tk.Checkbutton(bank_frame, text=bank, variable=var, font=('Arial', 10)).pack(side=tk.LEFT, padx=5)

        tk.Label(self.dialog, text='增量合并:', font=('Arial', 11)).pack(anchor=tk.W, **pad)
        self.incremental_var = tk.BooleanVar(value=True)
        inc_frame = tk.Frame(self.dialog)
        inc_frame.pack(anchor=tk.W, **pad)
        tk.Radiobutton(inc_frame, text='启用（推荐）', variable=self.incremental_var, value=True, font=('Arial', 10)).pack(side=tk.LEFT, padx=10)
        tk.Radiobutton(inc_frame, text='禁用（全量覆盖）', variable=self.incremental_var, value=False, font=('Arial', 10)).pack(side=tk.LEFT)

        btn_frame = tk.Frame(self.dialog)
        btn_frame.pack(fill=tk.X, padx=10, pady=20)

        tk.Button(btn_frame, text='保存', width=12, command=self._save,
                  bg='#4CAF50', fg='white', font=('Arial', 11, 'bold')).pack(side=tk.RIGHT, padx=5)
        tk.Button(btn_frame, text='取消', width=12, command=self.dialog.destroy,
                  bg='#9E9E9E', fg='white', font=('Arial', 11)).pack(side=tk.RIGHT)

    def _load_preset(self, preset):
        self.name_var.set(preset.get('name', ''))
        self.desc_text.delete(1.0, tk.END)
        self.desc_text.insert(1.0, preset.get('description', ''))
        self.output_var.set(preset.get('output_dir', ''))
        self.start_date_var.set(preset.get('start_date', ''))
        self.end_date_var.set(preset.get('end_date', ''))
        self.keep_var.set(preset.get('keep_strategy', 'keep_unprocessed'))
        self.incremental_var.set(preset.get('incremental', True))

        enabled_banks = preset.get('enabled_banks', BANK_PREFIXES)
        for bank in BANK_PREFIXES:
            self.bank_vars[bank].set(bank in enabled_banks)

    def _save(self):
        name = self.name_var.get().strip()
        if not name:
            messagebox.showerror('错误', '请输入预设名称', parent=self.dialog)
            return

        start_date = self.start_date_var.get().strip()
        end_date = self.end_date_var.get().strip()

        def validate_date(d):
            if not d:
                return True
            try:
                datetime.strptime(d, '%Y-%m-%d')
                return True
            except ValueError:
                return False

        if start_date and not validate_date(start_date):
            messagebox.showerror('错误', '开始日期格式错误，请使用 YYYY-MM-DD 格式', parent=self.dialog)
            return

        if end_date and not validate_date(end_date):
            messagebox.showerror('错误', '结束日期格式错误，请使用 YYYY-MM-DD 格式', parent=self.dialog)
            return

        enabled_banks = [bank for bank, var in self.bank_vars.items() if var.get()]
        if not enabled_banks:
            messagebox.showerror('错误', '请至少选择一个银行', parent=self.dialog)
            return

        preset_data = {
            'name': name,
            'description': self.desc_text.get(1.0, tk.END).strip(),
            'output_dir': self.output_var.get().strip(),
            'start_date': start_date,
            'end_date': end_date,
            'keep_strategy': self.keep_var.get(),
            'enabled_banks': enabled_banks,
            'incremental': self.incremental_var.get(),
        }

        if self.preset:
            preset_data['preset_id'] = self.preset['preset_id']

        preset_id = save_preset(preset_data, self.script_dir)

        messagebox.showinfo('成功', f'预设已保存\nID: {preset_id}', parent=self.dialog)
        self.dialog.destroy()

        if self.on_saved:
            self.on_saved()


def _list_presets_cli(presets):
    if not presets:
        print('\n暂无预设配置')
        return

    print('\n' + '-' * 80)
    print(f'{"ID":<22}{"名称":<20}{"描述":<25}{"更新时间":<20}')
    print('-' * 80)
    for p in presets:
        name = (p.get('name', '')[:18] + '..') if len(p.get('name', '')) > 20 else p.get('name', '')
        desc = (p.get('description', '')[:23] + '..') if len(p.get('description', '')) > 25 else p.get('description', '')
        print(f'{p["preset_id"]:<22}{name:<20}{desc:<25}{p.get("updated_at", ""):<20}')
    print('-' * 80)


def _save_preset_cli(script_dir):
    print('\n--- 保存新预设 ---')

    name = input('预设名称: ').strip()
    if not name:
        print('预设名称不能为空')
        return

    description = input('预设描述（可选）: ').strip()
    output_dir = input('输出目录（可选，留空使用默认）: ').strip()
    start_date = input('开始日期（YYYY-MM-DD，可选）: ').strip()
    end_date = input('结束日期（YYYY-MM-DD，可选）: ').strip()

    print('\n保留策略选项：')
    for key, desc in KEEP_STRATEGIES.items():
        print(f'  {key}: {desc}')
    keep_strategy = input(f'保留策略（默认: keep_unprocessed）: ').strip() or 'keep_unprocessed'
    if keep_strategy not in KEEP_STRATEGIES:
        print(f'无效的保留策略，使用默认: keep_unprocessed')
        keep_strategy = 'keep_unprocessed'

    print('\n可用银行列表：')
    for i, bank in enumerate(BANK_PREFIXES, 1):
        print(f'  {i}) {bank}')
    bank_input = input('选择启用的银行（输入编号，逗号分隔，回车全选）: ').strip()
    if bank_input:
        try:
            indices = [int(x.strip()) - 1 for x in bank_input.split(',')]
            enabled_banks = [BANK_PREFIXES[i] for i in indices if 0 <= i < len(BANK_PREFIXES)]
        except (ValueError, IndexError):
            print('输入无效，将启用所有银行')
            enabled_banks = list(BANK_PREFIXES)
    else:
        enabled_banks = list(BANK_PREFIXES)

    incremental_input = input('是否启用增量合并? (y/N): ').strip().lower()
    incremental = incremental_input == 'y'

    preset_data = {
        'name': name,
        'description': description,
        'output_dir': output_dir,
        'start_date': start_date,
        'end_date': end_date,
        'keep_strategy': keep_strategy,
        'enabled_banks': enabled_banks,
        'incremental': incremental,
    }

    preset_id = save_preset(preset_data, script_dir)
    print(f'\n✅ 预设已保存，ID: {preset_id}')


def _apply_preset_cli(script_dir):
    presets = list_presets(script_dir)
    if not presets:
        print('\n暂无预设可用')
        return

    _list_presets_cli(presets)
    preset_id = input('\n请输入要应用的预设ID: ').strip()

    preset = load_preset(preset_id, script_dir)
    if not preset:
        print(f'❌ 未找到预设: {preset_id}')
        return

    print(f'\n即将应用预设: {preset.get("name", "")}')
    _print_preset_detail(preset)

    folder = ask_directory('请选择银行流水文件夹')
    if not folder:
        print('未选择文件夹，取消应用')
        return

    confirm = input(f'\n确认应用预设处理文件夹「{folder}」? (Y/n): ').strip().lower()
    if confirm and confirm != 'y':
        print('已取消')
        return

    with AuditLogger('preset_pipeline', script_dir) as audit:
        audit.record_input(folder)
        audit.set_extra_info({'preset_id': preset_id, 'preset_name': preset.get('name', '')})

        result = apply_preset_to_pipeline(preset, folder, script_dir)
        audit.record_result(result)

        msg = format_result_message(result)
        msg += f'\n\n审计编号: {audit.audit_id}'
        msg += f'\n预设: {preset.get("name", "")} ({preset_id})'
        show_info('完成' if result.all_rows else '提示', msg)


def _delete_preset_cli(script_dir):
    presets = list_presets(script_dir)
    if not presets:
        print('\n暂无预设可删除')
        return

    _list_presets_cli(presets)
    preset_id = input('\n请输入要删除的预设ID: ').strip()

    preset = load_preset(preset_id, script_dir)
    if not preset:
        print(f'❌ 未找到预设: {preset_id}')
        return

    confirm = input(f'确认删除预设「{preset.get("name", "")}」? (y/N): ').strip().lower()
    if confirm != 'y':
        print('已取消')
        return

    if delete_preset(preset_id, script_dir):
        print('✅ 预设已删除')
    else:
        print('❌ 删除失败')


def _set_default_preset_cli(script_dir):
    presets = list_presets(script_dir)
    if not presets:
        print('\n暂无预设')
        return

    _list_presets_cli(presets)
    preset_id = input('\n请输入要设为默认的预设ID: ').strip()

    preset = load_preset(preset_id, script_dir)
    if not preset:
        print(f'❌ 未找到预设: {preset_id}')
        return

    set_default_preset(preset_id, script_dir)
    print(f'✅ 已将「{preset.get("name", "")}」设为默认预设')


def _show_preset_detail_cli(script_dir):
    presets = list_presets(script_dir)
    if not presets:
        print('\n暂无预设')
        return

    _list_presets_cli(presets)
    preset_id = input('\n请输入要查看的预设ID: ').strip()

    preset = load_preset(preset_id, script_dir)
    if not preset:
        print(f'❌ 未找到预设: {preset_id}')
        return

    _print_preset_detail(preset)


def _print_preset_detail(preset):
    print('\n' + '=' * 50)
    print(f'预设名称: {preset.get("name", "")}')
    print(f'预设ID: {preset.get("preset_id", "")}')
    print(f'描述: {preset.get("description", "无")}')
    print('-' * 50)
    print(f'输出目录: {preset.get("output_dir", "默认")}')
    print(f'开始日期: {preset.get("start_date", "不限制")}')
    print(f'结束日期: {preset.get("end_date", "不限制")}')
    print(f'保留策略: {KEEP_STRATEGIES.get(preset.get("keep_strategy"), "未知")}')
    print(f'启用银行: {", ".join(preset.get("enabled_banks", []))}')
    print(f'增量合并: {"是" if preset.get("incremental", True) else "否"}')
    print('-' * 50)
    print(f'创建时间: {preset.get("created_at", "")}')
    print(f'更新时间: {preset.get("updated_at", "")}')
    print('=' * 50)


# ──────────────────────────────────────────────
# 会计期间自动归属模块
# ──────────────────────────────────────────────

import calendar as _calendar


@dataclass
class AccountingPeriodConfig:
    period_type: str = 'monthly'
    cutoff_day: int = 25
    fiscal_year_start_month: int = 1
    period_name_format: str = 'YYYY-MM'


class AccountingPeriodManager:
    _instance = None
    _config_path = None
    _config: AccountingPeriodConfig = None
    _last_modified: float = 0.0

    def __new__(cls, config_path=None):
        if cls._instance is None:
            cls._instance = super(AccountingPeriodManager, cls).__new__(cls)
            cls._instance._initialized = False
        return cls._instance

    def __init__(self, config_path=None):
        if self._initialized:
            return
        self._initialized = True
        if config_path is None:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            config_path = os.path.join(script_dir, BANK_RULES_CONFIG_FILE)
        self._config_path = config_path
        self._config = AccountingPeriodConfig()
        self._last_modified = 0.0
        self.load_config()

    def load_config(self):
        logger = get_logger()
        if not os.path.exists(self._config_path):
            logger.warning('会计期间配置文件不存在: %s，使用默认配置', self._config_path)
            return False
        try:
            current_mtime = os.path.getmtime(self._config_path)
            if current_mtime == self._last_modified and self._config:
                return True
            with open(self._config_path, 'r', encoding='utf-8') as f:
                config_data = yaml.safe_load(f)
            period_cfg = config_data.get('accounting_period', {})
            if not period_cfg:
                logger.info('未配置 accounting_period 节，使用默认配置')
                return True
            cutoff = int(period_cfg.get('cutoff_day', 25))
            if cutoff < 1:
                cutoff = 1
            if cutoff > 31:
                cutoff = 31
            fysm = int(period_cfg.get('fiscal_year_start_month', 1))
            if fysm < 1:
                fysm = 1
            if fysm > 12:
                fysm = 12
            self._config = AccountingPeriodConfig(
                period_type=period_cfg.get('period_type', 'monthly'),
                cutoff_day=cutoff,
                fiscal_year_start_month=fysm,
                period_name_format=period_cfg.get('period_name_format', 'YYYY-MM'),
            )
            self._last_modified = current_mtime
            logger.info(
                '会计期间配置已加载: period_type=%s, cutoff_day=%d, fiscal_year_start_month=%d, format=%s',
                self._config.period_type, self._config.cutoff_day,
                self._config.fiscal_year_start_month, self._config.period_name_format,
            )
            return True
        except Exception as e:
            logger.error('加载会计期间配置失败: %s', e, exc_info=True)
            return False

    def get_config(self) -> AccountingPeriodConfig:
        self.load_config()
        return self._config


def get_accounting_period_manager():
    global _ap_manager_singleton
    if _ap_manager_singleton is None:
        _ap_manager_singleton = AccountingPeriodManager()
    return _ap_manager_singleton


_ap_manager_singleton = None


def _determine_monthly_period(year, month, day, cutoff_day):
    if cutoff_day >= 28:
        last_day = _calendar.monthrange(year, month)[1]
        effective_cutoff = min(cutoff_day, last_day)
    else:
        effective_cutoff = cutoff_day
    if day > effective_cutoff:
        if month == 12:
            return year + 1, 1
        return year, month + 1
    return year, month


def _determine_quarterly_period(year, month, day, cutoff_day, fiscal_year_start_month):
    offsets = [(fiscal_year_start_month - 1 + i) % 12 for i in range(12)]
    q_start_indices = [0, 3, 6, 9]
    quarter_of_month = None
    for qi in range(4):
        q_months = [offsets[q_start_indices[qi] + j] + 1 for j in range(3)]
        if month in q_months:
            quarter_of_month = qi
            break
    if quarter_of_month is None:
        return year, quarter_of_month
    q_months = [offsets[q_start_indices[quarter_of_month] + j] + 1 for j in range(3)]
    is_last_month_of_quarter = (month == q_months[2])
    if is_last_month_of_quarter:
        if day > cutoff_day:
            next_qi = quarter_of_month + 1
            if next_qi >= 4:
                next_q_first_month = offsets[0] + 1
                if next_q_first_month <= month:
                    return year + 1, 0
                return year, 0
            next_q_first_month = offsets[q_start_indices[next_qi]] + 1
            if next_q_first_month < month:
                return year + 1, next_qi
            return year, next_qi
        return year, quarter_of_month
    return year, quarter_of_month


def _quarter_label(year, quarter_index, fiscal_year_start_month, period_name_format):
    q_num = quarter_index + 1
    if period_name_format == 'YYYY-QN':
        return f'{year}-Q{q_num}'
    q_months_offsets = [(fiscal_year_start_month - 1 + i) % 12 for i in range(12)]
    q_start_indices = [0, 3, 6, 9]
    q_months = [q_months_offsets[q_start_indices[quarter_index] + j] + 1 for j in range(3)]
    return f'{year}-Q{q_num}({q_months[0]:02d}-{q_months[2]:02d})'


def assign_accounting_period(records, config=None):
    """
    根据可配置的账期截止日规则，将每笔交易自动归入对应会计月份或季度。

    Args:
        records: 交易记录列表，每条包含 '交易日期' 等字段
        config: AccountingPeriodConfig 实例，默认从配置文件加载

    Returns:
        tuple: (enriched_records, period_summary)
            - enriched_records: 每条记录新增 '会计期间' 字段
            - period_summary: Dict[str, List[Dict]] 按期间分组的记录
    """
    logger = get_logger()

    if config is None:
        manager = get_accounting_period_manager()
        config = manager.get_config()

    if not records:
        logger.warning('无交易记录，会计期间归属跳过')
        return [], {}

    period_summary: Dict[str, List[Dict]] = {}
    enriched = []

    for rec in records:
        trade_date = rec.get('交易日期')
        dt = _normalize_date(trade_date)
        period_label = '未知期间'

        if dt is not None:
            year = dt.year
            month = dt.month
            day = dt.day

            if config.period_type == 'quarterly':
                adj_year, q_index = _determine_quarterly_period(
                    year, month, day, config.cutoff_day, config.fiscal_year_start_month)
                period_label = _quarter_label(
                    adj_year, q_index, config.fiscal_year_start_month, config.period_name_format)
            else:
                adj_year, adj_month = _determine_monthly_period(
                    year, month, day, config.cutoff_day)
                if config.period_name_format == 'YYYY-QN':
                    q_index = (adj_month - 1) // 3
                    period_label = f'{adj_year}-Q{q_index + 1}'
                else:
                    period_label = f'{adj_year}-{adj_month:02d}'

        new_rec = dict(rec)
        new_rec['会计期间'] = period_label
        enriched.append(new_rec)

        if period_label not in period_summary:
            period_summary[period_label] = []
        period_summary[period_label].append(new_rec)

    logger.info(
        '会计期间归属完成: %d 条记录, %d 个期间',
        len(enriched), len(period_summary),
    )

    return enriched, period_summary


def _compute_period_aggregates(period_summary):
    result = {}
    for period, recs in period_summary.items():
        total_income = 0.0
        total_expense = 0.0
        income_count = 0
        expense_count = 0
        for rec in recs:
            receipt = to_float(rec.get('收款'))
            payment = to_float(rec.get('付款'))
            if receipt is not None and receipt > 0:
                total_income += receipt
                income_count += 1
            if payment is not None and payment < 0:
                total_expense += abs(payment)
                expense_count += 1
        net = round(total_income - total_expense, 2)
        result[period] = {
            '会计期间': period,
            '收入总额': round(total_income, 2),
            '支出总额': round(total_expense, 2),
            '净额': net,
            '交易笔数': len(recs),
            '收入笔数': income_count,
            '支出笔数': expense_count,
        }
    return result


def export_accounting_period_summary(enriched_records, period_summary, output_path,
                                      source_info=None):
    """
    导出按期间分 Sheet 的总表 Excel，便于结账。

    每个 Sheet 包含该期间的交易明细 + 期间汇总行。
    额外增加一个"期间汇总"Sheet，汇总所有期间的收支情况。

    Args:
        enriched_records: 含 '会计期间' 字段的记录列表
        period_summary: 按期间分组的记录字典
        output_path: 输出文件路径
        source_info: 数据源信息字典

    Returns:
        str: 输出文件路径
    """
    logger = get_logger()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    detail_columns = ['会计期间', '唯一id', '银行', '银行账号', '主体', '交易日期',
                       '付款', '收款', '摘要', '对方户名', '余额', '交易流水号']
    extra_keys = set()
    for rec in enriched_records:
        for k in rec:
            if k not in detail_columns:
                extra_keys.add(k)
    detail_columns.extend(sorted(extra_keys))

    sorted_periods = sorted(period_summary.keys())

    for period in sorted_periods:
        recs = period_summary[period]
        safe_title = period.replace('/', '-').replace('\\', '-')[:31]
        ws = wb.create_sheet(title=safe_title)

        for col_idx, col_name in enumerate(detail_columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)

        for row_idx, rec in enumerate(recs, 2):
            for col_idx, col_name in enumerate(detail_columns, 1):
                val = rec.get(col_name)
                ws.cell(row=row_idx, column=col_idx, value=val)

        summary_row = len(recs) + 3
        ws.cell(row=summary_row, column=1, value='期间汇总')
        ws.cell(row=summary_row, column=2, value=f'交易笔数: {len(recs)}')

        income_sum = sum(
            to_float(r.get('收款')) or 0
            for r in recs if to_float(r.get('收款')) is not None and to_float(r.get('收款')) > 0
        )
        expense_sum = sum(
            abs(to_float(r.get('付款')) or 0)
            for r in recs if to_float(r.get('付款')) is not None and to_float(r.get('付款')) < 0
        )
        ws.cell(row=summary_row, column=3, value=f'收入合计: {round(income_sum, 2)}')
        ws.cell(row=summary_row, column=4, value=f'支出合计: {round(expense_sum, 2)}')
        ws.cell(row=summary_row, column=5, value=f'净额: {round(income_sum - expense_sum, 2)}')

    summary_columns = ['会计期间', '收入总额', '支出总额', '净额', '交易笔数', '收入笔数', '支出笔数']
    ws_summary = wb.create_sheet(title='期间汇总', index=0)

    if source_info:
        ws_summary.cell(row=1, column=1, value='数据源信息')
        for i, (k, v) in enumerate(source_info.items(), 2):
            ws_summary.cell(row=i, column=1, value=k)
            ws_summary.cell(row=i, column=2, value=str(v))
        header_row = len(source_info) + 3
    else:
        header_row = 1

    for col_idx, col_name in enumerate(summary_columns, 1):
        ws_summary.cell(row=header_row, column=col_idx, value=col_name)

    aggregates = _compute_period_aggregates(period_summary)
    for row_offset, period in enumerate(sorted_periods):
        agg = aggregates[period]
        for col_idx, col_name in enumerate(summary_columns, 1):
            ws_summary.cell(row=header_row + 1 + row_offset, column=col_idx, value=agg.get(col_name, ''))

    wb.save(output_path)
    wb.close()
    logger.info('会计期间总表已导出: %s（%d 个期间）', output_path, len(sorted_periods))
    return output_path


def generate_accounting_period_report(records, output_dir=None, source_info=None, config=None):
    """
    从交易记录列表直接生成会计期间归属总表。

    Args:
        records: 交易记录列表
        output_dir: 输出目录，默认为当前脚本目录
        source_info: 数据源信息
        config: AccountingPeriodConfig 实例，默认从配置文件加载

    Returns:
        str: 生成的文件路径，如无数据则返回 None
    """
    logger = get_logger()

    if not records:
        logger.warning('无交易记录，跳过会计期间报告生成')
        return None

    if output_dir is None:
        output_dir = get_script_dir()

    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'会计期间总表_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, filename)

    enriched, period_summary = assign_accounting_period(records, config)

    if not enriched:
        logger.warning('会计期间归属后无有效记录，跳过导出')
        return None

    return export_accounting_period_summary(enriched, period_summary, output_path, source_info)


# ──────────────────────────────────────────────
# 主体维度汇总分析模块
# ──────────────────────────────────────────────

SUBJECT_SUMMARY_FILENAME = '主体维度汇总分析.xlsx'


@dataclass
class SubjectDimensionSummary:
    """主体维度汇总数据"""
    subject: str = ''
    bank: str = ''
    year_month: str = ''
    total_income: float = 0.0
    total_expense: float = 0.0
    net_amount: float = 0.0
    transaction_count: int = 0
    income_count: int = 0
    expense_count: int = 0


@dataclass
class SubjectSummaryResult:
    """汇总分析结果容器"""
    by_subject: List[Dict[str, Any]] = field(default_factory=list)
    by_subject_bank: List[Dict[str, Any]] = field(default_factory=list)
    by_subject_month: List[Dict[str, Any]] = field(default_factory=list)
    by_subject_bank_month: List[Dict[str, Any]] = field(default_factory=list)
    by_bank: List[Dict[str, Any]] = field(default_factory=list)
    by_month: List[Dict[str, Any]] = field(default_factory=list)
    overall_summary: Dict[str, Any] = field(default_factory=dict)
    unmatched_accounts: List[Dict[str, Any]] = field(default_factory=list)


def _extract_year_month(trade_date) -> str:
    """从交易日期提取年月，格式 YYYY-MM"""
    if trade_date is None:
        return '未知'

    s = str(trade_date).strip()
    if not s:
        return '未知'

    dt = _normalize_date(trade_date)
    if dt is None:
        if len(s) >= 7:
            if s[4] in '-/.' and len(s) >= 7:
                return s[:7]
            elif s.isdigit() and len(s) >= 6:
                return f'{s[:4]}-{s[4:6]}'
        return '未知'

    try:
        return dt.strftime('%Y-%m')
    except Exception:
        if len(s) >= 7 and s[4] in '-/.' and len(s) >= 7:
            return s[:7]
        return '未知'


def summarize_transactions(
    records: List[Dict[str, Any]],
    exclude_internal_transfers: bool = True,
) -> SubjectSummaryResult:
    """
    对交易记录进行多维度汇总分析。

    维度组合：
    1. 按主体
    2. 按主体 + 银行
    3. 按主体 + 月份
    4. 按主体 + 银行 + 月份
    5. 按银行
    6. 按月份

    每个维度统计：收入总额、支出总额、净额、交易笔数、收入笔数、支出笔数

    Args:
        records: 交易记录列表，每条包含 '主体', '银行', '交易日期', '付款', '收款' 等字段
        exclude_internal_transfers: 是否排除内部划转记录（默认 True，避免主体汇总重复计算）

    Returns:
        SubjectSummaryResult: 多维度汇总结果
    """
    logger = get_logger()

    empty_overall = {
        'total_income': 0.0, 'total_expense': 0.0, 'net_amount': 0.0,
        'transaction_count': 0, 'income_count': 0, 'expense_count': 0,
        'subject_count': 0, 'bank_count': 0, 'month_count': 0,
    }

    if not records:
        logger.warning('无交易记录可汇总')
        return SubjectSummaryResult(overall_summary=empty_overall)

    if exclude_internal_transfers:
        original_count = len(records)
        records = filter_internal_transfers_for_summary(records, exclude=True)
        filtered_count = original_count - len(records)
        if filtered_count > 0:
            logger.info(
                '主体汇总已排除 %d 条内部划转记录（划出+划入各 %d 条）',
                filtered_count, filtered_count // 2 if filtered_count % 2 == 0 else filtered_count,
            )

    if not records:
        logger.warning('排除内部划转后无交易记录可汇总')
        return SubjectSummaryResult(overall_summary=empty_overall)

    agg_3d: Dict[Tuple[str, str, str], SubjectDimensionSummary] = {}
    agg_subject: Dict[str, SubjectDimensionSummary] = {}
    agg_subject_bank: Dict[Tuple[str, str], SubjectDimensionSummary] = {}
    agg_subject_month: Dict[Tuple[str, str], SubjectDimensionSummary] = {}
    agg_bank: Dict[str, SubjectDimensionSummary] = {}
    agg_month: Dict[str, SubjectDimensionSummary] = {}

    total_income = 0.0
    total_expense = 0.0
    total_count = 0
    total_income_count = 0
    total_expense_count = 0

    unmatched_accounts_agg: Dict[str, Dict[str, Any]] = {}

    def _update(entry, income, expense, is_income, is_expense):
        entry.total_income += income
        entry.total_expense += expense
        entry.transaction_count += 1
        if is_income:
            entry.income_count += 1
        if is_expense:
            entry.expense_count += 1

    def _normalize_date_for_compare(date_val):
        if date_val is None:
            return None
        dt = _normalize_date(date_val)
        if dt is not None:
            return dt.strftime('%Y-%m-%d')
        s = str(date_val).strip()
        if len(s) >= 10 and s[4] in '-/.' and s[7] in '-/.':
            return s[:10].replace('/', '-').replace('.', '-')
        return None

    for rec in records:
        original_subject = str(rec.get('主体') or '').strip()
        subject = original_subject or '未指定主体'
        bank = str(rec.get('银行') or '').strip() or '未知银行'
        year_month = _extract_year_month(rec.get('交易日期'))
        bank_account = str(rec.get('银行账号') or '').strip()
        trade_date_str = _normalize_date_for_compare(rec.get('交易日期'))

        payment = to_float(rec.get('付款'))
        receipt = to_float(rec.get('收款'))

        income = 0.0
        expense = 0.0
        is_income = False
        is_expense = False

        if receipt is not None and receipt > 0:
            income = receipt
            is_income = True
        if payment is not None and payment < 0:
            expense = abs(payment)
            is_expense = True

        if not is_income and not is_expense:
            continue

        total_income += income
        total_expense += expense
        total_count += 1
        if is_income:
            total_income_count += 1
        if is_expense:
            total_expense_count += 1

        if not original_subject and bank_account:
            acc_key = bank_account
            if acc_key not in unmatched_accounts_agg:
                unmatched_accounts_agg[acc_key] = {
                    'account': bank_account,
                    'banks': set(),
                    'first_date': trade_date_str,
                    'last_date': trade_date_str,
                    'total_income': 0.0,
                    'total_expense': 0.0,
                    'transaction_count': 0,
                }
            ua = unmatched_accounts_agg[acc_key]
            ua['banks'].add(bank)
            if trade_date_str:
                if ua['first_date'] is None or trade_date_str < ua['first_date']:
                    ua['first_date'] = trade_date_str
                if ua['last_date'] is None or trade_date_str > ua['last_date']:
                    ua['last_date'] = trade_date_str
            ua['total_income'] += income
            ua['total_expense'] += expense
            ua['transaction_count'] += 1

        if subject not in agg_subject:
            agg_subject[subject] = SubjectDimensionSummary(subject=subject)
        _update(agg_subject[subject], income, expense, is_income, is_expense)

        if bank not in agg_bank:
            agg_bank[bank] = SubjectDimensionSummary(bank=bank)
        _update(agg_bank[bank], income, expense, is_income, is_expense)

        if year_month not in agg_month:
            agg_month[year_month] = SubjectDimensionSummary(year_month=year_month)
        _update(agg_month[year_month], income, expense, is_income, is_expense)

        key_sb = (subject, bank)
        if key_sb not in agg_subject_bank:
            agg_subject_bank[key_sb] = SubjectDimensionSummary(subject=subject, bank=bank)
        _update(agg_subject_bank[key_sb], income, expense, is_income, is_expense)

        key_sm = (subject, year_month)
        if key_sm not in agg_subject_month:
            agg_subject_month[key_sm] = SubjectDimensionSummary(subject=subject, year_month=year_month)
        _update(agg_subject_month[key_sm], income, expense, is_income, is_expense)

        key_3d = (subject, bank, year_month)
        if key_3d not in agg_3d:
            agg_3d[key_3d] = SubjectDimensionSummary(subject=subject, bank=bank, year_month=year_month)
        _update(agg_3d[key_3d], income, expense, is_income, is_expense)

    def finalize(entries):
        for e in entries:
            e.net_amount = round(e.total_income - e.total_expense, 2)
            e.total_income = round(e.total_income, 2)
            e.total_expense = round(e.total_expense, 2)

    finalize(agg_3d.values())
    finalize(agg_subject.values())
    finalize(agg_subject_bank.values())
    finalize(agg_subject_month.values())
    finalize(agg_bank.values())
    finalize(agg_month.values())

    def to_dict_list(entries, fields):
        result = []
        for e in entries:
            d = {}
            for f in fields:
                d[f] = getattr(e, f)
            result.append(d)
        return result

    fields_all = ['subject', 'bank', 'year_month', 'total_income', 'total_expense',
                  'net_amount', 'transaction_count', 'income_count', 'expense_count']
    fields_s = ['subject', 'total_income', 'total_expense', 'net_amount',
                'transaction_count', 'income_count', 'expense_count']
    fields_sb = ['subject', 'bank', 'total_income', 'total_expense', 'net_amount',
                 'transaction_count', 'income_count', 'expense_count']
    fields_sm = ['subject', 'year_month', 'total_income', 'total_expense', 'net_amount',
                 'transaction_count', 'income_count', 'expense_count']
    fields_b = ['bank', 'total_income', 'total_expense', 'net_amount',
                'transaction_count', 'income_count', 'expense_count']
    fields_m = ['year_month', 'total_income', 'total_expense', 'net_amount',
                'transaction_count', 'income_count', 'expense_count']

    unmatched_accounts_list = []
    for ua in unmatched_accounts_agg.values():
        unmatched_accounts_list.append({
            'account': ua['account'],
            'banks': '、'.join(sorted(ua['banks'])),
            'first_date': ua['first_date'] or '',
            'last_date': ua['last_date'] or '',
            'total_income': round(ua['total_income'], 2),
            'total_expense': round(ua['total_expense'], 2),
            'net_amount': round(ua['total_income'] - ua['total_expense'], 2),
            'transaction_count': ua['transaction_count'],
        })
    unmatched_accounts_list.sort(key=lambda x: x['transaction_count'], reverse=True)

    result = SubjectSummaryResult(
        by_subject=sorted(
            to_dict_list(agg_subject.values(), fields_s),
            key=lambda x: x['net_amount'], reverse=True
        ),
        by_subject_bank=sorted(
            to_dict_list(agg_subject_bank.values(), fields_sb),
            key=lambda x: (x['subject'], x['net_amount']), reverse=True
        ),
        by_subject_month=sorted(
            to_dict_list(agg_subject_month.values(), fields_sm),
            key=lambda x: (x['subject'], x['year_month'])
        ),
        by_subject_bank_month=sorted(
            to_dict_list(agg_3d.values(), fields_all),
            key=lambda x: (x['subject'], x['bank'], x['year_month'])
        ),
        by_bank=sorted(
            to_dict_list(agg_bank.values(), fields_b),
            key=lambda x: x['net_amount'], reverse=True
        ),
        by_month=sorted(
            to_dict_list(agg_month.values(), fields_m),
            key=lambda x: x['year_month']
        ),
        overall_summary={
            'total_income': round(total_income, 2),
            'total_expense': round(total_expense, 2),
            'net_amount': round(total_income - total_expense, 2),
            'transaction_count': total_count,
            'income_count': total_income_count,
            'expense_count': total_expense_count,
            'subject_count': len(agg_subject),
            'bank_count': len(agg_bank),
            'month_count': len(agg_month),
            'unmatched_account_count': len(unmatched_accounts_list),
        },
        unmatched_accounts=unmatched_accounts_list,
    )

    logger.info(
        '主体维度汇总完成: %d 条记录, %d 个主体, %d 家银行, %d 个月份, %d 个未匹配账号',
        total_count, len(agg_subject), len(agg_bank), len(agg_month),
        len(unmatched_accounts_list)
    )

    return result


CN_COLUMNS_S = {
    'subject': '主体',
    'total_income': '收入总额(元)',
    'total_expense': '支出总额(元)',
    'net_amount': '净额(元)',
    'transaction_count': '交易笔数',
    'income_count': '收入笔数',
    'expense_count': '支出笔数',
}

CN_COLUMNS_SB = {
    **CN_COLUMNS_S,
    'bank': '开户银行',
}

CN_COLUMNS_SM = {
    **CN_COLUMNS_S,
    'year_month': '月份',
}

CN_COLUMNS_ALL = {
    **CN_COLUMNS_SB,
    'year_month': '月份',
}

CN_COLUMNS_B = {
    'bank': '开户银行',
    'total_income': '收入总额(元)',
    'total_expense': '支出总额(元)',
    'net_amount': '净额(元)',
    'transaction_count': '交易笔数',
    'income_count': '收入笔数',
    'expense_count': '支出笔数',
}

CN_COLUMNS_M = {
    'year_month': '月份',
    'total_income': '收入总额(元)',
    'total_expense': '支出总额(元)',
    'net_amount': '净额(元)',
    'transaction_count': '交易笔数',
    'income_count': '收入笔数',
    'expense_count': '支出笔数',
}


def _rename_columns(df, col_map):
    """重命名 DataFrame 列名为中文，并调整列顺序"""
    existing = [c for c in col_map.keys() if c in df.columns]
    df = df[existing]
    df = df.rename(columns=col_map)
    return df


def _apply_number_format(ws, amount_cols, count_cols):
    """对工作表应用数字格式"""
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            col_letter = cell.column_letter
            if col_letter in amount_cols:
                cell.number_format = '#,##0.00'
            elif col_letter in count_cols:
                cell.number_format = '#,##0'


def export_subject_summary(summary_result: SubjectSummaryResult,
                           output_path: str,
                           source_info: Optional[Dict[str, Any]] = None) -> str:
    """
    将主体维度汇总分析结果导出为多 Sheet Excel 文件。

    输出的 Sheet 包括：
    1. 汇总总览 - 整体统计信息
    2. 按主体汇总 - 各主体收支净额统计
    3. 按主体+银行汇总 - 各主体在各银行的收支统计
    4. 按主体+月份汇总 - 各主体月度收支趋势
    5. 按主体+银行+月份汇总 - 最细粒度多维分析
    6. 按银行汇总 - 各银行总体收支
    7. 按月份汇总 - 全量月度收支趋势

    Args:
        summary_result: summarize_transactions 返回的汇总结果
        output_path: 输出 Excel 文件路径
        source_info: 可选，数据源信息（如总表路径、记录数等）

    Returns:
        str: 输出文件路径
    """
    logger = get_logger()

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            overview_data = []
            overall = summary_result.overall_summary

            overview_items = [
                ('统计项', '数值'),
                ('交易总笔数', overall.get('transaction_count', 0)),
                ('收入笔数', overall.get('income_count', 0)),
                ('支出笔数', overall.get('expense_count', 0)),
                ('收入总额(元)', overall.get('total_income', 0)),
                ('支出总额(元)', overall.get('total_expense', 0)),
                ('净额(元)', overall.get('net_amount', 0)),
                ('涉及主体数', overall.get('subject_count', 0)),
                ('涉及银行数', overall.get('bank_count', 0)),
                ('覆盖月份数', overall.get('month_count', 0)),
                ('未匹配账号数', overall.get('unmatched_account_count', 0)),
            ]
            if source_info:
                for k, v in source_info.items():
                    overview_items.append((k, v))

            overview_df = pd.DataFrame(overview_items[1:], columns=overview_items[0])
            overview_df.to_excel(writer, sheet_name='汇总总览', index=False)

            sheet_configs = [
                ('按主体汇总', summary_result.by_subject, CN_COLUMNS_S),
                ('按主体+银行', summary_result.by_subject_bank, CN_COLUMNS_SB),
                ('按主体+月份', summary_result.by_subject_month, CN_COLUMNS_SM),
                ('主体+银行+月份', summary_result.by_subject_bank_month, CN_COLUMNS_ALL),
                ('按银行汇总', summary_result.by_bank, CN_COLUMNS_B),
                ('按月份汇总', summary_result.by_month, CN_COLUMNS_M),
            ]

            for sheet_name, data, col_map in sheet_configs:
                if not data:
                    continue
                df = pd.DataFrame(data)
                df = _rename_columns(df, col_map)
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                ws = writer.sheets[sheet_name]
                amount_cols = set()
                count_cols = set()
                for idx, col_name in enumerate(df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    if '元' in str(col_name):
                        amount_cols.add(col_letter)
                    elif '笔数' in str(col_name):
                        count_cols.add(col_letter)
                    max_len = max(
                        len(str(col_name)),
                        max((len(str(v)) for v in df.iloc[:, idx - 1].astype(str)), default=0)
                    )
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

                _apply_number_format(ws, amount_cols, count_cols)

            unmatched = summary_result.unmatched_accounts
            if unmatched:
                ua_col_map = {
                    'account': '银行账号',
                    'banks': '涉及银行',
                    'first_date': '首次交易日期',
                    'last_date': '最后交易日期',
                    'total_income': '收入总额(元)',
                    'total_expense': '支出总额(元)',
                    'net_amount': '净额(元)',
                    'transaction_count': '交易笔数',
                }
                ua_df = pd.DataFrame(unmatched)
                ua_df = _rename_columns(ua_df, ua_col_map)
                ua_df.to_excel(writer, sheet_name='未匹配账号汇总', index=False)

                ws_ua = writer.sheets['未匹配账号汇总']
                ua_amount_cols = set()
                ua_count_cols = set()
                for idx, col_name in enumerate(ua_df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    if '元' in str(col_name):
                        ua_amount_cols.add(col_letter)
                    elif '笔数' in str(col_name):
                        ua_count_cols.add(col_letter)
                    max_len = max(
                        len(str(col_name)),
                        max((len(str(v)) for v in ua_df.iloc[:, idx - 1].astype(str)), default=0)
                    )
                    ws_ua.column_dimensions[col_letter].width = min(max_len + 4, 40)

                _apply_number_format(ws_ua, ua_amount_cols, ua_count_cols)

            ws_overview = writer.sheets['汇总总览']
            for col_idx in range(1, 3):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws_overview.column_dimensions[col_letter].width = 25

            for row in ws_overview.iter_rows(min_row=2):
                for cell in row:
                    if cell.column == 2:
                        val = cell.value
                        if isinstance(val, (int, float)):
                            if isinstance(val, float):
                                cell.number_format = '#,##0.00'
                            else:
                                cell.number_format = '#,##0'

        logger.info('主体维度汇总分析已导出: %s', output_path)
        return output_path

    except Exception as e:
        logger.error('导出主体维度汇总分析失败: %s', e, exc_info=True)
        raise


def generate_subject_summary_from_records(records: List[Dict[str, Any]],
                                          output_dir: Optional[str] = None,
                                          source_info: Optional[Dict[str, Any]] = None,
                                          exclude_internal_transfers: bool = True) -> Optional[str]:
    """
    从交易记录列表直接生成汇总分析 Excel 文件。

    Args:
        records: 交易记录列表
        output_dir: 输出目录，默认为当前脚本目录
        source_info: 数据源信息，会写入"汇总总览"Sheet
        exclude_internal_transfers: 是否排除内部划转记录（默认 True）

    Returns:
        str: 生成的文件路径，如无数据则返回 None
    """
    logger = get_logger()

    if not records:
        logger.warning('无交易记录，跳过汇总分析生成')
        return None

    if output_dir is None:
        output_dir = get_script_dir()

    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'主体维度汇总分析_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, filename)

    summary_result = summarize_transactions(
        records, exclude_internal_transfers=exclude_internal_transfers,
    )

    if not summary_result.overall_summary.get('transaction_count'):
        logger.warning('汇总结果为空，跳过导出')
        return None

    return export_subject_summary(summary_result, output_path, source_info)


def generate_subject_summary_from_total(total_path: str,
                                        output_dir: Optional[str] = None,
                                        exclude_internal_transfers: bool = True) -> Optional[str]:
    """
    从银行流水总表文件生成汇总分析 Excel。

    Args:
        total_path: 银行流水总表 Excel 文件路径
        output_dir: 输出目录，默认为总表所在目录
        exclude_internal_transfers: 是否排除内部划转记录（默认 True）

    Returns:
        str: 生成的文件路径，失败则返回 None
    """
    logger = get_logger()

    records = load_total_table(total_path)
    if not records:
        logger.warning('总表无数据: %s', total_path)
        return None

    if output_dir is None:
        output_dir = os.path.dirname(total_path) or get_script_dir()

    source_info = {
        '数据来源文件': os.path.basename(total_path),
        '总表记录数': len(records),
        '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    return generate_subject_summary_from_records(
        records, output_dir, source_info,
        exclude_internal_transfers=exclude_internal_transfers,
    )


def run_subject_summary_flow(script_dir):
    """主体维度汇总分析 CLI 流程"""
    logger = get_logger()
    logger.info('========== 主体维度汇总分析开始 ==========')

    print('\n' + '=' * 70)
    print('主体维度汇总分析 - 按主体/银行/月份统计收支与笔数')
    print('=' * 70)
    print('\n请选择数据来源：')
    print('  1) 从银行流水总表文件（Excel）')
    print('  2) 从数据库（按条件查询后汇总）')
    print('  0) 返回主菜单')

    choice = input('\n请输入选项（默认 1）: ').strip() or '1'

    records = []
    source_info = {}

    if choice == '0':
        return
    elif choice == '1':
        total_path = ask_file('请选择【银行流水总表】文件')
        if not total_path:
            show_info('提示', '未选择总表文件，返回。')
            return
        logger.info('用户选择总表文件: %s', total_path)
        records = load_total_table(total_path)
        if not records:
            show_warning('错误', '总表文件无数据或读取失败。')
            return
        source_info = {
            '数据来源文件': os.path.basename(total_path),
            '总表记录数': len(records),
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    elif choice == '2':
        if not HAS_DATABASE:
            show_warning('错误', '数据库模块不可用。')
            return

        print('\n输入查询条件（直接回车表示不限制）：')
        subject = input('主体名称: ').strip() or None
        bank = input('银行名称: ').strip() or None
        start_date = input('开始日期 (YYYY-MM-DD): ').strip() or None
        end_date = input('结束日期 (YYYY-MM-DD): ').strip() or None

        try:
            qr = db_module.query_transactions(
                subject=subject, bank=bank,
                start_date=start_date, end_date=end_date,
                limit=999999, script_dir=script_dir
            )
            records = [r.to_dict() for r in qr.records]
        except Exception as e:
            show_warning('错误', f'数据库查询失败: {e}')
            logger.error('数据库查询失败: %s', e, exc_info=True)
            return

        if not records:
            show_info('提示', '查询结果为空。')
            return

        source_info = {
            '数据来源': '数据库查询',
            '查询主体': subject or '全部',
            '查询银行': bank or '全部',
            '日期范围': f'{start_date or "不限"} ~ {end_date or "不限"}',
            '记录数': len(records),
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    else:
        print('无效选项')
        return

    summary_result = summarize_transactions(records)
    overall = summary_result.overall_summary

    print('\n' + '=' * 70)
    print('汇总总览')
    print('=' * 70)
    print(f'  交易总笔数: {overall.get("transaction_count", 0):,}')
    print(f'  收入笔数:   {overall.get("income_count", 0):,}')
    print(f'  支出笔数:   {overall.get("expense_count", 0):,}')
    print(f'  收入总额:   {overall.get("total_income", 0):>15,.2f} 元')
    print(f'  支出总额:   {overall.get("total_expense", 0):>15,.2f} 元')
    print(f'  净　　额:   {overall.get("net_amount", 0):>15,.2f} 元')
    print(f'  涉及主体:   {overall.get("subject_count", 0)} 个')
    print(f'  涉及银行:   {overall.get("bank_count", 0)} 家')
    print(f'  覆盖月份:   {overall.get("month_count", 0)} 个月')

    if summary_result.by_subject:
        print('\n' + '-' * 70)
        print('按主体汇总（前 10）')
        print('-' * 70)
        print(f'{"主体":<25}{"收入(元)":>18}{"支出(元)":>18}{"净额(元)":>18}{"笔数":>8}')
        for row in summary_result.by_subject[:10]:
            print(
                f'{str(row["subject"])[:23]:<25}'
                f'{row["total_income"]:>18,.2f}'
                f'{row["total_expense"]:>18,.2f}'
                f'{row["net_amount"]:>18,.2f}'
                f'{row["transaction_count"]:>8,}'
            )

    output_dir = input(f'\n请输入输出目录（回车使用当前目录）: ').strip()
    if not output_dir:
        output_dir = script_dir

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'主体维度汇总分析_{timestamp}.xlsx')

    try:
        export_subject_summary(summary_result, output_path, source_info)
        msg = f'汇总分析已导出！\n\n输出文件：{output_path}'
        show_info('导出成功', msg)
        logger.info('主体维度汇总分析导出完成: %s', output_path)
    except Exception as e:
        msg = f'导出失败：{e}'
        show_warning('导出失败', msg)
        logger.error('主体维度汇总分析导出失败: %s', e, exc_info=True)

    logger.info('========== 主体维度汇总分析结束 ==========')


# ──────────────────────────────────────────────
# 余额连续性校验模块
# ──────────────────────────────────────────────

@dataclass
class BalanceBreakRecord:
    """余额断裂记录"""
    bank_account: str
    subject: str
    bank: str
    transaction_date: Optional[datetime]
    prev_balance: float
    receipt: float
    payment: float
    expected_balance: float
    actual_balance: float
    diff_amount: float
    transaction_id: str
    summary: str


@dataclass
class BalanceCheckResult:
    """余额连续性校验结果"""
    total_accounts: int = 0
    checked_accounts: int = 0
    skipped_accounts: int = 0
    break_count: int = 0
    break_records: List[BalanceBreakRecord] = field(default_factory=list)
    accounts_with_breaks: List[str] = field(default_factory=list)
    check_summary: Dict[str, Any] = field(default_factory=dict)


def _parse_transaction_date(value) -> Optional[datetime]:
    """解析交易日期，支持多种格式，无效值返回 None"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, pd.Timestamp):
        if pd.isna(value):
            return None
        return value.to_pydatetime()
    parsed = _normalize_date(value)
    if parsed is None:
        return None
    if isinstance(parsed, pd.Timestamp) and pd.isna(parsed):
        return None
    if pd.isna(parsed):
        return None
    return parsed


def _safe_float(value) -> float:
    """安全转换为 float，空值或无效值返回 0.0"""
    result = to_float(value)
    return result if result is not None else 0.0


def check_balance_continuity(records: List[Dict[str, Any]],
                             tolerance: float = 0.01) -> BalanceCheckResult:
    """
    余额连续性校验。

    校验逻辑：
    1. 按银行账号分组
    2. 每组内按交易日期排序
    3. 逐笔核对：上期余额 + 收款 + 付款 = 当期余额
       （付款字段约定为负数，与银行解析 payment_sign=negative 一致）
    4. 对断裂或跳变的记录生成异常清单

    Args:
        records: 交易记录列表
        tolerance: 容差，默认 0.01 元

    Returns:
        BalanceCheckResult: 校验结果
    """
    logger = get_logger()
    result = BalanceCheckResult()

    if not records:
        logger.warning('无交易记录可校验')
        result.check_summary = {'status': '无数据'}
        return result

    account_groups: Dict[str, List[Dict[str, Any]]] = {}
    for record in records:
        account = str(record.get('银行账号', '')).strip()
        if not account:
            continue
        if account not in account_groups:
            account_groups[account] = []
        account_groups[account].append(record)

    result.total_accounts = len(account_groups)
    logger.info('共 %d 个账号待校验', result.total_accounts)

    for account, account_records in account_groups.items():
        logger.debug('校验账号: %s, 记录数: %d', account, len(account_records))

        parsed_records = []
        for r in account_records:
            parsed = r.copy()
            parsed['_date'] = _parse_transaction_date(r.get('交易日期'))
            parsed['_balance'] = _safe_float(r.get('余额'))
            parsed['_receipt'] = _safe_float(r.get('收款'))
            parsed['_payment'] = _safe_float(r.get('付款'))
            parsed_records.append(parsed)

        has_valid_date = any(r['_date'] is not None for r in parsed_records)
        if not has_valid_date:
            logger.warning('账号 %s 无有效交易日期，跳过', account)
            result.skipped_accounts += 1
            continue

        parsed_records.sort(key=lambda r: (r['_date'] or datetime.min, str(r.get('交易流水号', ''))))
        result.checked_accounts += 1

        has_break = False
        for i in range(1, len(parsed_records)):
            prev = parsed_records[i - 1]
            curr = parsed_records[i]

            prev_balance = prev['_balance']
            receipt = curr['_receipt']
            payment = curr['_payment']
            actual_balance = curr['_balance']

            expected_balance = prev_balance + receipt + payment
            diff = abs(actual_balance - expected_balance)

            if diff > tolerance:
                has_break = True
                result.break_count += 1

                break_record = BalanceBreakRecord(
                    bank_account=account,
                    subject=str(curr.get('主体', '')),
                    bank=str(curr.get('银行', '')),
                    transaction_date=curr['_date'],
                    prev_balance=prev_balance,
                    receipt=receipt,
                    payment=payment,
                    expected_balance=expected_balance,
                    actual_balance=actual_balance,
                    diff_amount=actual_balance - expected_balance,
                    transaction_id=str(curr.get('交易流水号', '')),
                    summary=str(curr.get('摘要', ''))
                )
                result.break_records.append(break_record)
                logger.debug(
                    '余额断裂 - 账号: %s, 日期: %s, 预期: %.2f, 实际: %.2f, 差异: %.2f',
                    account, curr['_date'], expected_balance, actual_balance, actual_balance - expected_balance
                )

        if has_break and account not in result.accounts_with_breaks:
            result.accounts_with_breaks.append(account)

    result.check_summary = {
        'total_accounts': result.total_accounts,
        'checked_accounts': result.checked_accounts,
        'skipped_accounts': result.skipped_accounts,
        'break_count': result.break_count,
        'accounts_with_breaks_count': len(result.accounts_with_breaks),
        'tolerance': tolerance,
        'check_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    logger.info(
        '余额连续性校验完成 - 总账号: %d, 已校验: %d, 跳过: %d, 异常笔数: %d, 异常账号: %d',
        result.total_accounts, result.checked_accounts, result.skipped_accounts,
        result.break_count, len(result.accounts_with_breaks)
    )

    return result


def export_balance_check_result(check_result: BalanceCheckResult,
                                output_path: str,
                                source_info: Optional[Dict[str, Any]] = None) -> str:
    """
    导出余额连续性校验结果为 Excel 文件。

    输出的 Sheet 包括：
    1. 校验总览 - 整体统计信息
    2. 异常明细 - 所有余额断裂的交易记录
    3. 异常账号清单 - 存在余额断裂的账号列表

    Args:
        check_result: check_balance_continuity 返回的校验结果
        output_path: 输出 Excel 文件路径
        source_info: 可选，数据源信息

    Returns:
        str: 输出文件路径
    """
    logger = get_logger()

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            overview_data = []
            summary = check_result.check_summary

            overview_items = [
                ('校验项', '数值'),
                ('账号总数', summary.get('total_accounts', 0)),
                ('已校验账号数', summary.get('checked_accounts', 0)),
                ('跳过账号数', summary.get('skipped_accounts', 0)),
                ('余额断裂笔数', summary.get('break_count', 0)),
                ('异常账号数', summary.get('accounts_with_breaks_count', 0)),
                ('容差(元)', summary.get('tolerance', 0.01)),
                ('校验时间', summary.get('check_time', '')),
            ]
            if source_info:
                for k, v in source_info.items():
                    overview_items.append((k, v))

            overview_df = pd.DataFrame(overview_items[1:], columns=overview_items[0])
            overview_df.to_excel(writer, sheet_name='校验总览', index=False)

            if check_result.break_records:
                break_data = []
                for br in check_result.break_records:
                    break_data.append({
                        '主体': br.subject,
                        '银行': br.bank,
                        '银行账号': br.bank_account,
                        '交易日期': br.transaction_date.strftime('%Y-%m-%d') if br.transaction_date else '',
                        '上期余额(元)': br.prev_balance,
                        '本期收款(元)': br.receipt,
                        '本期付款(元)': br.payment,
                        '预期余额(元)': br.expected_balance,
                        '实际余额(元)': br.actual_balance,
                        '差异(元)': br.diff_amount,
                        '交易流水号': br.transaction_id,
                        '摘要': br.summary,
                    })

                break_df = pd.DataFrame(break_data)
                break_df = break_df[[
                    '主体', '银行', '银行账号', '交易日期', '上期余额(元)',
                    '本期收款(元)', '本期付款(元)', '预期余额(元)', '实际余额(元)',
                    '差异(元)', '交易流水号', '摘要'
                ]]
                break_df.to_excel(writer, sheet_name='异常明细', index=False)

                ws = writer.sheets['异常明细']
                amount_cols = set()
                for idx, col_name in enumerate(break_df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    if '元' in str(col_name):
                        amount_cols.add(col_letter)
                    max_len = max(
                        len(str(col_name)),
                        max((len(str(v)) for v in break_df.iloc[:, idx - 1].astype(str)), default=0)
                    )
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        col_letter = cell.column_letter
                        if col_letter in amount_cols:
                            cell.number_format = '#,##0.00'

            if check_result.accounts_with_breaks:
                account_data = []
                for account in check_result.accounts_with_breaks:
                    account_breaks = [br for br in check_result.break_records if br.bank_account == account]
                    sample = account_breaks[0] if account_breaks else None
                    max_diff = max((abs(br.diff_amount) for br in account_breaks), default=0)
                    account_data.append({
                        '序号': len(account_data) + 1,
                        '主体': sample.subject if sample else '',
                        '银行': sample.bank if sample else '',
                        '银行账号': account,
                        '异常笔数': len(account_breaks),
                        '最大差异(元)': max_diff,
                        '首笔异常日期': account_breaks[0].transaction_date.strftime('%Y-%m-%d')
                        if account_breaks and account_breaks[0].transaction_date else '',
                    })

                account_df = pd.DataFrame(account_data)
                account_df.to_excel(writer, sheet_name='异常账号清单', index=False)

                ws = writer.sheets['异常账号清单']
                amount_cols = set()
                count_cols = set()
                for idx, col_name in enumerate(account_df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    if '元' in str(col_name):
                        amount_cols.add(col_letter)
                    elif '笔数' in str(col_name) or '序号' in str(col_name):
                        count_cols.add(col_letter)
                    max_len = max(
                        len(str(col_name)),
                        max((len(str(v)) for v in account_df.iloc[:, idx - 1].astype(str)), default=0)
                    )
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        col_letter = cell.column_letter
                        if col_letter in amount_cols:
                            cell.number_format = '#,##0.00'
                        elif col_letter in count_cols:
                            cell.number_format = '#,##0'

            ws_overview = writer.sheets['校验总览']
            for col_idx in range(1, 3):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws_overview.column_dimensions[col_letter].width = 25

            for row in ws_overview.iter_rows(min_row=2):
                for cell in row:
                    if cell.column == 2:
                        val = cell.value
                        if isinstance(val, (int, float)):
                            if isinstance(val, float):
                                cell.number_format = '#,##0.00'
                            else:
                                cell.number_format = '#,##0'

        logger.info('余额连续性校验结果已导出: %s', output_path)
        return output_path

    except Exception as e:
        logger.error('导出余额连续性校验结果失败: %s', e, exc_info=True)
        raise


def generate_balance_check_from_records(records: List[Dict[str, Any]],
                                        output_dir: Optional[str] = None,
                                        source_info: Optional[Dict[str, Any]] = None,
                                        tolerance: float = 0.01) -> Optional[str]:
    """
    从交易记录列表直接生成余额连续性校验报告。

    Args:
        records: 交易记录列表
        output_dir: 输出目录
        source_info: 数据源信息
        tolerance: 容差

    Returns:
        str: 生成的文件路径，如无数据则返回 None
    """
    logger = get_logger()

    if not records:
        logger.warning('无交易记录，跳过余额连续性校验')
        return None

    if output_dir is None:
        output_dir = get_script_dir()

    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'余额连续性校验报告_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, filename)

    check_result = check_balance_continuity(records, tolerance=tolerance)

    return export_balance_check_result(check_result, output_path, source_info)


def generate_balance_check_from_total(total_path: str,
                                      output_dir: Optional[str] = None,
                                      tolerance: float = 0.01) -> Optional[str]:
    """
    从银行流水总表文件生成余额连续性校验报告。

    Args:
        total_path: 银行流水总表 Excel 文件路径
        output_dir: 输出目录
        tolerance: 容差

    Returns:
        str: 生成的文件路径，失败则返回 None
    """
    logger = get_logger()

    records = load_total_table(total_path)
    if not records:
        logger.warning('总表无数据: %s', total_path)
        return None

    if output_dir is None:
        output_dir = os.path.dirname(total_path) or get_script_dir()

    source_info = {
        '数据来源文件': os.path.basename(total_path),
        '总表记录数': len(records),
        '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    return generate_balance_check_from_records(records, output_dir, source_info, tolerance)


def run_balance_check_flow(script_dir):
    """余额连续性校验 CLI 流程"""
    logger = get_logger()
    logger.info('========== 余额连续性校验开始 ==========')

    print('\n' + '=' * 70)
    print('余额连续性校验 - 逐笔核对余额连续性，识别断裂或跳变')
    print('=' * 70)
    print('\n请选择数据来源：')
    print('  1) 从银行流水总表文件（Excel）')
    print('  2) 从数据库（按条件查询后校验）')
    print('  0) 返回主菜单')

    choice = input('\n请输入选项（默认 1）: ').strip() or '1'

    records = []
    source_info = {}

    if choice == '0':
        return
    elif choice == '1':
        total_path = ask_file('请选择【银行流水总表】文件')
        if not total_path:
            show_info('提示', '未选择总表文件，返回。')
            return
        logger.info('用户选择总表文件: %s', total_path)
        records = load_total_table(total_path)
        if not records:
            show_warning('错误', '总表文件无数据或读取失败。')
            return
        source_info = {
            '数据来源文件': os.path.basename(total_path),
            '总表记录数': len(records),
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    elif choice == '2':
        if not HAS_DATABASE:
            show_warning('错误', '数据库模块不可用。')
            return

        print('\n输入查询条件（直接回车表示不限制）：')
        subject = input('主体名称: ').strip() or None
        bank = input('银行名称: ').strip() or None
        account = input('银行账号: ').strip() or None
        start_date = input('开始日期 (YYYY-MM-DD): ').strip() or None
        end_date = input('结束日期 (YYYY-MM-DD): ').strip() or None

        try:
            qr = db_module.query_transactions(
                subject=subject, bank=bank, account=account,
                start_date=start_date, end_date=end_date,
                limit=999999, script_dir=script_dir
            )
            records = [r.to_dict() for r in qr.records]
        except Exception as e:
            show_warning('错误', f'数据库查询失败: {e}')
            logger.error('数据库查询失败: %s', e, exc_info=True)
            return

        if not records:
            show_info('提示', '查询结果为空。')
            return

        source_info = {
            '数据来源': '数据库查询',
            '查询主体': subject or '全部',
            '查询银行': bank or '全部',
            '查询账号': account or '全部',
            '日期范围': f'{start_date or "不限"} ~ {end_date or "不限"}',
            '记录数': len(records),
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    else:
        print('无效选项')
        return

    tolerance_input = input('\n请输入容差（元，直接回车默认 0.01）: ').strip()
    tolerance = 0.01
    if tolerance_input:
        try:
            tolerance = float(tolerance_input)
        except ValueError:
            print('输入无效，使用默认容差 0.01 元')
            tolerance = 0.01

    print(f'\n开始校验，容差: {tolerance} 元...')
    check_result = check_balance_continuity(records, tolerance=tolerance)
    summary = check_result.check_summary

    print('\n' + '=' * 70)
    print('校验结果总览')
    print('=' * 70)
    print(f'  账号总数:     {summary.get("total_accounts", 0):,}')
    print(f'  已校验账号:   {summary.get("checked_accounts", 0):,}')
    print(f'  跳过账号:     {summary.get("skipped_accounts", 0):,}')
    print(f'  余额断裂笔数: {summary.get("break_count", 0):,}')
    print(f'  异常账号数:   {summary.get("accounts_with_breaks_count", 0):,}')

    if check_result.break_count > 0:
        print(f'\n  ⚠️  发现 {check_result.break_count} 笔余额异常，涉及 {len(check_result.accounts_with_breaks)} 个账号')
        for account in check_result.accounts_with_breaks[:10]:
            account_breaks = [br for br in check_result.break_records if br.bank_account == account]
            sample = account_breaks[0] if account_breaks else None
            print(f'    - {account} ({sample.subject if sample else "未知主体"}): {len(account_breaks)} 笔异常')
        if len(check_result.accounts_with_breaks) > 10:
            print(f'    ... 还有 {len(check_result.accounts_with_breaks) - 10} 个账号，详见导出文件')
    else:
        print(f'\n  ✅ 所有账号余额连续性校验通过！')

    output_dir = input('\n请输入输出目录（直接回车默认当前目录）: ').strip()
    if not output_dir:
        output_dir = script_dir

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'余额连续性校验报告_{timestamp}.xlsx')

    try:
        export_balance_check_result(check_result, output_path, source_info)
        msg = f'校验报告已导出！\n\n输出文件：{output_path}'
        show_info('导出成功', msg)
        logger.info('余额连续性校验报告导出完成: %s', output_path)
    except Exception as e:
        msg = f'导出失败：{e}'
        show_warning('导出失败', msg)
        logger.error('余额连续性校验报告导出失败: %s', e, exc_info=True)

    logger.info('========== 余额连续性校验结束 ==========')


# ──────────────────────────────────────────────
# 重复交易检测模块
# ──────────────────────────────────────────────

DUPLICATE_CHECK_FILENAME = '重复交易检测报告.xlsx'


@dataclass
class DuplicateRecord:
    """疑似重复记录"""
    group_id: int
    record_index: int
    bank_account: str
    trade_date: str
    payment: float
    receipt: float
    counterpart: str
    transaction_id: str
    bank: str
    subject: str
    summary: str
    balance: float
    unique_id: str
    import_batch: str
    match_type: str
    match_key: str


@dataclass
class DuplicateGroup:
    """重复组：包含两条及以上疑似重复的交易"""
    group_id: int
    match_type: str
    match_key: str
    records: List[DuplicateRecord] = field(default_factory=list)
    record_count: int = 0

    def __post_init__(self):
        self.record_count = len(self.records)


@dataclass
class DuplicateCheckResult:
    """重复交易检测结果"""
    total_records: int = 0
    duplicate_group_count: int = 0
    duplicate_record_count: int = 0
    groups: List[DuplicateGroup] = field(default_factory=list)
    match_type_stats: Dict[str, int] = field(default_factory=dict)
    check_summary: Dict[str, Any] = field(default_factory=dict)


def _normalize_for_key(value) -> str:
    if value is None:
        return ''
    if isinstance(value, float) and value != value:
        return ''
    if isinstance(value, float) and value == int(value):
        return str(int(value))
    s = str(value).strip()
    if not s:
        return ''
    if '.' in s:
        try:
            f = float(s)
            if f == int(f):
                return str(int(f))
        except (ValueError, TypeError, OverflowError):
            pass
    return s


def _make_transaction_id_key(row) -> str:
    bank_account = _account_key(row.get('银行账号'))
    transaction_id = _normalize_for_key(row.get('交易流水号'))
    if transaction_id:
        return f"tid::{bank_account}::{transaction_id}"
    return ''


def _make_amount_date_key(row) -> str:
    bank_account = _account_key(row.get('银行账号'))
    trade_date = _normalize_for_key(row.get('交易日期'))
    payment = to_float(row.get('付款'))
    receipt = to_float(row.get('收款'))
    p = f"{payment:.2f}" if payment is not None else ''
    r = f"{receipt:.2f}" if receipt is not None else ''
    return f"amt::{bank_account}::{trade_date}::{p}::{r}"


def _make_full_key(row) -> str:
    bank_account = _account_key(row.get('银行账号'))
    trade_date = _normalize_for_key(row.get('交易日期'))
    payment = to_float(row.get('付款'))
    receipt = to_float(row.get('收款'))
    counterpart = _normalize_for_key(row.get('对方户名'))
    p = f"{payment:.2f}" if payment is not None else ''
    r = f"{receipt:.2f}" if receipt is not None else ''
    return f"full::{bank_account}::{trade_date}::{p}::{r}::{counterpart}"


def detect_duplicates(records: List[Dict[str, Any]]) -> DuplicateCheckResult:
    """
    重复交易检测。

    检测逻辑（三级匹配，由严到宽）：
    1. 交易流水号匹配（match_type=transaction_id）：
       同一银行账号 + 同一交易流水号 → 银行重复记账或重复导入
    2. 金额日期匹配（match_type=amount_date）：
       同一银行账号 + 同一交易日期 + 同一付款金额 + 同一收款金额
       → 疑似重复（不同流水号但金额/日期/账号完全一致）
    3. 完全匹配（match_type=full）：
       同一银行账号 + 同一交易日期 + 同一付款 + 同一收款 + 同一对方户名
       → 最高疑似度重复

    同一条记录只归入最高优先级的匹配组，不会重复计数。

    Args:
        records: 交易记录列表

    Returns:
        DuplicateCheckResult: 检测结果
    """
    logger = get_logger()
    result = DuplicateCheckResult()
    result.total_records = len(records)

    if not records:
        logger.warning('无交易记录可检测重复')
        result.check_summary = {'status': '无数据'}
        return result

    tid_groups: Dict[str, List[int]] = {}
    amt_groups: Dict[str, List[int]] = {}
    full_groups: Dict[str, List[int]] = {}

    for idx, rec in enumerate(records):
        tid_key = _make_transaction_id_key(rec)
        if tid_key:
            tid_groups.setdefault(tid_key, []).append(idx)

        amt_key = _make_amount_date_key(rec)
        if amt_key:
            amt_groups.setdefault(amt_key, []).append(idx)

        full_key = _make_full_key(rec)
        if full_key:
            full_groups.setdefault(full_key, []).append(idx)

    claimed: set = set()
    group_id_counter = 0
    match_type_stats = {'transaction_id': 0, 'amount_date': 0, 'full': 0}

    for key, indices in tid_groups.items():
        if len(indices) < 2:
            continue
        group_id_counter += 1
        group = DuplicateGroup(
            group_id=group_id_counter,
            match_type='transaction_id',
            match_key=key,
        )
        for idx in indices:
            rec = records[idx]
            dup_rec = DuplicateRecord(
                group_id=group_id_counter,
                record_index=idx + 1,
                bank_account=str(rec.get('银行账号', '')),
                trade_date=str(rec.get('交易日期', '')),
                payment=_safe_float(rec.get('付款')),
                receipt=_safe_float(rec.get('收款')),
                counterpart=str(rec.get('对方户名', '')),
                transaction_id=str(rec.get('交易流水号', '')),
                bank=str(rec.get('银行', '')),
                subject=str(rec.get('主体', '')),
                summary=str(rec.get('摘要', '')),
                balance=_safe_float(rec.get('余额')),
                unique_id=str(rec.get('唯一id', '')),
                import_batch=str(rec.get('导入批次号', '')),
                match_type='transaction_id',
                match_key=key,
            )
            group.records.append(dup_rec)
            claimed.add(idx)
        group.record_count = len(group.records)
        result.groups.append(group)
        match_type_stats['transaction_id'] += 1

    for key, indices in amt_groups.items():
        if len(indices) < 2:
            continue
        unclaimed = [i for i in indices if i not in claimed]
        if len(unclaimed) < 2:
            continue
        group_id_counter += 1
        group = DuplicateGroup(
            group_id=group_id_counter,
            match_type='amount_date',
            match_key=key,
        )
        for idx in unclaimed:
            rec = records[idx]
            dup_rec = DuplicateRecord(
                group_id=group_id_counter,
                record_index=idx + 1,
                bank_account=str(rec.get('银行账号', '')),
                trade_date=str(rec.get('交易日期', '')),
                payment=_safe_float(rec.get('付款')),
                receipt=_safe_float(rec.get('收款')),
                counterpart=str(rec.get('对方户名', '')),
                transaction_id=str(rec.get('交易流水号', '')),
                bank=str(rec.get('银行', '')),
                subject=str(rec.get('主体', '')),
                summary=str(rec.get('摘要', '')),
                balance=_safe_float(rec.get('余额')),
                unique_id=str(rec.get('唯一id', '')),
                import_batch=str(rec.get('导入批次号', '')),
                match_type='amount_date',
                match_key=key,
            )
            group.records.append(dup_rec)
            claimed.add(idx)
        group.record_count = len(group.records)
        result.groups.append(group)
        match_type_stats['amount_date'] += 1

    for key, indices in full_groups.items():
        if len(indices) < 2:
            continue
        unclaimed = [i for i in indices if i not in claimed]
        if len(unclaimed) < 2:
            continue
        group_id_counter += 1
        group = DuplicateGroup(
            group_id=group_id_counter,
            match_type='full',
            match_key=key,
        )
        for idx in unclaimed:
            rec = records[idx]
            dup_rec = DuplicateRecord(
                group_id=group_id_counter,
                record_index=idx + 1,
                bank_account=str(rec.get('银行账号', '')),
                trade_date=str(rec.get('交易日期', '')),
                payment=_safe_float(rec.get('付款')),
                receipt=_safe_float(rec.get('收款')),
                counterpart=str(rec.get('对方户名', '')),
                transaction_id=str(rec.get('交易流水号', '')),
                bank=str(rec.get('银行', '')),
                subject=str(rec.get('主体', '')),
                summary=str(rec.get('摘要', '')),
                balance=_safe_float(rec.get('余额')),
                unique_id=str(rec.get('唯一id', '')),
                import_batch=str(rec.get('导入批次号', '')),
                match_type='full',
                match_key=key,
            )
            group.records.append(dup_rec)
            claimed.add(idx)
        group.record_count = len(group.records)
        result.groups.append(group)
        match_type_stats['full'] += 1

    result.duplicate_group_count = len(result.groups)
    result.duplicate_record_count = sum(g.record_count for g in result.groups)
    result.match_type_stats = match_type_stats

    tid_rec_count = sum(g.record_count for g in result.groups if g.match_type == 'transaction_id')
    amt_rec_count = sum(g.record_count for g in result.groups if g.match_type == 'amount_date')
    full_rec_count = sum(g.record_count for g in result.groups if g.match_type == 'full')

    result.check_summary = {
        'total_records': result.total_records,
        'duplicate_group_count': result.duplicate_group_count,
        'duplicate_record_count': result.duplicate_record_count,
        'duplicate_rate': round(result.duplicate_record_count / result.total_records * 100, 2) if result.total_records else 0,
        'transaction_id_groups': match_type_stats.get('transaction_id', 0),
        'transaction_id_records': tid_rec_count,
        'amount_date_groups': match_type_stats.get('amount_date', 0),
        'amount_date_records': amt_rec_count,
        'full_groups': match_type_stats.get('full', 0),
        'full_records': full_rec_count,
        'check_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    logger.info(
        '重复交易检测完成 - 总记录: %d, 疑似重复组: %d, 疑似重复记录: %d, 重复率: %.2f%%',
        result.total_records, result.duplicate_group_count,
        result.duplicate_record_count, result.check_summary['duplicate_rate']
    )

    return result


def export_duplicate_check_result(check_result: DuplicateCheckResult,
                                  output_path: str,
                                  source_info: Optional[Dict[str, Any]] = None) -> str:
    """
    导出重复交易检测结果为 Excel 文件。

    输出的 Sheet 包括：
    1. 检测总览 - 整体统计信息
    2. 疑似重复明细 - 所有疑似重复记录的详细列表
    3. 重复组汇总 - 按重复组汇总信息

    Args:
        check_result: detect_duplicates 返回的检测结果
        output_path: 输出 Excel 文件路径
        source_info: 可选，数据源信息

    Returns:
        str: 输出文件路径
    """
    logger = get_logger()

    MATCH_TYPE_LABELS = {
        'transaction_id': '交易流水号匹配',
        'amount_date': '金额+日期匹配',
        'full': '完全匹配(含对方户名)',
    }

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            overview_items = [
                ('检测项', '数值'),
                ('总记录数', check_result.check_summary.get('total_records', 0)),
                ('疑似重复组数', check_result.check_summary.get('duplicate_group_count', 0)),
                ('疑似重复记录数', check_result.check_summary.get('duplicate_record_count', 0)),
                ('重复率(%)', check_result.check_summary.get('duplicate_rate', 0)),
                ('流水号匹配组数', check_result.check_summary.get('transaction_id_groups', 0)),
                ('流水号匹配记录数', check_result.check_summary.get('transaction_id_records', 0)),
                ('金额日期匹配组数', check_result.check_summary.get('amount_date_groups', 0)),
                ('金额日期匹配记录数', check_result.check_summary.get('amount_date_records', 0)),
                ('完全匹配组数', check_result.check_summary.get('full_groups', 0)),
                ('完全匹配记录数', check_result.check_summary.get('full_records', 0)),
                ('检测时间', check_result.check_summary.get('check_time', '')),
            ]
            if source_info:
                for k, v in source_info.items():
                    overview_items.append((k, v))

            overview_df = pd.DataFrame(overview_items[1:], columns=overview_items[0])
            overview_df.to_excel(writer, sheet_name='检测总览', index=False)

            if check_result.groups:
                detail_data = []
                for group in check_result.groups:
                    for rec in group.records:
                        detail_data.append({
                            '重复组ID': rec.group_id,
                            '匹配类型': MATCH_TYPE_LABELS.get(rec.match_type, rec.match_type),
                            '序号(总表)': rec.record_index,
                            '唯一ID': rec.unique_id,
                            '主体': rec.subject,
                            '银行': rec.bank,
                            '银行账号': rec.bank_account,
                            '交易日期': rec.trade_date,
                            '付款(元)': rec.payment,
                            '收款(元)': rec.receipt,
                            '对方户名': rec.counterpart,
                            '余额(元)': rec.balance,
                            '交易流水号': rec.transaction_id,
                            '摘要': rec.summary,
                            '导入批次号': rec.import_batch,
                        })

                detail_df = pd.DataFrame(detail_data)
                detail_cols = [
                    '重复组ID', '匹配类型', '序号(总表)', '唯一ID',
                    '主体', '银行', '银行账号', '交易日期',
                    '付款(元)', '收款(元)', '对方户名', '余额(元)',
                    '交易流水号', '摘要', '导入批次号',
                ]
                detail_df = detail_df[[c for c in detail_cols if c in detail_df.columns]]
                detail_df.to_excel(writer, sheet_name='疑似重复明细', index=False)

                ws = writer.sheets['疑似重复明细']
                amount_cols = set()
                for idx, col_name in enumerate(detail_df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    if '元' in str(col_name):
                        amount_cols.add(col_letter)
                    max_len = max(
                        len(str(col_name)),
                        max((len(str(v)) for v in detail_df.iloc[:, idx - 1].astype(str)), default=0)
                    )
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        if cell.column_letter in amount_cols:
                            cell.number_format = '#,##0.00'

                group_data = []
                for group in check_result.groups:
                    first_rec = group.records[0] if group.records else None
                    same_batch = all(
                        r.import_batch == group.records[0].import_batch
                        for r in group.records
                    ) if group.records else True

                    duplicate_source = '同一批次导入' if same_batch else '跨批次导入'
                    if group.match_type == 'transaction_id':
                        duplicate_source += ' / 银行重复记账可能性高'
                    elif group.match_type == 'full':
                        duplicate_source += ' / 对方户名一致，疑似重复'
                    else:
                        duplicate_source += ' / 流水号不同但金额日期一致'

                    group_data.append({
                        '重复组ID': group.group_id,
                        '匹配类型': MATCH_TYPE_LABELS.get(group.match_type, group.match_type),
                        '重复记录数': group.record_count,
                        '主体': first_rec.subject if first_rec else '',
                        '银行': first_rec.bank if first_rec else '',
                        '银行账号': first_rec.bank_account if first_rec else '',
                        '交易日期': first_rec.trade_date if first_rec else '',
                        '付款(元)': first_rec.payment if first_rec else 0,
                        '收款(元)': first_rec.receipt if first_rec else 0,
                        '对方户名': first_rec.counterpart if first_rec else '',
                        '交易流水号': first_rec.transaction_id if first_rec else '',
                        '重复来源分析': duplicate_source,
                    })

                group_df = pd.DataFrame(group_data)
                group_df.to_excel(writer, sheet_name='重复组汇总', index=False)

                ws_group = writer.sheets['重复组汇总']
                amount_cols_g = set()
                count_cols_g = set()
                for idx, col_name in enumerate(group_df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    if '元' in str(col_name):
                        amount_cols_g.add(col_letter)
                    elif '数' in str(col_name) or 'ID' in str(col_name):
                        count_cols_g.add(col_letter)
                    max_len = max(
                        len(str(col_name)),
                        max((len(str(v)) for v in group_df.iloc[:, idx - 1].astype(str)), default=0)
                    )
                    ws_group.column_dimensions[col_letter].width = min(max_len + 4, 50)

                for row in ws_group.iter_rows(min_row=2):
                    for cell in row:
                        if cell.column_letter in amount_cols_g:
                            cell.number_format = '#,##0.00'
                        elif cell.column_letter in count_cols_g:
                            cell.number_format = '#,##0'

            ws_overview = writer.sheets['检测总览']
            for col_idx in range(1, 3):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws_overview.column_dimensions[col_letter].width = 25

            for row in ws_overview.iter_rows(min_row=2):
                for cell in row:
                    if cell.column == 2:
                        val = cell.value
                        if isinstance(val, (int, float)):
                            if isinstance(val, float):
                                cell.number_format = '#,##0.00'
                            else:
                                cell.number_format = '#,##0'

        logger.info('重复交易检测结果已导出: %s', output_path)
        return output_path

    except Exception as e:
        logger.error('导出重复交易检测结果失败: %s', e, exc_info=True)
        raise


def generate_duplicate_check_from_records(records: List[Dict[str, Any]],
                                          output_dir: Optional[str] = None,
                                          source_info: Optional[Dict[str, Any]] = None) -> Optional[str]:
    """
    从交易记录列表直接生成重复交易检测报告。

    Args:
        records: 交易记录列表
        output_dir: 输出目录
        source_info: 数据源信息

    Returns:
        str: 生成的文件路径，如无数据则返回 None
    """
    logger = get_logger()

    if not records:
        logger.warning('无交易记录，跳过重复交易检测')
        return None

    if output_dir is None:
        output_dir = get_script_dir()

    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'重复交易检测报告_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, filename)

    check_result = detect_duplicates(records)

    if not check_result.duplicate_group_count:
        logger.info('未检测到重复交易，仍导出报告')
        return export_duplicate_check_result(check_result, output_path, source_info)

    return export_duplicate_check_result(check_result, output_path, source_info)


def generate_duplicate_check_from_total(total_path: str,
                                        output_dir: Optional[str] = None) -> Optional[str]:
    """
    从银行流水总表文件生成重复交易检测报告。

    Args:
        total_path: 银行流水总表 Excel 文件路径
        output_dir: 输出目录

    Returns:
        str: 生成的文件路径，失败则返回 None
    """
    logger = get_logger()

    records = load_total_table(total_path)
    if not records:
        logger.warning('总表无数据: %s', total_path)
        return None

    if output_dir is None:
        output_dir = os.path.dirname(total_path) or get_script_dir()

    source_info = {
        '数据来源文件': os.path.basename(total_path),
        '总表记录数': len(records),
        '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    return generate_duplicate_check_from_records(records, output_dir, source_info)


def run_duplicate_check_flow(script_dir):
    """重复交易检测 CLI 流程"""
    logger = get_logger()
    logger.info('========== 重复交易检测开始 ==========')

    print('\n' + '=' * 70)
    print('重复交易检测 - 跨文件去重与疑似重复标记')
    print('=' * 70)
    print('\n请选择数据来源：')
    print('  1) 从银行流水总表文件（Excel）')
    print('  2) 从数据库（按条件查询后检测）')
    print('  0) 返回主菜单')

    choice = input('\n请输入选项（默认 1）: ').strip() or '1'

    records = []
    source_info = {}

    if choice == '0':
        return
    elif choice == '1':
        total_path = ask_file('请选择【银行流水总表】文件')
        if not total_path:
            show_info('提示', '未选择总表文件，返回。')
            return
        logger.info('用户选择总表文件: %s', total_path)
        records = load_total_table(total_path)
        if not records:
            show_warning('错误', '总表文件无数据或读取失败。')
            return
        source_info = {
            '数据来源文件': os.path.basename(total_path),
            '总表记录数': len(records),
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    elif choice == '2':
        if not HAS_DATABASE:
            show_warning('错误', '数据库模块不可用。')
            return

        print('\n输入查询条件（直接回车表示不限制）：')
        subject = input('主体名称: ').strip() or None
        bank = input('银行名称: ').strip() or None
        account = input('银行账号: ').strip() or None
        start_date = input('开始日期 (YYYY-MM-DD): ').strip() or None
        end_date = input('结束日期 (YYYY-MM-DD): ').strip() or None

        try:
            qr = db_module.query_transactions(
                subject=subject, bank=bank, account=account,
                start_date=start_date, end_date=end_date,
                limit=999999, script_dir=script_dir
            )
            records = [r.to_dict() for r in qr.records]
        except Exception as e:
            show_warning('错误', f'数据库查询失败: {e}')
            logger.error('数据库查询失败: %s', e, exc_info=True)
            return

        if not records:
            show_info('提示', '查询结果为空。')
            return

        source_info = {
            '数据来源': '数据库查询',
            '查询主体': subject or '全部',
            '查询银行': bank or '全部',
            '查询账号': account or '全部',
            '日期范围': f'{start_date or "不限"} ~ {end_date or "不限"}',
            '记录数': len(records),
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    else:
        print('无效选项')
        return

    print(f'\n开始检测，共 {len(records)} 条记录...')
    check_result = detect_duplicates(records)
    summary = check_result.check_summary

    MATCH_TYPE_LABELS = {
        'transaction_id': '交易流水号匹配',
        'amount_date': '金额+日期匹配',
        'full': '完全匹配(含对方户名)',
    }

    print('\n' + '=' * 70)
    print('检测结果总览')
    print('=' * 70)
    print(f'  总记录数:         {summary.get("total_records", 0):,}')
    print(f'  疑似重复组数:     {summary.get("duplicate_group_count", 0):,}')
    print(f'  疑似重复记录数:   {summary.get("duplicate_record_count", 0):,}')
    print(f'  重复率:           {summary.get("duplicate_rate", 0):.2f}%')
    print()
    print(f'  流水号匹配:       {summary.get("transaction_id_groups", 0)} 组 / {summary.get("transaction_id_records", 0)} 条')
    print(f'  金额日期匹配:     {summary.get("amount_date_groups", 0)} 组 / {summary.get("amount_date_records", 0)} 条')
    print(f'  完全匹配:         {summary.get("full_groups", 0)} 组 / {summary.get("full_records", 0)} 条')

    if check_result.duplicate_group_count > 0:
        print(f'\n  ⚠️  发现 {check_result.duplicate_group_count} 组疑似重复交易，涉及 {check_result.duplicate_record_count} 条记录')
        for group in check_result.groups[:10]:
            first = group.records[0] if group.records else None
            print(f'    - 组{group.group_id} [{MATCH_TYPE_LABELS.get(group.match_type, group.match_type)}]: '
                  f'{first.bank_account if first else ""} / {first.trade_date if first else ""} / '
                  f'{group.record_count}条')
        if len(check_result.groups) > 10:
            print(f'    ... 还有 {len(check_result.groups) - 10} 组，详见导出文件')
    else:
        print(f'\n  ✅ 未检测到重复交易！')

    output_dir = input('\n请输入输出目录（直接回车默认当前目录）: ').strip()
    if not output_dir:
        output_dir = script_dir

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'重复交易检测报告_{timestamp}.xlsx')

    try:
        export_duplicate_check_result(check_result, output_path, source_info)
        msg = f'检测报告已导出！\n\n输出文件：{output_path}'
        show_info('导出成功', msg)
        logger.info('重复交易检测报告导出完成: %s', output_path)
    except Exception as e:
        msg = f'导出失败：{e}'
        show_warning('导出失败', msg)
        logger.error('重复交易检测报告导出失败: %s', e, exc_info=True)

    logger.info('========== 重复交易检测结束 ==========')


# ──────────────────────────────────────────────
# 利息与手续费专项核对模块
# ──────────────────────────────────────────────

INTEREST_FEE_KEYWORDS = [
    '利息', '结息', '手续费', '服务费', '管理费', '年费', '工本费',
    '短信费', '网银费', '账户管理费', '结算手续费', '汇兑手续费',
    '转账手续费', '提现手续费', '支付手续费', '跨行手续费',
    'interest', 'fee', 'charge', 'commission',
]

TRANSACTION_TYPE_INTEREST = 'interest'
TRANSACTION_TYPE_FEE = 'fee'

FEE_CATEGORY_KEYWORDS = {
    '利息收入': ['利息收入', '存款利息', '活期利息', '定期利息', '结息'],
    '利息支出': ['利息支出', '贷款利息', '借款利息', '罚息'],
    '转账手续费': ['转账', '汇款', '汇兑', '跨行'],
    '提现手续费': ['提现', '取现'],
    '支付手续费': ['支付', '快捷支付', '网关支付'],
    '账户管理费': ['账户管理费', '年费', '账户维护费'],
    '网银服务费': ['网银', '网上银行', '手机银行'],
    '短信服务费': ['短信', '通知'],
    '其他手续费': ['手续费', '服务费', '工本费', '其他'],
}

EXPECTED_RATE_CONFIG = {
    '转账手续费': {'rate': 0.005, 'min': 1.0, 'max': 50.0, 'base': 'amount'},
    '提现手续费': {'rate': 0.001, 'min': 0.1, 'max': 100.0, 'base': 'amount'},
    '跨行手续费': {'rate': 0.003, 'min': 2.0, 'max': 50.0, 'base': 'amount'},
    '账户管理费': {'fixed': 10.0, 'base': 'monthly'},
    '年费': {'fixed': 100.0, 'base': 'yearly'},
    '短信服务费': {'fixed': 3.0, 'base': 'monthly'},
}

ANOMALY_THRESHOLD = {
    'vs_expected_rate': 0.3,
    'vs_historical_mean': 0.5,
    'period_over_period': 0.5,
}


@dataclass
class InterestFeeTransaction:
    """利息手续费交易记录"""
    transaction_id: str
    trade_date: Optional[datetime]
    bank: str
    bank_account: str
    subject: str
    summary: str
    amount: float
    counterpart: str
    transaction_type: str
    fee_category: str
    matched_keyword: str


@dataclass
class PeriodSummary:
    """期间汇总记录"""
    period: str
    period_type: str
    subject: str
    bank: str
    fee_category: str
    transaction_type: str
    transaction_count: int
    total_amount: float
    avg_amount: float
    max_amount: float
    min_amount: float
    historical_mean: float
    historical_std: float
    expected_amount: float
    deviation_from_expected: float
    deviation_from_historical: float
    period_over_period_change: float
    is_anomaly: bool
    anomaly_reasons: List[str]


@dataclass
class InterestFeeCheckResult:
    """利息手续费核对结果"""
    total_records: int = 0
    interest_records: int = 0
    fee_records: int = 0
    total_interest_amount: float = 0.0
    total_fee_amount: float = 0.0
    filtered_transactions: List[InterestFeeTransaction] = field(default_factory=list)
    period_summaries: List[PeriodSummary] = field(default_factory=list)
    anomaly_count: int = 0
    anomaly_summaries: List[PeriodSummary] = field(default_factory=list)
    check_summary: Dict[str, Any] = field(default_factory=dict)


def _classify_transaction(summary: str) -> Tuple[str, str, str]:
    """
    分类交易：识别交易类型和费用类别

    Returns:
        (transaction_type, fee_category, matched_keyword)
    """
    summary_lower = str(summary).lower() if summary else ''
    summary_cn = str(summary) if summary else ''

    matched_keyword = ''
    for kw in INTEREST_FEE_KEYWORDS:
        if kw.lower() in summary_lower or kw in summary_cn:
            matched_keyword = kw
            break

    if not matched_keyword:
        return '', '', ''

    transaction_type = TRANSACTION_TYPE_FEE
    fee_category = '其他手续费'

    if '利息' in summary_cn or 'interest' in summary_lower or '结息' in summary_cn:
        transaction_type = TRANSACTION_TYPE_INTEREST
        for category, keywords in FEE_CATEGORY_KEYWORDS.items():
            if '利息' in category:
                for kw in keywords:
                    if kw in summary_cn or kw.lower() in summary_lower:
                        fee_category = category
                        break
                if fee_category != '其他手续费':
                    break
        if fee_category == '其他手续费':
            fee_category = '利息收入'
    else:
        for category, keywords in FEE_CATEGORY_KEYWORDS.items():
            if '利息' not in category:
                for kw in keywords:
                    if kw in summary_cn or kw.lower() in summary_lower:
                        fee_category = category
                        break
                if fee_category != '其他手续费':
                    break

    return transaction_type, fee_category, matched_keyword


def _get_period_key(date_val: Optional[datetime], period_type: str = 'month') -> str:
    """获取期间键值"""
    if date_val is None:
        return '未知'
    if period_type == 'year':
        return date_val.strftime('%Y')
    elif period_type == 'quarter':
        quarter = (date_val.month - 1) // 3 + 1
        return f'{date_val.year}Q{quarter}'
    elif period_type == 'week':
        return date_val.strftime('%Y-W%W')
    else:
        return date_val.strftime('%Y-%m')


def filter_interest_fee_transactions(records: List[Dict[str, Any]]) -> List[InterestFeeTransaction]:
    """
    筛选摘要含利息、手续费等关键词的交易

    Args:
        records: 交易记录列表

    Returns:
        筛选后的利息手续费交易列表
    """
    logger = get_logger()
    result = []

    for record in records:
        summary = record.get('摘要', '')
        if not summary:
            continue

        trans_type, fee_category, matched_kw = _classify_transaction(summary)
        if not trans_type:
            continue

        payment = _safe_float(record.get('付款'))
        receipt = _safe_float(record.get('收款'))
        amount = receipt if receipt > 0 else abs(payment)

        trade_date = _parse_transaction_date(record.get('交易日期'))

        txn = InterestFeeTransaction(
            transaction_id=str(record.get('交易流水号', '')),
            trade_date=trade_date,
            bank=str(record.get('银行', '')),
            bank_account=str(record.get('银行账号', '')),
            subject=str(record.get('主体', '')),
            summary=str(summary),
            amount=amount,
            counterpart=str(record.get('对方户名', '')),
            transaction_type=trans_type,
            fee_category=fee_category,
            matched_keyword=matched_kw,
        )
        result.append(txn)

    logger.info('筛选出 %d 条利息手续费交易', len(result))
    return result


def _calculate_expected_amount(category: str, amount: float, period_type: str) -> float:
    """计算预期金额"""
    config = EXPECTED_RATE_CONFIG.get(category, {})
    if not config:
        return 0.0

    base = config.get('base', 'amount')
    if base == 'amount':
        rate = config.get('rate', 0)
        min_fee = config.get('min', 0)
        max_fee = config.get('max', float('inf'))
        expected = amount * rate
        return max(min_fee, min(expected, max_fee))
    elif base == 'monthly':
        return config.get('fixed', 0.0)
    elif base == 'yearly':
        if period_type == 'year':
            return config.get('fixed', 0.0)
        return config.get('fixed', 0.0) / 12

    return 0.0


def summarize_by_period(transactions: List[InterestFeeTransaction],
                        period_type: str = 'month') -> List[PeriodSummary]:
    """
    按期间汇总利息手续费交易

    Args:
        transactions: 利息手续费交易列表
        period_type: 期间类型 ('year', 'quarter', 'month', 'week')

    Returns:
        期间汇总列表
    """
    logger = get_logger()

    summary_map: Dict[Tuple[str, str, str, str, str], List[InterestFeeTransaction]] = {}
    for txn in transactions:
        period = _get_period_key(txn.trade_date, period_type)
        key = (period, txn.subject, txn.bank, txn.fee_category, txn.transaction_type)
        if key not in summary_map:
            summary_map[key] = []
        summary_map[key].append(txn)

    category_history: Dict[Tuple[str, str, str, str], List[float]] = {}
    period_totals: Dict[Tuple[str, str, str, str, str], float] = {}
    for key, txns in summary_map.items():
        period, subject, bank, category, txn_type = key
        total = sum(t.amount for t in txns)
        period_totals[key] = total
        hist_key = (subject, bank, category, txn_type)
        if hist_key not in category_history:
            category_history[hist_key] = []
        category_history[hist_key].append(total)

    sorted_periods = sorted({k[0] for k in summary_map.keys()})
    prev_period_totals: Dict[Tuple[str, str, str, str], float] = {}

    result = []
    for period in sorted_periods:
        for key, txns in summary_map.items():
            if key[0] != period:
                continue

            _, subject, bank, category, txn_type = key
            amounts = [t.amount for t in txns]
            total_amount = sum(amounts)
            count = len(amounts)

            hist_key = (subject, bank, category, txn_type)
            history = category_history.get(hist_key, [])
            if len(history) > 1:
                historical_mean = sum(history) / len(history)
                variance = sum((x - historical_mean) ** 2 for x in history) / len(history)
                historical_std = variance ** 0.5
            else:
                historical_mean = total_amount
                historical_std = 0.0

            expected_amount = 0.0
            for t in txns:
                expected_amount += _calculate_expected_amount(category, t.amount, period_type)

            deviation_from_expected = 0.0
            if expected_amount > 0:
                deviation_from_expected = (total_amount - expected_amount) / expected_amount

            deviation_from_historical = 0.0
            if historical_mean > 0:
                deviation_from_historical = (total_amount - historical_mean) / historical_mean

            pop_change = 0.0
            prev_total = prev_period_totals.get(hist_key, 0.0)
            if prev_total > 0:
                pop_change = (total_amount - prev_total) / prev_total
            prev_period_totals[hist_key] = total_amount

            anomaly_reasons = []
            is_anomaly = False

            if abs(deviation_from_expected) > ANOMALY_THRESHOLD['vs_expected_rate']:
                is_anomaly = True
                anomaly_reasons.append(f'与预期费率偏差{deviation_from_expected*100:.1f}%，超出阈值{ANOMALY_THRESHOLD["vs_expected_rate"]*100:.0f}%')

            if abs(deviation_from_historical) > ANOMALY_THRESHOLD['vs_historical_mean']:
                is_anomaly = True
                anomaly_reasons.append(f'与历史均值偏差{deviation_from_historical*100:.1f}%，超出阈值{ANOMALY_THRESHOLD["vs_historical_mean"]*100:.0f}%')

            if abs(pop_change) > ANOMALY_THRESHOLD['period_over_period']:
                is_anomaly = True
                anomaly_reasons.append(f'环比变动{pop_change*100:.1f}%，超出阈值{ANOMALY_THRESHOLD["period_over_period"]*100:.0f}%')

            summary = PeriodSummary(
                period=period,
                period_type=period_type,
                subject=subject,
                bank=bank,
                fee_category=category,
                transaction_type=txn_type,
                transaction_count=count,
                total_amount=total_amount,
                avg_amount=total_amount / count if count > 0 else 0.0,
                max_amount=max(amounts) if amounts else 0.0,
                min_amount=min(amounts) if amounts else 0.0,
                historical_mean=historical_mean,
                historical_std=historical_std,
                expected_amount=expected_amount,
                deviation_from_expected=deviation_from_expected,
                deviation_from_historical=deviation_from_historical,
                period_over_period_change=pop_change,
                is_anomaly=is_anomaly,
                anomaly_reasons=anomaly_reasons,
            )
            result.append(summary)

    logger.info('按%s汇总生成 %d 条期间汇总记录，异常 %d 条',
                period_type, len(result), sum(1 for s in result if s.is_anomaly))
    return result


def check_interest_fee(records: List[Dict[str, Any]],
                       period_type: str = 'month') -> InterestFeeCheckResult:
    """
    利息与手续费专项核对

    核对逻辑：
    1. 筛选摘要含利息、手续费等关键词的交易
    2. 按期间（月/季/年）汇总
    3. 与预期费率对比
    4. 与历史均值对比
    5. 环比波动分析
    6. 标记异常波动

    Args:
        records: 交易记录列表
        period_type: 汇总期间类型

    Returns:
        InterestFeeCheckResult: 核对结果
    """
    logger = get_logger()
    result = InterestFeeCheckResult()

    if not records:
        logger.warning('无交易记录可核对')
        result.check_summary = {'status': '无数据'}
        return result

    filtered = filter_interest_fee_transactions(records)
    result.filtered_transactions = filtered
    result.total_records = len(filtered)

    for txn in filtered:
        if txn.transaction_type == TRANSACTION_TYPE_INTEREST:
            result.interest_records += 1
            result.total_interest_amount += txn.amount
        else:
            result.fee_records += 1
            result.total_fee_amount += txn.amount

    summaries = summarize_by_period(filtered, period_type)
    result.period_summaries = summaries
    result.anomaly_summaries = [s for s in summaries if s.is_anomaly]
    result.anomaly_count = len(result.anomaly_summaries)

    result.check_summary = {
        'total_records': result.total_records,
        'interest_records': result.interest_records,
        'fee_records': result.fee_records,
        'total_interest_amount': result.total_interest_amount,
        'total_fee_amount': result.total_fee_amount,
        'period_count': len(summaries),
        'anomaly_count': result.anomaly_count,
        'period_type': period_type,
        'check_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    logger.info(
        '利息手续费核对完成 - 总记录: %d (利息: %d, 手续费: %d), '
        '期间汇总: %d, 异常: %d',
        result.total_records, result.interest_records, result.fee_records,
        len(summaries), result.anomaly_count
    )

    return result


def export_interest_fee_check_result(check_result: InterestFeeCheckResult,
                                     output_path: str,
                                     source_info: Optional[Dict[str, Any]] = None) -> str:
    """
    导出利息手续费核对结果为 Excel 文件

    输出的 Sheet 包括：
    1. 核对总览 - 整体统计信息
    2. 交易明细 - 所有筛选出的利息手续费交易
    3. 期间汇总 - 按期间汇总统计
    4. 异常清单 - 异常波动记录

    Args:
        check_result: 核对结果
        output_path: 输出文件路径
        source_info: 数据源信息

    Returns:
        str: 输出文件路径
    """
    logger = get_logger()

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            overview_data = []
            summary = check_result.check_summary

            overview_items = [
                ('核对项', '数值'),
                ('利息手续费交易总数', summary.get('total_records', 0)),
                ('利息类交易数', summary.get('interest_records', 0)),
                ('手续费类交易数', summary.get('fee_records', 0)),
                ('利息总金额(元)', summary.get('total_interest_amount', 0.0)),
                ('手续费总金额(元)', summary.get('total_fee_amount', 0.0)),
                ('期间汇总数', summary.get('period_count', 0)),
                ('异常汇总数', summary.get('anomaly_count', 0)),
                ('汇总期间', summary.get('period_type', 'month')),
                ('核对时间', summary.get('check_time', '')),
            ]
            if source_info:
                for k, v in source_info.items():
                    overview_items.append((k, v))

            overview_df = pd.DataFrame(overview_items[1:], columns=overview_items[0])
            overview_df.to_excel(writer, sheet_name='核对总览', index=False)

            if check_result.filtered_transactions:
                txn_data = []
                for txn in check_result.filtered_transactions:
                    txn_data.append({
                        '期间': _get_period_key(txn.trade_date, summary.get('period_type', 'month')),
                        '主体': txn.subject,
                        '银行': txn.bank,
                        '银行账号': txn.bank_account,
                        '交易日期': txn.trade_date.strftime('%Y-%m-%d') if txn.trade_date else '',
                        '交易类型': '利息' if txn.transaction_type == TRANSACTION_TYPE_INTEREST else '手续费',
                        '费用类别': txn.fee_category,
                        '金额(元)': txn.amount,
                        '摘要': txn.summary,
                        '对方户名': txn.counterpart,
                        '匹配关键词': txn.matched_keyword,
                        '交易流水号': txn.transaction_id,
                    })

                txn_df = pd.DataFrame(txn_data)
                txn_df = txn_df[[
                    '期间', '主体', '银行', '银行账号', '交易日期',
                    '交易类型', '费用类别', '金额(元)', '摘要',
                    '对方户名', '匹配关键词', '交易流水号'
                ]]
                txn_df.to_excel(writer, sheet_name='交易明细', index=False)

                ws = writer.sheets['交易明细']
                for idx, col_name in enumerate(txn_df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    max_len = max(
                        len(str(col_name)),
                        max((len(str(v)) for v in txn_df.iloc[:, idx - 1].astype(str)), default=0)
                    )
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        col_name = txn_df.columns[cell.column - 1]
                        if '金额' in str(col_name):
                            cell.number_format = '#,##0.00'

            if check_result.period_summaries:
                period_data = []
                for s in check_result.period_summaries:
                    period_data.append({
                        '期间': s.period,
                        '主体': s.subject,
                        '银行': s.bank,
                        '费用类别': s.fee_category,
                        '交易类型': '利息' if s.transaction_type == TRANSACTION_TYPE_INTEREST else '手续费',
                        '交易笔数': s.transaction_count,
                        '总金额(元)': s.total_amount,
                        '平均金额(元)': s.avg_amount,
                        '最大金额(元)': s.max_amount,
                        '最小金额(元)': s.min_amount,
                        '历史均值(元)': s.historical_mean,
                        '预期金额(元)': s.expected_amount,
                        '与预期偏差(%)': s.deviation_from_expected * 100,
                        '与历史偏差(%)': s.deviation_from_historical * 100,
                        '环比变动(%)': s.period_over_period_change * 100,
                        '是否异常': '是' if s.is_anomaly else '否',
                        '异常原因': '; '.join(s.anomaly_reasons),
                    })

                period_df = pd.DataFrame(period_data)
                period_df.to_excel(writer, sheet_name='期间汇总', index=False)

                ws = writer.sheets['期间汇总']
                for idx, col_name in enumerate(period_df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    max_len = max(
                        len(str(col_name)),
                        max((len(str(v)) for v in period_df.iloc[:, idx - 1].astype(str)), default=0)
                    )
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        col_name = period_df.columns[cell.column - 1]
                        if '金额' in str(col_name):
                            cell.number_format = '#,##0.00'
                        elif '偏差' in str(col_name) or '变动' in str(col_name):
                            cell.number_format = '0.00'

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        col_name = period_df.columns[cell.column - 1]
                        if col_name == '是否异常' and cell.value == '是':
                            cell.fill = openpyxl.styles.PatternFill(
                                start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'
                            )

            if check_result.anomaly_summaries:
                anomaly_data = []
                for i, s in enumerate(check_result.anomaly_summaries, 1):
                    anomaly_data.append({
                        '序号': i,
                        '期间': s.period,
                        '主体': s.subject,
                        '银行': s.bank,
                        '费用类别': s.fee_category,
                        '交易类型': '利息' if s.transaction_type == TRANSACTION_TYPE_INTEREST else '手续费',
                        '交易笔数': s.transaction_count,
                        '总金额(元)': s.total_amount,
                        '历史均值(元)': s.historical_mean,
                        '预期金额(元)': s.expected_amount,
                        '与预期偏差(%)': s.deviation_from_expected * 100,
                        '与历史偏差(%)': s.deviation_from_historical * 100,
                        '环比变动(%)': s.period_over_period_change * 100,
                        '异常原因': '; '.join(s.anomaly_reasons),
                    })

                anomaly_df = pd.DataFrame(anomaly_data)
                anomaly_df.to_excel(writer, sheet_name='异常清单', index=False)

                ws = writer.sheets['异常清单']
                for idx, col_name in enumerate(anomaly_df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(idx)
                    max_len = max(
                        len(str(col_name)),
                        max((len(str(v)) for v in anomaly_df.iloc[:, idx - 1].astype(str)), default=0)
                    )
                    ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        col_name = anomaly_df.columns[cell.column - 1]
                        if '金额' in str(col_name):
                            cell.number_format = '#,##0.00'
                        elif '偏差' in str(col_name) or '变动' in str(col_name):
                            cell.number_format = '0.00'
                        elif '序号' in str(col_name) or '笔数' in str(col_name):
                            cell.number_format = '#,##0'

            ws_overview = writer.sheets['核对总览']
            for col_idx in range(1, 3):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws_overview.column_dimensions[col_letter].width = 25

            for row in ws_overview.iter_rows(min_row=2):
                for cell in row:
                    if cell.column == 2:
                        val = cell.value
                        if isinstance(val, (int, float)):
                            if isinstance(val, float):
                                cell.number_format = '#,##0.00'
                            else:
                                cell.number_format = '#,##0'

        logger.info('利息手续费核对结果已导出: %s', output_path)
        return output_path

    except Exception as e:
        logger.error('导出利息手续费核对结果失败: %s', e, exc_info=True)
        raise


def generate_interest_fee_check_from_records(records: List[Dict[str, Any]],
                                             output_dir: Optional[str] = None,
                                             source_info: Optional[Dict[str, Any]] = None,
                                             period_type: str = 'month') -> Optional[str]:
    """
    从交易记录列表直接生成利息手续费核对报告

    Args:
        records: 交易记录列表
        output_dir: 输出目录
        source_info: 数据源信息
        period_type: 汇总期间类型

    Returns:
        str: 生成的文件路径，无数据则返回 None
    """
    logger = get_logger()

    if not records:
        logger.warning('无交易记录，跳过利息手续费核对')
        return None

    if output_dir is None:
        output_dir = get_script_dir()

    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'利息手续费核对报告_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, filename)

    check_result = check_interest_fee(records, period_type=period_type)

    if check_result.total_records == 0:
        logger.warning('未筛选到任何利息手续费交易，跳过报告生成')
        return None

    return export_interest_fee_check_result(check_result, output_path, source_info)


def generate_interest_fee_check_from_total(total_path: str,
                                           output_dir: Optional[str] = None,
                                           period_type: str = 'month') -> Optional[str]:
    """
    从银行流水总表文件生成利息手续费核对报告

    Args:
        total_path: 总表文件路径
        output_dir: 输出目录
        period_type: 汇总期间类型

    Returns:
        str: 生成的文件路径，无数据则返回 None
    """
    logger = get_logger()

    if not total_path or not os.path.exists(total_path):
        logger.error('总表文件不存在: %s', total_path)
        return None

    records = load_total_table(total_path)
    if not records:
        logger.warning('总表文件无数据: %s', total_path)
        return None

    source_info = {
        '数据来源文件': os.path.basename(total_path),
        '总表记录数': len(records),
        '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    return generate_interest_fee_check_from_records(
        records, output_dir, source_info, period_type
    )


def run_interest_fee_check_flow(script_dir):
    """利息手续费核对 CLI 流程"""
    logger = get_logger()
    logger.info('========== 利息手续费核对开始 ==========')

    print('\n' + '=' * 70)
    print('利息与手续费专项核对 - 筛选、汇总、费率对比与异常检测')
    print('=' * 70)
    print('\n请选择数据来源：')
    print('  1) 从银行流水总表文件（Excel）')
    print('  2) 从数据库（按条件查询后核对）')
    print('  0) 返回主菜单')

    choice = input('\n请输入选项（默认 1）: ').strip() or '1'

    records = []
    source_info = {}

    if choice == '0':
        return
    elif choice == '1':
        total_path = ask_file('请选择【银行流水总表】文件')
        if not total_path:
            show_info('提示', '未选择总表文件，返回。')
            return
        logger.info('用户选择总表文件: %s', total_path)
        records = load_total_table(total_path)
        if not records:
            show_warning('错误', '总表文件无数据或读取失败。')
            return
        source_info = {
            '数据来源文件': os.path.basename(total_path),
            '总表记录数': len(records),
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    elif choice == '2':
        if not HAS_DATABASE:
            show_warning('错误', '数据库模块不可用。')
            return

        print('\n输入查询条件（直接回车表示不限制）：')
        subject = input('主体名称: ').strip() or None
        bank = input('银行名称: ').strip() or None
        account = input('银行账号: ').strip() or None
        start_date = input('开始日期 (YYYY-MM-DD): ').strip() or None
        end_date = input('结束日期 (YYYY-MM-DD): ').strip() or None

        try:
            qr = db_module.query_transactions(
                subject=subject, bank=bank, account=account,
                start_date=start_date, end_date=end_date,
                limit=999999, script_dir=script_dir
            )
            records = [r.to_dict() for r in qr.records]
        except Exception as e:
            show_warning('错误', f'数据库查询失败: {e}')
            logger.error('数据库查询失败: %s', e, exc_info=True)
            return

        if not records:
            show_info('提示', '查询结果为空。')
            return

        source_info = {
            '数据来源': '数据库查询',
            '查询主体': subject or '全部',
            '查询银行': bank or '全部',
            '查询账号': account or '全部',
            '日期范围': f'{start_date or "不限"} ~ {end_date or "不限"}',
            '记录数': len(records),
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    else:
        print('无效选项')
        return

    print('\n请选择汇总期间类型：')
    print('  1) 按月汇总（默认）')
    print('  2) 按季度汇总')
    print('  3) 按年汇总')
    print('  4) 按周汇总')
    period_choice = input('\n请输入选项（默认 1）: ').strip() or '1'

    period_type_map = {
        '1': 'month',
        '2': 'quarter',
        '3': 'year',
        '4': 'week',
    }
    period_label_map = {
        '1': '按月',
        '2': '按季度',
        '3': '按年',
        '4': '按周',
    }
    period_type = period_type_map.get(period_choice, 'month')
    period_label = period_label_map.get(period_choice, '按月')

    print(f'\n开始核对，共 {len(records)} 条记录，{period_label}汇总...')
    check_result = check_interest_fee(records, period_type=period_type)
    summary = check_result.check_summary

    print('\n' + '=' * 70)
    print('核对结果总览')
    print('=' * 70)
    print(f'  总记录数:           {summary.get("total_records", 0):,}')
    print(f'  利息类交易:         {summary.get("interest_records", 0):,} 笔')
    print(f'  手续费类交易:       {summary.get("fee_records", 0):,} 笔')
    print(f'  利息总金额:         {summary.get("total_interest_amount", 0):,.2f} 元')
    print(f'  手续费总金额:       {summary.get("total_fee_amount", 0):,.2f} 元')
    print(f'  期间汇总数:         {summary.get("period_count", 0):,}')
    print(f'  异常汇总数:         {summary.get("anomaly_count", 0):,}')

    if check_result.anomaly_count > 0:
        print(f'\n  ⚠️  发现 {check_result.anomaly_count} 个期间存在异常波动')
        for anomaly in check_result.anomaly_summaries[:10]:
            print(f'    - [{anomaly.period}] {anomaly.subject}/{anomaly.bank}/{anomaly.fee_category}: '
                  f'{anomaly.total_amount:,.2f}元, 原因: {"; ".join(anomaly.anomaly_reasons)}')
        if len(check_result.anomaly_summaries) > 10:
            print(f'    ... 还有 {len(check_result.anomaly_summaries) - 10} 条异常，详见导出文件')
    else:
        print(f'\n  ✅ 未检测到异常波动！')

    output_dir = input('\n请输入输出目录（直接回车默认当前目录）: ').strip()
    if not output_dir:
        output_dir = script_dir

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'利息手续费核对报告_{timestamp}.xlsx')

    try:
        export_interest_fee_check_result(check_result, output_path, source_info)
        msg = f'核对报告已导出！\n\n输出文件：{output_path}'
        show_info('导出成功', msg)
        logger.info('利息手续费核对报告导出完成: %s', output_path)
    except Exception as e:
        msg = f'导出失败：{e}'
        show_warning('导出失败', msg)
        logger.error('利息手续费核对报告导出失败: %s', e, exc_info=True)

    logger.info('========== 利息手续费核对结束 ==========')


# ══════════════════════════════════════════════════════════════════════════════
# 电子表格协同编辑模块
# 功能：
#   1. 总表输出后自动生成带数据验证与下拉选项的 Excel 协同编辑模板
#   2. 财务在模板上补充凭证号、备注、会计科目等字段
#   3. 将补充后的 Excel 数据回写合并到总表/数据库
# ══════════════════════════════════════════════════════════════════════════════

COLLAB_TEMPLATE_SUFFIX = '_财务协同编辑版'

COLLAB_EDITABLE_COLUMNS = [
    '凭证号',
    '凭证日期',
    '会计科目编码',
    '会计科目名称',
    '备注',
    '制单人',
]

DEFAULT_SUBJECT_OPTIONS = [
    ('1001', '库存现金'),
    ('1002', '银行存款'),
    ('1012', '其他货币资金'),
    ('1122', '应收账款'),
    ('1123', '预付账款'),
    ('1221', '其他应收款'),
    ('1403', '原材料'),
    ('1405', '库存商品'),
    ('1511', '长期股权投资'),
    ('1601', '固定资产'),
    ('1602', '累计折旧'),
    ('2202', '应付账款'),
    ('2203', '预收账款'),
    ('2211', '应付职工薪酬'),
    ('2221', '应交税费'),
    ('2241', '其他应付款'),
    ('4001', '实收资本'),
    ('4002', '资本公积'),
    ('4101', '盈余公积'),
    ('4103', '本年利润'),
    ('4104', '利润分配'),
    ('5001', '生产成本'),
    ('5101', '制造费用'),
    ('6001', '主营业务收入'),
    ('6051', '其他业务收入'),
    ('6301', '营业外收入'),
    ('6401', '主营业务成本'),
    ('6402', '其他业务成本'),
    ('6403', '税金及附加'),
    ('6601', '销售费用'),
    ('6602', '管理费用'),
    ('6603', '财务费用'),
    ('6711', '营业外支出'),
    ('6801', '所得税费用'),
]

TAG_OPTIONS = ['黑名单', '白名单', '关注', '正常']

READONLY_COLUMN_FILL = openpyxl.styles.PatternFill(
    start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'
)
EDITABLE_COLUMN_FILL = openpyxl.styles.PatternFill(
    start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'
)
HEADER_FILL = openpyxl.styles.PatternFill(
    start_color='4472C4', end_color='4472C4', fill_type='solid'
)
HEADER_FONT = openpyxl.styles.Font(color='FFFFFF', bold=True)
BORDER_THIN = openpyxl.styles.Border(
    left=openpyxl.styles.Side(style='thin'),
    right=openpyxl.styles.Side(style='thin'),
    top=openpyxl.styles.Side(style='thin'),
    bottom=openpyxl.styles.Side(style='thin'),
)


def _collect_unique_values(records: List[Dict[str, Any]], field: str) -> List[str]:
    values = set()
    for rec in records:
        val = rec.get(field)
        if val is not None and str(val).strip():
            values.add(str(val).strip())
    return sorted(values)


def generate_collab_template(
    summary_path: str,
    output_dir: Optional[str] = None,
    lookup_source: Any = None,
    custom_subject_options: Optional[List[Tuple[str, str]]] = None,
) -> Optional[str]:
    """
    基于银行流水总表生成带数据验证和下拉选项的财务协同编辑模板。

    Args:
        summary_path: 原始银行流水总表 Excel 文件路径
        output_dir: 输出目录，默认与总表同目录
        lookup_source: 查找表源，用于获取主体列表
        custom_subject_options: 自定义会计科目选项 [(编码, 名称), ...]

    Returns:
        str: 生成的协同编辑模板文件路径，失败则返回 None
    """
    logger = get_logger()

    if not summary_path or not os.path.exists(summary_path):
        logger.error('总表文件不存在，无法生成协同编辑模板: %s', summary_path)
        return None

    try:
        df = pd.read_excel(summary_path, engine='openpyxl')
    except Exception as e:
        logger.error('读取总表失败: %s', e, exc_info=True)
        return None

    if df.empty:
        logger.warning('总表为空，跳过生成协同编辑模板')
        return None

    if output_dir is None:
        output_dir = os.path.dirname(summary_path) or get_script_dir()
    os.makedirs(output_dir, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(summary_path))[0]
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_filename = f'{base_name}{COLLAB_TEMPLATE_SUFFIX}_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, output_filename)

    records = df.to_dict('records')

    for col in COLLAB_EDITABLE_COLUMNS:
        if col not in df.columns:
            df[col] = ''

    base_columns = list(df.columns)

    try:
        wb = openpyxl.load_workbook(summary_path)
    except Exception as e:
        logger.error('openpyxl 加载总表失败: %s', e, exc_info=True)
        df.to_excel(output_path, index=False, engine='openpyxl')
        wb = openpyxl.load_workbook(output_path)

    ws = wb.active
    ws.title = '流水总表(协同编辑)'

    header_row_idx = 1
    col_count = ws.max_column

    editable_col_indices: Dict[str, int] = {}
    all_col_indices: Dict[str, int] = {}
    for col_idx in range(1, col_count + 1):
        header_val = ws.cell(row=header_row_idx, column=col_idx).value
        header_name = str(header_val).strip() if header_val is not None else ''
        all_col_indices[header_name] = col_idx
        if header_name in COLLAB_EDITABLE_COLUMNS:
            editable_col_indices[header_name] = col_idx

    for col_name in COLLAB_EDITABLE_COLUMNS:
        if col_name not in all_col_indices:
            new_col_idx = col_count + 1
            ws.cell(row=header_row_idx, column=new_col_idx, value=col_name)
            editable_col_indices[col_name] = new_col_idx
            all_col_indices[col_name] = new_col_idx
            col_count = new_col_idx

    for col_idx in range(1, col_count + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        cell.border = BORDER_THIN

    max_row = ws.max_row

    editable_col_set = set(editable_col_indices.values())
    for row_idx in range(2, max_row + 1):
        for col_idx in range(1, col_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = BORDER_THIN
            if col_idx in editable_col_set:
                cell.fill = EDITABLE_COLUMN_FILL
                cell.protection = openpyxl.styles.Protection(locked=False)
            else:
                cell.fill = READONLY_COLUMN_FILL
                cell.protection = openpyxl.styles.Protection(locked=True)

    bank_options = _collect_unique_values(records, '银行')
    if '银行' in all_col_indices and bank_options:
        dv_bank = openpyxl.worksheet.datavalidation.DataValidation(
            type='list',
            formula1=f'"{",".join(bank_options)}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle='银行名称无效',
            error='请从下拉列表中选择银行名称',
            showInputMessage=True,
            promptTitle='选择银行',
            prompt='从银行列表中选择',
        )
        ws.add_data_validation(dv_bank)
        dv_bank.add(f'{openpyxl.utils.get_column_letter(all_col_indices["银行"])}2:'
                     f'{openpyxl.utils.get_column_letter(all_col_indices["银行"])}{max_row}')

    subject_options_from_lookup = []
    if lookup_source is not None:
        try:
            lookup = _resolve_lookup(lookup_source) if not isinstance(lookup_source, dict) else lookup_source
            if 'all_entries' in lookup:
                for entry in lookup['all_entries']:
                    subj = entry.get('subject', '')
                    if subj and subj not in subject_options_from_lookup:
                        subject_options_from_lookup.append(subj)
        except Exception:
            pass

    subject_value_options = _collect_unique_values(records, '主体')
    for s in subject_options_from_lookup:
        if s not in subject_value_options:
            subject_value_options.append(s)
    if '主体' in all_col_indices and subject_value_options:
        dv_subject = openpyxl.worksheet.datavalidation.DataValidation(
            type='list',
            formula1=f'"{",".join(subject_value_options)}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle='主体名称无效',
            error='请从下拉列表中选择主体名称',
            showInputMessage=True,
            promptTitle='选择主体',
            prompt='从主体列表中选择',
        )
        ws.add_data_validation(dv_subject)
        dv_subject.add(f'{openpyxl.utils.get_column_letter(all_col_indices["主体"])}2:'
                        f'{openpyxl.utils.get_column_letter(all_col_indices["主体"])}{max_row}')

    if '凭证日期' in editable_col_indices:
        col_letter = openpyxl.utils.get_column_letter(editable_col_indices['凭证日期'])
        dv_date = openpyxl.worksheet.datavalidation.DataValidation(
            type='date',
            operator='between',
            formula1='1900-01-01',
            formula2='2100-12-31',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle='日期格式错误',
            error='请输入有效的日期，格式如: 2024-01-15',
            showInputMessage=True,
            promptTitle='录入凭证日期',
            prompt='请输入日期，格式如 2024-01-15',
        )
        ws.add_data_validation(dv_date)
        dv_date.add(f'{col_letter}2:{col_letter}{max_row}')
        for row_idx in range(2, max_row + 1):
            ws.cell(row=row_idx, column=editable_col_indices['凭证日期']).number_format = 'yyyy-mm-dd'

    subject_opts = custom_subject_options if custom_subject_options is not None else DEFAULT_SUBJECT_OPTIONS
    subject_code_opts = [code for code, _ in subject_opts]
    subject_name_opts = [name for _, name in subject_opts]

    if '会计科目编码' in editable_col_indices and subject_code_opts:
        col_letter = openpyxl.utils.get_column_letter(editable_col_indices['会计科目编码'])
        dv_code = openpyxl.worksheet.datavalidation.DataValidation(
            type='list',
            formula1=f'"{",".join(subject_code_opts)}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle='科目编码无效',
            error='请从下拉列表中选择标准会计科目编码',
            showInputMessage=True,
            promptTitle='选择会计科目编码',
            prompt='请选择标准会计科目编码，或自行输入',
        )
        ws.add_data_validation(dv_code)
        dv_code.add(f'{col_letter}2:{col_letter}{max_row}')

    if '会计科目名称' in editable_col_indices and subject_name_opts:
        col_letter = openpyxl.utils.get_column_letter(editable_col_indices['会计科目名称'])
        dv_name = openpyxl.worksheet.datavalidation.DataValidation(
            type='list',
            formula1=f'"{",".join(subject_name_opts)}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle='科目名称无效',
            error='请从下拉列表中选择标准会计科目名称',
            showInputMessage=True,
            promptTitle='选择会计科目名称',
            prompt='请选择标准会计科目名称，或自行输入',
        )
        ws.add_data_validation(dv_name)
        dv_name.add(f'{col_letter}2:{col_letter}{max_row}')

    if '黑白名单标签' in all_col_indices:
        col_letter = openpyxl.utils.get_column_letter(all_col_indices['黑白名单标签'])
        dv_tag = openpyxl.worksheet.datavalidation.DataValidation(
            type='list',
            formula1=f'"{",".join(TAG_OPTIONS)}"',
            allow_blank=True,
            showErrorMessage=False,
            showInputMessage=True,
            promptTitle='标签标记',
            prompt='可选择标签: 黑名单/白名单/关注/正常',
        )
        ws.add_data_validation(dv_tag)
        dv_tag.add(f'{col_letter}2:{col_letter}{max_row}')

    if '付款' in all_col_indices:
        col_letter = openpyxl.utils.get_column_letter(all_col_indices['付款'])
        dv_amt_pay = openpyxl.worksheet.datavalidation.DataValidation(
            type='decimal',
            operator='lessThanOrEqual',
            formula1='0',
            allow_blank=True,
            showInputMessage=True,
            promptTitle='付款金额',
            prompt='付款金额应为负数或0',
        )
        ws.add_data_validation(dv_amt_pay)
        dv_amt_pay.add(f'{col_letter}2:{col_letter}{max_row}')

    if '收款' in all_col_indices:
        col_letter = openpyxl.utils.get_column_letter(all_col_indices['收款'])
        dv_amt_recv = openpyxl.worksheet.datavalidation.DataValidation(
            type='decimal',
            operator='greaterThanOrEqual',
            formula1='0',
            allow_blank=True,
            showInputMessage=True,
            promptTitle='收款金额',
            prompt='收款金额应为正数或0',
        )
        ws.add_data_validation(dv_amt_recv)
        dv_amt_recv.add(f'{col_letter}2:{col_letter}{max_row}')

    sheet_opts = wb.create_sheet('下拉选项源')
    for i, (code, name) in enumerate(subject_opts, start=1):
        sheet_opts.cell(row=i, column=1, value=code)
        sheet_opts.cell(row=i, column=2, value=name)
    sheet_opts.cell(row=1, column=1, value='科目编码')
    sheet_opts.cell(row=1, column=2, value='科目名称')
    sheet_opts.cell(row=1, column=1).fill = HEADER_FILL
    sheet_opts.cell(row=1, column=1).font = HEADER_FONT
    sheet_opts.cell(row=1, column=2).fill = HEADER_FILL
    sheet_opts.cell(row=1, column=2).font = HEADER_FONT
    sheet_opts.sheet_state = 'hidden'

    ws.freeze_panes = 'A2'

    for col_idx in range(1, col_count + 1):
        max_len = 12
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        for row_idx in range(1, min(max_row + 1, 101)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None:
                val_len = len(str(val))
                if val_len > max_len:
                    max_len = min(val_len, 40)
        ws.column_dimensions[col_letter].width = max_len + 2

    ws.protection.sheet = True

    readme_ws = wb.create_sheet('使用说明', 0)
    readme_content = [
        ('A1', '银行流水总表 - 财务协同编辑模板'),
        ('A2', ''),
        ('A3', '使用说明'),
        ('A4', '1. 黄色背景列：可编辑（凭证号、凭证日期、会计科目、备注等）'),
        ('A5', '2. 灰色背景列：只读（原始流水数据，请勿修改）'),
        ('A6', '3. 下拉箭头：点击单元格右侧箭头从标准选项中选择'),
        ('A7', '4. 编辑完成后保存文件，然后在主程序中运行"回写合并"功能'),
        ('A8', ''),
        ('A9', '可编辑字段说明'),
        ('A10', '凭证号:     财务记账凭证编号，如 记-202401-0001'),
        ('A11', '凭证日期:   记账日期，格式 YYYY-MM-DD'),
        ('A12', '会计科目编码: 标准会计科目编码，如 1002（银行存款）'),
        ('A13', '会计科目名称: 标准会计科目名称，如 银行存款'),
        ('A14', '备注:       其他需补充的说明信息'),
        ('A15', '制单人:     录入人员姓名或编号'),
        ('A16', ''),
        ('A17', '数据验证规则'),
        ('A18', '银行/主体:  标准下拉选项，限定可选范围'),
        ('A19', '凭证日期:  合法日期校验'),
        ('A20', '付款金额:  小于等于0（负数表示支出）'),
        ('A21', '收款金额:  大于等于0（正数表示收入）'),
        ('A22', '会计科目:  标准科目下拉选项'),
        ('A23', ''),
        ('A24', '警告: 请勿修改唯一id、交易流水号等只读列，否则将无法回写合并！'),
    ]
    readme_ws['A1'].font = openpyxl.styles.Font(size=16, bold=True, color='1F4E79')
    readme_ws['A3'].font = openpyxl.styles.Font(size=13, bold=True)
    readme_ws['A9'].font = openpyxl.styles.Font(size=13, bold=True)
    readme_ws['A17'].font = openpyxl.styles.Font(size=13, bold=True)
    readme_ws['A24'].font = openpyxl.styles.Font(color='FF0000', bold=True)
    for cell_ref, content in readme_content:
        readme_ws[cell_ref] = content
    readme_ws.column_dimensions['A'].width = 80

    try:
        wb.save(output_path)
        wb.close()
    except Exception as e:
        logger.error('保存协同编辑模板失败: %s', e, exc_info=True)
        return None

    logger.info('财务协同编辑模板生成完成: %s（可编辑列: %s）',
                output_path, ', '.join(editable_col_indices.keys()))
    return output_path


def read_collab_edits(
    edited_path: str,
    match_by: str = '唯一id',
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    读取财务在协同编辑模板中补充的数据，返回待回写的修改字典列表。

    Args:
        edited_path: 已编辑的 Excel 文件路径
        match_by: 用于匹配记录的键列名（默认 '唯一id'）

    Returns:
        tuple: (edits_list, summary)
            - edits_list: 每个元素为 {match_key: value, 修改字段1: 值1, 修改字段2: 值2, ...}
            - summary: 统计信息字典
    """
    logger = get_logger()

    if not edited_path or not os.path.exists(edited_path):
        logger.error('协同编辑文件不存在: %s', edited_path)
        return [], {'error': '文件不存在'}

    try:
        wb = openpyxl.load_workbook(edited_path, data_only=True)
        ws = None
        for sheet_name in wb.sheetnames:
            if '协同' in sheet_name or '流水' in sheet_name or sheet_name == 'Sheet':
                ws = wb[sheet_name]
                break
        if ws is None:
            ws = wb.active
    except Exception as e:
        logger.error('读取协同编辑文件失败: %s', e, exc_info=True)
        return [], {'error': f'读取失败: {e}'}

    header_map: Dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col_idx).value
        if val is not None:
            header_map[str(val).strip()] = col_idx

    if match_by not in header_map:
        logger.error('协同编辑文件缺少匹配键列「%s」，无法回写', match_by)
        wb.close()
        return [], {'error': f'缺少匹配键列 {match_by}'}

    match_col_idx = header_map[match_by]

    editable_header_indices = {
        name: idx for name, idx in header_map.items()
        if name in COLLAB_EDITABLE_COLUMNS
    }

    edits: List[Dict[str, Any]] = []
    total_rows = 0
    edited_rows = 0
    unchanged_rows = 0
    empty_key_rows = 0

    for row_idx in range(2, ws.max_row + 1):
        total_rows += 1
        key_val = ws.cell(row=row_idx, column=match_col_idx).value
        if key_val is None or str(key_val).strip() == '':
            empty_key_rows += 1
            continue

        row_edit = {match_by: str(key_val).strip()}
        has_edit = False

        for field_name, col_idx in editable_header_indices.items():
            cell_val = ws.cell(row=row_idx, column=col_idx).value
            if cell_val is not None and str(cell_val).strip() != '':
                row_edit[field_name] = cell_val
                has_edit = True
            else:
                row_edit[field_name] = None

        if has_edit:
            edits.append(row_edit)
            edited_rows += 1
        else:
            unchanged_rows += 1

    wb.close()

    summary = {
        'total_rows': total_rows,
        'edited_rows': edited_rows,
        'unchanged_rows': unchanged_rows,
        'empty_key_rows': empty_key_rows,
        'edited_fields': list(editable_header_indices.keys()),
        'match_by': match_by,
    }
    logger.info('读取协同编辑完成: %d 行数据，其中 %d 行有修改，%d 行无修改，%d 行无匹配键',
                total_rows, edited_rows, unchanged_rows, empty_key_rows)
    return edits, summary


def apply_collab_edits_to_records(
    records: List[Dict[str, Any]],
    edits: List[Dict[str, Any]],
    match_by: str = '唯一id',
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    将协同编辑的修改应用到内存中的记录列表。

    Args:
        records: 原始记录列表（来自总表或数据库）
        edits: read_collab_edits() 返回的修改列表
        match_by: 匹配键列名

    Returns:
        tuple: (updated_records, stats)
    """
    logger = get_logger()

    if not records:
        logger.warning('无原始记录，跳过回写应用')
        return records, {'matched': 0, 'unmatched': len(edits)}

    key_to_records: Dict[str, List[int]] = {}
    for i, rec in enumerate(records):
        key = rec.get(match_by)
        key_str = str(key).strip() if key is not None else ''
        if key_str:
            if key_str not in key_to_records:
                key_to_records[key_str] = []
            key_to_records[key_str].append(i)

    matched = 0
    unmatched = 0
    fields_updated = {f: 0 for f in COLLAB_EDITABLE_COLUMNS}

    for edit in edits:
        edit_key = str(edit.get(match_by, '')).strip()
        if not edit_key:
            unmatched += 1
            continue

        rec_indices = key_to_records.get(edit_key)
        if not rec_indices:
            unmatched += 1
            logger.debug('未找到匹配记录，匹配键: %s', edit_key)
            continue

        for rec_idx in rec_indices:
            rec = records[rec_idx]
            for field in COLLAB_EDITABLE_COLUMNS:
                new_val = edit.get(field)
                if new_val is not None:
                    rec[field] = new_val
                    fields_updated[field] = fields_updated.get(field, 0) + 1
            matched += 1

    stats = {
        'matched_records': matched,
        'unmatched_edits': unmatched,
        'fields_updated_count': fields_updated,
    }
    logger.info('协同编辑回写应用完成: 匹配 %d 条记录，未匹配 %d 条修改',
                matched, unmatched)
    return records, stats


def merge_collab_edits_to_summary(
    edited_path: str,
    summary_path: Optional[str] = None,
    output_path: Optional[str] = None,
    script_dir: Optional[str] = None,
    lookup_source: Any = None,
) -> Dict[str, Any]:
    """
    完整流程：读取财务协同编辑后的 Excel → 合并回原始总表 → 输出新总表。

    Args:
        edited_path: 财务已编辑的 Excel 文件路径
        summary_path: 原始总表路径（可选，若 None 则从已编辑文件同目录查找）
        output_path: 合并后输出路径（可选，自动生成时间戳文件）
        script_dir: 脚本目录
        lookup_source: 查找表源

    Returns:
        dict: 包含 success, output_path, stats, error 等信息
    """
    logger = get_logger()

    if summary_path is None:
        if script_dir:
            candidate = get_summary_table_path(script_dir)
            if os.path.exists(candidate):
                summary_path = candidate

    if summary_path is None or not os.path.exists(summary_path):
        edited_dir = os.path.dirname(edited_path)
        candidate = os.path.join(edited_dir, SUMMARY_TABLE_FILENAME)
        if os.path.exists(candidate):
            summary_path = candidate
        else:
            for f in os.listdir(edited_dir):
                if f.startswith('银行流水总表') and f.endswith('.xlsx') \
                        and COLLAB_TEMPLATE_SUFFIX not in f:
                    summary_path = os.path.join(edited_dir, f)
                    break

    if summary_path is None or not os.path.exists(summary_path):
        msg = '找不到原始总表文件用于合并，请指定 summary_path'
        logger.error(msg)
        return {'success': False, 'error': msg, 'output_path': None}

    edits, read_stats = read_collab_edits(edited_path)
    if read_stats.get('error'):
        return {'success': False, 'error': read_stats['error'], 'output_path': None}

    if not edits:
        logger.warning('协同编辑文件中未读取到任何修改数据')
        return {
            'success': True,
            'warning': '协同编辑文件中无修改内容，未执行合并',
            'output_path': summary_path,
            'stats': read_stats,
            'applied_stats': None,
        }

    try:
        df = pd.read_excel(summary_path, engine='openpyxl')
    except Exception as e:
        msg = f'读取原始总表失败: {e}'
        logger.error(msg, exc_info=True)
        return {'success': False, 'error': msg, 'output_path': None}

    records = df.to_dict('records')

    updated_records, apply_stats = apply_collab_edits_to_records(records, edits)

    columns = get_summary_columns(updated_records, lookup_source)
    for col in COLLAB_EDITABLE_COLUMNS:
        if col not in columns:
            columns.append(col)

    if output_path is None:
        base_dir = os.path.dirname(summary_path) or (script_dir or get_script_dir())
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = os.path.join(base_dir, f'银行流水总表_合并财务编辑_{timestamp}.xlsx')

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '流水总表'
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            cell.border = BORDER_THIN

        for row_idx, rec in enumerate(updated_records, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                val = rec.get(col_name)
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = BORDER_THIN
                if col_name in COLLAB_EDITABLE_COLUMNS:
                    cell.fill = EDITABLE_COLUMN_FILL

        ws.freeze_panes = 'A2'
        for col_idx in range(1, len(columns) + 1):
            max_len = 12
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            for r in range(1, min(len(updated_records) + 2, 101)):
                v = ws.cell(row=r, column=col_idx).value
                if v is not None:
                    max_len = max(max_len, min(len(str(v)), 40))
            ws.column_dimensions[col_letter].width = max_len + 2

        wb.save(output_path)
        wb.close()
    except Exception as e:
        msg = f'保存合并后总表失败: {e}'
        logger.error(msg, exc_info=True)
        return {'success': False, 'error': msg, 'output_path': None}

    db_stats = None
    if HAS_DATABASE and updated_records:
        try:
            batch_id = f"COLLAB{datetime.now().strftime('%Y%m%d%H%M%S')}"
            db_inserted, db_duplicates = db_module.persist_transactions(
                updated_records,
                batch_id=batch_id,
                deduplicate=True,
                script_dir=script_dir,
            )
            db_stats = {
                'batch_id': batch_id,
                'inserted': db_inserted,
                'duplicates': db_duplicates,
            }
            logger.info('协同编辑结果已同步到数据库: 批次 %s, 插入 %d, 跳过 %d',
                        batch_id, db_inserted, db_duplicates)
        except Exception as e:
            logger.warning('协同编辑结果同步数据库失败: %s', e, exc_info=True)

    logger.info('协同编辑合并完成: 输出文件 %s，匹配 %d 条，未匹配 %d 条',
                output_path,
                apply_stats.get('matched_records', 0),
                apply_stats.get('unmatched_edits', 0))

    return {
        'success': True,
        'output_path': output_path,
        'read_stats': read_stats,
        'applied_stats': apply_stats,
        'db_stats': db_stats,
    }


# ──────────────────────────────────────────────
# 银企直连/网银导出目录对接集成
# ──────────────────────────────────────────────

try:
    from bank_directory_connector import (
        BankDirectoryConnector,
        ProcessingResult as DirectoryProcessingResult,
    )
    HAS_DIRECTORY_CONNECTOR = True
except ImportError as e:
    HAS_DIRECTORY_CONNECTOR = False
    logger = get_logger()
    logger.warning('目录对接模块不可用: %s', e)


@dataclass
class DirectoryPipelineResult:
    """目录流水线处理结果"""
    success: bool
    message: str = ''
    processed_files: List[str] = field(default_factory=list)
    error_files: List[Tuple[str, str]] = field(default_factory=list)
    output_path: Optional[str] = None
    archive_dir: Optional[str] = None
    pipeline_result: Optional[ProcessingResult] = None
    start_time: datetime = field(default_factory=datetime.now)
    end_time: Optional[datetime] = None


def run_directory_pipeline(
    script_dir: Optional[str] = None,
    config_path: Optional[str] = None,
    incremental: Optional[bool] = None,
    keep_strategy: Optional[str] = None,
) -> DirectoryPipelineResult:
    """
    运行目录对接流水线：自动处理 inbox 目录中的银行流水文件

    这是一个便捷函数，封装了 BankDirectoryConnector 的调用，
    与现有的 run_pipeline 函数保持一致的调用风格。

    处理流程：
    1. 扫描 inbox 目录中的银行流水文件
    2. 验证文件稳定性和锁定状态
    3. 移动到 processing 目录并调用 run_pipeline 处理
    4. 处理成功则归档到 outbox，失败则移动到 error

    Args:
        script_dir: 脚本目录，用于查找主体查找表等资源
        config_path: 目录对接配置文件路径
        incremental: 是否增量模式，None 则使用配置文件中的设置
        keep_strategy: 文件保留策略，None 则使用配置文件中的设置

    Returns:
        DirectoryPipelineResult 处理结果

    Example:
        >>> result = run_directory_pipeline('./backend')
        >>> print(result.message)
        >>> if result.success:
        ...     print(f'归档目录: {result.archive_dir}')
    """
    logger = get_logger()
    result = DirectoryPipelineResult(success=False)

    if not HAS_DIRECTORY_CONNECTOR:
        result.message = '目录对接模块不可用，请确保 bank_directory_connector.py 存在'
        logger.error(result.message)
        return result

    if script_dir is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    try:
        logger.info('========== 开始目录对接流水线 ==========')

        connector = BankDirectoryConnector(
            config_path=config_path,
            script_dir=script_dir,
        )

        if incremental is not None:
            connector._processing_config.incremental = incremental

        if keep_strategy is not None:
            connector._processing_config.keep_strategy = keep_strategy

        dir_result = connector.run_once()

        result.success = dir_result.success
        result.message = dir_result.message
        result.processed_files = dir_result.processed_files
        result.error_files = dir_result.error_files
        result.output_path = dir_result.output_path
        result.archive_dir = dir_result.archive_dir
        result.end_time = dir_result.end_time

        logger.info('目录对接流水线完成: %s', result.message)
        logger.info('========== 目录对接流水线结束 ==========')

        return result

    except Exception as e:
        logger.exception('目录对接流水线发生异常')
        result.message = f'处理异常: {e}'
        result.success = False
        result.end_time = datetime.now()
        return result


def run_directory_watch(
    script_dir: Optional[str] = None,
    config_path: Optional[str] = None,
    stop_on_empty: bool = False,
) -> None:
    """
    启动目录监控模式，持续监控 inbox 目录，自动处理新文件

    Args:
        script_dir: 脚本目录
        config_path: 配置文件路径
        stop_on_empty: 连续无文件时是否自动退出
    """
    logger = get_logger()

    if not HAS_DIRECTORY_CONNECTOR:
        logger.error('目录对接模块不可用')
        return

    if script_dir is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    try:
        connector = BankDirectoryConnector(
            config_path=config_path,
            script_dir=script_dir,
        )
        connector.watch(stop_on_first_empty=stop_on_empty)
    except KeyboardInterrupt:
        logger.info('目录监控已停止')
    except Exception as e:
        logger.exception('目录监控发生异常')


def get_directory_status(
    script_dir: Optional[str] = None,
    config_path: Optional[str] = None,
) -> Optional[Dict[str, Any]]:
    """
    获取目录对接的当前状态

    Args:
        script_dir: 脚本目录
        config_path: 配置文件路径

    Returns:
        状态字典，包含各目录的文件数、待处理文件列表等
    """
    if not HAS_DIRECTORY_CONNECTOR:
        return None

    if script_dir is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    try:
        connector = BankDirectoryConnector(
            config_path=config_path,
            script_dir=script_dir,
        )
        return connector.get_status()
    except Exception as e:
        logger = get_logger()
        logger.error('获取目录状态失败: %s', e)
        return None


def trigger_bank_download(
    bank_name: str,
    script_dir: Optional[str] = None,
    config_path: Optional[str] = None,
) -> Tuple[bool, str]:
    """
    触发指定银行的下载脚本

    Args:
        bank_name: 银行名称
        script_dir: 脚本目录
        config_path: 配置文件路径

    Returns:
        (success: bool, message: str)
    """
    if not HAS_DIRECTORY_CONNECTOR:
        return False, '目录对接模块不可用'

    if script_dir is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))

    try:
        connector = BankDirectoryConnector(
            config_path=config_path,
            script_dir=script_dir,
        )
        return connector.trigger_download(bank_name)
    except Exception as e:
        logger = get_logger()
        logger.error('触发银行下载失败: %s', e)
        return False, f'触发失败: {e}'


# ──────────────────────────────────────────────
# 期末余额与银行对账单核对模块
# ──────────────────────────────────────────────

BALANCE_RECONCILIATION_FILENAME = '期末余额与银行对账单核对报告.xlsx'

try:
    import pdfplumber
    HAS_PDFPLUMBER = True
except ImportError:
    HAS_PDFPLUMBER = False

try:
    from PIL import Image
    import pytesseract
    HAS_PYTESSERACT = True
except ImportError:
    HAS_PYTESSERACT = False


@dataclass
class BankStatementBalance:
    """银行对账单余额记录（从官方文件提取）"""
    bank_account: str
    bank_name: str = ''
    subject: str = ''
    statement_balance: float = 0.0
    statement_date: Optional[str] = None
    source_type: str = 'manual'
    source_file: str = ''
    extracted_at: str = ''
    remark: str = ''


@dataclass
class AccountEndBalance:
    """总表期末余额记录"""
    bank_account: str
    bank_name: str = ''
    subject: str = ''
    end_balance: float = 0.0
    end_date: Optional[str] = None
    transaction_count: int = 0
    last_transaction_id: str = ''


@dataclass
class BalanceReconciliationRecord:
    """单条余额比对记录"""
    bank_account: str
    bank_name: str = ''
    subject: str = ''
    total_balance: float = 0.0
    statement_balance: float = 0.0
    diff_amount: float = 0.0
    status: str = 'pending'
    diff_note: str = ''
    total_end_date: Optional[str] = None
    statement_date: Optional[str] = None
    source_file: str = ''


@dataclass
class BalanceReconciliationResult:
    """余额比对结果汇总"""
    total_accounts: int = 0
    matched_accounts: int = 0
    diff_accounts: int = 0
    missing_statement: int = 0
    missing_total: int = 0
    total_diff_amount: float = 0.0
    records: List[BalanceReconciliationRecord] = field(default_factory=list)
    check_summary: Dict[str, Any] = field(default_factory=dict)

    @property
    def match_rate(self) -> str:
        if self.total_accounts > 0:
            return f'{(self.matched_accounts / self.total_accounts * 100):.1f}%'
        return '0%'


def _parse_amount_from_text(text: str) -> Optional[float]:
    """
    从文本中解析金额，支持千分位、正负号、货币符号等格式。

    Args:
        text: 包含金额的文本

    Returns:
        float: 解析后的金额，解析失败返回 None
    """
    if not text:
        return None

    text = text.strip()

    patterns = [
        r'[-+]?\s*[¥￥$€£]?\s*[\d,，]+(?:\.\d{1,2})?',
        r'[-+]?\s*\d+(?:\.\d{1,2})?',
    ]

    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            amount_str = match.group(0)
            amount_str = amount_str.replace(',', '').replace('，', '')
            amount_str = re.sub(r'[¥￥$€£\s]', '', amount_str)
            try:
                return float(amount_str)
            except (ValueError, TypeError):
                continue

    return None


def _parse_account_from_text(text: str) -> Optional[str]:
    """
    从文本中解析银行账号，支持带分隔符的账号格式。

    Args:
        text: 包含账号的文本

    Returns:
        str: 解析后的账号，解析失败返回 None
    """
    if not text:
        return None

    patterns = [
        r'账号[：:\s]\s*([\d\- ]+)',
        r'账户[：:\s]\s*([\d\- ]+)',
        r'卡号[：:\s]\s*([\d\- ]+)',
        r'Account\s*[No\.]+\s*([\d\- ]+)',
        r'A/C\s*[No\.]+\s*([\d\- ]+)',
        r'\b(\d{10,25})\b',
    ]

    for pattern in patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            account = match.group(1)
            account = re.sub(r'[\- ]', '', account)
            if len(account) >= 10:
                return account

    return None


def _parse_date_from_text(text: str) -> Optional[str]:
    """
    从文本中解析日期，支持多种格式。

    Args:
        text: 包含日期的文本

    Returns:
        str: YYYY-MM-DD 格式的日期，解析失败返回 None
    """
    if not text:
        return None

    date_patterns = [
        (r'(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})', lambda m: f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"),
        (r'(\d{4})(\d{2})(\d{2})', lambda m: f"{m.group(1)}-{m.group(2)}-{m.group(3)}"),
        (r'(\d{2})[-/](\d{2})[-/](\d{4})', lambda m: f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"),
        (r'(\d{1,2})[-/月](\d{1,2})[-/日](\d{4})', lambda m: f"{m.group(3)}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"),
    ]

    for pattern, formatter in date_patterns:
        match = re.search(pattern, text)
        if match:
            try:
                return formatter(match)
            except (ValueError, IndexError):
                continue

    return None


def extract_end_balances_from_total(records: List[Dict[str, Any]]) -> Dict[str, AccountEndBalance]:
    """
    从交易记录中提取各账号的期末余额。

    按账号分组后，取每组中交易日期最新的记录的余额作为期末余额。
    如果日期相同，则取交易流水号排序后的最后一条。

    Args:
        records: 交易记录列表

    Returns:
        Dict[str, AccountEndBalance]: 账号到期末余额的映射字典
    """
    logger = get_logger()

    if not records:
        logger.warning('无交易记录，无法提取期末余额')
        return {}

    account_groups: Dict[str, List[Dict[str, Any]]] = {}
    for record in records:
        account = str(record.get('银行账号', '')).strip()
        if not account:
            continue
        if account not in account_groups:
            account_groups[account] = []
        account_groups[account].append(record)

    end_balances: Dict[str, AccountEndBalance] = {}

    for account, group_records in account_groups.items():
        def _sort_key(r):
            date_str = str(r.get('交易日期', '') or '')
            txn_id = str(r.get('交易流水号', '') or '')
            return (date_str, txn_id)

        sorted_records = sorted(group_records, key=_sort_key)
        last_record = sorted_records[-1]

        balance = to_float(last_record.get('余额'))
        if balance is None:
            logger.warning('账号 %s 最后一条记录无有效余额，跳过', account)
            continue

        end_balance = AccountEndBalance(
            bank_account=account,
            bank_name=str(last_record.get('银行', '')),
            subject=str(last_record.get('主体', '')),
            end_balance=balance,
            end_date=str(last_record.get('交易日期', '')) if last_record.get('交易日期') else None,
            transaction_count=len(group_records),
            last_transaction_id=str(last_record.get('交易流水号', '')),
        )
        end_balances[account] = end_balance
        logger.info('账号 %s 期末余额: %s (日期: %s, 笔数: %d)',
                    account, f'{balance:.2f}', end_balance.end_date, end_balance.transaction_count)

    logger.info('共提取 %d 个账号的期末余额', len(end_balances))
    return end_balances


def parse_bank_statement_pdf(pdf_path: str) -> Optional[BankStatementBalance]:
    """
    从PDF银行对账单中提取账号和期末余额。

    Args:
        pdf_path: PDF文件路径

    Returns:
        Optional[BankStatementBalance]: 解析成功返回余额记录，失败返回 None
    """
    logger = get_logger()

    if not HAS_PDFPLUMBER:
        logger.warning('未安装 pdfplumber 库，无法解析PDF文件。请运行: pip install pdfplumber')
        return None

    if not os.path.exists(pdf_path):
        logger.error('PDF文件不存在: %s', pdf_path)
        return None

    try:
        full_text = ''
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ''
                full_text += text + '\n'

        if not full_text.strip():
            logger.warning('PDF文件 %s 未提取到文本内容', pdf_path)
            return None

        account = _parse_account_from_text(full_text)
        if not account:
            logger.warning('未从PDF %s 中解析到银行账号', pdf_path)
            return None

        balance_patterns = [
            r'期末余额[：:]\s*([^\n]+)',
            r'余额[：:]\s*([^\n]+)',
            r'可用余额[：:]\s*([^\n]+)',
            r'账户余额[：:]\s*([^\n]+)',
            r'Closing\s*Balance[：:]\s*([^\n]+)',
            r'Ending\s*Balance[：:]\s*([^\n]+)',
            r'Balance[：:]\s*([^\n]+)',
        ]

        statement_balance = None
        for pattern in balance_patterns:
            match = re.search(pattern, full_text, re.IGNORECASE)
            if match:
                statement_balance = _parse_amount_from_text(match.group(1))
                if statement_balance is not None:
                    break

        if statement_balance is None:
            lines = full_text.split('\n')
            for line in reversed(lines[-50:]):
                if any(keyword in line for keyword in ['余额', 'Balance', 'balance']):
                    statement_balance = _parse_amount_from_text(line)
                    if statement_balance is not None:
                        break

        if statement_balance is None:
            logger.warning('未从PDF %s 中解析到期末余额', pdf_path)
            return None

        statement_date = _parse_date_from_text(full_text)

        bank_name = ''
        bank_patterns = [
            r'^(.+?)银行',
            r'(.+?)银行股份有限公司',
            r'^(.+?)(?:BANK|Bank)',
        ]
        for pattern in bank_patterns:
            match = re.search(pattern, full_text)
            if match:
                bank_name = match.group(1).strip()
                if '银行' not in bank_name:
                    bank_name += '银行'
                break

        record = BankStatementBalance(
            bank_account=account,
            bank_name=bank_name,
            statement_balance=statement_balance,
            statement_date=statement_date,
            source_type='pdf',
            source_file=os.path.basename(pdf_path),
            extracted_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        )

        logger.info('PDF解析成功: 账号 %s, 余额 %.2f, 日期 %s',
                    account, statement_balance, statement_date)
        return record

    except Exception as e:
        logger.error('解析PDF文件失败 %s: %s', pdf_path, e, exc_info=True)
        return None


def parse_bank_statement_image(image_path: str) -> Optional[BankStatementBalance]:
    """
    从银行余额截图中通过OCR提取账号和期末余额。

    Args:
        image_path: 图片文件路径

    Returns:
        Optional[BankStatementBalance]: 解析成功返回余额记录，失败返回 None
    """
    logger = get_logger()

    if not HAS_PYTESSERACT:
        logger.warning('未安装 pytesseract 和 PIL 库，无法进行OCR识别。'
                      '请运行: pip install pytesseract pillow')
        return None

    if not os.path.exists(image_path):
        logger.error('图片文件不存在: %s', image_path)
        return None

    try:
        image = Image.open(image_path)

        if image.mode != 'RGB':
            image = image.convert('RGB')

        ocr_text = pytesseract.image_to_string(image, lang='chi_sim+eng')

        if not ocr_text.strip():
            logger.warning('图片 %s OCR识别未提取到文本', image_path)
            return None

        account = _parse_account_from_text(ocr_text)
        if not account:
            logger.warning('未从图片 %s OCR结果中解析到银行账号', image_path)
            return None

        balance_keywords = ['期末余额', '余额', '可用余额', '账户余额', 'Balance', 'balance']
        statement_balance = None

        lines = ocr_text.split('\n')
        for line in lines:
            if any(keyword in line for keyword in balance_keywords):
                statement_balance = _parse_amount_from_text(line)
                if statement_balance is not None:
                    break

        if statement_balance is None:
            for line in reversed(lines):
                statement_balance = _parse_amount_from_text(line)
                if statement_balance is not None and abs(statement_balance) >= 1:
                    break

        if statement_balance is None:
            logger.warning('未从图片 %s OCR结果中解析到期末余额', image_path)
            return None

        statement_date = _parse_date_from_text(ocr_text)

        bank_name = ''
        for line in lines[:10]:
            if '银行' in line:
                match = re.search(r'(.+?银行)', line)
                if match:
                    bank_name = match.group(1).strip()
                    break

        record = BankStatementBalance(
            bank_account=account,
            bank_name=bank_name,
            statement_balance=statement_balance,
            statement_date=statement_date,
            source_type='image',
            source_file=os.path.basename(image_path),
            extracted_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        )

        logger.info('图片OCR解析成功: 账号 %s, 余额 %.2f, 日期 %s',
                    account, statement_balance, statement_date)
        return record

    except Exception as e:
        logger.error('解析图片文件失败 %s: %s', image_path, e, exc_info=True)
        return None


def manual_input_balance(account: str, balance: float,
                         bank_name: str = '', subject: str = '',
                         statement_date: Optional[str] = None) -> BankStatementBalance:
    """
    手动输入银行对账单余额。

    Args:
        account: 银行账号
        balance: 期末余额
        bank_name: 银行名称
        subject: 主体名称
        statement_date: 对账单日期

    Returns:
        BankStatementBalance: 手动输入的余额记录
    """
    return BankStatementBalance(
        bank_account=account,
        bank_name=bank_name,
        subject=subject,
        statement_balance=float(balance),
        statement_date=statement_date,
        source_type='manual',
        source_file='manual_input',
        extracted_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    )


def reconcile_balances(total_balances: Dict[str, AccountEndBalance],
                       statement_balances: Dict[str, BankStatementBalance],
                       tolerance: float = 0.01) -> BalanceReconciliationResult:
    """
    比对总表期末余额与银行对账单余额。

    比对逻辑：
    1. 按账号匹配总表余额和对账单余额
    2. 计算差异金额
    3. 根据容差判断是否一致
    4. 生成差异说明

    Args:
        total_balances: 总表期末余额字典（账号 -> AccountEndBalance）
        statement_balances: 银行对账单余额字典（账号 -> BankStatementBalance）
        tolerance: 容差，默认0.01元

    Returns:
        BalanceReconciliationResult: 比对结果
    """
    logger = get_logger()
    result = BalanceReconciliationResult()

    all_accounts = set(total_balances.keys()) | set(statement_balances.keys())
    result.total_accounts = len(all_accounts)

    if not all_accounts:
        logger.warning('无账号可比对')
        result.check_summary = {'status': '无数据'}
        return result

    records: List[BalanceReconciliationRecord] = []

    for account in sorted(all_accounts):
        total_info = total_balances.get(account)
        statement_info = statement_balances.get(account)

        if total_info is None and statement_info is None:
            continue

        record = BalanceReconciliationRecord(
            bank_account=account,
        )

        if total_info is not None:
            record.bank_name = total_info.bank_name
            record.subject = total_info.subject
            record.total_balance = total_info.end_balance
            record.total_end_date = total_info.end_date

        if statement_info is not None:
            if not record.bank_name:
                record.bank_name = statement_info.bank_name
            if not record.subject:
                record.subject = statement_info.subject
            record.statement_balance = statement_info.statement_balance
            record.statement_date = statement_info.statement_date
            record.source_file = statement_info.source_file

        if total_info is None:
            record.status = 'missing_total'
            record.diff_note = '总表中无此账号记录，无法比对'
            result.missing_total += 1
        elif statement_info is None:
            record.status = 'missing_statement'
            record.diff_note = '缺少该账号的银行对账单，请导入后重新比对'
            result.missing_statement += 1
        else:
            record.diff_amount = round(record.total_balance - record.statement_balance, 2)
            result.total_diff_amount += abs(record.diff_amount)

            if abs(record.diff_amount) <= tolerance:
                record.status = 'matched'
                record.diff_note = '余额一致，核对通过'
                result.matched_accounts += 1
            else:
                record.status = 'diff'
                result.diff_accounts += 1

                diff_notes = []
                diff_notes.append(f'总表余额: {record.total_balance:,.2f} 元')
                diff_notes.append(f'对账单余额: {record.statement_balance:,.2f} 元')
                diff_notes.append(f'差异金额: {record.diff_amount:,.2f} 元')

                if record.total_end_date and record.statement_date:
                    if record.total_end_date != record.statement_date:
                        diff_notes.append(f'⚠️  日期不一致: 总表日期 {record.total_end_date}，对账单日期 {record.statement_date}')

                if abs(record.diff_amount) > 10000:
                    diff_notes.append('⚠️  差异金额较大，建议逐笔核对交易明细')
                elif abs(record.diff_amount) > 1000:
                    diff_notes.append('⚠️  差异金额中等，建议检查是否有未达账项')

                record.diff_note = '；'.join(diff_notes)

        records.append(record)

    result.records = records
    result.total_diff_amount = round(result.total_diff_amount, 2)

    result.check_summary = {
        'total_accounts': result.total_accounts,
        'matched_accounts': result.matched_accounts,
        'diff_accounts': result.diff_accounts,
        'missing_statement': result.missing_statement,
        'missing_total': result.missing_total,
        'total_diff_amount': result.total_diff_amount,
        'match_rate': f'{(result.matched_accounts / result.total_accounts * 100):.1f}%' if result.total_accounts > 0 else '0%',
        'tolerance': tolerance,
    }

    logger.info('余额比对完成: 总账号 %d, 一致 %d, 差异 %d, 缺对账单 %d, 缺总表 %d, 累计差异 %s 元',
                result.total_accounts, result.matched_accounts, result.diff_accounts,
                result.missing_statement, result.missing_total, f'{result.total_diff_amount:.2f}')

    return result


def export_balance_reconciliation_result(recon_result: BalanceReconciliationResult,
                                         output_path: str,
                                         source_info: Optional[Dict[str, Any]] = None) -> str:
    """
    导出资余额核对结果为 Excel 文件。

    输出的 Sheet 包括：
    1. 核对总览 - 整体统计信息
    2. 比对明细 - 所有账号的比对详情
    3. 差异账号清单 - 存在余额差异的账号
    4. 待完善清单 - 缺少对账单或总表记录的账号

    Args:
        recon_result: 比对结果
        output_path: 输出文件路径
        source_info: 数据源信息

    Returns:
        str: 输出文件路径
    """
    logger = get_logger()

    try:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            overview_data = []
            overview_data.append({'核对项': '核对账号总数', '数值': recon_result.total_accounts})
            overview_data.append({'核对项': '余额一致账号数', '数值': recon_result.matched_accounts})
            overview_data.append({'核对项': '余额差异账号数', '数值': recon_result.diff_accounts})
            overview_data.append({'核对项': '缺少银行对账单账号数', '数值': recon_result.missing_statement})
            overview_data.append({'核对项': '总表无记录账号数', '数值': recon_result.missing_total})
            overview_data.append({'核对项': '累计差异金额(元)', '数值': recon_result.total_diff_amount})
            overview_data.append({'核对项': '核对一致率', '数值': recon_result.check_summary.get('match_rate', '0%')})
            overview_data.append({'核对项': '容差(元)', '数值': recon_result.check_summary.get('tolerance', 0.01)})

            if source_info:
                for key, value in source_info.items():
                    overview_data.append({'核对项': str(key), '数值': str(value)})

            overview_data.append({'核对项': '生成时间', '数值': datetime.now().strftime('%Y-%m-%d %H:%M:%S')})

            overview_df = pd.DataFrame(overview_data)
            overview_df.to_excel(writer, sheet_name='核对总览', index=False)

            detail_data = []
            for idx, record in enumerate(recon_result.records, 1):
                status_text = {
                    'matched': '✅ 一致',
                    'diff': '❌ 差异',
                    'missing_statement': '⚠️  缺对账单',
                    'missing_total': '⚠️  缺总表',
                    'pending': '待处理',
                }.get(record.status, record.status)

                detail_data.append({
                    '序号': idx,
                    '主体': record.subject,
                    '银行': record.bank_name,
                    '银行账号': record.bank_account,
                    '总表期末余额(元)': record.total_balance,
                    '银行对账单余额(元)': record.statement_balance,
                    '差异金额(元)': record.diff_amount,
                    '比对状态': status_text,
                    '总表期末日期': record.total_end_date or '',
                    '对账单日期': record.statement_date or '',
                    '来源文件': record.source_file,
                    '差异说明': record.diff_note,
                })

            detail_df = pd.DataFrame(detail_data)
            detail_df.to_excel(writer, sheet_name='比对明细', index=False)

            diff_records = [r for r in recon_result.records if r.status == 'diff']
            if diff_records:
                diff_data = []
                for idx, record in enumerate(diff_records, 1):
                    diff_data.append({
                        '序号': idx,
                        '主体': record.subject,
                        '银行': record.bank_name,
                        '银行账号': record.bank_account,
                        '总表余额(元)': record.total_balance,
                        '对账单余额(元)': record.statement_balance,
                        '差异金额(元)': record.diff_amount,
                        '总表日期': record.total_end_date or '',
                        '对账单日期': record.statement_date or '',
                        '差异说明': record.diff_note,
                    })
                diff_df = pd.DataFrame(diff_data)
                diff_df.to_excel(writer, sheet_name='差异账号清单', index=False)

            pending_records = [r for r in recon_result.records
                             if r.status in ('missing_statement', 'missing_total')]
            if pending_records:
                pending_data = []
                for idx, record in enumerate(pending_records, 1):
                    pending_type = '缺少银行对账单' if record.status == 'missing_statement' else '总表无此账号'
                    pending_data.append({
                        '序号': idx,
                        '主体': record.subject,
                        '银行': record.bank_name,
                        '银行账号': record.bank_account,
                        '问题类型': pending_type,
                        '总表余额(元)': record.total_balance,
                        '对账单余额(元)': record.statement_balance,
                        '说明': record.diff_note,
                    })
                pending_df = pd.DataFrame(pending_data)
                pending_df.to_excel(writer, sheet_name='待完善清单', index=False)

            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                df = None
                if sheet_name == '核对总览':
                    df = overview_df
                elif sheet_name == '比对明细':
                    df = detail_df
                elif sheet_name == '差异账号清单':
                    df = diff_df if diff_records else None
                elif sheet_name == '待完善清单':
                    df = pending_df if pending_records else None

                if df is not None and len(df) > 0:
                    amount_cols = set()
                    count_cols = set()
                    for col_idx, col_name in enumerate(df.columns, 1):
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        col_name_str = str(col_name)
                        if '金额' in col_name_str or '余额' in col_name_str:
                            amount_cols.add(col_letter)
                        elif '序号' in col_name_str or '数' in col_name_str:
                            count_cols.add(col_letter)

                        max_len = max(
                            len(col_name_str),
                            max((len(str(v)) for v in df.iloc[:, col_idx - 1].astype(str)), default=0)
                        )
                        ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

                    for row in ws.iter_rows(min_row=2):
                        for cell in row:
                            col_letter = cell.column_letter
                            if col_letter in amount_cols and isinstance(cell.value, (int, float)):
                                cell.number_format = '#,##0.00'
                            elif col_letter in count_cols and isinstance(cell.value, int):
                                cell.number_format = '#,##0'

        logger.info('期末余额与银行对账单核对报告已导出: %s', output_path)
        return output_path

    except Exception as e:
        logger.error('导出余额核对报告失败: %s', e, exc_info=True)
        raise


def generate_balance_reconciliation_from_total(total_input,
                                               statement_files: Optional[List[str]] = None,
                                               statement_balances: Optional[Dict[str, BankStatementBalance]] = None,
                                               output_dir: Optional[str] = None,
                                               tolerance: float = 0.01) -> Optional[str]:
    """
    从总表文件或交易记录生成期末余额与银行对账单核对报告。

    Args:
        total_input: 银行流水总表文件路径（str）或交易记录列表（List[Dict]）
        statement_files: 银行对账单文件列表（PDF或图片）
        statement_balances: 手动输入的银行对账单余额字典
        output_dir: 输出目录
        tolerance: 容差

    Returns:
        Optional[str]: 生成的报告文件路径，失败返回 None
    """
    logger = get_logger()

    total_path = None
    if isinstance(total_input, str):
        total_path = total_input
        records = load_total_table(total_path)
    else:
        records = total_input

    if not records:
        if total_path:
            logger.warning('总表无数据: %s', total_path)
        else:
            logger.warning('交易记录为空')
        return None

    total_balances = extract_end_balances_from_total(records)
    if not total_balances:
        logger.warning('未从总表提取到任何账号的期末余额')
        return None

    all_statement_balances: Dict[str, BankStatementBalance] = {}

    if statement_balances:
        all_statement_balances.update(statement_balances)

    if statement_files:
        for file_path in statement_files:
            record = None
            ext = os.path.splitext(file_path)[1].lower()
            if ext == '.pdf':
                record = parse_bank_statement_pdf(file_path)
            elif ext in ('.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.webp'):
                record = parse_bank_statement_image(file_path)
            else:
                logger.warning('不支持的文件类型: %s', file_path)
                continue

            if record:
                account_key = _account_key(record.bank_account)
                if account_key in all_statement_balances:
                    logger.warning('账号 %s 已有对账单记录，将覆盖', record.bank_account)
                all_statement_balances[account_key] = record

    normalized_total_balances: Dict[str, AccountEndBalance] = {}
    for account, balance_info in total_balances.items():
        key = _account_key(account)
        normalized_total_balances[key] = balance_info

    recon_result = reconcile_balances(
        normalized_total_balances,
        all_statement_balances,
        tolerance=tolerance
    )

    if output_dir is None:
        output_dir = os.path.dirname(total_path) if total_path else get_script_dir()

    os.makedirs(output_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'期末余额与银行对账单核对报告_{timestamp}.xlsx')

    source_info = {
        '总表文件': os.path.basename(total_path) if total_path else '内存数据',
        '总表记录数': len(records),
        '总表账号数': len(total_balances),
        '对账单文件数': len(statement_files or []),
        '手动输入账号数': len(statement_balances or {}),
        '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    return export_balance_reconciliation_result(recon_result, output_path, source_info)


def _scan_statement_files(folder: str) -> List[str]:
    """
    扫描文件夹中的银行对账单文件（PDF和图片）。

    Args:
        folder: 文件夹路径

    Returns:
        List[str]: 对账单文件路径列表
    """
    logger = get_logger()
    statement_files = []

    if not os.path.exists(folder):
        return statement_files

    supported_ext = ('.pdf', '.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.webp')

    for root, dirs, files in os.walk(folder):
        for f in files:
            if f.lower().endswith(supported_ext):
                full_path = os.path.join(root, f)
                statement_files.append(full_path)
                logger.debug('发现对账单文件: %s', full_path)

    logger.info('共扫描到 %d 个对账单文件', len(statement_files))
    return statement_files


def run_balance_reconciliation_flow(script_dir):
    """期末余额与银行对账单核对 CLI 流程"""
    logger = get_logger()
    logger.info('========== 期末余额与银行对账单核对开始 ==========')

    print('\n' + '=' * 70)
    print('期末余额与银行对账单核对 - 导入官方对账单，与总表按账号比对')
    print('=' * 70)

    total_path = ask_file('请选择【银行流水总表】文件')
    if not total_path:
        show_info('提示', '未选择总表文件，返回。')
        return
    logger.info('用户选择总表文件: %s', total_path)

    records = load_total_table(total_path)
    if not records:
        show_warning('错误', '总表文件无数据或读取失败。')
        return

    total_balances = extract_end_balances_from_total(records)
    if not total_balances:
        show_warning('错误', '未从总表提取到任何账号的期末余额。')
        return

    print(f'\n从总表提取到 {len(total_balances)} 个账号的期末余额：')
    for i, (account, info) in enumerate(sorted(total_balances.items()), 1):
        print(f'  {i:2d}. {info.subject or "未知主体"} - {info.bank_name or "未知银行"} '
              f'{account}: {info.end_balance:,.2f} 元 (日期: {info.end_date or "未知"})')

    print('\n' + '=' * 70)
    print('请选择银行对账单导入方式：')
    print('  1) 从文件夹批量导入（PDF/截图）')
    print('  2) 选择单个文件导入（PDF/截图）')
    print('  3) 手动输入对账单余额')
    print('  4) 混合模式（批量导入 + 手动补充）')
    print('  0) 返回主菜单')

    choice = input('\n请输入选项（默认 1）: ').strip() or '1'

    if choice == '0':
        return

    statement_balances: Dict[str, BankStatementBalance] = {}
    statement_files: List[str] = []

    if choice in ('1', '4'):
        print('\n请选择包含银行对账单的文件夹（PDF或截图）：')
        folder = ask_directory()
        if folder:
            statement_files = _scan_statement_files(folder)
            print(f'\n扫描到 {len(statement_files)} 个对账单文件')
            for f in statement_files:
                print(f'  - {os.path.basename(f)}')

            if not HAS_PDFPLUMBER:
                print('\n⚠️  未安装 pdfplumber，无法解析PDF文件。')
                print('   请运行: pip install pdfplumber')

            if not HAS_PYTESSERACT:
                print('\n⚠️  未安装 pytesseract，无法进行图片OCR。')
                print('   请运行: pip install pytesseract pillow')

    if choice in ('2', '4'):
        print('\n请选择对账单文件（PDF或截图）：')
        if HAS_TKINTER:
            root = tk.Tk()
            root.withdraw()
            files = filedialog.askopenfilenames(
                title='选择银行对账单文件',
                filetypes=[
                    ('PDF和图片文件', '*.pdf *.png *.jpg *.jpeg *.bmp *.tiff *.webp'),
                    ('PDF文件', '*.pdf'),
                    ('图片文件', '*.png *.jpg *.jpeg *.bmp *.tiff *.webp'),
                    ('所有文件', '*.*'),
                ]
            )
            root.destroy()
            for f in files:
                if f not in statement_files:
                    statement_files.append(f)
        else:
            while True:
                f = cli_askfile('请输入对账单文件路径（直接回车结束）: ')
                if not f:
                    break
                if f not in statement_files:
                    statement_files.append(f)

        if statement_files:
            print(f'\n共选择 {len(statement_files)} 个文件')

    if choice in ('3', '4'):
        print('\n=== 手动输入银行对账单余额 ===')
        print('（直接回车跳过输入）')
        while True:
            account = input('\n请输入银行账号: ').strip()
            if not account:
                break

            balance_input = input('请输入期末余额: ').strip()
            if not balance_input:
                continue

            try:
                balance = float(balance_input.replace(',', ''))
            except ValueError:
                print('⚠️  金额格式无效，请重新输入')
                continue

            bank_name = input('请输入银行名称（可选）: ').strip()
            subject = input('请输入主体名称（可选）: ').strip()
            statement_date = input('请输入对账单日期 YYYY-MM-DD（可选）: ').strip() or None

            record = manual_input_balance(
                account=account,
                balance=balance,
                bank_name=bank_name,
                subject=subject,
                statement_date=statement_date,
            )
            key = _account_key(account)
            statement_balances[key] = record
            print(f'✅ 已录入: {account} - 余额 {balance:,.2f} 元')

            more = input('\n是否继续录入？(y/N): ').strip().lower()
            if more != 'y':
                break

    if statement_files:
        print(f'\n开始解析 {len(statement_files)} 个对账单文件...')
        for i, file_path in enumerate(statement_files, 1):
            print(f'  [{i}/{len(statement_files)}] 正在解析: {os.path.basename(file_path)}')
            ext = os.path.splitext(file_path)[1].lower()
            record = None

            if ext == '.pdf':
                record = parse_bank_statement_pdf(file_path)
            elif ext in ('.png', '.jpg', '.jpeg', '.bmp', '.tiff', '.webp'):
                record = parse_bank_statement_image(file_path)

            if record:
                key = _account_key(record.bank_account)
                if key in statement_balances:
                    print(f'    ⚠️  账号 {record.bank_account} 已存在，将覆盖')
                statement_balances[key] = record
                print(f'    ✅ 解析成功: 账号 {record.bank_account}, 余额 {record.statement_balance:,.2f} 元')
            else:
                print(f'    ❌ 解析失败，请尝试手动录入')

    if not statement_balances:
        print('\n⚠️  未导入任何银行对账单余额，无法进行比对。')
        confirm = input('是否直接导出包含"待完善清单"的报告？(y/N): ').strip().lower()
        if confirm != 'y':
            return

    print(f'\n共导入 {len(statement_balances)} 个账号的银行对账单余额')

    tolerance_input = input('\n请输入容差（元，直接回车默认 0.01）: ').strip()
    tolerance = 0.01
    if tolerance_input:
        try:
            tolerance = float(tolerance_input)
        except ValueError:
            print('输入无效，使用默认容差 0.01 元')
            tolerance = 0.01

    print(f'\n开始比对，容差: {tolerance} 元...')

    normalized_total_balances: Dict[str, AccountEndBalance] = {}
    for account, balance_info in total_balances.items():
        key = _account_key(account)
        normalized_total_balances[key] = balance_info

    recon_result = reconcile_balances(
        normalized_total_balances,
        statement_balances,
        tolerance=tolerance
    )

    summary = recon_result.check_summary

    print('\n' + '=' * 70)
    print('比对结果总览')
    print('=' * 70)
    print(f'  核对账号总数:   {summary["total_accounts"]:,}')
    print(f'  余额一致账号:   {summary["matched_accounts"]:,}')
    print(f'  余额差异账号:   {summary["diff_accounts"]:,}')
    print(f'  缺少对账单:     {summary["missing_statement"]:,}')
    print(f'  总表无记录:     {summary["missing_total"]:,}')
    print(f'  核对一致率:     {summary["match_rate"]}')
    print(f'  累计差异金额:   {summary["total_diff_amount"]:,.2f} 元')

    if recon_result.diff_accounts > 0:
        print(f'\n  ⚠️  发现 {recon_result.diff_accounts} 个账号余额存在差异')
        diff_records = [r for r in recon_result.records if r.status == 'diff']
        for record in diff_records[:10]:
            print(f'    - {record.bank_account} ({record.subject}): '
                  f'总表 {record.total_balance:,.2f} vs 对账单 {record.statement_balance:,.2f}, '
                  f'差异 {record.diff_amount:,.2f} 元')
        if len(diff_records) > 10:
            print(f'    ... 还有 {len(diff_records) - 10} 个账号，详见导出文件')

    if recon_result.missing_statement > 0:
        print(f'\n  ⚠️  {recon_result.missing_statement} 个账号缺少银行对账单')
        missing_records = [r for r in recon_result.records if r.status == 'missing_statement']
        for record in missing_records[:10]:
            print(f'    - {record.bank_account} ({record.subject} - {record.bank_name}): '
                  f'总表余额 {record.total_balance:,.2f} 元')
        if len(missing_records) > 10:
            print(f'    ... 还有 {len(missing_records) - 10} 个账号，详见导出文件')

    if recon_result.matched_accounts == recon_result.total_accounts and recon_result.total_accounts > 0:
        print(f'\n  ✅ 太棒了！所有账号余额核对完全一致！')

    output_dir = input('\n请输入输出目录（直接回车默认当前目录）: ').strip()
    if not output_dir:
        output_dir = script_dir

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'期末余额与银行对账单核对报告_{timestamp}.xlsx')

    source_info = {
        '总表文件': os.path.basename(total_path),
        '总表记录数': len(records),
        '总表账号数': len(total_balances),
        '对账单文件数': len(statement_files),
        '手动输入账号数': sum(1 for r in statement_balances.values() if r.source_type == 'manual'),
        'PDF解析账号数': sum(1 for r in statement_balances.values() if r.source_type == 'pdf'),
        'OCR识别账号数': sum(1 for r in statement_balances.values() if r.source_type == 'image'),
        '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    try:
        export_balance_reconciliation_result(recon_result, output_path, source_info)
        msg = f'核对报告已导出！\n\n输出文件：{output_path}'
        show_info('导出成功', msg)
        logger.info('期末余额与银行对账单核对报告导出完成: %s', output_path)
    except Exception as e:
        msg = f'导出失败：{e}'
        show_warning('导出失败', msg)
        logger.error('期末余额与银行对账单核对报告导出失败: %s', e, exc_info=True)

    logger.info('========== 期末余额与银行对账单核对结束 ==========')


# ──────────────────────────────────────────────
# 非工作日交易标记模块
# ──────────────────────────────────────────────

CHINESE_HOLIDAYS = {
    '2020-01-01': '元旦',
    '2020-01-25': '春节',
    '2020-01-26': '春节',
    '2020-01-27': '春节',
    '2020-01-28': '春节',
    '2020-01-29': '春节',
    '2020-01-30': '春节',
    '2020-01-31': '春节',
    '2020-02-01': '春节',
    '2020-02-02': '春节',
    '2020-04-04': '清明节',
    '2020-04-05': '清明节',
    '2020-04-06': '清明节',
    '2020-05-01': '劳动节',
    '2020-05-02': '劳动节',
    '2020-05-03': '劳动节',
    '2020-05-04': '劳动节',
    '2020-05-05': '劳动节',
    '2020-06-25': '端午节',
    '2020-06-26': '端午节',
    '2020-06-27': '端午节',
    '2020-10-01': '国庆节/中秋节',
    '2020-10-02': '国庆节',
    '2020-10-03': '国庆节',
    '2020-10-04': '国庆节/中秋节',
    '2020-10-05': '国庆节',
    '2020-10-06': '国庆节',
    '2020-10-07': '国庆节',
    '2020-10-08': '国庆节',
    '2021-01-01': '元旦',
    '2021-01-02': '元旦',
    '2021-01-03': '元旦',
    '2021-02-11': '春节',
    '2021-02-12': '春节',
    '2021-02-13': '春节',
    '2021-02-14': '春节',
    '2021-02-15': '春节',
    '2021-02-16': '春节',
    '2021-02-17': '春节',
    '2021-04-03': '清明节',
    '2021-04-04': '清明节',
    '2021-04-05': '清明节',
    '2021-05-01': '劳动节',
    '2021-05-02': '劳动节',
    '2021-05-03': '劳动节',
    '2021-05-04': '劳动节',
    '2021-05-05': '劳动节',
    '2021-06-12': '端午节',
    '2021-06-13': '端午节',
    '2021-06-14': '端午节',
    '2021-09-19': '中秋节',
    '2021-09-20': '中秋节',
    '2021-09-21': '中秋节',
    '2021-10-01': '国庆节',
    '2021-10-02': '国庆节',
    '2021-10-03': '国庆节',
    '2021-10-04': '国庆节',
    '2021-10-05': '国庆节',
    '2021-10-06': '国庆节',
    '2021-10-07': '国庆节',
    '2022-01-01': '元旦',
    '2022-01-02': '元旦',
    '2022-01-03': '元旦',
    '2022-01-31': '春节',
    '2022-02-01': '春节',
    '2022-02-02': '春节',
    '2022-02-03': '春节',
    '2022-02-04': '春节',
    '2022-02-05': '春节',
    '2022-02-06': '春节',
    '2022-04-03': '清明节',
    '2022-04-04': '清明节',
    '2022-04-05': '清明节',
    '2022-04-30': '劳动节',
    '2022-05-01': '劳动节',
    '2022-05-02': '劳动节',
    '2022-05-03': '劳动节',
    '2022-05-04': '劳动节',
    '2022-06-03': '端午节',
    '2022-06-04': '端午节',
    '2022-06-05': '端午节',
    '2022-09-10': '中秋节',
    '2022-09-11': '中秋节',
    '2022-09-12': '中秋节',
    '2022-10-01': '国庆节',
    '2022-10-02': '国庆节',
    '2022-10-03': '国庆节',
    '2022-10-04': '国庆节',
    '2022-10-05': '国庆节',
    '2022-10-06': '国庆节',
    '2022-10-07': '国庆节',
    '2023-01-01': '元旦',
    '2023-01-02': '元旦',
    '2023-01-21': '春节',
    '2023-01-22': '春节',
    '2023-01-23': '春节',
    '2023-01-24': '春节',
    '2023-01-25': '春节',
    '2023-01-26': '春节',
    '2023-01-27': '春节',
    '2023-04-05': '清明节',
    '2023-04-29': '劳动节',
    '2023-04-30': '劳动节',
    '2023-05-01': '劳动节',
    '2023-05-02': '劳动节',
    '2023-05-03': '劳动节',
    '2023-06-22': '端午节',
    '2023-06-23': '端午节',
    '2023-06-24': '端午节',
    '2023-09-29': '中秋节/国庆节',
    '2023-09-30': '国庆节',
    '2023-10-01': '国庆节',
    '2023-10-02': '国庆节',
    '2023-10-03': '国庆节',
    '2023-10-04': '国庆节',
    '2023-10-05': '国庆节',
    '2023-10-06': '国庆节',
    '2024-01-01': '元旦',
    '2024-02-10': '春节',
    '2024-02-11': '春节',
    '2024-02-12': '春节',
    '2024-02-13': '春节',
    '2024-02-14': '春节',
    '2024-02-15': '春节',
    '2024-02-16': '春节',
    '2024-02-17': '春节',
    '2024-04-04': '清明节',
    '2024-04-05': '清明节',
    '2024-04-06': '清明节',
    '2024-05-01': '劳动节',
    '2024-05-02': '劳动节',
    '2024-05-03': '劳动节',
    '2024-05-04': '劳动节',
    '2024-05-05': '劳动节',
    '2024-06-08': '端午节',
    '2024-06-09': '端午节',
    '2024-06-10': '端午节',
    '2024-09-15': '中秋节',
    '2024-09-16': '中秋节',
    '2024-09-17': '中秋节',
    '2024-10-01': '国庆节',
    '2024-10-02': '国庆节',
    '2024-10-03': '国庆节',
    '2024-10-04': '国庆节',
    '2024-10-05': '国庆节',
    '2024-10-06': '国庆节',
    '2024-10-07': '国庆节',
    '2025-01-01': '元旦',
    '2025-01-28': '春节',
    '2025-01-29': '春节',
    '2025-01-30': '春节',
    '2025-01-31': '春节',
    '2025-02-01': '春节',
    '2025-02-02': '春节',
    '2025-02-03': '春节',
    '2025-02-04': '春节',
    '2025-04-04': '清明节',
    '2025-04-05': '清明节',
    '2025-04-06': '清明节',
    '2025-05-01': '劳动节',
    '2025-05-02': '劳动节',
    '2025-05-03': '劳动节',
    '2025-05-04': '劳动节',
    '2025-05-05': '劳动节',
    '2025-05-31': '端午节',
    '2025-06-01': '端午节',
    '2025-06-02': '端午节',
    '2025-10-01': '国庆节/中秋节',
    '2025-10-02': '国庆节',
    '2025-10-03': '国庆节',
    '2025-10-04': '国庆节',
    '2025-10-05': '国庆节',
    '2025-10-06': '国庆节',
    '2025-10-07': '国庆节',
    '2025-10-08': '国庆节',
    '2026-01-01': '元旦',
    '2026-01-02': '元旦',
    '2026-01-03': '元旦',
    '2026-02-17': '春节',
    '2026-02-18': '春节',
    '2026-02-19': '春节',
    '2026-02-20': '春节',
    '2026-02-21': '春节',
    '2026-02-22': '春节',
    '2026-02-23': '春节',
    '2026-04-04': '清明节',
    '2026-04-05': '清明节',
    '2026-04-06': '清明节',
    '2026-05-01': '劳动节',
    '2026-05-02': '劳动节',
    '2026-05-03': '劳动节',
    '2026-05-04': '劳动节',
    '2026-05-05': '劳动节',
    '2026-05-30': '端午节',
    '2026-05-31': '端午节',
    '2026-06-01': '端午节',
    '2026-10-01': '国庆节',
    '2026-10-02': '国庆节',
    '2026-10-03': '国庆节',
    '2026-10-04': '国庆节',
    '2026-10-05': '国庆节',
    '2026-10-06': '国庆节',
    '2026-10-07': '国庆节',
    '2026-10-08': '国庆节',
}

CHINESE_WORKDAY_ADJUSTMENTS = {
    '2020-01-19', '2020-04-26', '2020-05-09', '2020-06-28', '2020-09-27', '2020-10-10',
    '2021-02-07', '2021-02-20', '2021-04-25', '2021-05-08', '2021-09-18', '2021-09-26', '2021-10-09',
    '2022-01-29', '2022-01-30', '2022-04-02', '2022-04-24', '2022-05-07', '2022-10-08', '2022-10-09',
    '2023-01-28', '2023-01-29', '2023-04-23', '2023-05-06', '2023-06-25', '2023-10-07', '2023-10-08',
    '2024-02-04', '2024-02-18', '2024-04-07', '2024-04-28', '2024-05-11', '2024-09-14', '2024-09-29', '2024-10-12',
    '2025-01-26', '2025-02-08', '2025-04-27', '2025-09-28', '2025-10-11',
    '2026-02-14', '2026-02-15', '2026-04-26', '2026-09-27', '2026-10-10',
}

HOLIDAY_TAG_WORKDAY = '工作日'
HOLIDAY_TAG_WEEKEND = '周末'
HOLIDAY_TAG_HOLIDAY = '法定节假日'
HOLIDAY_TAG_ADJUSTED_WORKDAY = '调休工作日'


def _parse_trade_date(value) -> Optional[datetime]:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime()
    s = str(value).strip()
    if not s:
        return None
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y年%m月%d日', '%Y%m%d'):
        try:
            return datetime.strptime(s[:10] if len(s) >= 10 else s, fmt)
        except (ValueError, TypeError):
            continue
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d')
    except (ValueError, TypeError):
        return None


def classify_date(date_val) -> Tuple[str, str]:
    """
    判断给定日期的类型。

    Args:
        date_val: 日期值（datetime、str 或 pandas.Timestamp）

    Returns:
        (date_tag, holiday_name)
        - date_tag: HOLIDAY_TAG_WORKDAY / HOLIDAY_TAG_WEEKEND / HOLIDAY_TAG_HOLIDAY / HOLIDAY_TAG_ADJUSTED_WORKDAY
        - holiday_name: 节假日名称（仅法定节假日有值）
    """
    dt = _parse_trade_date(date_val)
    if dt is None:
        return HOLIDAY_TAG_WORKDAY, ''

    date_str = dt.strftime('%Y-%m-%d')

    if date_str in CHINESE_WORKDAY_ADJUSTMENTS:
        return HOLIDAY_TAG_ADJUSTED_WORKDAY, ''

    if date_str in CHINESE_HOLIDAYS:
        return HOLIDAY_TAG_HOLIDAY, CHINESE_HOLIDAYS[date_str]

    if dt.weekday() >= 5:
        return HOLIDAY_TAG_WEEKEND, ''

    return HOLIDAY_TAG_WORKDAY, ''


@dataclass
class HolidayMarkedRecord:
    record_index: int
    unique_id: str
    bank: str
    bank_account: str
    subject: str
    trade_date: str
    date_tag: str
    holiday_name: str
    payment: Optional[float] = None
    receipt: Optional[float] = None
    balance: Optional[float] = None
    counterpart: str = ''
    summary: str = ''
    transaction_id: str = ''
    import_batch: str = ''


@dataclass
class HolidayCheckResult:
    total_records: int = 0
    workday_count: int = 0
    weekend_count: int = 0
    holiday_count: int = 0
    adjusted_workday_count: int = 0
    non_workday_records: List[HolidayMarkedRecord] = field(default_factory=list)
    date_type_stats: Dict[str, int] = field(default_factory=dict)
    holiday_name_stats: Dict[str, int] = field(default_factory=dict)
    subject_stats: Dict[str, Dict[str, int]] = field(default_factory=dict)
    check_summary: Dict[str, Any] = field(default_factory=dict)


def mark_non_workday_transactions(records: List[Dict[str, Any]]) -> HolidayCheckResult:
    """
    对交易记录进行非工作日标记。

    判断每笔交易的交易日期是否为非工作日（周末或法定节假日），
    并生成标记结果。

    Args:
        records: 交易记录列表

    Returns:
        HolidayCheckResult: 标记结果
    """
    logger = get_logger()
    result = HolidayCheckResult()
    result.total_records = len(records)

    if not records:
        logger.warning('无交易记录可进行非工作日标记')
        result.check_summary = {'status': '无数据'}
        return result

    date_type_stats = {HOLIDAY_TAG_WORKDAY: 0, HOLIDAY_TAG_WEEKEND: 0,
                       HOLIDAY_TAG_HOLIDAY: 0, HOLIDAY_TAG_ADJUSTED_WORKDAY: 0}
    holiday_name_stats: Dict[str, int] = {}
    subject_stats: Dict[str, Dict[str, int]] = {}

    for idx, rec in enumerate(records):
        date_tag, holiday_name = classify_date(rec.get('交易日期'))

        date_type_stats[date_tag] = date_type_stats.get(date_tag, 0) + 1

        if holiday_name:
            holiday_name_stats[holiday_name] = holiday_name_stats.get(holiday_name, 0) + 1

        subject = str(rec.get('主体', ''))
        if subject and date_tag in (HOLIDAY_TAG_WEEKEND, HOLIDAY_TAG_HOLIDAY):
            if subject not in subject_stats:
                subject_stats[subject] = {HOLIDAY_TAG_WEEKEND: 0, HOLIDAY_TAG_HOLIDAY: 0}
            if date_tag == HOLIDAY_TAG_WEEKEND:
                subject_stats[subject][HOLIDAY_TAG_WEEKEND] += 1
            elif date_tag == HOLIDAY_TAG_HOLIDAY:
                subject_stats[subject][HOLIDAY_TAG_HOLIDAY] += 1

        if date_tag in (HOLIDAY_TAG_WEEKEND, HOLIDAY_TAG_HOLIDAY):
            marked = HolidayMarkedRecord(
                record_index=idx + 1,
                unique_id=str(rec.get('唯一id', '')),
                bank=str(rec.get('银行', '')),
                bank_account=str(rec.get('银行账号', '')),
                subject=subject,
                trade_date=str(rec.get('交易日期', '')),
                date_tag=date_tag,
                holiday_name=holiday_name,
                payment=_safe_float(rec.get('付款')),
                receipt=_safe_float(rec.get('收款')),
                balance=_safe_float(rec.get('余额')),
                counterpart=str(rec.get('对方户名', '')),
                summary=str(rec.get('摘要', '')),
                transaction_id=str(rec.get('交易流水号', '')),
                import_batch=str(rec.get('导入批次号', '')),
            )
            result.non_workday_records.append(marked)

    result.workday_count = date_type_stats.get(HOLIDAY_TAG_WORKDAY, 0)
    result.weekend_count = date_type_stats.get(HOLIDAY_TAG_WEEKEND, 0)
    result.holiday_count = date_type_stats.get(HOLIDAY_TAG_HOLIDAY, 0)
    result.adjusted_workday_count = date_type_stats.get(HOLIDAY_TAG_ADJUSTED_WORKDAY, 0)
    result.date_type_stats = date_type_stats
    result.holiday_name_stats = holiday_name_stats
    result.subject_stats = subject_stats

    non_workday_count = result.weekend_count + result.holiday_count
    non_workday_rate = (non_workday_count / result.total_records * 100) if result.total_records > 0 else 0.0

    result.check_summary = {
        'total_records': result.total_records,
        'workday_count': result.workday_count,
        'weekend_count': result.weekend_count,
        'holiday_count': result.holiday_count,
        'adjusted_workday_count': result.adjusted_workday_count,
        'non_workday_count': non_workday_count,
        'non_workday_rate': round(non_workday_rate, 2),
        'marked_records': len(result.non_workday_records),
        'check_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    logger.info('非工作日交易标记完成: 总记录 %d, 工作日 %d, 周末 %d, 法定节假日 %d, 调休工作日 %d',
                result.total_records, result.workday_count, result.weekend_count,
                result.holiday_count, result.adjusted_workday_count)

    return result


def apply_holiday_tags(records: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    """
    对交易记录列表打上非工作日标签，回写到记录字典中。

    Args:
        records: 交易记录列表

    Returns:
        (标记后的记录列表, 标记统计摘要)
    """
    logger = get_logger()

    if not records:
        return records, {'tagged_count': 0}

    tagged_count = 0
    for rec in records:
        date_tag, holiday_name = classify_date(rec.get('交易日期'))
        rec['非工作日标签'] = date_tag
        rec['节假日名称'] = holiday_name
        if date_tag != HOLIDAY_TAG_WORKDAY:
            tagged_count += 1

    summary = {
        'total_records': len(records),
        'tagged_count': tagged_count,
        'workday_count': sum(1 for r in records if r.get('非工作日标签') == HOLIDAY_TAG_WORKDAY),
        'weekend_count': sum(1 for r in records if r.get('非工作日标签') == HOLIDAY_TAG_WEEKEND),
        'holiday_count': sum(1 for r in records if r.get('非工作日标签') == HOLIDAY_TAG_HOLIDAY),
        'adjusted_workday_count': sum(1 for r in records if r.get('非工作日标签') == HOLIDAY_TAG_ADJUSTED_WORKDAY),
    }

    if tagged_count > 0:
        logger.info('非工作日交易打标: 总记录 %d, 非工作日 %d (周末 %d, 节假日 %d)',
                    len(records), tagged_count,
                    summary['weekend_count'], summary['holiday_count'])

    return records, summary


def export_holiday_check_result(check_result: HolidayCheckResult,
                                output_path: str,
                                source_info: Optional[Dict[str, Any]] = None) -> str:
    """
    导出非工作日交易标记结果为 Excel 文件。

    输出的 Sheet 包括：
    1. 标记总览 - 整体统计信息
    2. 非工作日交易明细 - 所有非工作日交易记录
    3. 节假日类型分布 - 按节假日名称统计
    4. 主体分布 - 按主体统计非工作日交易

    Args:
        check_result: mark_non_workday_transactions 返回的标记结果
        output_path: 输出 Excel 文件路径
        source_info: 可选，数据源信息

    Returns:
        str: 输出文件路径
    """
    logger = get_logger()

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            overview_items = [
                ('检测项', '数值'),
                ('总记录数', check_result.check_summary.get('total_records', 0)),
                ('工作日交易数', check_result.check_summary.get('workday_count', 0)),
                ('周末交易数', check_result.check_summary.get('weekend_count', 0)),
                ('法定节假日交易数', check_result.check_summary.get('holiday_count', 0)),
                ('调休工作日交易数', check_result.check_summary.get('adjusted_workday_count', 0)),
                ('非工作日交易合计', check_result.check_summary.get('non_workday_count', 0)),
                ('非工作日交易占比(%)', check_result.check_summary.get('non_workday_rate', 0)),
                ('标记记录数', check_result.check_summary.get('marked_records', 0)),
                ('检测时间', check_result.check_summary.get('check_time', '')),
            ]
            if source_info:
                for k, v in source_info.items():
                    overview_items.append((k, v))

            overview_df = pd.DataFrame(overview_items[1:], columns=overview_items[0])
            overview_df.to_excel(writer, sheet_name='标记总览', index=False)

            if check_result.non_workday_records:
                detail_data = []
                for rec in check_result.non_workday_records:
                    detail_data.append({
                        '序号(总表)': rec.record_index,
                        '唯一ID': rec.unique_id,
                        '主体': rec.subject,
                        '银行': rec.bank,
                        '银行账号': rec.bank_account,
                        '交易日期': rec.trade_date,
                        '日期类型': rec.date_tag,
                        '节假日名称': rec.holiday_name,
                        '付款(元)': rec.payment,
                        '收款(元)': rec.receipt,
                        '对方户名': rec.counterpart,
                        '余额(元)': rec.balance,
                        '交易流水号': rec.transaction_id,
                        '摘要': rec.summary,
                        '导入批次号': rec.import_batch,
                    })

                detail_df = pd.DataFrame(detail_data)
                detail_cols = [
                    '序号(总表)', '唯一ID', '主体', '银行', '银行账号',
                    '交易日期', '日期类型', '节假日名称',
                    '付款(元)', '收款(元)', '对方户名', '余额(元)',
                    '交易流水号', '摘要', '导入批次号',
                ]
                detail_df = detail_df[[c for c in detail_cols if c in detail_df.columns]]
                detail_df.to_excel(writer, sheet_name='非工作日交易明细', index=False)

                ws_detail = writer.sheets['非工作日交易明细']
                amount_cols = set()
                for col_idx, col_name in enumerate(detail_df.columns, 1):
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    if '元' in str(col_name):
                        amount_cols.add(col_letter)
                    max_len = max(
                        len(str(col_name)),
                        max((len(str(v)) for v in detail_df.iloc[:, col_idx - 1].astype(str)), default=0)
                    )
                    ws_detail.column_dimensions[col_letter].width = min(max_len + 4, 40)

                for row in ws_detail.iter_rows(min_row=2):
                    for cell in row:
                        if cell.column_letter in amount_cols:
                            cell.number_format = '#,##0.00'

                holiday_type_data = []
                for name, count in sorted(check_result.holiday_name_stats.items(), key=lambda x: -x[1]):
                    holiday_type_data.append({
                        '节假日名称': name,
                        '交易笔数': count,
                    })

                for tag, label in [(HOLIDAY_TAG_WEEKEND, '周末'), (HOLIDAY_TAG_HOLIDAY, '法定节假日')]:
                    count = check_result.date_type_stats.get(tag, 0)
                    if count > 0:
                        holiday_type_data.append({
                            '节假日名称': label,
                            '交易笔数': count,
                        })

                if holiday_type_data:
                    type_df = pd.DataFrame(holiday_type_data)
                    type_df.to_excel(writer, sheet_name='节假日类型分布', index=False)

                    ws_type = writer.sheets['节假日类型分布']
                    for col_idx, col_name in enumerate(type_df.columns, 1):
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        max_len = max(
                            len(str(col_name)),
                            max((len(str(v)) for v in type_df.iloc[:, col_idx - 1].astype(str)), default=0)
                        )
                        ws_type.column_dimensions[col_letter].width = min(max_len + 4, 30)

                subject_data = []
                for subject, stats in sorted(check_result.subject_stats.items()):
                    total = stats.get(HOLIDAY_TAG_WEEKEND, 0) + stats.get(HOLIDAY_TAG_HOLIDAY, 0)
                    subject_data.append({
                        '主体': subject,
                        '周末交易笔数': stats.get(HOLIDAY_TAG_WEEKEND, 0),
                        '节假日交易笔数': stats.get(HOLIDAY_TAG_HOLIDAY, 0),
                        '非工作日合计': total,
                    })

                if subject_data:
                    subject_df = pd.DataFrame(subject_data)
                    subject_df.to_excel(writer, sheet_name='主体分布', index=False)

                    ws_subj = writer.sheets['主体分布']
                    for col_idx, col_name in enumerate(subject_df.columns, 1):
                        col_letter = openpyxl.utils.get_column_letter(col_idx)
                        max_len = max(
                            len(str(col_name)),
                            max((len(str(v)) for v in subject_df.iloc[:, col_idx - 1].astype(str)), default=0)
                        )
                        ws_subj.column_dimensions[col_letter].width = min(max_len + 4, 40)

            ws_overview = writer.sheets['标记总览']
            for col_idx in range(1, 3):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                ws_overview.column_dimensions[col_letter].width = 25

            for row in ws_overview.iter_rows(min_row=2):
                for cell in row:
                    if cell.column == 2:
                        val = cell.value
                        if isinstance(val, (int, float)):
                            if isinstance(val, float):
                                cell.number_format = '#,##0.00'
                            else:
                                cell.number_format = '#,##0'

        logger.info('非工作日交易标记结果已导出: %s', output_path)
        return output_path

    except Exception as e:
        logger.error('导出非工作日交易标记结果失败: %s', e, exc_info=True)
        raise


def generate_holiday_check_from_records(records: List[Dict[str, Any]],
                                        output_dir: Optional[str] = None,
                                        source_info: Optional[Dict[str, Any]] = None) -> Optional[str]:
    """
    从交易记录列表直接生成非工作日交易标记报告。

    Args:
        records: 交易记录列表
        output_dir: 输出目录
        source_info: 数据源信息

    Returns:
        str: 生成的文件路径，如无数据则返回 None
    """
    logger = get_logger()

    if not records:
        logger.warning('无交易记录，跳过非工作日交易标记')
        return None

    if output_dir is None:
        output_dir = get_script_dir()

    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'非工作日交易标记报告_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, filename)

    check_result = mark_non_workday_transactions(records)

    if not check_result.non_workday_records:
        logger.info('未发现非工作日交易，仍导出报告')
        return export_holiday_check_result(check_result, output_path, source_info)

    return export_holiday_check_result(check_result, output_path, source_info)


def generate_holiday_check_from_total(total_path: str,
                                      output_dir: Optional[str] = None) -> Optional[str]:
    """
    从银行流水总表文件生成非工作日交易标记报告。

    Args:
        total_path: 银行流水总表 Excel 文件路径
        output_dir: 输出目录

    Returns:
        str: 生成的文件路径，失败则返回 None
    """
    logger = get_logger()

    records = load_total_table(total_path)
    if not records:
        logger.warning('总表无数据: %s', total_path)
        return None

    if output_dir is None:
        output_dir = os.path.dirname(total_path) or get_script_dir()

    source_info = {
        '数据来源文件': os.path.basename(total_path),
        '总表记录数': len(records),
        '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    return generate_holiday_check_from_records(records, output_dir, source_info)


def run_holiday_check_flow(script_dir):
    """非工作日交易标记 CLI 流程"""
    logger = get_logger()
    logger.info('========== 非工作日交易标记开始 ==========')

    print('\n' + '=' * 70)
    print('非工作日交易标记 - 周末与法定节假日交易自动打标')
    print('=' * 70)
    print('\n请选择数据来源：')
    print('  1) 从银行流水总表文件（Excel）')
    print('  2) 从数据库（按条件查询后标记）')
    print('  0) 返回主菜单')

    choice = input('\n请输入选项（默认 1）: ').strip() or '1'

    records = []
    source_info = {}

    if choice == '0':
        return
    elif choice == '1':
        total_path = ask_file('请选择【银行流水总表】文件')
        if not total_path:
            show_info('提示', '未选择总表文件，返回。')
            return
        logger.info('用户选择总表文件: %s', total_path)
        records = load_total_table(total_path)
        if not records:
            show_warning('错误', '总表文件无数据或读取失败。')
            return
        source_info = {
            '数据来源文件': os.path.basename(total_path),
            '总表记录数': len(records),
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    elif choice == '2':
        if not HAS_DATABASE:
            show_warning('错误', '数据库模块不可用。')
            return

        print('\n输入查询条件（直接回车表示不限制）：')
        subject = input('主体名称: ').strip() or None
        bank = input('银行名称: ').strip() or None
        account = input('银行账号: ').strip() or None
        start_date = input('开始日期 (YYYY-MM-DD): ').strip() or None
        end_date = input('结束日期 (YYYY-MM-DD): ').strip() or None

        try:
            qr = db_module.query_transactions(
                subject=subject, bank=bank, account=account,
                start_date=start_date, end_date=end_date,
                limit=999999, script_dir=script_dir
            )
            records = [r.to_dict() for r in qr.records]
        except Exception as e:
            show_warning('错误', f'数据库查询失败: {e}')
            logger.error('数据库查询失败: %s', e, exc_info=True)
            return

        if not records:
            show_info('提示', '查询结果为空。')
            return

        source_info = {
            '数据来源': '数据库查询',
            '查询主体': subject or '全部',
            '查询银行': bank or '全部',
            '查询账号': account or '全部',
            '日期范围': f'{start_date or "不限"} ~ {end_date or "不限"}',
            '记录数': len(records),
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }
    else:
        print('无效选项')
        return

    print(f'\n开始标记，共 {len(records)} 条记录...')
    check_result = mark_non_workday_transactions(records)
    summary = check_result.check_summary

    print('\n' + '=' * 70)
    print('标记结果总览')
    print('=' * 70)
    print(f'  总记录数:         {summary.get("total_records", 0):,}')
    print(f'  工作日交易:       {summary.get("workday_count", 0):,}')
    print(f'  周末交易:         {summary.get("weekend_count", 0):,}')
    print(f'  法定节假日交易:   {summary.get("holiday_count", 0):,}')
    print(f'  调休工作日交易:   {summary.get("adjusted_workday_count", 0):,}')
    print(f'  非工作日合计:     {summary.get("non_workday_count", 0):,}')
    print(f'  非工作日占比:     {summary.get("non_workday_rate", 0):.2f}%')

    if check_result.non_workday_records:
        print(f'\n  ⚠️  发现 {len(check_result.non_workday_records)} 笔非工作日交易')

        if check_result.holiday_name_stats:
            print('\n  节假日分布：')
            for name, count in sorted(check_result.holiday_name_stats.items(), key=lambda x: -x[1]):
                print(f'    - {name}: {count} 笔')

        top_subjects = sorted(check_result.subject_stats.items(),
                              key=lambda x: sum(x[1].values()), reverse=True)[:10]
        if top_subjects:
            print('\n  主体TOP10（非工作日交易数）：')
            for subject, stats in top_subjects:
                total = stats.get(HOLIDAY_TAG_WEEKEND, 0) + stats.get(HOLIDAY_TAG_HOLIDAY, 0)
                print(f'    - {subject}: {total} 笔 (周末 {stats.get(HOLIDAY_TAG_WEEKEND, 0)}, '
                      f'节假日 {stats.get(HOLIDAY_TAG_HOLIDAY, 0)})')
    else:
        print(f'\n  ✅ 未发现非工作日交易！')

    output_dir = input('\n请输入输出目录（直接回车默认当前目录）: ').strip()
    if not output_dir:
        output_dir = script_dir

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = os.path.join(output_dir, f'非工作日交易标记报告_{timestamp}.xlsx')

    try:
        export_holiday_check_result(check_result, output_path, source_info)
        msg = f'标记报告已导出！\n\n输出文件：{output_path}'
        show_info('导出成功', msg)
        logger.info('非工作日交易标记报告导出完成: %s', output_path)
    except Exception as e:
        msg = f'导出失败：{e}'
        show_warning('导出失败', msg)
        logger.error('非工作日交易标记报告导出失败: %s', e, exc_info=True)

    logger.info('========== 非工作日交易标记结束 ==========')


# ──────────────────────────────────────────────
# 现金流分类模块
# ──────────────────────────────────────────────

CASHFLOW_CATEGORIES = {
    'operating': '经营活动现金流',
    'investing': '投资活动现金流',
    'financing': '筹资活动现金流',
    'unclassified': '未分类',
}

CASHFLOW_SUBCATEGORIES = {
    'salary': '工资薪金',
    'tax': '税费缴纳',
    'supplier': '供应商付款',
    'customer': '客户收款',
    'operating_other': '其他经营活动',
    'investment_in': '投资收回',
    'investment_out': '投资支付',
    'investment_income': '投资收益',
    'fixed_asset': '固定资产',
    'loan_in': '取得借款',
    'loan_out': '偿还借款',
    'interest': '利息支出',
    'dividend': '分配股利',
    'capital_in': '实收资本',
    'financing_other': '其他筹资活动',
    'transfer': '内部转账',
    'unclassified': '未分类',
}

CASHFLOW_CATEGORY_HIERARCHY = {
    'salary': 'operating',
    'tax': 'operating',
    'supplier': 'operating',
    'customer': 'operating',
    'operating_other': 'operating',
    'investment_in': 'investing',
    'investment_out': 'investing',
    'investment_income': 'investing',
    'fixed_asset': 'investing',
    'loan_in': 'financing',
    'loan_out': 'financing',
    'interest': 'financing',
    'dividend': 'financing',
    'capital_in': 'financing',
    'financing_other': 'financing',
    'transfer': 'unclassified',
    'unclassified': 'unclassified',
}


@dataclass
class CashflowClassificationRule:
    """现金流分类规则"""
    rule_id: str
    name: str
    subcategory: str
    summary_keywords: List[str] = field(default_factory=list)
    counterpart_keywords: List[str] = field(default_factory=list)
    match_mode: str = 'contains'
    amount_direction: str = 'any'
    priority: int = 100
    enabled: bool = True
    description: Optional[str] = None
    created_at: str = ''
    updated_at: str = ''
    created_by: str = ''

    @property
    def main_category(self) -> str:
        return CASHFLOW_CATEGORY_HIERARCHY.get(self.subcategory, 'unclassified')


class CashflowRuleConfig:
    """现金流分类规则配置管理"""

    def __init__(self, script_dir=None):
        self.script_dir = script_dir or get_script_dir()
        self.config_path = os.path.join(self.script_dir, 'cashflow_rules.json')
        self._rules: List[CashflowClassificationRule] = []
        self.load_config()

    def load_config(self):
        logger = get_logger()
        if os.path.exists(self.config_path):
            try:
                with open(self.config_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                self._rules = [CashflowClassificationRule(**r) for r in data.get('rules', [])]
                logger.info('现金流分类规则已加载: %d 条规则', len(self._rules))
            except Exception as e:
                logger.error('加载现金流分类规则失败: %s', e)
                self._rules = []
        else:
            self._rules = []
            self._init_default_rules()
            self.save_config()

    def _init_default_rules(self):
        default_rules = get_default_cashflow_rules()
        self._rules = default_rules

    def save_config(self):
        logger = get_logger()
        try:
            data = {
                'rules': [vars(r) for r in self._rules],
            }
            with open(self.config_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            logger.info('现金流分类规则已保存: %s', self.config_path)
        except Exception as e:
            logger.error('保存现金流分类规则失败: %s', e)

    def get_rules(self, subcategory=None, enabled=None,
                  main_category=None) -> List[CashflowClassificationRule]:
        result = sorted(self._rules, key=lambda r: r.priority)
        if subcategory is not None:
            result = [r for r in result if r.subcategory == subcategory]
        if enabled is not None:
            result = [r for r in result if r.enabled == enabled]
        if main_category is not None:
            result = [r for r in result
                      if CASHFLOW_CATEGORY_HIERARCHY.get(r.subcategory) == main_category]
        return result

    def add_rule(self, rule: CashflowClassificationRule) -> str:
        if not rule.rule_id:
            rule.rule_id = f"CFR{datetime.now().strftime('%Y%m%d%H%M%S')}{uuid.uuid4().hex[:6].upper()}"
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        if not rule.created_at:
            rule.created_at = now
        rule.updated_at = now
        self._rules.append(rule)
        self.save_config()
        return rule.rule_id

    def update_rule(self, rule_id: str, updates: dict) -> bool:
        for i, r in enumerate(self._rules):
            if r.rule_id == rule_id:
                for k, v in updates.items():
                    if hasattr(r, k):
                        setattr(r, k, v)
                r.updated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.save_config()
                return True
        return False

    def delete_rule(self, rule_id: str) -> bool:
        original_len = len(self._rules)
        self._rules = [r for r in self._rules if r.rule_id != rule_id]
        if len(self._rules) < original_len:
            self.save_config()
            return True
        return False

    def toggle_rule(self, rule_id: str, enabled: bool) -> bool:
        for r in self._rules:
            if r.rule_id == rule_id:
                r.enabled = enabled
                r.updated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                self.save_config()
                return True
        return False


_cashflow_rule_config_instance = None


def get_cashflow_rule_config(script_dir=None) -> CashflowRuleConfig:
    global _cashflow_rule_config_instance
    if _cashflow_rule_config_instance is None:
        _cashflow_rule_config_instance = CashflowRuleConfig(script_dir)
    return _cashflow_rule_config_instance


def _reset_cashflow_singleton():
    global _cashflow_rule_config_instance
    _cashflow_rule_config_instance = None


def get_default_cashflow_rules() -> List[CashflowClassificationRule]:
    """获取预设的默认分类规则"""
    now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    rules = []

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='工资薪金',
        subcategory='salary',
        summary_keywords=['工资', '薪资', '薪酬', '奖金', '绩效', '年终奖', '薪', '代发工资'],
        counterpart_keywords=[],
        match_mode='contains',
        amount_direction='payment',
        priority=10,
        enabled=True,
        description='摘要包含工资、薪资等关键词的付款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='税费缴纳',
        subcategory='tax',
        summary_keywords=['税', '税费', '税金', '增值税', '所得税', '印花税', '城建税', '附加税',
                           '个税', '社保', '公积金', '保险', '缴款', '缴税'],
        counterpart_keywords=['税务局', '税务', '社保', '公积金', '国家税务总局'],
        match_mode='contains',
        amount_direction='payment',
        priority=15,
        enabled=True,
        description='摘要或对方户名包含税费、社保、公积金等关键词',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='客户收款',
        subcategory='customer',
        summary_keywords=['货款', '销售', '收入', '营收', '应收账款', '回款', '结算', '工程款',
                           '服务费', '咨询费', '技术服务费', '产品销售'],
        counterpart_keywords=[],
        match_mode='contains',
        amount_direction='receipt',
        priority=30,
        enabled=True,
        description='摘要包含货款、销售等关键词的收款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='供应商付款',
        subcategory='supplier',
        summary_keywords=['货款', '采购', '原材料', '库存商品', '应付账款', '进货', '采购款',
                           '材料款', '设备款', '工程款', '劳务费'],
        counterpart_keywords=[],
        match_mode='contains',
        amount_direction='payment',
        priority=30,
        enabled=True,
        description='摘要包含采购、货款等关键词的付款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='投资支付',
        subcategory='investment_out',
        summary_keywords=['投资', '股权投资', '对外投资', '长期投资', '短期投资', '理财',
                           '购买理财', '基金', '股票', '债券'],
        counterpart_keywords=['证券公司', '基金公司', '理财', '投资公司'],
        match_mode='contains',
        amount_direction='payment',
        priority=20,
        enabled=True,
        description='摘要或对方户名包含投资、理财等关键词的付款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='投资收回',
        subcategory='investment_in',
        summary_keywords=['投资收回', '赎回', '理财赎回', '基金赎回', '卖出股票', '处置投资',
                           '投资回款', '撤资', '股权转让'],
        counterpart_keywords=['证券公司', '基金公司'],
        match_mode='contains',
        amount_direction='receipt',
        priority=20,
        enabled=True,
        description='摘要包含投资收回、赎回等关键词的收款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='投资收益',
        subcategory='investment_income',
        summary_keywords=['利息', '股息', '分红', '投资收益', '理财收益', '基金分红',
                           '股票分红', '债券利息'],
        counterpart_keywords=[],
        match_mode='contains',
        amount_direction='receipt',
        priority=18,
        enabled=True,
        description='摘要包含利息、股息、分红等关键词的收款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='固定资产',
        subcategory='fixed_asset',
        summary_keywords=['固定资产', '设备', '房产', '土地', '车辆', '办公设备', '机器设备',
                           '在建工程', '工程物资'],
        counterpart_keywords=[],
        match_mode='contains',
        amount_direction='payment',
        priority=25,
        enabled=True,
        description='摘要包含固定资产、设备等关键词的付款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='取得借款',
        subcategory='loan_in',
        summary_keywords=['借款', '贷款', '融资', '授信', '流动资金贷款', '项目贷款',
                           '银行贷款', '发放贷款', '贷款到账', '借款到账'],
        counterpart_keywords=['贷款公司', '小额贷款', '金融公司', '信托', '融资租赁'],
        match_mode='contains',
        amount_direction='receipt',
        priority=15,
        enabled=True,
        description='摘要包含借款、贷款等关键词的收款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='偿还借款',
        subcategory='loan_out',
        summary_keywords=['还款', '偿还贷款', '归还借款', '还贷款', '还本付息', '偿还本金',
                           '还贷', '归还贷款', '清偿贷款', '贷款还款', '偿还借款本金',
                           '偿还银行贷款', '贷款本金', '归还本金', '偿付本金'],
        counterpart_keywords=['贷款公司', '小额贷款', '金融公司', '信托', '融资租赁'],
        match_mode='contains',
        amount_direction='payment',
        priority=15,
        enabled=True,
        description='摘要包含还款、还贷款等关键词的付款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='利息支出',
        subcategory='interest',
        summary_keywords=['利息', '贷款利息', '借款利息', '资金占用费', '财务费用', '罚息',
                           '支付利息', '季度利息', '月度利息', '贷款利息支出', '利息支出'],
        counterpart_keywords=['贷款公司', '小额贷款', '金融公司', '信托'],
        match_mode='contains',
        amount_direction='payment',
        priority=12,
        enabled=True,
        description='摘要包含利息、财务费用等关键词的付款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='分配股利',
        subcategory='dividend',
        summary_keywords=['股利', '分红', '利润分配', '应付股利', '股东分红'],
        counterpart_keywords=[],
        match_mode='contains',
        amount_direction='payment',
        priority=18,
        enabled=True,
        description='摘要包含股利、分红等关键词的付款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='实收资本',
        subcategory='capital_in',
        summary_keywords=['投资款', '注册资本', '实收资本', '出资', '股东出资', '增资'],
        counterpart_keywords=[],
        match_mode='contains',
        amount_direction='receipt',
        priority=18,
        enabled=True,
        description='摘要包含投资款、注册资本等关键词的收款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='内部转账',
        subcategory='transfer',
        summary_keywords=['转账', '划转', '调拨', '内部划转', '账户划转', '同名转账',
                           '转存', '提现', '存现', '划转至', '划转自'],
        counterpart_keywords=[],
        match_mode='contains',
        amount_direction='any',
        priority=60,
        enabled=True,
        description='摘要包含转账、划转等关键词的记录（不影响净现金流）',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='其他经营活动-收款',
        subcategory='operating_other',
        summary_keywords=['其他应收', '其他应付', '暂收', '暂付', '往来款', '备用金', '报销',
                           '差旅费', '办公费', '招待费', '水电费', '物业费', '租金',
                           '广告费', '宣传费', '培训费', '会议费'],
        counterpart_keywords=[],
        match_mode='contains',
        amount_direction='receipt',
        priority=50,
        enabled=True,
        description='其他经营活动相关的收款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    rules.append(CashflowClassificationRule(
        rule_id='',
        name='其他经营活动-付款',
        subcategory='operating_other',
        summary_keywords=['其他应收', '其他应付', '暂收', '暂付', '往来款', '备用金', '报销',
                           '差旅费', '办公费', '招待费', '水电费', '物业费', '租金',
                           '广告费', '宣传费', '培训费', '会议费'],
        counterpart_keywords=[],
        match_mode='contains',
        amount_direction='payment',
        priority=50,
        enabled=True,
        description='其他经营活动相关的付款记录',
        created_at=now,
        updated_at=now,
        created_by='system',
    ))

    return rules


def _match_cashflow_text(text: str, keywords: List[str],
                         match_mode: str = 'contains') -> Optional[str]:
    """匹配文本与关键词"""
    if not text or not keywords:
        return None
    text = str(text).strip()
    for kw in keywords:
        kw = str(kw).strip()
        if not kw:
            continue
        if match_mode == 'exact':
            if text == kw:
                return kw
        elif match_mode == 'startswith':
            if text.startswith(kw):
                return kw
        elif match_mode == 'endswith':
            if text.endswith(kw):
                return kw
        elif match_mode == 'regex':
            try:
                if re.search(kw, text):
                    return kw
            except re.error:
                pass
        else:
            if kw in text:
                return kw
    return None


def _check_amount_direction(record: Dict, direction: str) -> bool:
    """检查金额方向是否匹配"""
    if direction == 'any':
        return True
    payment = record.get('付款')
    receipt = record.get('收款')
    is_payment = payment is not None and payment < 0
    is_receipt = receipt is not None and receipt > 0
    if direction == 'payment':
        return is_payment
    if direction == 'receipt':
        return is_receipt
    return True


def apply_cashflow_classification(records: List[Dict],
                                   script_dir=None) -> Tuple[List[Dict], Dict[str, Any]]:
    """
    对交易记录应用现金流分类规则

    Args:
        records: 交易记录列表
        script_dir: 脚本目录

    Returns:
        (分类后的记录列表, 分类统计摘要)
    """
    logger = get_logger()
    config = get_cashflow_rule_config(script_dir)
    rules = config.get_rules(enabled=True)

    classified_count = 0
    unclassified_count = 0
    category_counts: Dict[str, int] = {}
    subcategory_counts: Dict[str, int] = {}
    rule_hit_counts: Dict[str, int] = {}

    for rec in records:
        matched_rule = None
        matched_kw = None
        matched_source = None

        for rule in rules:
            if not _check_amount_direction(rec, rule.amount_direction):
                continue

            summary = rec.get('摘要', '')
            counterpart = rec.get('对方户名', '')

            kw = None
            source = None

            if rule.summary_keywords:
                kw = _match_cashflow_text(summary, rule.summary_keywords, rule.match_mode)
                if kw:
                    source = 'summary'

            if kw is None and rule.counterpart_keywords:
                kw = _match_cashflow_text(counterpart, rule.counterpart_keywords, rule.match_mode)
                if kw:
                    source = 'counterpart'

            if kw is not None:
                matched_rule = rule
                matched_kw = kw
                matched_source = source
                break

        if matched_rule:
            main_category = matched_rule.main_category
            rec['现金流主类别'] = CASHFLOW_CATEGORIES.get(main_category, '未分类')
            rec['现金流子类别'] = CASHFLOW_SUBCATEGORIES.get(matched_rule.subcategory, '未分类')
            rec['现金流分类主类别编码'] = main_category
            rec['现金流分类子类别编码'] = matched_rule.subcategory
            rec['现金流分类规则名称'] = matched_rule.name
            rec['现金流分类匹配关键词'] = matched_kw
            rec['现金流分类匹配来源'] = matched_source
            classified_count += 1
            category_counts[main_category] = category_counts.get(main_category, 0) + 1
            subcategory_counts[matched_rule.subcategory] = subcategory_counts.get(
                matched_rule.subcategory, 0) + 1
            rule_hit_counts[matched_rule.name] = rule_hit_counts.get(matched_rule.name, 0) + 1
        else:
            rec['现金流主类别'] = '未分类'
            rec['现金流子类别'] = '未分类'
            rec['现金流分类主类别编码'] = 'unclassified'
            rec['现金流分类子类别编码'] = 'unclassified'
            rec['现金流分类规则名称'] = ''
            rec['现金流分类匹配关键词'] = ''
            rec['现金流分类匹配来源'] = ''
            unclassified_count += 1
            category_counts['unclassified'] = category_counts.get('unclassified', 0) + 1
            subcategory_counts['unclassified'] = subcategory_counts.get('unclassified', 0) + 1

    summary = {
        'total_records': len(records),
        'classified_count': classified_count,
        'unclassified_count': unclassified_count,
        'classification_rate': round(
            classified_count / len(records) * 100, 2) if records else 0,
        'category_counts': category_counts,
        'subcategory_counts': subcategory_counts,
        'rule_hit_counts': rule_hit_counts,
    }

    logger.info(
        '现金流分类完成: 总记录 %s, 已分类 %s, 未分类 %s, 分类率 %s%%',
        str(summary['total_records']), str(summary['classified_count']),
        str(summary['unclassified_count']), f"{summary['classification_rate']:.2f}",
    )

    return records, summary


def summarize_cashflow_by_category(records: List[Dict],
                                   group_by: str = 'subcategory'
                                   ) -> List[Dict[str, Any]]:
    """
    按现金流类别汇总统计

    Args:
        records: 已分类的交易记录列表
        group_by: 汇总维度 'main' 按主类别, 'sub' 按子类别, 'all' 按全部

    Returns:
        分类汇总列表
    """
    summary_map: Dict[str, Dict[str, Any]] = {}

    for rec in records:
        main_code = rec.get('现金流分类主类别编码', 'unclassified')
        sub_code = rec.get('现金流分类子类别编码', 'unclassified')
        main_name = rec.get('现金流主类别', '未分类')
        sub_name = rec.get('现金流子类别', '未分类')

        payment = rec.get('付款') or 0
        receipt = rec.get('收款') or 0

        if payment < 0:
            outflow = abs(payment)
            inflow = 0
        else:
            outflow = 0

        if receipt > 0:
            inflow = receipt
        else:
            inflow = 0

        net_amount = inflow - outflow

        if group_by in ['main', 'all']:
            key = f"main:{main_code}"
            if key not in summary_map:
                summary_map[key] = {
                    '汇总维度': '主类别',
                    '主类别编码': main_code,
                    '主类别名称': main_name,
                    '子类别编码': '',
                    '子类别名称': '',
                    '交易笔数': 0,
                    '流入金额': 0.0,
                    '流出金额': 0.0,
                    '净额': 0.0,
                }
            summary_map[key]['交易笔数'] += 1
            summary_map[key]['流入金额'] += inflow
            summary_map[key]['流出金额'] += outflow
            summary_map[key]['净额'] += net_amount

        if group_by in ['sub', 'subcategory', 'all']:
            key = f"sub:{sub_code}"
            if key not in summary_map:
                summary_map[key] = {
                    '汇总维度': '子类别',
                    '主类别编码': CASHFLOW_CATEGORY_HIERARCHY.get(sub_code, 'unclassified'),
                    '主类别名称': CASHFLOW_CATEGORIES.get(
                        CASHFLOW_CATEGORY_HIERARCHY.get(sub_code, 'unclassified'), '未分类'),
                    '子类别编码': sub_code,
                    '子类别名称': sub_name,
                    '交易笔数': 0,
                    '流入金额': 0.0,
                    '流出金额': 0.0,
                    '净额': 0.0,
                }
            summary_map[key]['交易笔数'] += 1
            summary_map[key]['流入金额'] += inflow
            summary_map[key]['流出金额'] += outflow
            summary_map[key]['净额'] += net_amount

    result = sorted(summary_map.values(), key=lambda x: (x['主类别编码'], x['子类别编码']))

    for item in result:
        item['流入金额'] = round(item['流入金额'], 2)
        item['流出金额'] = round(item['流出金额'], 2)
        item['净额'] = round(item['净额'], 2)

    return result


def export_cashflow_summary(records: List[Dict], output_path: str,
                            include_details: bool = True) -> str:
    """
    导出现金流分类汇总表

    Args:
        records: 已分类的交易记录列表
        output_path: 输出文件路径
        include_details: 是否包含明细数据

    Returns:
        输出文件路径
    """
    logger = get_logger()

    main_summary = summarize_cashflow_by_category(records, group_by='main')
    sub_summary = summarize_cashflow_by_category(records, group_by='sub')

    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        df_main = pd.DataFrame(main_summary)
        if not df_main.empty:
            df_main = df_main[[
                '主类别编码', '主类别名称', '交易笔数',
                '流入金额', '流出金额', '净额'
            ]]
            df_main.columns = [
                '主类别编码', '主类别名称', '交易笔数',
                '流入金额(元)', '流出金额(元)', '净额(元)'
            ]
        df_main.to_excel(writer, sheet_name='主类别汇总', index=False)

        df_sub = pd.DataFrame(sub_summary)
        if not df_sub.empty:
            df_sub = df_sub[[
                '主类别编码', '主类别名称', '子类别编码', '子类别名称',
                '交易笔数', '流入金额', '流出金额', '净额'
            ]]
            df_sub.columns = [
                '主类别编码', '主类别名称', '子类别编码', '子类别名称',
                '交易笔数', '流入金额(元)', '流出金额(元)', '净额(元)'
            ]
        df_sub.to_excel(writer, sheet_name='子类别汇总', index=False)

        if include_details:
            detail_cols = [
                '唯一id', '银行', '银行账号', '主体', '交易日期',
                '付款', '收款', '摘要', '对方户名', '余额',
                '现金流主类别', '现金流子类别',
                '现金流分类主类别编码', '现金流分类子类别编码',
                '现金流分类规则名称', '现金流分类匹配关键词',
                '现金流分类匹配来源', '交易流水号',
            ]
            available_cols = [c for c in detail_cols if c in records[0]] if records else []
            df_detail = pd.DataFrame(records, columns=available_cols)
            df_detail.to_excel(writer, sheet_name='交易明细', index=False)

    wb = openpyxl.load_workbook(output_path)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if sheet_name == '交易明细':
            for row_idx in range(2, ws.max_row + 1):
                main_cat = str(ws.cell(row=row_idx, column=11).value or '')
                if main_cat == '经营活动现金流':
                    fill = openpyxl.styles.PatternFill(
                        start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                elif main_cat == '投资活动现金流':
                    fill = openpyxl.styles.PatternFill(
                        start_color='FFF3E6', end_color='FFF3E6', fill_type='solid')
                elif main_cat == '筹资活动现金流':
                    fill = openpyxl.styles.PatternFill(
                        start_color='F3E6FF', end_color='F3E6FF', fill_type='solid')
                else:
                    fill = openpyxl.styles.PatternFill(
                        start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
                for col_idx in range(1, ws.max_column + 1):
                    ws.cell(row=row_idx, column=col_idx).fill = fill

        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for row_idx in range(1, ws.max_row + 1):
                try:
                    cell_value = str(ws.cell(row=row_idx, column=col_idx).value or '')
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        header_fill = openpyxl.styles.PatternFill(
            start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=1, column=col_idx).fill = header_fill
            ws.cell(row=1, column=col_idx).font = header_font

        for sheet_name in ['主类别汇总', '子类别汇总']:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for col_idx in range(4, 7):
                    for row_idx in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.number_format = '#,##0.00'

    wb.save(output_path)
    wb.close()

    logger.info('现金流分类汇总表已导出: %s', output_path)
    return output_path


def add_cashflow_rule(name: str, subcategory: str,
                      summary_keywords: Optional[List[str]] = None,
                      counterpart_keywords: Optional[List[str]] = None,
                      match_mode: str = 'contains',
                      amount_direction: str = 'any',
                      priority: int = 100,
                      description: Optional[str] = None,
                      script_dir=None,
                      username=None) -> str:
    """
    便捷添加现金流分类规则

    Args:
        name: 规则名称
        subcategory: 子类别编码
        summary_keywords: 摘要关键词列表
        counterpart_keywords: 对方户名关键词列表
        match_mode: 匹配模式
        amount_direction: 金额方向
        priority: 优先级（数值越小越优先）
        description: 规则描述
        script_dir: 脚本目录
        username: 用户名

    Returns:
        规则ID
    """
    config = get_cashflow_rule_config(script_dir)
    rule = CashflowClassificationRule(
        rule_id='',
        name=name,
        subcategory=subcategory,
        summary_keywords=summary_keywords or [],
        counterpart_keywords=counterpart_keywords or [],
        match_mode=match_mode,
        amount_direction=amount_direction,
        priority=priority,
        enabled=True,
        description=description,
        created_at='',
        updated_at='',
        created_by=username or get_current_user(),
    )
    return config.add_rule(rule)


def init_default_cashflow_rules(script_dir=None) -> List[str]:
    """
    初始化默认现金流分类规则（会覆盖现有配置）

    Args:
        script_dir: 脚本目录

    Returns:
        新增的规则ID列表
    """
    config = CashflowRuleConfig(script_dir=script_dir)
    default_rules = get_default_cashflow_rules()
    rule_ids = []
    for rule in default_rules:
        rule.rule_id = ''
        rule_id = config.add_rule(rule)
        rule_ids.append(rule_id)
    return rule_ids


def get_cashflow_classification(records: List[Dict],
                                script_dir=None) -> Tuple[List[Dict], Dict[str, Any], List[Dict]]:
    """
    一站式现金流分类：分类 + 汇总

    Args:
        records: 交易记录列表
        script_dir: 脚本目录

    Returns:
        (分类后的记录, 分类统计摘要, 子类别汇总列表)
    """
    classified_records, classification_summary = apply_cashflow_classification(
        records, script_dir=script_dir)
    summary = summarize_cashflow_by_category(classified_records, group_by='all')
    return classified_records, classification_summary, summary


# ══════════════════════════════════════════════════════════════════
# 跨账号内部划转识别模块
# ══════════════════════════════════════════════════════════════════

INTERNAL_TRANSFER_TAG_YES = '是'
INTERNAL_TRANSFER_TAG_NO = ''
INTERNAL_TRANSFER_DIRECTION_OUT = '划出'
INTERNAL_TRANSFER_DIRECTION_IN = '划入'

INTERNAL_TRANSFER_EXTRA_COLUMNS = [
    '内部划转标记',
    '内部划转配对ID',
    '内部划转方向',
    '内部划转备注',
]


@dataclass
class InternalTransferMatch:
    """一对匹配的内部划转记录"""
    match_id: str
    amount: float
    out_record_id: str
    in_record_id: str
    out_subject: str
    in_subject: str
    out_bank: str
    in_bank: str
    out_date: str
    in_date: str
    days_diff: int
    out_counterparty: str
    in_counterparty: str


@dataclass
class InternalTransferResult:
    """内部划转识别结果"""
    marked_records: List[Dict[str, Any]]
    matches: List[InternalTransferMatch]
    total_records: int = 0
    match_pairs: int = 0
    marked_out_count: int = 0
    marked_in_count: int = 0
    involved_subjects: List[str] = field(default_factory=list)
    involved_banks: List[str] = field(default_factory=list)


def _build_subject_index(records: List[Dict[str, Any]]) -> Dict[str, List[int]]:
    """
    构建主体名称到记录索引列表的映射。
    用于快速查找某主体的所有交易记录。
    """
    idx: Dict[str, List[int]] = {}
    for i, rec in enumerate(records):
        subject = str(rec.get('主体') or '').strip()
        if subject:
            idx.setdefault(subject, []).append(i)
    return idx


def _build_counterparty_subject_index(records: List[Dict[str, Any]]) -> Dict[str, List[int]]:
    """
    构建"对方户名==已知主体"的快速索引。
    对方户名为 key，对应的记录索引列表为 value。
    """
    all_subjects = {str(r.get('主体') or '').strip() for r in records}
    all_subjects.discard('')
    idx: Dict[str, List[int]] = {}
    for i, rec in enumerate(records):
        cp = str(rec.get('对方户名') or '').strip()
        if cp and cp in all_subjects:
            idx.setdefault(cp, []).append(i)
    return idx


def _abs_amount_equal(a: float, b: float, tolerance: float = 0.01) -> bool:
    """比较两个金额绝对值是否在容忍度范围内相等"""
    return abs(abs(a) - abs(b)) <= tolerance


def identify_internal_transfers(
    records: List[Dict[str, Any]],
    time_window_days: int = 7,
    amount_tolerance: float = 0.01,
    strict_counterparty_match: bool = True,
) -> InternalTransferResult:
    """
    识别跨账号内部划转交易。

    识别策略（三重匹配）：
    1. 对方户名匹配：划出记录的对方户名 == 划入记录的主体名称
                     划入记录的对方户名 == 划出记录的主体名称
    2. 金额对称性：划出金额的绝对值 与 划入金额的绝对值 相等（容忍误差内）
    3. 时间接近度：两笔交易的时间差不超过 time_window_days 天

    Args:
        records: 交易记录列表，每条包含 '主体','银行','交易日期','付款','收款','对方户名','唯一id'
        time_window_days: 时间窗口（天），默认 7 天
        amount_tolerance: 金额容忍度（元），默认 0.01 元
        strict_counterparty_match: 是否严格要求双向对方户名匹配，默认 True
                                   False 时只要求单向（划出->对方户名==划入主体）

    Returns:
        InternalTransferResult: 包含标记后的记录和所有匹配对
    """
    logger = get_logger()

    if not records:
        return InternalTransferResult(
            marked_records=[], matches=[], total_records=0,
        )

    result_records: List[Dict[str, Any]] = []
    for rec in records:
        new_rec = dict(rec)
        for col in INTERNAL_TRANSFER_EXTRA_COLUMNS:
            new_rec.setdefault(col, '')
        result_records.append(new_rec)

    subject_index = _build_subject_index(result_records)
    cp_subject_index = _build_counterparty_subject_index(result_records)

    matches: List[InternalTransferMatch] = []
    used_idx: set = set()
    match_counter = 0

    out_records_idx = []
    for i, rec in enumerate(result_records):
        payment = to_float(rec.get('付款'))
        if payment is not None and payment < 0:
            out_records_idx.append(i)

    for out_idx in out_records_idx:
        if out_idx in used_idx:
            continue

        out_rec = result_records[out_idx]
        out_subject = str(out_rec.get('主体') or '').strip()
        out_bank = str(out_rec.get('银行') or '').strip()
        out_counterparty = str(out_rec.get('对方户名') or '').strip()
        out_amount = abs(to_float(out_rec.get('付款')) or 0.0)
        out_date_raw = out_rec.get('交易日期')
        out_date_dt = _normalize_date(out_date_raw)
        out_date_str = str(out_date_raw or '').strip()

        if out_amount <= 0 or out_date_dt is None:
            continue

        candidate_in_idx = set()
        if out_counterparty and out_counterparty in subject_index:
            candidate_in_idx.update(subject_index[out_counterparty])

        if not strict_counterparty_match:
            if out_counterparty and out_counterparty in cp_subject_index:
                for cidx in cp_subject_index[out_counterparty]:
                    cand_subject = str(result_records[cidx].get('主体') or '').strip()
                    if cand_subject == out_counterparty:
                        candidate_in_idx.add(cidx)

        best_in_idx = -1
        best_days_diff = 10 ** 9

        for in_idx in candidate_in_idx:
            if in_idx == out_idx or in_idx in used_idx:
                continue

            in_rec = result_records[in_idx]
            receipt = to_float(in_rec.get('收款'))
            if receipt is None or receipt <= 0:
                continue

            in_subject = str(in_rec.get('主体') or '').strip()
            in_counterparty = str(in_rec.get('对方户名') or '').strip()
            in_amount = abs(receipt)
            in_date_dt = _normalize_date(in_rec.get('交易日期'))

            if in_date_dt is None:
                continue

            if strict_counterparty_match:
                if in_counterparty != out_subject:
                    continue
                if out_counterparty != in_subject:
                    continue
            else:
                if out_counterparty != in_subject and in_counterparty != out_subject:
                    continue

            if not _abs_amount_equal(out_amount, in_amount, amount_tolerance):
                continue

            days_diff = abs((out_date_dt - in_date_dt).days)
            if days_diff > time_window_days:
                continue

            if days_diff < best_days_diff:
                best_days_diff = days_diff
                best_in_idx = in_idx

        if best_in_idx < 0:
            continue

        in_rec = result_records[best_in_idx]
        in_date_str = str(in_rec.get('交易日期') or '').strip()
        in_bank = str(in_rec.get('银行') or '').strip()
        in_counterparty_final = str(in_rec.get('对方户名') or '').strip()
        in_subject_final = str(in_rec.get('主体') or '').strip()

        match_counter += 1
        match_id = f"IT{datetime.now().strftime('%Y%m%d')}{match_counter:05d}"

        result_records[out_idx]['内部划转标记'] = INTERNAL_TRANSFER_TAG_YES
        result_records[out_idx]['内部划转配对ID'] = match_id
        result_records[out_idx]['内部划转方向'] = INTERNAL_TRANSFER_DIRECTION_OUT
        result_records[out_idx]['内部划转备注'] = (
            f"内部划转至[{in_subject_final}]({in_bank})，"
            f"配对ID:{match_id}，时间差{best_days_diff}天"
        )

        result_records[best_in_idx]['内部划转标记'] = INTERNAL_TRANSFER_TAG_YES
        result_records[best_in_idx]['内部划转配对ID'] = match_id
        result_records[best_in_idx]['内部划转方向'] = INTERNAL_TRANSFER_DIRECTION_IN
        result_records[best_in_idx]['内部划转备注'] = (
            f"内部划转来自[{out_subject}]({out_bank})，"
            f"配对ID:{match_id}，时间差{best_days_diff}天"
        )

        out_id = str(result_records[out_idx].get('唯一id') or '')
        in_id = str(result_records[best_in_idx].get('唯一id') or '')

        matches.append(InternalTransferMatch(
            match_id=match_id,
            amount=round(out_amount, 2),
            out_record_id=out_id,
            in_record_id=in_id,
            out_subject=out_subject,
            in_subject=in_subject_final,
            out_bank=out_bank,
            in_bank=in_bank,
            out_date=out_date_str,
            in_date=in_date_str,
            days_diff=best_days_diff,
            out_counterparty=out_counterparty,
            in_counterparty=in_counterparty_final,
        ))

        used_idx.add(out_idx)
        used_idx.add(best_in_idx)

    involved_subjects_set = set()
    involved_banks_set = set()
    for m in matches:
        involved_subjects_set.add(m.out_subject)
        involved_subjects_set.add(m.in_subject)
        involved_banks_set.add(m.out_bank)
        involved_banks_set.add(m.in_bank)

    marked_out_count = sum(
        1 for r in result_records
        if r.get('内部划转标记') == INTERNAL_TRANSFER_TAG_YES
        and r.get('内部划转方向') == INTERNAL_TRANSFER_DIRECTION_OUT
    )
    marked_in_count = sum(
        1 for r in result_records
        if r.get('内部划转标记') == INTERNAL_TRANSFER_TAG_YES
        and r.get('内部划转方向') == INTERNAL_TRANSFER_DIRECTION_IN
    )

    result = InternalTransferResult(
        marked_records=result_records,
        matches=matches,
        total_records=len(result_records),
        match_pairs=len(matches),
        marked_out_count=marked_out_count,
        marked_in_count=marked_in_count,
        involved_subjects=sorted(involved_subjects_set),
        involved_banks=sorted(involved_banks_set),
    )

    logger.info(
        '内部划转识别完成: %d 条记录, 识别出 %d 对 (%d 划出 + %d 划入), '
        '涉及 %d 个主体, %d 家银行',
        result.total_records, result.match_pairs,
        result.marked_out_count, result.marked_in_count,
        len(result.involved_subjects), len(result.involved_banks),
    )

    return result


def filter_internal_transfers_for_summary(
    records: List[Dict[str, Any]],
    exclude: bool = True,
) -> List[Dict[str, Any]]:
    """
    在进行主体汇总前，过滤内部划转记录。

    Args:
        records: 交易记录列表
        exclude: True 排除内部划转记录（默认，用于主体汇总避免重复计算）
                 False 只保留内部划转记录（用于查看、审计）

    Returns:
        过滤后的记录列表
    """
    if exclude:
        return [
            r for r in records
            if r.get('内部划转标记') != INTERNAL_TRANSFER_TAG_YES
        ]
    else:
        return [
            r for r in records
            if r.get('内部划转标记') == INTERNAL_TRANSFER_TAG_YES
        ]


def export_internal_transfer_report(
    it_result: InternalTransferResult,
    output_path: str,
    source_info: Optional[Dict[str, Any]] = None,
) -> str:
    """
    将内部划转识别结果导出为 Excel 报告。

    输出 Sheet:
    1. 识别总览 - 统计信息
    2. 配对明细 - 每对内部划转的详细信息
    3. 标记明细 - 被标记的所有原始记录（含标记字段）

    Args:
        it_result: identify_internal_transfers 返回的结果
        output_path: 输出 Excel 路径
        source_info: 可选的数据源信息

    Returns:
        输出文件路径
    """
    logger = get_logger()

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)

    try:
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            overview_items = [
                ('统计项', '数值'),
                ('总记录数', it_result.total_records),
                ('识别配对数', it_result.match_pairs),
                ('划出记录数', it_result.marked_out_count),
                ('划入记录数', it_result.marked_in_count),
                ('涉及主体数', len(it_result.involved_subjects)),
                ('涉及银行数', len(it_result.involved_banks)),
                ('划转总金额(元)', round(sum(m.amount for m in it_result.matches), 2)),
            ]
            if source_info:
                for k, v in source_info.items():
                    overview_items.append((k, v))

            overview_df = pd.DataFrame(overview_items[1:], columns=overview_items[0])
            overview_df.to_excel(writer, sheet_name='识别总览', index=False)

            if it_result.matches:
                match_rows = []
                for m in it_result.matches:
                    match_rows.append({
                        '配对ID': m.match_id,
                        '划转金额(元)': m.amount,
                        '划出主体': m.out_subject,
                        '划出银行': m.out_bank,
                        '划出日期': m.out_date,
                        '划出记录唯一ID': m.out_record_id,
                        '划出对方户名': m.out_counterparty,
                        '划入主体': m.in_subject,
                        '划入银行': m.in_bank,
                        '划入日期': m.in_date,
                        '划入记录唯一ID': m.in_record_id,
                        '划入对方户名': m.in_counterparty,
                        '时间差(天)': m.days_diff,
                    })
                match_df = pd.DataFrame(match_rows)
                match_df.to_excel(writer, sheet_name='配对明细', index=False)

            tagged_records = filter_internal_transfers_for_summary(
                it_result.marked_records, exclude=False,
            )
            if tagged_records:
                base_cols = list(STANDARD_COLUMNS)
                extra_cols = [
                    c for c in INTERNAL_TRANSFER_EXTRA_COLUMNS
                    if c not in base_cols
                ]
                available_base = [c for c in base_cols if c in tagged_records[0]]
                available_extra = [c for c in extra_cols if c in tagged_records[0]]
                columns_order = available_base + available_extra
                detail_df = pd.DataFrame(tagged_records, columns=columns_order)
                detail_df.to_excel(writer, sheet_name='标记明细', index=False)

        wb = openpyxl.load_workbook(output_path)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_fill = openpyxl.styles.PatternFill(
                start_color='4472C4', end_color='4472C4', fill_type='solid',
            )
            header_font = openpyxl.styles.Font(bold=True, color='FFFFFF')
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=1, column=col_idx).fill = header_fill
                ws.cell(row=1, column=col_idx).font = header_font

            for col_idx in range(1, ws.max_column + 1):
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                for row_idx in range(1, ws.max_row + 1):
                    try:
                        val = str(ws.cell(row=row_idx, column=col_idx).value or '')
                        if len(val) > max_len:
                            max_len = len(val)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

            if sheet_name in ('配对明细', '标记明细'):
                amount_col_names = ['划转金额(元)', '付款', '收款', '余额']
                for col_idx in range(1, ws.max_column + 1):
                    col_name = str(ws.cell(row=1, column=col_idx).value or '')
                    col_letter = openpyxl.utils.get_column_letter(col_idx)
                    if any(ac in col_name for ac in amount_col_names):
                        for row_idx in range(2, ws.max_row + 1):
                            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

            if sheet_name == '识别总览':
                for row_idx in range(2, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=2)
                    val = cell.value
                    if isinstance(val, float):
                        cell.number_format = '#,##0.00'
                    elif isinstance(val, int):
                        cell.number_format = '#,##0'

        wb.save(output_path)
        wb.close()

        logger.info('内部划转识别报告已导出: %s', output_path)
        return output_path

    except Exception as e:
        logger.error('导出内部划转识别报告失败: %s', e, exc_info=True)
        raise


def generate_internal_transfer_from_records(
    records: List[Dict[str, Any]],
    output_dir: Optional[str] = None,
    source_info: Optional[Dict[str, Any]] = None,
    time_window_days: int = 7,
    amount_tolerance: float = 0.01,
    strict_counterparty_match: bool = True,
) -> Optional[str]:
    """
    从交易记录列表一站式生成内部划转识别报告，并返回带标记的记录结果。

    Args:
        records: 交易记录列表
        output_dir: 输出目录
        source_info: 数据源信息，写入总览 Sheet
        time_window_days: 时间窗口（天）
        amount_tolerance: 金额容忍度（元）
        strict_counterparty_match: 是否严格双向对方户名匹配

    Returns:
        (生成的报告文件路径, 标记后的记录列表) -> 这里只返回报告路径，
        标记结果通过调用方自行从返回的 result 获取（请使用 identify_internal_transfers）
    """
    logger = get_logger()

    if not records:
        logger.warning('无交易记录，跳过内部划转识别')
        return None

    if output_dir is None:
        output_dir = get_script_dir()

    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'内部划转识别报告_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, filename)

    it_result = identify_internal_transfers(
        records,
        time_window_days=time_window_days,
        amount_tolerance=amount_tolerance,
        strict_counterparty_match=strict_counterparty_match,
    )

    if it_result.match_pairs == 0:
        logger.info('未识别到任何内部划转配对，跳过报告生成')
        return None

    source_info_final = dict(source_info or {})
    source_info_final.setdefault('数据来源', '主流程自动生成')
    source_info_final.setdefault('记录数', len(records))
    source_info_final.setdefault('生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S'))

    return export_internal_transfer_report(it_result, output_path, source_info_final)


def identify_and_tag_internal_transfers(
    records: List[Dict[str, Any]],
    time_window_days: int = 7,
    amount_tolerance: float = 0.01,
    strict_counterparty_match: bool = True,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any], InternalTransferResult]:
    """
    便捷函数：识别内部划转并返回 (标记后的记录, 摘要dict, 完整result对象)。

    在主流程中调用此函数：
        final_rows, tag_summary, it_result = identify_and_tag_internal_transfers(final_rows)

    Args:
        records: 原始记录列表
        time_window_days: 时间窗口
        amount_tolerance: 金额容忍度
        strict_counterparty_match: 是否严格双向匹配

    Returns:
        (marked_records, summary_dict, it_result)
    """
    it_result = identify_internal_transfers(
        records,
        time_window_days=time_window_days,
        amount_tolerance=amount_tolerance,
        strict_counterparty_match=strict_counterparty_match,
    )
    summary = {
        'total_records': it_result.total_records,
        'match_pairs': it_result.match_pairs,
        'marked_out_count': it_result.marked_out_count,
        'marked_in_count': it_result.marked_in_count,
        'involved_subjects': it_result.involved_subjects,
        'involved_banks': it_result.involved_banks,
        'total_transfer_amount': round(sum(m.amount for m in it_result.matches), 2),
    }
    return it_result.marked_records, summary, it_result


if __name__ == '__main__':
    main()
