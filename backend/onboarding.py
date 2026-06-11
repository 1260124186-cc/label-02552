# -*- coding: utf-8 -*-
"""
内置帮助与示例引导模块
功能：
  1. 首次启动检测
  2. 操作步骤展示
  3. 查找表格式说明
  4. 一键演示模式（使用 samples 示例跑通全流程）
"""

import os
import sys
import shutil
import tempfile
from typing import Optional, Tuple

try:
    from i18n import t, init_i18n
    HAS_I18N = True
except ImportError:
    HAS_I18N = False
    def t(key, **kwargs):
        return key
    def init_i18n(lang=None):
        return None

if HAS_I18N:
    init_i18n()


FIRST_RUN_MARKER = '.bankcheck_onboarding_complete'


def get_script_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def get_program_dir():
    return get_script_dir()


def get_user_data_dir():
    app_name = 'bankcheck'
    if sys.platform.startswith('win'):
        base_dir = os.environ.get('APPDATA')
        if not base_dir:
            base_dir = os.path.expanduser('~\\AppData\\Roaming')
        return os.path.join(base_dir, app_name)
    elif sys.platform == 'darwin':
        return os.path.join(os.path.expanduser('~/Library/Application Support'), app_name)
    else:
        return os.path.join(os.path.expanduser('~'), '.' + app_name)


def get_writable_dir():
    program_dir = get_program_dir()
    if os.access(program_dir, os.W_OK):
        return program_dir
    user_data_dir = get_user_data_dir()
    os.makedirs(user_data_dir, exist_ok=True)
    return user_data_dir


def is_first_run() -> bool:
    """检测是否为首次启动"""
    writable_dir = get_writable_dir()
    marker_path = os.path.join(writable_dir, FIRST_RUN_MARKER)
    return not os.path.exists(marker_path)


def mark_first_run_complete() -> bool:
    """标记首次启动已完成"""
    try:
        writable_dir = get_writable_dir()
        marker_path = os.path.join(writable_dir, FIRST_RUN_MARKER)
        with open(marker_path, 'w', encoding='utf-8') as f:
            f.write('onboarding_completed')
        return True
    except Exception:
        return False


def reset_first_run_marker() -> bool:
    """重置首次启动标记（用于测试）"""
    try:
        writable_dir = get_writable_dir()
        marker_path = os.path.join(writable_dir, FIRST_RUN_MARKER)
        if os.path.exists(marker_path):
            os.remove(marker_path)
        return True
    except Exception:
        return False


def _print_separator(char='=', length=70):
    print(char * length)


def _print_title(title):
    _print_separator('=')
    print(f'  {title}')
    _print_separator('=')
    print()


def show_welcome():
    """显示欢迎信息"""
    _print_title(t('onboarding.welcome_title'))
    print(t('onboarding.welcome_message'))
    print()
    print(t('onboarding.welcome_features'))
    print()


def show_operation_steps():
    """显示操作步骤说明"""
    _print_title(t('onboarding.steps_title'))

    steps = [
        (t('onboarding.step1_title'), t('onboarding.step1_desc')),
        (t('onboarding.step2_title'), t('onboarding.step2_desc')),
        (t('onboarding.step3_title'), t('onboarding.step3_desc')),
        (t('onboarding.step4_title'), t('onboarding.step4_desc')),
        (t('onboarding.step5_title'), t('onboarding.step5_desc')),
    ]

    for i, (title, desc) in enumerate(steps, 1):
        print(f'  {t("onboarding.step_prefix", num=i)} {title}')
        print(f'     {desc}')
        print()

    print(t('onboarding.steps_note'))
    print()


def show_lookup_table_format():
    """显示查找表格式说明"""
    _print_title(t('onboarding.lookup_title'))

    print(t('onboarding.lookup_intro'))
    print()

    print(t('onboarding.lookup_file_name'))
    print('  - 主体查找表.xlsx (recommended)')
    print('  - 主体查找表.xls')
    print()

    print(t('onboarding.lookup_table_structure'))
    print()

    header = f'  | {"A列(主体名称)":<25} | {"B列(银行账号)":<25} | {"C列(优先级,可选)":<15} | {"D列...(扩展字段,可选)":<20} |'
    separator = '  |' + '-' * 27 + '|' + '-' * 27 + '|' + '-' * 17 + '|' + '-' * 22 + '|'

    print(separator)
    print(header)
    print(separator)

    sample_rows = [
        ('北京XX科技有限公司', '01090312345678901', '1', '北京分公司'),
        ('上海YY贸易有限公司', '38812345678', '0', '上海分公司'),
        ('深圳ZZ科技有限公司', '6222021234567890123', '0', ''),
    ]

    for subject, account, priority, extra in sample_rows:
        print(f'  | {subject:<25} | {account:<25} | {priority:<15} | {extra:<20} |')

    print(separator)
    print()

    print(t('onboarding.lookup_columns_desc'))
    print()

    columns = [
        (t('onboarding.col_subject'), t('onboarding.col_subject_desc')),
        (t('onboarding.col_account'), t('onboarding.col_account_desc')),
        (t('onboarding.col_priority'), t('onboarding.col_priority_desc')),
        (t('onboarding.col_extra'), t('onboarding.col_extra_desc')),
    ]

    for col_name, col_desc in columns:
        print(f'  • {col_name}: {col_desc}')
    print()

    print(t('onboarding.lookup_rules'))
    print()

    rules = [
        t('onboarding.rule1'),
        t('onboarding.rule2'),
        t('onboarding.rule3'),
        t('onboarding.rule4'),
    ]

    for i, rule in enumerate(rules, 1):
        print(f'  {i}. {rule}')
    print()


def show_demo_intro():
    """显示演示模式介绍"""
    _print_title(t('onboarding.demo_title'))
    print(t('onboarding.demo_intro'))
    print()
    print(t('onboarding.demo_what_includes'))
    print()

    includes = [
        t('onboarding.demo_include1'),
        t('onboarding.demo_include2'),
        t('onboarding.demo_include3'),
        t('onboarding.demo_include4'),
    ]

    for item in includes:
        print(f'  ✓ {item}')
    print()


def show_help_menu():
    """显示帮助菜单"""
    _print_title(t('onboarding.help_menu_title'))
    print(t('onboarding.help_menu_intro'))
    print()
    print(t('onboarding.help_option1'))
    print(t('onboarding.help_option2'))
    print(t('onboarding.help_option3'))
    print(t('onboarding.help_option4'))
    print(t('onboarding.help_option5'))
    print()


def create_demo_lookup_table(output_path: str) -> str:
    """创建演示用的查找表"""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '主体映射'

    ws['A1'] = '主体名称'
    ws['B1'] = '银行账号'
    ws['C1'] = '优先级'
    ws['D1'] = '备注'

    mappings = [
        ('北京XX科技有限公司', '01090312345678901', 1, '北京分公司'),
        ('上海YY贸易有限公司', '38812345678', 0, '上海分公司'),
    ]

    for i, (subject, account, priority, remark) in enumerate(mappings, 2):
        ws.cell(row=i, column=1, value=subject)
        ws.cell(row=i, column=2, value=account)
        ws.cell(row=i, column=3, value=priority)
        ws.cell(row=i, column=4, value=remark)

    wb.save(output_path)
    wb.close()
    return output_path


def prepare_demo_environment(script_dir: str) -> Tuple[Optional[str], Optional[str], str]:
    """
    准备演示环境
    Returns: (demo_folder, demo_lookup_path, temp_dir)
    """
    import bankcheck

    samples_dir = os.path.join(os.path.dirname(script_dir), 'samples')

    if not os.path.exists(samples_dir):
        return None, None, ''

    temp_dir = tempfile.mkdtemp(prefix='bankcheck_demo_')
    demo_folder = os.path.join(temp_dir, 'demo_bank_data')
    os.makedirs(demo_folder, exist_ok=True)

    sample_files = os.listdir(samples_dir)
    for f in sample_files:
        if f.endswith(('.xlsx', '.xls')) and not f.startswith('~$'):
            src = os.path.join(samples_dir, f)
            dst = os.path.join(demo_folder, f)
            shutil.copy2(src, dst)

    demo_lookup_path = os.path.join(temp_dir, '主体查找表.xlsx')
    create_demo_lookup_table(demo_lookup_path)

    return demo_folder, demo_lookup_path, temp_dir


def run_demo_mode(script_dir: str) -> bool:
    """
    运行演示模式 - 使用 samples 示例跑通全流程
    Returns: 是否成功
    """
    import bankcheck

    _print_title(t('onboarding.demo_running'))

    demo_folder, demo_lookup_path, temp_dir = prepare_demo_environment(script_dir)

    if demo_folder is None:
        print(t('onboarding.demo_no_samples'))
        return False

    print(t('onboarding.demo_prepared'))
    print(f'  {t("onboarding.demo_folder")}: {demo_folder}')
    print(f'  {t("onboarding.demo_lookup")}: {demo_lookup_path}')
    print()

    demo_script_dir = os.path.dirname(demo_lookup_path)
    demo_lookup_target = os.path.join(demo_script_dir, '主体查找表.xlsx')
    if demo_lookup_path != demo_lookup_target:
        shutil.copy2(demo_lookup_path, demo_lookup_target)

    try:
        print(t('onboarding.demo_processing'))
        _print_separator('-')

        result = bankcheck.run_pipeline(demo_folder, demo_script_dir, incremental=False)

        _print_separator('-')
        print()

        if result.folder_empty:
            print(t('onboarding.demo_result_empty'))
            return False

        if len(result.all_rows) == 0:
            print(t('onboarding.demo_result_no_records'))
            return False

        _print_title(t('onboarding.demo_result_title'))

        print(t('onboarding.demo_result_stats',
                processed=len(result.processed_files),
                records=len(result.all_rows),
                unprocessed=len(result.unprocessed_files)))
        print()

        print(t('onboarding.demo_result_data_preview'))
        print()

        header = f'  | {"序号":<4} | {"银行":<8} | {"主体":<20} | {"交易日期":<12} | {"付款(元)":>12} | {"收款(元)":>12} |'
        separator = '  |' + '-' * 6 + '|' + '-' * 10 + '|' + '-' * 22 + '|' + '-' * 14 + '|' + '-' * 14 + '|' + '-' * 14 + '|'

        print(separator)
        print(header)
        print(separator)

        for i, row in enumerate(result.all_rows[:10], 1):
            bank = str(row.get('银行', ''))[:8]
            subject = str(row.get('主体', ''))
            if len(subject) > 18:
                subject = subject[:16] + '...'
            else:
                subject = subject.ljust(20)
            date = str(row.get('交易日期', ''))[:10]
            payment = row.get('付款', '') or ''
            receipt = row.get('收款', '') or ''

            payment_str = f'{payment:>12,.2f}' if payment else ' ' * 12
            receipt_str = f'{receipt:>12,.2f}' if receipt else ' ' * 12

            print(f'  | {i:<4} | {bank:<8} | {subject:<20} | {date:<12} | {payment_str} | {receipt_str} |')

        print(separator)
        print()

        if result.output_path:
            print(t('onboarding.demo_result_output', path=result.output_path))
            print()

        print(t('onboarding.demo_success'))
        print()

        return True

    except Exception as e:
        print(t('onboarding.demo_error', error=str(e)))
        return False
    finally:
        try:
            shutil.rmtree(temp_dir, ignore_errors=True)
        except Exception:
            pass


def show_onboarding_menu() -> str:
    """
    显示首次启动引导菜单
    Returns: 用户选择 ('demo', 'steps', 'lookup', 'help', 'skip', 'exit')
    """
    show_welcome()

    print(t('onboarding.menu_intro'))
    print()
    print(t('onboarding.menu_option1'))
    print(t('onboarding.menu_option2'))
    print(t('onboarding.menu_option3'))
    print(t('onboarding.menu_option4'))
    print(t('onboarding.menu_option5'))
    print(t('onboarding.menu_option6'))
    print()

    choice = input(t('onboarding.menu_prompt')).strip().lower()

    if choice == '1':
        return 'demo'
    elif choice == '2':
        return 'steps'
    elif choice == '3':
        return 'lookup'
    elif choice == '4':
        return 'help'
    elif choice == '5':
        return 'skip'
    elif choice in ['6', 'q', 'quit', 'exit']:
        return 'exit'
    elif choice == '':
        return 'demo'
    else:
        return 'help'


def run_onboarding_flow(script_dir: str) -> str:
    """
    运行完整的首次启动引导流程
    Returns: 最终选择 ('continue' | 'exit')
    """
    if not is_first_run():
        return 'continue'

    while True:
        choice = show_onboarding_menu()

        if choice == 'demo':
            show_demo_intro()
            confirm = input(t('onboarding.demo_confirm')).strip().lower()
            if confirm in ['y', 'yes', '']:
                run_demo_mode(script_dir)
                print(t('onboarding.press_enter'))
                input()
        elif choice == 'steps':
            show_operation_steps()
            print(t('onboarding.press_enter'))
            input()
        elif choice == 'lookup':
            show_lookup_table_format()
            print(t('onboarding.press_enter'))
            input()
        elif choice == 'help':
            show_help_menu()
            print(t('onboarding.press_enter'))
            input()
        elif choice == 'skip':
            confirm = input(t('onboarding.skip_confirm')).strip().lower()
            if confirm in ['y', 'yes', '']:
                mark_first_run_complete()
                return 'continue'
        elif choice == 'exit':
            return 'exit'

    return 'continue'
