# -*- coding: utf-8 -*-
"""
主体查找表管理模块
提供银行账号与主体映射关系的增删改查、导入导出功能
"""

import os
import sys
import logging
from typing import List, Dict, Optional, Tuple
from dataclasses import dataclass

import openpyxl
import pandas as pd


def get_script_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def get_program_dir():
    return get_script_dir()


def is_writable(dir_path):
    import uuid
    if not os.path.isdir(dir_path):
        return False
    try:
        test_file = os.path.join(dir_path, '.lookup_write_test_' + uuid.uuid4().hex[:8])
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)
        return True
    except (OSError, IOError):
        return False


def get_user_data_dir():
    app_name = 'bankcheck'
    if sys.platform.startswith('win'):
        base_dir = os.environ.get('APPDATA')
        if not base_dir:
            base_dir = os.path.expanduser('~\\AppData\\Roaming')
        return os.path.join(base_dir, app_name)
    elif sys.platform == 'darwin':
        return os.path.join(os.path.expanduser('~/Library/Application Support'), app_name)
    else:
        return os.path.join(os.path.expanduser('~'), '.' + app_name)


def get_writable_dir():
    program_dir = get_program_dir()
    if is_writable(program_dir):
        return program_dir
    user_data_dir = get_user_data_dir()
    os.makedirs(user_data_dir, exist_ok=True)
    return user_data_dir


def get_output_dir(subdir=None):
    base_dir = get_writable_dir()
    if subdir:
        output_dir = os.path.join(base_dir, subdir)
    else:
        output_dir = base_dir
    os.makedirs(output_dir, exist_ok=True)
    return output_dir


def get_logger():
    return logging.getLogger('bankcheck')


LOOKUP_FILE_NAMES = ['主体查找表.xlsx', '主体查找表.xls']


def _find_lookup_in_dir(directory):
    if not directory or not os.path.isdir(directory):
        return None
    for name in LOOKUP_FILE_NAMES:
        candidate = os.path.join(directory, name)
        if os.path.isfile(candidate):
            return candidate
    excel_exts = ('.xlsx', '.xls')
    exclude_names = {'银行流水总表.xlsx', '银行流水总表.xls'}
    excel_files = []
    try:
        for f in os.listdir(directory):
            if f.startswith('~$'):
                continue
            if f in exclude_names:
                continue
            if f.lower().endswith(excel_exts):
                excel_files.append(os.path.join(directory, f))
    except OSError:
        return None
    if len(excel_files) == 1:
        return excel_files[0]
    return None


def _normalize_account_str(value):
    if value is None:
        return ''
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value != value:
            return ''
        if value == int(value):
            return str(int(value))
        return str(value)
    s = str(value).strip()
    if '.' in s:
        try:
            f = float(s)
            if f == int(f):
                return str(int(f))
        except (ValueError, TypeError, OverflowError):
            pass
    return s


def _account_key(value):
    normalized = _normalize_account_str(value)
    return normalized.lstrip('0') or '0'


@dataclass
class LookupEntry:
    """查找表条目：主体名称 + 银行账号"""
    subject: str
    account: str
    row_id: Optional[int] = None

    def to_dict(self) -> Dict[str, any]:
        return {
            'row_id': self.row_id,
            'subject': self.subject,
            'account': self.account,
        }


def _copy_lookup_to_output(program_lookup_path, output_dir):
    """将程序目录下的查找表复制到用户可写的输出目录"""
    logger = get_logger()
    if not program_lookup_path or not os.path.isfile(program_lookup_path):
        return None
    try:
        filename = os.path.basename(program_lookup_path)
        target_path = os.path.join(output_dir, filename)
        if os.path.exists(target_path):
            return target_path
        os.makedirs(output_dir, exist_ok=True)
        import shutil
        shutil.copy2(program_lookup_path, target_path)
        logger.info('已将查找表从程序目录复制到输出目录: %s -> %s',
                    program_lookup_path, target_path)
        return target_path
    except Exception as e:
        logger.warning('复制查找表到输出目录失败: %s', e)
        return None


def find_lookup_file(script_dir=None) -> Optional[str]:
    """
    智能查找主体查找表 Excel 文件。

    查找策略（按优先级）：
    1. 如果传入了 script_dir 参数，只在该目录查找（向后兼容）
    2. 否则优先在输出目录（可写目录）查找
    3. 如果在输出目录没找到，在程序目录查找
    4. 如果在程序目录找到但输出目录没找到，自动复制到输出目录
    """
    logger = get_logger()
    output_dir = get_output_dir()
    program_dir = get_program_dir()

    if script_dir is not None:
        custom_lookup = _find_lookup_in_dir(script_dir)
        if custom_lookup:
            logger.info('在指定目录找到主体查找表: %s', custom_lookup)
            return custom_lookup
        logger.warning('在指定目录未找到主体查找表: %s', script_dir)
        return None

    output_lookup = _find_lookup_in_dir(output_dir)
    if output_lookup:
        logger.info('在输出目录找到主体查找表: %s', output_lookup)
        return output_lookup

    program_lookup = _find_lookup_in_dir(program_dir)
    if program_lookup:
        logger.info('在程序目录找到主体查找表: %s', program_lookup)
        copied_path = _copy_lookup_to_output(program_lookup, output_dir)
        if copied_path:
            return copied_path
        return program_lookup

    logger.warning('未找到主体查找表文件')
    return None


def get_lookup_file_path(script_dir=None) -> str:
    """
    获取查找表文件路径，如果不存在则返回默认路径（主体查找表.xlsx）
    默认使用可写目录。
    """
    existing = find_lookup_file(script_dir)
    if existing:
        return existing

    output_dir = get_output_dir()
    return os.path.join(output_dir, '主体查找表.xlsx')


def _open_workbook_compat(filepath):
    """兼容打开 .xls 和 .xlsx 文件"""
    tmp_path = None
    if filepath.lower().endswith('.xls'):
        import xlrd
        import tempfile
        xls_book = xlrd.open_workbook(filepath)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '主体映射'
        for row in range(xls_book.sheet_by_index(0).nrows):
            for col in range(xls_book.sheet_by_index(0).ncols):
                cell_value = xls_book.sheet_by_index(0).cell_value(row, col)
                ws.cell(row=row + 1, column=col + 1, value=cell_value)
        tmp_path = tempfile.mktemp(suffix='.xlsx')
        wb.save(tmp_path)
        wb.close()
        wb = openpyxl.load_workbook(tmp_path)
    else:
        wb = openpyxl.load_workbook(filepath)
    return wb, tmp_path


def _cleanup_temp_file(tmp_path):
    if tmp_path and os.path.exists(tmp_path):
        try:
            os.remove(tmp_path)
        except Exception:
            pass


def read_lookup_entries(lookup_file=None) -> List[LookupEntry]:
    """
    读取查找表中的所有条目

    Args:
        lookup_file: 查找表文件路径，不传则自动查找

    Returns:
        LookupEntry 列表
    """
    logger = get_logger()

    if lookup_file is None:
        lookup_file = find_lookup_file()

    if lookup_file is None or not os.path.exists(lookup_file):
        logger.warning('查找表文件不存在: %s', lookup_file)
        return []

    entries = []
    tmp_path = None
    try:
        wb, tmp_path = _open_workbook_compat(lookup_file)
        ws = wb.active

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row,
                                                   min_col=1, max_col=2), start=2):
            subject_cell = row[0]
            account_cell = row[1]

            subject = subject_cell.value
            account = account_cell.value

            if subject is None and account is None:
                continue

            subject_str = str(subject).strip() if subject is not None else ''
            account_str = _normalize_account_str(account)

            if not account_str:
                continue

            entries.append(LookupEntry(
                subject=subject_str,
                account=account_str,
                row_id=row_idx
            ))

        wb.close()
        logger.info('从查找表读取到 %d 条记录', len(entries))
    except Exception as e:
        logger.error('读取查找表失败: %s', e, exc_info=True)
    finally:
        _cleanup_temp_file(tmp_path)

    return entries


def save_lookup_entries(entries: List[LookupEntry], lookup_file=None) -> bool:
    """
    保存条目列表到查找表
    完全覆盖原有内容，按现有格式写入（A列=主体，B列=账号，表头为"主体名称"和"银行账号"）

    Args:
        entries: 要保存的条目列表
        lookup_file: 查找表文件路径，不传则使用默认路径

    Returns:
        是否保存成功
    """
    logger = get_logger()

    if lookup_file is None:
        lookup_file = get_lookup_file_path()

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '主体映射'

        ws.cell(row=1, column=1, value='主体名称')
        ws.cell(row=1, column=2, value='银行账号')

        for i, entry in enumerate(entries, start=2):
            ws.cell(row=i, column=1, value=entry.subject)
            ws.cell(row=i, column=2, value=entry.account)

        wb.save(lookup_file)
        wb.close()
        logger.info('查找表已保存到 %s，共 %d 条记录', lookup_file, len(entries))
        return True
    except Exception as e:
        logger.error('保存查找表失败: %s', e, exc_info=True)
        return False


def _load_entries_with_duplicate_check() -> Tuple[List[LookupEntry], Dict[str, List[int]]]:
    """
    加载条目并检测重复账号
    Returns: (条目列表, 重复账号字典 {账号: [row_id列表]})
    """
    entries = read_lookup_entries()
    account_map: Dict[str, List[int]] = {}

    for idx, entry in enumerate(entries):
        key = _account_key(entry.account)
        if key not in account_map:
            account_map[key] = []
        account_map[key].append(idx)

    duplicates = {k: v for k, v in account_map.items() if len(v) > 1}
    return entries, duplicates


def get_entry_by_account(account: str, lookup_file=None) -> Optional[LookupEntry]:
    """
    根据账号查找条目

    Args:
        account: 银行账号
        lookup_file: 查找表文件路径

    Returns:
        找到的条目，未找到返回 None
    """
    entries = read_lookup_entries(lookup_file)
    target_key = _account_key(account)

    for entry in entries:
        if _account_key(entry.account) == target_key:
            return entry

    return None


def add_entry(subject: str, account: str, lookup_file=None) -> Tuple[bool, str]:
    """
    添加新条目

    Args:
        subject: 主体名称
        account: 银行账号
        lookup_file: 查找表文件路径

    Returns:
        (是否成功, 消息)
    """
    logger = get_logger()

    if not subject or not account:
        return False, '主体名称和银行账号不能为空'

    account_normalized = _normalize_account_str(account)
    if not account_normalized:
        return False, '银行账号格式无效'

    existing = get_entry_by_account(account, lookup_file)
    if existing is not None:
        return False, f'银行账号 {account} 已存在，对应主体: {existing.subject}'

    entries = read_lookup_entries(lookup_file)
    entries.append(LookupEntry(
        subject=subject.strip(),
        account=account_normalized
    ))

    success = save_lookup_entries(entries, lookup_file)
    if success:
        logger.info('已添加条目: %s -> %s', account_normalized, subject.strip())
        return True, '添加成功'
    return False, '保存失败'


def update_entry(old_account: str, new_subject: str, new_account: str,
                 lookup_file=None) -> Tuple[bool, str]:
    """
    更新条目

    Args:
        old_account: 原银行账号（用于定位）
        new_subject: 新的主体名称
        new_account: 新的银行账号
        lookup_file: 查找表文件路径

    Returns:
        (是否成功, 消息)
    """
    logger = get_logger()

    if not new_subject or not new_account:
        return False, '主体名称和银行账号不能为空'

    new_account_normalized = _normalize_account_str(new_account)
    if not new_account_normalized:
        return False, '银行账号格式无效'

    entries = read_lookup_entries(lookup_file)
    old_key = _account_key(old_account)
    new_key = _account_key(new_account_normalized)

    found_idx = None
    for idx, entry in enumerate(entries):
        if _account_key(entry.account) == old_key:
            found_idx = idx
            break

    if found_idx is None:
        return False, f'未找到账号 {old_account} 对应的条目'

    if old_key != new_key:
        for idx, entry in enumerate(entries):
            if idx != found_idx and _account_key(entry.account) == new_key:
                return False, f'新账号 {new_account} 已被使用，对应主体: {entry.subject}'

    entries[found_idx].subject = new_subject.strip()
    entries[found_idx].account = new_account_normalized

    success = save_lookup_entries(entries, lookup_file)
    if success:
        logger.info('已更新条目: %s -> %s (原账号: %s)',
                    new_account_normalized, new_subject.strip(), old_account)
        return True, '更新成功'
    return False, '保存失败'


def delete_entry(account: str, lookup_file=None) -> Tuple[bool, str]:
    """
    删除条目

    Args:
        account: 要删除的银行账号
        lookup_file: 查找表文件路径

    Returns:
        (是否成功, 消息)
    """
    logger = get_logger()

    entries = read_lookup_entries(lookup_file)
    target_key = _account_key(account)

    original_len = len(entries)
    entries = [e for e in entries if _account_key(e.account) != target_key]

    if len(entries) == original_len:
        return False, f'未找到账号 {account} 对应的条目'

    success = save_lookup_entries(entries, lookup_file)
    if success:
        deleted_count = original_len - len(entries)
        logger.info('已删除 %d 条条目，账号: %s', deleted_count, account)
        return True, f'成功删除 {deleted_count} 条记录'
    return False, '保存失败'


def import_from_excel(import_file: str, overwrite: bool = False,
                      lookup_file=None) -> Tuple[bool, str, Dict]:
    """
    从 Excel 文件导入条目

    Args:
        import_file: 要导入的 Excel 文件路径
        overwrite: 是否覆盖现有数据（True=全量覆盖，False=增量合并）
        lookup_file: 查找表文件路径

    Returns:
        (是否成功, 消息, 统计信息 {imported, updated, skipped, total})
    """
    logger = get_logger()

    if not os.path.exists(import_file):
        return False, f'导入文件不存在: {import_file}', {}

    stats = {'imported': 0, 'updated': 0, 'skipped': 0, 'total': 0}

    try:
        df = pd.read_excel(import_file, header=None)

        imported_entries: List[LookupEntry] = []
        header_keywords = {'主体名称', '银行账号', '主体', '账号', 'account', 'subject'}

        for row_idx, (_, row) in enumerate(df.iterrows()):
            subject = row.iloc[0] if len(row) > 0 else None
            account = row.iloc[1] if len(row) > 1 else None

            if subject is None and account is None:
                continue

            subject_str = str(subject).strip() if subject is not None else ''
            account_str = _normalize_account_str(account)

            if row_idx == 0:
                first_col = subject_str.lower()
                second_col = str(account).strip().lower() if account is not None else ''
                if (first_col in header_keywords or second_col in header_keywords):
                    logger.debug('检测到表头行，已跳过: %s, %s', subject_str, account_str)
                    continue

            if not account_str:
                stats['skipped'] += 1
                continue

            imported_entries.append(LookupEntry(
                subject=subject_str,
                account=account_str
            ))

        stats['total'] = len(imported_entries)

        if overwrite:
            success = save_lookup_entries(imported_entries, lookup_file)
            if success:
                stats['imported'] = len(imported_entries)
                return True, f'全量导入成功，共 {len(imported_entries)} 条记录', stats
            return False, '保存失败', stats

        existing_entries = read_lookup_entries(lookup_file)
        existing_map = {_account_key(e.account): idx for idx, e in enumerate(existing_entries)}

        for entry in imported_entries:
            key = _account_key(entry.account)
            if key in existing_map:
                idx = existing_map[key]
                if existing_entries[idx].subject != entry.subject:
                    existing_entries[idx].subject = entry.subject
                    stats['updated'] += 1
                else:
                    stats['skipped'] += 1
            else:
                existing_entries.append(entry)
                stats['imported'] += 1

        success = save_lookup_entries(existing_entries, lookup_file)
        if success:
            return True, (f'增量导入完成：新增 {stats["imported"]} 条，'
                         f'更新 {stats["updated"]} 条，跳过 {stats["skipped"]} 条'), stats
        return False, '保存失败', stats

    except Exception as e:
        logger.error('导入 Excel 失败: %s', e, exc_info=True)
        return False, f'导入失败: {str(e)}', stats


def export_to_excel(export_file: str, lookup_file=None) -> Tuple[bool, str]:
    """
    导出查找表到 Excel 文件

    Args:
        export_file: 导出文件路径
        lookup_file: 查找表文件路径

    Returns:
        (是否成功, 消息)
    """
    logger = get_logger()

    entries = read_lookup_entries(lookup_file)

    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '主体映射'

        ws.cell(row=1, column=1, value='主体名称')
        ws.cell(row=1, column=2, value='银行账号')

        for i, entry in enumerate(entries, start=2):
            ws.cell(row=i, column=1, value=entry.subject)
            ws.cell(row=i, column=2, value=entry.account)

        wb.save(export_file)
        wb.close()
        logger.info('查找表已导出到 %s，共 %d 条记录', export_file, len(entries))
        return True, f'导出成功，共 {len(entries)} 条记录'
    except Exception as e:
        logger.error('导出 Excel 失败: %s', e, exc_info=True)
        return False, f'导出失败: {str(e)}'


def search_entries(keyword: str = '', lookup_file=None) -> List[LookupEntry]:
    """
    搜索条目（按主体名称或账号模糊匹配）

    Args:
        keyword: 搜索关键词
        lookup_file: 查找表文件路径

    Returns:
        匹配的条目列表
    """
    entries = read_lookup_entries(lookup_file)

    if not keyword:
        return entries

    keyword = keyword.strip().lower()
    results = []

    for entry in entries:
        if (keyword in entry.subject.lower() or
                keyword in entry.account.lower() or
                keyword in _normalize_account_str(entry.account)):
            results.append(entry)

    return results


def get_duplicate_entries(lookup_file=None) -> List[Dict]:
    """
    获取重复的账号条目

    Args:
        lookup_file: 查找表文件路径

    Returns:
        重复条目列表，每项包含 {'account': 账号, 'subjects': [主体列表], 'count': 数量}
    """
    entries = read_lookup_entries(lookup_file)
    account_map: Dict[str, List[str]] = {}

    for entry in entries:
        key = _account_key(entry.account)
        if key not in account_map:
            account_map[key] = []
        account_map[key].append(entry.subject)

    duplicates = []
    for account, subjects in account_map.items():
        if len(subjects) > 1:
            duplicates.append({
                'account': account,
                'subjects': list(set(subjects)),
                'count': len(subjects)
            })

    return duplicates
