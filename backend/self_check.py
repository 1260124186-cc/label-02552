# -*- coding: utf-8 -*-
"""
程序启动自检模块
在程序启动时执行依赖库、查找表可读性、临时目录可写性等检查，
自检失败时给出明确修复指引，而非运行中途才报错。
"""

import os
import sys
import tempfile
import uuid
from dataclasses import dataclass, field
from typing import List, Optional, Tuple


REQUIRED_DEPENDENCIES = [
    {
        'name': 'openpyxl',
        'import_name': 'openpyxl',
        'purpose': 'Excel 文件读写（.xlsx 格式）',
        'fix_guide': 'pip install openpyxl>=3.1.0',
    },
    {
        'name': 'pandas',
        'import_name': 'pandas',
        'purpose': '数据处理与分析',
        'fix_guide': 'pip install pandas>=2.0.0',
    },
    {
        'name': 'PyYAML',
        'import_name': 'yaml',
        'purpose': 'YAML 配置文件解析',
        'fix_guide': 'pip install PyYAML>=6.0',
    },
    {
        'name': 'xlrd',
        'import_name': 'xlrd',
        'purpose': '旧版 Excel 文件读取（.xls 格式）',
        'fix_guide': 'pip install xlrd>=2.0.0',
    },
]

OPTIONAL_DEPENDENCIES = [
    {
        'name': 'Flask',
        'import_name': 'flask',
        'purpose': 'Web 服务（查找表管理、上传处理）',
        'fix_guide': 'pip install Flask>=3.0.0',
    },
    {
        'name': 'cryptography',
        'import_name': 'cryptography',
        'purpose': '文件加密与数字签名',
        'fix_guide': 'pip install cryptography>=42.0.0',
    },
    {
        'name': 'pdfplumber',
        'import_name': 'pdfplumber',
        'purpose': 'PDF 银行流水解析（可选）',
        'fix_guide': 'pip install pdfplumber>=0.10.0',
    },
    {
        'name': 'Pillow',
        'import_name': 'PIL',
        'purpose': '图片处理（OCR 辅助，可选）',
        'fix_guide': 'pip install Pillow>=10.0.0',
    },
    {
        'name': 'pytesseract',
        'import_name': 'pytesseract',
        'purpose': 'OCR 文字识别（可选，需额外安装 Tesseract）',
        'fix_guide': 'pip install pytesseract>=0.3.10，并安装 Tesseract OCR 引擎',
    },
    {
        'name': 'APScheduler',
        'import_name': 'apscheduler',
        'purpose': '定时任务调度（可选）',
        'fix_guide': 'pip install APScheduler>=3.10.0',
    },
    {
        'name': 'msoffcrypto-tool',
        'import_name': 'msoffcrypto',
        'purpose': '加密 Excel 文件解密（可选）',
        'fix_guide': 'pip install msoffcrypto-python>=4.12.0',
    },
]


@dataclass
class CheckResult:
    """单项检查结果"""
    name: str
    passed: bool
    severity: str = 'error'
    message: str = ''
    fix_guide: str = ''
    details: str = ''


@dataclass
class SelfCheckReport:
    """自检总报告"""
    passed: bool = True
    results: List[CheckResult] = field(default_factory=list)
    errors: List[CheckResult] = field(default_factory=list)
    warnings: List[CheckResult] = field(default_factory=list)

    def add_result(self, result: CheckResult):
        self.results.append(result)
        if not result.passed:
            if result.severity == 'warning':
                self.warnings.append(result)
            else:
                self.errors.append(result)
                self.passed = False


def _check_dependency(dep_info: dict, severity: str = 'error') -> CheckResult:
    """检查单个依赖库"""
    name = dep_info['name']
    import_name = dep_info['import_name']
    purpose = dep_info['purpose']
    fix_guide = dep_info['fix_guide']

    try:
        __import__(import_name)
        return CheckResult(
            name=name,
            passed=True,
            severity=severity,
            message=f'{name} 已安装，{purpose}',
            fix_guide='',
        )
    except ImportError as e:
        return CheckResult(
            name=name,
            passed=False,
            severity=severity,
            message=f'{name} 未安装：{purpose}',
            fix_guide=fix_guide,
            details=str(e),
        )


def check_required_dependencies() -> List[CheckResult]:
    """检查必需的依赖库"""
    results = []
    for dep in REQUIRED_DEPENDENCIES:
        results.append(_check_dependency(dep, severity='error'))
    return results


def check_optional_dependencies() -> List[CheckResult]:
    """检查可选的依赖库（仅警告，不阻断启动）"""
    results = []
    for dep in OPTIONAL_DEPENDENCIES:
        results.append(_check_dependency(dep, severity='warning'))
    return results


def _get_script_dir() -> str:
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def _find_lookup_file(script_dir: Optional[str] = None) -> Optional[str]:
    """查找主体查找表文件"""
    if script_dir is None:
        script_dir = _get_script_dir()

    lookup_names = ['主体查找表.xlsx', '主体查找表.xls']

    candidates = []

    output_dir = os.path.join(script_dir, 'output')
    candidates.extend([os.path.join(output_dir, n) for n in lookup_names])

    candidates.extend([os.path.join(script_dir, n) for n in lookup_names])

    try:
        if sys.platform.startswith('win'):
            appdata = os.environ.get('APPDATA', os.path.expanduser('~'))
            user_dir = os.path.join(appdata, 'bankcheck')
        elif sys.platform == 'darwin':
            user_dir = os.path.join(os.path.expanduser('~/Library/Application Support'), 'bankcheck')
        else:
            user_dir = os.path.join(os.path.expanduser('~'), '.bankcheck')
        candidates.extend([os.path.join(user_dir, n) for n in lookup_names])
    except Exception:
        pass

    for path in candidates:
        if os.path.isfile(path):
            return path

    return None


def check_lookup_table() -> CheckResult:
    """检查查找表可读性"""
    lookup_path = _find_lookup_file()

    if lookup_path is None:
        return CheckResult(
            name='查找表文件',
            passed=False,
            severity='warning',
            message='未找到主体查找表文件（主体查找表.xlsx / 主体查找表.xls）',
            fix_guide=(
                '请在程序目录下放置「主体查找表.xlsx」文件，'
                '包含至少两列：主体名称、银行账号。'
                '首次使用可通过 Web 管理界面创建。'
            ),
            details='查找位置：程序目录、output 目录、用户数据目录',
        )

    if not os.access(lookup_path, os.R_OK):
        return CheckResult(
            name='查找表文件',
            passed=False,
            severity='error',
            message=f'查找表文件不可读：{lookup_path}',
            fix_guide='请检查文件权限，确保当前用户拥有读取权限',
        )

    try:
        import openpyxl
        wb = openpyxl.load_workbook(lookup_path, read_only=True)
        ws = wb.active
        row_count = ws.max_row
        col_count = ws.max_column
        wb.close()

        if row_count < 2:
            return CheckResult(
                name='查找表文件',
                passed=False,
                severity='warning',
                message=f'查找表为空或只有表头：{lookup_path}',
                fix_guide='请在查找表中添加主体名称和银行账号数据',
                details=f'行数：{row_count}，列数：{col_count}',
            )

        return CheckResult(
            name='查找表文件',
            passed=True,
            severity='warning',
            message=f'查找表可读：{lookup_path}（{row_count - 1} 条记录）',
            fix_guide='',
            details=f'路径：{lookup_path}，行数：{row_count}，列数：{col_count}',
        )
    except ImportError:
        return CheckResult(
            name='查找表文件',
            passed=False,
            severity='error',
            message='无法解析查找表：缺少 openpyxl 依赖',
            fix_guide='pip install openpyxl>=3.1.0',
        )
    except Exception as e:
        return CheckResult(
            name='查找表文件',
            passed=False,
            severity='error',
            message=f'查找表读取失败：{lookup_path}',
            fix_guide='请检查文件是否损坏，或尝试用 Excel 打开后重新保存',
            details=str(e),
        )


def check_temp_directory() -> CheckResult:
    """检查临时目录可写性"""
    tmp_dir = tempfile.gettempdir()

    if not os.path.isdir(tmp_dir):
        return CheckResult(
            name='临时目录',
            passed=False,
            severity='error',
            message=f'临时目录不存在：{tmp_dir}',
            fix_guide='请检查系统临时目录配置，或设置 TMPDIR / TEMP 环境变量',
        )

    if not os.access(tmp_dir, os.W_OK):
        return CheckResult(
            name='临时目录',
            passed=False,
            severity='error',
            message=f'临时目录不可写：{tmp_dir}',
            fix_guide='请检查临时目录权限，或设置 TMPDIR 环境变量指向可写目录',
        )

    test_file = os.path.join(tmp_dir, f'.bankcheck_self_test_{uuid.uuid4().hex[:8]}')
    try:
        with open(test_file, 'w') as f:
            f.write('self_check_test')
        with open(test_file, 'r') as f:
            content = f.read()
        os.remove(test_file)

        if content != 'self_check_test':
            raise ValueError('写入内容与读取内容不一致')

        try:
            fd, tmp_path = tempfile.mkstemp(prefix='bankcheck_test_', suffix='.tmp')
            os.close(fd)
            os.remove(tmp_path)
        except Exception as e:
            return CheckResult(
                name='临时目录',
                passed=False,
                severity='error',
                message=f'临时目录创建临时文件失败：{tmp_dir}',
                fix_guide='请检查临时目录权限',
                details=str(e),
            )

        return CheckResult(
            name='临时目录',
            passed=True,
            severity='error',
            message=f'临时目录可写：{tmp_dir}',
            fix_guide='',
        )
    except Exception as e:
        if os.path.exists(test_file):
            try:
                os.remove(test_file)
            except Exception:
                pass
        return CheckResult(
            name='临时目录',
            passed=False,
            severity='error',
            message=f'临时目录读写测试失败：{tmp_dir}',
            fix_guide='请检查临时目录权限，或设置 TMPDIR 环境变量指向可写目录',
            details=str(e),
        )


def check_output_directory(script_dir: Optional[str] = None) -> CheckResult:
    """检查输出目录可写性"""
    if script_dir is None:
        script_dir = _get_script_dir()

    if os.access(script_dir, os.W_OK):
        output_dir = os.path.join(script_dir, 'output')
        os.makedirs(output_dir, exist_ok=True)
        if os.access(output_dir, os.W_OK):
            return CheckResult(
                name='输出目录',
                passed=True,
                severity='error',
                message=f'输出目录可写：{output_dir}',
                fix_guide='',
            )

    try:
        if sys.platform.startswith('win'):
            appdata = os.environ.get('APPDATA', os.path.expanduser('~'))
            user_dir = os.path.join(appdata, 'bankcheck')
        elif sys.platform == 'darwin':
            user_dir = os.path.join(os.path.expanduser('~/Library/Application Support'), 'bankcheck')
        else:
            user_dir = os.path.join(os.path.expanduser('~'), '.bankcheck')

        os.makedirs(user_dir, exist_ok=True)
        if os.access(user_dir, os.W_OK):
            return CheckResult(
                name='输出目录',
                passed=True,
                severity='error',
                message=f'输出目录可写（用户数据目录）：{user_dir}',
                fix_guide='',
            )
    except Exception:
        pass

    return CheckResult(
        name='输出目录',
        passed=False,
        severity='error',
        message='没有可写的输出目录',
        fix_guide='请确保程序目录或用户数据目录有写入权限',
    )


def run_self_check(
    include_optional: bool = False,
    script_dir: Optional[str] = None,
) -> SelfCheckReport:
    """
    执行完整的启动自检

    Args:
        include_optional: 是否检查可选依赖（仅警告）
        script_dir: 脚本目录，不传则自动检测

    Returns:
        SelfCheckReport 自检报告
    """
    report = SelfCheckReport()

    for result in check_required_dependencies():
        report.add_result(result)

    if include_optional:
        for result in check_optional_dependencies():
            report.add_result(result)

    report.add_result(check_lookup_table())
    report.add_result(check_temp_directory())
    report.add_result(check_output_directory(script_dir))

    return report


def format_report(report: SelfCheckReport, verbose: bool = False) -> str:
    """
    格式化自检报告为可读字符串

    Args:
        report: 自检报告
        verbose: 是否显示详细信息

    Returns:
        格式化的报告字符串
    """
    lines = []
    lines.append('=' * 60)
    lines.append('  程序启动自检报告')
    lines.append('=' * 60)
    lines.append('')

    if report.passed:
        lines.append('✅ 自检通过')
    else:
        lines.append(f'❌ 自检失败（{len(report.errors)} 项错误，{len(report.warnings)} 项警告）')
    lines.append('')

    if report.errors:
        lines.append('── 错误项 ──')
        lines.append('')
        for i, err in enumerate(report.errors, 1):
            lines.append(f'  [{i}] {err.name}')
            lines.append(f'      问题：{err.message}')
            lines.append(f'      修复：{err.fix_guide}')
            if verbose and err.details:
                lines.append(f'      详情：{err.details}')
            lines.append('')

    if report.warnings:
        lines.append('── 警告项 ──')
        lines.append('')
        for i, warn in enumerate(report.warnings, 1):
            lines.append(f'  [{i}] {warn.name}')
            lines.append(f'      问题：{warn.message}')
            if warn.fix_guide:
                lines.append(f'      建议：{warn.fix_guide}')
            if verbose and warn.details:
                lines.append(f'      详情：{warn.details}')
            lines.append('')

    if verbose:
        lines.append('── 全部检查项 ──')
        lines.append('')
        for result in report.results:
            status = '✅' if result.passed else '❌'
            lines.append(f'  {status} {result.name}：{result.message}')
        lines.append('')

    lines.append('=' * 60)

    return '\n'.join(lines)


def self_check_and_exit_if_failed(
    include_optional: bool = False,
    script_dir: Optional[str] = None,
    verbose: bool = True,
) -> SelfCheckReport:
    """
    执行自检，如果有严重错误则退出程序

    Args:
        include_optional: 是否检查可选依赖
        script_dir: 脚本目录
        verbose: 是否打印详细报告

    Returns:
        SelfCheckReport 自检报告
    """
    report = run_self_check(include_optional=include_optional, script_dir=script_dir)

    if verbose or not report.passed:
        print(format_report(report, verbose=verbose))

    if not report.passed:
        print('')
        print('❌ 自检未通过，程序无法启动。')
        print('请根据上述修复指引解决问题后重新运行。')
        print('')
        sys.exit(1)

    return report


if __name__ == '__main__':
    import argparse

    parser = argparse.ArgumentParser(description='银行流水检验工具 - 启动自检')
    parser.add_argument('--optional', action='store_true', help='检查可选依赖')
    parser.add_argument('--verbose', action='store_true', help='显示详细信息')
    args = parser.parse_args()

    report = run_self_check(include_optional=args.optional)
    print(format_report(report, verbose=args.verbose))

    if not report.passed:
        sys.exit(1)
