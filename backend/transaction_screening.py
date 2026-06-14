# -*- coding: utf-8 -*-
"""
大额与高频交易筛查模块
功能：
  1. 配置单笔金额阈值、单日笔数阈值与对方户名出现频次规则
  2. 筛查大额交易（单笔金额超过阈值）
  3. 筛查高频交易（单日交易笔数超过阈值）
  4. 筛查对方户名出现频次异常的交易
  5. 输出独立的风险提示清单供人工复核

使用方式：
  1. 直接调用 screening_from_records(records, config) 从交易记录列表筛查
  2. 调用 screening_from_total(total_path, config) 从银行流水总表文件筛查
  3. 调用 screening_from_db(config, db_backend) 从数据库筛查
"""

import os
import sys
import json
import logging
from dataclasses import dataclass, field, asdict
from datetime import datetime, timedelta
from typing import List, Optional, Dict, Any, Tuple
from collections import defaultdict


try:
    from pii_classifier import PIILogFilter
    HAS_PII_CLASSIFIER = True
except ImportError:
    HAS_PII_CLASSIFIER = False


try:
    import database as db_module
    from database import TransactionRecord, QueryResult
    HAS_DATABASE = True
except ImportError:
    HAS_DATABASE = False
    TransactionRecord = None
    QueryResult = None


try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False


def get_script_dir():
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def get_program_dir():
    return get_script_dir()


def is_writable(dir_path):
    import uuid
    if not os.path.isdir(dir_path):
        return False
    try:
        test_file = os.path.join(dir_path, '.ts_write_test_' + uuid.uuid4().hex[:8])
        with open(test_file, 'w') as f:
            f.write('test')
        os.remove(test_file)
        return True
    except (OSError, IOError):
        return False


def get_user_data_dir():
    app_name = 'bankcheck'
    if sys.platform.startswith('win'):
        base_dir = os.environ.get('APPDATA')
        if not base_dir:
            base_dir = os.path.expanduser('~\\AppData\\Roaming')
        return os.path.join(base_dir, app_name)
    elif sys.platform == 'darwin':
        return os.path.join(os.path.expanduser('~/Library/Application Support'), app_name)
    else:
        return os.path.join(os.path.expanduser('~'), '.' + app_name)


def get_writable_dir():
    program_dir = get_program_dir()
    if is_writable(program_dir):
        return program_dir
    user_data_dir = get_user_data_dir()
    os.makedirs(user_data_dir, exist_ok=True)
    return user_data_dir


def get_logger():
    logger = logging.getLogger('bankcheck.transaction_screening')
    if HAS_PII_CLASSIFIER:
        pii_filter = PIILogFilter()
        for h in logger.handlers:
            if not any(isinstance(f, PIILogFilter) for f in h.filters):
                h.addFilter(pii_filter)
    return logger


def to_float(value):
    if isinstance(value, (int, float)):
        return float(value)
    try:
        return float(str(value).strip())
    except (ValueError, TypeError):
        return None


def _get_amount(record: Dict[str, Any]) -> Optional[float]:
    payment = to_float(record.get('付款'))
    receipt = to_float(record.get('收款'))
    if payment is not None and payment != 0:
        return abs(payment)
    if receipt is not None and receipt != 0:
        return receipt
    return None


def _get_trade_date(record: Dict[str, Any]) -> Optional[str]:
    trade_date = record.get('交易日期')
    if trade_date is None:
        return None
    if isinstance(trade_date, datetime):
        return trade_date.strftime('%Y-%m-%d')
    date_str = str(trade_date).strip()
    if ' ' in date_str:
        date_str = date_str.split(' ')[0]
    return date_str


def _get_counterpart(record: Dict[str, Any]) -> Optional[str]:
    counterpart = record.get('对方户名')
    if counterpart is None:
        return None
    return str(counterpart).strip()


def _record_to_dict(record: Any) -> Dict[str, Any]:
    if isinstance(record, dict):
        return record
    if hasattr(record, 'to_dict'):
        return record.to_dict()
    if hasattr(record, '__dict__'):
        return record.__dict__
    return {}


# ──────────────────────────────────────────────
# 数据模型
# ──────────────────────────────────────────────

@dataclass
class CounterpartyFrequencyRule:
    """对方户名频次规则"""
    name: str = ''
    window_days: int = 30
    min_frequency: int = 10
    min_total_amount: Optional[float] = None
    enabled: bool = True

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'CounterpartyFrequencyRule':
        return cls(
            name=data.get('name', ''),
            window_days=int(data.get('window_days', 30)),
            min_frequency=int(data.get('min_frequency', 10)),
            min_total_amount=to_float(data.get('min_total_amount')),
            enabled=bool(data.get('enabled', True)),
        )


@dataclass
class ScreeningConfig:
    """筛查规则配置"""
    single_amount_threshold: Optional[float] = None
    daily_count_threshold: Optional[int] = None
    counterparty_rules: List[CounterpartyFrequencyRule] = field(default_factory=list)
    enabled: bool = True
    config_name: str = '默认配置'

    def to_dict(self) -> Dict[str, Any]:
        return {
            'config_name': self.config_name,
            'enabled': self.enabled,
            'single_amount_threshold': self.single_amount_threshold,
            'daily_count_threshold': self.daily_count_threshold,
            'counterparty_rules': [rule.to_dict() for rule in self.counterparty_rules],
        }

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'ScreeningConfig':
        rules_data = data.get('counterparty_rules', [])
        rules = [CounterpartyFrequencyRule.from_dict(r) for r in rules_data]
        return cls(
            config_name=data.get('config_name', '默认配置'),
            enabled=bool(data.get('enabled', True)),
            single_amount_threshold=to_float(data.get('single_amount_threshold')),
            daily_count_threshold=data.get('daily_count_threshold'),
            counterparty_rules=rules,
        )

    @classmethod
    def default(cls) -> 'ScreeningConfig':
        """创建默认筛查配置"""
        return cls(
            config_name='默认风险筛查配置',
            single_amount_threshold=500000.0,
            daily_count_threshold=50,
            counterparty_rules=[
                CounterpartyFrequencyRule(
                    name='30天内交易超过20笔',
                    window_days=30,
                    min_frequency=20,
                    min_total_amount=100000.0,
                    enabled=True,
                ),
                CounterpartyFrequencyRule(
                    name='7天内交易超过10笔',
                    window_days=7,
                    min_frequency=10,
                    min_total_amount=None,
                    enabled=True,
                ),
            ],
        )


class RiskType:
    """风险类型常量"""
    LARGE_AMOUNT = 'large_amount'
    HIGH_FREQUENCY = 'high_frequency'
    COUNTERPARTY_FREQUENCY = 'counterparty_frequency'
    AMOUNT_ANOMALY = 'amount_anomaly'

    LABELS = {
        LARGE_AMOUNT: '大额交易',
        HIGH_FREQUENCY: '高频交易',
        COUNTERPARTY_FREQUENCY: '对方户名频次异常',
        AMOUNT_ANOMALY: '金额字段异常',
    }


@dataclass
class RiskAlert:
    """风险提示项"""
    risk_id: str = ''
    risk_type: str = ''
    risk_level: str = 'medium'
    description: str = ''
    transaction_id: Optional[str] = None
    trade_date: Optional[str] = None
    amount: Optional[float] = None
    counterpart: Optional[str] = None
    bank: Optional[str] = None
    account: Optional[str] = None
    subject: Optional[str] = None
    summary: Optional[str] = None
    matched_rule: Optional[str] = None
    rule_details: Optional[Dict[str, Any]] = None
    related_transaction_ids: List[str] = field(default_factory=list)
    created_at: str = field(default_factory=lambda: datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
    review_status: str = 'pending'
    reviewer: Optional[str] = None
    review_comment: Optional[str] = None

    def to_dict(self) -> Dict[str, Any]:
        data = asdict(self)
        if self.rule_details:
            data['rule_details'] = json.dumps(self.rule_details, ensure_ascii=False)
        return data

    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> 'RiskAlert':
        alert = cls()
        for key, value in data.items():
            if hasattr(alert, key):
                setattr(alert, key, value)
        if isinstance(data.get('rule_details'), str):
            try:
                alert.rule_details = json.loads(data['rule_details'])
            except (json.JSONDecodeError, TypeError):
                alert.rule_details = None
        return alert


@dataclass
class ScreeningResult:
    """筛查结果"""
    config: ScreeningConfig
    total_records: int = 0
    large_amount_alerts: List[RiskAlert] = field(default_factory=list)
    high_frequency_alerts: List[RiskAlert] = field(default_factory=list)
    counterparty_alerts: List[RiskAlert] = field(default_factory=list)
    source_info: Optional[Dict[str, Any]] = None

    @property
    def all_alerts(self) -> List[RiskAlert]:
        return (
            self.large_amount_alerts
            + self.high_frequency_alerts
            + self.counterparty_alerts
        )

    @property
    def alert_count(self) -> int:
        return len(self.all_alerts)

    @property
    def high_risk_count(self) -> int:
        return sum(1 for a in self.all_alerts if a.risk_level == 'high')

    @property
    def medium_risk_count(self) -> int:
        return sum(1 for a in self.all_alerts if a.risk_level == 'medium')

    @property
    def low_risk_count(self) -> int:
        return sum(1 for a in self.all_alerts if a.risk_level == 'low')

    def to_dict(self) -> Dict[str, Any]:
        return {
            'config': self.config.to_dict(),
            'total_records': self.total_records,
            'large_amount_alerts': [a.to_dict() for a in self.large_amount_alerts],
            'high_frequency_alerts': [a.to_dict() for a in self.high_frequency_alerts],
            'counterparty_alerts': [a.to_dict() for a in self.counterparty_alerts],
            'source_info': self.source_info,
            'summary': {
                'total_alerts': self.alert_count,
                'high_risk': self.high_risk_count,
                'medium_risk': self.medium_risk_count,
                'low_risk': self.low_risk_count,
            },
        }


# ──────────────────────────────────────────────
# 筛查逻辑
# ──────────────────────────────────────────────

def _generate_risk_id(prefix: str) -> str:
    timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
    import random
    suffix = ''.join(random.choices('0123456789', k=6))
    return f'{prefix}_{timestamp}_{suffix}'


def screen_large_amount(
    records: List[Dict[str, Any]],
    threshold: float,
) -> List[RiskAlert]:
    """
    筛查大额交易

    Args:
        records: 交易记录列表
        threshold: 单笔金额阈值

    Returns:
        List[RiskAlert]: 大额交易风险提示列表
    """
    logger = get_logger()
    alerts = []

    if threshold is None or threshold <= 0:
        logger.warning('大额交易阈值未配置或无效，跳过筛查')
        return alerts

    logger.info('开始大额交易筛查，阈值: %.2f 元', threshold)

    for record in records:
        amount = _get_amount(record)
        if amount is None:
            continue

        if amount >= threshold:
            trade_date = _get_trade_date(record)
            counterpart = _get_counterpart(record)

            multiple = amount / threshold
            if multiple >= 5:
                risk_level = 'high'
            elif multiple >= 2:
                risk_level = 'medium'
            else:
                risk_level = 'low'

            alert = RiskAlert(
                risk_id=_generate_risk_id('LARGE'),
                risk_type=RiskType.LARGE_AMOUNT,
                risk_level=risk_level,
                description=f'单笔交易金额 {amount:,.2f} 元，超过阈值 {threshold:,.2f} 元',
                transaction_id=record.get('交易流水号') or record.get('唯一id'),
                trade_date=trade_date,
                amount=amount,
                counterpart=counterpart,
                bank=record.get('银行'),
                account=record.get('银行账号'),
                subject=record.get('主体'),
                summary=record.get('摘要'),
                matched_rule=f'单笔金额阈值 {threshold:,.2f} 元',
                rule_details={
                    'threshold': threshold,
                    'actual_amount': amount,
                    'multiple': round(multiple, 2),
                },
            )
            alerts.append(alert)

    logger.info('大额交易筛查完成，发现 %d 条风险', len(alerts))
    return alerts


def screen_high_frequency(
    records: List[Dict[str, Any]],
    threshold: int,
) -> List[RiskAlert]:
    """
    筛查高频交易（单日交易笔数超过阈值）

    Args:
        records: 交易记录列表
        threshold: 单日交易笔数阈值

    Returns:
        List[RiskAlert]: 高频交易风险提示列表
    """
    logger = get_logger()
    alerts = []

    if threshold is None or threshold <= 0:
        logger.warning('高频交易阈值未配置或无效，跳过筛查')
        return alerts

    logger.info('开始高频交易筛查，单日笔数阈值: %d', threshold)

    date_records: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for record in records:
        trade_date = _get_trade_date(record)
        if trade_date:
            date_records[trade_date].append(record)

    for trade_date, day_records in date_records.items():
        count = len(day_records)
        if count >= threshold:
            total_amount = sum(
                _get_amount(r) or 0 for r in day_records
            )

            multiple = count / threshold
            if multiple >= 3:
                risk_level = 'high'
            elif multiple >= 1.5:
                risk_level = 'medium'
            else:
                risk_level = 'low'

            transaction_ids = [
                r.get('交易流水号') or r.get('唯一id') for r in day_records
                if r.get('交易流水号') or r.get('唯一id')
            ]

            alert = RiskAlert(
                risk_id=_generate_risk_id('HIGH'),
                risk_type=RiskType.HIGH_FREQUENCY,
                risk_level=risk_level,
                description=f'{trade_date} 当日交易 {count} 笔，超过阈值 {threshold} 笔',
                trade_date=trade_date,
                amount=round(total_amount, 2),
                bank=day_records[0].get('银行') if day_records else None,
                account=day_records[0].get('银行账号') if day_records else None,
                subject=day_records[0].get('主体') if day_records else None,
                matched_rule=f'单日笔数阈值 {threshold} 笔',
                rule_details={
                    'threshold': threshold,
                    'actual_count': count,
                    'multiple': round(multiple, 2),
                    'total_amount': round(total_amount, 2),
                    'transaction_count': count,
                },
                related_transaction_ids=transaction_ids,
            )
            alerts.append(alert)

    logger.info('高频交易筛查完成，发现 %d 条风险', len(alerts))
    return alerts


def screen_counterparty_frequency(
    records: List[Dict[str, Any]],
    rules: List[CounterpartyFrequencyRule],
) -> List[RiskAlert]:
    """
    筛查对方户名出现频次异常的交易

    Args:
        records: 交易记录列表
        rules: 对方户名频次规则列表

    Returns:
        List[RiskAlert]: 对方户名频次异常风险提示列表
    """
    logger = get_logger()
    alerts = []

    if not rules:
        logger.warning('对方户名频次规则未配置，跳过筛查')
        return alerts

    enabled_rules = [r for r in rules if r.enabled]
    if not enabled_rules:
        logger.warning('对方户名频次规则均未启用，跳过筛查')
        return alerts

    logger.info('开始对方户名频次筛查，启用规则数: %d', len(enabled_rules))

    counterpart_records: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for record in records:
        counterpart = _get_counterpart(record)
        if counterpart:
            counterpart_records[counterpart].append(record)

    for rule in enabled_rules:
        window_days = rule.window_days
        min_frequency = rule.min_frequency
        min_total_amount = rule.min_total_amount

        logger.info('执行规则 [%s]：%d天内交易>=%d笔，总金额>=%s',
                    rule.name, window_days, min_frequency,
                    f'{min_total_amount:,.2f}元' if min_total_amount else '不限')

        for counterpart, cpt_records in counterpart_records.items():
            dated_records = []
            for record in cpt_records:
                trade_date = _get_trade_date(record)
                if not trade_date:
                    continue
                try:
                    record_date = datetime.strptime(trade_date, '%Y-%m-%d')
                    dated_records.append((record_date, record))
                except ValueError:
                    continue

            if len(dated_records) < min_frequency:
                continue

            dated_records.sort(key=lambda x: x[0])

            n = len(dated_records)
            for i in range(n):
                window_end = dated_records[i][0]
                window_start = window_end - timedelta(days=window_days - 1)

                window_records = []
                for j in range(n):
                    if window_start <= dated_records[j][0] <= window_end:
                        window_records.append(dated_records[j][1])

                count = len(window_records)
                total_amount = sum(
                    _get_amount(r) or 0 for r in window_records
                )

                if count < min_frequency:
                    continue
                if min_total_amount is not None and total_amount < min_total_amount:
                    continue

                multiple = count / min_frequency
                if multiple >= 3:
                    risk_level = 'high'
                elif multiple >= 1.5:
                    risk_level = 'medium'
                else:
                    risk_level = 'low'

                transaction_ids = [
                    r.get('交易流水号') or r.get('唯一id') for r in window_records
                    if r.get('交易流水号') or r.get('唯一id')
                ]

                alert = RiskAlert(
                    risk_id=_generate_risk_id('COUNTER'),
                    risk_type=RiskType.COUNTERPARTY_FREQUENCY,
                    risk_level=risk_level,
                    description=(
                        f'对方户名 [{counterpart}] 在 {window_days} 天内'
                        f'（{window_start.strftime("%Y-%m-%d")} 至 '
                        f'{window_end.strftime("%Y-%m-%d")}）发生交易 {count} 笔，'
                        f'总金额 {total_amount:,.2f} 元'
                    ),
                    trade_date=window_end.strftime('%Y-%m-%d'),
                    amount=round(total_amount, 2),
                    counterpart=counterpart,
                    matched_rule=rule.name,
                    rule_details={
                        'rule_name': rule.name,
                        'window_days': window_days,
                        'min_frequency': min_frequency,
                        'min_total_amount': min_total_amount,
                        'actual_count': count,
                        'actual_total_amount': round(total_amount, 2),
                        'multiple': round(multiple, 2),
                        'window_start_date': window_start.strftime('%Y-%m-%d'),
                        'window_end_date': window_end.strftime('%Y-%m-%d'),
                    },
                    related_transaction_ids=transaction_ids,
                )
                alerts.append(alert)

    seen = set()
    unique_alerts = []
    for alert in alerts:
        key = (alert.risk_type, alert.counterpart, alert.trade_date, alert.matched_rule)
        if key not in seen:
            seen.add(key)
            unique_alerts.append(alert)

    logger.info('对方户名频次筛查完成，发现 %d 条风险', len(unique_alerts))
    return unique_alerts


def run_screening(
    records: List[Dict[str, Any]],
    config: ScreeningConfig,
    source_info: Optional[Dict[str, Any]] = None,
) -> ScreeningResult:
    """
    执行完整的交易筛查流程

    Args:
        records: 交易记录列表
        config: 筛查配置
        source_info: 数据源信息

    Returns:
        ScreeningResult: 筛查结果
    """
    logger = get_logger()

    if not config.enabled:
        logger.warning('筛查配置未启用，跳过筛查')
        return ScreeningResult(config=config, source_info=source_info)

    if not records:
        logger.warning('无交易记录，跳过筛查')
        return ScreeningResult(config=config, source_info=source_info)

    logger.info('========== 交易筛查开始 ==========')
    logger.info('配置名称: %s', config.config_name)
    logger.info('待筛查记录数: %d', len(records))

    result = ScreeningResult(
        config=config,
        total_records=len(records),
        source_info=source_info,
    )

    if config.single_amount_threshold:
        result.large_amount_alerts = screen_large_amount(
            records, config.single_amount_threshold
        )

    if config.daily_count_threshold:
        result.high_frequency_alerts = screen_high_frequency(
            records, config.daily_count_threshold
        )

    if config.counterparty_rules:
        result.counterparty_alerts = screen_counterparty_frequency(
            records, config.counterparty_rules
        )

    logger.info('========== 交易筛查完成 ==========')
    logger.info('总计风险提示: %d 条', result.alert_count)
    logger.info('  - 大额交易: %d 条', len(result.large_amount_alerts))
    logger.info('  - 高频交易: %d 条', len(result.high_frequency_alerts))
    logger.info('  - 对方户名频次异常: %d 条', len(result.counterparty_alerts))
    logger.info('风险等级分布: 高 %d, 中 %d, 低 %d',
                result.high_risk_count, result.medium_risk_count, result.low_risk_count)

    return result


# ──────────────────────────────────────────────
# 报告导出
# ──────────────────────────────────────────────

def _apply_header_style(ws, row_num, col_count):
    if not HAS_OPENPYXL:
        return

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border


def _apply_alert_row_style(ws, row_num, col_count, risk_level):
    if not HAS_OPENPYXL:
        return

    fill_colors = {
        'high': 'F8CBAD',
        'medium': 'FFE699',
        'low': 'C6E0B4',
    }
    fill_color = fill_colors.get(risk_level, 'FFFFFF')
    cell_fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = cell_fill
        cell.border = thin_border
        if col in (4, 5, 6):
            cell.alignment = left_align
        else:
            cell.alignment = center_align


def export_screening_result(
    result: ScreeningResult,
    output_path: str,
    source_info: Optional[Dict[str, Any]] = None,
) -> Optional[str]:
    """
    导出筛查结果为 Excel 报告

    Args:
        result: 筛查结果
        output_path: 输出文件路径
        source_info: 数据源信息

    Returns:
        str: 输出文件路径，失败则返回 None
    """
    logger = get_logger()

    if not HAS_OPENPYXL:
        logger.error('缺少 openpyxl 依赖，无法导出 Excel 报告')
        return None

    if result.alert_count == 0:
        logger.warning('无风险提示，跳过报告生成')
        return None

    logger.info('开始生成风险提示清单: %s', output_path)

    wb = openpyxl.Workbook()

    ws_summary = wb.active
    ws_summary.title = '筛查总览'

    summary_headers = [
        '项目', '数值', '说明'
    ]
    for col, header in enumerate(summary_headers, 1):
        ws_summary.cell(row=1, column=col, value=header)
    _apply_header_style(ws_summary, 1, len(summary_headers))

    summary_data = [
        ('配置名称', result.config.config_name, '筛查规则配置名称'),
        ('筛查记录总数', result.total_records, '参与筛查的交易记录总数'),
        ('风险提示总数', result.alert_count, '所有类型风险提示数量合计'),
        ('高风险数量', result.high_risk_count, '需要立即关注的高风险项'),
        ('中风险数量', result.medium_risk_count, '需要重点关注的中风险项'),
        ('低风险数量', result.low_risk_count, '需要留意的低风险项'),
        ('大额交易风险', len(result.large_amount_alerts), '单笔金额超过阈值的交易'),
        ('高频交易风险', len(result.high_frequency_alerts), '单日交易笔数超过阈值'),
        ('对方户名频次异常', len(result.counterparty_alerts), '特定户名交易频次异常'),
        ('大额交易阈值',
         f'{result.config.single_amount_threshold:,.2f} 元'
         if result.config.single_amount_threshold else '未配置',
         '单笔金额筛查阈值'),
        ('单日笔数阈值',
         f'{result.config.daily_count_threshold} 笔'
         if result.config.daily_count_threshold else '未配置',
         '单日交易笔数筛查阈值'),
        ('对方户名规则数',
         len([r for r in result.config.counterparty_rules if r.enabled]),
         '已启用的对方户名频次规则数量'),
        ('生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S'), '报告生成时间'),
    ]

    for row_idx, (item, value, desc) in enumerate(summary_data, 2):
        ws_summary.cell(row=row_idx, column=1, value=item)
        ws_summary.cell(row=row_idx, column=2, value=value)
        ws_summary.cell(row=row_idx, column=3, value=desc)

    for col in range(1, 4):
        ws_summary.column_dimensions[chr(64 + col)].width = 30
    ws_summary.freeze_panes = 'A2'

    if source_info or result.source_info:
        info = source_info or result.source_info
        if info:
            ws_info = wb.create_sheet('数据来源')
            info_headers = ['信息项', '内容']
            for col, header in enumerate(info_headers, 1):
                ws_info.cell(row=1, column=col, value=header)
            _apply_header_style(ws_info, 1, len(info_headers))
            for row_idx, (key, value) in enumerate(info.items(), 2):
                ws_info.cell(row=row_idx, column=1, value=key)
                ws_info.cell(row=row_idx, column=2, value=str(value))
            ws_info.column_dimensions['A'].width = 25
            ws_info.column_dimensions['B'].width = 60

    alert_columns = [
        '风险编号', '风险类型', '风险等级', '风险描述',
        '交易日期', '交易金额(元)', '对方户名', '银行',
        '主体', '匹配规则', '交易流水号', '关联交易数',
        '复核状态', '创建时间',
    ]

    risk_type_sheets = [
        (RiskType.LARGE_AMOUNT, '大额交易风险', result.large_amount_alerts),
        (RiskType.HIGH_FREQUENCY, '高频交易风险', result.high_frequency_alerts),
        (RiskType.COUNTERPARTY_FREQUENCY, '对方户名频次异常', result.counterparty_alerts),
    ]

    for risk_type, sheet_title, alerts in risk_type_sheets:
        if not alerts:
            continue

        ws = wb.create_sheet(sheet_title)

        for col, header in enumerate(alert_columns, 1):
            ws.cell(row=1, column=col, value=header)
        _apply_header_style(ws, 1, len(alert_columns))

        for row_idx, alert in enumerate(alerts, 2):
            risk_type_label = RiskType.LABELS.get(alert.risk_type, alert.risk_type)
            risk_level_label = {'high': '高', 'medium': '中', 'low': '低'}.get(alert.risk_level, alert.risk_level)

            ws.cell(row=row_idx, column=1, value=alert.risk_id)
            ws.cell(row=row_idx, column=2, value=risk_type_label)
            ws.cell(row=row_idx, column=3, value=risk_level_label)
            ws.cell(row=row_idx, column=4, value=alert.description)
            ws.cell(row=row_idx, column=5, value=alert.trade_date or '')
            ws.cell(row=row_idx, column=6, value=f'{alert.amount:,.2f}' if alert.amount is not None else '')
            ws.cell(row=row_idx, column=7, value=alert.counterpart or '')
            ws.cell(row=row_idx, column=8, value=alert.bank or '')
            ws.cell(row=row_idx, column=9, value=alert.subject or '')
            ws.cell(row=row_idx, column=10, value=alert.matched_rule or '')
            ws.cell(row=row_idx, column=11, value=alert.transaction_id or '')
            ws.cell(row=row_idx, column=12, value=len(alert.related_transaction_ids))
            ws.cell(row=row_idx, column=13, value={'pending': '待复核', 'confirmed': '已确认', 'ignored': '已忽略'}.get(alert.review_status, alert.review_status))
            ws.cell(row=row_idx, column=14, value=alert.created_at)

            _apply_alert_row_style(ws, row_idx, len(alert_columns), alert.risk_level)

        col_widths = [28, 12, 10, 50, 12, 15, 25, 12, 20, 25, 25, 12, 10, 20]
        for col, width in enumerate(col_widths, 1):
            ws.column_dimensions[chr(64 + col)].width = width
        ws.freeze_panes = 'A2'

    ws_all = wb.create_sheet('全部风险提示', 1)
    for col, header in enumerate(alert_columns, 1):
        ws_all.cell(row=1, column=col, value=header)
    _apply_header_style(ws_all, 1, len(alert_columns))

    all_alerts_sorted = sorted(
        result.all_alerts,
        key=lambda a: {'high': 0, 'medium': 1, 'low': 2}.get(a.risk_level, 3),
    )

    for row_idx, alert in enumerate(all_alerts_sorted, 2):
        risk_type_label = RiskType.LABELS.get(alert.risk_type, alert.risk_type)
        risk_level_label = {'high': '高', 'medium': '中', 'low': '低'}.get(alert.risk_level, alert.risk_level)

        ws_all.cell(row=row_idx, column=1, value=alert.risk_id)
        ws_all.cell(row=row_idx, column=2, value=risk_type_label)
        ws_all.cell(row=row_idx, column=3, value=risk_level_label)
        ws_all.cell(row=row_idx, column=4, value=alert.description)
        ws_all.cell(row=row_idx, column=5, value=alert.trade_date or '')
        ws_all.cell(row=row_idx, column=6, value=f'{alert.amount:,.2f}' if alert.amount is not None else '')
        ws_all.cell(row=row_idx, column=7, value=alert.counterpart or '')
        ws_all.cell(row=row_idx, column=8, value=alert.bank or '')
        ws_all.cell(row=row_idx, column=9, value=alert.subject or '')
        ws_all.cell(row=row_idx, column=10, value=alert.matched_rule or '')
        ws_all.cell(row=row_idx, column=11, value=alert.transaction_id or '')
        ws_all.cell(row=row_idx, column=12, value=len(alert.related_transaction_ids))
        ws_all.cell(row=row_idx, column=13, value={'pending': '待复核', 'confirmed': '已确认', 'ignored': '已忽略'}.get(alert.review_status, alert.review_status))
        ws_all.cell(row=row_idx, column=14, value=alert.created_at)

        _apply_alert_row_style(ws_all, row_idx, len(alert_columns), alert.risk_level)

    col_widths = [28, 12, 10, 50, 12, 15, 25, 12, 20, 25, 25, 12, 10, 20]
    for col, width in enumerate(col_widths, 1):
        ws_all.column_dimensions[chr(64 + col)].width = width
    ws_all.freeze_panes = 'A2'

    ws_rules = wb.create_sheet('筛查规则配置')
    rule_headers = ['规则类型', '规则名称', '配置值', '说明']
    for col, header in enumerate(rule_headers, 1):
        ws_rules.cell(row=1, column=col, value=header)
    _apply_header_style(ws_rules, 1, len(rule_headers))

    rule_data = [
        ('大额交易', '单笔金额阈值',
         f'{result.config.single_amount_threshold:,.2f} 元'
         if result.config.single_amount_threshold else '未启用',
         '单笔交易金额超过此值即触发风险提示'),
        ('高频交易', '单日笔数阈值',
         f'{result.config.daily_count_threshold} 笔'
         if result.config.daily_count_threshold else '未启用',
         '单日交易笔数超过此值即触发风险提示'),
    ]

    for rule in result.config.counterparty_rules:
        rule_data.append((
            '对方户名频次',
            rule.name,
            f'{rule.window_days}天内 >= {rule.min_frequency}笔' +
            (f'，总金额 >= {rule.min_total_amount:,.2f}元' if rule.min_total_amount else ''),
            '已启用' if rule.enabled else '已禁用',
        ))

    for row_idx, (rule_type, rule_name, value, desc) in enumerate(rule_data, 2):
        ws_rules.cell(row=row_idx, column=1, value=rule_type)
        ws_rules.cell(row=row_idx, column=2, value=rule_name)
        ws_rules.cell(row=row_idx, column=3, value=value)
        ws_rules.cell(row=row_idx, column=4, value=desc)

    ws_rules.column_dimensions['A'].width = 15
    ws_rules.column_dimensions['B'].width = 25
    ws_rules.column_dimensions['C'].width = 35
    ws_rules.column_dimensions['D'].width = 30

    try:
        wb.save(output_path)
        wb.close()
        logger.info('风险提示清单已生成: %s', output_path)
        return output_path
    except Exception as e:
        logger.error('生成风险提示清单失败: %s', e)
        return None


# ──────────────────────────────────────────────
# 便捷入口函数
# ──────────────────────────────────────────────

def load_total_table(total_path: str) -> List[Dict[str, Any]]:
    """
    加载银行流水总表文件

    Args:
        total_path: 总表文件路径

    Returns:
        List[Dict[str, Any]]: 交易记录列表
    """
    logger = get_logger()

    if not HAS_PANDAS:
        logger.error('缺少 pandas 依赖，无法读取 Excel 文件')
        return []

    if not total_path or not os.path.exists(total_path):
        logger.error('总表文件不存在: %s', total_path)
        return []

    try:
        df = pd.read_excel(total_path, engine='openpyxl')
        records = df.to_dict('records')
        logger.info('已加载总表记录 %d 条', len(records))
        return records
    except Exception as e:
        logger.error('读取总表文件失败: %s', e)
        return []


def screening_from_records(
    records: List[Any],
    config: Optional[ScreeningConfig] = None,
    output_dir: Optional[str] = None,
    source_info: Optional[Dict[str, Any]] = None,
) -> Tuple[Optional[ScreeningResult], Optional[str]]:
    """
    从交易记录列表执行筛查并生成报告

    Args:
        records: 交易记录列表
        config: 筛查配置，为空则使用默认配置
        output_dir: 输出目录，为空则使用脚本目录
        source_info: 数据源信息

    Returns:
        Tuple[Optional[ScreeningResult], Optional[str]]: (筛查结果, 报告文件路径)
    """
    logger = get_logger()

    if config is None:
        config = ScreeningConfig.default()

    record_dicts = [_record_to_dict(r) for r in records]

    result = run_screening(record_dicts, config, source_info)

    if result.alert_count == 0:
        logger.info('未发现任何风险，无需生成报告')
        return result, None

    if output_dir is None:
        output_dir = get_writable_dir()

    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'风险提示清单_{timestamp}.xlsx'
    output_path = os.path.join(output_dir, filename)

    report_path = export_screening_result(result, output_path, source_info)
    return result, report_path


def screening_from_total(
    total_path: str,
    config: Optional[ScreeningConfig] = None,
    output_dir: Optional[str] = None,
) -> Tuple[Optional[ScreeningResult], Optional[str]]:
    """
    从银行流水总表文件执行筛查并生成报告

    Args:
        total_path: 总表文件路径
        config: 筛查配置，为空则使用默认配置
        output_dir: 输出目录，为空则使用脚本目录

    Returns:
        Tuple[Optional[ScreeningResult], Optional[str]]: (筛查结果, 报告文件路径)
    """
    logger = get_logger()

    records = load_total_table(total_path)
    if not records:
        logger.warning('总表文件无数据，跳过筛查')
        return None, None

    source_info = {
        '数据来源文件': os.path.basename(total_path),
        '总表记录数': len(records),
        '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    }

    return screening_from_records(records, config, output_dir, source_info)


def screening_from_db(
    config: Optional[ScreeningConfig] = None,
    output_dir: Optional[str] = None,
    db_backend=None,
    **query_kwargs,
) -> Tuple[Optional[ScreeningResult], Optional[str]]:
    """
    从数据库查询交易记录并执行筛查

    Args:
        config: 筛查配置，为空则使用默认配置
        output_dir: 输出目录，为空则使用脚本目录
        db_backend: 数据库后端实例，为空则创建默认 SQLite 后端
        **query_kwargs: 查询参数，传递给 query_records

    Returns:
        Tuple[Optional[ScreeningResult], Optional[str]]: (筛查结果, 报告文件路径)
    """
    logger = get_logger()

    if not HAS_DATABASE:
        logger.error('数据库模块不可用')
        return None, None

    if db_backend is None:
        db_backend = db_module.SQLiteBackend()

    need_close = False
    try:
        if db_backend.conn is None:
            db_backend.connect()
            need_close = True

        query_result = db_backend.query_records(**query_kwargs)
        records = [r.to_dict() for r in query_result.records]

        if not records:
            logger.warning('数据库查询无结果，跳过筛查')
            return None, None

        source_info = {
            '数据来源': '数据库',
            '查询条件': str(query_kwargs),
            '查询记录数': len(records),
            '生成时间': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        }

        return screening_from_records(records, config, output_dir, source_info)
    finally:
        if need_close and db_backend.conn is not None:
            db_backend.disconnect()
